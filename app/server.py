import os, sqlite3, hashlib, secrets, urllib.parse, json, csv, io, shutil, socket, threading, webbrowser, zipfile, tempfile, hmac, time, gc, base64, uuid, smtplib, ssl, random, re, unicodedata
from email.parser import BytesParser
from email.message import EmailMessage
from email.policy import default
from http.server import ThreadingHTTPServer, BaseHTTPRequestHandler
from http import cookies
from datetime import datetime, date, timedelta
from pathlib import Path
from PIL import Image, ImageOps
from cryptography.hazmat.primitives.asymmetric.ed25519 import Ed25519PublicKey
from cryptography.exceptions import InvalidSignature

PROGRAM_DIR=Path(__file__).resolve().parent
DEFAULT_DATA_ROOT=Path(os.environ.get('LOCALAPPDATA') or Path.home())/'CiftlikPro'
DATA_ROOT=Path(os.environ.get('CIFTLIKPRO_DATA_DIR') or DEFAULT_DATA_ROOT)
DATA_ROOT.mkdir(parents=True,exist_ok=True)
DB=DATA_ROOT/'ciftlik.db'
BACKUPS=DATA_ROOT/'backups'
UPLOADS=DATA_ROOT/'uploads'
PORT=int(os.environ.get('CIFTLIKPRO_PORT','8953'))
SESSIONS={}
ANIMAL_IMPORT_PREVIEWS={}
ANIMAL_IMPORT_LOCK=threading.Lock()

APP_NAME='ÇiftlikPro Enterprise'
APP_VERSION='3.9.20'
APP_CHANNEL='RELEASE'
APP_LABEL='v3.9.20'

LICENSE_FILE=DATA_ROOT/'ciftlikpro.license'
LICENSE_PUBLIC_KEY_B64='Z9rGVotpzHR7eNxdVtFX3ztjrxhzhSYBHweob5EYqHE='

def device_id():
    parts=[]
    if os.name=='nt':
        try:
            import winreg
            with winreg.OpenKey(winreg.HKEY_LOCAL_MACHINE,r'SOFTWARE\Microsoft\Cryptography') as k:
                parts.append(str(winreg.QueryValueEx(k,'MachineGuid')[0]))
        except Exception: pass
    if not parts:
        parts.extend([socket.gethostname(),str(uuid.getnode())])
    digest=hashlib.sha256(('|'.join(parts)+'|CiftlikPro').encode('utf-8')).hexdigest().upper()
    return 'CF-'+digest[:4]+'-'+digest[4:8]+'-'+digest[8:12]+'-'+digest[12:16]

def license_payload_bytes(payload):
    return json.dumps(payload,ensure_ascii=False,sort_keys=True,separators=(',',':')).encode('utf-8')

def license_doc_to_key(doc):
    import zlib
    raw=json.dumps(doc,ensure_ascii=False,separators=(",",":")).encode("utf-8")
    packed=zlib.compress(raw,9)
    text=base64.b32encode(packed).decode("ascii").rstrip("=")
    return "CFP-"+"-".join(text[i:i+5] for i in range(0,len(text),5))

def _device_id_from_raw(raw8):
    hx=raw8.hex().upper()
    return 'CF-'+hx[0:4]+'-'+hx[4:8]+'-'+hx[8:12]+'-'+hx[12:16]

def _license_date(value):
    if not value:return ''
    return (date(2020,1,1)+timedelta(days=int(value)-1)).isoformat()

def license_key_to_bytes(key):
    import zlib,struct,re
    raw_key=str(key or '').upper().strip()
    if raw_key.startswith('CFS-'):
        clean=re.sub(r'[^A-Z2-7]','',raw_key[4:])
        if not clean:raise ValueError('Lisans anahtarı boş.')
        clean+='='*((8-len(clean)%8)%8)
        blob=base64.b32decode(clean,casefold=True)
        if len(blob)<78 or blob[0]!=1:raise ValueError('Kısa lisans anahtarı biçimi geçersiz.')
        owner_len=blob[13]
        if len(blob)!=(14+owner_len+64):raise ValueError('Kısa lisans anahtarı eksik veya bozuk.')
        issued_days,expires_days=struct.unpack('>HH',blob[9:13])
        payload={
            'product':'CiftlikPro Enterprise',
            'licensee':blob[14:14+owner_len].decode('utf-8'),
            'device_id':_device_id_from_raw(blob[1:9]),
            'license_type':'Süresiz' if not expires_days else 'Süreli',
            'issued_on':_license_date(issued_days),
            'expires_on':_license_date(expires_days),
        }
        doc={'payload':payload,'signature':base64.b64encode(blob[14+owner_len:]).decode('ascii')}
        return json.dumps(doc,ensure_ascii=False,separators=(',',':')).encode('utf-8')
    clean=raw_key
    if clean.startswith('CFP-'):clean=clean[4:]
    clean=re.sub(r'[^A-Z2-7]','',clean)
    if not clean:raise ValueError('Lisans anahtarı boş.')
    clean+='='*((8-len(clean)%8)%8)
    return zlib.decompress(base64.b32decode(clean,casefold=True))

def validate_license_bytes(raw=None):
    try:
        if raw is None:
            if not LICENSE_FILE.exists(): return False,None,'Lisans dosyası bulunamadı.'
            raw=LICENSE_FILE.read_bytes()
        doc=json.loads(raw.decode('utf-8'))
        payload=doc.get('payload') or {}
        signature=base64.b64decode(doc.get('signature',''),validate=True)
        Ed25519PublicKey.from_public_bytes(base64.b64decode(LICENSE_PUBLIC_KEY_B64)).verify(signature,license_payload_bytes(payload))
        if payload.get('product')!='CiftlikPro Enterprise': return False,payload,'Bu lisans farklı bir ürün için.'
        if str(payload.get('device_id','')).upper()!=device_id().upper(): return False,payload,'Lisans bu bilgisayara ait değil.'
        expires=str(payload.get('expires_on') or '').strip()
        if expires and date.today()>date.fromisoformat(expires): return False,payload,'Lisans süresi dolmuş.'
        return True,payload,'Aktif'
    except InvalidSignature:return False,None,'Lisans imzası geçersiz.'
    except Exception as exc:return False,None,'Lisans doğrulanamadı: '+str(exc)

_LICENSE_CACHE={'stamp':None,'checked':0.0,'value':None}
def license_status():
    """Aynı lisans dosyasının imza/cihaz doğrulamasını her sayfa isteğinde tekrarlama."""
    try: stamp=(LICENSE_FILE.stat().st_mtime_ns,LICENSE_FILE.stat().st_size)
    except Exception: stamp=None
    now=time.monotonic()
    if _LICENSE_CACHE['value'] is not None and _LICENSE_CACHE['stamp']==stamp and now-_LICENSE_CACHE['checked']<30.0:
        return _LICENSE_CACHE['value']
    value=validate_license_bytes();_LICENSE_CACHE.update(stamp=stamp,checked=now,value=value);return value


CSS='''
:root{--g:#176b3a;--g2:#228b4f;--bg:#f3f6f4;--card:#fff;--txt:#203127;--mut:#6b7b70;--red:#c8392b;--orange:#e58c16;--blue:#2e6fc2}
*{box-sizing:border-box}body{margin:0;font-family:Segoe UI,Arial,sans-serif;background:var(--bg);color:var(--txt)}
a{text-decoration:none;color:inherit}.top{height:64px;background:linear-gradient(90deg,var(--g),var(--g2));color:#fff;display:flex;align-items:center;justify-content:space-between;padding:0 20px;position:fixed;top:0;left:0;right:0;z-index:30}.brand{font-weight:800;font-size:20px}.ver{font-size:12px;background:#ffffff2b;padding:6px 10px;border-radius:20px}.layout{display:block;min-height:100vh;padding-top:64px}.side{position:fixed;top:64px;left:0;bottom:0;width:220px;background:#153d28;color:#e9fff1;padding:18px 12px;overflow-y:auto;z-index:20}.side a{display:block;padding:12px;border-radius:10px;margin:5px 0}.side a:hover,.side a.on{background:#ffffff18}.main{margin-left:220px;padding:22px;min-height:calc(100vh - 64px)}.grid{display:grid;grid-template-columns:repeat(4,minmax(170px,1fr));gap:14px}.card{background:var(--card);border-radius:16px;padding:18px;box-shadow:0 4px 18px #14271b12}.stat b{font-size:27px;display:block;margin-top:8px}.mut{color:var(--mut)}.actions{display:flex;gap:10px;flex-wrap:wrap;margin:12px 0}.btn{display:inline-block;border:0;border-radius:10px;padding:10px 14px;cursor:pointer;background:var(--g);color:#fff;font-weight:700}.btn.alt{background:#eef4ef;color:var(--g)}.btn.red{background:var(--red)}.btn.blue{background:var(--blue)}.btn.orange{background:var(--orange)}.inline-form{display:inline}.costbox{background:#f7fbf8;border:1px solid #d8e7dc;border-radius:14px;padding:14px;margin-top:12px}
table{width:100%;border-collapse:collapse;background:#fff;border-radius:12px;overflow:hidden}th,td{padding:11px;border-bottom:1px solid #e7ece8;text-align:left;font-size:14px}th{background:#edf5ef}.form{display:grid;grid-template-columns:repeat(2,minmax(180px,1fr));gap:12px}.form label{font-size:13px;font-weight:700}.form input,.form select,.form textarea{width:100%;padding:10px;border:1px solid #cfd9d1;border-radius:9px;margin-top:5px}.full{grid-column:1/-1}.flash{padding:12px;border-radius:10px;background:#e8f7ec;color:#175f34;margin-bottom:14px}.err{background:#fdebea;color:#a52d25}.login{max-width:420px;margin:9vh auto;background:#fff;padding:28px;border-radius:18px;box-shadow:0 10px 35px #1a3b2720}.login h1{margin-top:0}.login input{width:100%;padding:12px;margin:7px 0 14px;border:1px solid #ccd7cf;border-radius:10px}.chart{display:flex;align-items:end;gap:8px;height:190px;padding-top:16px}.bar{flex:1;background:linear-gradient(#2c9660,#176b3a);border-radius:8px 8px 0 0;min-width:18px;position:relative}.bar span{position:absolute;bottom:-24px;font-size:11px;width:100%;text-align:center}.bar i{position:absolute;top:-20px;font-style:normal;font-size:10px;width:100%;text-align:center}.two{display:grid;grid-template-columns:1.2fr .8fr;gap:14px}.taglink{font-weight:800;color:var(--g);text-decoration:underline}.profile{display:grid;grid-template-columns:180px 1fr;gap:18px}.photo{width:180px;height:180px;border-radius:16px;object-fit:cover;background:#e8efe9;display:flex;align-items:center;justify-content:center;font-size:54px}.pill{display:inline-block;padding:6px 10px;border-radius:20px;background:#eaf4ed;margin:3px;font-size:13px}.preg{font-weight:800}.preg.pos{color:var(--g)}.preg.neg{color:var(--red)}.hero{background:linear-gradient(135deg,#123f29,#238a50);color:white;border-radius:20px;padding:24px;margin-bottom:16px;display:flex;justify-content:space-between;gap:16px;align-items:center}.hero h1{margin:0 0 6px}.metric{border-left:5px solid var(--g)}.metric.red{border-left-color:var(--red)}.metric.blue{border-left-color:var(--blue)}.metric.orange{border-left-color:var(--orange)}.gallery{display:grid;grid-template-columns:repeat(auto-fill,minmax(150px,1fr));gap:12px}.gallery figure{margin:0;background:#fff;border-radius:14px;overflow:hidden;box-shadow:0 3px 14px #0001}.gallery img{width:100%;height:150px;object-fit:cover;display:block}.gallery figcaption{padding:8px;font-size:12px}.alertlist{display:grid;gap:8px}.alertitem{padding:10px;border-radius:10px;background:#f3f7f4;border-left:4px solid var(--g)}.mini-chart{display:flex;align-items:end;gap:10px;height:180px;padding:20px 5px 28px}.mini-col{flex:1;display:flex;gap:3px;align-items:end;height:100%;position:relative}.mini-col b{flex:1;border-radius:6px 6px 0 0;background:#2c9660;min-height:2px}.mini-col i{flex:1;border-radius:6px 6px 0 0;background:#d95b4e;min-height:2px}.mini-col span{position:absolute;bottom:-22px;width:100%;text-align:center;font-size:11px}.uploadbox{border:2px dashed #b8c9bd;border-radius:14px;padding:14px;background:#f9fbf9}.camera-note{font-size:12px;color:var(--mut)}.photo-upload-status{display:none;margin-top:8px;padding:10px;border-radius:10px;background:#eef5ef;font-size:12px}.photo-upload-status.on{display:block}.photo-upload-status.error{background:#fdebea;color:#a52d25}.upload-progress{height:9px;background:#dbe6dd;border-radius:99px;overflow:hidden;margin-top:7px}.upload-progress-bar{height:100%;width:0;background:linear-gradient(90deg,var(--g2),var(--blue));transition:width .15s ease}.btn[disabled]{opacity:.62;cursor:not-allowed}
.side .nav-home{font-weight:800}.nav-group{margin:5px 0}.nav-group summary{list-style:none;cursor:pointer;padding:12px;border-radius:10px;font-weight:800;display:flex;align-items:center;justify-content:space-between;user-select:none}.nav-group summary::-webkit-details-marker{display:none}.nav-group summary:hover,.nav-group.open-group summary{background:#ffffff10}.nav-group summary:after{content:"›";font-size:20px;transition:transform .18s ease}.nav-group[open] summary:after{transform:rotate(90deg)}.nav-children{padding:2px 0 4px 10px;border-left:1px solid #ffffff22;margin-left:13px}.side .nav-children a{padding:9px 11px;margin:2px 0;font-size:13px}.menu-toggle{display:none;border:0;background:#ffffff22;color:#fff;border-radius:9px;padding:8px 11px;font-size:20px;cursor:pointer}.top-left{display:flex;align-items:center;gap:10px}

.estrus-decision-badge{display:inline-block;padding:8px 11px;border-radius:12px;font-weight:800;margin:0 7px 7px 0;line-height:1.25}
.estrus-decision-badge small{font-weight:600;opacity:.8}
.estrus-skipped{background:#f1ecff;color:#6542a6}
.estrus-done{background:#e7f6eb;color:#176b3a}


.sortable-insem th{white-space:nowrap}
.sort-head{appearance:none;border:0;background:transparent;color:inherit;font:inherit;font-weight:800;padding:8px 4px;cursor:pointer;display:inline-flex;align-items:center;gap:6px}
.sort-head span{color:#718278;font-size:13px}
.sort-head:hover{color:#08783e}
.sort-head.active{color:#08783e}
.sort-head.active span{color:#08783e}

.license-shell{max-width:760px;margin:55px auto;padding:18px}.license-card{background:#fff;border-radius:22px;padding:28px;box-shadow:0 12px 38px #183c2820;border:1px solid #dfe9e2}
.device-code{font-family:Consolas,monospace;font-size:20px;font-weight:800;letter-spacing:1px;background:#eef6f0;padding:14px;border-radius:12px;word-break:break-all}
.license-ok{display:inline-block;background:#e7f6eb;color:#176b3a;padding:7px 11px;border-radius:999px;font-weight:800}.license-bad{display:inline-block;background:#fff0ee;color:#a92f24;padding:7px 11px;border-radius:999px;font-weight:800}

@media(max-width:760px){.sort-head{padding:6px 2px;font-size:12px;white-space:normal;text-align:left}.sort-head span{font-size:11px}}

/* UX14 Desktop Experience */
.summary-grid{gap:9px!important;align-items:stretch!important}.summary-grid .card{padding:9px 13px!important;min-height:120px;height:120px;border-radius:12px!important;overflow:hidden!important;display:flex!important;flex-direction:column!important;align-items:flex-start!important}.summary-grid .stat b{font-size:21px!important;margin:4px 0!important;line-height:1.05}.summary-grid .metric-icon{font-size:16px!important;margin-bottom:2px!important;flex:0 0 auto}.summary-grid .metric-title{display:block;font-size:14px;line-height:1.16;font-weight:500;min-height:2.32em;max-height:2.32em;overflow:hidden}.summary-grid .metric small{margin-top:auto!important;white-space:nowrap}.dashboard-section-title{margin:15px 0 7px!important}@media(max-width:700px){.summary-grid{grid-auto-rows:116px!important}.summary-grid .card{height:116px!important;min-height:116px!important;padding:10px 11px!important}.summary-grid .metric-title{font-size:13px!important;line-height:1.14;min-height:2.28em;max-height:2.28em}.summary-grid .stat b{font-size:21px!important;margin:3px 0!important}.summary-grid .metric small{font-size:11px!important}}@media(max-width:380px){.summary-grid .card{padding:9px!important}.summary-grid .metric-title{font-size:12.5px!important}}.taglink{text-decoration:none!important;border-bottom:1px dotted #7ba58a}.taglink:hover{border-bottom-style:solid}.animal-tag-btn{display:inline-flex;align-items:center;gap:6px;padding:7px 11px;border:1px solid #b9d4c2;border-radius:9px;background:#eef7f1;color:#086b35!important;font-weight:800;text-decoration:none!important;line-height:1;white-space:nowrap;box-shadow:0 1px 2px #173b2810;transition:.15s ease}.animal-tag-btn:hover{background:#dff1e5;border-color:#6fa985;transform:translateY(-1px);box-shadow:0 3px 8px #173b2818}
.archive-animal-table .animal-tag-btn,.mobile-animal-table .animal-tag-btn{min-width:112px;justify-content:space-between}
.archive-animal-table td:first-child,.mobile-animal-table td:first-child{white-space:nowrap}.animal-tag-btn:before{content:"🐄";font-size:13px}.animal-tag-btn:after{content:"›";font-size:16px;line-height:10px;color:#4f8c67;margin-left:1px}.profile{grid-template-columns:140px 1fr!important;gap:16px!important;padding:16px!important;border:1px solid #dbe5de!important;border-radius:12px!important;box-shadow:0 5px 18px #173b2810!important;background:#fff!important}.profile .photo{width:140px!important;height:140px!important;border-radius:10px!important}.profile h1{font-size:25px!important;margin:0 0 2px!important}.profile h2{font-size:16px!important;margin:0 0 8px!important;color:#617168!important}.profile .pill{border-radius:7px!important;padding:5px 8px!important;margin:2px!important}.profile .quick-metrics{margin-top:8px!important}.profile .costbox{border-radius:10px!important}.profile~.two .card,.profile~.card{border:1px solid #dfe8e2!important;box-shadow:none!important;border-radius:11px!important}@media(max-width:850px){.profile{grid-template-columns:90px 1fr!important}.profile .photo{width:90px!important;height:90px!important}}
@media(max-width:650px){.profile{grid-template-columns:1fr}.photo{width:100%;height:220px}}
/* V3.7.6 Besi Kârlılık */
.perf-hero{background:linear-gradient(135deg,#173f2b,#245f3e);color:#fff;border-radius:24px;padding:24px 26px;display:flex;align-items:center;justify-content:space-between;gap:18px;box-shadow:0 14px 34px rgba(22,72,45,.16)}
.perf-hero h1{margin:0 0 6px;font-size:28px}.perf-hero p{margin:0;color:#dcece2}.perf-hero .btn{background:#fff;color:#18492f}
.perf-tabs{display:flex;gap:8px;flex-wrap:wrap;margin:16px 0}.perf-tab{display:inline-flex;padding:10px 15px;border-radius:999px;background:#edf4ef;color:#355747;font-weight:800;text-decoration:none;border:1px solid #dce8df}.perf-tab.active{background:#1f6b42;color:#fff;border-color:#1f6b42}
/* V3.9.15 unified ration workspace */
.target-workspace{margin-top:10px}.target-controlbar{display:flex;align-items:flex-end;gap:10px;justify-content:space-between;flex-wrap:wrap;padding:8px 10px;background:#fff;border:1px solid #dce8df;border-radius:12px}.target-controlbar .target-head{min-width:190px}.target-controlbar .target-head h3{margin:0;font-size:16px}.target-controlbar .target-context{margin-top:2px;font-size:12px}.target-controlbar .target-form{display:flex;align-items:flex-end;gap:7px;flex:1;justify-content:flex-end;flex-wrap:wrap}.target-controlbar .target-form label{font-size:11px;min-width:110px}.target-controlbar .target-form input,.target-controlbar .target-form select{padding:6px 8px;min-height:34px}.target-compare-sticky{position:sticky;top:72px;z-index:25;margin-top:8px;padding:7px 8px;background:rgba(248,251,249,.97);backdrop-filter:blur(7px);border:1px solid #bcd8c5;border-radius:12px;box-shadow:0 5px 18px rgba(28,76,49,.08)}.target-compare-title{display:flex;align-items:center;justify-content:space-between;gap:10px;margin:0 2px 5px;font-size:12px}.target-compare-title span{color:#617269}.nutri-mini-grid{grid-template-columns:repeat(8,minmax(92px,1fr))!important;gap:5px!important}.nutri-mini{padding:6px 7px!important;min-height:64px!important}.nutri-mini span{font-size:10px!important}.nutri-mini b{font-size:16px!important;margin-top:1px!important}.nutri-mini small,.nutri-mini i,.nutri-mini em{font-size:9.5px!important;line-height:1.15!important}.compact-changebar{margin:4px 0 6px!important;padding:0 2px}.ration-workbench-table th,.ration-workbench-table td{padding:6px 7px!important}.ration-workbench-table{font-size:12px}.ration-workbench-table tbody tr:nth-child(even){background:#f4f8f5}.ration-workbench-table tbody tr:nth-child(odd){background:#fff}.ration-workbench-table tbody tr:hover{background:#eaf5ee}.quick-feed-card:not([open]){display:none}.quick-feed-card[open]{display:block;position:fixed;z-index:80;left:max(12px,calc(50% - 430px));right:max(12px,calc(50% - 430px));top:8vh;max-height:84vh;overflow:auto;margin:0!important;box-shadow:0 24px 70px rgba(0,0,0,.28);border:1px solid #b9d7c3!important}.quick-feed-card[open]::before{content:'';position:fixed;inset:0;background:rgba(17,35,25,.38);z-index:-1}.quick-feed-head{position:sticky;top:0;background:#fff;z-index:2;padding:4px 0}.ration-savebar{padding:7px!important}.workbench-page-head{display:flex;align-items:flex-start;justify-content:space-between;gap:12px;margin-bottom:8px}.workbench-page-head h1{margin-bottom:3px}
.ration-picker-grid{gap:8px!important}.ration-picker-card{min-height:88px!important;padding:10px 12px!important}
@media(max-width:1100px){.nutri-mini-grid{grid-template-columns:repeat(4,minmax(100px,1fr))!important}.target-compare-sticky{top:72px}.target-controlbar{align-items:stretch}.target-controlbar .target-form{justify-content:flex-start}}
@media(max-width:650px){.nutri-mini-grid{grid-template-columns:repeat(2,minmax(110px,1fr))!important}.target-compare-title span{display:none}.target-controlbar .target-form label{min-width:calc(50% - 6px);flex:1}.target-compare-sticky{padding:6px}.nutri-mini{min-height:58px!important}.quick-feed-card[open]{left:8px;right:8px;top:4vh;max-height:92vh}}
.ration-stepper{display:flex;align-items:center;gap:6px;white-space:nowrap}.ration-qty{width:92px;padding:8px;border:1px solid #bfd0c4;border-radius:9px;text-align:center;font-weight:800}.ration-savebar{position:sticky;bottom:8px;background:#fffffff2;padding:10px;text-align:right;border-top:1px solid #e1ebe4;z-index:9}#ration-workbench{scroll-margin-top:82px}.ration-section-collapse details>summary{cursor:pointer;font-weight:800}.ration-live{position:sticky;top:42px;z-index:7;background:#f8fbf9;border:1px solid #cfe3d5;border-radius:12px;padding:10px;margin:10px 0}.ration-live-grid{display:grid;grid-template-columns:repeat(7,minmax(100px,1fr));gap:8px}.ration-live-metric{background:#fff;border:1px solid #e0e8e2;border-radius:9px;padding:8px}.ration-live-metric span{display:block;color:#607067;font-size:12px}.ration-live-metric b{display:block;margin-top:3px}.ration-live-metric small{display:block;margin-top:2px}.ration-dirty{outline:2px solid #f0a126;background:#fff9ec}.ration-changebar{display:flex;align-items:center;justify-content:space-between;gap:10px;flex-wrap:wrap;margin-top:8px}.ration-dirty-text{font-weight:800;color:#9a5b00}@media(max-width:1000px){.ration-live-grid{grid-template-columns:repeat(2,minmax(120px,1fr))}.ration-live{position:static}}
.perf-filter-card{border:1px solid #dce8df;background:linear-gradient(180deg,#fff,#f8fbf9)}.perf-filter-head{display:flex;align-items:center;justify-content:space-between;gap:12px;margin-bottom:14px}.perf-filter-head h2{margin:0}
.perf-filter-grid{display:grid;grid-template-columns:1.2fr repeat(3,1fr);gap:10px;align-items:end}.perf-filter-grid label{display:flex;flex-direction:column;gap:5px;font-weight:800;font-size:13px}.perf-filter-actions{display:flex;gap:8px}
.perf-summary{display:grid;grid-template-columns:repeat(5,minmax(0,1fr));gap:12px;margin-top:14px}.perf-summary small{display:block;margin-top:7px;color:#6d8075}.perf-table-wrap{overflow:auto;border:1px solid #e1e9e4;border-radius:16px}.performance-table{min-width:1450px;margin:0}.performance-table th{position:sticky;top:0;background:#edf5ef;z-index:1;white-space:nowrap}.performance-table td{white-space:nowrap}.performance-table tbody tr:hover{background:#f7fbf8}
.profit-pill{display:inline-flex;padding:6px 10px;border-radius:999px;font-weight:900}.profit-pill.good{background:#e5f6eb;color:#14723d}.profit-pill.bad{background:#fff0ed;color:#b33a2e}.profit-pill.wait{background:#eef2f0;color:#65756d}
@media(max-width:1100px){.perf-filter-grid{grid-template-columns:repeat(2,1fr)}.perf-summary{grid-template-columns:repeat(2,1fr)}}@media(max-width:650px){.perf-hero{padding:19px;align-items:flex-start;flex-direction:column}.perf-filter-grid{grid-template-columns:1fr}.perf-summary{grid-template-columns:1fr 1fr}.perf-summary .card:last-child{grid-column:1/-1}.perf-tabs{overflow-x:auto;flex-wrap:nowrap}.perf-tab{white-space:nowrap}}
@media(max-width:900px){.menu-toggle{display:inline-block}.side{transform:translateX(-105%);transition:transform .2s ease;width:260px;box-shadow:8px 0 24px #0003}.side.mobile-open{transform:translateX(0)}.main{margin-left:0;padding-top:18px}.grid{grid-template-columns:repeat(2,1fr)}.two{grid-template-columns:1fr}}@media(max-width:560px){.grid,.form{grid-template-columns:1fr}.main{padding:12px}.top{padding:0 12px}.brand{font-size:17px}}




.finance-toolbar{display:flex;gap:10px;align-items:end;flex-wrap:wrap}
.finance-toolbar label{display:flex;flex-direction:column;gap:5px;font-weight:700}
.finance-table-wrap{overflow-x:auto;border-radius:14px}
.finance-table{width:100%;border-collapse:collapse;min-width:900px}
.finance-table th{white-space:nowrap;background:#edf5ef}
.finance-table td{vertical-align:middle}
.finance-table .finance-actions{display:flex;gap:6px;align-items:center;white-space:nowrap}
.finance-table .finance-actions form{margin:0}
.finance-table .btn{padding:8px 12px}
/* HOTFIX 6.16: Finans tablosu kompakt satırlar + işlem butonları tek hizada */
.finance-table td{padding:8px 10px!important;line-height:1.22!important;vertical-align:middle!important}
.finance-table th{padding:9px 10px!important}
.finance-table td:last-child,.finance-table th:last-child{width:158px;min-width:158px}
.finance-table .finance-actions{display:flex!important;flex-direction:row!important;flex-wrap:nowrap!important;gap:6px!important;align-items:center!important;justify-content:flex-start!important}
.finance-table .finance-actions form{display:block!important;margin:0!important;flex:0 0 auto!important}
.finance-table .finance-actions .btn{min-height:34px!important;height:34px!important;padding:6px 9px!important;font-size:12px!important;line-height:1!important;white-space:nowrap!important}

.dashboard-editbar{display:flex;justify-content:flex-end;gap:10px;margin:10px 0 14px}
.dashboard-slot{position:relative;min-width:0}
.dashboard-slot.editing{padding-top:0}
.dashboard-slot.editing>.summary-link{outline:2px dashed #83ad90;outline-offset:3px}
.dashboard-slot-plus{position:absolute;right:-7px;top:-9px;z-index:4;width:34px;height:34px;border-radius:50%;border:3px solid #fff;background:#167a43;color:#fff;font-size:23px;line-height:27px;font-weight:900;cursor:pointer;box-shadow:0 4px 12px #173b2828;display:flex;align-items:center;justify-content:center}
.dashboard-slot-plus:hover{transform:scale(1.08);background:#0f6838}
.dashboard-empty-slot{min-height:145px;border:2px dashed #91b59c;border-radius:18px;display:flex;align-items:center;justify-content:center;background:linear-gradient(135deg,#f8fcf9,#edf7f0);color:#176b3a;cursor:pointer;transition:.18s}
.dashboard-empty-slot:hover{border-color:#176b3a;background:#eaf7ee;transform:translateY(-2px)}
.dashboard-empty-slot .plus-icon{width:52px;height:52px;border-radius:50%;background:#176b3a;color:#fff;display:flex;align-items:center;justify-content:center;font-size:35px;margin:0 auto 8px;box-shadow:0 7px 18px #176b3a2c}
.dashboard-empty-slot small{display:block;font-size:14px;color:#607869;font-weight:700;text-align:center}
.dashboard-picker-backdrop{display:none;position:fixed;inset:0;background:#10271b8c;z-index:100;align-items:center;justify-content:center;padding:18px;backdrop-filter:blur(3px)}
.dashboard-picker-backdrop.open{display:flex}
.dashboard-picker{width:min(760px,96vw);max-height:88vh;overflow:auto;background:#fff;border-radius:24px;padding:22px;box-shadow:0 25px 80px #07150d55}
.dashboard-picker-head{display:flex;justify-content:space-between;gap:15px;align-items:flex-start;margin-bottom:16px}
.dashboard-picker-head h2{margin:0;color:#173c29}.dashboard-picker-head p{margin:4px 0 0;color:#718276}
.dashboard-picker-close{border:0;background:#edf4ef;width:42px;height:42px;border-radius:50%;font-size:22px;cursor:pointer}
.dashboard-card-gallery{display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:11px}
.dashboard-card-choice{display:flex;align-items:center;gap:12px;text-align:left;border:2px solid #e0eae3;background:#fbfdfb;border-radius:16px;padding:14px;cursor:pointer;font:inherit;color:#183a29}
.dashboard-card-choice:hover,.dashboard-card-choice.active{border-color:#248952;background:#eaf7ee;transform:translateY(-1px)}
.dashboard-card-choice .choice-icon{width:44px;height:44px;border-radius:13px;background:#edf6ef;display:flex;align-items:center;justify-content:center;font-size:24px}
.dashboard-card-choice b{display:block}.dashboard-card-choice small{color:#6f8174}
.dashboard-picker-footer{display:flex;gap:9px;justify-content:space-between;align-items:center;margin-top:16px;flex-wrap:wrap}
@media(max-width:600px){.dashboard-card-gallery{grid-template-columns:1fr}.dashboard-picker{padding:16px;border-radius:18px}.dashboard-slot-plus{right:-3px;top:-6px}}
@media(max-width:700px){
 .finance-toolbar{align-items:stretch}.finance-toolbar>*{width:100%}.finance-toolbar input,.finance-toolbar select,.finance-toolbar .btn{width:100%}
 .finance-table{min-width:760px}
}

.bulk-animal-box{display:none}
.bulk-picker{border:1px solid #d8e5dc;border-radius:16px;background:#fbfdfb;padding:14px}
.bulk-picker-head{display:flex;gap:10px;align-items:center;justify-content:space-between;flex-wrap:wrap;margin-bottom:10px}
.bulk-search{max-width:420px;width:100%}
.bulk-list{max-height:300px;overflow:auto;border:1px solid #dfe9e2;border-radius:12px;background:white}
.bulk-row{display:grid;grid-template-columns:34px minmax(130px,1fr) minmax(110px,1fr);gap:10px;align-items:center;padding:10px 12px;border-bottom:1px solid #edf2ee;cursor:pointer}
.bulk-row:last-child{border-bottom:0}.bulk-row:hover{background:#f5faf6}.bulk-row.selected{background:#eaf7ee}
.bulk-row input{width:20px;height:20px}.bulk-row .tag{font-weight:800;color:#176b3a}.bulk-row .nick{color:#6d8074;font-size:13px}
.bulk-summary{display:flex;gap:10px;flex-wrap:wrap;align-items:center;margin-top:12px}
.bulk-summary .pill{background:#edf6ef;padding:9px 12px;border-radius:999px}
.bulk-selected-preview{font-size:13px;color:#5f7466;margin-top:8px}
.finance-savebar{display:flex;gap:10px;align-items:center;flex-wrap:wrap}
.finance-savebar .btn{min-width:180px}

.farm-profile-head{display:flex;gap:18px;align-items:center;flex-wrap:wrap}
.farm-logo-preview{width:120px;height:120px;border-radius:18px;object-fit:contain;background:#f4f7f5;border:1px solid #dbe5de;padding:8px}
.farm-logo-placeholder{width:120px;height:120px;border-radius:18px;background:#eaf4ed;display:flex;align-items:center;justify-content:center;font-size:50px}
.farm-hero{display:flex;align-items:center;gap:16px}.farm-hero-logo{width:76px;height:76px;border-radius:16px;object-fit:contain;background:#ffffff18;padding:5px}
@media(max-width:560px){.farm-hero{align-items:flex-start}.farm-hero-logo{width:58px;height:58px}}

.pro-form-head{display:flex;justify-content:space-between;gap:12px;align-items:center;flex-wrap:wrap;margin-bottom:14px}
.type-chip{padding:7px 11px;border-radius:999px;background:#eaf4ed;color:var(--g);font-weight:800}
.livebox{display:flex;gap:8px;align-items:center;margin:12px 0}
.livebox input{flex:1;max-width:540px;padding:11px;border:1px solid #cfd9d1;border-radius:10px}
.empty-state{display:none;padding:16px;text-align:center;color:var(--mut)}
.quick-metrics{display:grid;grid-template-columns:repeat(4,minmax(120px,1fr));gap:10px;margin-top:12px}
.quick-metrics .pill{display:block;text-align:center;padding:12px}
@media(max-width:700px){.quick-metrics{grid-template-columns:repeat(2,1fr)}}

.summary-link{display:block;transition:transform .15s ease,box-shadow .15s ease}.summary-link:hover{transform:translateY(-2px);box-shadow:0 8px 24px #14271b20}.summary-link:focus{outline:2px solid var(--g2);outline-offset:2px}.summary-grid{grid-template-columns:repeat(4,minmax(145px,1fr))}.summary-grid .card{padding:15px}.summary-grid .stat b{font-size:24px}.summary-grid .metric-icon{font-size:21px;margin-bottom:6px}@media(max-width:1250px){.summary-grid{grid-template-columns:repeat(3,1fr)}}@media(max-width:700px){.summary-grid{grid-template-columns:repeat(2,1fr)}}
.metric-icon{font-size:24px;display:block;margin-bottom:8px}.metric small{display:block;margin-top:5px;color:var(--mut);font-size:12px;font-weight:600}.metric.green{border-left-color:#2c9660}.metric.purple{border-left-color:#7b5cc7}.metric.teal{border-left-color:#178c91}.cost-visual{display:grid;grid-template-columns:.9fr 1.1fr;gap:18px;align-items:center}.donut{width:190px;height:190px;border-radius:50%;margin:auto;position:relative;background:conic-gradient(var(--blue) 0 var(--purchase-pct),var(--orange) var(--purchase-pct) 100%)}.donut:after{content:"";position:absolute;inset:31px;background:var(--card);border-radius:50%}.donut-center{position:absolute;inset:0;display:flex;z-index:2;align-items:center;justify-content:center;flex-direction:column;text-align:center}.donut-center b{font-size:20px}.legend-row{display:grid;grid-template-columns:14px 1fr auto;gap:8px;align-items:center;margin:11px 0}.legend-dot{width:12px;height:12px;border-radius:4px}.dot-blue{background:var(--blue)}.dot-orange{background:var(--orange)}.progress-list{display:grid;gap:12px}.progress-item{display:grid;gap:5px}.progress-head{display:flex;justify-content:space-between;gap:10px;font-size:13px}.progress-track{height:11px;background:#e8eee9;border-radius:99px;overflow:hidden}.progress-fill{height:100%;border-radius:99px;background:linear-gradient(90deg,var(--g2),var(--blue))}.health-group{border:1px solid #dfe9e2;border-radius:14px;margin:10px 0;background:#fbfdfb;overflow:hidden}.health-group summary{display:flex;justify-content:space-between;gap:12px;align-items:center;padding:14px 16px;cursor:pointer;list-style:none}.health-group summary::-webkit-details-marker{display:none}.health-group[open] summary{background:#eef7f0;border-bottom:1px solid #dfe9e2}.dash-fold{margin-top:14px}.dash-fold>summary{list-style:none;cursor:pointer;display:flex;justify-content:space-between;align-items:center;padding:15px 18px;background:#fff;border:1px solid #e0e9e2;border-radius:15px;font-size:18px;font-weight:900;color:#183c2a;box-shadow:0 3px 12px #173b2810}.dash-fold>summary::-webkit-details-marker{display:none}.dash-fold[open]>summary{border-radius:15px 15px 0 0}.dash-fold-content{padding:14px;background:#f8fbf9;border:1px solid #e0e9e2;border-top:0;border-radius:0 0 15px 15px}.dashboard-section-title{display:flex;justify-content:space-between;align-items:end;gap:12px;margin:22px 0 10px}.dashboard-section-title h2{margin:0}.dashboard-section-title span{color:var(--mut);font-size:13px}@media(max-width:760px){.cost-visual{grid-template-columns:1fr}.donut{width:165px;height:165px}}

.performance-card{border-left:5px solid var(--blue)}.status-good{color:#176b3a;background:#e8f7ec}.status-watch{color:#8a5a00;background:#fff4d6}.status-low{color:#a52d25;background:#fdebea}.status-none{color:var(--mut);background:#f0f3f1}.perf-badge{display:inline-block;padding:6px 10px;border-radius:999px;font-weight:800;font-size:12px}.weight-chart{width:100%;height:240px;border-radius:14px;background:linear-gradient(180deg,#f7fbf8,#fff);border:1px solid #e0e9e2}.weight-chart text{font-family:Segoe UI,Arial,sans-serif;font-size:11px;fill:#6b7b70}.weight-chart .axis{stroke:#bccac0;stroke-width:1}.weight-chart .gridline{stroke:#e2e9e4;stroke-width:1}.weight-chart .trend{fill:none;stroke:var(--blue);stroke-width:4;stroke-linecap:round;stroke-linejoin:round}.weight-chart .point{fill:var(--g);stroke:#fff;stroke-width:3}.warning-panel{border-left:5px solid var(--red);background:#fff7f6}.performance-table tr.low-row td{background:#fff3f2}.performance-table tr.watch-row td{background:#fff9e8}.performance-table tr.good-row td{background:#f2fbf5}.setting-box{background:#f7fbf8;border:1px solid #d8e7dc;border-radius:14px;padding:16px}
.insem-head{display:flex;justify-content:space-between;align-items:center;gap:14px;flex-wrap:wrap}.insem-head h1{margin-bottom:4px}.insem-search{display:flex;gap:8px;align-items:center;min-width:min(100%,460px)}.insem-search input{flex:1;padding:11px;border:1px solid #cfd9d1;border-radius:10px}.insem-stats{grid-template-columns:repeat(4,minmax(150px,1fr));margin:14px 0}.insem-table details summary{cursor:pointer;color:var(--g);font-weight:800}.insem-history{margin:10px 0 2px;background:#f8fbf9;border-radius:10px}.status-badge{display:inline-block;padding:5px 9px;border-radius:999px;font-weight:800;font-size:12px}.status-preg{background:#e8f7ec;color:#176b3a}.status-wait{background:#eaf2ff;color:#2e6fc2}.status-neg{background:#fdebea;color:#a52d25}.status-unknown{background:#fff4d6;color:#8a5a00}.animal-picker-note{font-size:12px;color:var(--mut);margin-top:5px}.attempt-preview{display:flex;align-items:center;min-height:42px;padding:10px;border:1px solid #cfd9d1;border-radius:9px;margin-top:5px;background:#f7fbf8;font-weight:800}.future-warning{font-size:12px;color:var(--red);display:none;margin-top:5px}.row-actions{display:flex;gap:6px;flex-wrap:wrap}.compact-btn{padding:7px 10px;font-size:12px}.insem-empty{display:none;padding:18px;text-align:center;color:var(--mut)}.animal-picker{position:relative}.animal-suggestions{display:none;position:absolute;left:0;right:0;top:100%;z-index:60;background:#fff;border:1px solid #cfd9d1;border-radius:10px;box-shadow:0 10px 26px #14271b24;max-height:260px;overflow-y:auto;margin-top:4px}.animal-suggestions.open{display:block}.animal-suggestion{display:block;width:100%;border:0;background:#fff;text-align:left;padding:11px 12px;cursor:pointer;border-bottom:1px solid #edf1ee;color:var(--txt)}.animal-suggestion:last-child{border-bottom:0}.animal-suggestion:hover,.animal-suggestion:focus{background:#edf7f0;outline:none}.animal-suggestion b{display:block;color:var(--g)}.animal-suggestion span{font-size:12px;color:var(--mut)}
@media(max-width:800px){.insem-stats{grid-template-columns:repeat(2,1fr)}.insem-search{min-width:100%}.insem-head{align-items:stretch}.insem-search .btn{flex:0 0 auto}}
@media(max-width:560px){.top-user{font-size:12px;max-width:46vw;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}.ver{display:none}.card{padding:14px;border-radius:13px}.btn{min-height:44px;padding:11px 13px}.form input,.form select,.form textarea,.livebox input,.insem-search input{min-height:44px;font-size:16px}.insem-stats{grid-template-columns:1fr 1fr;gap:8px}.insem-stats .card{padding:12px}.insem-stats .stat b{font-size:22px}.insem-head h1{font-size:25px}.insem-search{flex-direction:column;align-items:stretch}.insem-search .btn{width:100%}.animal-suggestions{max-height:220px}.row-actions .btn{width:100%;text-align:center}.row-actions .inline-form{width:100%}.row-actions .inline-form .btn{width:100%}table{min-width:620px}.insem-table{min-width:680px}.insem-history{min-width:650px}.card{overflow-x:auto}.animal-picker{overflow:visible}.login{margin:4vh 12px;padding:20px}.actions>input,.actions>select,.actions>label{max-width:100%}.actions{align-items:stretch}.actions .btn{flex:1 1 auto}}


/* V3.1.9 — gerçek mobil arayüz */
.business-summary-grid{grid-template-columns:repeat(3,minmax(0,1fr))}
.mobile-animal-table .inline-form{display:inline-block}
@media(max-width:700px){
  .top{height:58px;padding:0 12px}.layout{padding-top:58px}.side{top:58px}.main{padding:14px 12px 24px}
  .top-left{gap:8px}.menu-toggle{font-size:22px;padding:9px 12px;border-radius:12px}.brand{font-size:22px}.top-user{font-size:13px;max-width:42vw}
  h1{font-size:30px;line-height:1.08;margin:16px 0 18px}h2{font-size:21px}.card{padding:14px;border-radius:18px}
  .livebox{gap:8px;margin:8px 0 14px}.livebox input{max-width:none;width:100%;min-width:0;height:46px;font-size:16px;padding:10px 42px 10px 13px}.live-clear{padding:10px 12px;min-height:46px}
  .business-summary-grid{grid-template-columns:1fr 1fr!important;gap:10px}.business-summary-grid .card{min-width:0;padding:13px}.business-summary-grid .card:last-child{grid-column:1/-1}.business-summary-grid .stat{font-size:13px}.business-summary-grid .stat b{font-size:23px;line-height:1.08;overflow-wrap:anywhere}
  .summary-grid{grid-template-columns:1fr 1fr;gap:10px}.summary-grid .card{padding:12px}.summary-grid .stat b{font-size:22px}
  .actions{gap:8px}.actions .btn{flex:1 1 auto;text-align:center}
  .card:has(> .mobile-animal-table){padding:8px;background:transparent;box-shadow:none;overflow:visible}
  .mobile-animal-table,.mobile-animal-table tbody,.mobile-animal-table tr,.mobile-animal-table td{display:block;width:100%!important;min-width:0!important}
  .mobile-animal-table{background:transparent;min-width:0!important}.mobile-animal-table thead{display:none}
  .mobile-animal-table tr.data-row{background:#fff;border:1px solid #e2ebe5;border-radius:16px;margin:0 0 10px;padding:14px;box-shadow:0 3px 13px #14271b0d}
  .mobile-animal-table tr.data-row td{border:0;padding:4px 0;font-size:14px;display:grid;grid-template-columns:104px 1fr;gap:10px;align-items:center}
  .mobile-animal-table tr.data-row td:before{font-size:12px;font-weight:800;color:var(--mut)}
  .mobile-animal-table .taglink{font-size:18px}.mobile-animal-table .btn{min-height:40px;padding:9px 11px;font-size:13px}
  .mobile-animal-table tr.data-row td:last-child{display:flex;gap:7px;flex-wrap:wrap;padding-top:10px;margin-top:6px;border-top:1px solid #edf1ee}
  .mobile-animal-table tr.data-row td:last-child:before{display:none}.mobile-animal-table tr.data-row td:last-child .btn{flex:1 1 auto;text-align:center}.mobile-animal-table tr.data-row td:last-child .inline-form{flex:0 0 auto}.mobile-animal-table tr.data-row td:last-child .inline-form .btn{width:auto}
  .all-animal-table td:nth-child(1):before{content:'Küpe'}.all-animal-table td:nth-child(2):before{content:'Takma Ad'}.all-animal-table td:nth-child(3):before{content:'Tür'}.all-animal-table td:nth-child(4):before{content:'Irk'}.all-animal-table td:nth-child(5):before{content:'Padok'}.all-animal-table td:nth-child(6):before{content:'Yaş'}
    .female-table td:nth-child(1):before{content:'Küpe'}.female-table td:nth-child(2):before{content:'Takma Ad'}.female-table td:nth-child(3):before{content:'Cinsiyet'}.female-table td:nth-child(4):before{content:'Irk'}.female-table td:nth-child(5):before{content:'Padok'}.female-table td:nth-child(6):before{content:'Yaş'}
  .male-table td:nth-child(1):before{content:'Küpe'}.male-table td:nth-child(2):before{content:'Takma Ad'}.male-table td:nth-child(3):before{content:'Irk'}.male-table td:nth-child(4):before{content:'Padok'}.male-table td:nth-child(5):before{content:'Bizde Kalma'}.male-table td:nth-child(6):before{content:'Alış'}.male-table td:nth-child(7):before{content:'Anlık Maliyet'}.male-table td:nth-child(8):before{content:'Hedef Kâr'}
  .calf-table td:nth-child(1):before{content:'Küpe'}.calf-table td:nth-child(2):before{content:'Anne'}.calf-table td:nth-child(3):before{content:'Baba'}.calf-table td:nth-child(4):before{content:'Doğum'}.calf-table td:nth-child(5):before{content:'Yaş'}.calf-table td:nth-child(6):before{content:'Cinsiyet'}
  .insem-head{gap:12px}.insem-head h1{font-size:28px}.insem-search{width:100%;flex-direction:row}.insem-search input{min-width:0;width:100%}.insem-search .live-clear{width:auto!important;flex:0 0 auto}
  .insem-stats{grid-template-columns:1fr 1fr!important;gap:9px}.insem-stats .card{min-width:0;padding:12px}.insem-stats .stat{font-size:13px}.insem-stats .stat b{font-size:22px}.insem-stats small{font-size:11px}
  .insem-table,.insem-table tbody,.insem-table tr.data-row,.insem-table tr.data-row>td{display:block;width:100%;min-width:0!important}.insem-table{background:transparent}.insem-table>thead{display:none}
  .insem-table tr.data-row{background:#fff;border:1px solid #e1ebe4;border-radius:16px;margin-bottom:10px;padding:14px;box-shadow:0 3px 13px #14271b0d}
  .insem-table tr.data-row>td{border:0;padding:5px 0;display:grid;grid-template-columns:112px 1fr;gap:9px;align-items:center}
  .insem-table tr.data-row>td:before{font-size:12px;font-weight:800;color:var(--mut)}
  .insem-table tr.data-row>td:nth-child(1):before{content:'Hayvan'}.insem-table tr.data-row>td:nth-child(2):before{content:'Son Deneme'}.insem-table tr.data-row>td:nth-child(3):before{content:'Tohumlama'}.insem-table tr.data-row>td:nth-child(4):before{content:'Durum'}.insem-table tr.data-row>td:nth-child(5):before{content:'Tahmini Doğum'}.insem-table tr.data-row>td:nth-child(6){display:block;border-top:1px solid #edf1ee;margin-top:7px;padding-top:10px}.insem-table tr.data-row>td:nth-child(6):before{display:none}
  .insem-table details summary{padding:8px 0;font-weight:800;color:var(--g)}.insem-history{min-width:570px!important}
}

.estrus-actions{display:flex;gap:8px;flex-wrap:wrap;margin-top:9px}.estrus-actions form{margin:0}.estrus-window-now{background:#fff6e8;border-left-color:#e27b1f}.estrus-window-next{background:#f3f7f4;border-left-color:#238a50}
@media(max-width:700px){
  .estrus-table,.estrus-table tbody,.estrus-table tr.data-row,.estrus-table tr.data-row>td{display:block;width:100%;min-width:0!important}.estrus-table{background:transparent}.estrus-table>thead{display:none}
  .estrus-table tr.data-row{background:#fff;border:1px solid #e1ebe4;border-radius:16px;margin-bottom:10px;padding:14px;box-shadow:0 3px 13px #14271b0d}
  .estrus-table tr.data-row>td{border:0;padding:5px 0;display:grid;grid-template-columns:112px 1fr;gap:9px;align-items:start}
  .estrus-table tr.data-row>td:before{font-size:12px;font-weight:800;color:var(--mut)}
  .estrus-table tr.data-row>td:nth-child(1):before{content:'Hayvan'}.estrus-table tr.data-row>td:nth-child(2):before{content:'Gözlem'}.estrus-table tr.data-row>td:nth-child(3):before{content:'Belirtiler'}.estrus-table tr.data-row>td:nth-child(4):before{content:'Pencere'}.estrus-table tr.data-row>td:nth-child(5):before{content:'21. Gün'}.estrus-table tr.data-row>td:nth-child(6):before{content:'Not'}
  .estrus-table tr.data-row>td:last-child{display:flex;gap:7px;flex-wrap:wrap;border-top:1px solid #edf1ee;margin-top:7px;padding-top:10px}.estrus-table tr.data-row>td:last-child:before{display:none}
}

@media(max-width:430px){
  .brand{font-size:20px}.top-user{font-size:12px}.main{padding-left:10px;padding-right:10px}h1{font-size:28px}
  .live-clear{font-size:0;width:46px!important}.live-clear:after{content:'×';font-size:24px;line-height:1}
  .business-summary-grid .stat b{font-size:21px}.mobile-animal-table tr.data-row td{grid-template-columns:92px 1fr}
}

/* V3.7.6 Görsel Standartlar */
.btn{min-height:40px;border-radius:11px;padding:9px 14px;display:inline-flex;align-items:center;justify-content:center;gap:6px;line-height:1.15;text-decoration:none;transition:transform .12s ease,filter .12s ease,box-shadow .12s ease}
.btn:hover{filter:brightness(.97);box-shadow:0 5px 14px rgba(20,70,43,.10)}
.btn:active{transform:translateY(1px)}
.btn.red,.btn.danger{background:#d83a2e;color:#fff}.btn.secondary{background:#eef4ef;color:#176b3a}.compact-btn{min-height:34px;padding:7px 10px;border-radius:9px}
.row-actions,.finance-actions,.actions{gap:8px;align-items:center;flex-wrap:wrap}.inline-form{margin:0}
.finance-primary-actions{display:flex;align-items:center;gap:14px;margin:14px 0 18px;flex-wrap:wrap}.finance-new-btn{font-size:15px;padding:13px 20px;border-radius:12px;box-shadow:0 8px 20px rgba(15,112,61,.18)}
.finance-drawer-backdrop{position:fixed;inset:0;background:rgba(13,38,26,.38);backdrop-filter:blur(2px);z-index:1090;display:none}.finance-drawer-backdrop.open{display:block}.finance-drawer{position:fixed;top:0;right:0;height:100vh;width:min(760px,94vw);background:#f5f8f6;z-index:1100;transform:translateX(105%);transition:transform .22s ease;box-shadow:-20px 0 50px rgba(12,45,28,.20);display:flex;flex-direction:column}.finance-drawer.open{transform:translateX(0)}.finance-drawer-head{display:flex;justify-content:space-between;gap:15px;align-items:flex-start;padding:22px 24px 17px;background:#fff;border-bottom:1px solid #dce8df;position:sticky;top:0;z-index:2}.finance-drawer-close{width:42px;height:42px;border:0;border-radius:50%;font-size:29px;background:#eaf3ed;color:#145b34;cursor:pointer}.finance-drawer-body{padding:18px 20px 35px;overflow:auto}.finance-entry-card{margin:0!important;box-shadow:none!important}.finance-drawer .bulk-list{max-height:310px;overflow:auto}
@media(max-width:700px){.finance-primary-actions{position:sticky;top:62px;z-index:15;background:#f4f7f5;padding:8px 0;margin:5px 0 12px}.finance-new-btn{width:100%}.finance-primary-actions .mut{display:none}.finance-drawer{width:100vw}.finance-drawer-head{padding:16px}.finance-drawer-body{padding:12px 12px 28px}.finance-drawer .bulk-list{max-height:42vh}}
.finance-filter-card{background:linear-gradient(180deg,#fff,#f8fbf9);border:1px solid #dce8df}
.finance-filter-title{display:flex;align-items:center;justify-content:space-between;gap:12px;margin-bottom:14px}.finance-filter-title h2{margin:0 0 3px}
.filter-count-pill{display:inline-flex;padding:7px 11px;border-radius:999px;background:#edf5ef;color:#315943;font-weight:800;white-space:nowrap}
.finance-toolbar-modern{display:grid;grid-template-columns:1fr 1fr .9fr 1.15fr auto;gap:10px;align-items:end}
.finance-toolbar-modern label{display:flex;flex-direction:column;gap:5px;font-size:13px;font-weight:800}.finance-toolbar-modern label>span{color:#415f50}
.finance-toolbar-modern input,.finance-toolbar-modern select{min-height:42px;border:1px solid #ccd9d0;border-radius:10px;background:#fff;padding:8px 10px}
.finance-filter-actions{display:flex;gap:7px;align-items:center;white-space:nowrap}.export-btn{background:#2f74c7;color:#fff}
@media(max-width:1100px){.finance-toolbar-modern{grid-template-columns:repeat(2,minmax(0,1fr))}.finance-filter-actions{grid-column:1/-1}}
@media(max-width:600px){.finance-filter-title{align-items:flex-start;flex-direction:column}.finance-toolbar-modern{grid-template-columns:1fr}.finance-filter-actions{grid-column:auto;display:grid;grid-template-columns:1fr 1fr}.finance-filter-actions .export-btn{grid-column:1/-1}.btn{min-height:44px}}

/* UX14.4 — Professional Desktop UI polish */
:root{--panel:#ffffff;--line:#dce6df;--line-strong:#c9d8ce;--soft:#f7faf8;--soft-green:#edf6f0;--shadow-soft:0 3px 12px rgba(20,55,35,.055);--shadow-hover:0 7px 18px rgba(20,55,35,.09)}
body{background:#f2f5f3}
.top{box-shadow:0 1px 0 rgba(255,255,255,.12),0 4px 16px rgba(14,55,32,.10)}
.brand{letter-spacing:-.2px}.ver{border:1px solid #ffffff24;background:#ffffff1e;backdrop-filter:blur(5px)}
.side{background:linear-gradient(180deg,#123e27 0%,#103721 100%);border-right:1px solid #0b2d1b;padding-top:14px}
.side a,.nav-group summary{position:relative;transition:background .14s ease,color .14s ease,transform .14s ease}
.side a:hover,.nav-group summary:hover{background:#ffffff12;transform:translateX(1px)}
.side a.on,.side .nav-home.on{background:#ffffff17;color:#fff;box-shadow:inset 3px 0 0 #65c88e}
.nav-group[open]>summary{background:#ffffff0d}.nav-children{border-left-color:#ffffff16}
.main{background:linear-gradient(180deg,#f4f7f5 0,#f2f5f3 220px)}
h1,h2,h3{letter-spacing:-.25px}h1{color:#173324}h2{color:#193727}
.card{border:1px solid var(--line);box-shadow:var(--shadow-soft);border-radius:13px}
.card:hover{border-color:#d2dfd6}.costbox,.setting-box,.alertitem{border-color:var(--line)}
.summary-grid .card{box-shadow:0 2px 7px rgba(20,55,35,.045)!important;border:1px solid var(--line)!important;transition:transform .14s ease,box-shadow .14s ease,border-color .14s ease}
.summary-grid .summary-link:hover .card{transform:translateY(-1px);box-shadow:var(--shadow-hover)!important;border-color:#c8dacf!important}
.summary-grid .metric-title{font-weight:650;color:#213a2c}.summary-grid .stat b{font-weight:850;color:#102b1d}.summary-grid .metric small{color:#60736a;font-weight:650}
.dashboard-section-title h2{font-size:22px}.dashboard-section-title span{font-size:12px}
.btn{border:1px solid transparent;font-weight:750;box-shadow:0 1px 2px rgba(20,55,35,.08)}
.btn:hover{filter:none;transform:translateY(-1px);box-shadow:0 5px 12px rgba(20,55,35,.11)}
.btn:active{transform:translateY(0);box-shadow:0 1px 3px rgba(20,55,35,.08)}
.btn.alt,.btn.secondary{background:#f1f6f3;color:#176b3a;border-color:#d7e5dc;box-shadow:none}.btn.alt:hover,.btn.secondary:hover{background:#e8f2ec;border-color:#c7d9ce}
.btn.red,.btn.danger{background:#d94336;border-color:#cc3b30}.btn.red:hover,.btn.danger:hover{background:#ce3c31}
.btn.blue{background:#316ebc;border-color:#2b63ac}.btn.orange{background:#e58a12;border-color:#d77f0b}
.form input,.form select,.form textarea,.insem-search input,.livebox input{background:#fff;border-color:#cbd8cf;box-shadow:inset 0 1px 2px rgba(15,49,29,.025);transition:border-color .14s ease,box-shadow .14s ease,background .14s ease}
.form input:focus,.form select:focus,.form textarea:focus,.insem-search input:focus,.livebox input:focus{outline:none;border-color:#4c9a69;box-shadow:0 0 0 3px rgba(36,139,79,.11);background:#fff}
.form label{color:#30483a}
table{border:1px solid var(--line);box-shadow:0 2px 8px rgba(20,55,35,.035)}
th{background:#edf4ef;color:#2a4535;font-weight:800;border-bottom:1px solid #d3e0d7}
tbody tr:nth-child(even){background:#fbfcfb}tbody tr:hover{background:#f0f7f3}td{border-bottom-color:#e5ece7}
.row-actions .btn,.mobile-animal-table .btn{white-space:nowrap}
.animal-tag-btn{border-color:#b8d5c2;background:#f0f8f3;box-shadow:none}.animal-tag-btn:hover{background:#e3f2e8;border-color:#88b99a;box-shadow:0 3px 8px rgba(23,59,40,.09)}
.pill,.status-badge,.perf-badge{border:1px solid rgba(23,107,58,.08)}
.flash{border:1px solid #cce4d4;box-shadow:0 2px 8px rgba(20,55,35,.04)}.flash.err{border-color:#f0c9c5}
.profile{box-shadow:var(--shadow-soft)!important}.profile~.two .card,.profile~.card{background:#fff!important}
.dash-fold>summary{box-shadow:var(--shadow-soft);border-color:var(--line)}.dash-fold>summary:hover{background:#fbfdfb;border-color:#cfddd3}
.menu-toggle{border:1px solid #ffffff1d}
@media(max-width:900px){.side{border-right:0}.main{background:#f3f6f4}}
@media(max-width:700px){
  .summary-grid{gap:8px!important}.summary-grid .card{border-radius:11px!important;box-shadow:0 1px 5px rgba(20,55,35,.045)!important}
  .summary-grid .metric-title{font-weight:700}.summary-grid .metric small{font-size:10.5px!important}
  .mobile-animal-table tr.data-row,.insem-table tr.data-row,.estrus-table tr.data-row{border-radius:12px;box-shadow:0 2px 8px rgba(20,55,35,.05)}
  .animal-tag-btn{min-height:40px;padding:8px 10px}
}

/* UX15 — Premium Dashboard Cards + linked feed/finance safety */
.summary-grid .summary-link.card{position:relative;isolation:isolate;background:linear-gradient(145deg,#fff 0%,#fbfdfc 68%,#f4f9f6 100%)!important;border-left-width:4px!important;padding:12px 14px!important}
.summary-grid .summary-link.card:after{content:"";position:absolute;right:-22px;top:-30px;width:92px;height:92px;border-radius:50%;background:currentColor;opacity:.045;z-index:-1}
.summary-grid .metric-icon{display:grid!important;place-items:center;width:31px;height:31px;border-radius:9px;background:#edf5f0;border:1px solid #dce9e0;font-size:17px!important;margin:0 0 5px!important}
.summary-grid .metric-title{min-height:auto!important;height:auto!important;max-height:none!important;white-space:nowrap;text-overflow:ellipsis;overflow:hidden;width:100%;font-size:13px!important;color:#425a4b!important}
.summary-grid .stat b{font-size:24px!important;letter-spacing:-.6px;margin:4px 0 2px!important}
.summary-grid .metric small{font-size:10.5px!important;opacity:.9}
.summary-grid .metric.green{border-left-color:#2e9b5e!important}.summary-grid .metric.blue{border-left-color:#3479c8!important}.summary-grid .metric.orange{border-left-color:#e59622!important}.summary-grid .metric.teal{border-left-color:#23999b!important}.summary-grid .metric.purple{border-left-color:#8061cc!important}.summary-grid .metric.red{border-left-color:#d64b40!important}
.summary-grid .metric.green .metric-icon{background:#edf8f1}.summary-grid .metric.blue .metric-icon{background:#eef5fd}.summary-grid .metric.orange .metric-icon{background:#fff6e8}.summary-grid .metric.teal .metric-icon{background:#eaf8f8}.summary-grid .metric.purple .metric-icon{background:#f3effd}.summary-grid .metric.red .metric-icon{background:#fff0ef}
.linked-feed-box{grid-column:1/-1;border:1px solid #bddbc7;background:linear-gradient(180deg,#f3fbf6,#edf7f1);border-radius:12px;padding:13px 14px}.linked-feed-box h3{margin:0 0 5px;font-size:15px}.linked-feed-grid{display:grid;grid-template-columns:1.2fr 1fr 1fr;gap:10px;margin-top:10px}.linked-total{display:flex;align-items:center;justify-content:space-between;gap:12px;background:#fff;border:1px solid #d7e7dc;border-radius:10px;padding:10px 12px;margin-top:9px}.linked-total b{font-size:18px;color:#176b3a}
@media(max-width:700px){.summary-grid .summary-link.card{height:116px!important;min-height:116px!important;padding:9px 10px!important}.summary-grid .metric-icon{width:27px;height:27px;font-size:15px!important;margin-bottom:3px!important}.summary-grid .stat b{font-size:21px!important}.summary-grid .metric-title{font-size:12px!important}.linked-feed-grid{grid-template-columns:1fr}.linked-total{align-items:flex-start;flex-direction:column}}

'''

def db():
    c=sqlite3.connect(DB)
    c.row_factory=sqlite3.Row
    return c

def is_pregnant_value(value):
    normalized=str(value or "").strip().lower()
    if normalized.startswith("gebe (satın alındığında"):
        return True
    return normalized in {
        "pozitif","gebe","evet","yes","true","1","olumlu",
        "gebelik pozitif","pozitif (gebe)","pregnant"
    }


def recalculate_animal_exit_status(con, animal_id):
    if not animal_id:
        return
    row=con.execute(
        """select animal_status_action,tx_date,category,amount
           from finance
           where animal_id=? and animal_status_action in ('Satıldı','Kesildi')
           order by tx_date desc,id desc limit 1""",
        (animal_id,)
    ).fetchone()
    if row:
        con.execute(
            "update animals set status=?,exit_date=?,exit_reason=?,sold_price=? where id=?",
            (row["animal_status_action"],row["tx_date"],row["category"],row["amount"],animal_id)
        )
    else:
        con.execute(
            "update animals set status='Aktif',exit_date='',exit_reason='',sold_price=0 where id=?",
            (animal_id,)
        )

_ARCHIVE_SCHEMA_READY=False
_ARCHIVE_SCHEMA_LOCK=threading.Lock()
def ensure_archive_schema():
    """Şema kontrolünü süreç başına bir kez yap. Önceki DEV'de her GET isteğinde
    PRAGMA/UPDATE çalışması login -> dashboard geçişini gereksiz yere yavaşlatıyordu."""
    global _ARCHIVE_SCHEMA_READY
    if _ARCHIVE_SCHEMA_READY:return
    with _ARCHIVE_SCHEMA_LOCK:
        if _ARCHIVE_SCHEMA_READY:return
        with db() as c:
            cols={r[1] for r in c.execute("pragma table_info(animals)").fetchall()}
            for col,typ in [
                ("status","TEXT DEFAULT 'Aktif'"),
                ("exit_date","TEXT DEFAULT ''"),
                ("exit_reason","TEXT DEFAULT ''"),
                ("sold_price","REAL DEFAULT 0")
            ]:
                if col not in cols:c.execute(f"ALTER TABLE animals ADD COLUMN {col} {typ}")
            fcols={r[1] for r in c.execute("pragma table_info(finance)").fetchall()}
            if "animal_status_action" not in fcols:c.execute("ALTER TABLE finance ADD COLUMN animal_status_action TEXT DEFAULT ''")
            c.execute("update animals set status='Aktif' where status is null or trim(status)=''")
        _ARCHIVE_SCHEMA_READY=True




def next_estrus_cycle(con, rec, today=None):
    today=today or date.today()
    try: observed=date.fromisoformat(rec['estrus_date'])
    except Exception: return None
    cycle=1
    while cycle<=24:
        d=con.execute("select decision from estrus_decisions where estrus_id=? and cycle_no=?",(rec['id'],cycle)).fetchone()
        if not d:
            center=observed+timedelta(days=21*cycle)
            return {'cycle_no':cycle,'start':center-timedelta(days=3),'center':center,'end':center+timedelta(days=3)}
        if str(d['decision'])=='Atlandı':
            cycle+=1
            continue
        return None
    return None

def password_hash(password):
    salt=secrets.token_bytes(16); rounds=240000
    digest=hashlib.pbkdf2_hmac('sha256',password.encode('utf-8'),salt,rounds)
    return f'pbkdf2_sha256${rounds}${salt.hex()}${digest.hex()}'

def password_verify(password,stored):
    stored=str(stored or '')
    if stored.startswith('pbkdf2_sha256$'):
        try:
            _,rounds,salt_hex,digest_hex=stored.split('$',3)
            candidate=hashlib.pbkdf2_hmac('sha256',password.encode('utf-8'),bytes.fromhex(salt_hex),int(rounds)).hex()
            return hmac.compare_digest(candidate,digest_hex)
        except Exception:return False
    return hmac.compare_digest(hashlib.sha256(('farm-v05'+password).encode()).hexdigest(),stored)

def audit(username,action,detail='',ip_address=''):
    try:
        with db() as c:c.execute('insert into audit_log(username,action,detail,created_at,ip_address) values(?,?,?,?,?)',(username or 'sistem',action,detail,datetime.now().strftime('%Y-%m-%d %H:%M:%S'),ip_address or ''))
    except Exception:pass


def setting_get(key,default=''):
    try:
        with db() as c:
            r=c.execute('select setting_value from settings where setting_key=?',(key,)).fetchone()
        return (r['setting_value'] if r else default) or default
    except Exception:
        return default

def setting_set(key,value):
    with db() as c:
        c.execute("""insert into settings(setting_key,setting_value) values(?,?)
                     on conflict(setting_key) do update set setting_value=excluded.setting_value""",(key,str(value or '')))

def smtp_config():
    return {
        'host':setting_get('smtp_host','smtp.gmail.com'),
        'port':int(setting_get('smtp_port','587') or 587),
        'username':setting_get('smtp_username',''),
        'password':setting_get('smtp_password',''),
        'sender':setting_get('smtp_sender','') or setting_get('smtp_username',''),
        'security':setting_get('smtp_security','starttls') or 'starttls',
    }

def send_reset_email(to_email,full_name,code):
    cfg=smtp_config()
    if not cfg['host'] or not cfg['sender']:
        raise RuntimeError('E-posta sunucusu ayarları tamamlanmamış.')
    msg=EmailMessage();msg['Subject']='ÇiftlikPro şifre sıfırlama kodu';msg['From']=cfg['sender'];msg['To']=to_email
    display=(full_name or 'ÇiftlikPro kullanıcısı').strip()
    msg.set_content(f"Merhaba {display},\n\nÇiftlikPro şifre sıfırlama doğrulama kodunuz:\n\n{code}\n\nBu kod 5 dakika geçerlidir ve yalnızca bir kez kullanılabilir.\nBu işlemi siz başlatmadıysanız bu e-postayı dikkate almayın.\n\nÇiftlikPro Enterprise")
    context=ssl.create_default_context()
    if cfg['security']=='ssl':
        with smtplib.SMTP_SSL(cfg['host'],cfg['port'],timeout=20,context=context) as server:
            if cfg['username']:server.login(cfg['username'],cfg['password'])
            server.send_message(msg)
    else:
        with smtplib.SMTP(cfg['host'],cfg['port'],timeout=20) as server:
            server.ehlo()
            if cfg['security']=='starttls':server.starttls(context=context);server.ehlo()
            if cfg['username']:server.login(cfg['username'],cfg['password'])
            server.send_message(msg)

def reset_code_hash(salt,code):return hashlib.sha256((str(salt)+'|'+str(code)).encode('utf-8')).hexdigest()
def reset_token_hash(token):return hashlib.sha256(str(token).encode('utf-8')).hexdigest()

def finance_request_fingerprint(username,form):
    keys=('tx_date','tx_type','category','amount','description','payment_method','animal_id','animal_ids','milk_animal_ids')
    payload='|'.join(str(form.get(k,'')).strip() for k in keys)
    return hashlib.sha256((str(username)+'|'+payload).encode('utf-8')).hexdigest()

def claim_request_once(con,fingerprint,ttl_seconds=15):
    cutoff=(datetime.now()-timedelta(seconds=ttl_seconds)).isoformat(timespec='seconds')
    con.execute('delete from request_dedupe where created_at<?',(cutoff,))
    try:
        con.execute('insert into request_dedupe(fingerprint,created_at) values(?,?)',
                    (fingerprint,datetime.now().isoformat(timespec='seconds')))
        return True
    except sqlite3.IntegrityError:
        return False


def active_admin_count():
    with db() as c:return c.execute("select count(*) from users where role='admin' and active=1").fetchone()[0]

def init_db():
    BACKUPS.mkdir(exist_ok=True)
    UPLOADS.mkdir(exist_ok=True)
    with db() as c:
        c.executescript('''
        CREATE TABLE IF NOT EXISTS users(id INTEGER PRIMARY KEY, username TEXT UNIQUE, password TEXT, role TEXT DEFAULT 'admin');
        CREATE TABLE IF NOT EXISTS animals(id INTEGER PRIMARY KEY, tag TEXT UNIQUE NOT NULL, nickname TEXT, gender TEXT NOT NULL, breed TEXT, birth_date TEXT, notes TEXT);
        CREATE TABLE IF NOT EXISTS inseminations(id INTEGER PRIMARY KEY, animal_id INTEGER, attempt INTEGER, insemination_date TEXT, pregnancy_result TEXT, due_date TEXT, UNIQUE(animal_id,attempt));
        CREATE TABLE IF NOT EXISTS estrus_records(id INTEGER PRIMARY KEY, animal_id INTEGER NOT NULL, estrus_date TEXT NOT NULL, signs TEXT, notes TEXT, created_at TEXT NOT NULL);
        CREATE TABLE IF NOT EXISTS estrus_decisions(id INTEGER PRIMARY KEY, estrus_id INTEGER NOT NULL, cycle_no INTEGER NOT NULL, decision TEXT NOT NULL, decision_date TEXT NOT NULL, insemination_id INTEGER, notes TEXT, UNIQUE(estrus_id,cycle_no));
        CREATE TABLE IF NOT EXISTS calves(id INTEGER PRIMARY KEY, tag TEXT UNIQUE NOT NULL, mother_id INTEGER NOT NULL, father_tag TEXT, birth_date TEXT NOT NULL, gender TEXT, notes TEXT);
        CREATE TABLE IF NOT EXISTS health(id INTEGER PRIMARY KEY, animal_id INTEGER, kind TEXT, product TEXT, applied_date TEXT, next_date TEXT, cost REAL DEFAULT 0, notes TEXT);
        CREATE TABLE IF NOT EXISTS health_courses(id INTEGER PRIMARY KEY,kind TEXT NOT NULL,product TEXT NOT NULL,scope_type TEXT NOT NULL DEFAULT 'single',paddock_id INTEGER,start_date TEXT NOT NULL,treatment_days INTEGER DEFAULT 1,times_per_day INTEGER DEFAULT 1,dose_count INTEGER DEFAULT 1,interval_days INTEGER DEFAULT 0,cost_per_application REAL DEFAULT 0,notes TEXT,created_at TEXT NOT NULL);
        CREATE TABLE IF NOT EXISTS health_tasks(id INTEGER PRIMARY KEY,course_id INTEGER NOT NULL,animal_id INTEGER,calf_id INTEGER,planned_date TEXT NOT NULL,dose_no INTEGER DEFAULT 1,dose_total INTEGER DEFAULT 1,day_no INTEGER DEFAULT 1,day_total INTEGER DEFAULT 1,application_no INTEGER DEFAULT 1,applications_per_day INTEGER DEFAULT 1,status TEXT DEFAULT 'Bekliyor',completed_date TEXT,cost REAL DEFAULT 0,notes TEXT);
        CREATE INDEX IF NOT EXISTS idx_health_tasks_due ON health_tasks(status,planned_date,course_id);
        CREATE TABLE IF NOT EXISTS finance(id INTEGER PRIMARY KEY, tx_date TEXT NOT NULL, tx_type TEXT NOT NULL, category TEXT NOT NULL, amount REAL NOT NULL, description TEXT, payment_method TEXT, animal_id INTEGER, created_at TEXT NOT NULL);
        CREATE TABLE IF NOT EXISTS backups(id INTEGER PRIMARY KEY, filename TEXT, created_at TEXT, size_bytes INTEGER);
        CREATE TABLE IF NOT EXISTS weights(id INTEGER PRIMARY KEY, animal_id INTEGER NOT NULL, measure_date TEXT NOT NULL, weight REAL NOT NULL, notes TEXT);
        CREATE TABLE IF NOT EXISTS milk(id INTEGER PRIMARY KEY, animal_id INTEGER NOT NULL, measure_date TEXT NOT NULL, liters REAL NOT NULL, notes TEXT);
        CREATE TABLE IF NOT EXISTS animal_photos(id INTEGER PRIMARY KEY, animal_id INTEGER NOT NULL, filename TEXT NOT NULL, created_at TEXT NOT NULL, caption TEXT);
        CREATE TABLE IF NOT EXISTS audit_log(id INTEGER PRIMARY KEY, username TEXT, action TEXT, detail TEXT, created_at TEXT, ip_address TEXT);
        CREATE TABLE IF NOT EXISTS settings(setting_key TEXT PRIMARY KEY, setting_value TEXT);
        CREATE TABLE IF NOT EXISTS paddocks(id INTEGER PRIMARY KEY,name TEXT UNIQUE NOT NULL,code TEXT,type TEXT,capacity INTEGER DEFAULT 0,notes TEXT,active INTEGER DEFAULT 1,created_at TEXT NOT NULL);
        CREATE TABLE IF NOT EXISTS paddock_history(id INTEGER PRIMARY KEY,animal_source TEXT NOT NULL,animal_id INTEGER NOT NULL,from_paddock_id INTEGER,to_paddock_id INTEGER,moved_at TEXT NOT NULL,notes TEXT);
        CREATE TABLE IF NOT EXISTS feed_catalog(id INTEGER PRIMARY KEY,name TEXT UNIQUE NOT NULL,category TEXT,dm_pct REAL DEFAULT 0,ndf_pct REAL DEFAULT 0,effective_ndf_pct REAL DEFAULT 0,cp_pct REAL DEFAULT 0,tdn_pct REAL DEFAULT 0,me_mcal_kg REAL DEFAULT 0,nem_mcal_kg REAL DEFAULT 0,neg_mcal_kg REAL DEFAULT 0,starch_pct REAL DEFAULT 0,fat_pct REAL DEFAULT 0,ash_pct REAL DEFAULT 0,ca_pct REAL DEFAULT 0,p_pct REAL DEFAULT 0,mg_pct REAL DEFAULT 0,k_pct REAL DEFAULT 0,na_pct REAL DEFAULT 0,s_pct REAL DEFAULT 0,label_cp_pct_as_fed REAL DEFAULT 0,label_me_kcal_kg_as_fed REAL DEFAULT 0,label_crude_fiber_pct_as_fed REAL DEFAULT 0,label_fat_pct_as_fed REAL DEFAULT 0,label_ash_pct_as_fed REAL DEFAULT 0,label_sodium_pct_as_fed REAL DEFAULT 0,starch_degradability_pct REAL DEFAULT 0,ndf_digestibility_pct REAL DEFAULT 0,rdp_pct_cp REAL DEFAULT 0,rup_pct_cp REAL DEFAULT 0,inra_ufv REAL DEFAULT 0,inra_pdi_g_kg_dm REAL DEFAULT 0,inra_pdia_g_kg_dm REAL DEFAULT 0,inra_rpb_g_kg_dm REAL DEFAULT 0,inra_fill_unit REAL DEFAULT 0,processing_method TEXT,solver_min_kg_day REAL DEFAULT 0,solver_max_kg_day REAL DEFAULT 0,constraint_source TEXT,source TEXT,active INTEGER DEFAULT 1);
        CREATE TABLE IF NOT EXISTS feed_prices(id INTEGER PRIMARY KEY,feed_id INTEGER NOT NULL,effective_date TEXT NOT NULL,price_per_kg REAL NOT NULL,notes TEXT);
        CREATE TABLE IF NOT EXISTS feed_stock_transactions(id INTEGER PRIMARY KEY,feed_id INTEGER NOT NULL,tx_date TEXT NOT NULL,tx_type TEXT NOT NULL,quantity_kg REAL NOT NULL,unit_price REAL DEFAULT 0,notes TEXT);
        CREATE TABLE IF NOT EXISTS rations(id INTEGER PRIMARY KEY,name TEXT UNIQUE NOT NULL,target_group TEXT,notes TEXT,active INTEGER DEFAULT 1,created_at TEXT NOT NULL);
        CREATE TABLE IF NOT EXISTS ration_items(id INTEGER PRIMARY KEY,ration_id INTEGER NOT NULL,feed_id INTEGER NOT NULL,kg_per_head_day REAL NOT NULL,UNIQUE(ration_id,feed_id));
        CREATE TABLE IF NOT EXISTS ration_item_history(id INTEGER PRIMARY KEY,ration_id INTEGER NOT NULL,feed_id INTEGER NOT NULL,effective_date TEXT NOT NULL,kg_per_head_day REAL NOT NULL,created_at TEXT NOT NULL,notes TEXT);
        CREATE INDEX IF NOT EXISTS idx_ration_item_history_lookup ON ration_item_history(ration_id,feed_id,effective_date,id);
        CREATE TABLE IF NOT EXISTS paddock_rations(id INTEGER PRIMARY KEY,paddock_id INTEGER NOT NULL,ration_id INTEGER NOT NULL,start_date TEXT NOT NULL,end_date TEXT,active INTEGER DEFAULT 1,notes TEXT);
        ''')
        user_cols={r[1] for r in c.execute('pragma table_info(users)').fetchall()}
        for col,typ in [('full_name','TEXT'),('active','INTEGER DEFAULT 1'),('last_login','TEXT'),('password_changed_at','TEXT'),('recovery_email','TEXT')]:
            if col not in user_cols:c.execute(f'ALTER TABLE users ADD COLUMN {col} {typ}')
        c.execute("update users set active=1 where active is null")
        c.execute("update users set full_name=username where full_name is null or trim(full_name)=''")
        insemination_cols={r[1] for r in c.execute('pragma table_info(inseminations)').fetchall()}
        for col,typ in [('bull_tag','TEXT'),('bull_name','TEXT'),('inseminator','TEXT')]:
            if col not in insemination_cols:c.execute(f'ALTER TABLE inseminations ADD COLUMN {col} {typ}')
        feed_cols={r[1] for r in c.execute('pragma table_info(feed_catalog)').fetchall()}
        endf_added=False
        if 'effective_ndf_pct' not in feed_cols:
            c.execute('ALTER TABLE feed_catalog ADD COLUMN effective_ndf_pct REAL DEFAULT 0'); endf_added=True
        # DEV4.12 bilimsel veri katmanı. Bu alanlar NASEM çekirdeğini INRA/CNCPS
        # değerleriyle karıştırmaz; veri mevcut olduğunda ayrı doğrulama ve güvenlik
        # hesabına temel olur. Sıfır değer "bilinmiyor" anlamındadır.
        advanced_feed_cols=(
            ('starch_degradability_pct','REAL DEFAULT 0'),('ndf_digestibility_pct','REAL DEFAULT 0'),
            ('rdp_pct_cp','REAL DEFAULT 0'),('rup_pct_cp','REAL DEFAULT 0'),
            ('inra_ufv','REAL DEFAULT 0'),('inra_pdi_g_kg_dm','REAL DEFAULT 0'),
            ('inra_pdia_g_kg_dm','REAL DEFAULT 0'),('inra_rpb_g_kg_dm','REAL DEFAULT 0'),
            ('inra_fill_unit','REAL DEFAULT 0'),('processing_method','TEXT'),
            ('solver_min_kg_day','REAL DEFAULT 0'),('solver_max_kg_day','REAL DEFAULT 0'),
            ('constraint_source','TEXT'),
            ('label_cp_pct_as_fed','REAL DEFAULT 0'),('label_me_kcal_kg_as_fed','REAL DEFAULT 0'),
            ('label_crude_fiber_pct_as_fed','REAL DEFAULT 0'),('label_fat_pct_as_fed','REAL DEFAULT 0'),
            ('label_ash_pct_as_fed','REAL DEFAULT 0'),('label_sodium_pct_as_fed','REAL DEFAULT 0'))
        for col,typ in advanced_feed_cols:
            if col not in feed_cols:c.execute(f'ALTER TABLE feed_catalog ADD COLUMN {col} {typ}')
        # Hotfix3 migrasyonu: mevcut kurulumda yeni eNDF sütununu katalogdaki Excel referans değerleriyle doldur.
        # Diğer yem/besin alanlarına dokunulmaz; kullanıcının eski düzenlemeleri korunur.
        if endf_added:
            try:
                catalog_file=PROGRAM_DIR/'feed_catalog.json'
                if catalog_file.exists():
                    for x in json.loads(catalog_file.read_text(encoding='utf-8')):
                        c.execute('update feed_catalog set effective_ndf_pct=? where name=? and coalesce(effective_ndf_pct,0)=0',(float(x.get('effective_ndf_pct') or 0),x.get('name','')))
            except Exception as exc:
                print('eNDF migrasyonu uygulanamadı:',exc)
        c.execute("insert or ignore into settings(setting_key,setting_value) values('male_min_daily_gain','1.0')")
        c.execute("insert or ignore into settings(setting_key,setting_value) values('male_warning_ratio','0.90')")
        ration_cols={r[1] for r in c.execute('pragma table_info(rations)').fetchall()}
        for col,typ in [('target_weight_kg','REAL DEFAULT 450'),('target_adg_kg','REAL DEFAULT 1.3'),('animal_type',"TEXT DEFAULT 'Besi Erkek'"),('ration_type',"TEXT DEFAULT 'Besi'"),('target_milk_l','REAL DEFAULT 25'),('milk_fat_pct','REAL DEFAULT 3.8'),('milk_protein_pct','REAL DEFAULT 3.2'),('target_age_months','REAL DEFAULT 0'),('target_beef_phase',"TEXT DEFAULT 'Otomatik'")]:
            if col not in ration_cols:c.execute(f'ALTER TABLE rations ADD COLUMN {col} {typ}')
        # Rasyon miktar geçmişi: mevcut rasyonları başlangıç revizyonu olarak koru.
        # Böylece padoka atama tarihinden sonraki maliyetler ilerideki rasyon değişiklikleriyle geriye dönük bozulmaz.
        for ri in c.execute('''select ri.ration_id,ri.feed_id,ri.kg_per_head_day,r.created_at
                               from ration_items ri join rations r on r.id=ri.ration_id''').fetchall():
            exists=c.execute('select 1 from ration_item_history where ration_id=? and feed_id=? limit 1',(ri['ration_id'],ri['feed_id'])).fetchone()
            if not exists:
                eff=str(ri['created_at'] or date.today().isoformat())[:10]
                c.execute('insert into ration_item_history(ration_id,feed_id,effective_date,kg_per_head_day,created_at,notes) values(?,?,?,?,?,?)',
                          (ri['ration_id'],ri['feed_id'],eff,float(ri['kg_per_head_day'] or 0),datetime.now().isoformat(timespec='seconds'),'V3.9.8 maliyet entegrasyonu başlangıç revizyonu'))
        c.execute("""CREATE TABLE IF NOT EXISTS password_reset_codes(
            id INTEGER PRIMARY KEY,user_id INTEGER NOT NULL,code_hash TEXT NOT NULL,salt TEXT NOT NULL,expires_at TEXT NOT NULL,
            attempts INTEGER DEFAULT 0,used INTEGER DEFAULT 0,reset_token_hash TEXT,reset_token_expires TEXT,created_at TEXT NOT NULL,ip_address TEXT)""")
        c.execute("""CREATE TABLE IF NOT EXISTS finance_animals(
            finance_id INTEGER NOT NULL,animal_id INTEGER NOT NULL,relation_type TEXT DEFAULT 'İlgili',PRIMARY KEY(finance_id,animal_id))""")
        c.execute("""CREATE TABLE IF NOT EXISTS feed_finance_links(
            id INTEGER PRIMARY KEY,feed_id INTEGER NOT NULL,stock_tx_id INTEGER,finance_id INTEGER,quantity_kg REAL DEFAULT 0,unit_price REAL DEFAULT 0,created_at TEXT NOT NULL,UNIQUE(stock_tx_id),UNIQUE(finance_id))""")
        for k,v in [('smtp_host','smtp.gmail.com'),('smtp_port','587'),('smtp_security','starttls'),('smtp_username',''),('smtp_password',''),('smtp_sender','')]:
            c.execute("insert or ignore into settings(setting_key,setting_value) values(?,?)",(k,v))
        c.execute("""CREATE TABLE IF NOT EXISTS request_dedupe(
            fingerprint TEXT PRIMARY KEY,
            created_at TEXT NOT NULL
        )""")

        health_cols={r[1] for r in c.execute('pragma table_info(health)').fetchall()}
        if 'calf_id' not in health_cols:c.execute('ALTER TABLE health ADD COLUMN calf_id INTEGER')
        calf_cols={r[1] for r in c.execute('pragma table_info(calves)').fetchall()}
        if 'promoted_animal_id' not in calf_cols:c.execute('ALTER TABLE calves ADD COLUMN promoted_animal_id INTEGER')
        if 'promoted_at' not in calf_cols:c.execute('ALTER TABLE calves ADD COLUMN promoted_at TEXT')
        for col,typ in [('nickname','TEXT'),('breed','TEXT'),('paddock','TEXT'),('photo_url','TEXT'),('purchase_date','TEXT'),('purchase_price','REAL DEFAULT 0'),('purchase_payment_method',"TEXT DEFAULT 'Nakit'")]:
            if col not in calf_cols:c.execute(f'ALTER TABLE calves ADD COLUMN {col} {typ}')
        c.execute("CREATE TABLE IF NOT EXISTS calf_weights(id INTEGER PRIMARY KEY,calf_id INTEGER NOT NULL,measure_date TEXT NOT NULL,weight REAL NOT NULL,notes TEXT)")
        c.execute("CREATE TABLE IF NOT EXISTS calf_photos(id INTEGER PRIMARY KEY,calf_id INTEGER NOT NULL,filename TEXT NOT NULL,created_at TEXT NOT NULL,caption TEXT)")
        cols={r[1] for r in c.execute('pragma table_info(animals)').fetchall()}
        for col,typ in [('paddock','TEXT'),('photo_url','TEXT'),('sold_price','REAL DEFAULT 0'),('status',"TEXT DEFAULT 'Aktif'"),('exit_date','TEXT'),('exit_reason','TEXT'),('purchase_date','TEXT'),('purchase_price','REAL DEFAULT 0'),('purchase_weight','REAL DEFAULT 0'),('daily_feed_cost','REAL DEFAULT 0'),('daily_care_cost','REAL DEFAULT 0'),('target_sale_price','REAL DEFAULT 0'),('pregnancy_source','TEXT DEFAULT \'\''),('pregnancy_age_months_at_entry','REAL DEFAULT 0'),('pregnancy_entry_date','TEXT DEFAULT \'\'')]:
            if col not in cols:c.execute(f'ALTER TABLE animals ADD COLUMN {col} {typ}')
        # V3.9.0 Padok + Yem/Rasyon veri modeli
        calf_cols={r[1] for r in c.execute('pragma table_info(calves)').fetchall()}
        if 'paddock_id' not in calf_cols:c.execute('ALTER TABLE calves ADD COLUMN paddock_id INTEGER')
        animal_cols={r[1] for r in c.execute('pragma table_info(animals)').fetchall()}
        if 'paddock_id' not in animal_cols:c.execute('ALTER TABLE animals ADD COLUMN paddock_id INTEGER')
        # Eski serbest metin padokları kaybetmeden gerçek padok kayıtlarına dönüştür.
        legacy_names=set()
        for rr in c.execute("select distinct trim(coalesce(paddock,'')) p from animals where trim(coalesce(paddock,''))<>''").fetchall(): legacy_names.add(rr['p'])
        for rr in c.execute("select distinct trim(coalesce(paddock,'')) p from calves where trim(coalesce(paddock,''))<>''").fetchall(): legacy_names.add(rr['p'])
        for name in sorted(legacy_names):
            c.execute('insert or ignore into paddocks(name,code,type,capacity,notes,active,created_at) values(?,?,?,?,?,?,?)',(name,'','Genel',0,'V3.8.x serbest metin padok kaydından aktarıldı',1,datetime.now().isoformat(timespec='seconds')))
        c.execute("update animals set paddock_id=(select id from paddocks where paddocks.name=trim(animals.paddock)) where paddock_id is null and trim(coalesce(paddock,''))<>''")
        c.execute("update calves set paddock_id=(select id from paddocks where paddocks.name=trim(calves.paddock)) where paddock_id is null and trim(coalesce(paddock,''))<>''")
        # Besi_V5.02.xlsm'den yalnız besin referans verilerini ilk kurulumda yükle; fiyatlar özellikle taşınmaz.
        if c.execute('select count(*) from feed_catalog').fetchone()[0]==0:
            catalog_file=PROGRAM_DIR/'feed_catalog.json'
            if catalog_file.exists():
                try:
                    for x in json.loads(catalog_file.read_text(encoding='utf-8')):
                        c.execute('''insert or ignore into feed_catalog(name,category,dm_pct,ndf_pct,effective_ndf_pct,cp_pct,tdn_pct,me_mcal_kg,nem_mcal_kg,neg_mcal_kg,starch_pct,fat_pct,ash_pct,ca_pct,p_pct,mg_pct,k_pct,na_pct,s_pct,source,active) values(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,1)''',
                                  (x.get('name',''),x.get('category',''),x.get('dm_pct',0),x.get('ndf_pct',0),x.get('effective_ndf_pct',0),x.get('cp_pct',0),x.get('tdn_pct',0),x.get('me_mcal_kg',0),x.get('nem_mcal_kg',0),x.get('neg_mcal_kg',0),x.get('starch_pct',0),x.get('fat_pct',0),x.get('ash_pct',0),x.get('ca_pct',0),x.get('p_pct',0),x.get('mg_pct',0),x.get('k_pct',0),x.get('na_pct',0),x.get('s_pct',0),x.get('source','')))
                except Exception as exc:
                    print('Yem kataloğu yüklenemedi:',exc)
        # V3.9.16: NASEM ile birebir eşleştirilebilen temel yemleri mevcut kurulumlarda da güncelle.
        # Kullanıcının daha önce elle/laboratuvar analiziyle değiştirdiği kayıtları ezmemek için yalnız eski Besi_V5.02 kaynaklı satırlar güncellenir.
        try:
            catalog_file=PROGRAM_DIR/'feed_catalog.json'
            if catalog_file.exists():
                modern_names={'ARPA, AĞIR','ARPA, HAFİF','ARPA SİLAJI','BUĞDAY, ÖĞÜTÜLMÜŞ','BUĞDAY SAMANI','MISIR SİLAJI, %33-40 KM'}
                for x in json.loads(catalog_file.read_text(encoding='utf-8')):
                    if x.get('name') not in modern_names: continue
                    c.execute('''update feed_catalog set category=?,dm_pct=?,ndf_pct=?,effective_ndf_pct=?,cp_pct=?,tdn_pct=?,me_mcal_kg=?,nem_mcal_kg=?,neg_mcal_kg=?,starch_pct=?,fat_pct=?,ash_pct=?,ca_pct=?,p_pct=?,mg_pct=?,k_pct=?,na_pct=?,s_pct=?,source=? where name=? and (source is null or source='' or source like 'Besi_V5.02%')''',
                              (x.get('category',''),x.get('dm_pct',0),x.get('ndf_pct',0),x.get('effective_ndf_pct',0),x.get('cp_pct',0),x.get('tdn_pct',0),x.get('me_mcal_kg',0),x.get('nem_mcal_kg',0),x.get('neg_mcal_kg',0),x.get('starch_pct',0),x.get('fat_pct',0),x.get('ash_pct',0),x.get('ca_pct',0),x.get('p_pct',0),x.get('mg_pct',0),x.get('k_pct',0),x.get('na_pct',0),x.get('s_pct',0),x.get('source',''),x.get('name','')))
        except Exception as exc:
            print('NASEM yem kataloğu güncellemesi uygulanamadı:',exc)
        # HOTFIX 6.13: Mevcut DEV/veritabanlarında feed_catalog.json yalnız ilk kurulumda
        # yüklenmiş olduğu için 6.12'de düzeltilen JSON değerleri eski DB satırlarına ulaşmıyordu.
        # Burada yalnız bariz eski/şüpheli kayıtlar gerçek referans değerlerine migrate edilir.
        # Kullanıcının makul laboratuvar/elle analiz değerleri korunur.
        try:
            def _patch_feed(name, values, suspicious_sql=None, suspicious_args=()):
                row=c.execute("select id,source,starch_pct from feed_catalog where upper(name)=upper(?)",(name,)).fetchone()
                if not row: return
                src=(row['source'] or '')
                old_source=((not src) or src.startswith('Besi_V5.02') or src.startswith('ÇiftlikPro 4.11')
                            or src.startswith('ÇiftlikPro 6.12') or src.startswith('ÇiftlikPro 6.13')
                            or src.startswith('ÇiftlikPro DEV4.14') or src.startswith('Sunar/Çukoyem')
                            or src.startswith('Sunar resmi') or src.startswith('Sunar ürün etiketi'))
                suspicious=bool(suspicious_sql and c.execute('select 1 from feed_catalog where id=? and '+suspicious_sql,(row['id'],*suspicious_args)).fetchone())
                if not (old_source or suspicious): return
                cols=list(values)
                c.execute('update feed_catalog set '+','.join(k+'=?' for k in cols)+' where id=?',tuple(values[k] for k in cols)+(row['id'],))

            _patch_feed('ARPA SAMANI',{
                'category':'Kuru Kaba Yemler','dm_pct':90.9,'ndf_pct':80.5,'effective_ndf_pct':100.0,
                'cp_pct':3.8,'starch_pct':2.0,'fat_pct':1.4,'ash_pct':7.5,'ca_pct':0.46,'p_pct':0.10,
                'source':'ÇiftlikPro 6.13 · Feedipedia barley straw referansı'
            },'starch_pct>20 OR ndf_pct<65')
            _patch_feed('BUĞDAY KEPEĞİ',{
                'category':'Kesif Yemler','dm_pct':87.0,'ndf_pct':45.2,'cp_pct':17.3,'starch_pct':23.1,
                'fat_pct':3.9,'ash_pct':5.6,'ca_pct':0.13,'p_pct':1.10,
                'source':'ÇiftlikPro 6.13 · Feedipedia wheat bran referansı'
            },'starch_pct>45')
            _patch_feed('BUZAĞI BÜYÜTME YEMİ',{
                'name':'SUNAR BUZAĞI BÜYÜTME ÖZEL DÖNEM YEMİ','category':'Ticari Karma Yem','starch_pct':32.0,
                'processing_method':'Ticari buzağı büyütme yemi',
                'constraint_source':'Sunar 2020 katalog: 61-120 gün serbest; sayısal etiket alt/üst dozu yayımlanmamış',
                'source':'Sunar resmi ürün sayfası ve 2020 katalog: 60-120 gün, serbest tüketim; besin profili ÇiftlikPro/Besi_V5.02 referans tahminidir, ürün etiketi veya laboratuvar analiziyle doğrulanmalı'
            },'starch_pct>55')
            # DEV4.18: Kullanıcının 19-20 Ağustos 2026 tarihli gerçek Sunar
            # etiketleri. Çuval değerleri ürün bazında aynen, solver değerleri ise
            # %88,35 referans KM üzerinden tutulur. Etikette bulunmayan NDF,
            # nişasta, KM ve Ca/P alanları JSON profilinde açıkça tahmin olarak kalır.
            sunar_beef_1526={
                'name':'SUNAR 15.26 GELİŞTİRME BESİ YEMİ','category':'Ticari Karma Yem',
                'cp_pct':16.978,'me_mcal_kg':2.943,'nem_mcal_kg':1.984,'neg_mcal_kg':1.333,
                'starch_pct':35.0,'fat_pct':3.396,'ash_pct':8.749,'na_pct':0.306,
                'label_cp_pct_as_fed':15.0,'label_me_kcal_kg_as_fed':2600.0,
                'label_crude_fiber_pct_as_fed':9.27,'label_fat_pct_as_fed':3.0,
                'label_ash_pct_as_fed':7.73,'label_sodium_pct_as_fed':0.27,
                'solver_max_kg_day':10.0,'processing_method':'Ticari besi yemi',
                'constraint_source':'Sunar ürün etiketi 20/08/2026: üst doz 10 kg/baş/gün; alt doz kat yerinde okunamadığı için kesin alt sınır girilmemiştir',
                'source':"Sunar ürün etiketi 20/08/2026: ürün bazında %15 HP, %3,00 yağ, %9,27 ham selüloz, %7,73 kül ve %0,27 sodyum. Sunar resmi ürün adı 15.26 Geliştirme Besi Yemi'dir; 2600 kcal/kg enerji sınıfı ürün kodundan alınmıştır, güncel etiketin analitik bileşenler bölümünde ME ayrıca yazmamaktadır. Solver alanları %88,35 referans KM ile dönüştürülmüştür. KM, NDF, nişasta, Ca/P ve ileri alanlar referans tahminidir; laboratuvar analiziyle doğrulanmalıdır"
            }
            for old_name in ('ÇUKOYEM GELİŞTİRME BESİ YEMİ,15,2650','SIĞIR BESİ YEMİ,15,2700'):
                _patch_feed(old_name,sunar_beef_1526,'starch_pct>55 OR category like ?',('Sulu Kaba%',))

            sunar_dairy_1927={
                'name':'SUNAR KARDELEN 19.27 SÜT YEMİ','category':'Ticari Karma Yem',
                'cp_pct':21.505,'me_mcal_kg':3.056,'nem_mcal_kg':2.078,'neg_mcal_kg':1.413,
                'starch_pct':32.0,'fat_pct':3.962,'ash_pct':7.799,'na_pct':0.374,
                'processing_method':'Ticari süt yemi','label_cp_pct_as_fed':19.0,
                'label_me_kcal_kg_as_fed':2700.0,'label_crude_fiber_pct_as_fed':9.07,
                'label_fat_pct_as_fed':3.50,'label_ash_pct_as_fed':6.89,
                'label_sodium_pct_as_fed':0.33,'solver_min_kg_day':6.0,'solver_max_kg_day':12.0,
                'constraint_source':'Sunar ürün etiketi 19/08/2026: 6-12 kg/baş/gün; 3-4 öğünde toplam rasyona karıştırılarak ve 10 günlük geçişle kullanılmalıdır',
                'source':"Sunar ürün etiketi 19/08/2026: ürün bazında %19 HP, %3,50 yağ, %9,07 ham selüloz, %6,89 kül ve %0,33 sodyum. Sunar resmi ürün adı Kardelen 19.27'dir; 2700 kcal/kg enerji sınıfı ürün kodu ve önceki resmi katalogla uyumludur, güncel etiketin analitik bileşenler bölümünde ME ayrıca yazmamaktadır. Solver alanları %88,35 referans KM ile dönüştürülmüştür. KM, NDF, nişasta, Ca/P ve ileri alanlar referans tahminidir; laboratuvar analiziyle doğrulanmalıdır"
            }
            for old_name in ('SUNAR KARDELEN SÜT YEMİ,19,2700','SIĞIR SÜT YEMİ'):
                _patch_feed(old_name,sunar_dairy_1927,'starch_pct>55 OR category like ?',('Sulu Kaba%',))
            # DEV4.14: Katalogdaki bütün jenerik ticari yemlerin mevcut kurulumlara
            # temel besin profilini eksiksiz taşı. Bunlar marka/parti analizi değildir;
            # ileri rumen/INRA ve etiket doz alanları bilinmiyorsa sıfır bırakılır.
            # Kullanıcı veya laboratuvar kaynaklı bir satır kesinlikle ezilmez.
            commercial_names={
                'BUZAĞI BAŞLANGIÇ YEMİ','SUNAR BUZAĞI BÜYÜTME ÖZEL DÖNEM YEMİ',
                'SUNAR KARDELEN 19.27 SÜT YEMİ',
                'SIĞIR BESİ YEMİ,13,2700','SIĞIR BESİ YEMİ,14,2800',
                'SUNAR 15.26 GELİŞTİRME BESİ YEMİ','SIĞIR BESİ YEMİ,14,2600'}
            catalog_rows={x.get('name'):x for x in json.loads(catalog_file.read_text(encoding='utf-8')) if x.get('name') in commercial_names}
            basic_cols=('category','dm_pct','ndf_pct','effective_ndf_pct','cp_pct','tdn_pct',
                        'me_mcal_kg','nem_mcal_kg','neg_mcal_kg','starch_pct','fat_pct','ash_pct',
                        'ca_pct','p_pct','mg_pct','k_pct','na_pct','s_pct',
                        'label_cp_pct_as_fed','label_me_kcal_kg_as_fed',
                        'label_crude_fiber_pct_as_fed','label_fat_pct_as_fed',
                        'label_ash_pct_as_fed','label_sodium_pct_as_fed','source')
            for commercial_name,x in catalog_rows.items():
                current=c.execute('select id,source from feed_catalog where upper(name)=upper(?)',(commercial_name,)).fetchone()
                if not current:
                    c.execute('''insert or ignore into feed_catalog(
                        name,category,dm_pct,ndf_pct,effective_ndf_pct,cp_pct,tdn_pct,
                        me_mcal_kg,nem_mcal_kg,neg_mcal_kg,starch_pct,fat_pct,ash_pct,
                        ca_pct,p_pct,mg_pct,k_pct,na_pct,s_pct,processing_method,
                        label_cp_pct_as_fed,label_me_kcal_kg_as_fed,
                        label_crude_fiber_pct_as_fed,label_fat_pct_as_fed,
                        label_ash_pct_as_fed,label_sodium_pct_as_fed,
                        solver_min_kg_day,solver_max_kg_day,constraint_source,source,active
                    ) values(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,1)''',(
                        x.get('name',''),x.get('category',''),x.get('dm_pct',0),x.get('ndf_pct',0),
                        x.get('effective_ndf_pct',0),x.get('cp_pct',0),x.get('tdn_pct',0),
                        x.get('me_mcal_kg',0),x.get('nem_mcal_kg',0),x.get('neg_mcal_kg',0),
                        x.get('starch_pct',0),x.get('fat_pct',0),x.get('ash_pct',0),x.get('ca_pct',0),
                        x.get('p_pct',0),x.get('mg_pct',0),x.get('k_pct',0),x.get('na_pct',0),
                        x.get('s_pct',0),x.get('processing_method') or 'Ticari karma yem',
                        x.get('label_cp_pct_as_fed',0),x.get('label_me_kcal_kg_as_fed',0),
                        x.get('label_crude_fiber_pct_as_fed',0),x.get('label_fat_pct_as_fed',0),
                        x.get('label_ash_pct_as_fed',0),x.get('label_sodium_pct_as_fed',0),
                        x.get('solver_min_kg_day',0),x.get('solver_max_kg_day',0),
                        x.get('constraint_source') or 'Ürün etiketi veya laboratuvar analizi gerekli',
                        x.get('source','')
                    ))
                    current=c.execute('select id,source from feed_catalog where upper(name)=upper(?)',(commercial_name,)).fetchone()
                if not current:continue
                src=str(current['source'] or '')
                reference_source=(not src or src.startswith('Besi_V5.02') or src.startswith('ÇiftlikPro 4.11')
                                  or src.startswith('ÇiftlikPro 6.12') or src.startswith('ÇiftlikPro 6.13')
                                  or src.startswith('ÇiftlikPro DEV4.14') or src.startswith('Sunar/Çukoyem')
                                  or src.startswith('Sunar resmi') or src.startswith('Sunar ürün etiketi'))
                if not reference_source:continue
                values={k:x.get(k,0) for k in basic_cols}
                values.update({'processing_method':'Ticari karma yem',
                               'constraint_source':x.get('constraint_source') or 'Ürün etiketi veya laboratuvar analizi gerekli'})
                if float(x.get('solver_min_kg_day') or 0)>0:
                    values['solver_min_kg_day']=float(x['solver_min_kg_day'])
                if float(x.get('solver_max_kg_day') or 0)>0:
                    values['solver_max_kg_day']=float(x['solver_max_kg_day'])
                cols=list(values)
                c.execute('update feed_catalog set '+','.join(k+'=?' for k in cols)+' where id=?',
                          tuple(values[k] for k in cols)+(current['id'],))
            print('HOTFIX 6.13 yem kataloğu gerçek DB migrasyonu kontrol edildi.')
        except Exception as exc:
            print('HOTFIX 6.13 yem kataloğu migrasyonu uygulanamadı:',exc)
        # DEV10: Ezme adları JSON'da kalmasın; mevcut kullanıcı veritabanlarına da ekle.
        # Besin değerleri uygulamanın kendi feed_catalog.json kaydından gelir; mevcut kullanıcı yemleri değiştirilmez.
        try:
            alias_names={'ARPA EZMESİ','BUĞDAY EZMESİ','MISIR EZMESİ','YULAF EZMESİ','ÇAVDAR EZMESİ'}
            catalog_file=PROGRAM_DIR/'feed_catalog.json'
            if catalog_file.exists():
                for x in json.loads(catalog_file.read_text(encoding='utf-8')):
                    if x.get('name') not in alias_names: continue
                    sql_alias='insert or ignore into feed_catalog(name,category,dm_pct,ndf_pct,effective_ndf_pct,cp_pct,tdn_pct,me_mcal_kg,nem_mcal_kg,neg_mcal_kg,starch_pct,fat_pct,ash_pct,ca_pct,p_pct,mg_pct,k_pct,na_pct,s_pct,source,active) values(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,1)'
                    c.execute(sql_alias,(x.get('name',''),'Kesif Yemler',x.get('dm_pct',0),x.get('ndf_pct',0),x.get('effective_ndf_pct',0),x.get('cp_pct',0),x.get('tdn_pct',0),x.get('me_mcal_kg',0),x.get('nem_mcal_kg',0),x.get('neg_mcal_kg',0),x.get('starch_pct',0),x.get('fat_pct',0),x.get('ash_pct',0),x.get('ca_pct',0),x.get('p_pct',0),x.get('mg_pct',0),x.get('k_pct',0),x.get('na_pct',0),x.get('s_pct',0),x.get('source','ÇiftlikPro katalog adı eşlemesi')))
        except Exception as exc:
            print('DEV10 ezme yemleri DB migrasyonu uygulanamadı:',exc)
        # DEV4.12: INRA 2018 değerleri NASEM gereksinim hesabına eklenmez. Yalnız
        # ayrı bir veri-kapsama ve rumen fermantasyon kontrolünde kullanılmak üzere,
        # açıkça eşleşen standart tanelere referans alanları doldurulur. Kullanıcının
        # girdiği değer (sıfırdan büyükse) hiçbir zaman ezilmez.
        try:
            scientific_refs={
                'ARPA, AĞIR':(86.0,68.0,1.11,87.0,30.0,-23.0,0.194,'Tane'),
                'ARPA, HAFİF':(86.0,68.0,1.11,87.0,30.0,-23.0,0.194,'Tane'),
                'ARPA EZMESİ':(86.0,68.0,1.11,87.0,30.0,-23.0,0.194,'Ezme'),
                'BUĞDAY, ÖĞÜTÜLMÜŞ':(90.0,68.0,1.22,89.0,28.0,-13.0,0.153,'Öğütülmüş'),
                'BUĞDAY EZMESİ':(90.0,68.0,1.22,89.0,28.0,-13.0,0.153,'Ezme'),
                'MISIR, 72 kg/hektolitre':(61.0,79.0,1.33,95.0,42.0,-55.0,0.336,'Tane'),
                'MISIR EZMESİ':(61.0,79.0,1.33,95.0,42.0,-55.0,0.336,'Ezme'),
            }
            for name,(sd,ndfd,ufv,pdi,pdia,rpb,fill,processing) in scientific_refs.items():
                c.execute('''update feed_catalog set
                    starch_degradability_pct=case when coalesce(starch_degradability_pct,0)=0 then ? else starch_degradability_pct end,
                    ndf_digestibility_pct=case when coalesce(ndf_digestibility_pct,0)=0 then ? else ndf_digestibility_pct end,
                    inra_ufv=case when coalesce(inra_ufv,0)=0 then ? else inra_ufv end,
                    inra_pdi_g_kg_dm=case when coalesce(inra_pdi_g_kg_dm,0)=0 then ? else inra_pdi_g_kg_dm end,
                    inra_pdia_g_kg_dm=case when coalesce(inra_pdia_g_kg_dm,0)=0 then ? else inra_pdia_g_kg_dm end,
                    inra_rpb_g_kg_dm=case when coalesce(inra_rpb_g_kg_dm,0)=0 then ? else inra_rpb_g_kg_dm end,
                    inra_fill_unit=case when coalesce(inra_fill_unit,0)=0 then ? else inra_fill_unit end,
                    processing_method=case when trim(coalesce(processing_method,''))='' then ? else processing_method end
                    where upper(name)=upper(?)''',(sd,ndfd,ufv,pdi,pdia,rpb,fill,processing,name))
        except Exception as exc:
            print('DEV4.12 bilimsel yem referansı migrasyonu uygulanamadı:',exc)
        # V3.9.18: ÇiftlikPro 4 Fazlı Besi reçeteleri.
        # Reçeteler mevcut yem kataloğundaki kayıtları kullanır; besin değerlerini kopyalamaz.
        # Böylece kullanıcı bir yemin KM/HP/NDF/ME vb. değerini düzenlediğinde rasyon hesabı anında yeni değeri kullanır.
        try:
            def _feed_key(txt):
                txt=(txt or '').strip().upper()
                tr=str.maketrans({'İ':'I','Ş':'S','Ğ':'G','Ü':'U','Ö':'O','Ç':'C'})
                return ' '.join(txt.translate(tr).replace('_',' ').replace('-',' ').split())

            active_feeds=c.execute("select id,name from feed_catalog where active=1 order by id").fetchall()
            keyed=[(r['id'],r['name'],_feed_key(r['name'])) for r in active_feeds]

            def _pick_feed(*aliases):
                alias_keys=[_feed_key(a) for a in aliases if a]
                for ak in alias_keys:
                    for fid,name,k in keyed:
                        if k==ak:return fid,name
                for ak in alias_keys:
                    toks=[t for t in ak.split() if len(t)>1]
                    for fid,name,k in keyed:
                        if toks and all(t in k for t in toks):return fid,name
                return None,None

            feed_refs={
                'straw':_pick_feed('ARPA SAMANI'),
                'barley':_pick_feed('ARPA EZMESİ','ARPA EZMESI','ARPA, AĞIR','ARPA, HAFİF'),
                'wheat':_pick_feed('BUĞDAY EZMESİ','BUGDAY EZMESI','BUĞDAY, ÖĞÜTÜLMÜŞ'),
                'alfalfa':_pick_feed('YONCA',"YONCA KURU OTU, KM'de %17-19 HP, %40-44 NDF","YONCA KURU OTU, KM'de %15-17 HP, %44-48 NDF"),
                'cob_silage':_pick_feed('MISIR KOÇANI SİLAJI','MISIR KOCANI SILAJI'),
                'beef15':_pick_feed('SUNAR 15.26 GELİŞTİRME BESİ YEMİ','15 PROTEİN 2600 ME BESİ YEMİ','%15 BESİ YEMİ','15 BESİ YEMİ'),
                'dairy19':_pick_feed('SUNAR KARDELEN 19.27 SÜT YEMİ','19 PROTEİN 2700 ME SÜT YEMİ','%19 SÜT YEMİ','19 SÜT YEMİ'),
            }

            phase_recipes=[
                ('ÇiftlikPro Faz 1 · 250–350 kg Büyütme','250–350 kg · Büyütme',300,1.00,
                 {'straw':1.20,'barley':2.00,'wheat':0.30,'alfalfa':1.50,'cob_silage':4.00,'beef15':1.00,'dairy19':0.50},
                 'Hazır başlangıç reçetesi · hedef CAA yaklaşık 0,9–1,1 kg/gün. Kaba yem ve protein ağırlıklı büyütme fazı.'),
                ('ÇiftlikPro Faz 2 · 350–450 kg Geliştirme','350–450 kg · Geliştirme',400,1.20,
                 {'straw':1.20,'barley':3.50,'wheat':0.50,'alfalfa':1.40,'cob_silage':4.50,'beef15':1.40,'dairy19':0.30},
                 'Hazır başlangıç reçetesi · hedef CAA yaklaşık 1,1–1,3 kg/gün. Enerji kademeli yükseltilir.'),
                ('ÇiftlikPro Faz 3 · 450–550 kg Yoğun Besi','450–550 kg · Yoğun Besi',500,1.50,
                 {'straw':1.00,'barley':6.50,'wheat':1.00,'alfalfa':1.00,'cob_silage':4.00,'beef15':2.00,'dairy19':0.00},
                 'Hazır başlangıç reçetesi · hedef CAA yaklaşık 1,4–1,5 kg/gün. Buğday 1,0 kg/gün üst sınırında tutulur.'),
                ('ÇiftlikPro Faz 4 · 550 kg+ Bitirme','550 kg+ · Bitirme',600,1.35,
                 {'straw':0.80,'barley':8.00,'wheat':1.00,'alfalfa':1.00,'cob_silage':5.00,'beef15':2.00,'dairy19':0.00},
                 'Hazır başlangıç reçetesi · hedef CAA yaklaşık 1,2–1,5 kg/gün. Ekonomik bitirme ve yüksek enerji yoğunluğu hedeflenir.'),
            ]

            now_iso=datetime.now().isoformat(timespec='seconds')
            for rname,target_group,target_w,target_adg,items,rnote in phase_recipes:
                rr=c.execute('select id from rations where name=?',(rname,)).fetchone()
                if rr:
                    oldnote=c.execute('select notes from rations where id=?',(rr['id'],)).fetchone()
                    if oldnote and 'Hazır başlangıç reçetesi' in str(oldnote['notes'] or ''):
                        rid=int(rr['id'])
                        c.execute('delete from ration_items where ration_id=?',(rid,))
                        c.execute('update rations set target_group=?,target_weight_kg=?,target_adg_kg=?,notes=? where id=?',(target_group,float(target_w),float(target_adg),rnote+' V3.9.19 dengelenmiş hazır reçete · NASEM ortalama yem değerleri esaslıdır.',rid))
                        for key,kg in items.items():
                            fid,_fname=feed_refs[key]
                            if fid and float(kg)>0:
                                c.execute('insert or replace into ration_items(ration_id,feed_id,kg_per_head_day) values(?,?,?)',(rid,fid,float(kg)))
                                c.execute('insert into ration_item_history(ration_id,feed_id,effective_date,kg_per_head_day,created_at,notes) values(?,?,?,?,?,?)',(rid,fid,date.today().isoformat(),float(kg),now_iso,'V3.9.19 dengelenmiş hazır 4 faz reçetesi'))
                        continue
                    continue  # Kullanıcının kendi reçetesini/değiştirdiği kaydı ezme.
                missing=[]
                for key in items:
                    if not feed_refs[key][0]:missing.append(key)
                note=rnote+' Besin hesapları doğrudan aktif Yem Kataloğu değerlerinden yapılır.'
                if missing:
                    labels={'straw':'Arpa samanı','barley':'Arpa ezmesi','wheat':'Buğday ezmesi','alfalfa':'Yonca','cob_silage':'Mısır koçanı silajı','beef15':'Sunar 15.26 besi yemi','dairy19':'Sunar Kardelen 19.27 süt yemi'}
                    note+=' Eksik katalog eşleşmesi: '+', '.join(labels[x] for x in missing)+'. Bu yemleri katalogda ekledikten sonra reçeteye manuel ekleyebilirsiniz.'
                cur=c.execute('''insert into rations(name,target_group,notes,active,created_at,target_weight_kg,target_adg_kg,animal_type,ration_type,target_milk_l,milk_fat_pct,milk_protein_pct)
                                 values(?,?,?,1,?,?,?,?,?,?,?,?)''',
                              (rname,target_group,note,now_iso,float(target_w),float(target_adg),'Besi Erkek','Besi',25,3.8,3.2))
                rid=cur.lastrowid
                for key,kg in items.items():
                    fid,_fname=feed_refs[key]
                    if not fid or float(kg)<=0:continue
                    c.execute('insert or ignore into ration_items(ration_id,feed_id,kg_per_head_day) values(?,?,?)',(rid,fid,float(kg)))
                    c.execute('insert into ration_item_history(ration_id,feed_id,effective_date,kg_per_head_day,created_at,notes) values(?,?,?,?,?,?)',
                              (rid,fid,date.today().isoformat(),float(kg),now_iso,'V3.9.18 hazır 4 fazlı besi reçetesi başlangıç değeri'))
        except Exception as exc:
            print('4 fazlı hazır besi reçeteleri oluşturulamadı:',exc)

        finance_cols={r[1] for r in c.execute('pragma table_info(finance)').fetchall()}
        if 'animal_status_action' not in finance_cols:c.execute("ALTER TABLE finance ADD COLUMN animal_status_action TEXT DEFAULT ''")
        n=c.execute('select count(*) from users').fetchone()[0]
        if not n:
            c.execute('insert into users(username,password,role,full_name,active,password_changed_at) values(?,?,?,?,?,?)',('admin',password_hash('admin123'),'admin','Yönetici',1,datetime.now().strftime('%Y-%m-%d %H:%M:%S')))


FARM_PROFILE_KEYS = (
    'farm_name','owner_name','phone','email','province','district','address',
    'business_no','tax_or_tc','vet_name','vet_phone','vet_email','notes','farm_logo'
)

def farm_profile():
    profile={k:'' for k in FARM_PROFILE_KEYS}
    try:
        with db() as c:
            rows=c.execute("select setting_key,setting_value from settings").fetchall()
        for r in rows:
            if r['setting_key'] in profile:
                profile[r['setting_key']]=r['setting_value'] or ''
    except Exception:
        pass
    return profile

def farm_display_name(profile=None):
    p=profile or farm_profile()
    return (p.get('farm_name') or '').strip() or 'ÇiftlikPro'


ANIMAL_IMPORT_ALIASES={
    'tag':('kupe no','kupe numarasi','kupe','tag','ear tag'),
    'herd_no':('suru no','suru numarasi','herd no'),
    'species':('tur','hayvan turu','species'),
    'breed':('irk','breed'),
    'gender':('cinsiyet','gender'),
    'birth_date':('dogum tarihi','dogum','birth date'),
    'mother_tag':('ana no','anne no','anne kupe no','ana kupe no','mother tag'),
    'arrival_date':('gelis tarihi','alis tarihi','isletmeye gelis tarihi','arrival date'),
    'nickname':('takma ad','hayvan adi','nickname'),
    'paddock':('padok','ahir','padok ahir'),
    'notes':('not','notlar','aciklama','notes'),
}

def normalized_heading(value):
    text=unicodedata.normalize('NFD',str(value or '').strip().lower())
    text=''.join(ch for ch in text if unicodedata.category(ch)!='Mn').replace('ı','i')
    return re.sub(r'[^a-z0-9]+',' ',text).strip()

def normalize_animal_tag(value):
    if value is None:return ''
    if isinstance(value,float) and value.is_integer():value=int(value)
    return re.sub(r'\s+','',str(value).strip()).upper()

def parse_import_date(value):
    if value is None or value=='':return ''
    if isinstance(value,datetime):return value.date().isoformat()
    if isinstance(value,date):return value.isoformat()
    text=str(value).strip()
    for fmt in ('%Y-%m-%d','%d/%m/%Y','%d.%m.%Y','%d-%m-%Y','%d/%m/%y','%d.%m.%y'):
        try:return datetime.strptime(text[:10],fmt).date().isoformat()
        except Exception:pass
    return None

def _tabular_import_rows(matrix):
    matrix=[list(r) for r in matrix if any(str(v or '').strip() for v in r)]
    if not matrix:raise ValueError('Dosyada okunabilir satır bulunamadı.')
    aliases={alias:key for key,values in ANIMAL_IMPORT_ALIASES.items() for alias in values}
    header_index=-1; mapping={}
    for idx,row in enumerate(matrix[:30]):
        candidate={}
        for col,value in enumerate(row):
            key=aliases.get(normalized_heading(value))
            if key and key not in candidate:candidate[key]=col
        if 'tag' in candidate and len(candidate)>=3:
            header_index=idx;mapping=candidate;break
    if header_index<0:raise ValueError('Küpe No başlığını içeren hayvan tablosu bulunamadı.')
    rows=[]
    for row in matrix[header_index+1:]:
        item={key:(row[col] if col<len(row) else '') for key,col in mapping.items()}
        if any(str(v or '').strip() for v in item.values()):rows.append(item)
    if not rows:raise ValueError('Başlık bulundu ancak hayvan satırı bulunamadı.')
    return rows

def _official_pdf_rows(text):
    rows=[]
    tag=r'(?:[A-Z]{1,3})?\d{8,15}'
    date_rx=r'\d{2}[./-]\d{2}[./-]\d{2,4}'
    pattern=re.compile(
        rf'^\s*(?P<tag>{tag})\s+(?P<herd_no>\d+)\s+(?P<species>SIĞIR|SIGIR|MANDA)\s+'
        rf'(?P<breed>.*?)\s+(?P<gender>DİŞİ|DIŞI|DISI|ERKEK)\s+(?P<birth_date>{date_rx})'
        rf'(?:\s+(?P<mother_tag>{tag}|-))?\s+(?P<arrival_date>{date_rx})\s*$',re.I)
    for raw in str(text or '').splitlines():
        line=re.sub(r'\s+',' ',raw).strip()
        match=pattern.match(line)
        if match:rows.append(match.groupdict())
    return rows

def parse_animal_import_file(filename,content):
    name=str(filename or '').lower(); raw_rows=[]
    if len(content)>12*1024*1024:raise ValueError('Dosya 12 MB sınırını aşıyor.')
    if name.endswith('.xlsx'):
        from openpyxl import load_workbook
        wb=load_workbook(io.BytesIO(content),read_only=True,data_only=True)
        ws=wb.active
        try:raw_rows=_tabular_import_rows(ws.iter_rows(values_only=True))
        finally:wb.close()
    elif name.endswith('.csv'):
        decoded=None
        for enc in ('utf-8-sig','cp1254','latin-1'):
            try:decoded=content.decode(enc);break
            except UnicodeDecodeError:pass
        if decoded is None:raise ValueError('CSV metin kodlaması okunamadı.')
        try:
            dialect=csv.Sniffer().sniff(decoded[:4096],delimiters=';,\t,')
            matrix=csv.reader(io.StringIO(decoded),dialect)
        except Exception:matrix=csv.reader(io.StringIO(decoded),delimiter=';')
        raw_rows=_tabular_import_rows(matrix)
    elif name.endswith('.pdf'):
        from pypdf import PdfReader
        reader=PdfReader(io.BytesIO(content));parts=[]
        for pdf_page in reader.pages:
            try:parts.append(pdf_page.extract_text(extraction_mode='layout') or '')
            except TypeError:parts.append(pdf_page.extract_text() or '')
        text='\n'.join(parts)
        if not text.strip():raise ValueError('PDF taranmış görüntü biçiminde. Orijinal dijital PDF veya Excel kullanın.')
        raw_rows=_official_pdf_rows(text)
        if not raw_rows:raise ValueError('PDF içindeki hayvan tablosu güvenli biçimde ayrıştırılamadı. Excel dosyasını kullanın veya PDF örneğini kontrol edin.')
    else:raise ValueError('Yalnızca XLSX, CSV veya PDF dosyası desteklenir.')
    if len(raw_rows)>5000:raise ValueError('Tek seferde en fazla 5.000 hayvan aktarılabilir.')
    return raw_rows

def prepare_animal_import(filename,raw_rows):
    with db() as c:
        existing={normalize_animal_tag(r['tag']) for r in c.execute('select tag from animals').fetchall()}
        existing.update(normalize_animal_tag(r['tag']) for r in c.execute('select tag from calves').fetchall())
        existing_females={normalize_animal_tag(r['tag']) for r in c.execute("select tag from animals where gender='Dişi'").fetchall()}
    incoming_females=set()
    for raw in raw_rows:
        if normalized_heading(raw.get('gender')) in ('disi','female','f'):
            incoming_females.add(normalize_animal_tag(raw.get('tag')))
    known_females=existing_females|incoming_females
    prepared=[];seen=set()
    for index,raw in enumerate(raw_rows,1):
        tag=normalize_animal_tag(raw.get('tag'));mother=normalize_animal_tag(raw.get('mother_tag'))
        gender_raw=normalized_heading(raw.get('gender'))
        gender='Dişi' if gender_raw in ('disi','female','f') else 'Erkek' if gender_raw in ('erkek','male','m') else ''
        birth=parse_import_date(raw.get('birth_date'));arrival=parse_import_date(raw.get('arrival_date'))
        issues=[];errors=[]
        if not tag:errors.append('Küpe numarası boş')
        elif len(tag)<6:errors.append('Küpe numarası çok kısa')
        if tag in seen:errors.append('Dosya içinde mükerrer küpe')
        if tag:seen.add(tag)
        if tag in existing:errors.append('Küpe sistemde zaten kayıtlı')
        if not gender:errors.append('Cinsiyet okunamadı')
        if raw.get('birth_date') not in (None,'') and birth is None:errors.append('Doğum tarihi geçersiz')
        if raw.get('arrival_date') not in (None,'') and arrival is None:issues.append('Geliş tarihi okunamadı')
        record_type=gender or 'Dişi'
        if birth and 0<=months_old(birth)<10:
            if mother and mother in known_females:
                record_type='Buzağı'
            else:issues.append('10 aydan küçük; anne kaydı bulunamadığı için cinsiyet listesine aktarılacak')
        if not str(raw.get('breed') or '').strip():issues.append('Irk boş')
        state='error' if errors else ('warning' if issues else 'ready')
        prepared.append({
            'row_no':index,'tag':tag,'herd_no':clean_text(raw.get('herd_no')),
            'species':clean_text(raw.get('species')) or 'Sığır','breed':clean_text(raw.get('breed')),
            'gender':gender,'birth_date':birth or '','mother_tag':mother,'arrival_date':arrival or '',
            'nickname':clean_text(raw.get('nickname')),'paddock':clean_text(raw.get('paddock')),
            'notes':clean_text(raw.get('notes')),'record_type':record_type,'state':state,
            'issues':errors+issues,'valid':not errors,
        })
    return prepared

def animal_report_rows(group='all',status='Aktif',search='',paddock=''):
    with db() as c:
        adults=c.execute('''select a.*,
            (select m.tag from calves oldc join animals m on m.id=oldc.mother_id where oldc.promoted_animal_id=a.id limit 1) mother_tag
            from animals a order by a.tag''').fetchall()
        calves=c.execute('''select ca.*,m.tag mother_tag from calves ca join animals m on m.id=ca.mother_id
                            where ca.promoted_animal_id is null order by ca.tag''').fetchall()
    result=[]
    for r in adults:
        rec_status=str(r['status'] or 'Aktif'); rec_group='Dişi' if r['gender']=='Dişi' else 'Erkek'
        if status!='Tümü' and rec_status!=status:continue
        result.append({'tag':r['tag'],'nickname':r['nickname'] or '','group':rec_group,'species':'Sığır','breed':r['breed'] or '',
                       'gender':r['gender'] or '','birth_date':r['birth_date'] or '','mother_tag':r['mother_tag'] or '',
                       'arrival_date':r['purchase_date'] or '','paddock':r['paddock'] or '','status':rec_status})
    if status in ('Aktif','Tümü'):
        for r in calves:
            result.append({'tag':r['tag'],'nickname':r['nickname'] or '','group':'Buzağı','species':'Sığır','breed':r['breed'] or '',
                           'gender':r['gender'] or '','birth_date':r['birth_date'] or '','mother_tag':r['mother_tag'] or '',
                           'arrival_date':r['purchase_date'] or '','paddock':r['paddock'] or '','status':'Aktif'})
    wanted={'female':'Dişi','male':'Erkek','calves':'Buzağı'}.get(group)
    if wanted:result=[r for r in result if r['group']==wanted]
    if paddock:result=[r for r in result if str(r['paddock']).strip()==paddock]
    term=normalized_heading(search)
    if term:
        result=[r for r in result if term in normalized_heading(' '.join(str(r.get(k,'')) for k in ('tag','nickname','breed','gender','mother_tag','paddock')))]
    return sorted(result,key=lambda r:(r['group'],r['tag']))

ANIMAL_REPORT_COLUMNS = [
    ('row_no','Sıra',7,7),
    ('tag','Küpe No',20,31),
    ('nickname','Takma Ad',18,32),
    ('group','Grup',12,16),
    ('species','Tür',10,13),
    ('breed','Irk',24,29),
    ('gender','Cinsiyet',12,18),
    ('birth_date','Doğum Tarihi',15,23),
    ('mother_tag','Ana No',20,31),
    ('arrival_date','Geliş Tarihi',15,23),
    ('paddock','Padok',18,17),
    ('status','Durum',12,17),
]
ANIMAL_REPORT_COLUMN_KEYS = {x[0] for x in ANIMAL_REPORT_COLUMNS}
ANIMAL_REPORT_DEFAULT_COLUMNS = [x[0] for x in ANIMAL_REPORT_COLUMNS]

def animal_report_selected_columns(query=None):
    """Sütun sırasını sabit tutar ve geçersiz sorgu değerlerini dışarıda bırakır."""
    query=query or {}
    requested=query.get('columns',[]) if query.get('columns_mode',[''])[0]=='custom' else ANIMAL_REPORT_DEFAULT_COLUMNS
    requested={str(x) for x in requested if str(x) in ANIMAL_REPORT_COLUMN_KEYS}
    requested.add('tag')  # Küpe numarası her raporda kaydı ayırt etmek için bulunur.
    return [x for x in ANIMAL_REPORT_COLUMNS if x[0] in requested]

def animal_report_display_value(row,key,index):
    if key=='row_no':return str(index)
    if key in ('birth_date','arrival_date'):return fmt_date(row.get(key)) or '-'
    return str(row.get(key) or '-')

def animal_report_xlsx(rows,profile,columns=None):
    from openpyxl import Workbook
    from openpyxl.styles import Font,PatternFill,Border,Side,Alignment
    columns=columns or ANIMAL_REPORT_COLUMNS
    last_col=max(1,len(columns));last_letter=''
    n=last_col
    while n:
        n,rem=divmod(n-1,26);last_letter=chr(65+rem)+last_letter
    wb=Workbook();ws=wb.active;ws.title='Hayvanlar'
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=last_col);ws['A1']=farm_display_name(profile)+' · Tüm Hayvanlar Raporu';ws['A1'].font=Font(size=16,bold=True,color='176B3A')
    ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=last_col);ws['A2']=f"İşletme No: {profile.get('business_no') or '-'}   |   Rapor Tarihi: {date.today().strftime('%d/%m/%Y')}   |   Toplam: {len(rows)}"
    for col,(_,value,_,_) in enumerate(columns,1):
        cell=ws.cell(4,col,value);cell.font=Font(bold=True,color='FFFFFF');cell.fill=PatternFill('solid',fgColor='176B3A');cell.alignment=Alignment(horizontal='center')
    for idx,row in enumerate(rows,1):
        for col,(key,_,_,_) in enumerate(columns,1):
            value=idx if key=='row_no' else row.get(key) or ''
            if key in ('birth_date','arrival_date') and value:
                try:value=date.fromisoformat(value);ws.cell(idx+4,col).number_format='dd/mm/yyyy'
                except Exception:pass
            ws.cell(idx+4,col,value)
    for idx,(_,_,width,_) in enumerate(columns,1):
        col_letter='';n=idx
        while n:n,rem=divmod(n-1,26);col_letter=chr(65+rem)+col_letter
        ws.column_dimensions[col_letter].width=width
    thin=Side(style='thin',color='D9E5DD')
    for row in ws.iter_rows(min_row=4,max_row=4+len(rows),min_col=1,max_col=last_col):
        for cell in row:cell.border=Border(bottom=thin);cell.alignment=Alignment(vertical='center',wrap_text=True)
    ws.freeze_panes='A5';ws.auto_filter.ref=f'A4:{last_letter}{max(4,4+len(rows))}';ws.print_title_rows='1:4';ws.sheet_properties.pageSetUpPr.fitToPage=True
    ws.page_setup.orientation='landscape';ws.page_setup.fitToWidth=1;ws.page_setup.fitToHeight=0
    output=io.BytesIO();wb.save(output);wb.close();return output.getvalue()

def animal_report_pdf(rows,profile,subtitle='Aktif Kayıtlar · Tüm Hayvanlar',columns=None):
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_LEFT,TA_RIGHT
    from reportlab.lib.pagesizes import A4,landscape
    from reportlab.lib.styles import ParagraphStyle,getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.platypus import SimpleDocTemplate,Table,TableStyle,Paragraph,Spacer,Image as PdfImage
    import reportlab
    fonts_dir=Path(reportlab.__file__).resolve().parent/'fonts'
    regular=fonts_dir/'Vera.ttf';bold=fonts_dir/'VeraBd.ttf'
    try:
        if 'CiftlikVera' not in pdfmetrics.getRegisteredFontNames():pdfmetrics.registerFont(TTFont('CiftlikVera',str(regular)))
        if 'CiftlikVeraBold' not in pdfmetrics.getRegisteredFontNames():pdfmetrics.registerFont(TTFont('CiftlikVeraBold',str(bold)))
        font_name='CiftlikVera';bold_name='CiftlikVeraBold'
    except Exception:
        font_name='Helvetica';bold_name='Helvetica-Bold'
    columns=columns or ANIMAL_REPORT_COLUMNS
    output=io.BytesIO();page_size=landscape(A4)
    doc=SimpleDocTemplate(output,pagesize=page_size,leftMargin=9*mm,rightMargin=9*mm,topMargin=9*mm,bottomMargin=12*mm,
                          title='Tüm Hayvanlar Raporu',author='ÇiftlikPro Enterprise')
    styles=getSampleStyleSheet()
    title=ParagraphStyle('ReportTitle',parent=styles['Title'],fontName=bold_name,fontSize=17,leading=20,textColor=colors.HexColor('#173b28'),alignment=TA_LEFT,spaceAfter=2)
    subtitle_style=ParagraphStyle('ReportSubtitle',parent=styles['Normal'],fontName=font_name,fontSize=8.5,leading=11,textColor=colors.HexColor('#607168'))
    meta_style=ParagraphStyle('ReportMeta',parent=styles['Normal'],fontName=font_name,fontSize=8.2,leading=12,alignment=TA_RIGHT,textColor=colors.HexColor('#243d30'))
    cell_style=ParagraphStyle('ReportCell',parent=styles['Normal'],fontName=font_name,fontSize=6.7,leading=8,textColor=colors.HexColor('#1c3427'))
    cell_bold=ParagraphStyle('ReportCellBold',parent=cell_style,fontName=bold_name)
    head_style=ParagraphStyle('ReportHead',parent=cell_style,fontName=bold_name,fontSize=6.7,leading=8,textColor=colors.white)
    brand_parts=[]
    logo_url=(profile.get('farm_logo') or '').strip();logo_path=UPLOADS/os.path.basename(logo_url) if logo_url.startswith('/uploads/') else None
    if logo_path and logo_path.exists():
        try:brand_parts.append(PdfImage(str(logo_path),width=18*mm,height=18*mm,kind='proportional'))
        except Exception:pass
    title_block=[Paragraph(h(farm_display_name(profile)),title),Paragraph('İşletmede Bulunan Hayvanlar Raporu',ParagraphStyle('ReportName',parent=title,fontSize=12,leading=14,spaceAfter=1)),Paragraph(h(subtitle),subtitle_style)]
    title_table=Table([[brand_parts[0] if brand_parts else Paragraph('CP',ParagraphStyle('LogoFallback',parent=title,fontSize=18,textColor=colors.HexColor('#176b3a'))),title_block]],colWidths=[21*mm,115*mm])
    title_table.setStyle(TableStyle([('VALIGN',(0,0),(-1,-1),'MIDDLE'),('LEFTPADDING',(0,0),(-1,-1),0),('RIGHTPADDING',(0,0),(0,0),3*mm),('TOPPADDING',(0,0),(-1,-1),0),('BOTTOMPADDING',(0,0),(-1,-1),0)]))
    location=' / '.join(x for x in ((profile.get('province') or '').strip(),(profile.get('district') or '').strip()) if x) or '-'
    meta=Paragraph(f"<b>Rapor Tarihi:</b> {date.today().strftime('%d/%m/%Y')}<br/><b>İşletme No:</b> {h(profile.get('business_no') or '-')}<br/><b>İşletme Sahibi:</b> {h(profile.get('owner_name') or '-')}<br/>{h(location)}",meta_style)
    header=Table([[title_table,meta]],colWidths=[doc.width*.68,doc.width*.32])
    header.setStyle(TableStyle([('VALIGN',(0,0),(-1,-1),'MIDDLE'),('LEFTPADDING',(0,0),(-1,-1),0),('RIGHTPADDING',(0,0),(-1,-1),0),('TOPPADDING',(0,0),(-1,-1),0),('BOTTOMPADDING',(0,0),(-1,-1),3*mm),('LINEBELOW',(0,0),(-1,-1),1.5,colors.HexColor('#176b3a'))]))
    female=sum(1 for r in rows if r['group']=='Dişi');male=sum(1 for r in rows if r['group']=='Erkek');calf=sum(1 for r in rows if r['group']=='Buzağı')
    summary=Paragraph(f'<b>Toplam Hayvan: {len(rows)}</b> &nbsp;&nbsp;&nbsp; Dişi: {female} &nbsp;&nbsp;&nbsp; Erkek: {male} &nbsp;&nbsp;&nbsp; Buzağı: {calf}',ParagraphStyle('Summary',parent=subtitle_style,fontSize=8.5,textColor=colors.HexColor('#173b28')))
    data=[[Paragraph(h(label),head_style) for _,label,_,_ in columns]]
    for idx,row in enumerate(rows,1):
        data.append([Paragraph(h(animal_report_display_value(row,key,idx)),cell_bold if key=='tag' else cell_style) for key,_,_,_ in columns])
    raw_widths=[pdf_width for _,_,_,pdf_width in columns];scale=doc.width/sum(raw_widths)
    widths=[value*scale for value in raw_widths]
    table=Table(data,colWidths=widths,repeatRows=1,splitByRow=1,hAlign='LEFT')
    table.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.HexColor('#176b3a')),('VALIGN',(0,0),(-1,-1),'MIDDLE'),('GRID',(0,0),(-1,-1),0.35,colors.HexColor('#cfded4')),('LEFTPADDING',(0,0),(-1,-1),2.4),('RIGHTPADDING',(0,0),(-1,-1),2.4),('TOPPADDING',(0,0),(-1,-1),3),('BOTTOMPADDING',(0,0),(-1,-1),3),('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white,colors.HexColor('#f5f8f6')])]))
    def footer(canvas,document):
        canvas.saveState();canvas.setStrokeColor(colors.HexColor('#d2ded6'));canvas.line(document.leftMargin,8*mm,page_size[0]-document.rightMargin,8*mm)
        canvas.setFont(font_name,6.7);canvas.setFillColor(colors.HexColor('#68776e'));canvas.drawString(document.leftMargin,4.8*mm,'ÇiftlikPro Enterprise · '+farm_display_name(profile));canvas.drawRightString(page_size[0]-document.rightMargin,4.8*mm,f'Sayfa {canvas.getPageNumber()} · {len(rows)} kayıt');canvas.restoreState()
    doc.build([header,Spacer(1,3*mm),summary,Spacer(1,2.5*mm),table],onFirstPage=footer,onLaterPages=footer)
    return output.getvalue()


DASHBOARD_CARD_OPTIONS = [
    ('active_total','🐄 Toplam Aktif Hayvan'),
    ('female','🐮 Dişi Hayvan'),
    ('male','🐂 Erkek Hayvan'),
    ('pregnant','🤰 Gebe Hayvan'),
    ('calves','🐮 Buzağı'),
    ('due','📅 Yaklaşan Doğum'),
    ('estrus','🌸 Yaklaşan Kızgınlık'),
    ('health_due','💉 Yaklaşan Sağlık İşlemleri'),
    ('income','📥 Toplam Gelir'),
    ('expense','📤 Toplam Gider'),
    ('net','⚖️ Net Durum'),
]
DASHBOARD_DEFAULT_LAYOUT=['active_total','female','male','pregnant','calves','due','estrus','health_due']

def dashboard_layout(username):
    key='dashboard_layout_'+str(username)
    try:
        with db() as c:
            r=c.execute("select setting_value from settings where setting_key=?",(key,)).fetchone()
        if r and r['setting_value']:
            vals=(r['setting_value'].split(',')+['']*8)[:8]
            if vals[:7]==['active_total','female','male','pregnant','calves','due','estrus'] and not vals[7]:
                vals[7]='health_due'
            valid={x[0] for x in DASHBOARD_CARD_OPTIONS}
            return [v if v in valid else '' for v in vals]
    except Exception:
        pass
    return DASHBOARD_DEFAULT_LAYOUT[:]

def h(s):
    return str(s or '').replace('&','&amp;').replace('<','&lt;').replace('>','&gt;').replace('"','&quot;')

def money(v):
    return f"₺{float(v or 0):,.2f}".replace(',','X').replace('.',',').replace('X','.')

# Kullanıcıya gösterilen tarihler GG/AA/YYYY; form/veritabanı ISO kalır.
def fmt_date(v):
    if not v:return ''
    try:return datetime.strptime(str(v)[:10],'%Y-%m-%d').strftime('%d/%m/%Y')
    except Exception:return str(v)

def fmt_datetime(v):
    if not v:return ''
    try:
        x=str(v).replace('T',' ')
        d=datetime.fromisoformat(x)
        return d.strftime('%d/%m/%Y %H:%M')
    except Exception:
        return fmt_date(v)

def current_pregnancy_record(c, animal_id):
    latest=c.execute("select * from inseminations where animal_id=? order by insemination_date desc,id desc limit 1",(animal_id,)).fetchone()
    if not latest or not is_pregnant_value(latest['pregnancy_result']): return None
    ins_date=(latest['insemination_date'] or '').strip()
    if ins_date and c.execute("select id from calves where mother_id=? and birth_date>=? order by birth_date desc,id desc limit 1",(animal_id,ins_date)).fetchone(): return None
    return latest

def is_currently_pregnant(c, animal_id):
    return current_pregnancy_record(c,animal_id) is not None

def age_text(d):
    if not d:return '-'
    try:
        b=date.fromisoformat(d); days=(date.today()-b).days
        if days<60:return f'{days} gün'
        if days<730:return f'{days//30} ay'
        return f'{days//365} yıl {(days%365)//30} ay'
    except:return '-'

def months_old(d):
    try:
        b=date.fromisoformat(d); t=date.today()
        return (t.year-b.year)*12 + t.month-b.month - (1 if t.day < b.day else 0)
    except:return -1

def record_ration_item_history(c, ration_id, feed_id, kg_per_head_day, effective_date=None, notes='Rasyon düzenleme'):
    # Bir rasyon kaleminin tarihli durumunu saklar; kg=0 silinmiş kalemdir.
    eff=(effective_date or date.today().isoformat())[:10]
    kg=max(0.0,float(kg_per_head_day or 0))
    # Aynı gün aynı yem için son durum yeterlidir; geçmiş günleri değiştirmeyiz.
    c.execute('insert into ration_item_history(ration_id,feed_id,effective_date,kg_per_head_day,created_at,notes) values(?,?,?,?,?,?)',
              (int(ration_id),int(feed_id),eff,kg,datetime.now().isoformat(timespec='seconds'),notes))


def ration_items_on_date(c, ration_id, on_date):
    d=str(on_date)[:10]
    return c.execute('''select h.feed_id,h.kg_per_head_day,f.*
        from ration_item_history h
        join feed_catalog f on f.id=h.feed_id
        where h.ration_id=? and h.effective_date<=?
          and h.id=(select h2.id from ration_item_history h2
                    where h2.ration_id=h.ration_id and h2.feed_id=h.feed_id and h2.effective_date<=?
                    order by h2.effective_date desc,h2.id desc limit 1)
          and h.kg_per_head_day>0
        order by f.name''',(ration_id,d,d)).fetchall()


def ration_cost_on_date(c, ration_id, on_date):
    d=str(on_date)[:10]
    total=0.0
    for r in ration_items_on_date(c,ration_id,d):
        total += float(r['kg_per_head_day'] or 0) * current_feed_price(r['feed_id'],c,d)
    return total


def ration_cost_between(c, ration_id, start_day, end_day):
    # [start_day,end_day) aralığındaki tarihsel rasyon yem maliyetini hesaplar.
    if end_day<=start_day:return 0.0
    s0=start_day.isoformat(); e0=end_day.isoformat()
    # Rasyon miktarı veya ilgili yem fiyatı değiştiği gün yeni maliyet dönemi başlar.
    feed_ids=[int(r['feed_id']) for r in c.execute('select distinct feed_id from ration_item_history where ration_id=?',(ration_id,)).fetchall()]
    bounds={start_day,end_day}
    for r in c.execute('select distinct effective_date from ration_item_history where ration_id=? and effective_date>? and effective_date<?',(ration_id,s0,e0)).fetchall():
        try:bounds.add(date.fromisoformat(r['effective_date'][:10]))
        except Exception:pass
    if feed_ids:
        ph=','.join('?'*len(feed_ids))
        rows=c.execute(f'select distinct effective_date from feed_prices where feed_id in ({ph}) and effective_date>? and effective_date<?',(*feed_ids,s0,e0)).fetchall()
        for r in rows:
            try:bounds.add(date.fromisoformat(r['effective_date'][:10]))
            except Exception:pass
    pts=sorted(bounds); total=0.0
    for a,b in zip(pts,pts[1:]):
        days=max(0,(b-a).days)
        if days: total += days*ration_cost_on_date(c,ration_id,a.isoformat())
    return total


def animal_paddock_intervals(c, a, start_day, end_day):
    # Hayvanın [start,end) dönemindeki padoklarını tarih aralıkları halinde döndürür.
    aid=int(a['id']); current_pid=a['paddock_id'] if 'paddock_id' in a.keys() else None
    rows=c.execute("select * from paddock_history where animal_source='animal' and animal_id=? order by moved_at,id",(aid,)).fetchall()
    # Başlangıç günündeki padoku bul.
    pid=current_pid
    before=[]; after=[]
    for r in rows:
        try:md=date.fromisoformat(str(r['moved_at'])[:10])
        except Exception:continue
        if md<=start_day: before.append((md,r))
        else: after.append((md,r))
    if before:
        pid=before[-1][1]['to_paddock_id']
    elif after:
        # İlk kayıt, hareket öncesindeki padoku da biliyor.
        pid=after[0][1]['from_paddock_id']
    intervals=[]; cur=start_day
    for md,r in after:
        if md>=end_day:break
        if md>cur:intervals.append((cur,md,pid))
        pid=r['to_paddock_id'];cur=max(cur,md)
    if cur<end_day:intervals.append((cur,end_day,pid))
    return intervals


def paddock_feed_cost_between(c, paddock_id, start_day, end_day, manual_daily_feed):
    # Padokta rasyon tanımlı dönemlerde rasyon, boşluklarda manuel sabit maliyet kullanılır.
    if end_day<=start_day:return 0.0,0,0
    if not paddock_id:
        days=(end_day-start_day).days
        return days*manual_daily_feed,0,days
    rows=c.execute('''select * from paddock_rations
        where paddock_id=? and start_date<? and (end_date is null or end_date='' or end_date>=?)
        order by start_date,id''',(paddock_id,end_day.isoformat(),start_day.isoformat())).fetchall()
    total=0.0;ration_days=0;manual_days=0;cur=start_day
    for pr in rows:
        try:rs=max(start_day,date.fromisoformat(pr['start_date'][:10]))
        except Exception:continue
        try:re=min(end_day,date.fromisoformat(pr['end_date'][:10])+timedelta(days=1)) if pr['end_date'] else end_day
        except Exception:re=end_day
        if re<=rs:continue
        if rs>cur:
            d=(rs-cur).days;total+=d*manual_daily_feed;manual_days+=d
        seg_start=max(cur,rs)
        if re>seg_start:
            d=(re-seg_start).days
            total+=ration_cost_between(c,int(pr['ration_id']),seg_start,re);ration_days+=d
            cur=max(cur,re)
        if cur>=end_day:break
    if cur<end_day:
        d=(end_day-cur).days;total+=d*manual_daily_feed;manual_days+=d
    return total,ration_days,manual_days


def animal_current_feed_context(a, con=None, on_date=None):
    own=con is None;c=con or db().__enter__()
    try:
        manual=float(a['daily_feed_cost'] or 0) if 'daily_feed_cost' in a.keys() else 0.0
        pid=a['paddock_id'] if 'paddock_id' in a.keys() else None
        d=(on_date or date.today().isoformat())[:10]
        if pid:
            pr=c.execute('''select pr.*,r.name ration_name from paddock_rations pr join rations r on r.id=pr.ration_id
                where pr.paddock_id=? and pr.start_date<=? and (pr.end_date is null or pr.end_date='' or pr.end_date>=?)
                order by pr.start_date desc,pr.id desc limit 1''',(pid,d,d)).fetchone()
            if pr:
                return {'feed_cost':ration_cost_on_date(c,int(pr['ration_id']),d),'source':'ration','ration_name':pr['ration_name'],'start_date':pr['start_date'],'paddock_id':pid}
        return {'feed_cost':manual,'source':'manual','ration_name':'','start_date':'','paddock_id':pid}
    finally:
        if own:c.close()


def animal_cost_values(a, con=None):
    purchase=float(a['purchase_price'] or 0) if 'purchase_price' in a.keys() else 0.0
    manual_feed=float(a['daily_feed_cost'] or 0) if 'daily_feed_cost' in a.keys() else 0.0
    care=float(a['daily_care_cost'] or 0) if 'daily_care_cost' in a.keys() else 0.0
    start=(a['purchase_date'] if 'purchase_date' in a.keys() else '') or (a['birth_date'] if 'birth_date' in a.keys() else '')
    status=str(a['status'] or 'Aktif') if 'status' in a.keys() else 'Aktif'
    exit_date=(a['exit_date'] if 'exit_date' in a.keys() else '') or ''
    try:
        start_date=date.fromisoformat(start);end_date=date.today()
        if status!='Aktif' and exit_date:end_date=min(date.today(),date.fromisoformat(exit_date))
        days=max(0,(end_date-start_date).days)
    except Exception:
        days=0;start_date=end_date=date.today()
    own=con is None;c=con or db().__enter__()
    try:
        feed_acc=0.0
        for a0,b0,pid in animal_paddock_intervals(c,a,start_date,end_date):
            part,_,_=paddock_feed_cost_between(c,pid,a0,b0,manual_feed);feed_acc+=part
        accumulated=feed_acc + days*care
        ctx=animal_current_feed_context(a,c,end_date.isoformat())
        daily=float(ctx['feed_cost'])+care
        return days,daily,accumulated,purchase+accumulated
    finally:
        if own:c.close()


def pregnancy_vaccine_tasks(con, animal_id=None, horizon_days=7):
    # Aktif gebelikler için 7. ve 8. ay aşı görevlerini üretir.
    params=[]
    where="a.gender='Dişi' and coalesce(a.status,'Aktif')='Aktif'"
    if animal_id is not None:
        where+=' and a.id=?'; params.append(animal_id)
    rows=con.execute(f'''select i.*,a.tag,a.nickname,a.status
        from inseminations i join animals a on a.id=i.animal_id
        where {where}
        order by i.animal_id,i.insemination_date desc,i.id desc''',params).fetchall()
    latest_by_animal={}
    for r in rows:
        if r['animal_id'] not in latest_by_animal:
            latest_by_animal[r['animal_id']]=r
    today=date.today(); horizon=today+timedelta(days=horizon_days)
    tasks=[]
    for r in latest_by_animal.values():
        if not is_pregnant_value(r['pregnancy_result']):
            continue
        try:
            insemination_day=date.fromisoformat(r['insemination_date'])
            due_day=date.fromisoformat(r['due_date']) if r['due_date'] else insemination_day+timedelta(days=280)
        except Exception:
            continue
        if due_day < today:
            continue
        for month,offset in ((7,210),(8,240)):
            task_day=insemination_day+timedelta(days=offset)
            if task_day>horizon:
                continue
            token=f'GEBELIK_ASI|{r["id"]}|{month}'
            done=con.execute("select 1 from health where animal_id=? and notes like ? limit 1",(r['animal_id'],token+'%')).fetchone()
            if done:
                continue
            days_left=(task_day-today).days
            tasks.append({'animal_id':r['animal_id'],'insemination_id':r['id'],'tag':r['tag'],'nickname':r['nickname'],'month':month,'task_date':task_day.isoformat(),'days_left':days_left,'token':token,'overdue':days_left<0,'today':days_left==0})
    return sorted(tasks,key=lambda x:(x['task_date'],x['tag']))


def setting_float(key, default):
    try:
        with db() as c:
            row=c.execute('select setting_value from settings where setting_key=?',(key,)).fetchone()
        return float(row['setting_value']) if row and row['setting_value'] not in (None,'') else float(default)
    except Exception:
        return float(default)


def current_feed_price(feed_id, con=None, on_date=None):
    own=con is None
    c=con or db().__enter__()
    try:
        d=(on_date or date.today().isoformat())[:10]
        r=c.execute("select price_per_kg from feed_prices where feed_id=? and effective_date<=? order by effective_date desc,id desc limit 1",(feed_id,d)).fetchone()
        return float(r['price_per_kg'] or 0) if r else 0.0
    finally:
        if own:c.close()

def feed_stock_kg(feed_id, con=None):
    own=con is None
    c=con or db().__enter__()
    try:
        r=c.execute("select coalesce(sum(case when tx_type in ('Giriş','Sayım +') then quantity_kg when tx_type in ('Çıkış','Tüketim','Sayım -') then -quantity_kg else 0 end),0) qty from feed_stock_transactions where feed_id=?",(feed_id,)).fetchone()
        return float(r['qty'] or 0)
    finally:
        if own:c.close()

def feed_group(feed):
    """Rasyon kaba/kesif oranı için pratik yem grubu sınıflaması.
    Mineral/katkılar oran hesabına dahil edilmez. Katalog kategorileri eski Excel'de
    tek tip geldiği için isim + NDF tabanlı güvenli bir ön sınıflama kullanılır.
    """
    name=str(feed['name'] or '').upper()
    ndf=float(feed['ndf_pct'] or 0)
    additive_words=('TUZ','FOSFAT','KİREÇ','KIREC','MERMER','KALSİYUM','KALSIYUM','PREMİKS','PREMIKS','MİNERAL','MINERAL','VİTAMİN','VITAMIN','VİTAMİN PREMİKS','VITAMIN PREMIX','VİT.','VIT.','VİT.-MİN','VIT.-MIN','KAVIMIX','SODYUM BİKARBONAT','SODYUM BIKARBONAT','AMONYUM','ÜRE','URE')
    if any(w in name for w in additive_words): return 'Katkı'
    # HOTFIX 6.12: pamuk tohumu / pamuk küspesi yüksek NDF içerdiği için eski
    # NDF>=35 geri dönüşü bunları yanlışlıkla kaba yem sayıyordu. Bunlar rasyonda
    # enerji/protein konsantresi olarak değerlendirilir; pamuk kabuğu/çırçır artığı kaba yemdir.
    if 'PAMUK TOHUMU' in name and not any(w in name for w in ('KAPÇIĞI','KAPCIGI','KABUĞU','KABUGU','ÇIRÇIR','CIRCIR','HULL')):
        return 'Kesif'
    rough_words=('SAMAN','SİLAJ','SILAJ','KURU OT','YONCA','MERA','ÇAYIR','OTU','HASIL','FİĞ','FIG','ÇİM','CIM')
    concentrate_words=('YEMİ','YEMI','KÜSPE','KUSPE','SOYA','KANOLA','AYÇİÇEĞİ','AYCICEGI','ARPA,','MISIR,','MISIR DANE','MISIR KIRMA','BUĞDAY,','BUGDAY,','KEPEK','KEPEĞ','KEPEG','MELAS','FLAKED','PULU')
    if any(w in name for w in rough_words): return 'Kaba'
    if any(w in name for w in concentrate_words): return 'Kesif'
    return 'Kaba' if ndf>=35 else 'Kesif'

def ration_summary(ration_id, con=None):
    own=con is None
    c=con or db().__enter__()
    try:
        sql="""select ri.id item_id,ri.kg_per_head_day,f.*,coalesce((select fp.price_per_kg from feed_prices fp where fp.feed_id=f.id and fp.effective_date<=? order by fp.effective_date desc,fp.id desc limit 1),0) price
                 from ration_items ri join feed_catalog f on f.id=ri.feed_id where ri.ration_id=? order by f.name"""
        rows=c.execute(sql,(date.today().isoformat(),ration_id)).fetchall()
        out={'as_fed_kg':0.0,'dm_kg':0.0,'cp_kg':0.0,'ndf_kg':0.0,'endf_kg':0.0,'starch_kg':0.0,
             'rapid_starch_kg':0.0,'known_degradability_starch_kg':0.0,'tdn_kg':0.0,'me_mcal':0.0,'nem_mcal':0.0,'neg_mcal':0.0,
             'ca_g':0.0,'p_g':0.0,'cost':0.0,'roughage_dm_kg':0.0,'concentrate_dm_kg':0.0,'additive_dm_kg':0.0,'items':rows}
        for r in rows:
            kg=float(r['kg_per_head_day'] or 0); dm=kg*float(r['dm_pct'] or 0)/100.0
            out['as_fed_kg']+=kg; out['dm_kg']+=dm
            out['cp_kg']+=dm*float(r['cp_pct'] or 0)/100.0
            out['ndf_kg']+=dm*float(r['ndf_pct'] or 0)/100.0
            out['endf_kg']+=dm*float(r['ndf_pct'] or 0)/100.0*float(r['effective_ndf_pct'] or 0)/100.0
            # Ekran, öneriler ve solver aynı güvenli nişasta değerini kullanır.
            # Böylece eski katalog aktarımlarındaki bariz kolon kaymaları rasyon özetini bozmaz.
            feed_starch=dm*_solver_starch_pct(r)/100.0; out['starch_kg']+=feed_starch
            starch_deg,known_deg=_feed_starch_degradability(r)
            if known_deg:
                out['known_degradability_starch_kg']+=feed_starch
                out['rapid_starch_kg']+=feed_starch*starch_deg/100.0
            out['tdn_kg']+=dm*float(r['tdn_pct'] or 0)/100.0
            out['me_mcal']+=dm*float(r['me_mcal_kg'] or 0)
            out['nem_mcal']+=dm*float(r['nem_mcal_kg'] or 0)
            out['neg_mcal']+=dm*float(r['neg_mcal_kg'] or 0)
            out['ca_g']+=dm*float(r['ca_pct'] or 0)*10.0
            out['p_g']+=dm*float(r['p_pct'] or 0)*10.0
            out['cost']+=kg*float(r['price'] or 0)
            grp=feed_group(r)
            if grp=='Kaba': out['roughage_dm_kg']+=dm
            elif grp=='Kesif': out['concentrate_dm_kg']+=dm
            else: out['additive_dm_kg']+=dm
        out['cp_pct_dm']=(out['cp_kg']/out['dm_kg']*100) if out['dm_kg'] else 0.0
        out['ndf_pct_dm']=(out['ndf_kg']/out['dm_kg']*100) if out['dm_kg'] else 0.0
        out['endf_pct_dm']=(out['endf_kg']/out['dm_kg']*100) if out['dm_kg'] else 0.0
        out['starch_pct_dm']=(out['starch_kg']/out['dm_kg']*100) if out['dm_kg'] else 0.0
        out['rapid_starch_pct_dm']=(out['rapid_starch_kg']/out['dm_kg']*100) if out['dm_kg'] else 0.0
        out['starch_degradability_coverage']=(out['known_degradability_starch_kg']/out['starch_kg']) if out['starch_kg'] else 1.0
        out['me_per_kg_dm']=(out['me_mcal']/out['dm_kg']) if out['dm_kg'] else 0.0
        out['nem_density']=(out['nem_mcal']/out['dm_kg']) if out['dm_kg'] else 0.0
        out['neg_density']=(out['neg_mcal']/out['dm_kg']) if out['dm_kg'] else 0.0
        rc_dm=out['roughage_dm_kg']+out['concentrate_dm_kg']
        out['roughage_pct_dm']=(out['roughage_dm_kg']/rc_dm*100) if rc_dm else 0.0
        out['concentrate_pct_dm']=(out['concentrate_dm_kg']/rc_dm*100) if rc_dm else 0.0
        return out
    finally:
        if own:c.close()

def _interp(x, xs, ys):
    """Küçük referans tabloları için doğrusal enterpolasyon/ölçülü ekstrapolasyon."""
    x=float(x)
    if x<=xs[0]:
        x0,x1=xs[0],xs[1]; y0,y1=ys[0],ys[1]
    elif x>=xs[-1]:
        x0,x1=xs[-2],xs[-1]; y0,y1=ys[-2],ys[-1]
    else:
        for i in range(len(xs)-1):
            if xs[i]<=x<=xs[i+1]: x0,x1=xs[i],xs[i+1]; y0,y1=ys[i],ys[i+1]; break
    return y0+(x-x0)*(y1-y0)/(x1-x0)

def _beef_phase_from_weight_age(weight_kg, age_months=0):
    """Yaş tek başına gereksinim belirlemez; ağırlık ana parametredir. Yaş yalnız faz kontrolüne yardım eder."""
    w=float(weight_kg); age=float(age_months or 0)
    # DEV3: üç ana besi fazı. Uyum/hazırlık ağırlıktan bağımsız bir yönetim
    # dönemi olduğundan kullanıcı ileride 'uyum' seçeneği verdiğinde ayrıca uygulanabilir.
    if w<300: phase='Besi Başlangıç'
    elif w<450: phase='Besi Geliştirme'
    else: phase='Besi Bitirme'
    age_note=''
    if age>0:
        if w<300 and age>=15: age_note='Canlı ağırlık yaşa göre düşük; tartım, sağlık ve büyüme performansı kontrol edilmeli.'
        elif w>=500 and age<=12: age_note='Canlı ağırlık yaşa göre yüksek; hayvan tipi/ırk ve tartım doğrulanmalı.'
    return phase,age_note

def nasem_dynamic_dmi(sbw_kg, nem_density, age_months=0, weight_kg=0):
    """NASEM 2016 growing cattle DMI prediction (kg DM/day).

    DMI is diet-energy-density dependent.  The calf/yearling equations differ only
    in the intercept term.  Age is used when supplied; otherwise weight is used
    as a conservative phase proxy.  Environmental/breed/additive multipliers are
    left at 1.0 until those inputs exist in CiftlikPro.
    """
    sbw=max(1.0,float(sbw_kg or 0)); nem=max(0.70,min(float(nem_density or 1.55),2.50))
    nema=max(nem,0.95)
    age=float(age_months or 0); w=float(weight_kg or 0)
    is_yearling=(age>=12.0) if age>0 else (w>=300.0)
    intercept=0.0869 if is_yearling else 0.1128
    dmi=(sbw**0.75)*(0.2435*nema-0.0466*(nema**2)-intercept)/nema
    return max(0.0,dmi)

def _reference_nem_density(weight_kg, age_months=0):
    """Pre-solver target-card reference only; actual DMI is recalculated from solved diet NEm."""
    w=float(weight_kg or 0)
    if w<300:return 1.60
    if w<400:return 1.68
    if w<500:return 1.76
    return 1.84

def beef_starch_targets(phase):
    """Besi dönemi için muhafazakâr çalışma/uyarı bandı (% KM).

    İdeal bant solver için yumuşak yönlendirmedir; enerji, NDF/eNDF ve kaba/kesif
    hedeflerinin önüne geçmez. ``starch_max`` adı geriye uyumluluk için korunur,
    fakat değer evrensel fizyolojik üst sınır değildir: üzerinde tek başına çözüm
    reddedilmez; hızlı nişasta, etkili lif, işleme ve adaptasyonla birlikte uyarı
    eşiği olarak değerlendirilir.
    """
    return {
        'Besi Başlangıç':{'starch_min':20.0,'starch_ideal_max':24.0,'starch_max':28.0},
        'Besi Geliştirme':{'starch_min':23.0,'starch_ideal_max':27.0,'starch_max':30.0},
        'Besi Bitirme':{'starch_min':25.0,'starch_ideal_max':29.0,'starch_max':31.0},
    }.get(str(phase),{'starch_min':23.0,'starch_ideal_max':27.0,'starch_max':30.0})

def _beef_animal_profile(animal_type):
    """NASEM Chapter 20 hayvan profilini kullanıcı seçimiyle eşleştirir.

    Türkiye'deki ``Besi Erkek`` kaydı pratikte çoğunlukla kastre edilmemiş
    tosun/boğayı ifade eder. Önceki sürüm bu alanı tamamen görmezden geliyor ve
    her hayvanı orta çerçeveli kastre erkek/düve tablosuyla hesaplıyordu. Bu da
    özellikle bakım enerjisini, büyüme enerjisini, MP ve mineral kartlarını aynı
    anda yanlış profile bağlıyordu.
    """
    name=str(animal_type or '').strip().upper()
    is_intact_male=any(token in name for token in ('BESİ ERKEK','BESI ERKEK','TOSUN','BOĞA','BOGA','BULL'))
    is_castrated=any(token in name for token in ('KASTRE','STEER'))
    if is_intact_male and not is_castrated:
        return {
            'key':'bull','label':'Tosun / Boğa (kastre edilmemiş)',
            'mature_sbw_kg':900.0,'nem_coefficient':0.0885,
            'neg_coefficient':0.0400,
        }
    return {
        'key':'steer_heifer','label':'Düve / kastre erkek / genel büyüyen sığır',
        'mature_sbw_kg':550.0,'nem_coefficient':0.0770,
        # Table 20-1 gereksinim satırlarıyla uyumlu eşdeğer ağırlık katsayısı.
        'neg_coefficient':0.0575,
    }

def ration_requirement_targets(weight_kg=450.0, target_adg=1.3, animal_type='Besi Erkek', age_months=0, phase_override='Otomatik'):
    """Besi Hayvanı İhtiyaç Motoru V1.

    Enerji çekirdeği NASEM 2016 büyüyen/bitirilen sığır yaklaşımına dayanır ve
    seçilen hayvan profiline göre Table 20-1 veya Table 20-2 kullanır:
      SBW = BW*0.96, EBW = 0.891*SBW, EBG = 0.956*ADG
      NEm = profil_katsayısı*SBW^0.75
      RE/NEg = profil_katsayısı*EBW^0.75*EBG^1.097
    MP, Ca ve P için NASEM 2016 Chapter 20 errata referans tablosu enterpolasyonla kullanılır.
    Katalogda RDP/RUP olmadığı için CP hedefi MP'nin tam karşılığı değildir; solverda ayrıca güvenlik taraması olarak tutulur.
    """
    w=max(150.0,min(float(weight_kg or 450),900.0)); adg=max(0.2,min(float(target_adg or 1.3),2.2)); age=max(0.0,min(float(age_months or 0),36.0))
    profile=_beef_animal_profile(animal_type)
    sbw=w*0.96; ebw=0.891*sbw; ebg=max(0.05,0.956*adg)
    nem_req=profile['nem_coefficient']*(sbw**0.75)
    neg_req=profile['neg_coefficient']*(ebw**0.75)*(ebg**1.097)

    adg_grid=[0.4,0.8,1.2,1.6,2.0]
    if profile['key']=='bull':
        # NASEM 2016 Chapter 20 Errata, Table 20-2: growing bulls.
        sbw_grid=[300,400,500,600,700,800]
        mp_maint_vals=[274,340,402,461,517,572]
        ca_maint_vals=[9.2,12.3,15.4,18.5,21.6,24.6]
        p_maint_vals=[7.1,9.4,11.8,14.1,16.5,18.8]
        mp_gain_by_sbw={
            300:[123,241,355,469,581],400:[125,242,355,467,576],500:[128,245,358,468,575],
            600:[126,239,347,451,553],700:[115,215,310,400,487],800:[104,192,273,350,423],
        }
        ca_gain_by_sbw={
            300:[11.4,22.3,32.9,43.4,53.8],400:[10.5,20.3,29.9,39.2,48.4],500:[9.6,18.5,27.0,35.3,43.4],
            600:[8.8,16.7,24.2,31.5,38.6],700:[8.0,15.0,21.6,27.9,34.0],800:[7.3,13.4,19.1,24.5,29.6],
        }
        p_gain_by_sbw={
            300:[4.6,9.0,13.3,17.5,21.7],400:[4.2,8.2,12.1,15.8,19.6],500:[3.9,7.5,10.9,14.2,17.5],
            600:[3.6,6.8,9.8,12.7,15.6],700:[3.2,6.1,8.7,11.3,13.7],800:[2.9,5.4,7.7,9.9,11.9],
        }
        eval_adg={
            300:[1.13,1.38,1.57,1.69],400:[1.31,1.58,1.79,1.92],500:[1.46,1.75,1.96,2.10],
            600:[1.58,1.88,2.11,2.26],700:[1.69,2.01,2.25,2.40],800:[1.79,2.12,2.36,2.52],
        }
        eval_cp={
            300:[11.7,13.3,14.8,16.3],400:[10.6,11.9,13.1,14.4],500:[9.8,10.9,12.0,13.0],
            600:[9.0,9.9,10.9,11.8],700:[8.1,8.8,9.6,10.4],800:[7.3,8.0,8.6,9.3],
        }
    else:
        # NASEM 2016 Chapter 20 Errata, Table 20-1: growing/finishing cattle.
        sbw_grid=[250,300,350,400,450,500]
        mp_maint_vals=[239,274,307,340,371,402]
        ca_maint_vals=[7.7,9.2,10.8,12.3,13.9,15.4]
        p_maint_vals=[5.9,7.1,8.2,9.4,10.6,11.8]
        mp_gain_by_sbw={
            250:[125,242,355,467,576],300:[127,244,357,467,575],350:[129,246,358,466,571],
            400:[120,226,326,423,516],450:[111,207,296,381,463],500:[102,188,267,341,412],
        }
        ca_gain_by_sbw={
            250:[10.4,20.1,29.6,38.9,48.0],300:[9.7,18.6,27.2,35.6,43.8],350:[9.0,17.2,25.0,32.5,39.9],
            400:[8.4,15.8,22.8,29.5,36.1],450:[7.7,14.4,20.7,26.6,32.4],500:[7.1,13.1,18.6,23.8,28.8],
        }
        p_gain_by_sbw={
            250:[4.2,8.1,12.0,15.7,19.4],300:[3.9,7.5,11.0,14.4,17.7],350:[3.6,6.9,10.1,13.1,16.1],
            400:[3.4,6.4,9.2,11.9,14.6],450:[3.1,5.8,8.4,10.8,13.1],500:[2.9,5.3,7.5,9.6,11.6],
        }
        eval_adg={250:[.86,1.04,1.17,1.25],300:[.94,1.12,1.26,1.35],350:[1.00,1.19,1.34,1.43],400:[1.06,1.26,1.41,1.51],450:[1.12,1.32,1.48,1.58],500:[1.17,1.38,1.54,1.64]}
        eval_cp={250:[11.5,12.9,14.3,15.6],300:[10.8,12.0,13.2,14.4],350:[10.2,11.3,12.4,13.5],400:[9.4,10.3,11.2,12.2],450:[8.7,9.4,10.2,11.0],500:[8.0,8.7,9.4,10.1]}

    mp_maint=_interp(sbw,sbw_grid,mp_maint_vals)
    ca_maint=_interp(sbw,sbw_grid,ca_maint_vals)
    p_maint=_interp(sbw,sbw_grid,p_maint_vals)

    # Aynı errata tablolarındaki büyüme değerleri ağırlık ve ADG boyunca enterpole edilir.
    mp_by_weight=[]
    for sw in sbw_grid: mp_by_weight.append(_interp(adg,adg_grid,mp_gain_by_sbw[sw]))
    mp_gain=_interp(sbw,sbw_grid,mp_by_weight)
    mp_req=max(0.0,mp_maint+mp_gain)

    ca_gain=_interp(sbw,sbw_grid,[_interp(adg,adg_grid,ca_gain_by_sbw[x]) for x in sbw_grid])
    p_gain=_interp(sbw,sbw_grid,[_interp(adg,adg_grid,p_gain_by_sbw[x]) for x in sbw_grid])
    ca_g=max(0.0,ca_maint+ca_gain); p_g=max(0.0,p_maint+p_gain)

    # Chapter 20 diet-evaluation satırlarından DMI ve CP tarama hedefi. Bunlar seçilen diyetin enerji yoğunluğuna göre
    # değişebildiğinden solver ayrıca NEm/NEg yoğunluğundan dinamik gerekli KM hesabı yapar.
    # DMI sabit bir BW yüzdesi değildir. NASEM 2016'da genç sığırlarda diyetin NEm
    # yoğunluğuna bağlıdır. Çözümden önce kartlarda faza uygun referans NEm kullanılır;
    # solver her aday rasyonda DMI bütçesini gerçek aday NEm yoğunluğundan tekrar hesaplar.
    eval_sbw=max(sbw_grid[0],min(sbw,sbw_grid[-1]))
    reference_nem=_reference_nem_density(w,age)
    dmi_kg=nasem_dynamic_dmi(sbw,reference_nem,age,w)
    dmi_kg=max(w*.018,min(w*.035,dmi_kg))
    dmi_pct=dmi_kg/w*100.0
    cp_pct=max(8.0,min(16.0,_interp(eval_sbw,sbw_grid,[_interp(max(min(adg,max(eval_adg[x])),min(eval_adg[x])),eval_adg[x],eval_cp[x]) for x in sbw_grid])))
    phase,age_note=_beef_phase_from_weight_age(w,age)
    requested_phase=str(phase_override or 'Otomatik').strip()
    if requested_phase in ('Besi Başlangıç','Besi Geliştirme','Besi Bitirme'):
        phase=requested_phase
    if w<300: ndf_min,ndf_max=28.0,42.0
    elif w<500: ndf_min,ndf_max=25.0,40.0
    else: ndf_min,ndf_max=22.0,38.0
    # DEV4 faz standardı (KM bazında): canlı ağırlık önce besi dönemini seçer.
    # Başlangıç 50/50, geliştirme 40/60, bitirme ise yem kalitesi ve rumen
    # güvenliğine göre %30-40 kaba / %60-70 kesif koridorunda çalışır.
    rough_target={'Besi Başlangıç':50.0,'Besi Geliştirme':40.0,'Besi Bitirme':35.0}[phase]
    rough_min={'Besi Başlangıç':47.0,'Besi Geliştirme':37.0,'Besi Bitirme':30.0}[phase]
    rough_max={'Besi Başlangıç':53.0,'Besi Geliştirme':43.0,'Besi Bitirme':40.0}[phase]
    endf_min={'Besi Başlangıç':12.0,'Besi Geliştirme':11.5,'Besi Bitirme':10.5}[phase]
    starch_targets=beef_starch_targets(phase)

    # HOTFIX 6.17: Besi_V5.02 Excel'deki enerji dönüşüm mantığı.
    # Eski tek katsayı ((NEm+NEg)/0.65) özellikle genç besi hayvanlarında ME hedefini
    # yapay olarak düşük gösteriyor, solver da KM/HP'yi tuttururken ME kartını sürekli
    # +%20-30 fazla gösteriyordu. Excel modelinde bakım ve büyüme için ayrı ME kullanım
    # etkinlikleri, hedef ADG'ye bağlı referans ME yoğunluğundan türetiliyor.
    me_conc_ref=(0.0857*adg**2 + 0.5357*adg + 1.6021) + (0 if w<=400 else (0.0000035*w**2 - 0.0016*w + 0.0539))
    k_m=0.35*((me_conc_ref*4.184)/18.4)+0.503
    k_g=0.78*((me_conc_ref*4.184)/18.4)+0.006
    me_mcal_day=(nem_req/max(k_m,.01))+(neg_req/max(k_g,.01))
    return {'mode':'Besi','engine':'NASEM 2016 enerji + Chapter 20 hayvan profili','weight_kg':w,'age_months':age,'adg':adg,
            'animal_profile':profile['key'],'animal_profile_label':profile['label'],'mature_sbw_kg':profile['mature_sbw_kg'],
            'nem_coefficient':profile['nem_coefficient'],'gain_energy_coefficient':profile['neg_coefficient'],
            'sbw_kg':sbw,'ebw_kg':ebw,'ebg_kg':ebg,'phase':phase,'phase_mode':('Manuel' if requested_phase!='Otomatik' else 'Otomatik'),'age_note':age_note,
            'dmi_pct_bw':dmi_pct,'dmi_kg':dmi_kg,'dmi_reference_nem':reference_nem,'cp_pct':cp_pct,'mp_req_g':mp_req,'mp_maint_g':mp_maint,'mp_gain_g':mp_gain,
            'nem_req_mcal':nem_req,'neg_req_mcal':neg_req,'me_mcal_day':me_mcal_day,'me_mcal_kg':me_mcal_day/max(dmi_kg,.01),
            'ca_g':ca_g,'p_g':p_g,'ca_pct':ca_g/max(dmi_kg,.01)/10,'p_pct':p_g/max(dmi_kg,.01)/10,
            'ndf_min':ndf_min,'ndf_max':ndf_max,'endf_min':endf_min,'roughage_target':rough_target,'roughage_min':rough_min,'roughage_max':rough_max,
            **starch_targets}

def beef_phase_limits(weight_kg, target_adg, target_dm, age_months=0, phase_override='Otomatik'):
    """Besi fazına göre rumen ve yem güvenlik rayları (KM bazında)."""
    w=float(weight_kg); phase,_=_beef_phase_from_weight_age(w,age_months)
    if str(phase_override or 'Otomatik').strip() in ('Besi Başlangıç','Besi Geliştirme','Besi Bitirme'):
        phase=str(phase_override).strip()
    # DEV4: faz önce gelir. Kaba/kesif koridoru canlı ağırlığa göre seçilir;
    # saman zorunlu değildir, kaba payı yonca/silaj/kaliteli kuru otla tamamlanabilir.
    if phase=='Besi Başlangıç': rough_min,rough_max,endf_min,silage_dm_max=47.0,53.0,12.0,0.45
    elif phase=='Besi Geliştirme': rough_min,rough_max,endf_min,silage_dm_max=37.0,43.0,11.5,0.40
    else: rough_min,rough_max,endf_min,silage_dm_max=30.0,40.0,10.5,0.35
    starch_targets=beef_starch_targets(phase)
    return {'phase':phase,'roughage_min':rough_min,'roughage_max':rough_max,'endf_min':endf_min,**starch_targets,
            'silage_dm_max_frac':silage_dm_max,'salt_dm_frac':0.0030,'target_dm':target_dm}

def _is_commercial_compound_feed(feed):
    """Besi ve süt için üretilmiş ticari karma yemleri tek yerde tanır.

    Eski katalogda bazı süt yemleri ``Sulu Kaba Yemler`` kategorisine aktarılmıştır;
    bu nedenle sınıflama kategoriye değil ürün adına dayanır.
    """
    name=str(_rowval(feed,'name','')).upper()
    commercial_words=(
        'BUZAĞI BAŞLANGIÇ YEMİ','BUZAGI BASLANGIC YEMI',
        'BUZAĞI BÜYÜTME','BUZAGI BUYUTME',
        'SIĞIR BESİ YEMİ','SIGIR BESI YEMI','BESİ YEMİ','BESI YEMI',
        'SIĞIR SÜT YEMİ','SIGIR SUT YEMI','SÜT YEMİ','SUT YEMI',
        'SAĞMAL YEMİ','SAGMAL YEMI','KARMA YEM'
    )
    return any(x in name for x in commercial_words)

def _commercial_feed_kind(feed):
    """Ticari yemin hedef hayvan profilini adından güvenli biçimde ayırır."""
    name=str(_rowval(feed,'name','')).upper()
    if any(x in name for x in ('SÜT YEMİ','SUT YEMI','SAĞMAL','SAGMAL','KARDELEN')):
        return 'dairy'
    if any(x in name for x in ('BESİ YEMİ','BESI YEMI','BUZAĞI','BUZAGI','GELİŞTİRME BESİ','GELISTIRME BESI')):
        return 'beef'
    return 'neutral'

def _commercial_profile_penalty(feeds, qtys, animal_type, target_dm):
    """Karşı profile ait ticari yemi yasaklamadan güçlü biçimde geri plana atar."""
    animal=str(animal_type or '').upper()
    wanted='dairy' if any(x in animal for x in ('SÜT','SUT','SAĞMAL','SAGMAL')) else 'beef'
    mismatch_dm=0.0
    for feed,qty in zip(feeds,qtys):
        if not _is_commercial_compound_feed(feed): continue
        kind=_commercial_feed_kind(feed)
        if kind!='neutral' and kind!=wanted:
            mismatch_dm+=max(0.0,float(qty or 0))*max(float(_rowval(feed,'dm_pct',0))/100.0,.01)
    return mismatch_dm/max(float(target_dm or 0),.1)

def _apply_commercial_profile_bounds(feeds, bounds, animal_type):
    """Besi yemi varken aynı besi çözümünde süt yeminin kullanılmasını kapatır."""
    animal_upper=str(animal_type or '').upper()
    is_dairy=any(x in animal_upper for x in ('SÜT','SUT','SAĞMAL','SAGMAL'))
    has_beef=(not is_dairy and any(
        _is_commercial_compound_feed(f) and _commercial_feed_kind(f)=='beef' for f in feeds))
    if not has_beef:return list(bounds)
    return [(0.0,0.0) if (_is_commercial_compound_feed(f) and _commercial_feed_kind(f)=='dairy') else b
            for f,b in zip(feeds,bounds)]


def _solver_feed_role(feed):
    """Solver için pratik yem rolü; yalnız optimizasyon sınır/kalite katmanında kullanılır."""
    name=str(_rowval(feed,'name','')).upper(); grp=feed_group(feed)
    cp=float(_solver_nutrient(feed,'cp_pct') or 0); starch=float(_solver_starch_pct(feed) or 0)
    if grp=='Katkı': return 'additive'
    if grp=='Kaba':
        if 'SAMAN' in name: return 'straw'
        if 'SİLAJ' in name or 'SILAJ' in name: return 'silage'
        if any(x in name for x in ('YONCA','KURU OT','ÇAYIR','CAYIR','MERA','FİĞ','FIG','HASIL')): return 'hay'
        return 'roughage'
    if _is_commercial_compound_feed(feed): return 'commercial'
    if cp>=30 or any(x in name for x in ('SOYA KÜSPESİ','SOYA KUSPESI','KANOLA KÜSPESİ','KANOLA KUSPESI','AYÇİÇEĞİ KÜSPESİ','AYCICEGI KUSPESI','PAMUK TOHUMU KÜSPESİ','PAMUK TOHUMU KUSPESI')):
        return 'protein'
    if starch>=35 or any(x in name for x in ('ARPA EZMESİ','ARPA EZMESI','BUĞDAY EZMESİ','BUGDAY EZMESI','BUĞDAY, ÖĞÜTÜLMÜŞ','BUGDAY, OGUTULMUS','MISIR DANE','MISIR KIRMA')):
        return 'grain'
    return 'concentrate'

def _feed_quality_penalty(feeds, qtys, target_dm):
    """Hedefleri tutturan adaylar arasında saha kalitesini maliyetten önce sıralar.
    Ceza, düşük kaliteli kaba yem/tahıl dominansını ve tek yeme aşırı yüklenmeyi azaltır;
    kaliteli protein ve ticari yemi zorunlu kılmaz, yalnız makul kullanımına alan açar.
    """
    d=max(float(target_dm or 0),0.1); penalty=0.0; active=0
    role_dm={}
    for f,q in zip(feeds,qtys):
        q=max(0.0,float(q or 0)); dmfrac=max(float(_rowval(f,'dm_pct',0))/100.0,.01); dm=q*dmfrac
        if dm<=1e-6: continue
        active+=1; share=dm/d
        role=_solver_feed_role(f); role_dm[role]=role_dm.get(role,0.0)+dm
        me=float(_solver_nutrient(f,'me_mcal_kg') or 0); ndf=float(_solver_nutrient(f,'ndf_pct') or 0)
        # Tek bir kalemin rasyonu ele geçirmesi kararsız/tek yönlü çözümler üretir.
        preferred={'straw':0.08,'silage':0.24,'hay':0.22,'roughage':0.22,'grain':0.18,'protein':0.12,'commercial':0.30,'concentrate':0.20}.get(role,0.22)
        if share>preferred:
            mult={'straw':70,'grain':42,'roughage':24,'silage':18,'protein':14,'commercial':8,'hay':12,'concentrate':16}.get(role,16)
            penalty += mult*((share-preferred)/max(preferred,.05))**2
        # Çok düşük enerji + çok yüksek NDF kaba yem (özellikle saman) KM doldurmak için seçilmesin.
        if role in ('straw','roughage') and me<1.75 and ndf>65 and share>0.06:
            penalty += 55*((share-0.06)/0.06)**2
    # Toplam tahıl dominansı; nişasta kartı güvenli kalsa bile pratik çeşitliliği korur.
    grain_share=role_dm.get('grain',0.0)/d
    if grain_share>0.34: penalty += 28*((grain_share-0.34)/0.10)**2
    # Buğday nişastası arpa/mısıra göre daha hızlı fermente olur. Buğday yasak
    # değildir; fakat seçili tane yem KM'sinin yarısından fazlasını oluşturduğunda
    # solver daha dengeli (arpa/mısır içeren) adayı tercih eder. Toplam rasyon KM'si
    # için ayrıca smart_feed_bounds içinde muhafazakâr %30 güvenlik tavanı vardır.
    grain_dm,wheat_dm,_=_grain_mix_dm(feeds,qtys)
    wheat_mix_share=wheat_dm/max(grain_dm,.01)
    # ÇiftlikPro muhafazakâr saha hedefi: işleme/yıkılabilirlik bilinmiyorsa
    # buğday toplam tane-yem KM'sinin %30'unu geçmesin. %30-40 dikkat bandıdır;
    # %40 üzeri solverın sert güvenlik kapısında ayrıca reddedilir.
    if grain_dm>0.05 and wheat_mix_share>0.30:
        penalty += 90*((wheat_mix_share-0.30)/0.10)**2
    # 2-3 yeme çöken veya seçili tüm yemleri minik miktarda taşıyan çözümleri yumuşakça engelle.
    if active<4: penalty += (4-active)*4.0
    return penalty


def _grain_mix_dm(feeds, qtys):
    """Tahıl karışımını yaş kg değil KM kg olarak döndürür.

    Sonuç: (toplam tahıl KM, buğday KM, arpa KM). Fabrika karma yemleri tahıl
    karışımına dahil edilmez; kendi ticari-yem KM bütçelerinde izlenir.
    """
    total=wheat=barley=0.0
    for feed,qty in zip(feeds,qtys):
        if _solver_feed_role(feed)!='grain': continue
        dm=max(0.0,float(qty or 0))*max(float(_rowval(feed,'dm_pct',0))/100.0,.01)
        total+=dm
        name=str(_rowval(feed,'name','')).upper()
        if 'BUĞDAY' in name or 'BUGDAY' in name: wheat+=dm
        elif 'ARPA' in name: barley+=dm
    return total,wheat,barley

def _apply_explicit_feed_limits(feed, automatic_bounds):
    """Etiket/uzman alt-üst sınırını otomatik pratik sınırın önüne geçirir.

    Sıfır alan "tanımlı değil" demektir. Kullanıcı bir üst sınır girdiyse solver
    onu asla aşmaz; alt sınır girdiyse seçili yemi o dozun altında kullanmaz.
    """
    auto_lo,auto_hi=automatic_bounds
    label_lo=max(0.0,float(_rowval(feed,'solver_min_kg_day',0) or 0))
    label_hi=max(0.0,float(_rowval(feed,'solver_max_kg_day',0) or 0))
    lo=label_lo if label_lo>0 else auto_lo
    hi=min(auto_hi,label_hi) if label_hi>0 else auto_hi
    if label_hi>0 and label_lo>label_hi:
        # Geçersiz katalog girişi güvenli tarafta tek sabit doza kapanır; veri uyarısı
        # kullanıcıya hangi kaydın düzeltilmesi gerektiğini ayrıca bildirir.
        lo=hi=label_hi
    return max(0.0,lo),max(max(0.0,lo),hi)

def smart_feed_bounds(feed, weight_kg, target_dm, target_adg=1.3, age_months=0, phase_override='Otomatik'):
    """Hayvan/faz + etiket/uzman sınırları (kg/baş/gün, yaş baz).

    Evrensel bir arpa/buğday/fabrika-yemi yüzdesi sert kısıt değildir. Toplam
    nişasta, etkin nişasta yıkılabilirliği, eNDF ve kaba/kesif dengesi rasyon
    düzeyinde korunur; ürün etiketi varsa kesin alt/üst sınır olarak uygulanır.
    """
    dm=max(float(feed['dm_pct'] or 0)/100.0,.05); grp=feed_group(feed); name=str(feed['name'] or '').upper(); starch=_solver_starch_pct(feed)
    lim=beef_phase_limits(weight_kg,target_adg,target_dm,age_months,phase_override)
    if grp=='Katkı':
        if 'TUZ' in name:
            dose=target_dm*lim['salt_dm_frac']; return _apply_explicit_feed_limits(feed,(max(0.0,dose*0.85),dose*1.15))
        if 'REVITAMIN BT-SACC' in name or 'BT-SACC' in name: return _apply_explicit_feed_limits(feed,(0.10,0.10))
        if any(x in name for x in ('VİTAMİN','VITAMIN','PREMİKS','PREMIKS','PREMIX','VİT.-MİN','VIT.-MIN')): return _apply_explicit_feed_limits(feed,(0.0,0.05))
        if any(x in name for x in ('ÜRE','URE')): return _apply_explicit_feed_limits(feed,(0.0,min(0.10,target_dm*0.008)))
        if any(x in name for x in ('BİKARBONAT','BIKARBONAT')): return _apply_explicit_feed_limits(feed,(0.0,min(0.20,target_dm*0.015)))
        if any(x in name for x in ('MERMER','KİREÇ','KIREC','KALSİYUM','KALSIYUM')): return _apply_explicit_feed_limits(feed,(0.0,min(0.08,target_dm*0.008)))
        if any(x in name for x in ('FOSFAT','MİNERAL','MINERAL')): return _apply_explicit_feed_limits(feed,(0.0,min(0.12,target_dm*0.010)))
        return _apply_explicit_feed_limits(feed,(0.0,min(0.10,target_dm*0.008)))
    if grp=='Kaba':
        role=_solver_feed_role(feed)
        if role=='silage':
            dm_cap=target_dm*lim['silage_dm_max_frac']; return _apply_explicit_feed_limits(feed,(0.0,max(0.5,dm_cap/dm)))
        if role=='straw':
            # Saman fiziksel lif sağlar fakat KM/enerji açığını dolduran ana yem haline gelmemeli.
            # Canlı ağırlık ve besi ilerledikçe rasyondaki azami saman KM payı kademeli düşer.
            frac=0.10 if weight_kg<300 else (0.08 if weight_kg<450 else 0.06)
            return _apply_explicit_feed_limits(feed,(0.0,max(0.25,min(weight_kg*0.010,target_dm*frac/dm))))
        if role=='hay':
            frac=0.34 if weight_kg<300 else (0.28 if weight_kg<450 else 0.18)
            return _apply_explicit_feed_limits(feed,(0.0,max(0.40,min(weight_kg*0.018,target_dm*frac/dm))))
        return _apply_explicit_feed_limits(feed,(0.0,max(0.5,min(weight_kg*0.018,target_dm*0.45/dm))))
    # Ticari tam/karma buzağı-besi yemleri tek bir tahıl değildir. Eski katalogdaki
    # nişasta alanı hatalı yüksek olsa bile bunları tane yem sınırına sıkıştırma.
    is_commercial=_is_commercial_compound_feed(feed)
    if is_commercial:
        # Ticari karma yemin gerçek sınırı ürün etiketidir. Etiket girilmemişse
        # kaba/kesif koridorunun izin verdiği geniş kapasite kullanılır; grup yüzdesi
        # uydurulmaz. Nişasta/lif güvenliği rasyon toplamında değerlendirilir.
        return _apply_explicit_feed_limits(feed,(0.0,max(0.50,min(weight_kg*0.025,target_dm*0.70/dm))))
    role=_solver_feed_role(feed)
    if role=='protein':
        # Küspeler kaliteli protein kaynağıdır; legacy nişasta hatası yüzünden tahıl sınırına düşmez.
        dm_frac=0.18 if weight_kg<300 else 0.16
        return _apply_explicit_feed_limits(feed,(0.0,max(0.25,min(weight_kg*0.006,target_dm*dm_frac/dm))))
    if role=='grain':
        # Buğday hızlı fermente olan bir tahıldır. Muhafazakâr saha rayı olarak
        # toplam rasyon KM'sinin en çok %30'u buğday olabilir; karışım içindeki
        # buğday dominansı ayrıca kalite sıralamasında cezalandırılır. Arpa/mısır
        # için toplam nişasta, eNDF ve rumen-risk kapıları belirleyicidir.
        if 'BUĞDAY' in name or 'BUGDAY' in name:
            return _apply_explicit_feed_limits(feed,(0.0,max(0.30,min(weight_kg*0.018,target_dm*0.30/dm))))
        return _apply_explicit_feed_limits(feed,(0.0,max(0.30,min(weight_kg*0.018,target_dm*0.65/dm))))
    if starch>=55: dm_frac=0.22
    elif starch>=35: dm_frac=0.30
    else: dm_frac=0.36
    return _apply_explicit_feed_limits(feed,(0.0,max(0.30,min(weight_kg*(0.012 if starch>=55 else 0.016),target_dm*max(dm_frac,.50)/dm))))

def practical_feed_min(feed, weight_kg, target_dm):
    """Sahada anlamlı en düşük kullanım miktarı (kg/baş/gün, yaş baz).
    Bu bir zorunlu minimum değildir: yem ya 0 olur ya da bu eşiğin üzerinde kullanılır.
    Tuz/premiks/mineral gram ölçeğinde ayrı tutulur."""
    name=str(_rowval(feed,'name','')).upper(); grp=feed_group(feed)
    if grp=='Katkı':
        if 'BT-SACC' in name:return 0.10
        if 'TUZ' in name:return 0.005
        return 0.005
    if 'SİLAJ' in name or 'SILAJ' in name:return max(0.40,min(1.00,float(weight_kg)*0.0025))
    if grp=='Kaba':return max(0.15,min(0.35,float(weight_kg)*0.0010))
    # Tahıl ve ticari yemlerde 50-100 gramlık matematiksel kalemler yerine uygulanabilir miktar.
    return max(0.15,min(0.30,float(weight_kg)*0.0008))

def _practical_qty_penalty(feeds,qtys,weight_kg,target_dm):
    penalty=0.0
    for f,q in zip(feeds,qtys):
        q=float(q or 0); mn=practical_feed_min(f,weight_kg,target_dm)
        if q>0.001 and q<mn:
            # Eşiğe çok yakın küçük sapma hafif, birkaç gramlık kullanım çok ağır cezalı.
            penalty += 55.0*((mn-q)/max(mn,0.01))**2 + 8.0
    return penalty

def _rowval(row,key,default=0.0):
    try:
        if hasattr(row,'keys') and key not in row.keys(): return default
        v=row[key]
        return default if v is None else v
    except Exception:return default



def _solver_nutrient(feed,key):
    """HOTFIX 6.12 katalog kalite katmanı.
    Eski Besi_V5.02 aktarımında özellikle nişasta kolonunda belirgin kaymalar var.
    Solver ham veriyi silmez; yalnız bariz hatalı/etiketle çelişen kayıtları güvenilir
    çalışma değerleriyle normalize eder. Kullanıcının sonradan girdiği gerçek analiz
    makul aralıktaysa aynen kullanılır.
    """
    raw=max(0.0,float(_rowval(feed,key,0)))
    name=str(_rowval(feed,'name','')).upper()
    if key=='starch_pct':
        return _solver_starch_pct(feed)
    # Kullanıcının 20/08/2026 tarihli gerçek Sunar 15.26 etiketi: %15 HP.
    # Ürün adındaki 15.26 enerji sınıfı 2600 kcal/kg ürün bazına karşılık gelir;
    # %88,35 referans KM ile 2,943 Mcal/kg KM ve türetilmiş NEm/NEg kullanılır.
    # Eski Çukoyem adı da DB geçişi tamamlanana kadar aynı profile bağlanır.
    if (('SUNAR' in name and '15.26' in name) or
        (('ÇUKOYEM' in name or 'CUKOYEM' in name) and '15,2650' in name)):
        if key=='cp_pct': return 16.978
        if key=='me_mcal_kg': return 2.943
        if key=='nem_mcal_kg': return 1.984
        if key=='neg_mcal_kg': return 1.333
    # Sunar Kardelen 19.27: ürün adı %19 HP ve 2700 kcal/kg ME bilgisini verir.
    # Referans %88,35 KM ile ME yaklaşık 3,056 Mcal/kg KM; net enerjiler NRC
    # dönüşümüyle türetilmiştir. NDF/nişasta/mineraller etiket gelene kadar
    # katalogdaki açıkça işaretlenmiş referans profilden okunur.
    if 'KARDELEN' in name and ('19,2700' in name or '19.27' in name):
        if key=='cp_pct': return 21.505
        if key=='me_mcal_kg': return 3.056
        if key=='nem_mcal_kg': return 2.078
        if key=='neg_mcal_kg': return 1.413
    # Büyütme yemi: mevcut enerji/protein analizi makul, yalnız eski nişasta kolonu bozuk.
    # Pamuk tohumu yüksek lintli için OSU tipik whole-cottonseed değerlerine yakın,
    # daha muhafazakâr net enerji ve fiziksel etkinlik kullanılır.
    if 'PAMUK TOHUMU, YÜKSEK LİNTLİ' in name or 'PAMUK TOHUMU, YUKSEK LINTLI' in name:
        vals={'ndf_pct':53.0,'effective_ndf_pct':80.0,'tdn_pct':77.0,
              'me_mcal_kg':2.78,'nem_mcal_kg':1.70,'neg_mcal_kg':1.39,
              'cp_pct':24.0,'fat_pct':19.0}
        if key in vals:return vals[key]
    if 'PAMUK TOHUMU, BÜTÜN' in name or 'PAMUK TOHUMU, BUTUN' in name:
        vals={'ndf_pct':53.0,'effective_ndf_pct':80.0,'tdn_pct':77.0,
              'nem_mcal_kg':1.70,'neg_mcal_kg':1.39,'cp_pct':24.0,'fat_pct':19.0}
        if key in vals:return vals[key]
    return raw

def _solver_starch_pct(feed):
    """Legacy katalogdaki bariz nişasta hatalarını solver sırasında güvenli biçimde sınırlar.
    Ham katalog verisini değiştirmez; yalnız optimizasyon değerlendirmesinde kullanılır.
    Amaç 100% nişasta samanı veya 80%+ nişasta ticari yem gibi fiziksel olarak şüpheli
    değerlerin optimizasyonu kilitlemesini önlemektir.
    """
    raw=max(0.0,float(_rowval(feed,'starch_pct',0)))
    name=str(_rowval(feed,'name','')).upper(); grp=feed_group(feed)
    if grp=='Kaba':
        if 'SAMAN' in name: return min(raw,3.0)
        if any(x in name for x in ('KURU OT','YONCA','ÇAYIR','MERA','FİĞ','FIG')): return min(raw,12.0)
        if 'SİLAJ' in name or 'SILAJ' in name: return min(raw,35.0)
        return min(raw,20.0)
    # Yan ürünlerde eski Excel aktarımındaki kolon kaymaları özellikle nişastayı abartabiliyor.
    # Ham katalog değerini değiştirmiyoruz; solver güvenli çalışma değeriyle hesaplıyor.
    if 'KEPEK' in name and raw>45.0: return 25.0
    if 'PAMUK TOHUMU' in name and raw>25.0: return 2.0
    # Protein küspelerinde legacy Excel aktarımı nişasta kolonunu sık sık 75-100 aralığına taşıdı.
    # Bu yemleri tahıl gibi cezalandırmak kaliteli protein kaynaklarını solverdan gereksiz yere uzaklaştırıyordu.
    if any(x in name for x in ('SOYA KÜSPESİ','SOYA KUSPESI','KANOLA KÜSPESİ','KANOLA KUSPESI','AYÇİÇEĞİ KÜSPESİ','AYCICEGI KUSPESI','PAMUK TOHUMU KÜSPESİ','PAMUK TOHUMU KUSPESI')):
        return min(raw,8.0)
    if any(x in name for x in ('PANCAR POSASI','BEET PULP')) and raw>25.0: return 12.0
    # Ticari tam/karma yemlerde 55% üzeri nişasta değeri çoğu zaman eski Excel alan eşleşmesi hatasıdır.
    # Gerçek ürün analizi girilene kadar buzağı yeminde %32, besi yeminde %35 muhafazakâr çalışma değeri kullanılır.
    if ('BUZAĞI' in name or 'BUZAGI' in name) and ('YEMİ' in name or 'YEMI' in name) and raw>55.0: return 32.0
    if ('SÜT YEMİ' in name or 'SUT YEMI' in name) and raw>55.0: return 32.0
    if ('BESİ' in name or 'BESI' in name) and ('YEMİ' in name or 'YEMI' in name) and raw>55.0: return 35.0
    if ('YEMİ' in name or 'YEMI' in name) and raw>55.0: return 38.0
    return min(raw,75.0)

def _feed_starch_degradability(feed):
    """INRA/CNCPS benzeri fermantasyon kontrolü için (değer, biliniyor) döndürür.

    İşleme biçimi rumen nişasta yıkılabilirliğini ciddi ölçüde değiştirebildiği için
    bilinmeyen bir yeme yalnız adından kesin oran uydurulmaz. ``init_db`` yalnız
    açıkça eşleşen standart arpa/buğday/mısır kayıtlarını INRA 2018 referansıyla
    doldurur; özel yemlerde laboratuvar/etiket alanı beklenir.
    """
    value=max(0.0,min(100.0,float(_rowval(feed,'starch_degradability_pct',0) or 0)))
    return value,value>0

def _rumen_risk_assessment(metrics, limits):
    """Tek bir sahte pH sayısı yerine veri-kapsamalı göreli asidoz risk sınıfı.

    Bu bir klinik tanı veya gerçek rumen pH tahmini değildir. Toplam nişasta,
    bilinen kısmın etkin rumen yıkılabilirliği ve fiziksel etkili lif göstergesini
    birlikte yorumlar. İşleme, adaptasyon ve yemleme yönetimi kullanıcı uyarısında
    ayrıca korunur.
    """
    starch=float(metrics.get('starch_pct_dm') or 0)
    endf=float(metrics.get('endf_pct_dm') or 0)
    rapid=float(metrics.get('rapid_starch_pct_dm') or 0)
    coverage=float(metrics.get('starch_degradability_coverage') or 0)
    ideal=float(limits.get('starch_ideal_max') or 100)
    maximum=float(limits.get('starch_max') or 100)
    endf_min=float(limits.get('endf_min') or 0)
    score=0
    reasons=[]
    # Analiz ve ekranda yuvarlama belirsizliği nedeniyle ideal bandın yalnız
    # 0,1-0,5 puan üstü tek başına risk sınıfını yükseltmez.
    starch_attention_buffer=.5
    if starch>maximum:
        score+=2; reasons.append('toplam nişasta dikkat eşiğinin üzerinde')
    elif starch>ideal+starch_attention_buffer:
        score+=1; reasons.append('toplam nişasta dikkat bandında')
    if endf_min and endf<endf_min*.80:
        score+=2; reasons.append('fiziksel etkili lif çok düşük')
    elif endf_min and endf<endf_min:
        score+=1; reasons.append('fiziksel etkili lif düşük')
    # Yalnız veri kapsamı yeterliyse hızlı fermente olan nişasta ayrı risk sinyali olur.
    if coverage>=.70 and rapid>ideal*.75:
        score+=1; reasons.append('rumende etkin yıkılan nişasta yüksek')
    level='Yüksek' if score>=3 else ('Orta' if score>=1 else 'Düşük')
    confidence='Yüksek' if coverage>=.85 else ('Orta' if coverage>=.50 else 'Düşük')
    return {'level':level,'score':score,'confidence':confidence,'coverage':coverage,
            'reasons':reasons,'clinical_ph_prediction':False}

def _solver_feasibility_report(metrics, targets, limits):
    """Profesyonel optimizerlerdeki "limiting constraints" yaklaşımının özeti.

    Güvensiz veya birden çok temel hedefi ciddi kaçıran aday "çözüm" diye
    kaydedilmez. Küçük saha sapmaları ise açık uyarıyla kullanılabilir kalır.
    """
    dm_target=max(float(metrics.get('predicted_dmi_kg') or _predicted_dmi_for_metrics(metrics,targets)),.01)
    metrics['predicted_dmi_kg']=dm_target
    cp_target=max(float(targets.get('cp_pct') or 0),.01)
    me_target=max(float(targets.get('me_mcal_day') or 0),.01)
    dm_signed=(float(metrics.get('dm_kg') or 0)-dm_target)/dm_target
    cp_signed=(float(metrics.get('cp_pct_dm') or 0)-cp_target)/cp_target
    me_signed=(float(metrics.get('me_mcal') or 0)-me_target)/me_target
    adg_target=max(float(targets.get('adg') or 0),0.0)
    adg_value=float(metrics.get('achievable_adg_kg') or 0)
    adg_signed=(adg_value-adg_target)/adg_target if adg_target>0 else 0.0
    rough=float(metrics.get('roughage_pct_dm') or 0)
    unsafe=[]; blockers=[]; warnings=[]
    rumen_risk=_rumen_risk_assessment(metrics,limits)
    # DEV4.17: İdeal üstü ile sert üst arasındaki dikkat bandı kullanılabilir;
    # fazın sert nişasta üst sınırı aşılırsa çözüm kaydedilmez.
    starch=float(metrics.get('starch_pct_dm') or 0)
    starch_max=float(limits.get('starch_max') or 100)
    if starch>starch_max+.05:
        unsafe.append(f'nişasta %{starch:.1f}; faz üst sınırı %{starch_max:.1f}')
    if rumen_risk['level']=='Yüksek' and (
        starch>starch_max
        or float(metrics.get('endf_pct_dm') or 0)<float(limits.get('endf_min') or 0)):
        unsafe.append('yüksek göreli asidoz riski (nişasta/etkin lif birlikte)')
    if float(metrics.get('endf_pct_dm') or 0)<float(limits.get('endf_min') or 0)*.80:
        unsafe.append('fiziksel etkili lif ciddi düşük')
    if float(metrics.get('ndf_pct_dm') or 0)<float(targets.get('ndf_min') or 0)*.80:
        unsafe.append('NDF ciddi düşük')
    wheat_grain=float(metrics.get('wheat_grain_pct_dm') or 0)
    if wheat_grain>40.0:
        unsafe.append(f'buğday tahıl KM payı %{wheat_grain:.1f}; sert üst sınır %40')
    elif wheat_grain>30.0:
        warnings.append(f'buğday tahıl KM payı %{wheat_grain:.1f}; hedef en çok %30')
    cap=float(metrics.get('ca_g') or 0)/max(float(metrics.get('p_g') or 0),.01)
    if cap<1.0 or cap>4.0:unsafe.append('Ca:P oranı güvenli pencerenin dışında')
    mw=_mineral_windows(targets,dm_target)
    if float(metrics.get('ca_g') or 0)>mw['ca_hard'] or float(metrics.get('p_g') or 0)>mw['p_hard']:
        unsafe.append('mineral sert üst sınırı aşılıyor')

    if abs(dm_signed)>.08:blockers.append(f'KM {dm_signed*100:+.1f}%')
    elif abs(dm_signed)>.035:warnings.append(f'KM {dm_signed*100:+.1f}%')
    # Chapter 20 CP değeri bir asgari tarama düzeyidir; RDP/RUP ve MP arzı
    # bilinmeden hedefin üstündeki HP'yi fizibilite hatası saymak doğru değildir.
    if cp_signed<-.08:blockers.append(f'HP tabanı {cp_signed*100:+.1f}%')
    elif cp_signed<-.035:warnings.append(f'HP tabanı {cp_signed*100:+.1f}%')
    elif float(metrics.get('cp_pct_dm') or 0)>max(cp_target*1.45,16.0):
        warnings.append('HP yüksek; maliyet ve azot yükü kontrol edilmeli')
    # Toplam ME arzı tek başına büyüme yeterliliği değildir. Enerji fizibilitesi
    # aşağıdaki NEm/NEg tabanlı GCAA kapasitesiyle değerlendirilir.
    if adg_target>0:
        if adg_signed<-.01:blockers.append(f'GCAA kapasitesi {adg_signed*100:+.1f}%')
        elif adg_signed<-.005:warnings.append(f'GCAA kapasitesi {adg_signed*100:+.1f}%')
    if rough<float(limits.get('roughage_min') or 0)-5 or rough>float(limits.get('roughage_max') or 100)+5:
        blockers.append(f'kaba yem %{rough:.1f}')
    elif rough<float(limits.get('roughage_min') or 0) or rough>float(limits.get('roughage_max') or 100):
        warnings.append(f'kaba yem %{rough:.1f}')

    # Bir tek ciddi kartta sapma varsa kullanıcıya "sınırlı" çözüm gösterilebilir;
    # iki veya daha fazlası seçili yem setinin birlikte fizibil olmadığını gösterir.
    growth_blocked=adg_target>0 and adg_signed<-.01
    core_critical=(abs(dm_signed)>.10 or cp_signed<-.10)
    status='unsafe' if unsafe else ('infeasible' if growth_blocked or core_critical or len(blockers)>=2 else ('limited' if blockers or warnings else 'feasible'))
    return {'status':status,'unsafe':unsafe,'blockers':blockers,'warnings':warnings,
            'dm_signed':dm_signed,'cp_signed':cp_signed,'me_signed':me_signed,'adg_signed':adg_signed,
            'rumen_risk':rumen_risk}

def _feed_data_warnings(feeds):
    out=[]
    for f in feeds:
        name=str(_rowval(f,'name','Yem'))
        raw=float(_rowval(f,'starch_pct',0)); safe=_solver_starch_pct(f)
        if raw>safe+5:
            out.append(f'{name}: katalog nişastası %{raw:.1f} şüpheli; çözümde %{safe:.1f} güvenlik değeri kullanıldı.')
        dm=float(_rowval(f,'dm_pct',0)); cp=float(_rowval(f,'cp_pct',0)); ndf=float(_rowval(f,'ndf_pct',0))
        if not (5<=dm<=100): out.append(f'{name}: KM %{dm:.1f} değeri kontrol edilmeli.')
        if cp>45: out.append(f'{name}: HP %{cp:.1f} değeri kontrol edilmeli.')
        if ndf>90: out.append(f'{name}: NDF %{ndf:.1f} değeri kontrol edilmeli.')
        label_min=float(_rowval(f,'solver_min_kg_day',0) or 0); label_max=float(_rowval(f,'solver_max_kg_day',0) or 0)
        if label_min>0 and label_max>0 and label_min>label_max:
            out.append(f'{name}: etiket alt sınırı üst sınırdan büyük; Yem Kataloğu kaydını düzeltin.')
    return out

def _scientific_feed_coverage(feeds,qtys):
    """Seçili çözümün ileri model alanlarındaki ağırlıklı veri kapsamını bildirir."""
    total_dm=0.0; starch_total=0.0; starch_known=0.0; pdi_dm=0.0; ndfd_dm=0.0; limits_dm=0.0
    for f,q in zip(feeds,qtys):
        dm=max(0.0,float(q or 0))*max(float(_rowval(f,'dm_pct',0) or 0)/100.0,.01)
        if dm<=1e-9:continue
        total_dm+=dm
        starch=dm*_solver_starch_pct(f)/100.0; starch_total+=starch
        if float(_rowval(f,'starch_degradability_pct',0) or 0)>0:starch_known+=starch
        if float(_rowval(f,'inra_pdi_g_kg_dm',0) or 0)>0:pdi_dm+=dm
        if float(_rowval(f,'ndf_digestibility_pct',0) or 0)>0:ndfd_dm+=dm
        if float(_rowval(f,'solver_min_kg_day',0) or 0)>0 or float(_rowval(f,'solver_max_kg_day',0) or 0)>0:limits_dm+=dm
    return {
        'starch_degradability':starch_known/starch_total if starch_total else 1.0,
        'inra_pdi':pdi_dm/total_dm if total_dm else 0.0,
        'ndf_digestibility':ndfd_dm/total_dm if total_dm else 0.0,
        'explicit_limits':limits_dm/total_dm if total_dm else 0.0,
    }

def _ration_assistant(feeds,m,t,lim):
    """Çözüm güvenlik kapısına takılırsa seçili yemleri tanıyarak uygulanabilir öneri üretir."""
    tips=[]
    adg_target=max(float(t.get('adg') or 0),0.0)
    adg_capacity=float(m.get('achievable_adg_kg') or _achievable_adg(m,t) or 0)
    commercial_feed=next((x for x in feeds if _is_commercial_compound_feed(x)),None)
    commercial=str(_rowval(commercial_feed,'name','')) if commercial_feed is not None else None
    if adg_target>0 and adg_capacity<adg_target*.995:
        gap=max(0.0,(adg_target-adg_capacity)/adg_target*100)
        # Enerji açığını yüksek HP'li ticari yemi körlemesine artırarak kapatma.
        # Seçili yemler içinden NEg/KM'si yüksek ve HP yükü görece düşük olanı bul.
        energy_candidates=[x for x in feeds
                           if (str(_rowval(x,'category','')).startswith('Kesif') or _is_commercial_compound_feed(x))
                           and _solver_nutrient(x,'neg_mcal_kg')>0]
        energy_candidates.sort(key=lambda x:(_solver_nutrient(x,'neg_mcal_kg')-.012*max(0,_solver_nutrient(x,'cp_pct')-t['cp_pct'])),reverse=True)
        energy_name=str(_rowval(energy_candidates[0],'name','uygun enerji yemi')) if energy_candidates else 'arpa/mısır gibi uygun enerji yemi'
        tips.append(f'GCAA kapasitesi hedefin %{gap:.1f} altında: seçili yemlerden {energy_name} enerji açığı yönünde artırılmalı; aynı KM korunarak düşük enerjili veya gereksiz HP yükselten yem azaltılmalı. Nişasta, eNDF ve kaba/kesif rayları birlikte korunur.')
    if m['cp_pct_dm']<t['cp_pct']*.90:
        if commercial: tips.append(f'Protein yetersiz: seçili {commercial} protein kaynağı olarak kullanılıyor; yeterli değilse soya/kanola/ayçiçeği küspesi ekleyin.')
        else: tips.append('Protein yetersiz: soya/kanola/ayçiçeği küspesi veya uygun proteinli besi yemi ekleyin.')
    if m['endf_pct_dm']<lim['endf_min']:
        tips.append('Etkili lif düşük: saman/yonca/uygun kuru ot gibi fiziksel etkili kaba yem ekleyin veya tahılı azaltın.')
    elif m['roughage_pct_dm']<lim['roughage_min']:
        tips.append(f'Kaba yem KM oranı faz koridorunun altında (%{m["roughage_pct_dm"]:.1f}; hedef %{lim["roughage_min"]:.0f}–%{lim["roughage_max"]:.0f}). eNDF yeterli olsa da kaba/kesif dağılımını düzeltmek için uygun kaba yemi artırın veya kesif yemi azaltın.')
    elif m['roughage_pct_dm']>lim['roughage_max']:
        tips.append(f'Kaba yem KM oranı faz koridorunun üzerinde (%{m["roughage_pct_dm"]:.1f}; hedef %{lim["roughage_min"]:.0f}–%{lim["roughage_max"]:.0f}). Enerji yoğunluğunu ve GCAA kapasitesini koruyarak kaba/kesif dağılımını yeniden dengeleyin.')
    if m['starch_pct_dm']>lim['starch_max']:
        tips.append('Nişasta yüksek: buğday/arpa gibi hızlı fermente tahılı azaltın; kaba yem veya daha düşük nişastalı enerji kaynağı ekleyin.')
    elif m['starch_pct_dm']>lim['starch_ideal_max']+.5:
        tips.append(f'Nişasta ideal bandın üzerinde (%{lim["starch_ideal_max"]:.0f}–%{lim["starch_max"]:.0f} dikkat bölgesi): arpa/buğday miktarı, yem işleme inceliği ve eNDF birlikte kontrol edilmeli.')
    cap=m['ca_g']/m['p_g'] if m['p_g']>0 else 99
    if cap<1.25: tips.append('Ca:P düşük: kalsiyum kaynağı/mineral dengeleyici ekleyin.')
    elif cap>3.0: tips.append('Ca:P yüksek: yüksek kalsiyumlu yem/minerali azaltın veya fosfor dengesini kontrol edin.')
    mw=_mineral_windows(t,m.get('predicted_dmi_kg') or m.get('dm_kg') or t.get('dmi_kg'))
    if m['ca_g']>mw['ca_hard'] or m['p_g']>mw['p_hard']: tips.append(_selected_mineral_tip(feeds))
    if not tips: tips.append('Seçili yemler yeniden dengelendi ancak hedef pencereleri aynı anda kapanmadı; önce seçili yemlerin sınırlarını ve katalog kalite uyarılarını kontrol edin.')
    return ' '.join(tips)

def smart_ration_metrics(feeds, qty):
    z={'as_fed_kg':0.0,'dm_kg':0.0,'cp_kg':0.0,'ndf_kg':0.0,'endf_kg':0.0,'starch_kg':0.0,
       'rapid_starch_kg':0.0,'known_degradability_starch_kg':0.0,'tdn_kg':0.0,
       'me_mcal':0.0,'nem_mcal':0.0,'neg_mcal':0.0,'ca_g':0.0,'p_g':0.0,'na_g':0.0,'cl_g':0.0,'cost':0.0,'rough_dm':0.0,'conc_dm':0.0}
    for f,kg in zip(feeds,qty):
        dm=kg*float(_rowval(f,'dm_pct'))/100.0; z['as_fed_kg']+=kg; z['dm_kg']+=dm
        z['cp_kg']+=dm*_solver_nutrient(f,'cp_pct')/100; z['ndf_kg']+=dm*_solver_nutrient(f,'ndf_pct')/100
        z['endf_kg']+=dm*_solver_nutrient(f,'ndf_pct')/100*_solver_nutrient(f,'effective_ndf_pct')/100
        feed_starch=dm*_solver_starch_pct(f)/100; z['starch_kg']+=feed_starch
        starch_deg,known_deg=_feed_starch_degradability(f)
        if known_deg:
            z['known_degradability_starch_kg']+=feed_starch
            z['rapid_starch_kg']+=feed_starch*starch_deg/100
        z['tdn_kg']+=dm*_solver_nutrient(f,'tdn_pct')/100; z['me_mcal']+=dm*_solver_nutrient(f,'me_mcal_kg')
        z['nem_mcal']+=dm*_solver_nutrient(f,'nem_mcal_kg'); z['neg_mcal']+=dm*_solver_nutrient(f,'neg_mcal_kg')
        z['ca_g']+=dm*float(_rowval(f,'ca_pct'))*10; z['p_g']+=dm*float(_rowval(f,'p_pct'))*10
        z['na_g']+=dm*float(_rowval(f,'na_pct'))*10; z['cl_g']+=dm*float(_rowval(f,'cl_pct'))*10; z['cost']+=kg*float(_rowval(f,'price'))
        if feed_group(f)=='Kaba':z['rough_dm']+=dm
        elif feed_group(f)=='Kesif':z['conc_dm']+=dm
    dm=z['dm_kg']; rc=z['rough_dm']+z['conc_dm']; z['cp_pct_dm']=z['cp_kg']/dm*100 if dm else 0; z['ndf_pct_dm']=z['ndf_kg']/dm*100 if dm else 0
    z['endf_pct_dm']=z['endf_kg']/dm*100 if dm else 0; z['starch_pct_dm']=z['starch_kg']/dm*100 if dm else 0
    z['rapid_starch_pct_dm']=z['rapid_starch_kg']/dm*100 if dm else 0
    z['starch_degradability_coverage']=z['known_degradability_starch_kg']/z['starch_kg'] if z['starch_kg'] else 1.0
    z['roughage_pct_dm']=z['rough_dm']/rc*100 if rc else 0; z['concentrate_pct_dm']=z['conc_dm']/rc*100 if rc else 0
    z['nem_density']=z['nem_mcal']/dm if dm else 0; z['neg_density']=z['neg_mcal']/dm if dm else 0
    grain_dm,wheat_dm,barley_dm=_grain_mix_dm(feeds,qty)
    z['grain_dm_kg']=grain_dm;z['wheat_dm_kg']=wheat_dm;z['barley_dm_kg']=barley_dm
    z['wheat_grain_pct_dm']=wheat_dm/max(grain_dm,.01)*100 if grain_dm>.01 else 0.0
    return z

def _energy_balance(m,t):
    """Net enerji sisteminde bakım için gereken KM ve kalan KM'nin büyümeye sağlayabildiği NEg."""
    dm=float(m.get('dm_kg') or 0);nem=float(m.get('nem_density') or 0);neg=float(m.get('neg_density') or 0)
    if dm<=0 or nem<=0 or neg<=0:return 0.0,0.0,999.0
    maint_dm=t['nem_req_mcal']/nem
    gain_supply=max(0.0,dm-maint_dm)*neg
    required_dm=maint_dm+t['neg_req_mcal']/neg
    return gain_supply,required_dm,maint_dm

def _predicted_dmi_for_metrics(metrics, targets):
    """Kart, skor ve fizibilite için tek dinamik KM tüketim tahmini.

    NASEM DMI denklemi diyetin NEm yoğunluğuna bağlıdır. Önceki kart sabit bir
    referans yoğunluk kullanırken solver aday diyet yoğunluğunu kullanıyordu;
    böylece aynı rasyon için iki farklı KM hedefi görünüyordu.
    """
    # Eski kayıtlar ve bazı yardımcı/test çağrıları SBW taşımayabilir. Bu durumda
    # uydurma 0,04 kg gibi bir tahmin yerine kaydedilmiş referans KM korunur.
    if float(targets.get('sbw_kg') or 0)<=0 or float(targets.get('weight_kg') or 0)<=0:
        return max(float(targets.get('dmi_kg') or 0),.01)
    nem=float(metrics.get('nem_density') or targets.get('dmi_reference_nem') or 1.6)
    dmi=nasem_dynamic_dmi(targets.get('sbw_kg',0),nem,targets.get('age_months',0),targets.get('weight_kg',0))
    weight=max(float(targets.get('weight_kg') or 0),1.0)
    return max(weight*.018,min(weight*.035,dmi))

def _achievable_adg(metrics, targets):
    """Çözülmüş diyetin NEm/NEg arzından karşılanabilir canlı ağırlık artışı."""
    gain,_,_=_energy_balance(metrics,targets)
    if gain<=0:return 0.0
    ebw=max(float(targets.get('ebw_kg') or 0),1.0)
    gain_coefficient=max(float(targets.get('gain_energy_coefficient') or 0.0635),.001)
    ebg=(gain/max(gain_coefficient*(ebw**0.75),1e-9))**(1/1.097)
    adg=max(0.0,ebg/0.956)
    cp_target=max(float(targets.get('cp_pct') or 0),.1)
    cp=float(metrics.get('cp_pct_dm') or 0)
    if cp<cp_target*.95:adg*=max(.70,cp/max(cp_target*.95,.1))
    return adg

def _mineral_windows(t, dmi_kg):
    """Formülasyon için hedef + yumuşak üst + pratik sert üst pencereleri.
    Minimum gereksinimi hedefte tutar; ticari karma yem nedeniyle hedefin biraz üstünü
    otomatik olarak 'başarısız' saymaz. Sert üstler güvenlik kapısıdır.
    """
    d=max(float(dmi_kg or 0),0.1)
    ca_target=max(float(t.get('ca_g',0) or 0),1.0); p_target=max(float(t.get('p_g',0) or 0),1.0)
    ca_soft=max(ca_target*1.50,d*10*0.85); p_soft=max(p_target*1.50,d*10*0.50)
    ca_hard=max(ca_target*2.25,d*10*1.20); p_hard=max(p_target*2.25,d*10*0.70)
    return {'ca_target':ca_target,'p_target':p_target,'ca_soft':ca_soft,'p_soft':p_soft,'ca_hard':ca_hard,'p_hard':p_hard}

def _selected_mineral_tip(feeds):
    names=[str(_rowval(f,'name','')).upper() for f in feeds]
    if any('MERMER' in n or 'KİREÇ' in n or 'KIREC' in n for n in names):
        return 'Mineral fazlalığı: seçili mermer/kireç kaynağı gerekmiyorsa solver bunu sıfıra kadar indirebilir; ticari yemlerden gelen Ca/P de toplam hesaba katılır.'
    return 'Mineral fazlalığı: yüksek Ca/P sağlayan seçili yemlerin miktarı azaltılıp enerji/protein daha düşük mineralli kaynaklardan tamamlanmalı.'

def smart_ration_score(m,t,lim):
    def rel(v,x):return abs(v-x)/max(abs(x),.01)
    gain_supply,energy_required_dm,maint_dm=_energy_balance(m,t)
    dynamic_dmi=nasem_dynamic_dmi(t['sbw_kg'],m['nem_density'] or t.get('dmi_reference_nem',1.6),t.get('age_months',0),t['weight_kg'])
    dynamic_dmi=max(t['weight_kg']*.018,min(t['weight_kg']*.035,dynamic_dmi))
    m['predicted_dmi_kg']=dynamic_dmi
    # Önce enerji + gerçek DMI uyumu. Solver enerji açığını sadece daha çok KM yedirerek kapatamaz;
    # fakat DMI bütçesi de diyet enerji yoğunluğuna göre adaydan adaya değişir.
    cp_deficit=max(0.0,(t['cp_pct']-m['cp_pct_dm'])/max(t['cp_pct'],.1))
    ca_deficit=max(0.0,(t['ca_g']-m['ca_g'])/max(t['ca_g'],1))
    p_deficit=max(0.0,(t['p_g']-m['p_g'])/max(t['p_g'],1))
    score=20*rel(gain_supply,t['neg_req_mcal']) + 15*rel(m['dm_kg'],dynamic_dmi) + 8*cp_deficit + 4*ca_deficit+4*p_deficit
    if m['dm_kg'] > dynamic_dmi*1.08: score += 55*(m['dm_kg']-dynamic_dmi*1.08)/max(dynamic_dmi,.01)
    if m['dm_kg'] < dynamic_dmi*0.90: score += 25*(dynamic_dmi*.90-m['dm_kg'])/max(dynamic_dmi,.01)
    if m['dm_kg'] < maint_dm: score += 45*rel(m['dm_kg'],maint_dm)
    if gain_supply < t['neg_req_mcal']*.97: score += 38*(t['neg_req_mcal']*.97-gain_supply)/max(t['neg_req_mcal'],.01)
    if energy_required_dm > dynamic_dmi*1.08: score += 28*(energy_required_dm-dynamic_dmi*1.08)/max(dynamic_dmi,.01)
    if m['cp_pct_dm'] < t['cp_pct']*.95: score += 18*(t['cp_pct']*.95-m['cp_pct_dm'])/t['cp_pct']
    # HP asgari gereksinimdir; küçük bir güvenlik payı normaldir. Fakat gereksinimin
    # %20 üstünden sonrası enerji hedefini iyileştirmeden azot yükünü ve maliyeti
    # artırır, bu nedenle kademeli olarak cezalanır.
    if m['cp_pct_dm'] > t['cp_pct']*1.20:
        score += 12*(m['cp_pct_dm']-t['cp_pct']*1.20)/max(t['cp_pct'],.1)
    if m['ndf_pct_dm']<t['ndf_min']:score+=10*(t['ndf_min']-m['ndf_pct_dm'])/t['ndf_min']
    if m['ndf_pct_dm']>t['ndf_max']:score+=5*(m['ndf_pct_dm']-t['ndf_max'])/t['ndf_max']
    if m['roughage_pct_dm']<lim['roughage_min']:score+=18*(lim['roughage_min']-m['roughage_pct_dm'])/lim['roughage_min']
    if m['roughage_pct_dm']>lim['roughage_max']:score+=5*(m['roughage_pct_dm']-lim['roughage_max'])/lim['roughage_max']
    if m['endf_pct_dm']<lim['endf_min']:score+=25*(lim['endf_min']-m['endf_pct_dm'])/lim['endf_min']
    # İdeal nişasta bandı yumuşak hedeftir: ME yeterliyse düşük nişasta zorla tahılla tamamlanmaz.
    # İdeal üstünde kademeli dikkat cezası, güvenlik sınırında daha güçlü ceza uygulanır.
    if m['starch_pct_dm']<lim['starch_min'] and gain_supply<t['neg_req_mcal']*1.01:
        score+=2.5*(lim['starch_min']-m['starch_pct_dm'])/max(lim['starch_min'],1)
    if m['starch_pct_dm']>lim['starch_ideal_max']:
        score+=3.5*(m['starch_pct_dm']-lim['starch_ideal_max'])/max(lim['starch_ideal_max'],1)
    if m['starch_pct_dm']>lim['starch_max']:score+=8*(m['starch_pct_dm']-lim['starch_max'])/max(lim['starch_max'],1)
    if m['starch_pct_dm']>lim['starch_max']+2.0:score+=20*(m['starch_pct_dm']-lim['starch_max']-2.0)/max(lim['starch_max'],1)
    risk=_rumen_risk_assessment(m,lim)
    if risk['level']=='Yüksek':score+=30
    elif risk['level']=='Orta':score+=5
    cap=m['ca_g']/m['p_g'] if m['p_g']>0 else 99
    if cap<1.25:score+=10*(1.25-cap)
    if cap>3.0:score+=5*(cap-3.0)
    # Mineral hedefleri minimum gereksinimdir; ticari yemlerle bir miktar üstüne çıkmak olağandır.
    # Hedef üstü kademeli cezalanır, pratik sert pencereye yaklaştıkça ceza hızlanır.
    mw=_mineral_windows(t,dynamic_dmi)
    if m['ca_g']>mw['ca_soft']: score+=14*(m['ca_g']-mw['ca_soft'])/max(mw['ca_soft'],1)
    if m['p_g']>mw['p_soft']: score+=14*(m['p_g']-mw['p_soft'])/max(mw['p_soft'],1)
    if m['ca_g']>mw['ca_hard']: score+=70*(m['ca_g']-mw['ca_hard'])/max(mw['ca_hard'],1)
    if m['p_g']>mw['p_hard']: score+=70*(m['p_g']-mw['p_hard'])/max(mw['p_hard'],1)
    return score+0.002*m['cost']

def solve_smart_ration(feeds, weight_kg, target_adg, animal_type='Besi Erkek', age_months=0, phase_override='Otomatik'):
    """DEV4.18 - NASEM hayvan profilli, gerçek etiketli besi optimizerı.

    Excel dosyasindaki gercek Solver modeli incelendi: yem miktarlari karar degiskenidir,
    hedef kartlari *ceza puani* degil toleransli kisitlardir ve uygun cozumler arasinda
    maliyet dusurulur. CiftlikPro'da kullanicinin sectigi normal yemler korunur.

    Ana saha kısıtları:
      - Diyetin NEm yoğunluğuna bağlı kuru madde tüketimi
      - Minimum ham protein tarama düzeyi
      - NEm/NEg arzından karşılanabilir GCAA
      - Fazın kaba yem KM koridoru

    NDF/eNDF, toplam ve etkin rumen nişastası ile mineral pencereleri güvenlik
    raylarıdır. Güvensiz veya birden çok ana hedefi ciddi kaçıran aday kaydedilmez;
    kullanıcıya çözümü sınırlayan kısıtlar bildirilir.
    """
    import random as _random, time as _time
    t=ration_requirement_targets(weight_kg,target_adg,animal_type,age_months,phase_override)
    lim=beef_phase_limits(t['weight_kg'],target_adg,t['dmi_kg'],age_months,phase_override)
    t['roughage_min'],t['roughage_max']=lim['roughage_min'],lim['roughage_max']
    t['endf_min']=lim['endf_min'];t['starch_min']=lim['starch_min'];t['starch_ideal_max']=lim['starch_ideal_max'];t['starch_max']=lim['starch_max'];t['phase']=lim['phase']
    n=len(feeds)
    if n<2:return None,t,'En az 2 yem seçin.'

    bounds=[smart_feed_bounds(f,t['weight_kg'],t['dmi_kg'],target_adg,age_months,phase_override) for f in feeds]
    # Besi için uygun ticari yem seçilmişse süt yemini matematiksel kestirme olarak
    # kullanma. Besi yemi yoksa kullanıcı seçimini tamamen kilitlemeden profil
    # cezası mekanizması yalnız zorunlu durumda kullanımına izin verir.
    bounds=_apply_commercial_profile_bounds(feeds,bounds,animal_type)
    # Normal yemlerde 0 kullanımına izin verilir. Eski sürüm seçilen her kaba/kesif yeme
    # zorunlu minimum vererek saman/tahıl gibi düşük kaliteli kalemleri rasyonda tutabiliyordu.
    # 0 ile pratik minimum arasındaki anlamsız gramajlar _practical_qty_penalty ile cezalanır.

    rough=[i for i,f in enumerate(feeds) if feed_group(f)=='Kaba']
    conc=[i for i,f in enumerate(feeds) if feed_group(f)=='Kesif']
    if not rough or not conc:return None,t,'En az bir kaba ve bir kesif yem seçin.'

    dmfrac=[max(float(_rowval(f,'dm_pct',0))/100.0,.01) for f in feeds]
    fixed=[abs(hi-lo)<1e-10 for lo,hi in bounds]
    movable=[i for i in range(n) if not fixed[i]]
    rng=_random.Random(6172026)
    tol=.035
    adg_tol=.01
    rough_tol_pp={'Besi Başlangıç':3.0,'Besi Geliştirme':3.0,'Besi Bitirme':5.0}.get(t.get('phase'),3.5)
    target_dm=float(t['dmi_kg'])
    target_cp=float(t['cp_pct'])
    target_me=float(t['me_mcal_day'])
    target_rough=float(t.get('roughage_target') or 50.0)

    def clip(q):
        return [max(bounds[i][0],min(bounds[i][1],float(q[i] or 0))) for i in range(n)]

    def achieved_adg(m):
        return _achievable_adg(m,t)

    def safety_vector(m):
        # İzleme/gösterim rayları. Küçük bir NDF veya rumen-risk sapması dört ana besin
        # hedefinin önüne geçirilmez; ciddi riskler hard_safety_vector ile ayrılır.
        vals=[]
        vals.append(max(0.0,(t['ndf_min']-m['ndf_pct_dm'])/max(t['ndf_min'],1)))
        vals.append(max(0.0,(m['ndf_pct_dm']-t['ndf_max'])/max(t['ndf_max'],1)))
        vals.append(max(0.0,(lim['endf_min']-m['endf_pct_dm'])/max(lim['endf_min'],1)))
        vals.append(max(0.0,(m['starch_pct_dm']-lim['starch_max'])/max(lim['starch_max'],1)))
        cap=m['ca_g']/m['p_g'] if m['p_g']>0 else 99.0
        vals.append(max(0.0,(1.10-cap)/1.10)); vals.append(max(0.0,(cap-3.70)/3.70))
        mw=_mineral_windows(t,target_dm)
        vals.append(max(0.0,(m['ca_g']-mw['ca_hard'])/max(mw['ca_hard'],1)))
        vals.append(max(0.0,(m['p_g']-mw['p_hard'])/max(mw['p_hard'],1)))
        return vals

    def hard_safety_vector(m,q):
        """Ana hedeflerden önce gelen yalnız ciddi rumen/mineral riskleri.

        Önceki sıralama türetilmiş bir pH göstergesini veya çok küçük bir eNDF sapmasını bile KM/HP/ME
        fizibilitesinden üstün tutuyordu. Bu da solverın çok sayıda hedefi kaçıran,
        tahıl ağırlıklı bir adayı seçmesine yol açabiliyordu.
        """
        vals=[]
        vals.append(max(0.0,(t['ndf_min']*.85-m['ndf_pct_dm'])/max(t['ndf_min'],1)))
        vals.append(max(0.0,(m['ndf_pct_dm']-t['ndf_max']*1.15)/max(t['ndf_max'],1)))
        vals.append(max(0.0,(lim['endf_min']*.80-m['endf_pct_dm'])/max(lim['endf_min'],1)))
        # Fazın nişasta üst sınırı kayıt öncesi sert güvenlik kapısıdır.
        vals.append(max(0.0,(m['starch_pct_dm']-lim['starch_max'])/max(lim['starch_max'],1)))
        risk=_rumen_risk_assessment(m,lim)
        vals.append(1.0 if risk['level']=='Yüksek' else 0.0)
        grain_dm,wheat_dm,_=_grain_mix_dm(feeds,q)
        wheat_grain_share=wheat_dm/max(grain_dm,.01)
        vals.append(max(0.0,(wheat_grain_share-.40)/.40) if grain_dm>.05 else 0.0)
        cap=m['ca_g']/m['p_g'] if m['p_g']>0 else 99.0
        vals.append(max(0.0,(1.00-cap)/1.00)); vals.append(max(0.0,(cap-4.00)/4.00))
        mw=_mineral_windows(t,target_dm)
        vals.append(max(0.0,(m['ca_g']-mw['ca_hard'])/max(mw['ca_hard'],1)))
        vals.append(max(0.0,(m['p_g']-mw['p_hard'])/max(mw['p_hard'],1)))
        return vals

    def target_devs(m):
        dynamic_dm=_predicted_dmi_for_metrics(m,t)
        m['predicted_dmi_kg']=dynamic_dm
        dm=abs(m['dm_kg']-dynamic_dm)/max(dynamic_dm,.01)
        # HP, Ca ve P asgari gereksinimdir. Hedefin üstü eşitlik sapması değildir.
        cp=max(0.0,target_cp-m['cp_pct_dm'])/max(target_cp,.01)
        # Enerji başarısı toplam ME yerine NEm/NEg'den hesaplanan GCAA kapasitesidir.
        me=0.0
        adg=max(0.0,target_adg-float(m.get('achievable_adg_kg') or 0))/max(target_adg,.01)
        rough=abs(m['roughage_pct_dm']-target_rough)  # yuzde puani
        return dm,cp,me,adg,rough

    def rank(q):
        q=clip(q); m=smart_ration_metrics(feeds,q)
        m['achievable_adg_kg']=achieved_adg(m)
        dm,cp,me,adg,rough=target_devs(m)
        starch=float(m.get('starch_pct_dm') or 0)
        starch_soft=0.0
        if starch>lim['starch_ideal_max']:
            starch_soft=(starch-lim['starch_ideal_max'])/max(lim['starch_ideal_max'],1)
        if starch>lim['starch_max']:
            starch_soft+=3.0*(starch-lim['starch_max'])/max(lim['starch_max'],1)
        elif starch<lim['starch_min'] and me<target_me*.99:
            starch_soft=.25*(lim['starch_min']-starch)/max(lim['starch_min'],1)
        cp_signed=(m['cp_pct_dm']-target_cp)/max(target_cp,.01)
        me_signed=(m['me_mcal']-target_me)/max(target_me,.01)
        adg_signed=(m['achievable_adg_kg']-target_adg)/max(target_adg,.01)
        # Asgari hedefleri eşitlik kısıtına çevirmeden gereksiz yüksek arzı azalt.
        # HP'de %20 güvenlik payı, enerjiye göre GCAA kapasitesinde %5 saha payı
        # serbesttir; sonrası güvenlik hedeflerinden sonra, maliyetten önce sıralanır.
        cp_surplus_soft=max(0.0,cp_signed-.20)/.20
        adg_surplus_soft=max(0.0,adg_signed-.05)/.05
        surplus_soft=cp_surplus_soft+adg_surplus_soft
        outs=[max(0.0,dm-tol)/tol,
              max(0.0,-cp_signed-tol)/tol,
              0.0,
              max(0.0,abs(adg_signed)-adg_tol)/adg_tol,
              max(0.0,rough-rough_tol_pp)/rough_tol_pp]
        sv=safety_vector(m)
        hsv=hard_safety_vector(m,q)
        rumen_risk=_rumen_risk_assessment(m,lim)
        risk_soft=(2 if rumen_risk['level']=='Yüksek' else (1 if rumen_risk['level']=='Orta' else 0))
        if rumen_risk['confidence']=='Düşük':risk_soft*=.35
        safety_count=sum(v>1e-10 for v in sv)
        hard_safety_count=sum(v>1e-10 for v in hsv)
        target_count=sum(v>1e-10 for v in outs)
        grain_dm,wheat_dm,_=_grain_mix_dm(feeds,q)
        wheat_share=wheat_dm/max(grain_dm,.01) if grain_dm>.05 else 0.0
        wheat_soft=max(0.0,(wheat_share-.30)/.10)
        profile_soft=_commercial_profile_penalty(feeds,q,animal_type,target_dm)
        # Önce ciddi güvenlik, sonra KM/HP/ME/kaba fizibilitesi. Hafif NDF/eNDF/rumen-risk
        # sapmaları ancak ana hedeflerden sonra karşılaştırılır; maliyet en son gelir.
        key=(hard_safety_count,
             round(max(hsv) if hsv else 0.0,10),
             target_count,
             round(max(outs),10),
             round(sum(v*v for v in outs),10),
             round(profile_soft,8),
             round(starch_soft,8),
             round(wheat_soft,8),
             round(_practical_qty_penalty(feeds,q,t['weight_kg'],target_dm),8),
             safety_count,
             round(max(sv) if sv else 0.0,10),
             round(risk_soft,8),
             round(surplus_soft,8),
             # Aynı fizibilite seviyesinde önce uygulanabilirlik/kalite, sonra milimetrik kart farkı.
             # Bu sıra 0.01 kg gibi matematiksel ama saha dışı kalemleri ve ucuz yem dominansını azaltır.
             round(_feed_quality_penalty(feeds,q,target_dm),8),
             round(dm+cp+me+adg+rough/100.0,10),
             round(m['cost'],6))
        return key,m

    def allocate_group(q, inds, dm_need, weights):
        remaining=max(0.0,dm_need)
        active=list(inds)
        # alt sinirlardaki KM zaten q icinde; burada ilave KM dagitilir.
        for _ in range(12):
            if remaining<=1e-8 or not active:break
            sw=sum(max(weights.get(i,1.0),1e-6) for i in active) or 1.0
            used=0.0; nxt=[]
            for i in active:
                want=remaining*max(weights.get(i,1.0),1e-6)/sw
                cap=max(0.0,(bounds[i][1]-q[i])*dmfrac[i])
                take=min(want,cap)
                if take>0:q[i]+=take/dmfrac[i]; used+=take
                if cap-take>1e-7:nxt.append(i)
            if used<=1e-10:break
            remaining-=used; active=nxt
        return remaining

    def make_seed(rough_share, mode='balanced'):
        q=[lo for lo,hi in bounds]
        base_dm=sum(q[i]*dmfrac[i] for i in range(n))
        need=max(0.0,target_dm-base_dm)
        rcbase=sum(q[i]*dmfrac[i] for i in rough+conc)
        roughbase=sum(q[i]*dmfrac[i] for i in rough)
        target_rough_dm=max(0.0,target_dm*rough_share)
        rneed=max(0.0,target_rough_dm-roughbase)
        cneed=max(0.0,need-rneed)
        # Feed weights: Excel'in min-maliyet mantigina ek olarak dengeli baslangiclar.
        rw={}; cw={}
        for i in rough:
            f=feeds[i]; me=max(.05,_solver_nutrient(f,'me_mcal_kg')); cp=max(.1,_solver_nutrient(f,'cp_pct'))
            endf=max(.1,_solver_nutrient(f,'ndf_pct')*_solver_nutrient(f,'effective_ndf_pct')/100.0)
            price=max(.01,float(_rowval(f,'price',0) or .01))
            if mode=='energy': val=me+.02*cp+.005*endf
            elif mode=='cost': val=1.0/price
            else: val=.55*me+.015*cp+.008*endf
            rw[i]=max(.01,val)*(0.65+rng.random()*.70)
        for i in conc:
            f=feeds[i]; me=max(.05,_solver_nutrient(f,'me_mcal_kg')); cp=max(.1,_solver_nutrient(f,'cp_pct'))
            starch=_solver_starch_pct(f); price=max(.01,float(_rowval(f,'price',0) or .01))
            if mode=='protein': val=.10*cp+.35*me
            elif mode=='cost': val=1.0/price
            else: val=.65*me+.035*cp-.004*starch
            cw[i]=max(.01,val)*(0.65+rng.random()*.70)
        allocate_group(q,rough,rneed,rw)
        allocate_group(q,conc,cneed,cw)
        # kalan KM herhangi bir gruptaki kapasiteye dagitilir
        cur=sum(q[i]*dmfrac[i] for i in range(n)); rem=max(0.0,target_dm-cur)
        if rem>1e-8:
            allw={**rw,**cw}; allocate_group(q,rough+conc,rem,allw)
        return clip(q)

    # Geniş, deterministik coklu baslangic. Hedef kaba oraninin cevresinde ve faz raylari icinde.
    shares=[]
    for pp in (-10,-7,-4,-2,0,2,4,7,10):
        sh=max(lim['roughage_min'],min(lim['roughage_max'],target_rough+pp))/100.0
        if all(abs(sh-x)>.005 for x in shares): shares.append(sh)
    seeds=[]
    for sh in shares:
        for mode in ('balanced','energy','protein','cost'):
            for _ in range(2):seeds.append(make_seed(sh,mode))

    # Rastgele fizibilite taramasi: karar degiskenleri yem miktarlaridir; DMI ve kaba hedefi
    # etrafinda farkli dagilimlar olusturulur. Sabit RNG nedeniyle sonuc tekrarlanabilirdir.
    for _ in range(450):
        sh=max(lim['roughage_min'],min(lim['roughage_max'],target_rough+rng.uniform(-12,12)))/100.0
        seeds.append(make_seed(sh,rng.choice(('balanced','energy','protein','cost'))))

    scored=[]
    for q in seeds:
        k,m=rank(q); scored.append((k,q))
    scored.sort(key=lambda x:x[0])
    starts=[q for k,q in scored[:18]]

    def local_opt(q,deadline):
        q=clip(q); k,_=rank(q)
        # Excel Solver'in karar degiskeni mantigina yakin: iki yem arasinda ayni KM'yi takas et,
        # sonra tek yem koordinatinda toplam KM'yi ince ayarla. Buyuk adimdan kucuge iner.
        for dmstep in (.80,.55,.35,.22,.14,.09,.055,.035,.020,.010):
            improved=True; rounds=0
            while improved and rounds<8 and _time.perf_counter()<deadline:
                improved=False; rounds+=1; bestq=q; bestk=k
                # once tum pairwise KM takaslari
                order=list(movable); rng.shuffle(order)
                for i in order:
                    for j in order:
                        if i==j:continue
                        cand=q[:]; cand[i]+=dmstep/dmfrac[i]; cand[j]-=dmstep/dmfrac[j]; cand=clip(cand)
                        ck,_=rank(cand)
                        if ck<bestk:bestq,bestk=cand,ck
                # sonra DMI acigini/kacagini duzeltecek tek koordinat adimlari
                for i in order:
                    for sign in (-1,1):
                        cand=q[:]; cand[i]+=sign*dmstep/dmfrac[i]; cand=clip(cand)
                        ck,_=rank(cand)
                        if ck<bestk:bestq,bestk=cand,ck
                if bestk<k:q,k=bestq,bestk; improved=True
        return q,k

    started=_time.perf_counter(); deadline=started+4.5
    best=None; bestk=None
    for st in starts:
        if _time.perf_counter()>deadline and best is not None:break
        q,k=local_opt(st,deadline)
        if bestk is None or k<bestk:best,bestk=q,k

    if best is None:return None,t,'Rasyon optimizasyonu başlatılamadı.'
    # Kullanici sectigi normal yemleri kaybetmesin; katkilar kendi doz sinirinda kalir.
    best=[round(max(bounds[i][0],min(bounds[i][1],x)),3 if feed_group(feeds[i])=='Katkı' else 2) for i,x in enumerate(best)]
    finalk,bm=rank(best)
    bm['predicted_dmi_kg']=_predicted_dmi_for_metrics(bm,t)
    bm['achievable_adg_kg']=achieved_adg(bm)
    bm['solver_seconds']=_time.perf_counter()-started
    bm['solver_engine']='v3.9.20 Solver DEV4.18 · gerçek Sunar etiketi→NASEM hayvan profili→dinamik KM/GCAA→sert nişasta/profil→buğday→kalite/maliyet'
    bm['roughage_target_pct']=target_rough
    bm['rumen_risk']=_rumen_risk_assessment(bm,lim)
    bm['scientific_coverage']=_scientific_feed_coverage(feeds,best)
    feasibility=_solver_feasibility_report(bm,t,lim)
    bm['feasibility']=feasibility

    if feasibility['status'] in ('unsafe','infeasible'):
        reasons=feasibility['unsafe']+feasibility['blockers']
        detail='; '.join(reasons) if reasons else 'seçili yem kısıtları hedefleri aynı anda karşılamıyor'
        advice=_ration_assistant(feeds,bm,t,lim)
        return None,t,'Çözüm kaydedilmedi: '+detail+'. Sınırlayan kısıtlar düzeltilmeli. Öneri: '+advice

    warnings=[]
    warnings.extend(_feed_data_warnings(feeds))
    dm,cp,me,adg,rough=target_devs(bm)
    residual=[]
    final_dmi=max(float(bm.get('predicted_dmi_kg') or target_dm),.01)
    if dm>tol:residual.append(f'KM {(bm["dm_kg"]-final_dmi)/final_dmi*100:+.1f}%')
    cp_signed=(bm['cp_pct_dm']-target_cp)/max(target_cp,.01)
    if cp_signed < -tol:residual.append(f'HP tabanı {cp_signed*100:+.1f}%')
    adg_signed=(bm['achievable_adg_kg']-target_adg)/max(target_adg,.01)
    if adg_signed<-.005:residual.append(f'GCAA kapasitesi {adg_signed*100:+.1f}%')
    if rough>rough_tol_pp:residual.append(f'Kaba {bm["roughage_pct_dm"]-target_rough:+.1f} puan')
    sv=safety_vector(bm)
    if residual:
        warnings.insert(0,'En iyi saha çözümü üretildi; ±%3,5 bandı dışında kalan kartlar: '+', '.join(residual)+'.')
    if any(x>1e-10 for x in sv):
        warnings.append('Güvenlik raylarından biri sınırda; NDF/eNDF, nişasta, göreli asidoz riski ve mineral kartlarını kontrol edin.')
    grain_dm,wheat_dm,_=_grain_mix_dm(feeds,best)
    if grain_dm>.01 and wheat_dm/grain_dm>.30:
        wheat_share=wheat_dm/grain_dm*100
        band='sert %40 sınırının üzerinde' if wheat_share>40 else 'muhafazakâr %30 hedefinin üzerinde'
        warnings.append(f'Buğday tahıl karışımının %{wheat_share:.0f} KM payında; {band}. İşleme inceliği, kademeli adaptasyon, toplam nişasta ve eNDF birlikte kontrol edilmelidir.')
    used_commercial=[f for f,q in zip(feeds,best) if q>.01 and _is_commercial_compound_feed(f)]
    # Etiket eksikliği ancak solver ürünü otomatik güvenlik sınırının %90'ına kadar
    # zorluyorsa kullanıcıyı meşgul eder. Düşük/orta kullanımda uzun genel uyarı yoktur.
    unlabeled=[]
    for f,q in zip(feeds,best):
        if q<=.01 or not _is_commercial_compound_feed(f) or float(_rowval(f,'solver_max_kg_day',0) or 0)>0:continue
        _,auto_hi=smart_feed_bounds(f,t['weight_kg'],target_dm,target_adg,age_months,phase_override)
        if auto_hi>0 and q>=auto_hi*.90:unlabeled.append(str(_rowval(f,'name','Yem')))
    if unlabeled:
        warnings.append('Otomatik güvenlik sınırına yaklaşan ticari yemde etiket üst dozu eksik: '+', '.join(unlabeled)+'. Gerçek ürün etiketini Yem Kataloğu → Düzenle alanına ekleyin.')
    if bm['scientific_coverage']['starch_degradability']<.70 and bm['starch_kg']>.10:
        warnings.append(f'Nişasta yıkılabilirliği veri kapsamı %{bm["scientific_coverage"]["starch_degradability"]*100:.0f}; rumen risk güveni düşüktür. İşleme biçimi ve analiz değerlerini tamamlayın.')
    if feasibility['blockers'] or feasibility['warnings']:
        warnings.insert(0,'Sınırlı saha çözümü: '+', '.join(feasibility['blockers']+feasibility['warnings'])+'.')
    if residual or any(x>1e-10 for x in sv):
        warnings.append('Öneri: '+_ration_assistant(feeds,bm,t,lim))
    return (best,bm,bounds),t,' '.join(warnings)

def dairy_requirement_targets(weight_kg=650.0, target_milk_l=25.0, milk_fat_pct=3.7, milk_protein_pct=3.1):
    """Sağmal inek için Süt_V5.01 / INRA-NASEM yaklaşımından sadeleştirilmiş hedef motoru.

    Kullanıcı arayüzünde yalnız canlı ağırlık + hedef süt istenir. Yağ/protein için
    saha varsayımları kullanılır; ileride laboratuvar/sürü değerleri bağlanabilir.
    DMI hesabı örnek Excel'deki sağmal inek denkleminin orta laktasyon (100 DIM)
    varsayımıyla uygulanmış halidir.
    """
    import math
    w=max(350.0,min(float(weight_kg or 650),900.0)); milk=max(0.0,min(float(target_milk_l or 25),70.0))
    fat=max(2.5,min(float(milk_fat_pct or 3.7),6.5)); prot=max(2.5,min(float(milk_protein_pct or 3.1),5.0))
    # 4% FCM ve Süt_V5.01'deki sağmal DMI denklemi; kullanıcıdan DIM istememek için 100 gün orta laktasyon varsayımı.
    fcm=milk*(0.40+0.15*fat)
    dim=100.0
    dmi_kg=(((w**0.75)*0.0968)+(0.372*fcm)-0.293)*(1-math.exp(-0.192*(dim/7.0+3.67)))
    dmi_kg=max(w*0.020,min(w*0.045,dmi_kg))
    lactose=4.8
    milk_nel_per_kg=0.0929*fat+0.0563*prot+0.0395*lactose
    maintenance_nel=0.080*(w**0.75)
    # ME referans kartı: NEL gereksinimini yaklaşık laktasyon kullanım verimine dönüştürür.
    me_day=(maintenance_nel+milk*milk_nel_per_kg)/0.64
    cp_pct=max(14.5,min(18.5,14.3+milk*0.075))
    ca_pct=max(0.62,min(0.90,0.62+milk*0.004)); p_pct=max(0.34,min(0.46,0.34+milk*0.0022))
    # Süt rasyonunda fiziksel etkili lif korunur; süt yükseldikçe kesif payı kontrollü artabilir.
    rough_target=max(42.0,min(58.0,55.0-0.45*max(0.0,milk-20.0)))
    rough_min=max(40.0,rough_target-6.0); rough_max=min(62.0,rough_target+6.0)
    return {'mode':'Süt','weight_kg':w,'milk_l':milk,'milk_fat_pct':fat,'milk_protein_pct':prot,
            'dmi_pct_bw':dmi_kg/w*100,'dmi_kg':dmi_kg,'cp_pct':cp_pct,'me_mcal_day':me_day,
            'me_mcal_kg':me_day/max(dmi_kg,.01),'ca_g':dmi_kg*ca_pct*10,'p_g':dmi_kg*p_pct*10,
            'ca_pct':ca_pct,'p_pct':p_pct,'ndf_min':28.0,'ndf_max':34.0,
            'roughage_target':rough_target,'roughage_min':rough_min,'roughage_max':rough_max,
            'endf_min':18.0,'starch_min':22.0,'starch_ideal_max':27.0,'starch_max':30.0,
            'source_note':'Süt_V5.01 / INRA-NASEM saha hedef motoru'}


def dairy_feed_bounds(feed, weight_kg, target_dm):
    """Süt rasyonu için pratik günlük yaş-yem ve ürün etiketi sınırları."""
    dm=max(float(_rowval(feed,'dm_pct',0))/100.0,.05); grp=feed_group(feed); name=str(_rowval(feed,'name','')).upper()
    if grp=='Katkı':
        if 'TUZ' in name:return _apply_explicit_feed_limits(feed,(0.0,min(0.15,target_dm*0.008)))
        if any(x in name for x in ('VİTAMİN','VITAMIN','PREMİKS','PREMIKS','PREMIX')):return _apply_explicit_feed_limits(feed,(0.0,0.25))
        if any(x in name for x in ('BİKARBONAT','BIKARBONAT')):return _apply_explicit_feed_limits(feed,(0.0,min(0.30,target_dm*0.015)))
        if any(x in name for x in ('MERMER','KİREÇ','KIREC','KALSİYUM','KALSIYUM')):return _apply_explicit_feed_limits(feed,(0.0,min(0.15,target_dm*0.008)))
        return _apply_explicit_feed_limits(feed,(0.0,min(0.20,target_dm*0.010)))
    if grp=='Kaba':
        if 'SİLAJ' in name or 'SILAJ' in name:return _apply_explicit_feed_limits(feed,(0.0,max(1.0,target_dm*0.70/dm)))
        return _apply_explicit_feed_limits(feed,(0.0,max(0.5,target_dm*0.55/dm)))
    # Tam karma/süt yemleri enerji-protein taşıyıcıdır; tek tahıl gibi sıkıştırılmaz.
    if _is_commercial_compound_feed(feed) or any(x in name for x in ('YEMİ','YEMI')):
        return _apply_explicit_feed_limits(feed,(0.0,max(1.0,target_dm*0.58/dm)))
    starch=_solver_starch_pct(feed)
    frac=0.30 if starch>=45 else 0.40
    return _apply_explicit_feed_limits(feed,(0.0,max(0.5,target_dm*frac/dm)))


def solve_smart_dairy_ration(feeds, weight_kg, target_milk_l):
    """Aynı ÇiftlikPro yem kataloğunu kullanarak otomatik süt rasyonu miktar optimizasyonu."""
    import time as _time
    t=dairy_requirement_targets(weight_kg,target_milk_l)
    n=len(feeds)
    if n<2:return None,t,'Süt rasyonu çözmek için en az 2 yem seçin.'
    bounds=[dairy_feed_bounds(f,t['weight_kg'],t['dmi_kg']) for f in feeds]
    dmfrac=[max(float(_rowval(f,'dm_pct',0))/100.0,.05) for f in feeds]
    fixed=[feed_group(f)=='Katkı' for f in feeds]
    rough=[i for i,f in enumerate(feeds) if feed_group(f)=='Kaba']
    conc=[i for i,f in enumerate(feeds) if feed_group(f)=='Kesif']
    normal=[i for i in range(n) if not fixed[i]]
    # Kullanıcının seçtiği normal yemler rasyonda kalır; katkılar gereksinime göre 0 olabilir.
    mins=[]
    for i,f in enumerate(feeds):
        if fixed[i]:mins.append(bounds[i][0])
        else:mins.append(max(bounds[i][0],min(bounds[i][1],practical_feed_min(f,t['weight_kg'],t['dmi_kg']))))
    def clip(q):return [max(mins[i],min(bounds[i][1],float(q[i]))) for i in range(n)]
    def score(q):
        m=smart_ration_metrics(feeds,q)
        def rel(v,x):return abs(float(v)-float(x))/max(abs(float(x)),.01)
        # Dört ana süt hedefi: KM, HP, ME, kaba/kesif. NDF/eNDF ve göreli
        # asidoz riski güvenlik raylarıdır; klinik pH tahmini yapılmaz.
        sc=13*rel(m['dm_kg'],t['dmi_kg'])+14*rel(m['cp_pct_dm'],t['cp_pct'])+14*rel(m['me_mcal'],t['me_mcal_day'])+10*rel(m['roughage_pct_dm'],t['roughage_target'])
        if m['ndf_pct_dm']<t['ndf_min']:sc+=18*(t['ndf_min']-m['ndf_pct_dm'])/t['ndf_min']
        if m['ndf_pct_dm']>t['ndf_max']:sc+=7*(m['ndf_pct_dm']-t['ndf_max'])/t['ndf_max']
        if m['endf_pct_dm']<t['endf_min']:sc+=22*(t['endf_min']-m['endf_pct_dm'])/t['endf_min']
        if m['starch_pct_dm']>t['starch_max']:sc+=12*(m['starch_pct_dm']-t['starch_max'])/t['starch_max']
        risk=_rumen_risk_assessment(m,t)
        if risk['level']=='Yüksek':sc+=35
        elif risk['level']=='Orta':sc+=6
        if m['roughage_pct_dm']<t['roughage_min']:sc+=15*(t['roughage_min']-m['roughage_pct_dm'])/t['roughage_min']
        if m['roughage_pct_dm']>t['roughage_max']:sc+=8*(m['roughage_pct_dm']-t['roughage_max'])/t['roughage_max']
        # Mineral aşırılıklarını ikincil ama görünür şekilde cezalandır.
        if m['ca_g']>t['ca_g']*1.5:sc+=4*(m['ca_g']/max(t['ca_g'],1)-1.5)
        if m['p_g']>t['p_g']*1.5:sc+=4*(m['p_g']/max(t['p_g'],1)-1.5)
        return sc+0.0015*m['cost'],m
    # Başlangıç: minler + kalan KM'yi kaba/kesif hedefe göre dağıt.
    q=clip(mins[:]); curdm=sum(q[i]*dmfrac[i] for i in range(n)); need=max(0.0,t['dmi_kg']-curdm)
    rneed=need*(t['roughage_target']/100.0); cneed=max(0.0,need-rneed)
    def distribute(indices,dm_need):
        nonlocal q
        if not indices or dm_need<=0:return
        # Enerji + protein yoğun yemlere biraz daha ağırlık ver, ancak hepsini kullan.
        weights=[]
        for i in indices:
            f=feeds[i]; wgt=max(0.3,float(_solver_nutrient(f,'me_mcal_kg') or 0)+0.03*float(_solver_nutrient(f,'cp_pct') or 0))
            weights.append(wgt)
        sw=sum(weights) or 1
        for i,wgt in zip(indices,weights):
            adddm=dm_need*wgt/sw; q[i]=min(bounds[i][1],q[i]+adddm/dmfrac[i])
    distribute(rough,rneed); distribute(conc,cneed); q=clip(q)
    best=q[:]; bestscore,bm=score(best); deadline=_time.perf_counter()+2.5
    # Pairwise DM takasları ve küçük toplam KM ayarları.
    for stepdm in (0.45,0.25,0.12,0.06,0.03):
        improved=True; loops=0
        while improved and loops<6 and _time.perf_counter()<deadline:
            improved=False; loops+=1; cand_best=None; cand_score=bestscore
            for i in normal:
                # Tek yemi artır/azalt: toplam KM hedefi de hareket edebilsin.
                for sign in (-1,1):
                    cand=best[:]; cand[i]+=sign*stepdm/dmfrac[i]; cand=clip(cand); sc,_=score(cand)
                    if sc+1e-9<cand_score:cand_best,cand_score=cand,sc
                for j in normal:
                    if i==j:continue
                    cand=best[:]; cand[i]+=stepdm/dmfrac[i]; cand[j]-=stepdm/dmfrac[j]; cand=clip(cand); sc,_=score(cand)
                    if sc+1e-9<cand_score:cand_best,cand_score=cand,sc
            if cand_best is not None:best,bestscore=cand_best,cand_score; improved=True
    best=[round(max(mins[i],min(bounds[i][1],best[i])),3 if fixed[i] else 2) for i in range(n)]
    _,bm=score(best)
    bm['rumen_risk']=_rumen_risk_assessment(bm,t)
    bm['solver_engine']='6.16 Süt_V5.01/INRA-NASEM · KM+HP+ME+kaba/kesif çok hedefli süt optimizer'
    notes=[]
    for label,val,target in [('KM',bm['dm_kg'],t['dmi_kg']),('HP',bm['cp_pct_dm'],t['cp_pct']),('ME',bm['me_mcal'],t['me_mcal_day']),('Kaba',bm['roughage_pct_dm'],t['roughage_target'])]:
        pct=(float(val)-float(target))/max(abs(float(target)),.01)*100
        if abs(pct)>5:notes.append(f'{label} {pct:+.1f}%')
    if notes:notes.insert(0,'En iyi süt rasyonu oluşturuldu; hedef bandı dışında:')
    return (best,bm,bounds),t,' '.join(notes)

def ration_targets_for_record(rr):
    typ=(rr['ration_type'] if 'ration_type' in rr.keys() and rr['ration_type'] else 'Besi').strip()
    if typ.lower().startswith(('süt','sut')):
        return dairy_requirement_targets(rr['target_weight_kg'],rr['target_milk_l'],rr['milk_fat_pct'],rr['milk_protein_pct'])
    return ration_requirement_targets(rr['target_weight_kg'],rr['target_adg_kg'],rr['animal_type'],_rowval(rr,'target_age_months',0),_rowval(rr,'target_beef_phase','Otomatik'))

def is_real_protein_feed(feed):
    """Gerçek yem proteini ile NPN/mineral eşdeğerlerini ayırır.
    Protein önerisi için yalnız besleyici yem hammaddeleri kullanılır.
    """
    name=str(feed['name'] or '').upper()
    cp=float(feed['cp_pct'] or 0); me=float(feed['me_mcal_kg'] or 0); tdn=float(feed['tdn_pct'] or 0); ash=float(feed['ash_pct'] or 0) if 'ash_pct' in feed.keys() else 0.0
    blocked=('AMONYUM','ÜRE','URE','BİÜRE','BIURE','NPN','FOSFAT','KİREÇ','KIREC','TUZ','SODYUM BİKARBONAT','SODYUM BIKARBONAT')
    if any(x in name for x in blocked): return False
    # %100'ün üstündeki HP değerleri çoğunlukla NPN protein eşdeğeridir.
    if cp<=0 or cp>70: return False
    if me<=0 and tdn<=0: return False
    if ash>=35: return False
    return True

def nutrient_status(actual,target,tol=0.05,upper_tol=None):
    if target<=0:return ('-', 'mut')
    upper_tol=tol if upper_tol is None else float(upper_tol)
    ratio=actual/target
    if ratio < 1-tol:return (f'⚠️ Eksik %{(1-ratio)*100:.0f}','red')
    if ratio > 1+upper_tol:return (f'⬆️ Fazla %{(ratio-1)*100:.0f}','orange')
    return ('✅ Uygun','preg pos')

def _ration_requirement_panel_legacy(rr,sm):
    t=ration_targets_for_record(rr)
    cp_s,cp_c=nutrient_status(sm['cp_pct_dm'],t['cp_pct'],0.05,0.10); me_s,me_c=nutrient_status(sm['me_mcal'],t['me_mcal_day'],0.08,0.10); dm_s,dm_c=nutrient_status(sm['dm_kg'],t['dmi_kg'],0.10); ca_s,ca_c=nutrient_status(sm['ca_g'],t['ca_g'],0.10); p_s,p_c=nutrient_status(sm['p_g'],t['p_g'],0.10)
    ndf=sm['ndf_pct_dm']; ndf_s='✅ Uygun' if t['ndf_min']<=ndf<=t['ndf_max'] else ('⚠️ Düşük' if ndf<t['ndf_min'] else '⚠️ Yüksek')
    starch=sm.get('starch_pct_dm',0);starch_min=t.get('starch_min',0);starch_ideal_max=t.get('starch_ideal_max',t.get('starch_max',100));starch_max=t.get('starch_max',100)
    starch_s='✅ Uygun' if starch_min<=starch<=starch_ideal_max else ('ℹ️ İdeal altı' if starch<starch_min else ('⚠️ Sınıra yakın' if starch<=starch_max else '🔴 Yüksek'))
    rough=sm.get('roughage_pct_dm',0); conc=sm.get('concentrate_pct_dm',0); rc_s='✅ Uygun' if (t['roughage_min']-0.05)<=rough<=(t['roughage_max']+0.05) else ('⚠️ Kaba yem düşük' if rough<t['roughage_min'] else '⚠️ Kaba yem yüksek')
    if t['mode']=='Süt':
        form_fields=f"""<input type='hidden' name='ration_type' value='Süt'><label>Canlı Ağırlık (kg)<input type='number' min='350' max='900' step='1' name='target_weight_kg' value='{t['weight_kg']:.0f}'></label><label>Hedef Süt (L/gün)<input type='number' min='0' max='70' step='0.5' name='target_milk_l' value='{t['milk_l']:.1f}'></label><details class='target-more'><summary>🥛 Gelişmiş</summary><div class='target-more-grid'><label>Süt Yağı %<input type='number' min='2.5' max='6.5' step='0.1' name='milk_fat_pct' value='{t['milk_fat_pct']:.1f}'></label><label>Süt Proteini %<input type='number' min='2.5' max='5' step='0.1' name='milk_protein_pct' value='{t['milk_protein_pct']:.1f}'></label></div></details>"""
        title='🥛 Akıllı Süt Rasyonu Hedefi'; ctx=f"<div class='target-context compact'><b>{t['weight_kg']:.0f} kg</b> · hedef <b>{t['milk_l']:.1f} L/gün süt</b></div>"
    else:
        at=h(rr['animal_type'] or 'Besi Erkek')
        form_fields=f"""<input type='hidden' name='ration_type' value='Besi'><label>Canlı Ağırlık (kg)<input type='number' min='150' max='900' step='1' name='target_weight_kg' value='{t['weight_kg']:.0f}'></label><label>Hedef Artış (kg/gün)<input type='number' min='0.2' max='2.2' step='0.05' name='target_adg_kg' value='{t['adg']:.2f}'></label><label>Yaş (ay, opsiyonel)<input type='number' min='0' max='36' step='1' name='target_age_months' value='{t.get('age_months',0):.0f}'></label><label>Besi Dönemi<select name='target_beef_phase'><option value='Otomatik' {'selected' if _rowval(rr,'target_beef_phase','Otomatik')=='Otomatik' else ''}>Otomatik (canlı ağırlığa göre)</option><option {'selected' if _rowval(rr,'target_beef_phase','Otomatik')=='Besi Başlangıç' else ''}>Besi Başlangıç</option><option {'selected' if _rowval(rr,'target_beef_phase','Otomatik')=='Besi Geliştirme' else ''}>Besi Geliştirme</option><option {'selected' if _rowval(rr,'target_beef_phase','Otomatik')=='Besi Bitirme' else ''}>Besi Bitirme</option></select></label><label>Hayvan Tipi<input name='animal_type' value='{at}'></label>"""
        capacity=_achievable_adg(sm,t) if sm.get('nem_density',0)>0 and sm.get('neg_density',0)>0 else 0.0
        capacity_text=f" · rasyon enerji kapasitesi <b>{capacity:.2f} kg/gün</b>" if capacity>0 else ''
        title='🎯 Akıllı Besi Rasyonu Hedefi'; ctx=f"<div class='target-context compact'><b>{t['weight_kg']:.0f} kg</b> · hedef <b>{t['adg']:.2f} kg/gün artış</b>{capacity_text} · <b>{t.get('phase','')}</b> ({t.get('phase_mode','Otomatik')}) · faz K/K <b>%{t.get('roughage_min',0):.0f}–{t.get('roughage_max',0):.0f} / %{100-t.get('roughage_max',0):.0f}–{100-t.get('roughage_min',0):.0f}</b> · NEm {t.get('nem_req_mcal',0):.1f} / NEg {t.get('neg_req_mcal',0):.1f} Mcal</div>"
    def diff_text(actual,target,unit='',digits=1):
        if not target:return ''
        d=float(actual)-float(target); pct=d/float(target)*100; sign='+' if d>0 else ''
        return f"{sign}{d:.{digits}f}{unit} ({sign}{pct:.0f}%)"
    def box(key,label,current,target,status,diff=''):
        cls='ok' if ('Uygun' in status or 'Canlı' in status) else 'warn'
        return f"""<div class='nutri-mini nutri-compare-card {cls}' id='target-mini-{key}'>
            <div class='nutri-card-title'>{label}</div>
            <div class='nutri-compare-body'>
                <div class='nutri-side nutri-target-side'><span>HEDEF</span><b>{target}</b></div>
                <div class='nutri-side nutri-current-side'><span>RASYON</span><b id='target-mini-{key}-current'>{current}</b></div>
            </div>
            <div class='nutri-card-footer'><i class='nutri-diff' id='target-mini-{key}-diff'>{diff}</i><em id='target-mini-{key}-status'>{status}</em></div>
        </div>"""
    mineral_status='✅ Uygun' if ('Uygun' in ca_s and 'Uygun' in p_s) else '⚠️ Mineral dengesi'
    mineral_cls='ok' if 'Uygun' in mineral_status else 'warn'
    mineral=f"""<div class='nutri-mini nutri-compare-card mineral-card {mineral_cls}' id='target-mini-mineral'>
        <div class='nutri-card-title'>Ca + P</div>
        <div class='mineral-pairs'>
          <div class='nutri-pair'><span>Ca</span><div><small>HEDEF</small><b>{t['ca_g']:.0f} g</b></div><div><small>RASYON</small><b id='target-mini-ca-current'>{sm['ca_g']:.0f} g</b></div></div>
          <div class='nutri-pair'><span>P</span><div><small>HEDEF</small><b>{t['p_g']:.0f} g</b></div><div><small>RASYON</small><b id='target-mini-p-current'>{sm['p_g']:.0f} g</b></div></div>
        </div>
        <div class='nutri-card-footer mineral-footer'><span><em id='target-mini-ca-status'>{ca_s}</em><i class='nutri-diff' id='target-mini-ca-diff'>{diff_text(sm['ca_g'],t['ca_g'],' g',0)}</i></span><span><em id='target-mini-p-status'>{p_s}</em><i class='nutri-diff' id='target-mini-p-diff'>{diff_text(sm['p_g'],t['p_g'],' g',0)}</i></span></div>
    </div>"""
    rumen_risk=_rumen_risk_assessment(sm,t)
    risk_s={'Düşük':'✅ Düşük','Orta':'⚠️ Orta','Yüksek':'🔴 Yüksek'}[rumen_risk['level']]
    starch_rumen_cls='ok' if ('Uygun' in starch_s and rumen_risk['level']=='Düşük') else 'warn'
    starch_rumen=f"""<div class='nutri-mini nutri-compare-card mineral-card {starch_rumen_cls}' id='target-mini-starch-rumen'>
        <div class='nutri-card-title' title='Göreli risk göstergesidir; klinik rumen pH tahmini değildir'>Nişasta + Rumen ⓘ</div>
        <div class='mineral-pairs'>
          <div class='nutri-pair'><span>NİŞASTA</span><div><small>HEDEF</small><b>%{starch_min:.0f}–{starch_ideal_max:.0f}</b></div><div><small>RASYON</small><b id='target-mini-starch-current'>%{starch:.1f}</b></div></div>
          <div class='nutri-pair'><span>ASİDOZ RİSKİ</span><div><small>HEDEF</small><b>Düşük</b></div><div><small>GÖSTERGE</small><b id='target-mini-ph-current'>{rumen_risk['level']}</b></div></div>
        </div>
        <div class='nutri-card-footer mineral-footer'><span><em id='target-mini-starch-status'>{starch_s}</em><i class='nutri-diff' id='target-mini-starch-diff'>{sm.get('starch_kg',0):.2f} kg/baş</i></span><span><em id='target-mini-ph-status'>{risk_s}</em><i class='nutri-diff' id='target-mini-ph-diff'>Veri güveni {rumen_risk['confidence']}</i></span></div>
    </div>"""
    mini=''.join([
        box('dm','KM',f"{sm['dm_kg']:.2f} kg",f"{t['dmi_kg']:.2f} kg",dm_s,diff_text(sm['dm_kg'],t['dmi_kg'],' kg',2)),
        box('cp','HP',f"%{sm['cp_pct_dm']:.1f}",f"%{t['cp_pct']:.1f}",cp_s,diff_text(sm['cp_pct_dm'],t['cp_pct'],' puan',1)),
        box('me','ME (referans)',f"{sm['me_mcal']:.1f} Mcal",f"{t['me_mcal_day']:.1f} Mcal",me_s,diff_text(sm['me_mcal'],t['me_mcal_day'],' Mcal',1)),
        box('ndf','NDF',f"%{ndf:.1f}",f"%{t['ndf_min']:.0f}–{t['ndf_max']:.0f}",ndf_s,'Aralık içi' if t['ndf_min']<=ndf<=t['ndf_max'] else ('Alt sınırın altında' if ndf<t['ndf_min'] else 'Üst sınırın üzerinde')),
        mineral,
        box('rc','Kaba/Kesif',f"%{rough:.0f} / %{conc:.0f}",f"Kaba %{t['roughage_min']:.0f}–{t['roughage_max']:.0f}",rc_s,'Aralık içi' if (t['roughage_min']-0.05)<=rough<=(t['roughage_max']+0.05) else ('Kaba yem düşük' if rough<t['roughage_min'] else 'Kaba yem yüksek')),
        starch_rumen,
        box('cost','Maliyet',money(sm['cost']),money(sm['cost']),'💰 Canlı','Değişiklik yok'),
    ])
    detail=f"""<details class='nutri-detail'><summary>📋 Besin detaylarını göster</summary><div class='table-compact-wrap'><table class='ration-target-table compact-table zebra'><thead><tr><th>Besin</th><th>Hedef</th><th>Mevcut</th><th>Durum</th></tr></thead><tbody><tr><td>Kuru Madde</td><td>{t['dmi_kg']:.2f} kg/gün</td><td>{sm['dm_kg']:.2f} kg</td><td>{dm_s}</td></tr><tr><td>Ham Protein</td><td>%{t['cp_pct']:.1f} KM</td><td>%{sm['cp_pct_dm']:.1f} KM</td><td>{cp_s}</td></tr><tr><td>Metabolik Enerji</td><td>{t['me_mcal_day']:.1f} Mcal/gün</td><td>{sm['me_mcal']:.1f} Mcal</td><td>{me_s}</td></tr><tr><td>NDF</td><td>%{t['ndf_min']:.0f}–{t['ndf_max']:.0f}</td><td>%{ndf:.1f}</td><td>{ndf_s}</td></tr><tr><td>eNDF</td><td>En az %{t.get('endf_min',0):.1f}</td><td>%{sm.get('endf_pct_dm',0):.1f}</td><td>{'✅ Uygun' if sm.get('endf_pct_dm',0)>=t.get('endf_min',0) else '⚠️ Düşük'}</td></tr><tr><td>Nişasta</td><td>İdeal %{starch_min:.0f}–{starch_ideal_max:.0f} · üst %{starch_max:.0f}</td><td>%{starch:.1f} KM · {sm.get('starch_kg',0):.2f} kg</td><td>{starch_s}</td></tr><tr><td>Etkin rumen nişastası</td><td>Veri kapsamıyla izlenir</td><td>%{sm.get('rapid_starch_pct_dm',0):.1f} KM · kapsam %{sm.get('starch_degradability_coverage',0)*100:.0f}</td><td>{risk_s}</td></tr><tr><td>Asidoz riski</td><td>Düşük</td><td>{rumen_risk['level']} · güven {rumen_risk['confidence']}</td><td>Klinik pH tahmini değildir</td></tr><tr><td>Kalsiyum</td><td>{t['ca_g']:.0f} g</td><td>{sm['ca_g']:.0f} g</td><td>{ca_s}</td></tr><tr><td>Fosfor</td><td>{t['p_g']:.0f} g</td><td>{sm['p_g']:.0f} g</td><td>{p_s}</td></tr><tr><td>Kaba/Kesif</td><td>Kaba %{t['roughage_min']:.0f}–{t['roughage_max']:.0f}</td><td>%{rough:.0f} / %{conc:.0f}</td><td>{rc_s}</td></tr></tbody></table></div></details>"""
    return f"""<style>
body:has(.workbench-shell) .nutri-mini-grid{{grid-template-columns:repeat(4,minmax(0,1fr))!important;gap:7px!important}}
body:has(.workbench-shell) .nutri-mini.nutri-compare-card{{height:100px!important;min-height:100px!important;grid-template-rows:23px 1fr 28px!important}}
body:has(.workbench-shell) .nutri-card-title{{padding:4px 7px!important;font-size:12px!important;line-height:15px!important}}
body:has(.workbench-shell) .nutri-side{{padding:3px 5px!important;justify-content:center!important}}
body:has(.workbench-shell) .nutri-side span{{font-size:8px!important}}
body:has(.workbench-shell) .nutri-side b{{font-size:17px!important;white-space:nowrap!important;overflow:hidden!important;text-overflow:ellipsis!important}}
body:has(.workbench-shell) .nutri-card-footer{{height:28px!important;min-height:28px!important;padding:2px 5px!important;display:grid!important;grid-template-columns:auto 1fr!important;align-items:center!important;gap:4px!important}}
body:has(.workbench-shell) .nutri-card-footer em{{font-size:10px!important;white-space:nowrap!important;overflow:visible!important;text-overflow:clip!important;min-height:0!important}}
body:has(.workbench-shell) .nutri-card-footer .nutri-diff{{font-size:8.5px!important;white-space:nowrap!important;overflow:hidden!important;text-overflow:ellipsis!important;min-height:0!important;text-align:right!important}}
body:has(.workbench-shell) .mineral-card{{grid-template-rows:23px 1fr 28px!important}}
body:has(.workbench-shell) .mineral-pairs{{display:grid;grid-template-columns:1fr 1fr;min-height:0}}
body:has(.workbench-shell) .nutri-pair{{display:grid;grid-template-columns:1fr 1fr;grid-template-rows:10px 1fr;align-items:center;text-align:center;padding:1px 3px;border-right:1px solid #e0e9e3;min-width:0}}
body:has(.workbench-shell) .nutri-pair:last-child{{border-right:0}}
body:has(.workbench-shell) .nutri-pair>span{{grid-column:1/-1;font-size:7.5px!important;font-weight:900;color:#28724c;line-height:1}}
body:has(.workbench-shell) .nutri-pair>div{{min-width:0}}
body:has(.workbench-shell) .nutri-pair small{{display:block;font-size:7px!important;color:#668074;line-height:1}}
body:has(.workbench-shell) .nutri-pair b{{font-size:14px!important;line-height:1.03;font-weight:900;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}}
body:has(.workbench-shell) .mineral-footer{{display:grid!important;grid-template-columns:1fr 1fr!important;gap:5px!important}}
body:has(.workbench-shell) .mineral-footer>span{{display:flex;align-items:center;justify-content:center;gap:4px;min-width:0}}
body:has(.workbench-shell) .mineral-footer em,body:has(.workbench-shell) .mineral-footer .nutri-diff{{font-size:7.8px!important;min-height:0!important;line-height:1!important;width:auto!important}}
@media(max-width:1180px){{body:has(.workbench-shell) .nutri-mini-grid{{grid-template-columns:repeat(2,minmax(145px,1fr))!important}}}}
@media(max-width:820px){{body:has(.workbench-shell) .nutri-mini-grid{{grid-template-columns:repeat(2,minmax(0,1fr))!important}}body:has(.workbench-shell) .nutri-mini.nutri-compare-card{{height:102px!important;min-height:102px!important}}body:has(.workbench-shell) .nutri-side b{{font-size:16px!important}}body:has(.workbench-shell) .nutri-card-footer em{{font-size:9px!important}}body:has(.workbench-shell) .nutri-card-footer .nutri-diff{{font-size:8px!important}}}}
</style><div class='target-workspace'><div class='target-controlbar'><div class='target-head'><h3>{title}</h3>{ctx}</div><form method='post' action='/ration/target' class='target-form'><input type='hidden' name='ration_id' value='{rr['id']}'>{form_fields}<button class='btn blue compact-target-btn'>Güncelle</button></form></div><div class='target-compare-sticky'><div class='target-compare-title'><b>🎯 Hedef ↔ Mevcut</b><span id='target-live-note'>Göreli asidoz göstergesi · Klinik pH tahmini değildir</span></div><div class='nutri-mini-grid'>{mini}</div></div></div>"""

def ration_requirement_panel(rr,sm):
    """Besi hedeflerini gereksinim, arz ve performans olarak doğru anlamda gösterir.

    Eski kartlar CP/Ca/P minimumlarını eşitlik hedefi, toplam ME'yi ise doğrudan
    büyüme başarısı gibi gösteriyordu. Bu panel solver ile aynı dinamik DMI ve
    NEm/NEg tabanlı GCAA kapasitesini kullanır. Süt paneli mevcut motor
    yenilenene kadar kendi doğrulanmış eski görünümünü korur.
    """
    t=ration_targets_for_record(rr)
    if t.get('mode')=='Süt':
        return _ration_requirement_panel_legacy(rr,sm)

    predicted_dmi=_predicted_dmi_for_metrics(sm,t)
    sm['predicted_dmi_kg']=predicted_dmi
    capacity=_achievable_adg(sm,t) if sm.get('nem_density',0)>0 and sm.get('neg_density',0)>0 else 0.0
    gain_supply,_,_=_energy_balance(sm,t)
    limits=beef_phase_limits(t['weight_kg'],t['adg'],predicted_dmi,t.get('age_months',0),t.get('phase','Otomatik'))
    rumen_risk=_rumen_risk_assessment(sm,limits)
    mw=_mineral_windows(t,predicted_dmi)

    def symmetric_status(actual,target,soft=.05,hard=.10):
        if target<=0:return 'ℹ️ Veri yok'
        delta=(actual-target)/target
        if delta < -hard:return f'🔴 Eksik %{abs(delta)*100:.0f}'
        if delta > hard:return f'⚠️ Yüksek %{delta*100:.0f}'
        if abs(delta)>soft:return '⚠️ Sınıra yakın'
        return '✅ Uygun'

    def minimum_status(actual,minimum,soft=.05,hard=.10):
        if minimum<=0:return 'ℹ️ Veri yok'
        delta=(actual-minimum)/minimum
        if delta < -hard:return f'🔴 Eksik %{abs(delta)*100:.0f}'
        # Asgari hedefte %5'lik "yeşil tolerans" yanıltıcıdır. Yalnız ölçüm ve
        # yuvarlama payı olan %0,5 korunur; bunun altındaki arz açıkça uyarılır.
        if delta < -.005:return f'⚠️ Hedef altı %{abs(delta)*100:.1f}'
        return '✅ Yeterli'

    def protein_status(actual,minimum):
        base=minimum_status(actual,minimum,.05,.10)
        if minimum<=0 or actual<minimum:return base
        surplus=(actual-minimum)/minimum
        if surplus>.25:return f'⚠️ Yüksek %{surplus*100:.0f}'
        if surplus>.15:return f'ℹ️ Güvenlik payı +%{surplus*100:.0f}'
        return '✅ Yeterli'

    def capacity_status(actual,target):
        if target<=0:return 'ℹ️ Veri yok'
        delta=(actual-target)/target
        if delta<-.10:return f'🔴 Eksik %{abs(delta)*100:.0f}'
        if delta<-.005:return f'⚠️ Hedef altı %{abs(delta)*100:.1f}'
        if delta>.10:return f'⚠️ Kapasite +%{delta*100:.0f}'
        if delta>.05:return f'ℹ️ Kapasite +%{delta*100:.0f}'
        return '✅ Hedefe yakın'

    def mineral_status(actual,key):
        minimum=mw[f'{key}_target'];soft=mw[f'{key}_soft'];hard=mw[f'{key}_hard']
        if actual<minimum*.90:return f'🔴 Eksik %{(1-actual/minimum)*100:.0f}'
        if actual<minimum:return '⚠️ Minimuma yakın'
        if actual>hard:return '🔴 Güvenlik üstü'
        if actual>soft:return '⚠️ Yüksek'
        return '✅ Yeterli'

    def range_status(actual,low,high,label=''):
        if actual<low:return '🔴 Düşük'
        if actual>high:return '⚠️ Yüksek'
        return '✅ Aralıkta'

    dm_s=symmetric_status(sm['dm_kg'],predicted_dmi,.05,.10)
    adg_s=capacity_status(capacity,t['adg']) if capacity>0 else 'ℹ️ NEm/NEg verisi yok'
    cp_s=protein_status(sm['cp_pct_dm'],t['cp_pct'])
    ndf_s=range_status(sm['ndf_pct_dm'],t['ndf_min'],t['ndf_max'])
    endf_s=minimum_status(sm.get('endf_pct_dm',0),t.get('endf_min',0),.05,.15)
    rough=sm.get('roughage_pct_dm',0);conc=sm.get('concentrate_pct_dm',0)
    rough_s=range_status(rough,t['roughage_min'],t['roughage_max'])
    starch=sm.get('starch_pct_dm',0);starch_min=t.get('starch_min',0);starch_ideal=t.get('starch_ideal_max',100);starch_max=t.get('starch_max',100)
    starch_s='✅ İdeal bant' if starch_min<=starch<=starch_ideal else ('ℹ️ İdeal altı' if starch<starch_min else ('⚠️ Sınırda' if starch<=starch_ideal+.5 else ('⚠️ Dikkat bandı' if starch<=starch_max else '🔴 Yüksek')))
    risk_s={'Düşük':'✅ Düşük','Orta':'⚠️ Orta','Yüksek':'🔴 Yüksek'}[rumen_risk['level']]
    ca_s=mineral_status(sm['ca_g'],'ca');p_s=mineral_status(sm['p_g'],'p')
    cap_ratio=sm['ca_g']/sm['p_g'] if sm['p_g']>0 else 0.0
    ratio_s='✅ Dengeli' if 1.2<=cap_ratio<=3.0 else ('🔴 Düşük' if cap_ratio<1.0 else '⚠️ Kontrol')

    def kind(status):
        if '🔴' in status:return 'bad'
        if '⚠️' in status:return 'warn'
        if '✅' in status:return 'ok'
        return 'info'

    def row(key,label,target,current,status,note=''):
        return f"""<div class='science-target-row {kind(status)}' id='target-mini-{key}'>
          <div class='science-target-label'><b>{label}</b><small>{note}</small></div>
          <div class='science-target-values'><span><small>GEREKSİNİM</small><b id='target-mini-{key}-target'>{target}</b></span><span><small>RASYON</small><b id='target-mini-{key}-current'>{current}</b></span></div>
          <div class='science-target-state'><em id='target-mini-{key}-status'>{status}</em><i id='target-mini-{key}-diff'></i></div>
        </div>"""

    energy_card=f"""<section class='science-target-card'><h4>⚖️ Tüketim & Performans</h4>
      {row('dm','KM tüketimi',f'≈ {predicted_dmi:.2f} kg',f'{sm["dm_kg"]:.2f} kg',dm_s,'Diyet NEm yoğunluğuna göre dinamik')}
      {row('adg','Enerjiye göre GCAA kapasitesi',f'≈ {t["adg"]:.2f} kg',f'{capacity:.2f} kg' if capacity>0 else '—',adg_s,'Gerçekleşen GCAA tahmini değildir')}
      <span id='target-mini-me-current' hidden>{sm['me_mcal']:.1f}</span>
    </section>"""
    protein_card=f"""<section class='science-target-card'><h4>🌿 Protein & Lif</h4>
      {row('cp','Ham protein',f'≥ %{t["cp_pct"]:.1f}',f'%{sm["cp_pct_dm"]:.1f}',cp_s,f'MP gereksinimi {t["mp_req_g"]:.0f} g/gün; HP bir tabandır')}
      {row('ndf','NDF çalışma bandı',f'%{t["ndf_min"]:.0f}–{t["ndf_max"]:.0f}',f'%{sm["ndf_pct_dm"]:.1f}',ndf_s,'Faz bandı; tek başına güvenlik kararı değildir')}
      {row('endf','Etkili NDF',f'≥ %{t.get("endf_min",0):.1f}',f'%{sm.get("endf_pct_dm",0):.1f}',endf_s,'Fiziksel etkili lif göstergesi')}
    </section>"""
    rumen_card=f"""<section class='science-target-card'><h4>🧪 Rumen Dengesi</h4>
      {row('rc','Kaba / Kesif',f'Kaba %{t["roughage_min"]:.0f}–{t["roughage_max"]:.0f}',f'%{rough:.0f} / %{conc:.0f}',rough_s,'KM bazında')}
      {row('starch','Nişasta',f'İdeal %{starch_min:.0f}–{starch_ideal:.0f}',f'%{starch:.1f}',starch_s,f'{sm.get("starch_kg",0):.2f} kg/baş/gün')}
      {row('ph','Göreli asidoz riski','Düşük',rumen_risk['level'],risk_s,f'Veri güveni {rumen_risk["confidence"]}; klinik pH değildir')}
    </section>"""
    mineral_card=f"""<section class='science-target-card'><h4>🧂 Mineral & Ekonomi</h4>
      {row('ca','Kalsiyum',f'≥ {t["ca_g"]:.0f} g',f'{sm["ca_g"]:.0f} g',ca_s,'Minimum gereksinim')}
      {row('p','Fosfor',f'≥ {t["p_g"]:.0f} g',f'{sm["p_g"]:.0f} g',p_s,'Minimum gereksinim')}
      {row('cap','Ca:P oranı','1,2–3,0',f'{cap_ratio:.2f}',ratio_s,'Güvenlik penceresi')}
      {row('cost','Günlük maliyet','—',money(sm['cost']),'💰 Canlı','Besleme uygunluğundan sonra değerlendirilir')}
    </section>"""

    current_type=str(_rowval(rr,'animal_type','Besi Erkek') or 'Besi Erkek')
    animal_choices=[('Besi Erkek','Besi Erkek (Tosun / Boğa)'),('Düve','Düve'),('Kastre Erkek','Kastre Erkek'),('Genel Büyüyen Sığır','Genel Büyüyen Sığır')]
    animal_options=''.join(f"<option value='{h(value)}' {'selected' if current_type==value else ''}>{h(label)}</option>" for value,label in animal_choices)
    if current_type not in [value for value,_ in animal_choices]:
        animal_options=f"<option value='{h(current_type)}' selected>{h(current_type)}</option>"+animal_options
    form_fields=f"""<input type='hidden' name='ration_type' value='Besi'><label>Canlı Ağırlık (kg)<input type='number' min='150' max='900' step='1' name='target_weight_kg' value='{t['weight_kg']:.0f}'></label><label>Hedef Artış (kg/gün)<input type='number' min='0.2' max='2.2' step='0.05' name='target_adg_kg' value='{t['adg']:.2f}'></label><label>Yaş (ay, opsiyonel)<input type='number' min='0' max='36' step='1' name='target_age_months' value='{t.get('age_months',0):.0f}'></label><label>Besi Dönemi<select name='target_beef_phase'><option value='Otomatik' {'selected' if _rowval(rr,'target_beef_phase','Otomatik')=='Otomatik' else ''}>Otomatik</option><option {'selected' if _rowval(rr,'target_beef_phase','Otomatik')=='Besi Başlangıç' else ''}>Besi Başlangıç</option><option {'selected' if _rowval(rr,'target_beef_phase','Otomatik')=='Besi Geliştirme' else ''}>Besi Geliştirme</option><option {'selected' if _rowval(rr,'target_beef_phase','Otomatik')=='Besi Bitirme' else ''}>Besi Bitirme</option></select></label><label>Hayvan Profili<select name='animal_type'>{animal_options}</select></label>"""
    ctx=f"<div class='target-context compact'><b>{t['weight_kg']:.0f} kg</b> · hedef <b>{t['adg']:.2f} kg/gün</b> · <b>{h(t['animal_profile_label'])}</b> · {h(t['phase'])}</div>"

    details=f"""<details class='science-target-details'><summary>📋 Hesap ayrıntılarını göster</summary><div class='table-compact-wrap'><table class='ration-target-table compact-table zebra'><thead><tr><th>Değer</th><th>Gereksinim / bant</th><th>Rasyon arzı</th><th>Yorum</th></tr></thead><tbody>
      <tr><td>KM tüketimi</td><td>≈ {predicted_dmi:.2f} kg/gün</td><td>{sm['dm_kg']:.2f} kg</td><td>{dm_s}</td></tr>
      <tr><td>GCAA kapasitesi</td><td>≥ {t['adg']:.2f} kg/gün</td><td>{capacity:.2f} kg/gün</td><td>{adg_s}</td></tr>
      <tr><td>NEm / NEg</td><td>{t['nem_req_mcal']:.1f} / {t['neg_req_mcal']:.1f} Mcal</td><td>NEm yoğunluk {sm.get('nem_density',0):.2f} · büyümeye NEg {gain_supply:.1f} Mcal</td><td>Performans hesabının enerji temeli</td></tr>
      <tr><td>Ham protein</td><td>En az %{t['cp_pct']:.1f} KM · MP {t['mp_req_g']:.0f} g/gün</td><td>%{sm['cp_pct_dm']:.1f} KM</td><td>{cp_s}; RDP/RUP analizi MP doğruluğunu artırır</td></tr>
      <tr><td>ME</td><td>{t['me_mcal_day']:.1f} Mcal/gün referans</td><td>{sm['me_mcal']:.1f} Mcal</td><td>Bilgi amaçlı; fizibilite GCAA kapasitesiyle verilir</td></tr>
      <tr><td>NDF / eNDF</td><td>%{t['ndf_min']:.0f}–{t['ndf_max']:.0f} / en az %{t.get('endf_min',0):.1f}</td><td>%{sm['ndf_pct_dm']:.1f} / %{sm.get('endf_pct_dm',0):.1f}</td><td>{ndf_s} · {endf_s}</td></tr>
      <tr><td>Nişasta / risk</td><td>İdeal %{starch_min:.0f}–{starch_ideal:.0f}</td><td>%{starch:.1f} · {rumen_risk['level']}</td><td>{starch_s}; klinik pH tahmini değildir</td></tr>
      <tr><td>Ca / P</td><td>En az {t['ca_g']:.0f} / {t['p_g']:.0f} g</td><td>{sm['ca_g']:.0f} / {sm['p_g']:.0f} g</td><td>{ca_s} · {p_s} · oran {cap_ratio:.2f}</td></tr>
    </tbody></table></div></details>"""

    return f"""<style>
    body:has(.workbench-shell) .science-target-grid{{display:grid!important;grid-template-columns:repeat(4,minmax(0,1fr))!important;gap:8px!important;overflow:visible!important}}
    .science-target-card{{min-width:0;border:1px solid #cfe1d3;border-radius:10px;background:#fff;overflow:hidden}}
    .science-target-card h4{{margin:0;padding:7px 9px;background:#edf6f0;border-bottom:1px solid #d8e6dc;font-size:12px;color:#173f2b}}
    .science-target-row{{display:grid;grid-template-columns:minmax(72px,.8fr) minmax(128px,1.35fr);grid-template-areas:'label values' 'state state';gap:2px 7px;padding:5px 7px;border-bottom:1px solid #edf2ee;min-width:0;border-left:3px solid #94a69a}}
    .science-target-row:last-child{{border-bottom:0}}.science-target-row.ok{{border-left-color:#21a45b}}.science-target-row.warn{{border-left-color:#e6a11a;background:#fffaf0}}.science-target-row.bad{{border-left-color:#d6453d;background:#fff5f4}}
    .science-target-label{{grid-area:label;min-width:0}}.science-target-label>b{{display:block;font-size:11px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}}.science-target-label>small{{display:block;font-size:7.5px;line-height:1.1;color:#6b7b72;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;margin-top:1px}}
    .science-target-values{{grid-area:values;display:grid;grid-template-columns:1fr 1fr;gap:5px;min-width:0}}.science-target-values span{{min-width:0}}.science-target-values small{{display:block;font-size:7px;color:#668074;font-weight:800}}.science-target-values b{{display:block;font-size:12px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}}
    .science-target-state{{grid-area:state;display:flex;align-items:center;justify-content:space-between;gap:6px;min-width:0}}.science-target-state em{{font-style:normal;font-size:9px;font-weight:900;white-space:nowrap;min-width:0;overflow:hidden;text-overflow:ellipsis}}.science-target-state i{{font-style:normal;font-size:8px;color:#68766d;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;min-width:0}}
    .science-target-details{{margin-top:7px;font-size:11px}}.science-target-details>summary{{cursor:pointer;font-weight:800;color:#315d46}}
    @media(min-width:1001px) and (max-width:1450px){{.science-target-row{{grid-template-columns:minmax(78px,.9fr) minmax(100px,1.2fr);gap:2px 4px;padding:4px 5px}}.science-target-values{{gap:3px}}.science-target-values b{{font-size:11px}}.science-target-state i{{display:none}}}}
    @media(max-width:1000px){{body:has(.workbench-shell) .science-target-grid{{grid-template-columns:repeat(2,minmax(0,1fr))!important}}}}
    @media(max-width:650px){{body:has(.workbench-shell) .science-target-grid{{display:grid!important;grid-auto-flow:column!important;grid-template-columns:none!important;grid-auto-columns:minmax(280px,86vw)!important;overflow-x:auto!important;scroll-snap-type:x mandatory;padding-bottom:5px!important}}.science-target-card{{scroll-snap-align:start}}.science-target-row{{padding:6px 7px}}}}
    </style><div class='target-workspace'><div class='target-controlbar'><div class='target-head'><h3>🎯 Bilimsel Hedef Özeti</h3>{ctx}</div><form method='post' action='/ration/target' class='target-form'><input type='hidden' name='ration_id' value='{rr['id']}'>{form_fields}<button class='btn blue compact-target-btn'>Güncelle</button></form></div><div class='target-compare-sticky science-target-shell'><div class='target-compare-title'><b>Gereksinim ↔ Rasyon Arzı</b><span id='target-live-note'>NASEM hayvan profili · Dinamik KM · NEm/NEg performansı</span></div><div class='science-target-grid'>{energy_card}{protein_card}{rumen_card}{mineral_card}</div>{details}</div></div>"""

def ration_smart_recommendations(rr, sm, con=None, limit=6):
    """Katalogdaki yemleri mevcut besin açıklarını iyileştirme potansiyeline göre sıralar.
    Bu bir karar-destek simülasyonudur; reçete değildir."""
    own=con is None; c=con or db().__enter__()
    try:
        t=ration_targets_for_record(rr)
        predicted_dmi=_predicted_dmi_for_metrics(sm,t)
        gain_supply,_,_=_energy_balance(sm,t)
        deficits={
            'dm':max(0,predicted_dmi-sm['dm_kg'])/max(predicted_dmi,.01),
            'cp':max(0,t['cp_pct']-sm['cp_pct_dm'])/max(t['cp_pct'],.01),
            'energy':max(0,t.get('neg_req_mcal',0)-gain_supply)/max(t.get('neg_req_mcal',0),.01),
            'ca':max(0,t['ca_g']-sm['ca_g'])/max(t['ca_g'],.01),
            'p':max(0,t['p_g']-sm['p_g'])/max(t['p_g'],.01),
        }
        rows=c.execute("""select f.*,coalesce((select fp.price_per_kg from feed_prices fp where fp.feed_id=f.id and fp.effective_date<=? order by fp.effective_date desc,fp.id desc limit 1),0) price from feed_catalog f where f.active=1""",(date.today().isoformat(),)).fetchall()
        out=[]
        for f in rows:
            dm=float(f['dm_pct'] or 0)/100; cp=float(f['cp_pct'] or 0)/100; ndf=float(f['ndf_pct'] or 0)/100; neg=float(f['neg_mcal_kg'] or 0); ca=float(f['ca_pct'] or 0); ph=float(f['p_pct'] or 0)
            if dm<=0: continue
            # 1 kg yaş yem katkısının açık kapatma yönündeki göreli puanı
            protein_factor=(dm*cp*5 if is_real_protein_feed(f) else 0.0)
            score=(deficits['dm']*dm + deficits['cp']*protein_factor + deficits['energy']*dm*neg/1.5 + deficits['ca']*dm*ca/1.0 + deficits['p']*dm*ph/0.6)
            # NDF zaten yüksekse yüksek-NDF yemleri; protein zaten fazlaysa çok proteinli yemleri geri it
            if sm['ndf_pct_dm']>t['ndf_max'] and float(f['ndf_pct'] or 0)>40: score*=0.55
            if sm['cp_pct_dm']>max(t['cp_pct']*1.45,16.0) and float(f['cp_pct'] or 0)>18: score*=0.60
            if score<=0: continue
            price=float(f['price'] or 0); value=score/(1+(price/20 if price>0 else .15))
            reasons=[]
            if deficits['energy']>.03 and neg>=1.0: reasons.append('büyüme enerjisi')
            if deficits['cp']>.03 and is_real_protein_feed(f) and float(f['cp_pct'] or 0)>=14: reasons.append('protein')
            if deficits['dm']>.05: reasons.append('kuru madde')
            if deficits['ca']>.05 and ca>=.5: reasons.append('kalsiyum')
            if deficits['p']>.05 and ph>=.3: reasons.append('fosfor')
            out.append((value,f,', '.join(reasons) or 'denge'))
        out.sort(key=lambda x:x[0],reverse=True)
        return out[:limit]
    finally:
        if own:c.close()

def ration_balance_error(targets, sm):
    """Rasyonu hedefe uzaklığına göre puanlar. Daha düşük daha iyidir.
    Bu skor reçete değildir; öneri sıralamasında karar desteği için kullanılır."""
    def rel(actual, target):
        return abs(float(actual)-float(target))/max(abs(float(target)), 0.01)
    def deficit(actual, minimum):
        return max(0.0,float(minimum)-float(actual))/max(abs(float(minimum)),0.01)
    score=0.0
    predicted_dmi=_predicted_dmi_for_metrics(sm,targets)
    sm['predicted_dmi_kg']=predicted_dmi
    score += rel(sm.get('dm_kg',0), predicted_dmi)*1.10
    # Chapter 20 HP, Ca ve P değerleri minimum gereksinimdir. Hedefi makul
    # ölçüde aşmak, hedefe eşit olmamaktan doğan bir formülasyon hatası değildir.
    score += deficit(sm.get('cp_pct_dm',0), targets['cp_pct'])*1.20
    if targets.get('mode')=='Besi' and sm.get('nem_density',0)>0 and sm.get('neg_density',0)>0:
        score += deficit(_achievable_adg(sm,targets),targets.get('adg',0))*1.45
    else:
        score += deficit(sm.get('me_mcal',0),targets.get('me_mcal_day',0))*1.45
    score += deficit(sm.get('ca_g',0), targets['ca_g'])*0.65
    score += deficit(sm.get('p_g',0), targets['p_g'])*0.65
    mw=_mineral_windows(targets,predicted_dmi)
    if float(sm.get('ca_g',0) or 0)>mw['ca_soft']:
        score += (float(sm.get('ca_g',0))-mw['ca_soft'])/max(mw['ca_soft'],1)*.25
    if float(sm.get('p_g',0) or 0)>mw['p_soft']:
        score += (float(sm.get('p_g',0))-mw['p_soft'])/max(mw['p_soft'],1)*.25
    cp_high=max(float(targets.get('cp_pct',0))*1.45,16.0)
    if float(sm.get('cp_pct_dm',0) or 0)>cp_high:
        score += (float(sm.get('cp_pct_dm',0))-cp_high)/max(cp_high,1)*.15
    ndf=float(sm.get('ndf_pct_dm',0) or 0)
    if ndf < targets['ndf_min']:
        score += (targets['ndf_min']-ndf)/max(targets['ndf_min'],1)*1.10
    elif ndf > targets['ndf_max']:
        score += (ndf-targets['ndf_max'])/max(targets['ndf_max'],1)*1.10
    rough=float(sm.get('roughage_pct_dm',0) or 0)
    if rough < targets.get('roughage_min',0):
        score += (targets['roughage_min']-rough)/max(targets['roughage_min'],1)*0.90
    elif rough > targets.get('roughage_max',100):
        score += (rough-targets['roughage_max'])/max(targets['roughage_max'],1)*0.90
    starch=float(sm.get('starch_pct_dm',0) or 0)
    if starch > targets.get('starch_max',100):
        score += (starch-targets['starch_max'])/max(targets['starch_max'],1)*1.35
    elif starch > targets.get('starch_ideal_max',100):
        score += (starch-targets['starch_ideal_max'])/max(targets['starch_ideal_max'],1)*0.45
    return score

def ration_reduction_recommendations(rr, sm, con=None, limit=6):
    """Mevcut rasyondaki fazlalıkları azaltmaya yardımcı olabilecek yem/miktarları sıralar."""
    own=con is None; c=con or db().__enter__()
    try:
        t=ration_targets_for_record(rr)
        base_dmi=_predicted_dmi_for_metrics(sm,t); mw=_mineral_windows(t,base_dmi)
        base_err=ration_balance_error(t,sm); out=[]
        for it in sm['items']:
            current=float(it['kg_per_head_day'] or 0)
            if current <= 0.05: continue
            best=None
            for step in (0.25,0.50,1.00):
                delta=-min(step,current)
                if current+delta < -0.001: continue
                ss=ration_simulated_summary(rr['id'],it['id'] if False else it['id'],delta,c)
                err=ration_balance_error(t,ss); improve=base_err-err
                if improve <= 0.0005: continue
                reasons=[]
                if sm['cp_pct_dm']>max(t['cp_pct']*1.45,16.0) and float(it['cp_pct'] or 0)>=t['cp_pct']: reasons.append('çok yüksek proteini düşürür')
                if sm['ca_g']>mw['ca_soft'] and float(it['ca_pct'] or 0)>=0.35: reasons.append('yüksek kalsiyumu düşürür')
                if sm['p_g']>mw['p_soft'] and float(it['p_pct'] or 0)>=0.25: reasons.append('yüksek fosforu düşürür')
                if sm['ndf_pct_dm']>t['ndf_max'] and float(it['ndf_pct'] or 0)>t['ndf_max']: reasons.append('NDF yükünü azaltır')
                # Bir değeri düzeltirken enerji/KM açığını büyütüyorsa cezalandır.
                warnings=[]
                if _achievable_adg(ss,t) < t.get('adg',0)*.90: warnings.append('GCAA kapasitesi düşebilir')
                if ss['dm_kg'] < _predicted_dmi_for_metrics(ss,t)*0.90: warnings.append('KM düşebilir')
                candidate=(improve, it, delta, ', '.join(reasons) or 'genel dengeyi iyileştirir', ', '.join(warnings), ss)
                if best is None or candidate[0]>best[0]: best=candidate
            if best: out.append(best)
        out.sort(key=lambda x:x[0],reverse=True)
        return out[:limit]
    finally:
        if own:c.close()

def ration_addition_recommendations(rr, sm, con=None, limit=30):
    """Tüm katalogdaki yemleri +0.50 kg simülasyonuyla değerlendirir ve hedefe yaklaştıranları sıralar."""
    own=con is None; c=con or db().__enter__()
    try:
        t=ration_targets_for_record(rr)
        base_dmi=_predicted_dmi_for_metrics(sm,t); mw=_mineral_windows(t,base_dmi)
        base_adg=_achievable_adg(sm,t)
        base_err=ration_balance_error(t,sm)
        rows=c.execute("""select f.*,coalesce((select fp.price_per_kg from feed_prices fp where fp.feed_id=f.id and fp.effective_date<=? order by fp.effective_date desc,fp.id desc limit 1),0) price,
            coalesce((select sum(case when st.tx_type in ('Giriş','Sayım +') then st.quantity_kg when st.tx_type in ('Çıkış','Tüketim','Sayım -') then -st.quantity_kg else 0 end) from feed_stock_transactions st where st.feed_id=f.id),0) stock
            from feed_catalog f where f.active=1""",(date.today().isoformat(),)).fetchall()
        out=[]
        for f in rows:
            if float(f['dm_pct'] or 0)<=0: continue
            ss=ration_simulated_summary(rr['id'],f['id'],0.50,c)
            err=ration_balance_error(t,ss); improve=base_err-err
            # HP açığı varsa gerçek protein hammaddelerini ayrıca hedefe göre puanla.
            # Böylece enerji zaten yüksek olsa bile soya/kanola/ayçiçeği gibi kaynaklar
            # çözüm adayı olarak görülebilir; NPN/mineral ürünleri protein diye önerilmez.
            targeted=0.0
            cp_gap=max(0.0,t['cp_pct']-sm['cp_pct_dm'])
            if cp_gap>0.15 and is_real_protein_feed(f):
                cp_gain=max(0.0,ss['cp_pct_dm']-sm['cp_pct_dm'])
                targeted += min(cp_gain/max(cp_gap,0.01),1.5)*0.90
            rank_score=improve+targeted
            if rank_score<=0.0005: continue
            reasons=[]
            if base_adg<t.get('adg',0)*.98 and _achievable_adg(ss,t)>base_adg: reasons.append('GCAA kapasitesini artırır')
            if sm['cp_pct_dm']<t['cp_pct']*0.98 and is_real_protein_feed(f) and ss['cp_pct_dm']>sm['cp_pct_dm']+0.03: reasons.append('protein açığını azaltır')
            if sm['dm_kg']<base_dmi*.95 and ss['dm_kg']>sm['dm_kg']: reasons.append('KM desteği')
            if sm['ca_g']<t['ca_g']*0.95 and ss['ca_g']>sm['ca_g']: reasons.append('Ca desteği')
            if sm['p_g']<t['p_g']*0.95 and ss['p_g']>sm['p_g']: reasons.append('P desteği')
            warnings=[]
            if sm['cp_pct_dm']>max(t['cp_pct']*1.45,16.0) and ss['cp_pct_dm']>sm['cp_pct_dm']+0.05: warnings.append('çok yüksek proteini artırır')
            if sm['ca_g']>mw['ca_soft'] and ss['ca_g']>sm['ca_g']+1: warnings.append('yüksek Ca arzını artırır')
            if sm['p_g']>mw['p_soft'] and ss['p_g']>sm['p_g']+1: warnings.append('yüksek P arzını artırır')
            out.append((rank_score,f,', '.join(reasons) or 'genel dengeyi iyileştirir',', '.join(warnings),ss))
        # Önce denge katkısı; aynı katkıda fiyatı bilinen ve daha ucuz olan öne gelsin.
        out.sort(key=lambda x:(x[0], 1 if float(x[1]['price'] or 0)>0 else 0, -float(x[1]['price'] or 999999)),reverse=True)
        return out[:limit]
    finally:
        if own:c.close()

def ration_simulated_multi_summary(ration_id, changes, con=None):
    """Birden fazla yem değişikliğini kaydetmeden aynı rasyon üzerinde birlikte simüle eder."""
    own=con is None; c=con or db().__enter__()
    try:
        base=ration_summary(ration_id,c)
        out={k:v for k,v in base.items() if k!='items'}; out['items']=base['items']
        for feed_id,delta_kg in changes:
            f=c.execute("""select f.*,coalesce((select fp.price_per_kg from feed_prices fp where fp.feed_id=f.id and fp.effective_date<=? order by fp.effective_date desc,fp.id desc limit 1),0) price from feed_catalog f where f.id=?""",(date.today().isoformat(),int(feed_id))).fetchone()
            if not f: continue
            kg=float(delta_kg or 0); dm=kg*float(f['dm_pct'] or 0)/100
            feed_starch=dm*_solver_starch_pct(f)/100
            starch_deg,known_deg=_feed_starch_degradability(f)
            out['as_fed_kg']+=kg;out['dm_kg']+=dm;out['cp_kg']+=dm*float(f['cp_pct'] or 0)/100;out['ndf_kg']+=dm*float(f['ndf_pct'] or 0)/100;out['endf_kg']+=dm*float(f['ndf_pct'] or 0)/100*float(f['effective_ndf_pct'] or 0)/100;out['starch_kg']+=feed_starch;out['tdn_kg']+=dm*float(f['tdn_pct'] or 0)/100;out['me_mcal']+=dm*float(f['me_mcal_kg'] or 0);out['nem_mcal']+=dm*float(f['nem_mcal_kg'] or 0);out['neg_mcal']+=dm*float(f['neg_mcal_kg'] or 0);out['ca_g']+=dm*float(f['ca_pct'] or 0)*10;out['p_g']+=dm*float(f['p_pct'] or 0)*10;out['cost']+=kg*float(f['price'] or 0)
            if known_deg:
                out['known_degradability_starch_kg']+=feed_starch
                out['rapid_starch_kg']+=feed_starch*starch_deg/100
            grp=feed_group(f)
            if grp=='Kaba': out['roughage_dm_kg']=out.get('roughage_dm_kg',0)+dm
            elif grp=='Kesif': out['concentrate_dm_kg']=out.get('concentrate_dm_kg',0)+dm
            else: out['additive_dm_kg']=out.get('additive_dm_kg',0)+dm
        for key in ('as_fed_kg','dm_kg','cp_kg','ndf_kg','endf_kg','starch_kg','rapid_starch_kg','known_degradability_starch_kg','me_mcal','nem_mcal','neg_mcal','ca_g','p_g','cost','roughage_dm_kg','concentrate_dm_kg'):
            if key in out: out[key]=max(0.0,float(out[key] or 0))
        out['cp_pct_dm']=out['cp_kg']/out['dm_kg']*100 if out['dm_kg'] else 0;out['ndf_pct_dm']=out['ndf_kg']/out['dm_kg']*100 if out['dm_kg'] else 0;out['endf_pct_dm']=out['endf_kg']/out['dm_kg']*100 if out['dm_kg'] else 0;out['starch_pct_dm']=out['starch_kg']/out['dm_kg']*100 if out['dm_kg'] else 0;out['me_per_kg_dm']=out['me_mcal']/out['dm_kg'] if out['dm_kg'] else 0
        out['rapid_starch_pct_dm']=out['rapid_starch_kg']/out['dm_kg']*100 if out['dm_kg'] else 0;out['starch_degradability_coverage']=out['known_degradability_starch_kg']/out['starch_kg'] if out['starch_kg'] else 1.0
        out['nem_density']=out['nem_mcal']/out['dm_kg'] if out['dm_kg'] else 0;out['neg_density']=out['neg_mcal']/out['dm_kg'] if out['dm_kg'] else 0
        rc=out.get('roughage_dm_kg',0)+out.get('concentrate_dm_kg',0); out['roughage_pct_dm']=out.get('roughage_dm_kg',0)/rc*100 if rc else 0; out['concentrate_pct_dm']=out.get('concentrate_dm_kg',0)/rc*100 if rc else 0
        return out
    finally:
        if own:c.close()

def ration_effect_text(targets,before,after,max_items=3):
    """Öneriyi ölçülebilir ve kısa bir dille açıklar."""
    impacts=[]
    before_dmi=_predicted_dmi_for_metrics(before,targets);after_dmi=_predicted_dmi_for_metrics(after,targets)
    dm_gain=abs(before['dm_kg']-before_dmi)/max(before_dmi,.01)-abs(after['dm_kg']-after_dmi)/max(after_dmi,.01)
    if dm_gain>0.01:impacts.append((dm_gain,f"KM {before['dm_kg']:.2f} → {after['dm_kg']:.2f} kg"))
    minimums=[('HP',before['cp_pct_dm'],after['cp_pct_dm'],targets['cp_pct'],'%{:.1f}'),('Ca',before['ca_g'],after['ca_g'],targets['ca_g'],'{:.0f} g'),('P',before['p_g'],after['p_g'],targets['p_g'],'{:.0f} g')]
    for name,b,a,t,fmt in minimums:
        if t<=0:continue
        gain=max(0,t-b)/t-max(0,t-a)/t
        if gain>0.01:impacts.append((gain,f"{name} {fmt.format(b)} → {fmt.format(a)}"))
    if targets.get('mode')=='Besi':
        b_adg=_achievable_adg(before,targets);a_adg=_achievable_adg(after,targets);goal=max(targets.get('adg',0),.01)
        gain=max(0,goal-b_adg)/goal-max(0,goal-a_adg)/goal
        if gain>0.01:impacts.append((gain,f"GCAA kapasitesi {b_adg:.2f} → {a_adg:.2f} kg"))
    def add_range(name,b,a,lo,hi,fmt):
        def dist(x): return (lo-x)/max(lo,1) if x<lo else (x-hi)/max(hi,1) if x>hi else 0
        gain=dist(b)-dist(a)
        if gain>0.01: impacts.append((gain,f"{name} {fmt.format(b)} → {fmt.format(a)}"))
    add_range('NDF',before['ndf_pct_dm'],after['ndf_pct_dm'],targets['ndf_min'],targets['ndf_max'],'%{:.1f}')
    add_range('Nişasta',before.get('starch_pct_dm',0),after.get('starch_pct_dm',0),targets.get('starch_min',0),targets.get('starch_ideal_max',100),'%{:.1f}')
    add_range('Kaba',before.get('roughage_pct_dm',0),after.get('roughage_pct_dm',0),targets.get('roughage_min',0),targets.get('roughage_max',100),'%{:.0f}')
    impacts.sort(key=lambda x:x[0],reverse=True)
    return ' · '.join(x[1] for x in impacts[:max_items]) or 'Genel denge hedefe yaklaşıyor'

def ration_combined_recommendations(rr, sm, reductions, additions, limit=5):
    """Azalt + ekle çiftlerini gerçek birlikte simülasyonla puanlar."""
    t=ration_targets_for_record(rr); base_err=ration_balance_error(t,sm); out=[]
    for red in reductions[:5]:
        for add in additions[:12]:
            if int(red[1]['id'])==int(add[1]['id']): continue
            ss=ration_simulated_multi_summary(rr['id'],[(red[1]['id'],red[2]),(add[1]['id'],0.50)])
            improve=base_err-ration_balance_error(t,ss)
            if improve<=0.0005: continue
            effect=ration_effect_text(t,sm,ss,3)
            out.append((improve,red,add,ss,effect))
    out.sort(key=lambda x:(x[0],-float(x[3].get('cost',999999))),reverse=True)
    return out[:limit]

def ration_simulated_summary(ration_id, feed_id, delta_kg, con=None):
    own=con is None; c=con or db().__enter__()
    try:
        base=ration_summary(ration_id,c); out={k:v for k,v in base.items() if k!='items'}; out['items']=base['items']
        f=c.execute("""select f.*,coalesce((select fp.price_per_kg from feed_prices fp where fp.feed_id=f.id and fp.effective_date<=? order by fp.effective_date desc,fp.id desc limit 1),0) price from feed_catalog f where f.id=?""",(date.today().isoformat(),feed_id)).fetchone()
        if not f:return base
        kg=float(delta_kg); dm=kg*float(f['dm_pct'] or 0)/100
        feed_starch=dm*_solver_starch_pct(f)/100
        starch_deg,known_deg=_feed_starch_degradability(f)
        out['as_fed_kg']+=kg;out['dm_kg']+=dm;out['cp_kg']+=dm*float(f['cp_pct'] or 0)/100;out['ndf_kg']+=dm*float(f['ndf_pct'] or 0)/100;out['endf_kg']+=dm*float(f['ndf_pct'] or 0)/100*float(f['effective_ndf_pct'] or 0)/100;out['starch_kg']+=feed_starch;out['tdn_kg']+=dm*float(f['tdn_pct'] or 0)/100;out['me_mcal']+=dm*float(f['me_mcal_kg'] or 0);out['nem_mcal']+=dm*float(f['nem_mcal_kg'] or 0);out['neg_mcal']+=dm*float(f['neg_mcal_kg'] or 0);out['ca_g']+=dm*float(f['ca_pct'] or 0)*10;out['p_g']+=dm*float(f['p_pct'] or 0)*10;out['cost']+=kg*float(f['price'] or 0)
        if known_deg:
            out['known_degradability_starch_kg']+=feed_starch
            out['rapid_starch_kg']+=feed_starch*starch_deg/100
        grp=feed_group(f)
        if grp=='Kaba': out['roughage_dm_kg']=out.get('roughage_dm_kg',0)+dm
        elif grp=='Kesif': out['concentrate_dm_kg']=out.get('concentrate_dm_kg',0)+dm
        else: out['additive_dm_kg']=out.get('additive_dm_kg',0)+dm
        for key in ('as_fed_kg','dm_kg','cp_kg','ndf_kg','endf_kg','starch_kg','rapid_starch_kg','known_degradability_starch_kg','me_mcal','nem_mcal','neg_mcal','ca_g','p_g','cost','roughage_dm_kg','concentrate_dm_kg'):
            if key in out:out[key]=max(0.0,float(out[key] or 0))
        out['cp_pct_dm']=out['cp_kg']/out['dm_kg']*100 if out['dm_kg'] else 0;out['ndf_pct_dm']=out['ndf_kg']/out['dm_kg']*100 if out['dm_kg'] else 0;out['endf_pct_dm']=out['endf_kg']/out['dm_kg']*100 if out['dm_kg'] else 0;out['starch_pct_dm']=out['starch_kg']/out['dm_kg']*100 if out['dm_kg'] else 0;out['me_per_kg_dm']=out['me_mcal']/out['dm_kg'] if out['dm_kg'] else 0
        out['rapid_starch_pct_dm']=out['rapid_starch_kg']/out['dm_kg']*100 if out['dm_kg'] else 0;out['starch_degradability_coverage']=out['known_degradability_starch_kg']/out['starch_kg'] if out['starch_kg'] else 1.0
        out['nem_density']=out['nem_mcal']/out['dm_kg'] if out['dm_kg'] else 0;out['neg_density']=out['neg_mcal']/out['dm_kg'] if out['dm_kg'] else 0
        rc=out.get('roughage_dm_kg',0)+out.get('concentrate_dm_kg',0); out['roughage_pct_dm']=out.get('roughage_dm_kg',0)/rc*100 if rc else 0; out['concentrate_pct_dm']=out.get('concentrate_dm_kg',0)/rc*100 if rc else 0
        return out
    finally:
        if own:c.close()

def paddock_population(paddock_id, con=None):
    own=con is None
    c=con or db().__enter__()
    try:
        adults=c.execute("select count(*) n from animals where paddock_id=? and coalesce(status,'Aktif')='Aktif'",(paddock_id,)).fetchone()['n']
        calves=c.execute("select count(*) n from calves where paddock_id=? and promoted_animal_id is null",(paddock_id,)).fetchone()['n']
        return int(adults or 0)+int(calves or 0)
    finally:
        if own:c.close()

def sync_paddock_text(c, source, animal_id, paddock_id):
    name=''
    if paddock_id:
        r=c.execute('select name from paddocks where id=?',(paddock_id,)).fetchone(); name=r['name'] if r else ''
    table='animals' if source=='animal' else 'calves'
    c.execute(f'update {table} set paddock_id=?,paddock=? where id=?',(paddock_id or None,name,animal_id))

def male_weight_performance(animal_id, con=None):
    own=con is None
    c=con or db()
    try:
        rows=c.execute('select measure_date,weight,notes from weights where animal_id=? order by measure_date,id',(animal_id,)).fetchall()
        if len(rows)<2:
            return {'rows':rows,'gain':None,'days':0,'daily':None,'monthly':None,'status':'none'}
        prev,last=rows[-2],rows[-1]
        try: days=(date.fromisoformat(last['measure_date'])-date.fromisoformat(prev['measure_date'])).days
        except Exception: days=0
        gain=float(last['weight'])-float(prev['weight'])
        daily=gain/days if days>0 else None
        monthly=daily*30 if daily is not None else None
        target=setting_float('male_min_daily_gain',1.0)
        warn_ratio=setting_float('male_warning_ratio',0.90)
        if daily is None: status='none'
        elif daily < target*warn_ratio: status='low'
        elif daily < target: status='watch'
        else: status='good'
        return {'rows':rows,'previous':prev,'latest':last,'gain':gain,'days':days,'daily':daily,'monthly':monthly,'target':target,'status':status}
    finally:
        if own:c.close()

def weight_chart_svg(rows):
    if len(rows)<2:return '<p class="mut">Grafik için en az iki tartım kaydı gerekir.</p>'
    data=[]
    for r in rows:
        try:data.append((date.fromisoformat(r['measure_date']),float(r['weight'])))
        except Exception:pass
    if len(data)<2:return '<p class="mut">Geçerli tarih ve kilo verisi yetersiz.</p>'
    w,hgt,pad=720,240,42
    minv=min(v for _,v in data);maxv=max(v for _,v in data)
    if maxv==minv:maxv=minv+1
    span=max(1,(data[-1][0]-data[0][0]).days)
    pts=[]
    for d,v in data:
        x=pad+((d-data[0][0]).days/span)*(w-2*pad)
        y=hgt-pad-((v-minv)/(maxv-minv))*(hgt-2*pad)
        pts.append((x,y,d,v))
    poly=' '.join(f'{x:.1f},{y:.1f}' for x,y,_,_ in pts)
    grid=''.join(f'<line class="gridline" x1="{pad}" y1="{pad+i*(hgt-2*pad)/4:.1f}" x2="{w-pad}" y2="{pad+i*(hgt-2*pad)/4:.1f}"/>' for i in range(5))
    dots=''.join(f'<circle class="point" cx="{x:.1f}" cy="{y:.1f}" r="6"><title>{d.isoformat()} · {v:.1f} kg</title></circle>' for x,y,d,v in pts)
    labels=f'<text x="{pad}" y="{hgt-10}">{data[0][0].strftime("%d.%m.%Y")}</text><text x="{w-pad}" y="{hgt-10}" text-anchor="end">{data[-1][0].strftime("%d.%m.%Y")}</text><text x="8" y="{pad+4}">{maxv:.1f} kg</text><text x="8" y="{hgt-pad+4}">{minv:.1f} kg</text>'
    return f'<svg class="weight-chart" viewBox="0 0 {w} {hgt}" role="img" aria-label="Kilo gelişim grafiği">{grid}<line class="axis" x1="{pad}" y1="{hgt-pad}" x2="{w-pad}" y2="{hgt-pad}"/><polyline class="trend" points="{poly}"/>{dots}{labels}</svg>'

def promote_mature_calves():
    """10 ayını dolduran buzağıları yetişkin karta geçirir; geçmiş kayıtları korur."""
    with db() as c:
        rows=c.execute("select * from calves where promoted_animal_id is null and birth_date is not null and birth_date<>''").fetchall()
        for calf in rows:
            if months_old(calf['birth_date']) < 10: continue
            existing=c.execute('select id from animals where tag=?',(calf['tag'],)).fetchone()
            if existing: aid=existing['id']
            else:
                cur=c.execute('insert into animals(tag,nickname,gender,breed,birth_date,notes,paddock,photo_url,sold_price,status,purchase_date,purchase_price) values(?,?,?,?,?,?,?,?,?,?,?,?)',
                    (calf['tag'],calf['nickname'] or '',calf['gender'] or '',calf['breed'] or '',calf['birth_date'],calf['notes'] or '',calf['paddock'] or '',calf['photo_url'] or '',0,'Aktif',calf['purchase_date'] or '',float(calf['purchase_price'] or 0)))
                aid=cur.lastrowid
            c.execute('update health set animal_id=?,calf_id=null where calf_id=?',(aid,calf['id']))
            for w in c.execute('select * from calf_weights where calf_id=? order by measure_date,id',(calf['id'],)).fetchall():
                if not c.execute('select 1 from weights where animal_id=? and measure_date=?',(aid,w['measure_date'])).fetchone():
                    c.execute('insert into weights(animal_id,measure_date,weight,notes) values(?,?,?,?)',(aid,w['measure_date'],w['weight'],w['notes']))
            for ph in c.execute('select * from calf_photos where calf_id=? order by id',(calf['id'],)).fetchall():
                if not c.execute('select 1 from animal_photos where animal_id=? and filename=?',(aid,ph['filename'])).fetchone():
                    c.execute('insert into animal_photos(animal_id,filename,created_at,caption) values(?,?,?,?)',(aid,ph['filename'],ph['created_at'],ph['caption']))
            c.execute('update calves set promoted_animal_id=?,promoted_at=? where id=?',(aid,datetime.now().isoformat(timespec='seconds'),calf['id']))


NAV=[('Dashboard','/'),('📈 Besi Performansı','/performance'),('➕ Hayvan Ekle','/animal-add'),('Dişi Hayvanlar','/animals'),('Erkek Hayvanlar','/males'),('Satılan Hayvanlar','/archive/sold'),('Kesilen Hayvanlar','/archive/slaughtered'),('Buzağılar','/calves'),('Kızgınlık Takibi','/estrus'),('Tohumlama','/inseminations'),('Sağlık','/health'),('Finans','/finance'),('Raporlar','/reports'),('Veri Aktarımı','/data'),('💾 Yedekleme Merkezi','/backups'),('🔐 Şifremi Değiştir','/password-change')]
ADMIN_NAV=[('👥 Kullanıcı Yönetimi','/users'),('📜 İşlem Günlüğü','/audit-log')]


# 6.17 Desktop ERP visual shell only. Backend / routes / solver / network unchanged.
DESKTOP_ERP_CSS = r'''
:root{--erp-side:198px;--erp-top:34px;--erp-cmd:58px;--erp-tabs:36px;--erp-status:28px}
html,body{height:100%}body{overflow-x:hidden;background:#f3f5f4}
.top{height:var(--erp-top)!important;left:0!important;background:#fbfcfb!important;color:#16231b!important;border-bottom:1px solid #d9dfdb!important;padding:0 12px!important;box-shadow:none!important;z-index:60!important}
.top .brand{display:none!important}
.top-user{font-size:12px!important;color:#17251d!important}.top .ver{display:none!important}
.erp-commandbar{position:fixed;left:var(--erp-side);right:0;top:var(--erp-top);height:var(--erp-cmd);z-index:55;background:#fff;border-bottom:1px solid #d9dfdb;display:flex;align-items:stretch;padding:0 14px}
.erp-commandbar a{min-width:78px;padding:6px 13px;display:flex;flex-direction:column;align-items:center;justify-content:center;gap:2px;border-right:1px solid #e5e9e6;font-size:11px;font-weight:600;color:#17251d}.erp-commandbar a:hover{background:#f0f6f2}.erp-commandbar .ico{font-size:17px;line-height:18px}
.erp-tabs{position:fixed;left:var(--erp-side);right:0;top:calc(var(--erp-top) + var(--erp-cmd));height:var(--erp-tabs);z-index:54;background:#f5f7f6;border-bottom:1px solid #d7ded9;display:flex;align-items:end;padding:0 10px}.erp-tab{height:31px;min-width:190px;max-width:280px;background:#fff;border:1px solid #d7ded9;border-bottom:2px solid #087443;border-radius:6px 6px 0 0;padding:7px 13px;font-size:12px;font-weight:700;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.layout{padding-top:calc(var(--erp-top) + var(--erp-cmd) + var(--erp-tabs))!important;min-height:100vh!important}
.side{top:var(--erp-top)!important;bottom:var(--erp-status)!important;width:var(--erp-side)!important;background:linear-gradient(180deg,#08683f 0%,#075c38 100%)!important;color:#fff!important;padding:10px 8px 44px!important;border-right:1px solid #075232!important;overflow-y:auto!important}.side:before{display:none!important}.side a,.side summary{color:#fff!important}.side>a,.nav-group summary{padding:10px 11px!important;margin:2px 0!important;border-radius:5px!important;font-size:13px!important}.side a:hover,.side a.on,.nav-group.open-group summary{background:rgba(255,255,255,.13)!important}.nav-children{border-left:0!important;margin-left:9px!important;padding-left:8px!important}.side .nav-children a{padding:7px 10px!important;font-size:12px!important}.nav-group summary:after{font-size:16px!important}
.main{margin-left:var(--erp-side)!important;padding:10px 12px calc(var(--erp-status) + 12px)!important;min-width:0!important;width:calc(100% - var(--erp-side))!important;min-height:calc(100vh - var(--erp-top) - var(--erp-cmd) - var(--erp-tabs))!important;overflow-x:hidden!important}
.card{border-radius:7px!important;border:1px solid #dce2de!important;box-shadow:0 1px 3px rgba(21,48,32,.05)!important;padding:12px!important}.hero{border-radius:8px!important;padding:16px 18px!important;margin-bottom:10px!important;min-height:92px!important}.grid,.dashboard-grid{gap:9px!important}table{border-radius:5px!important}th,td{padding:7px 8px!important;font-size:12px!important}.btn{border-radius:5px!important;padding:8px 12px!important;font-size:12px!important}.form input,.form select,.form textarea{border-radius:5px!important;padding:8px!important}
.erp-statusbar{position:fixed;left:0;right:0;bottom:0;height:var(--erp-status);z-index:70;background:#fff;border-top:1px solid #d7ded9;display:flex;align-items:center;gap:42px;padding:0 15px;font-size:11px;color:#25372d}.erp-statusbar span:before{content:'●';color:#26a269;margin-right:7px}.erp-statusbar .erp-version{margin-left:auto}.erp-statusbar .erp-version:before{display:none}
.workbench-page-head{margin:0 0 6px!important}.workbench-page-head h1,.ration-page-title{font-size:19px!important;margin:3px 0!important}.target-controlbar{border-radius:6px!important;padding:7px 9px!important}.target-compare-sticky{border-radius:6px!important;margin-top:5px!important;padding:5px!important}.nutri-mini-grid{gap:5px!important;padding:5px!important}.nutri-mini.nutri-compare-card{border-radius:5px!important;min-height:100px!important;height:auto!important}.nutri-card-title{font-size:12px!important;padding:5px 7px!important}.nutri-side b{font-size:15px!important}.nutri-card-footer{min-height:29px!important;padding:4px 6px!important}.ration-workbench-table th,.ration-workbench-table td{padding:5px 6px!important}.ration-stepper input{height:28px!important}.ration-stepper button{min-height:28px!important}
@media(max-width:900px){:root{--erp-side:0px}.erp-commandbar,.erp-tabs{left:0}.side{width:220px!important;transform:translateX(-105%);z-index:80!important;padding-top:72px!important}.side.mobile-open{transform:translateX(0)}.main{margin-left:0!important;width:100%!important}.menu-toggle{display:inline-block!important}.erp-commandbar a{min-width:58px;padding:4px 7px}.erp-commandbar{overflow-x:auto}.erp-statusbar{gap:12px;overflow:hidden}}
'''

# 6.17 FINAL Desktop ERP visual language for all modules.
ERP_ALL_MODULES_FINAL_CSS = r'''
/* ---------- Global ERP workspace ---------- */
body{font-size:13px;color:#14231b;background:#f3f5f4}
.main>h1,.main>h2,.main>h3{letter-spacing:-.25px}
.main>h1:first-child,.main>h2:first-child{margin-top:2px!important}
.main .mut{color:#69786f!important}
.main hr{border:0;border-top:1px solid #dce4df;margin:10px 0}
.main .card,.main .panel,.main .section,.main fieldset{background:#fff;border:1px solid #dce4df;border-radius:7px;box-shadow:0 1px 3px rgba(18,52,33,.04)}
.main fieldset{padding:10px 12px}
.main fieldset legend{font-weight:800;color:#153c28;padding:0 5px}
.main details:not(.nav-group):not(.quick-feed-card){background:#fff;border:1px solid #dce4df;border-radius:6px;margin:6px 0;overflow:hidden}
.main details:not(.nav-group):not(.quick-feed-card)>summary{padding:8px 10px;background:#f8faf9;font-weight:800;color:#213b2d;border-bottom:1px solid transparent}
.main details:not(.nav-group):not(.quick-feed-card)[open]>summary{border-bottom-color:#e2e9e4}
.main input:not([type=checkbox]):not([type=radio]),.main select,.main textarea{border:1px solid #bccbc1!important;border-radius:4px!important;background:#fff!important;box-shadow:inset 0 1px 1px rgba(13,47,27,.03)!important;min-height:31px}
.main input:focus,.main select:focus,.main textarea:focus{outline:2px solid rgba(15,130,73,.14)!important;border-color:#1b8a50!important}
.main label{color:#294537}
.main .btn,.main button:not(.ration-stepper button){border-radius:4px!important;box-shadow:none!important;font-weight:700}
.main .btn:hover,.main button:hover{filter:brightness(.985)}
.main table{width:100%;border-collapse:separate!important;border-spacing:0!important;background:#fff;border:1px solid #dce4df!important;border-radius:5px!important;overflow:hidden;box-shadow:none!important}
.main thead th,.main table th{background:#edf3ef!important;color:#213b2d!important;border-bottom:1px solid #d4ded7!important;font-size:11.5px!important;text-transform:none!important;font-weight:800!important;position:sticky;top:0;z-index:2}
.main tbody td{border-bottom:1px solid #e6ebe8!important}
.main tbody tr:nth-child(even){background:#f8faf9}
.main tbody tr:hover{background:#edf7f0!important}
.main tbody tr:last-child td{border-bottom:0!important}
.main .table-wrap,.main .table-compact-wrap{border-radius:5px;overflow:auto}
.main .toolbar,.main .filters,.main .filterbar,.main .actions,.main .card-actions{background:#fbfcfb;border:1px solid #dfe6e1;border-radius:6px;padding:7px 8px;gap:7px}
.main .toolbar .btn,.main .actions .btn{min-height:31px}
.dashboard-grid{grid-template-columns:repeat(4,minmax(180px,1fr))!important;gap:8px!important}
.dashboard-grid .card{min-height:88px!important;border-radius:7px!important;padding:11px 13px!important;box-shadow:0 1px 3px rgba(20,58,36,.05)!important}
.dashboard-grid .card b,.dashboard-grid .card strong{letter-spacing:-.3px}
.dashboard-hero,.hero{background:linear-gradient(90deg,#075b39,#088451)!important;color:#fff!important;border:0!important;box-shadow:none!important}
.dashboard-hero h1,.dashboard-hero h2,.dashboard-hero h3,.hero h1,.hero h2,.hero h3{color:#fff!important}
.dashboard-hero .mut,.hero .mut{color:#e7f5ec!important}
.animal-table td,.animal-table th,.health-table td,.health-table th{vertical-align:middle!important}
.animal-tag,.ear-tag,.tag,.badge{border-radius:4px!important}
.photo-thumb{border-radius:5px!important}
.finance-summary-grid,.finance-kpi-grid{gap:8px!important}
.finance-summary-grid .card,.finance-kpi-grid .card{padding:10px!important;min-height:78px!important}
.finance-table td,.finance-table th{padding-top:6px!important;padding-bottom:6px!important}
.finance-table .btn,.finance-table button{min-height:28px!important;padding:4px 8px!important}
.finance-table td:last-child{white-space:nowrap!important;vertical-align:middle!important}
.feed-grid,.feed-list{gap:7px!important}
.feed-card{border-radius:6px!important;box-shadow:none!important}
.report-card,.backup-card,.user-card,.audit-card{border-radius:7px!important;box-shadow:none!important}
.main pre,.main code{font-family:Consolas,'Cascadia Mono',monospace}
.main>div,.main>section,.main>form{max-width:none}
.main p{margin-top:5px;margin-bottom:8px}
.main h1{font-size:22px!important;margin:7px 0 5px!important}.main h2{font-size:18px!important}.main h3{font-size:15px!important}
@media(max-width:1300px){.dashboard-grid{grid-template-columns:repeat(3,minmax(170px,1fr))!important}.main{padding-left:8px!important;padding-right:8px!important}}
@media(max-width:1000px){.dashboard-grid{grid-template-columns:repeat(2,minmax(160px,1fr))!important}}
@media(max-width:700px){.dashboard-grid{grid-template-columns:1fr!important}.erp-statusbar span:nth-child(2),.erp-statusbar span:nth-child(3){display:none}.main table{font-size:11px}}

'''

WORKBENCH_REFERENCE_UI_V3 = r"""
<style id="dev8-erp-repro-finance">
/* DEV8 — Üreme ERP + finans okunabilirlik */
.estrus-erp-shell,.insem-head{margin-bottom:10px}.estrus-erp-shell h1,.insem-head h1{margin:0 0 3px!important}
body:has(#estrusLiveTable) .two{display:grid!important;grid-template-columns:minmax(320px,390px) minmax(0,1fr)!important;gap:10px!important;align-items:start}
body:has(#estrusLiveTable) .two>.card{margin:0!important;border-radius:8px!important;box-shadow:0 1px 3px rgba(18,54,34,.05)!important}
body:has(#estrusLiveTable) #estrusLiveTable{font-size:12px}.estrus-actions .btn{min-height:32px;padding:6px 9px;font-size:11px}
body:has(#inseminationLiveTable) .insem-stats{grid-template-columns:repeat(4,minmax(0,1fr))!important;gap:8px!important}
body:has(#inseminationLiveTable) .insem-stats .card{border-radius:8px!important;padding:10px!important;box-shadow:0 1px 3px rgba(18,54,34,.05)!important}
body:has(#inseminationLiveTable) .insem-stats .stat b{font-size:24px!important}
body:has(#inseminationLiveTable) #inseminationForm{display:grid!important;grid-template-columns:1.35fr .65fr 1fr 1fr 1fr 1fr!important;gap:8px!important;align-items:end}
body:has(#inseminationLiveTable) #inseminationForm .full{grid-column:1/-1}
body:has(#inseminationLiveTable) .card:has(#inseminationForm){border-radius:8px!important;padding:12px!important}
.bulk-summary{display:grid!important;grid-template-columns:max-content minmax(210px,max-content) max-content!important;align-items:center!important;gap:10px!important}
.bulk-summary .pill{display:inline-flex!important;align-items:center!important;gap:5px!important;white-space:nowrap!important;min-width:max-content!important;max-width:none!important}
#bulkShare{display:inline-block!important;min-width:105px!important;text-align:right!important;font-variant-numeric:tabular-nums!important;overflow:visible!important;text-overflow:clip!important}
.bulk-selected-preview{white-space:normal!important;overflow-wrap:anywhere!important;line-height:1.4!important;min-height:20px}
@media(max-width:1050px){body:has(#estrusLiveTable) .two{grid-template-columns:1fr!important}body:has(#inseminationLiveTable) #inseminationForm{grid-template-columns:repeat(2,minmax(0,1fr))!important}.bulk-summary{grid-template-columns:1fr 1fr!important}.bulk-summary .btn{grid-column:1/-1}}
@media(max-width:700px){body:has(#inseminationLiveTable) .insem-stats{grid-template-columns:1fr 1fr!important}body:has(#inseminationLiveTable) #inseminationForm{grid-template-columns:1fr!important}.bulk-summary{grid-template-columns:1fr!important}.bulk-summary .pill{justify-content:space-between;width:100%;box-sizing:border-box}.bulk-summary .btn{grid-column:auto;width:100%}#bulkShare{min-width:0!important}.finance-drawer-body{padding-left:12px!important;padding-right:12px!important}}
</style>
<script id="dev9-feed-live-search">
(()=>{
  const norm=v=>(v||'').toLocaleLowerCase('tr-TR').normalize('NFD').replace(/[\u0300-\u036f]/g,'').replace(/ı/g,'i').replace(/ş/g,'s').replace(/ğ/g,'g').replace(/ü/g,'u').replace(/ö/g,'o').replace(/ç/g,'c');
  function applyFeedFilter(input){
    const box=input.closest('#quick-feed-add')||document.getElementById('quick-feed-add'); if(!box)return;
    const q=norm(input.value.trim()); let shown=0;
    box.querySelectorAll('.quick-feed-result').forEach(r=>{
      const hay=norm((r.dataset.search||'')+' '+(r.dataset.feedName||'')+' '+(r.textContent||''));
      const ok=!q||hay.includes(q); const visible=ok&&(q||shown<30);
      r.style.setProperty('display',visible?'flex':'none','important'); if(visible)shown++;
    });
  }
  document.addEventListener('input',e=>{if(e.target&&e.target.id==='quick-feed-search')applyFeedFilter(e.target)});
  document.addEventListener('click',e=>{const b=e.target.closest('.quick-filter');if(!b)return;const q=document.getElementById('quick-feed-search');if(!q)return;q.value=b.dataset.filter||'';applyFeedFilter(q);q.focus();});
  setTimeout(()=>{const q=document.getElementById('quick-feed-search');if(q)applyFeedFilter(q)},0);
})();
</script>


<style id="dev9-settings-center">
.settings-page-head{margin:0 0 10px}.settings-page-head h1{margin:0 0 3px}.settings-groups{display:grid;grid-template-columns:repeat(4,minmax(0,1fr));gap:8px;margin-bottom:10px}.settings-group{background:#fff;border:1px solid #dce6df;border-radius:8px;overflow:hidden}.settings-group h3{margin:0;padding:9px 11px;background:#f7faf8;border-bottom:1px solid #e2e9e4;font-size:12px}.settings-group a{display:flex;flex-direction:column;gap:2px;padding:9px 11px;border-bottom:1px solid #edf1ee}.settings-group a:last-child{border-bottom:0}.settings-group a:hover{background:#f1f8f3}.settings-group a b{font-size:11.5px}.settings-group a span{font-size:9.5px;color:#65736b}.farm-profile-head{display:flex;align-items:center;gap:14px}.farm-logo-placeholder,.farm-logo-preview{width:72px!important;height:72px!important;object-fit:cover;border-radius:10px}.settings-page-head+.settings-groups+.card{margin-top:0!important}
@media(max-width:1100px){.settings-groups{grid-template-columns:repeat(2,minmax(0,1fr))}}@media(max-width:650px){.settings-groups{grid-template-columns:1fr}.settings-group a{min-height:44px;justify-content:center}}
</style>
<style id="dev9-brand-polish">
@media(min-width:901px){
  .side{padding-top:0!important}
  .erp-side-brand{height:54px!important;display:flex!important;align-items:center!important;gap:8px!important;margin:0 -8px 8px!important;padding:0 16px!important;font-size:18px!important;border-bottom:1px solid rgba(255,255,255,.16)!important;background:rgba(0,0,0,.04)!important;box-sizing:border-box!important}
}
@media(max-width:900px){.erp-side-brand{display:flex!important;align-items:center!important;gap:8px!important;min-height:46px!important}}
</style>
<style id="workbench-reference-ui-v3">
body.erp-ration-reference .main{padding:8px 10px 30px!important;overflow-x:hidden!important}
body.erp-ration-reference .workbench-page-head{display:none!important}
body.erp-ration-reference .workbench-shell{padding:0!important;border:0!important;background:transparent!important;box-shadow:none!important;margin:0!important;overflow:visible!important}
body.erp-ration-reference .workbench-shell>.erp-ration-titlebar{display:flex;align-items:center;justify-content:space-between;gap:12px;background:#fff;border:1px solid #dfe7e2;border-radius:8px;padding:8px 10px;margin-bottom:8px}
body.erp-ration-reference .erp-ration-titlebar h2{font-size:15px!important;margin:0!important}.erp-ration-titlebar .mut{font-size:10px!important}
body.erp-ration-reference .erp-ration-layout{display:grid;grid-template-columns:255px minmax(0,1fr);gap:8px;align-items:start;min-width:0}
body.erp-ration-reference .erp-ration-left,body.erp-ration-reference .erp-ration-center,body.erp-ration-reference .erp-ration-right{min-width:0}
body.erp-ration-reference .erp-panel{background:#fff;border:1px solid #dfe7e2;border-radius:8px;box-shadow:0 1px 3px rgba(18,54,34,.05);overflow:hidden}
body.erp-ration-reference .erp-panel-head{display:flex;align-items:center;justify-content:space-between;gap:8px;padding:8px 10px;border-bottom:1px solid #e5ebe7;background:#fbfcfb;font-size:12px;font-weight:900}
body.erp-ration-reference .erp-panel-body{padding:8px 10px}
body.erp-ration-reference .target-workspace{margin:0!important}.erp-ration-reference .target-controlbar{margin:0 0 8px!important;padding:8px!important;border-radius:8px!important;display:flex!important;align-items:flex-end!important;gap:8px!important}.erp-ration-reference .target-controlbar .target-head{display:none!important}.erp-ration-reference .target-controlbar .target-form{justify-content:flex-start!important;display:grid!important;grid-template-columns:repeat(4,minmax(105px,1fr)) auto!important;gap:7px!important;width:100%!important}.erp-ration-reference .target-controlbar .target-form label{min-width:0!important;font-size:10px!important}.erp-ration-reference .target-controlbar input,.erp-ration-reference .target-controlbar select{min-height:30px!important;padding:5px 7px!important;font-size:11px!important}.erp-ration-reference .target-controlbar button{min-height:30px!important;padding:5px 12px!important}
body.erp-ration-reference .target-compare-sticky{position:static!important;margin:0 0 8px!important;padding:8px!important;box-shadow:none!important;border-radius:8px!important;background:#fff!important}.erp-ration-reference .target-compare-title{margin:0 0 6px!important}.erp-ration-reference .target-compare-title span{display:none!important}.erp-ration-reference .nutri-mini-grid{display:grid!important;grid-template-columns:repeat(4,minmax(0,1fr))!important;gap:6px!important;height:auto!important;overflow:visible!important}.erp-ration-reference .nutri-mini{height:92px!important;min-height:92px!important;border-radius:7px!important;padding:6px!important;grid-template-rows:16px 1fr 28px!important}.erp-ration-reference .nutri-card-title{font-size:11px!important;line-height:16px!important}.erp-ration-reference .nutri-side b{font-size:15px!important}.erp-ration-reference .nutri-card-footer{height:28px!important;min-height:28px!important}.erp-ration-reference .nutri-card-footer em{font-size:10px!important}.erp-ration-reference .nutri-card-footer .nutri-diff{font-size:8.5px!important}
body.erp-ration-reference #ration-workbench{margin:0!important;padding:0!important;border-radius:8px!important;overflow:hidden!important;box-shadow:none!important;border:1px solid #dfe7e2!important;background:#fff!important}.erp-ration-reference #ration-workbench .workbench-head{padding:8px 10px!important;margin:0!important;background:#fbfcfb;border-bottom:1px solid #e5ebe7;align-items:center!important}.erp-ration-reference #ration-workbench .workbench-head h3{font-size:12px!important}.erp-ration-reference #ration-workbench .workbench-head .mut{display:none!important}.erp-ration-reference #ration-workbench .workbench-actions .btn{font-size:10px!important;min-height:28px!important;padding:5px 8px!important}.erp-ration-reference #ration-workbench form{padding:0 8px 8px}.erp-ration-reference .compact-changebar{margin:5px 0!important}.erp-ration-reference .ration-workbench-table{font-size:10.5px!important;min-width:690px}.erp-ration-reference .ration-workbench-table th,.erp-ration-reference .ration-workbench-table td{padding:4px 5px!important;line-height:1.12!important}.erp-ration-reference .ration-workbench-table .ration-qty{width:60px!important;height:26px!important;padding:3px!important;font-size:12px!important}.erp-ration-reference .ration-workbench-table .btn{min-height:26px!important;padding:3px 7px!important}.erp-ration-reference .ration-savebar{position:static!important;bottom:auto!important;padding:6px 0 0!important;background:#fff!important}
body.erp-ration-reference #quick-feed-add{display:block!important;position:static!important;inset:auto!important;max-height:none!important;margin:0!important;padding:0!important;border:0!important;border-radius:0!important;box-shadow:none!important;background:#fff!important}.erp-ration-reference #quick-feed-add::before{display:none!important}.erp-ration-reference #quick-feed-add>summary{display:none!important}.erp-ration-reference #quick-feed-add .quick-feed-body{display:block!important;padding:0!important}.erp-ration-reference #quick-feed-add .quick-feed-tools{margin:0 0 7px!important;display:block!important}.erp-ration-reference #quick-feed-add .quick-feed-tools>input{min-width:0!important;width:100%!important;padding:7px 8px!important;border-radius:6px!important;font-size:11px!important}.erp-ration-reference #quick-feed-add .quick-feed-shortcuts{margin-top:5px;gap:3px!important}.erp-ration-reference #quick-feed-add .quick-feed-shortcuts .btn{font-size:9px!important;padding:3px 6px!important;min-height:24px!important}.erp-ration-reference #quick-feed-add .quick-feed-results{display:block!important;max-height:430px!important;overflow:auto!important;margin:0!important;border:1px solid #e3e9e5;border-radius:6px}.erp-ration-reference #quick-feed-add .quick-feed-result{display:flex!important;border:0!important;border-bottom:1px solid #edf1ee!important;border-radius:0!important;padding:7px 8px!important;gap:6px!important}.erp-ration-reference #quick-feed-add .quick-feed-result:last-child{border-bottom:0!important}.erp-ration-reference #quick-feed-add .quick-feed-result b{font-size:10.5px!important}.erp-ration-reference #quick-feed-add .quick-feed-result small{font-size:8.5px!important;line-height:1.15!important}.erp-ration-reference #quick-feed-add .quick-feed-side{display:none!important}.erp-ration-reference #quick-feed-add .quick-feed-selected{display:grid!important;grid-template-columns:1fr!important;gap:6px!important;margin:7px 0 0!important;padding:7px!important;border-radius:6px!important}.erp-ration-reference #quick-feed-add .quick-feed-selected .ration-stepper{justify-content:center!important}.erp-ration-reference #quick-feed-add .quick-feed-selected .btn{justify-content:center!important}.erp-ration-reference #quick-feed-add #quick-feed-close{display:none!important}
body.erp-ration-reference .erp-summary-list{display:grid;gap:0}.erp-ration-reference .erp-summary-row{display:flex;align-items:center;justify-content:space-between;gap:8px;padding:7px 0;border-bottom:1px solid #edf1ee;font-size:10.5px}.erp-ration-reference .erp-summary-row:last-child{border-bottom:0}.erp-ration-reference .erp-summary-row b{font-size:11.5px}.erp-ration-reference .erp-summary-row.total{padding-top:9px}.erp-ration-reference .erp-summary-row.total b{font-size:15px}.erp-ration-reference .erp-safety-list{display:grid;gap:0}.erp-ration-reference .erp-safety-row{display:flex;align-items:center;justify-content:space-between;gap:8px;padding:7px 0;border-bottom:1px solid #edf1ee;font-size:10px}.erp-ration-reference .erp-safety-row:last-child{border-bottom:0}.erp-ration-reference .erp-safety-state{font-weight:900;color:#16864a}.erp-ration-reference .erp-safety-state.warn{color:#b76b00}
body.erp-ration-reference .erp-secondary{margin-top:8px}.erp-ration-reference .erp-secondary>.card,.erp-ration-reference .erp-secondary>details{border-radius:8px!important;box-shadow:none!important}.erp-ration-reference #smart-balance{margin-top:8px!important}
@media(max-width:1300px){body.erp-ration-reference .erp-ration-layout{grid-template-columns:220px minmax(650px,1fr) 180px}.erp-ration-reference .nutri-mini-grid{grid-template-columns:repeat(4,minmax(0,1fr))!important}}
@media(max-width:1050px){body.erp-ration-reference .erp-ration-layout{grid-template-columns:1fr}.erp-ration-reference .erp-ration-left,.erp-ration-reference .erp-ration-right{position:static}.erp-ration-reference #quick-feed-add .quick-feed-results{max-height:260px!important}.erp-ration-reference .target-controlbar .target-form{grid-template-columns:repeat(2,minmax(120px,1fr))!important}.erp-ration-reference .nutri-mini-grid{grid-template-columns:repeat(2,minmax(0,1fr))!important}}
</style>
<script id="workbench-reference-ui-v3-script">
(()=>{
 // DEV4.7: masaüstü ERP DOM dönüşümü yalnız geniş ekranda. Mobil, yerel mobil kokpiti kullanır.
 if(window.matchMedia && window.matchMedia('(max-width: 900px)').matches) return;
 const shell=document.querySelector('.workbench-shell'); if(!shell||document.querySelector('.erp-ration-layout')) return;
 document.body.classList.add('erp-ration-reference');
 const original=[...shell.children];
 const title=original[0]; if(title){title.classList.add('erp-ration-titlebar');}
 const target=document.querySelector('.target-workspace');
 const table=document.getElementById('ration-workbench');
 const feed=document.getElementById('quick-feed-add');
 if(feed) feed.open=true;
 const layout=document.createElement('div'); layout.className='erp-ration-layout';
 const left=document.createElement('aside'); left.className='erp-ration-left erp-panel'; left.innerHTML='<div class="erp-panel-head"><span>Yem Havuzu</span><span class="mut">Katalog</span></div><div class="erp-panel-body" id="erp-feed-slot"></div>';
 const center=document.createElement('section'); center.className='erp-ration-center';
 layout.append(left,center);
 if(title&&title.parentNode===shell) title.after(layout); else shell.prepend(layout);
 if(feed) left.querySelector('#erp-feed-slot').append(feed);
 if(target) center.append(target);
 if(table) center.append(table);
 const secondary=document.createElement('div'); secondary.className='erp-secondary'; center.append(secondary);
 [...shell.children].forEach(el=>{if(el!==title&&el!==layout&&!layout.contains(el)) secondary.append(el)});
 const txt=id=>document.getElementById(id)?.textContent?.trim()||'—';
 const status=id=>document.getElementById(id)?.textContent?.trim()||'';
 function sync(){
   const data=[['Toplam KM',txt('target-mini-dm-current')],['Toplam ME',txt('target-mini-me-current')+' Mcal'],['Toplam HP',txt('target-mini-cp-current')],['Toplam NDF',txt('target-mini-ndf-current')],['Kaba / Kesif',txt('target-mini-rc-current')],['Günlük Maliyet',txt('target-mini-cost-current')]];
   // DEV8: Rasyon Özeti kaldırıldı; aynı değerler Hedef ↔ Mevcut kartlarında zaten gösteriliyor.
 }
 sync();
 const obs=new MutationObserver(sync); document.querySelectorAll('[id^="target-mini-"]').forEach(x=>obs.observe(x,{subtree:true,childList:true,characterData:true}));
})();
</script>
"""

# FINAL5 UI polish
FINAL5_UI_CSS = r"""
.dashboard-tabs{display:none!important}
.erp-side-brand{font-size:18px!important;padding:13px 12px!important;margin:0 0 12px!important;background:transparent!important;color:#fff!important;border-bottom:1px solid rgba(255,255,255,.14)!important;border-radius:0!important}
.erp-side-brand:hover{background:rgba(255,255,255,.08)!important}
.animal-tag-btn{border-radius:6px!important;background:#eef4ef!important;border:1px solid #cbdacf!important;padding:8px 12px!important;color:#126b3a!important;font-weight:800!important;box-shadow:none!important}
.animal-tag-btn:hover{background:#e3eee6!important;border-color:#9ebda8!important;transform:none!important}
.animal-tag-btn:before{display:none!important}.animal-tag-btn:after{content:'›'!important;margin-left:8px!important}
.top-user a{color:inherit!important;text-decoration:none!important}.top-user a:hover{text-decoration:underline!important}
.home-hero-link{color:#fff!important;text-decoration:none!important;display:flex!important;align-items:center!important;gap:12px!important;min-width:0}.home-hero-link:hover{opacity:.96}.farm-hero-logo{cursor:pointer}
.settings-hub{display:grid;grid-template-columns:repeat(3,minmax(0,1fr));gap:9px;margin:0 0 12px}.settings-tile{display:flex;flex-direction:column;gap:4px;padding:12px;border:1px solid #d8e3dc;border-radius:7px;background:#fff;color:#173b28!important;text-decoration:none!important}.settings-tile:hover,.settings-tile.active{border-color:#7fb493;background:#f2f8f4}.settings-tile span{font-size:11px;color:#6b7b72}
.today-work-grid{display:grid;grid-template-columns:repeat(3,minmax(0,1fr));gap:9px}.today-work-card{padding:0!important;overflow:hidden;display:flex;flex-direction:column;min-height:210px}.today-work-head{display:flex;justify-content:space-between;gap:10px;align-items:center;padding:11px 12px;border-bottom:1px solid #e1e8e3;background:#f8faf9}.today-work-head span{min-width:28px;height:28px;border-radius:5px;background:#e8f3ec;color:#126b3a;font-weight:900;display:grid;place-items:center;padding:0 6px}.today-work-body{padding:8px 10px;flex:1;max-height:145px;overflow:auto}.compact-alerts .alertitem{font-size:11px!important;padding:7px!important;margin:0!important}.compact-alerts .actions{display:none!important}.today-work-action{display:block;padding:9px 12px;border-top:1px solid #e1e8e3;background:#fbfcfb;color:#126b3a!important;font-weight:800;text-decoration:none!important}.today-finance-row{display:flex;justify-content:space-between;padding:9px 12px;border-bottom:1px solid #edf1ee}.today-finance-row.net{background:#f4f8f5}.today-title{margin-top:14px!important}
@media(max-width:900px){.side{padding-top:10px!important}.erp-side-brand{display:block!important}.top .brand{display:inline-flex!important;font-size:15px!important}.top-left{min-width:0}.top-user{font-size:0!important;white-space:nowrap}.top-user a{font-size:12px!important;margin-left:8px}.top-user b{display:none}.erp-tabs{display:none!important}.layout{padding-top:calc(var(--erp-top) + var(--erp-cmd))!important}.main{min-height:calc(100vh - var(--erp-top) - var(--erp-cmd))!important}.today-work-grid{grid-template-columns:1fr}.settings-hub{grid-template-columns:repeat(2,minmax(0,1fr))}.erp-commandbar{scrollbar-width:thin}.erp-commandbar a{flex:0 0 auto}.finance-toolbar-modern{display:grid!important;grid-template-columns:1fr 1fr!important}.finance-filter-actions{grid-column:1/-1!important;display:flex!important;flex-wrap:wrap!important}.finance-filter-actions .btn{flex:1 1 auto!important;text-align:center!important}.bulk-picker-head{align-items:stretch!important;flex-direction:column!important}.bulk-search{width:100%!important;max-width:none!important;box-sizing:border-box!important}}
@media(max-width:560px){.settings-hub{grid-template-columns:1fr}.top-user a:first-child{display:none}.erp-statusbar span:nth-child(2){display:none}.finance-toolbar-modern{grid-template-columns:1fr!important}.finance-filter-actions{grid-column:1!important}.hero{align-items:flex-start!important;flex-direction:column!important}.home-hero-link{width:100%}.dashboard-section-title{align-items:flex-start!important;flex-direction:column!important}.today-work-card{min-height:0}}
/* DEV7 mobile navigation + dashboard spacing hotfix */
.mobile-dashboard-command{display:none!important}
.dashboard-layout{padding-top:calc(var(--erp-top) + var(--erp-cmd))!important}
.dashboard-layout .main{min-height:calc(100vh - var(--erp-top) - var(--erp-cmd))!important}
@media(max-width:900px){
 .menu-toggle{display:inline-grid!important;place-items:center!important;color:#08683f!important;background:#edf5f0!important;border:1px solid #b9d2c2!important;width:38px!important;height:30px!important;padding:0!important;border-radius:5px!important;font-size:20px!important;line-height:1!important;flex:0 0 auto!important}
 .menu-toggle:hover{background:#dfeee4!important}
 .mobile-dashboard-command{display:flex!important}
 .side{top:var(--erp-top)!important;padding-top:8px!important;box-shadow:8px 0 24px rgba(0,0,0,.22)!important}
 .side.mobile-open{transform:translateX(0)!important}
 .erp-side-brand{margin-top:0!important}
}
"""

# DEV10 — Toparlama: global ERP görünümü / ayarlar / üreme / finans / sidebar
DEV10_GLOBAL_FIX = r"""
<style id="dev10-global-fix">
@media(min-width:901px){
  .side{top:0!important;padding-top:0!important}
  .erp-side-brand{height:78px!important;min-height:78px!important;margin:0 0 8px!important;padding:0 18px!important;font-size:20px!important;display:flex!important;align-items:center!important;background:linear-gradient(180deg,#087643,#066739)!important;border-bottom:1px solid rgba(255,255,255,.18)!important}
}
.erp-side-brand{cursor:pointer!important;text-decoration:none!important}
body.erp-ration-reference .erp-ration-layout{grid-template-columns:240px minmax(0,1fr)!important}
body.erp-ration-reference .erp-ration-right{display:none!important}
body.erp-ration-reference .erp-ration-center{min-width:0!important;width:100%!important}
.settings-page-head{margin:4px 0 14px!important}.settings-page-head h1{font-size:24px!important;margin:0 0 4px!important}.settings-page-head p{margin:0!important}
.settings-groups{display:grid!important;grid-template-columns:repeat(4,minmax(0,1fr))!important;gap:10px!important;margin:0 0 14px!important}
.settings-group{display:block!important;background:#fff!important;border:1px solid #d8e4dc!important;border-radius:10px!important;overflow:hidden!important;box-shadow:0 1px 3px rgba(18,58,36,.05)!important}
.settings-group h3{margin:0!important;padding:10px 12px!important;background:#f3f8f5!important;border-bottom:1px solid #dfe8e2!important;font-size:13px!important;color:#1b4a30!important}
.settings-group a{display:flex!important;flex-direction:column!important;gap:3px!important;padding:11px 12px!important;border-bottom:1px solid #edf2ee!important;text-decoration:none!important;color:#173c27!important;min-height:54px!important;box-sizing:border-box!important}
.settings-group a:last-child{border-bottom:0!important}.settings-group a:hover{background:#eef7f1!important}.settings-group a b{font-size:12px!important}.settings-group a span{font-size:10.5px!important;color:#65756b!important;line-height:1.25!important}
.farm-profile-head{display:flex!important;align-items:center!important;gap:14px!important}.farm-logo-placeholder,.farm-logo-preview{width:76px!important;height:76px!important;object-fit:cover!important;border-radius:12px!important}
@media(max-width:1150px){.settings-groups{grid-template-columns:repeat(2,minmax(0,1fr))!important}}@media(max-width:650px){.settings-groups{grid-template-columns:1fr!important}}
body:has(#estrusLiveTable) .taglink,body:has(#inseminationLiveTable) .taglink,body:has(#estrusLiveTable) td:first-child a,body:has(#inseminationLiveTable) td:first-child a{display:inline-flex!important;align-items:center!important;gap:6px!important;padding:7px 10px!important;border:1px solid #bfd6c6!important;border-radius:6px!important;background:#eef6f0!important;color:#08713d!important;font-weight:850!important;text-decoration:none!important;white-space:nowrap!important}
body:has(#estrusLiveTable) td:first-child a:hover,body:has(#inseminationLiveTable) td:first-child a:hover{background:#e3f0e7!important;border-color:#8eb79b!important}
#bulkAnimalBox .bulk-summary{display:grid!important;grid-template-columns:max-content minmax(280px,max-content) max-content!important;gap:10px!important;align-items:center!important;overflow:visible!important}
#bulkAnimalBox .bulk-summary .pill{overflow:visible!important;max-width:none!important;min-width:0!important;white-space:nowrap!important}
#bulkShare{display:inline-block!important;min-width:150px!important;max-width:none!important;overflow:visible!important;white-space:nowrap!important;text-align:right!important;font-size:15px!important;font-weight:900!important;font-variant-numeric:tabular-nums!important}
@media(max-width:760px){#bulkAnimalBox .bulk-summary{grid-template-columns:1fr!important}#bulkAnimalBox .bulk-summary .pill{width:100%!important;justify-content:space-between!important;box-sizing:border-box!important}#bulkShare{min-width:0!important;font-size:16px!important}}
/* DEV11.1: compact sidebar brand, optically centered in the reserved header area */
@media(min-width:901px){.side{padding-top:0!important}.erp-side-brand{height:78px!important;min-height:78px!important;margin:0 0 8px!important;padding:0 16px!important;display:flex!important;align-items:center!important;justify-content:center!important;text-align:center!important;font-size:18px!important;line-height:1!important;box-sizing:border-box!important}}
/* DEV11: target cockpit keeps the exact original left/right alignment while floating */
@media(min-width:1181px){body:has(.workbench-shell) .target-compare-sticky.is-floating{box-sizing:border-box!important;max-width:none!important}}
</style>
<script id="dev10-global-js">
(()=>{const removeRight=()=>document.querySelectorAll('.erp-ration-right').forEach(x=>x.remove());if(document.readyState==='loading')document.addEventListener('DOMContentLoaded',removeRight);else removeRight();})();
</script>
"""

def page(title,body,path='/',user='admin',flash=''):
    try:
        with db() as c:account=c.execute('select role,full_name from users where username=?',(user,)).fetchone()
        role=account['role'] if account else 'personel';display=account['full_name'] if account and account['full_name'] else user
    except Exception:role='personel';display=user
    def nav_link(name,url):
        return f'<a class="{"on" if path==url else ""}" href="{url}">{name}</a>'
    groups=[
        ('🐄 Hayvanlar',[('Dişi Hayvanlar','/animals'),('Erkek Hayvanlar','/males'),('Buzağılar','/calves'),('Kesilen Hayvanlar','/archive/slaughtered'),('Satılan Hayvanlar','/archive/sold'),('➕ Hayvan Ekle','/animal-add')]),
        ('🐂 Besi',[('🏠 Padok Yönetimi','/paddocks'),('🌾 Yem Kataloğu','/feeds'),('🥣 Rasyon Yönetimi','/rations'),('Besi Performansı','/performance')]),
        ('🩺 Üreme & Sağlık',[('Kızgınlık Takibi','/estrus'),('Tohumlama','/inseminations'),('Sağlık','/health')]),
        ('💰 Finans',[('Finans','/finance'),('Raporlar','/reports')]),
        ('🗄️ Veri & Sistem',[('Veri Aktarımı','/data'),('💾 Yedekleme Merkezi','/backups'),('📝 Sürüm Notları','/version-notes')]),
        ('⚙️ Yönetim',[('⚙️ Program Ayarları','/farm-profile'),('🔐 Şifremi Değiştir','/password-change')]+([('🔐 Lisans Bilgileri','/license-info'),('👥 Kullanıcı Yönetimi','/users'),('📜 İşlem Günlüğü','/audit-log')] if role=='admin' else []))
    ]
    nav='<a class="erp-side-brand" href="/" title="Ana Sayfa / Dashboard">🐄 <b>ÇiftlikPro</b></a>'+nav_link('🏠 Dashboard','/')
    for label,items in groups:
        active=any(path==url or (url=='/performance' and path.startswith('/performance')) for _,url in items)
        children=''.join(nav_link(name,url) for name,url in items)
        nav+=f'<details class="nav-group {"open-group" if active else ""}" {"open" if active else ""}><summary>{label}</summary><div class="nav-children">{children}</div></details>'
    if flash and path=='/rations' and len(flash)>260:
        flash_summary=flash.split('. ',1)[0].strip()
        if len(flash_summary)>180:
            flash_summary=flash_summary[:177].rsplit(' ',1)[0]+'…'
        fl=(f'<details class="flash ration-result-flash"><summary>{h(flash_summary)} <span>Ayrıntılar</span></summary>'
            f'<div class="ration-result-flash-detail">{h(flash)}</div></details>')
    else:
        fl=f'<div class="flash">{h(flash)}</div>' if flash else ''
    return f"""<!doctype html><html lang="tr"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>{h(title)}</title><style>{CSS}{DESKTOP_ERP_CSS}{ERP_ALL_MODULES_FINAL_CSS}{FINAL5_UI_CSS}
/* V3.7.8 Finance Drawer Hotfix */
#financeDrawerBackdrop.finance-drawer-backdrop{{position:fixed!important;inset:0!important;z-index:2090!important;display:none!important;background:rgba(9,34,22,.42)!important;backdrop-filter:blur(2px)}}
#financeDrawerBackdrop.finance-drawer-backdrop.open{{display:block!important}}
#financeDrawer.finance-drawer{{position:fixed!important;top:0!important;right:0!important;bottom:0!important;left:auto!important;width:min(760px,94vw)!important;height:100dvh!important;max-width:none!important;margin:0!important;padding:0!important;z-index:2100!important;display:flex!important;flex-direction:column!important;background:#f5f8f6!important;border:0!important;border-radius:0!important;box-shadow:-22px 0 55px rgba(12,45,28,.24)!important;transform:translateX(105%)!important;visibility:hidden!important;pointer-events:none!important;transition:transform .24s ease,visibility 0s linear .24s!important;overflow:hidden!important}}
#financeDrawer.finance-drawer.open{{transform:translateX(0)!important;visibility:visible!important;pointer-events:auto!important;transition:transform .24s ease!important}}
#financeDrawer .finance-drawer-head{{display:flex!important;align-items:flex-start!important;justify-content:space-between!important;gap:16px!important;padding:22px 24px 17px!important;flex:0 0 auto!important;background:#fff!important;border-bottom:1px solid #dce8df!important}}
#financeDrawer .finance-drawer-close{{appearance:none!important;display:grid!important;place-items:center!important;width:42px!important;height:42px!important;min-width:42px!important;padding:0!important;margin:0!important;border:0!important;border-radius:50%!important;background:#e8f3ec!important;color:#126c3b!important;font:700 28px/1 Arial,sans-serif!important;cursor:pointer!important}}
#financeDrawer .finance-drawer-body{{flex:1 1 auto!important;min-height:0!important;overflow-y:auto!important;padding:18px 20px 36px!important}}
#financeDrawer .finance-entry-card{{margin:0!important;width:100%!important;box-sizing:border-box!important}}
#financeDrawer .bulk-list{{max-height:310px!important;overflow-y:auto!important}}
@media(max-width:700px){{#financeDrawer.finance-drawer{{width:100vw!important;max-width:100vw!important}}#financeDrawer .finance-drawer-head{{padding:16px!important}}#financeDrawer .finance-drawer-body{{padding:12px 12px 28px!important}}#financeDrawer .bulk-list{{max-height:42vh!important}}}}

.ration-result-flash>summary{{font-weight:800;line-height:1.45;cursor:pointer}}.ration-result-flash>summary span{{white-space:nowrap;font-size:12px;color:#226b40;text-decoration:underline}}.ration-result-flash-detail{{margin-top:10px;padding-top:10px;border-top:1px solid #cfe2d5;line-height:1.5}}.ration-new-collapsed>details>summary{{cursor:pointer}}.target-context{{background:#f4faf6;border:1px solid #d8e9dd;border-radius:10px;padding:10px 12px;margin-bottom:12px}}.smart-solution-grid{{display:grid;grid-template-columns:repeat(3,minmax(0,1fr));gap:10px;margin-top:12px}}.smart-solution{{border:1px solid #dce8df;background:#f8fbf9;border-radius:12px;padding:12px}}.smart-solution b{{display:block;color:#164f31}}.smart-solution small{{display:block;margin-top:4px;font-weight:700}}.smart-solution p{{margin:8px 0;font-size:13px}}.ration-section-collapse>details>summary,details>summary{{cursor:pointer}}@media(max-width:900px){{.smart-solution-grid{{grid-template-columns:1fr}}}}

.quick-feed-card:not([open]){{padding:11px 14px}}.quick-feed-card[open]{{position:fixed;z-index:95;left:max(210px,8vw);right:4vw;top:58px;bottom:4vh;background:#fff;overflow:auto;padding:14px 16px!important;border-radius:16px!important;box-shadow:0 24px 70px rgba(0,0,0,.28)!important}}.quick-feed-card[open]::before{{content:'';position:fixed;inset:0;background:rgba(8,38,23,.38);z-index:-1}}.quick-feed-card[open] .quick-feed-head{{position:sticky;top:-14px;z-index:3;background:#fff;padding:10px 0;border-bottom:1px solid #e2ece5}}.quick-feed-card[open] .quick-feed-results{{max-height:42vh}}.quick-feed-card:not([open]) .quick-feed-head{{min-height:28px}}.quick-feed-card:not([open]) .quick-feed-head .mut{{font-size:12px}}@media(max-width:820px){{.quick-feed-card[open]{{left:8px;right:8px;top:48px;bottom:8px}}}}.quick-feed-card{{padding:16px}}.quick-feed-head,.quick-feed-selected{{display:flex;align-items:center;justify-content:space-between;gap:12px;flex-wrap:wrap}}.quick-feed-tools{{display:flex;gap:10px;align-items:center;flex-wrap:wrap;margin-top:12px}}.quick-feed-tools>input{{flex:1;min-width:280px;padding:11px 13px;border:1px solid #cbd8cf;border-radius:10px}}.quick-feed-shortcuts{{display:flex;gap:6px;flex-wrap:wrap}}.quick-feed-results{{margin-top:10px;display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:7px;max-height:290px;overflow:auto}}.quick-feed-result{{border:1px solid #dce8df;background:#fff;border-radius:10px;padding:10px 12px;text-align:left;display:flex;justify-content:space-between;gap:12px;cursor:pointer;color:inherit}}.quick-feed-result:hover,.quick-feed-result.selected{{border-color:#23824a;background:#f0f8f3}}.quick-feed-result span{{display:flex;flex-direction:column;gap:3px}}.quick-feed-result small{{color:#66736b}}.quick-feed-side{{text-align:right;white-space:nowrap}}.quick-feed-selected{{margin-top:12px;padding:11px;background:#f6faf7;border:1px solid #dce8df;border-radius:10px}}.quick-feed-selected>div:first-child{{display:flex;flex-direction:column}}.quick-feed-selected .ration-stepper input{{width:92px}}@media(max-width:760px){{.quick-feed-results{{grid-template-columns:1fr}}.quick-feed-tools>input{{min-width:100%}}.quick-feed-selected{{align-items:stretch}}.quick-feed-selected .ration-stepper{{justify-content:center}}.quick-feed-selected .btn{{justify-content:center}}}}
.target-card{{padding:14px 16px}}.target-head{{display:flex;align-items:center;justify-content:space-between;gap:10px;flex-wrap:wrap}}.target-head h3{{margin:0}}.target-context.compact{{margin:0;padding:7px 10px;font-size:13px}}.target-form{{display:grid;grid-template-columns:repeat(3,minmax(140px,1fr)) auto;gap:8px;align-items:end;margin-top:10px}}.target-form label{{font-size:12px}}.target-form input,.target-form select{{padding:7px 9px}}.compact-target-btn{{height:36px}}.target-more{{grid-column:auto}}.target-more summary{{font-size:12px;font-weight:700;padding:8px}}.target-more-grid{{display:grid;grid-template-columns:1fr 1fr;gap:6px;margin-top:6px}}.nutri-mini-grid{{display:grid;grid-template-columns:repeat(7,minmax(96px,1fr));gap:6px;margin-top:8px}}.nutri-mini{{border:1px solid #dce8df;border-radius:10px;padding:7px 8px;background:#fbfdfb;min-width:0}}.nutri-mini span{{display:block;font-size:11px;color:#587064}}.nutri-mini b{{display:block;font-size:15px;line-height:1.15;margin:2px 0}}.nutri-mini small{{display:block;font-size:11px;color:#3f5148;font-weight:800;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}}.nutri-mini em{{display:block;font-style:normal;font-size:10px;font-weight:700;margin-top:3px}}.nutri-diff{{display:block;font-style:normal;font-size:10px;font-weight:800;margin-top:4px;color:#607067}}.nutri-mini.warn .nutri-diff{{color:#bf4d32}}.target-live-note{{margin-top:7px;padding:7px 9px;border-radius:8px;background:#eef8f1;color:#2d6242;font-size:11px;font-weight:700}}.nutri-mini.ok{{border-left:3px solid #27a45e}}.nutri-mini.ok em{{color:#17733d}}.nutri-mini.warn{{border-left:3px solid #e6a11a;background:#fffdf7}}.nutri-mini.warn em{{color:#a96700}}.nutri-detail{{margin-top:8px}}.nutri-detail summary{{font-size:12px;color:#345b47}}.compact-note{{margin:7px 0 0;font-size:11px}}.table-compact-wrap{{overflow:auto;margin-top:7px}}.compact-table th,.compact-table td{{font-size:13px!important}}.ration-workbench-table th,.ration-workbench-table td,.smart-tech-table th,.smart-tech-table td{{padding:5px 7px;font-size:12px;vertical-align:middle}}.zebra tbody tr:nth-child(even),.ration-workbench-table tbody tr:nth-child(even),.smart-tech-table tbody tr:nth-child(even){{background:#f1f8f3}}.zebra tbody tr:nth-child(odd),.ration-workbench-table tbody tr:nth-child(odd),.smart-tech-table tbody tr:nth-child(odd){{background:#fff}}.ration-workbench-table tbody tr:hover,.smart-tech-table tbody tr:hover{{background:#e4f2e9}}.ration-workbench-table th:nth-child(1){{width:32%}}.ration-workbench-table th:nth-child(2){{width:22%}}.ration-workbench-table .ration-stepper{{gap:4px}}.ration-workbench-table .ration-qty{{width:72px;padding:6px}}.ration-live-grid{{display:none!important}}.ration-live{{position:static!important;padding:6px 8px!important;margin:5px 0!important;background:#f8fbf9!important}}.ration-live-metric{{padding:7px 8px!important}}.smart-solution{{padding:10px}}.smart-solution .effect{{font-size:11px;color:#456859;background:#eef7f1;border-radius:7px;padding:6px 7px;margin:7px 0}}.combo-apply{{margin-top:7px}}.smart-solution-grid{{gap:8px}}.ration-section-collapse{{padding:12px 14px}}.smart-tech-table th,.smart-tech-table td{{white-space:normal}}.ration-savebar{{margin-top:7px}}.ration-page-steps{{margin-bottom:8px}}
@media(max-width:1250px){{.nutri-mini-grid{{grid-template-columns:repeat(4,minmax(110px,1fr))}}.target-form{{grid-template-columns:repeat(2,minmax(150px,1fr)) auto}}.ration-workbench-table th,.ration-workbench-table td{{padding:5px 6px;font-size:11.5px}}}}
@media(max-width:820px){{.nutri-mini-grid{{grid-template-columns:repeat(2,minmax(120px,1fr))}}.target-form{{grid-template-columns:1fr}}.target-more{{grid-column:1}}.compact-target-btn{{width:100%}}.ration-workbench-table th:nth-child(3),.ration-workbench-table td:nth-child(3),.ration-workbench-table th:nth-child(4),.ration-workbench-table td:nth-child(4),.ration-workbench-table th:nth-child(5),.ration-workbench-table td:nth-child(5){{display:none}}}}
/* V3.9.15 - hedef çalışma masasıyla birlikte görünür */
.target-card{{padding-bottom:8px}}.nutri-mini-grid{{grid-template-columns:repeat(8,minmax(108px,1fr));position:sticky;top:48px;z-index:18;background:#f4f7f5;padding:7px;border:1px solid #cfe3d5;border-radius:12px;box-shadow:0 6px 18px rgba(24,73,46,.08)}}.nutri-mini{{padding:6px 7px}}.nutri-mini b{{font-size:16px}}.nutri-mini small{{font-weight:900;color:#203c2d}}.target-live-note{{display:none}}#ration-workbench{{margin-top:8px!important}}.workbench-head{{display:flex;justify-content:space-between;align-items:center;gap:10px;flex-wrap:wrap}}.workbench-actions{{display:flex;gap:8px;align-items:center;flex-wrap:wrap}}.quick-feed-card:not([open]){{display:none}}.quick-feed-card[open]{{display:block}}
@media(max-width:1350px){{.nutri-mini-grid{{grid-template-columns:repeat(4,minmax(120px,1fr));position:static}}}}
@media(max-width:820px){{.nutri-mini-grid{{grid-template-columns:repeat(2,minmax(120px,1fr));position:static}}.workbench-actions{{width:100%}}.workbench-actions .btn{{flex:1}}}}
/* V3.9.20 Hotfix3 UX - Hedef/Mevcut tek satır, daha dolu ve okunaklı */
.nutri-mini-grid{{grid-template-columns:repeat(8,minmax(132px,1fr))!important;gap:7px!important;position:static!important;padding:8px!important}}
.nutri-mini.nutri-compare-card{{padding:0!important;min-height:118px;border-radius:11px!important;overflow:hidden;background:#fff!important;display:flex;flex-direction:column}}
.nutri-card-title{{padding:7px 9px 6px;font-size:15px;font-weight:900;color:#183d2a;border-bottom:1px solid #e1ebe4;line-height:1.05}}
.nutri-compare-body{{display:grid;grid-template-columns:1fr 1fr;flex:1;min-height:56px}}
.nutri-side{{display:flex;flex-direction:column;justify-content:center;align-items:center;gap:3px;padding:6px 4px;text-align:center;min-width:0}}
.nutri-side+ .nutri-side{{border-left:1px solid #e1ebe4}}
.nutri-side span{{font-size:10.5px!important;font-weight:900;letter-spacing:.25px;color:#3b7d58!important}}
.nutri-side b{{font-size:18px!important;line-height:1.05!important;margin:0!important;color:#13271c;white-space:normal;overflow-wrap:anywhere}}
.nutri-card-footer{{display:flex;align-items:center;justify-content:space-between;gap:5px;min-height:34px;padding:6px 8px;background:#eef8f1;border-top:1px solid #dcebe1}}
.nutri-mini.warn .nutri-card-footer{{background:#fff7e9}}
.nutri-card-footer .nutri-diff{{margin:0!important;font-size:12px!important;font-weight:900;line-height:1.12}}
.nutri-card-footer em{{margin:0!important;font-size:12px!important;font-weight:900!important;line-height:1.12;text-align:right}}
@media(max-width:1450px){{.nutri-mini-grid{{grid-template-columns:repeat(4,minmax(150px,1fr))!important}}}}
@media(max-width:820px){{.nutri-mini-grid{{grid-template-columns:repeat(2,minmax(145px,1fr))!important}}}}
@media(max-width:560px){{.nutri-mini-grid{{grid-template-columns:1fr!important}}.nutri-mini.nutri-compare-card{{min-height:112px}}.nutri-side b{{font-size:19px!important}}.nutri-card-footer .nutri-diff,.nutri-card-footer em{{font-size:12.5px!important}}}}
/* V3.9.18 - Rasyon sayfası kompakt üst alan */
.ration-page-title{{margin:8px 0 4px!important;font-size:25px!important;line-height:1.05}}.ration-page-subtitle{{margin:0 0 6px!important;font-size:13px}}.ration-page-steps{{display:flex;gap:10px;flex-wrap:wrap;margin:3px 0 8px!important;font-size:12px;color:#486457}}.ration-new-collapsed{{padding:8px 12px!important;margin-bottom:8px}}.ration-new-collapsed>details>summary{{display:flex;align-items:center;gap:7px;min-height:30px}}.ration-new-collapsed>details>summary h2{{font-size:18px!important;line-height:1.1}}.ration-new-collapsed>details>summary .mut{{font-size:11px!important;margin-left:2px!important}}.ration-create-grid{{gap:8px!important;margin-top:8px!important}}.ration-create-grid>.card{{padding:10px!important}}@media(max-width:700px){{.ration-page-title{{font-size:23px!important}}.ration-new-collapsed>details>summary{{flex-wrap:wrap}}}}
.ration-picker-grid{{display:grid;grid-template-columns:repeat(auto-fill,minmax(205px,235px));gap:10px;margin-top:10px}}.ration-picker-card{{display:block;min-height:0;padding:12px 14px!important;border:1px solid #dce8df!important;text-decoration:none}}.ration-picker-card.active{{border:2px solid #176b3a!important;padding:11px 13px!important}}.ration-picker-head{{display:flex;align-items:center;justify-content:space-between;gap:8px}}.ration-picker-head h3{{margin:0;font-size:17px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}}.ration-picker-type{{font-size:12px;color:#60766a;white-space:nowrap}}.ration-picker-main{{display:flex;align-items:center;gap:6px;margin-top:8px;font-size:14px}}.ration-picker-ratio{{margin-top:5px;font-size:12px;color:#526b5e}}@media(max-width:700px){{.ration-picker-grid{{grid-template-columns:repeat(2,minmax(0,1fr))}}.ration-picker-card{{padding:10px!important}}.ration-picker-main{{font-size:13px;flex-wrap:wrap}}}}@media(max-width:430px){{.ration-picker-grid{{grid-template-columns:1fr}}}}
.ration-action-row{{display:grid;grid-template-columns:minmax(280px,430px) minmax(340px,520px);gap:10px;align-items:start;margin:10px 0 8px}}
.ration-action-row>.card{{margin:0!important;min-width:0}}
.ration-action-row .ration-new-collapsed{{margin-bottom:0!important}}
.smart-solve-card{{padding:8px 12px!important;border:1px solid #b9d8c4;background:linear-gradient(135deg,#f8fcf9,#eef8f1)}}.smart-solve-card summary{{cursor:pointer;display:flex;align-items:center;gap:7px;min-height:30px}}.smart-solve-card summary h2{{font-size:18px!important;line-height:1.1}}.smart-solve-card summary .mut{{font-size:11px!important;margin-left:2px!important}}
.solve-assistant-top{{position:sticky;top:0;z-index:8;margin:0 0 12px;padding:13px 14px;border:2px solid #e0a52b;border-radius:13px;background:#fff8e7;box-shadow:0 5px 16px rgba(126,87,6,.10)}}.solve-assistant-title{{display:flex;align-items:center;gap:9px;margin-bottom:7px}}.solve-assistant-title>span{{font-size:24px}}.solve-assistant-title b{{display:block;color:#6d4b00;font-size:16px}}.solve-assistant-title small{{display:block;color:#8b6b24;font-weight:700;margin-top:1px}}.solve-assistant-text{{font-size:13px;line-height:1.45;color:#4e3b10;font-weight:650;white-space:normal}}
.solve-search-wrap{{position:sticky;top:0;z-index:2;background:#f8fcf9;padding:2px 0 7px}}.solve-search{{width:100%;padding:10px 12px;border:1px solid #b9d2c1;border-radius:10px;font-size:14px;background:#fff}}.solve-search:focus{{outline:2px solid #9ecdb0;border-color:#4e9a6c}}
.solve-feed-grid{{display:grid;grid-template-columns:repeat(3,minmax(210px,1fr));gap:8px;max-height:330px;overflow:auto;padding:6px}}.solve-feed{{display:flex!important;gap:9px;align-items:flex-start;padding:10px;border:1px solid #dce8df;border-radius:11px;background:#fff}}.solve-feed input{{width:auto!important;margin:3px 0!important}}.solve-feed span{{display:block}}.solve-feed small{{display:block;color:#607067;margin-top:3px}}
@media(max-width:1050px){{.ration-action-row{{grid-template-columns:1fr 1fr}}.solve-feed-grid{{grid-template-columns:repeat(2,minmax(190px,1fr))}}}}
@media(max-width:800px){{.ration-action-row{{grid-template-columns:1fr}}.solve-feed-grid{{grid-template-columns:1fr}}}}

/* V3.9.20 Hotfix3 FINAL UX - tek satır hedef kartları + pH + eşit durum kutuları */
.target-compare-sticky{{position:sticky!important;top:8px!important;z-index:30!important;background:rgba(248,251,249,.98)!important;box-shadow:0 7px 22px rgba(24,73,46,.12)!important}}
.nutri-mini-grid{{display:grid!important;grid-template-columns:repeat(9,minmax(132px,1fr))!important;gap:7px!important;overflow-x:auto!important;overflow-y:hidden!important;padding:8px!important;align-items:stretch!important;scrollbar-width:thin}}
.nutri-mini.nutri-compare-card{{height:126px!important;min-height:126px!important;display:grid!important;grid-template-rows:auto 1fr 42px!important;padding:0!important;overflow:hidden!important}}
.nutri-card-title{{min-height:29px!important;display:flex!important;align-items:center!important;padding:7px 9px!important;font-size:15px!important}}
.nutri-compare-body{{min-height:0!important;height:auto!important}}
.nutri-side{{padding:5px 4px!important;gap:4px!important}}
.nutri-side span{{font-size:10.5px!important}}
.nutri-side b{{font-size:20px!important;line-height:1.02!important;font-weight:900!important}}
.nutri-card-footer{{height:42px!important;min-height:42px!important;box-sizing:border-box!important;padding:5px 7px!important;display:grid!important;grid-template-columns:minmax(0,1fr) minmax(0,1fr)!important;align-items:center!important;gap:5px!important}}
.nutri-card-footer .nutri-diff,.nutri-card-footer em{{font-size:12.5px!important;line-height:1.08!important;font-weight:900!important;margin:0!important;display:flex!important;align-items:center!important;min-height:30px!important}}
.nutri-card-footer .nutri-diff{{justify-content:flex-start!important;text-align:left!important}}
.nutri-card-footer em{{justify-content:flex-end!important;text-align:right!important}}
.nutri-mini.ok .nutri-card-footer{{background:#eaf7ef!important}}
.nutri-mini.warn .nutri-card-footer{{background:#fff4df!important}}
@media(max-width:1500px){{.nutri-mini-grid{{grid-template-columns:repeat(9,minmax(128px,1fr))!important}}.nutri-side b{{font-size:19px!important}}}}
@media(max-width:900px){{.target-compare-sticky{{top:4px!important}}.nutri-mini-grid{{grid-template-columns:repeat(9,minmax(145px,145px))!important}}.nutri-mini.nutri-compare-card{{height:124px!important}}}}
/* V3.9.20 Hotfix3 DEV UX2 — Drawer + dengeli hedef kartları */
body.ration-drawer-open{{overflow:hidden!important}}
.ration-action-launchers{{display:flex;gap:12px;align-items:stretch;margin:12px 0 10px;max-width:820px}}
.ration-launch-card{{appearance:none;border:1px solid #d7e5dc;background:#fff;border-radius:16px;padding:16px 18px;display:flex;align-items:center;gap:13px;text-align:left;cursor:pointer;min-width:260px;box-shadow:0 5px 16px rgba(18,62,38,.06);transition:.16s ease}}
.ration-launch-card:hover{{transform:translateY(-1px);box-shadow:0 8px 22px rgba(18,62,38,.10)}}
.ration-launch-card.add{{border-color:#b9ddc6;background:linear-gradient(135deg,#fbfffc,#f0faf4)}}.ration-launch-card.solve{{border-color:#b9d0f5;background:linear-gradient(135deg,#fbfdff,#f1f6ff)}}
.ration-launch-card .launch-icon{{width:46px;height:46px;border-radius:14px;display:grid;place-items:center;font-size:25px;font-weight:900;flex:0 0 46px}}.ration-launch-card.add .launch-icon{{background:#dff4e6;color:#087d3f}}.ration-launch-card.solve .launch-icon{{background:#e1ebff;color:#1259d8}}
.ration-launch-card b{{display:block;font-size:18px;line-height:1.1;color:#143422}}.ration-launch-card small{{display:block;margin-top:4px;color:#627468;font-size:12px}}
.ration-drawer-backdrop{{position:fixed;inset:0;z-index:2290;background:rgba(13,38,25,.38);backdrop-filter:blur(2px);opacity:0;visibility:hidden;transition:.22s ease}}.ration-drawer-backdrop.open{{opacity:1;visibility:visible}}
.ration-drawer{{position:fixed;top:0;right:0;bottom:0;width:min(760px,92vw);z-index:2300;background:#f7faf8;box-shadow:-24px 0 64px rgba(10,44,27,.23);transform:translateX(104%);visibility:hidden;transition:transform .24s ease,visibility 0s linear .24s;display:flex;flex-direction:column;overflow:hidden}}.ration-drawer.open{{transform:translateX(0);visibility:visible;transition:transform .24s ease}}
.ration-drawer-head{{padding:20px 22px 16px;background:#fff;border-bottom:1px solid #dce8df;display:flex;justify-content:space-between;align-items:flex-start;gap:16px;flex:0 0 auto}}.ration-drawer-head h2{{margin:0;font-size:24px}}.ration-drawer-head p{{margin:5px 0 0;color:#66756c;font-size:13px}}.ration-drawer-close{{width:42px;height:42px;border:0;border-radius:12px;background:#edf3ef;font-size:28px;line-height:1;color:#315342;cursor:pointer}}

.solve-selected-box{{position:sticky;top:50px;z-index:2;margin:0 0 10px;padding:10px 12px;border:1px solid #cfe2d5;border-radius:12px;background:#f7fcf8;box-shadow:0 3px 12px rgba(20,80,45,.06)}}
.solve-selected-head{{display:flex;align-items:center;justify-content:space-between;gap:10px;margin-bottom:7px;color:#244c34}}.solve-selected-head span{{font-size:12px;font-weight:800;color:#557062}}
.solve-selected-chips{{display:flex;gap:6px;flex-wrap:wrap;max-height:100px;overflow:auto}}.solve-chip{{display:grid;grid-template-columns:auto auto;grid-template-areas:'name x' 'cat x';align-items:center;gap:0 7px;border:1px solid #9bc9aa;border-radius:10px;background:#fff;padding:5px 7px 5px 9px;color:#183c27;cursor:pointer;text-align:left}}.solve-chip span{{grid-area:name;font-weight:800;font-size:12px;max-width:210px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}}.solve-chip small{{grid-area:cat;font-size:10px;color:#668071}}.solve-chip strong{{grid-area:x;font-size:18px;color:#9a3b32}}.solve-chip:hover{{background:#fff3f1;border-color:#d7a29d}}

.ration-drawer-form{{display:flex;flex-direction:column;min-height:0;flex:1}}.ration-drawer-body{{padding:18px 20px 26px;overflow:auto;min-height:0;flex:1}}.ration-drawer-foot{{padding:12px 18px;background:#fff;border-top:1px solid #dce8df;display:flex;align-items:center;gap:10px;justify-content:flex-end;flex:0 0 auto}}.ration-drawer-foot>span{{margin-right:auto;font-weight:800;color:#456153}}
.solve-target-grid{{display:grid;grid-template-columns:1fr 1fr;gap:12px}}.solve-target-grid label{{font-weight:800;font-size:13px}}.solve-target-grid input{{margin-top:6px;width:100%;box-sizing:border-box}}.solve-section-title{{margin:20px 0 9px;display:flex;flex-direction:column;gap:3px}}.solve-section-title b{{font-size:17px}}.solve-section-title span{{font-size:12px;color:#66756c}}.drawer-feed-grid{{grid-template-columns:1fr!important;max-height:none!important;overflow:visible!important;padding:0!important;gap:8px!important}}.drawer-feed-grid .solve-feed{{min-height:54px;align-items:center!important}}.drawer-feed-grid .solve-feed.solve-filter-hidden{{display:none!important}}.drawer-feed-grid .solve-feed:hover{{border-color:#7db692;background:#f2faf5}}.drawer-feed-grid .solve-feed:has(input:checked){{border-color:#2f9659;background:#edf9f1;box-shadow:0 0 0 1px #2f9659 inset}}.solve-search-wrap{{position:sticky!important;top:-18px!important;padding:10px 0!important;background:#f7faf8!important;z-index:3!important}}.solve-search{{font-size:15px!important;padding:12px 14px!important}}
.drawer-section-card{{background:#fff;border:1px solid #dce8df;border-radius:15px;padding:16px;margin-bottom:14px}}.drawer-section-card h3{{margin:0 0 12px;font-size:18px}}
.target-compare-sticky{{position:sticky!important;top:6px!important;z-index:30!important;background:rgba(248,251,249,.985)!important;border:1px solid #cfe3d5!important;border-radius:14px!important;padding:8px!important;box-shadow:0 8px 24px rgba(24,73,46,.11)!important}}
.target-compare-title{{display:flex!important;align-items:center!important;justify-content:space-between!important;gap:10px!important;padding:2px 4px 6px!important;font-size:13px!important}}.target-compare-title span{{font-size:11px!important;color:#678074!important}}
.nutri-mini-grid{{display:grid!important;grid-template-columns:repeat(5,minmax(150px,1fr))!important;gap:8px!important;overflow:visible!important;padding:0!important;align-items:stretch!important}}
.nutri-mini.nutri-compare-card{{height:148px!important;min-height:148px!important;display:grid!important;grid-template-rows:34px 1fr 52px!important;padding:0!important;border-radius:12px!important;background:#fff!important;overflow:hidden!important}}
.nutri-card-title{{height:34px!important;min-height:34px!important;box-sizing:border-box!important;padding:8px 10px!important;display:flex!important;align-items:center!important;font-size:16px!important;line-height:1!important}}
.nutri-compare-body{{display:grid!important;grid-template-columns:1fr 1fr!important;min-height:0!important}}.nutri-side{{padding:7px 5px!important;gap:5px!important}}.nutri-side span{{font-size:10px!important;letter-spacing:.35px!important}}.nutri-side b{{font-size:22px!important;line-height:1.05!important;font-weight:900!important;white-space:normal!important;overflow-wrap:normal!important;word-break:normal!important}}
.nutri-card-footer{{height:52px!important;min-height:52px!important;box-sizing:border-box!important;display:flex!important;flex-direction:column!important;align-items:center!important;justify-content:center!important;gap:2px!important;padding:5px 7px!important;text-align:center!important}}.nutri-card-footer em{{order:1!important;width:100%!important;justify-content:center!important;text-align:center!important;font-size:14px!important;font-weight:950!important;line-height:1.05!important;min-height:16px!important}}.nutri-card-footer .nutri-diff{{order:2!important;width:100%!important;justify-content:center!important;text-align:center!important;font-size:11.5px!important;font-weight:850!important;line-height:1.05!important;min-height:14px!important}}
.nutri-mini.ok .nutri-card-footer{{background:#eaf7ef!important}}.nutri-mini.warn .nutri-card-footer{{background:#fff3df!important}}
@media(min-width:1500px){{.nutri-mini-grid{{grid-template-columns:repeat(5,minmax(160px,1fr))!important}}}}
@media(max-width:1180px){{.nutri-mini-grid{{grid-template-columns:repeat(3,minmax(160px,1fr))!important}}.target-compare-sticky{{position:static!important}}}}

.solve-progress{{display:none;align-items:center;gap:8px;margin-left:auto;padding:7px 10px;border-radius:10px;background:#eaf3ff;color:#245b9d;font-weight:800;font-size:12px}}.solve-progress.on{{display:flex}}.solve-spinner{{width:16px;height:16px;border:2px solid #9bbce4;border-top-color:#2e6fc2;border-radius:50%;animation:solveSpin .7s linear infinite}}@keyframes solveSpin{{to{{transform:rotate(360deg)}}}}.solve-submit.solving{{opacity:.8;cursor:wait}}
@media(max-width:820px){{.ration-action-launchers{{max-width:none;display:grid;grid-template-columns:1fr 1fr}}.ration-launch-card{{min-width:0;padding:12px}}.ration-launch-card small{{display:none}}.ration-drawer{{width:100vw}}.solve-target-grid{{grid-template-columns:1fr}}.nutri-mini-grid{{grid-template-columns:repeat(2,minmax(145px,1fr))!important}}.nutri-side b{{font-size:20px!important}}}}
@media(max-width:520px){{.ration-action-launchers{{grid-template-columns:1fr}}.nutri-mini-grid{{grid-template-columns:1fr!important}}.ration-drawer-head{{padding:15px}}.ration-drawer-body{{padding:12px}}.ration-drawer-foot{{padding:10px;flex-wrap:wrap}}.ration-drawer-foot .btn.blue{{width:100%}}}}
\n/* DEV UX3 — tek sıra kompakt sticky hedef kartları */
.target-compare-sticky{{position:sticky!important;top:70px!important;z-index:29!important;padding:6px 7px!important;border-radius:12px!important;background:rgba(248,251,249,.985)!important;box-shadow:0 8px 22px rgba(24,73,46,.12)!important}}
.target-compare-title{{padding:0 3px 4px!important;margin:0!important;font-size:12px!important;min-height:18px!important}}
.nutri-mini-grid{{display:grid!important;grid-template-columns:repeat(9,minmax(0,1fr))!important;gap:6px!important;padding:0!important;overflow:visible!important;align-items:stretch!important}}
.nutri-mini.nutri-compare-card{{height:112px!important;min-height:112px!important;display:grid!important;grid-template-rows:28px 1fr 36px!important;border-radius:10px!important}}
.nutri-card-title{{height:28px!important;min-height:28px!important;padding:6px 8px!important;font-size:14px!important;line-height:1!important}}
.nutri-compare-body{{min-height:0!important}}.nutri-side{{padding:4px 3px!important;gap:2px!important}}.nutri-side span{{font-size:9px!important;line-height:1!important}}.nutri-side b{{font-size:18px!important;line-height:1!important;font-weight:900!important;white-space:normal!important;overflow-wrap:normal!important;word-break:normal!important}}
.nutri-card-footer{{height:36px!important;min-height:36px!important;padding:3px 5px!important;gap:1px!important;display:flex!important;flex-direction:column!important;justify-content:center!important;align-items:center!important}}
.nutri-card-footer em{{font-size:12px!important;line-height:1!important;min-height:13px!important;white-space:nowrap!important;overflow:hidden!important;text-overflow:ellipsis!important}}
.nutri-card-footer .nutri-diff{{font-size:9.5px!important;line-height:1!important;min-height:11px!important;white-space:nowrap!important;overflow:hidden!important;text-overflow:ellipsis!important}}
@media(max-width:1500px){{.nutri-mini-grid{{grid-template-columns:repeat(9,minmax(128px,1fr))!important;overflow-x:auto!important;padding-bottom:3px!important}}.target-compare-sticky{{overflow:hidden!important}}.nutri-mini.nutri-compare-card{{min-width:128px!important}}}}
@media(max-width:1180px){{.target-compare-sticky{{position:sticky!important;top:70px!important}}.nutri-mini-grid{{grid-template-columns:repeat(9,132px)!important;overflow-x:auto!important}}}}
@media(max-width:820px){{.target-compare-sticky{{position:static!important}}.nutri-mini-grid{{grid-template-columns:repeat(2,minmax(145px,1fr))!important;overflow:visible!important}}.nutri-mini.nutri-compare-card{{height:118px!important}}.nutri-side b{{font-size:19px!important}}}}
@media(max-width:520px){{.nutri-mini-grid{{grid-template-columns:1fr!important}}.nutri-mini.nutri-compare-card{{height:112px!important}}}}


/* DEV UX5 — 2 satır hedef kartları + sağlam floating summary */
.target-compare-sticky{{position:relative!important;top:auto!important;z-index:29!important;padding:7px 8px!important;border-radius:12px!important;background:rgba(248,251,249,.99)!important;box-shadow:0 7px 20px rgba(24,73,46,.10)!important;overflow:visible!important}}
.target-compare-sticky.is-floating{{position:fixed!important;z-index:2050!important;margin:0!important;box-shadow:0 10px 28px rgba(18,62,38,.18)!important;backdrop-filter:blur(8px)!important}}
.target-compare-placeholder{{display:none}}.target-compare-placeholder.active{{display:block}}
.nutri-mini-grid{{display:grid!important;grid-template-columns:repeat(5,minmax(0,1fr))!important;gap:7px!important;padding:0!important;overflow:visible!important;align-items:stretch!important}}
.nutri-mini.nutri-compare-card{{height:116px!important;min-height:116px!important;min-width:0!important;display:grid!important;grid-template-rows:29px 1fr 38px!important;border-radius:10px!important;overflow:hidden!important}}
.nutri-card-title{{height:29px!important;min-height:29px!important;padding:6px 8px!important;font-size:14px!important;line-height:1!important}}
.nutri-compare-body{{display:grid!important;grid-template-columns:1fr 1fr!important;min-height:0!important}}
.nutri-side{{padding:4px 4px!important;gap:2px!important;min-width:0!important}}.nutri-side span{{font-size:9.5px!important;line-height:1!important}}.nutri-side b{{font-size:19px!important;line-height:1.03!important;font-weight:900!important;white-space:normal!important;overflow-wrap:anywhere!important}}
.nutri-card-footer{{height:38px!important;min-height:38px!important;padding:3px 5px!important;gap:1px!important;display:flex!important;flex-direction:column!important;justify-content:center!important;align-items:center!important;box-sizing:border-box!important}}
.nutri-card-footer em{{font-size:12px!important;line-height:1!important;min-height:13px!important;white-space:nowrap!important;overflow:hidden!important;text-overflow:ellipsis!important;width:100%!important;text-align:center!important;justify-content:center!important}}.nutri-card-footer .nutri-diff{{font-size:9.5px!important;line-height:1!important;min-height:11px!important;white-space:nowrap!important;overflow:hidden!important;text-overflow:ellipsis!important;width:100%!important;text-align:center!important;justify-content:center!important}}
@media(max-width:1180px){{.nutri-mini-grid{{grid-template-columns:repeat(3,minmax(145px,1fr))!important}}.target-compare-sticky.is-floating{{position:relative!important;left:auto!important;top:auto!important;width:auto!important}}.target-compare-placeholder.active{{display:none!important}}}}
@media(max-width:820px){{.nutri-mini-grid{{grid-template-columns:repeat(2,minmax(140px,1fr))!important}}.nutri-mini.nutri-compare-card{{height:114px!important}}}}
@media(max-width:520px){{.nutri-mini-grid{{grid-template-columns:1fr!important}}}}


/* DEV UX6 — gerçek mobil Rasyon Kokpiti */
.target-compare-sticky{{position:sticky!important;top:70px!important;z-index:28!important;margin-top:8px!important;overflow:visible!important;background:rgba(248,251,249,.985)!important;backdrop-filter:blur(8px)!important;box-shadow:0 8px 24px rgba(20,68,42,.13)!important}}
.target-compare-sticky.is-floating{{position:sticky!important;left:auto!important;top:70px!important;width:auto!important;margin-top:8px!important}}
.target-compare-placeholder{{display:none!important}}

/* Masaüstü: okunaklı iki satır (5 + 4) */
.nutri-mini-grid{{grid-template-columns:repeat(5,minmax(0,1fr))!important;gap:7px!important;overflow:visible!important}}
.nutri-mini.nutri-compare-card{{height:114px!important;min-height:114px!important}}

@media(max-width:1180px){{
  .nutri-mini-grid{{grid-template-columns:repeat(3,minmax(145px,1fr))!important}}
  .target-compare-sticky{{position:sticky!important;top:70px!important}}
}}

@media(max-width:820px){{
  .target-compare-sticky{{position:sticky!important;top:62px!important;padding:6px!important;border-radius:12px!important}}
  .target-compare-title{{padding:0 2px 5px!important;font-size:12px!important}}
  .target-compare-title span{{display:none!important}}
  .nutri-mini-grid{{grid-template-columns:repeat(2,minmax(0,1fr))!important;gap:6px!important}}
  .nutri-mini.nutri-compare-card{{height:103px!important;min-height:103px!important;grid-template-rows:25px 1fr 33px!important}}
  .nutri-card-title{{height:25px!important;min-height:25px!important;padding:5px 7px!important;font-size:13px!important}}
  .nutri-side{{padding:3px 2px!important;gap:1px!important}}
  .nutri-side span{{font-size:8px!important}}
  .nutri-side b{{font-size:16px!important;line-height:1!important;white-space:nowrap!important;overflow:hidden!important;text-overflow:ellipsis!important;max-width:100%!important}}
  .nutri-card-footer{{height:33px!important;min-height:33px!important;padding:3px 4px!important}}
  .nutri-card-footer em{{font-size:10.5px!important;min-height:11px!important}}
  .nutri-card-footer .nutri-diff{{font-size:8.8px!important;min-height:9px!important}}

  /* Çalışma masasını mobil kart listesine dönüştür. */
  #ration-workbench{{padding:12px!important;overflow:visible!important}}
  #ration-workbench .workbench-head{{display:block!important}}
  #ration-workbench .workbench-head h3{{font-size:22px!important;line-height:1.1!important}}
  #ration-workbench .workbench-head .mut{{display:block;font-size:13px!important;line-height:1.35!important;margin-top:4px}}
  #ration-workbench .workbench-actions{{width:100%!important;margin-top:10px!important}}
  #ration-workbench .workbench-actions .btn{{width:100%!important;min-height:46px!important}}
  #ration-workbench form>div[style*="overflow:auto"]{{overflow:visible!important;margin-top:7px!important}}
  .ration-workbench-table,.ration-workbench-table tbody{{display:block!important;width:100%!important;min-width:0!important}}
  .ration-workbench-table thead{{display:none!important}}
  .ration-workbench-table tr.ration-row{{display:grid!important;grid-template-columns:1fr 1fr!important;grid-template-areas:'name name' 'qty qty' 'price daily' 'remove remove'!important;gap:7px 10px!important;width:100%!important;min-width:0!important;margin:0 0 9px!important;padding:12px!important;border:1px solid #dce8df!important;border-radius:14px!important;background:#fff!important;box-shadow:0 3px 12px rgba(20,65,40,.06)!important}}
  .ration-workbench-table tr.ration-row:nth-child(even){{background:#f6faf7!important}}
  .ration-workbench-table tr.ration-row td{{display:none!important;border:0!important;padding:0!important;width:auto!important;min-width:0!important;font-size:12px!important}}
  .ration-workbench-table tr.ration-row td:nth-child(1){{display:block!important;grid-area:name!important;font-size:14px!important;line-height:1.25!important;padding-bottom:5px!important;border-bottom:1px solid #e6eee8!important;overflow-wrap:anywhere!important}}
  .ration-workbench-table tr.ration-row td:nth-child(2){{display:block!important;grid-area:qty!important}}
  .ration-workbench-table tr.ration-row td:nth-child(6){{display:flex!important;grid-area:price!important;align-items:center!important;padding:7px 8px!important;background:#f1f6f2!important;border-radius:9px!important}}
  .ration-workbench-table tr.ration-row td:nth-child(6)::before{{content:'Fiyat';display:block;margin-right:auto;color:#718077;font-size:10px;font-weight:800}}
  .ration-workbench-table tr.ration-row td:nth-child(7){{display:flex!important;grid-area:daily!important;align-items:center!important;justify-content:flex-end!important;padding:7px 8px!important;background:#f1f6f2!important;border-radius:9px!important}}
  .ration-workbench-table tr.ration-row td:nth-child(7)::before{{content:'Günlük';display:block;margin-right:auto;color:#718077;font-size:10px;font-weight:800}}
  .ration-workbench-table tr.ration-row td:nth-child(8){{display:block!important;grid-area:remove!important;text-align:right!important}}
  .ration-workbench-table .ration-stepper{{display:grid!important;grid-template-columns:46px minmax(90px,1fr) 46px!important;gap:7px!important;width:100%!important;align-items:center!important}}
  .ration-workbench-table .ration-stepper .btn{{width:46px!important;height:44px!important;min-height:44px!important;padding:0!important;font-size:20px!important}}
  .ration-workbench-table .ration-qty{{width:100%!important;height:44px!important;min-width:0!important;font-size:18px!important;padding:7px!important}}
  .ration-workbench-table .qty-delta{{display:block!important;text-align:center!important;margin-top:4px!important;font-size:11px!important}}
  .ration-workbench-table .qty-zero{{width:100%!important;min-height:42px!important}}
  .ration-changebar{{display:block!important;margin:8px 0!important}}
  .ration-changebar #dirty-status{{display:block!important;margin-bottom:7px!important}}
  .ration-changebar #ration-reset{{width:100%!important}}
  .ration-savebar{{position:sticky!important;bottom:8px!important;z-index:18!important;padding:8px!important;margin:8px -4px -4px!important;background:rgba(255,255,255,.96)!important;border-radius:12px!important;box-shadow:0 -5px 18px rgba(18,60,37,.09)!important}}
  .ration-savebar .btn{{width:100%!important;min-height:48px!important;font-size:15px!important}}

  /* Drawer telefonda gerçek tam ekran çalışma alanı. */
  .ration-drawer{{width:100vw!important;max-width:100vw!important}}
  .ration-drawer-head{{padding:14px 14px 11px!important}}
  .ration-drawer-head h2{{font-size:21px!important}}
  .ration-drawer-body{{padding:10px 12px 18px!important}}
  .ration-drawer-foot{{padding:9px 10px!important;display:grid!important;grid-template-columns:1fr 1fr!important}}
  .ration-drawer-foot>span{{grid-column:1/-1!important;margin:0!important}}
  .ration-drawer-foot .btn.blue{{grid-column:1/-1!important;width:100%!important}}
  .drawer-feed-grid{{grid-template-columns:1fr!important}}
}}

@media(max-width:520px){{
  .main{{padding-left:8px!important;padding-right:8px!important}}
  .target-compare-sticky{{top:60px!important;padding:5px!important;margin-left:-2px!important;margin-right:-2px!important}}
  .nutri-mini-grid{{grid-template-columns:repeat(2,minmax(0,1fr))!important;gap:5px!important}}
  .nutri-mini.nutri-compare-card{{height:98px!important;min-height:98px!important;grid-template-rows:23px 1fr 31px!important}}
  .nutri-card-title{{height:23px!important;min-height:23px!important;font-size:12px!important;padding:4px 6px!important}}
  .nutri-side b{{font-size:15px!important}}
  .nutri-card-footer{{height:31px!important;min-height:31px!important}}
  .nutri-card-footer em{{font-size:10px!important}}
  .nutri-card-footer .nutri-diff{{font-size:8px!important}}
  #ration-workbench{{padding:10px!important;border-radius:14px!important}}
  #ration-workbench .workbench-head h3{{font-size:20px!important}}
  .ration-workbench-table tr.ration-row{{padding:10px!important}}
}}


/* DEV FINAL UX: çalışma masası solda, hedef ayar şeridi üstte, Canlı Analiz sağda tam görünür */
body:has(.workbench-shell) .workbench-shell{{display:grid;grid-template-columns:minmax(0,1fr) 370px;grid-template-rows:auto auto auto;gap:12px 14px;align-items:start;padding:14px!important;overflow:visible!important}}
body:has(.workbench-shell) .workbench-shell>div:first-child{{grid-column:1/-1;grid-row:1}}
/* target-workspace kabını kaldır; içindeki iki bölüm ana grid'e katılsın */
body:has(.workbench-shell) .workbench-shell>.target-workspace{{display:contents!important}}
/* Hedef düzenleme alanı sağ panelden çıktı: tüm genişlikte ince şerit */
body:has(.workbench-shell) .target-controlbar{{grid-column:1/-1;grid-row:2;display:grid!important;grid-template-columns:auto minmax(0,1fr);align-items:center!important;gap:12px!important;padding:8px 10px!important;margin:0!important;background:#fff;border:1px solid #dce8df;border-radius:12px;box-shadow:none!important}}
body:has(.workbench-shell) .target-controlbar .target-head{{display:flex;align-items:center;gap:9px;min-width:0}}
body:has(.workbench-shell) .target-controlbar .target-head h3{{font-size:14px!important;white-space:nowrap;margin:0!important}}
body:has(.workbench-shell) .target-controlbar .target-context{{margin:0!important;padding:5px 8px!important;font-size:12px!important;white-space:nowrap}}
body:has(.workbench-shell) .target-controlbar .target-form{{display:grid!important;grid-template-columns:repeat(3,minmax(118px,1fr)) auto!important;gap:7px!important;align-items:end!important;justify-content:end!important;margin:0!important}}
body:has(.workbench-shell) .target-controlbar .target-form label{{font-size:10px!important;min-width:0!important;margin:0!important}}
body:has(.workbench-shell) .target-controlbar .target-form input,body:has(.workbench-shell) .target-controlbar .target-form select{{min-height:32px!important;height:32px!important;padding:5px 7px!important;font-size:12px!important}}
body:has(.workbench-shell) .target-controlbar .compact-target-btn{{height:32px!important;min-height:32px!important;padding:5px 12px!important;grid-column:auto!important;width:auto!important}}
body:has(.workbench-shell) .target-controlbar .target-more{{display:none!important}}
/* Sol çalışma alanı */
body:has(.workbench-shell) .workbench-shell>#ration-workbench{{grid-column:1;grid-row:3;margin-top:0!important;min-width:0}}
body:has(.workbench-shell) .workbench-shell>#ration-workbench~*:not(script){{grid-column:1!important;min-width:0}}
/* Sağ Canlı Analiz: ekranın içinde tamamen görünür ve sayfa kayarken sabit */
body:has(.workbench-shell) .target-compare-sticky{{grid-column:2;grid-row:3;position:sticky!important;top:64px!important;align-self:start!important;height:calc(100vh - 82px)!important;max-height:calc(100vh - 82px)!important;overflow:hidden!important;margin:0!important;padding:8px!important;background:#f8fbf9!important;border:1px solid #cfe3d5!important;border-radius:14px!important;box-shadow:0 8px 24px rgba(24,73,46,.10)!important;z-index:18!important}}
body:has(.workbench-shell) .target-compare-title{{height:28px;display:flex!important;align-items:center!important;justify-content:space-between!important;gap:6px!important;margin:0 2px 6px!important;font-size:13px!important}}
body:has(.workbench-shell) .target-compare-title>b{{font-size:14px!important}}
body:has(.workbench-shell) .target-compare-title span{{font-size:9px!important;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;max-width:175px}}
body:has(.workbench-shell) .nutri-mini-grid{{position:static!important;display:grid!important;grid-template-columns:repeat(2,minmax(0,1fr))!important;grid-template-rows:repeat(5,minmax(0,1fr))!important;grid-auto-rows:minmax(0,1fr)!important;gap:6px!important;height:calc(100% - 34px)!important;padding:0!important;margin:0!important;border:0!important;box-shadow:none!important;background:transparent!important;overflow:hidden!important}}
body:has(.workbench-shell) .nutri-mini{{min-height:0!important;height:auto!important;padding:5px 6px!important;display:grid!important;grid-template-rows:auto minmax(0,1fr) auto!important;overflow:hidden!important;border-radius:10px!important}}
body:has(.workbench-shell) .nutri-card-title{{font-size:11px!important;line-height:1.05!important;margin:0 0 2px!important}}
body:has(.workbench-shell) .nutri-compare-body{{min-height:0!important;display:grid!important;grid-template-columns:1fr 1fr!important;align-items:center!important}}
body:has(.workbench-shell) .nutri-side{{padding:2px 3px!important;min-width:0!important}}
body:has(.workbench-shell) .nutri-side span{{font-size:8px!important;line-height:1!important}}
body:has(.workbench-shell) .nutri-side b{{font-size:14px!important;line-height:1.05!important;margin-top:2px!important;white-space:normal!important;overflow-wrap:anywhere!important}}
body:has(.workbench-shell) .nutri-card-footer{{min-height:25px!important;padding-top:2px!important;display:grid!important;align-content:center!important;gap:0!important}}
body:has(.workbench-shell) .nutri-card-footer em{{font-size:9.5px!important;line-height:1.05!important;font-weight:900!important}}
body:has(.workbench-shell) .nutri-card-footer .nutri-diff{{font-size:8px!important;line-height:1.05!important;margin-top:1px!important;white-space:normal!important}}
body:has(.workbench-shell) .nutri-detail,body:has(.workbench-shell) .target-live-note{{display:none!important}}
/* Dokuzuncu kart tek kalınca sol yarıda kalsın; genişletip ritmi bozma */
body:has(.workbench-shell) #target-mini-cost{{grid-column:1}}
@media(max-width:1280px){{
 body:has(.workbench-shell) .workbench-shell{{grid-template-columns:minmax(0,1fr) 340px}}
 body:has(.workbench-shell) .target-controlbar{{grid-template-columns:1fr}}
 body:has(.workbench-shell) .target-controlbar .target-form{{grid-template-columns:repeat(3,minmax(105px,1fr)) auto!important}}
}}
@media(max-width:1080px){{
 body:has(.workbench-shell) .workbench-shell{{grid-template-columns:1fr!important}}
 body:has(.workbench-shell) .target-controlbar,body:has(.workbench-shell) .target-compare-sticky,body:has(.workbench-shell) .workbench-shell>#ration-workbench,body:has(.workbench-shell) .workbench-shell>#ration-workbench~*:not(script){{grid-column:1!important;grid-row:auto!important}}
 body:has(.workbench-shell) .target-compare-sticky{{position:static!important;height:auto!important;max-height:none!important;overflow:visible!important}}
 body:has(.workbench-shell) .nutri-mini-grid{{height:auto!important;grid-template-columns:repeat(3,minmax(0,1fr))!important;grid-template-rows:none!important;grid-auto-rows:auto!important}}
 body:has(.workbench-shell) #target-mini-cost{{grid-column:auto}}
}}
@media(max-width:820px){{
 body:has(.workbench-shell) .workbench-shell{{padding:8px!important;gap:9px!important}}
 body:has(.workbench-shell) .target-controlbar{{display:block!important}}
 body:has(.workbench-shell) .target-controlbar .target-head{{display:flex!important;justify-content:space-between!important;flex-wrap:wrap!important;margin-bottom:7px!important}}
 body:has(.workbench-shell) .target-controlbar .target-form{{grid-template-columns:1fr!important}}
 body:has(.workbench-shell) .target-controlbar .compact-target-btn{{width:100%!important}}
 body:has(.workbench-shell) .nutri-mini-grid{{grid-template-columns:repeat(2,minmax(0,1fr))!important}}
}}

/* HOTFIX4 UX: hedef kartları çalışma masasının üstünde, iki satır ve okunaklı sabit kokpit */
body:has(.workbench-shell) .workbench-shell{{display:block!important;padding:14px!important;overflow:visible!important}}
body:has(.workbench-shell) .workbench-shell>.target-workspace{{display:block!important}}
body:has(.workbench-shell) .target-controlbar{{display:grid!important;grid-template-columns:auto minmax(0,1fr)!important;gap:12px!important;margin:8px 0!important;padding:8px 10px!important}}
body:has(.workbench-shell) .target-compare-sticky{{position:sticky!important;top:62px!important;height:auto!important;max-height:none!important;overflow:visible!important;margin:8px 0 10px!important;padding:9px 10px!important;background:rgba(248,251,249,.985)!important;backdrop-filter:blur(8px)!important;z-index:40!important;box-shadow:0 8px 22px rgba(20,65,40,.14)!important}}
body:has(.workbench-shell) .target-compare-title{{height:auto!important;margin:0 2px 7px!important;font-size:13px!important}}
body:has(.workbench-shell) .target-compare-title>b{{font-size:14px!important}}
body:has(.workbench-shell) .target-compare-title span{{font-size:10px!important;max-width:none!important}}
body:has(.workbench-shell) .nutri-mini-grid{{position:static!important;display:grid!important;grid-template-columns:repeat(5,minmax(0,1fr))!important;grid-template-rows:none!important;grid-auto-rows:112px!important;height:auto!important;gap:7px!important;padding:0!important;margin:0!important;border:0!important;background:transparent!important;box-shadow:none!important;overflow:visible!important}}
body:has(.workbench-shell) .nutri-mini{{min-height:112px!important;height:112px!important;padding:7px 8px!important;display:grid!important;grid-template-rows:18px 1fr 34px!important;border-radius:10px!important;overflow:hidden!important}}
body:has(.workbench-shell) .nutri-card-title{{font-size:13px!important;line-height:18px!important;font-weight:900!important;margin:0!important}}
body:has(.workbench-shell) .nutri-compare-body{{display:grid!important;grid-template-columns:1fr 1fr!important;align-items:center!important;min-height:0!important}}
body:has(.workbench-shell) .nutri-side{{padding:2px 5px!important;min-width:0!important;text-align:center!important}}
body:has(.workbench-shell) .nutri-side span{{font-size:9px!important;line-height:1!important;font-weight:800!important}}
body:has(.workbench-shell) .nutri-side b{{font-size:18px!important;line-height:1.05!important;margin-top:4px!important;white-space:normal!important;overflow-wrap:anywhere!important}}
body:has(.workbench-shell) .nutri-card-footer{{min-height:34px!important;height:34px!important;padding:4px 3px 2px!important;display:grid!important;grid-template-columns:1fr!important;align-content:center!important;justify-items:center!important;gap:1px!important}}
body:has(.workbench-shell) .nutri-card-footer em{{font-size:12px!important;line-height:1!important;font-weight:900!important;white-space:nowrap!important}}
body:has(.workbench-shell) .nutri-card-footer .nutri-diff{{font-size:10px!important;line-height:1.05!important;margin:1px 0 0!important;white-space:nowrap!important;font-weight:800!important}}
body:has(.workbench-shell) #target-mini-cost{{grid-column:auto!important}}
body:has(.workbench-shell) #ration-workbench{{margin-top:0!important}}
@media(max-width:1350px){{body:has(.workbench-shell) .nutri-mini-grid{{grid-template-columns:repeat(4,minmax(0,1fr))!important}}}}
@media(max-width:1080px){{body:has(.workbench-shell) .nutri-mini-grid{{grid-template-columns:repeat(3,minmax(0,1fr))!important}}body:has(.workbench-shell) .target-compare-sticky{{position:static!important}}}}
@media(max-width:820px){{body:has(.workbench-shell) .nutri-mini-grid{{grid-template-columns:repeat(2,minmax(0,1fr))!important;grid-auto-rows:116px!important}}body:has(.workbench-shell) .nutri-mini{{height:116px!important;min-height:116px!important}}}}

/* HOTFIX4 DEV FINAL STICKY TARGET COCKPIT */
body:has(.workbench-shell) .workbench-shell{{overflow:visible!important;position:relative!important}}
body:has(.workbench-shell) .target-workspace{{overflow:visible!important;position:relative!important}}
body:has(.workbench-shell) .target-compare-sticky{{position:sticky!important;top:70px!important;z-index:120!important;margin:8px 0 12px!important;padding:7px 8px!important;background:rgba(248,251,249,.985)!important;border:1px solid #cfe3d5!important;border-radius:13px!important;box-shadow:0 8px 22px rgba(19,66,40,.14)!important;backdrop-filter:blur(8px)!important;overflow:visible!important}}
body:has(.workbench-shell) .target-compare-title{{min-height:18px!important;height:auto!important;margin:0 2px 5px!important;padding:0!important;display:flex!important;align-items:center!important;justify-content:space-between!important;gap:10px!important;font-size:12px!important;line-height:1.1!important}}
body:has(.workbench-shell) .target-compare-title>b{{font-size:13px!important}}
body:has(.workbench-shell) .target-compare-title span{{font-size:9.5px!important;white-space:nowrap!important;overflow:hidden!important;text-overflow:ellipsis!important}}
body:has(.workbench-shell) .nutri-mini-grid{{display:grid!important;grid-template-columns:repeat(5,minmax(0,1fr))!important;gap:7px!important;padding:0!important;margin:0!important;overflow:visible!important;align-items:stretch!important}}
body:has(.workbench-shell) .nutri-mini.nutri-compare-card{{height:101px!important;min-height:101px!important;min-width:0!important;display:grid!important;grid-template-rows:24px 1fr 31px!important;padding:0!important;border-radius:10px!important;overflow:hidden!important;background:#fff!important}}
body:has(.workbench-shell) .nutri-card-title{{height:24px!important;min-height:24px!important;padding:5px 8px!important;box-sizing:border-box!important;display:flex!important;align-items:center!important;font-size:13px!important;line-height:1!important;font-weight:900!important}}
body:has(.workbench-shell) .nutri-compare-body{{display:grid!important;grid-template-columns:1fr 1fr!important;min-height:0!important}}
body:has(.workbench-shell) .nutri-side{{min-width:0!important;padding:3px 4px!important;gap:2px!important;display:flex!important;flex-direction:column!important;justify-content:center!important;align-items:center!important}}
body:has(.workbench-shell) .nutri-side span{{font-size:8.5px!important;line-height:1!important;font-weight:900!important;white-space:nowrap!important}}
body:has(.workbench-shell) .nutri-side b{{font-size:18px!important;line-height:1.02!important;font-weight:950!important;margin:2px 0 0!important;white-space:normal!important;overflow-wrap:normal!important;word-break:normal!important;text-align:center!important;max-width:100%!important}}
body:has(.workbench-shell) .nutri-card-footer{{height:31px!important;min-height:31px!important;padding:3px 5px!important;box-sizing:border-box!important;display:grid!important;grid-template-columns:1fr!important;grid-template-rows:14px 11px!important;gap:0!important;align-items:center!important;justify-items:center!important;text-align:center!important}}
body:has(.workbench-shell) .nutri-card-footer em{{width:100%!important;min-height:14px!important;margin:0!important;font-size:11.5px!important;line-height:1!important;font-weight:950!important;display:block!important;text-align:center!important;white-space:nowrap!important;overflow:hidden!important;text-overflow:ellipsis!important}}
body:has(.workbench-shell) .nutri-card-footer .nutri-diff{{width:100%!important;min-height:11px!important;margin:0!important;font-size:8.8px!important;line-height:1!important;font-weight:850!important;display:block!important;text-align:center!important;white-space:nowrap!important;overflow:hidden!important;text-overflow:ellipsis!important}}
body:has(.workbench-shell) .nutri-mini.ok .nutri-card-footer{{background:#e9f7ee!important}}
body:has(.workbench-shell) .nutri-mini.warn .nutri-card-footer{{background:#fff3dc!important}}
body:has(.workbench-shell) #target-mini-cost{{grid-column:auto!important}}
body:has(.workbench-shell) #ration-workbench{{margin-top:0!important}}
@media(max-width:1250px){{body:has(.workbench-shell) .nutri-mini-grid{{grid-template-columns:repeat(4,minmax(0,1fr))!important}}body:has(.workbench-shell) .nutri-mini.nutri-compare-card{{height:104px!important;min-height:104px!important}}}}
@media(max-width:920px){{body:has(.workbench-shell) .target-compare-sticky{{position:static!important;top:auto!important}}body:has(.workbench-shell) .nutri-mini-grid{{grid-template-columns:repeat(2,minmax(0,1fr))!important}}body:has(.workbench-shell) .nutri-side b{{font-size:19px!important}}}}
@media(max-width:520px){{body:has(.workbench-shell) .nutri-mini-grid{{grid-template-columns:1fr!important}}}}

/* HOTFIX4 UX10 — ana görünüm ve FIXED görünüm birebir 4+4 */
@media(min-width:1181px){{
  body:has(.workbench-shell) .target-compare-sticky,
  body:has(.workbench-shell) .target-compare-sticky.is-floating{{box-sizing:border-box!important}}
  body:has(.workbench-shell) .nutri-mini-grid,
  body:has(.workbench-shell) .target-compare-sticky.is-floating .nutri-mini-grid{{grid-template-columns:repeat(4,minmax(0,1fr))!important}}
}}

/* HOTFIX4 FINAL FIX — hedef kokpit gerçek fixed: yalnızca çalışma masası kayar */
@media(min-width:1181px){{
  body:has(.workbench-shell) .target-compare-sticky{{position:relative!important;top:auto!important;z-index:29!important}}
  body:has(.workbench-shell) .target-compare-sticky.is-floating{{position:fixed!important;top:70px!important;z-index:2050!important;margin:0!important;background:rgba(248,251,249,.99)!important;backdrop-filter:blur(8px)!important;box-shadow:0 10px 28px rgba(18,62,38,.18)!important;overflow:visible!important}}
  body:has(.workbench-shell) .target-compare-placeholder{{display:none!important}}
  body:has(.workbench-shell) .target-compare-placeholder.active{{display:block!important}}
}}
@media(max-width:1180px){{
  body:has(.workbench-shell) .target-compare-sticky,
  body:has(.workbench-shell) .target-compare-sticky.is-floating{{position:static!important;left:auto!important;top:auto!important;width:auto!important;margin-top:8px!important}}
  body:has(.workbench-shell) .target-compare-placeholder.active{{display:none!important}}
}}

/* HOTFIX4 UX11 — fixed durumda yalnız konum değişir; kart ölçüleri birebir korunur */
@media(min-width:1181px){{
  body:has(.workbench-shell) .target-compare-sticky.is-floating{{box-sizing:border-box!important;min-width:0!important;max-width:none!important}}
  body:has(.workbench-shell) .target-compare-sticky .nutri-mini-grid,
  body:has(.workbench-shell) .target-compare-sticky.is-floating .nutri-mini-grid{{grid-template-columns:repeat(4,minmax(0,1fr))!important}}
}}

/* HOTFIX4 UX12 — fixed kokpit: ana içerik içinde eşit sol/sağ boşluk + rasyon hazırlama raporu */
@media(min-width:1181px){{
  body:has(.workbench-shell) .target-compare-sticky.is-floating{{width:auto!important;max-width:none!important;box-sizing:border-box!important}}
}}
.ration-prep-report{{border:1px solid #cfe3d5!important;background:#fff!important}}
.prep-report-head{{display:flex;align-items:flex-end;justify-content:space-between;gap:14px;flex-wrap:wrap}}
.prep-report-controls{{display:flex;align-items:flex-end;gap:9px;flex-wrap:wrap}}
.prep-report-controls label{{font-size:12px;font-weight:800;color:#365b48}}
.prep-report-controls input{{display:block;width:130px;margin-top:4px;padding:9px 10px;border:1px solid #bcd0c2;border-radius:9px;font-size:16px;font-weight:900}}
.prep-report-summary{{display:grid;grid-template-columns:repeat(4,minmax(150px,1fr));gap:8px;margin:12px 0}}
.prep-report-summary>div{{background:#f2f8f4;border:1px solid #d9e9de;border-radius:10px;padding:9px 11px}}
.prep-report-summary span{{display:block;font-size:10px;color:#65766d;font-weight:800}}.prep-report-summary b{{display:block;margin-top:3px;font-size:18px;color:#163c29}}
.prep-report-table-wrap{{overflow:auto}}.prep-report-table{{width:100%;border-collapse:collapse}}.prep-report-table th,.prep-report-table td{{padding:8px 9px;border-bottom:1px solid #e1ebe4;text-align:left}}.prep-report-table th{{background:#eaf4ed;font-size:12px}}.prep-report-note{{margin-top:8px;font-size:11px}}
@media(max-width:820px){{.prep-report-summary{{grid-template-columns:repeat(2,minmax(0,1fr))}}.prep-report-controls{{width:100%}}.prep-report-controls label{{flex:1}}.prep-report-controls input{{width:100%}}.prep-report-controls .btn{{flex:1}}}}
@media(max-width:520px){{.prep-report-summary{{grid-template-columns:1fr}}.prep-report-table th:nth-child(4),.prep-report-table td:nth-child(4){{display:none}}}}
</style></head><body><div class="top"><div class="top-left"><button class="menu-toggle" id="menuToggle" aria-label="Menüyü aç">☰</button><a class="brand" href="/" title="Ana Sayfa">🐄 ÇiftlikPro</a></div><div class="top-user"><a href="/#approaching-estrus">🔔 Bildirimler</a> &nbsp;&nbsp; <a href="/farm-profile">⚙ Ayarlar</a> &nbsp;&nbsp; <b>{h(display)}</b> · <a href="/logout">Çıkış</a></div></div><div class="erp-commandbar"><a class="mobile-dashboard-command" href="/"><span class="ico">⌂</span>Dashboard</a><a href="/animal-add"><span class="ico">＋</span>Yeni Kayıt</a><a href="/rations"><span class="ico">⚖</span>Rasyon</a><a href="/feeds"><span class="ico">🌾</span>Yem Kataloğu</a><a href="/finance"><span class="ico">₺</span>Finans</a><a href="/reports"><span class="ico">▥</span>Raporlar</a><a href="/data"><span class="ico">⇄</span>Veri</a></div><div class="erp-tabs {'dashboard-tabs' if path=='/' else ''}"><div class="erp-tab">{h(title)}</div></div><div class="layout {'dashboard-layout' if path=='/' else ''}"><aside class="side" id="sideMenu">{nav}</aside><main class="main">{fl}{body}</main></div><div class="erp-statusbar"><span>Durum: Hazır</span><span>Veritabanı: Bağlı</span><span>Aktif Kullanıcı: {h(display)}</span><span class="erp-version">v3.9.20</span></div><script>
(function(){{
 const btn=document.getElementById("menuToggle"),side=document.getElementById("sideMenu");
 if(btn&&side){{btn.addEventListener("click",function(){{side.classList.toggle("mobile-open");}});side.querySelectorAll("a").forEach(function(a){{a.addEventListener("click",function(){{side.classList.remove("mobile-open");}});}});}}
 document.querySelectorAll(".nav-group").forEach(function(d){{d.addEventListener("toggle",function(){{if(!d.open)return;document.querySelectorAll(".nav-group").forEach(function(o){{if(o!==d)o.open=false;}});}});}});
 const c=document.getElementById("financeCategory"),a=document.getElementById("financeAnimal"),w=document.getElementById("statusWarning"),bulk=document.getElementById("bulkAnimalIds");if(c&&a&&w){{function x(){{const r=c.value==="Hayvan Satışı"||c.value==="Kesim Geliri";w.style.display=r?"block":"none";a.required=r&&!bulk;}}c.addEventListener("change",x);x();}}
}})();
function copyDeviceSimple(id,btn){{
 var el=document.getElementById(id);if(!el)return;var v=(el.innerText||el.textContent||'').trim();
 var ta=document.createElement('textarea');ta.value=v;ta.style.position='fixed';ta.style.opacity='0';document.body.appendChild(ta);ta.focus();ta.select();
 try{{document.execCommand('copy');var old=btn.innerText;btn.innerText='✓ Kopyalandı';setTimeout(function(){{btn.innerText=old}},1400)}}
 catch(e){{window.prompt('Cihaz Kimliğini kopyalayın:',v)}}
 document.body.removeChild(ta);
}}
function toggleAnimalFields(){{
 var type=document.getElementById('recordType');if(!type)return;var calf=type.value==='Buzağı';
 document.querySelectorAll('.calf-only').forEach(function(e){{e.style.display=calf?'block':'none';}});
 document.querySelectorAll('.adult-only').forEach(function(e){{e.style.display=calf?'none':'block';}});
 document.querySelectorAll('.female-pregnancy').forEach(function(e){{e.style.display=type.value==='Dişi'?'block':'none';}});
 var badge=document.getElementById('recordTypeBadge');if(badge)badge.textContent=type.options[type.selectedIndex].text;
 toggleEntryPregnancy();
}}
function toggleEntryPregnancy(){{
 var type=document.getElementById('recordType'),status=document.getElementById('entryPregnancyStatus'),mode=document.getElementById('pregnancyInfoMode');
 if(!type||!status)return;var pregnant=type.value==='Dişi'&&status.value==='Gebe',age=pregnant&&mode&&mode.value==='age';
 ['pregnancyInfoModeLabel','pregnancyEntryDateLabel','pregnancyEntryHint'].forEach(function(id){{var e=document.getElementById(id);if(e)e.style.display=pregnant?'block':'none';}});
 var kd=document.getElementById('knownInseminationLabel'),ag=document.getElementById('pregnancyAgeLabel');
 if(kd)kd.style.display=pregnant&&!age?'block':'none';if(ag)ag.style.display=age?'block':'none';
}}
function liveTableFilter(inputId,tableId,emptyId){{
 var input=document.getElementById(inputId),table=document.getElementById(tableId),empty=document.getElementById(emptyId);if(!input||!table)return;
 input.addEventListener('input',function(){{var q=(input.value||'').toLocaleLowerCase('tr-TR').trim(),visible=0;table.querySelectorAll('tbody tr.data-row').forEach(function(row){{var ok=!q||row.textContent.toLocaleLowerCase('tr-TR').includes(q);row.style.display=ok?'':'none';if(ok)visible++;}});if(empty)empty.style.display=visible?'none':'block';}});
}}
async function optimizePhotoFile(file,status,text,bar){{
 if(!file||!file.type||!file.type.startsWith('image/'))return file;
 const maxSide=1024,quality=.58;
 if(text)text.textContent='Fotoğraf hazırlanıyor…';if(status)status.classList.add('on');if(bar)bar.style.width='8%';
 try{{
   const bitmap=await createImageBitmap(file);
   let w=bitmap.width,h=bitmap.height,scale=Math.min(1,maxSide/Math.max(w,h));w=Math.max(1,Math.round(w*scale));h=Math.max(1,Math.round(h*scale));
   const canvas=document.createElement('canvas');canvas.width=w;canvas.height=h;const ctx=canvas.getContext('2d');ctx.drawImage(bitmap,0,0,w,h);if(bitmap.close)bitmap.close();
   if(bar)bar.style.width='22%';
   const blob=await new Promise(function(resolve){{canvas.toBlob(resolve,'image/webp',quality);}});
   if(!blob)throw new Error('Fotoğraf dönüştürülemedi');
   const base=(file.name||'hayvan').replace(/[.][^.]+$/,'');
   return new File([blob],base+'.webp',{{type:'image/webp',lastModified:Date.now()}});
 }}catch(err){{
   console.warn('Fotoğraf optimizasyonu atlandı:',err);return file;
 }}
}}
function bindSmartPhotoForms(){{
 document.querySelectorAll('form[data-smart-photo-form="1"]').forEach(function(form){{
   if(form.dataset.smartBound==='1')return;form.dataset.smartBound='1';
   const input=form.querySelector('input[type="file"][name="photo_file"]'),btn=form.querySelector('button[type="submit"],button:not([type])'),status=form.querySelector('[data-upload-status]'),text=form.querySelector('[data-upload-text]'),bar=form.querySelector('[data-upload-bar]');
   if(input)input.addEventListener('change',function(){{if(status){{status.classList.remove('error');status.classList.add('on');}}if(text){{const f=input.files&&input.files[0];text.textContent=f?'Seçildi: '+f.name+' · Kaydederken otomatik küçültülecek.':'Fotoğraf hazırlanıyor…';}}if(bar)bar.style.width='0';}});
   form.addEventListener('submit',async function(ev){{
     ev.preventDefault();if(form.dataset.submitting==='1')return;if(!form.checkValidity()){{form.reportValidity();return;}}
     form.dataset.submitting='1';if(btn){{btn.disabled=true;btn.dataset.oldText=btn.textContent;btn.textContent='Kaydediliyor…';}}
     if(status){{status.classList.remove('error');status.classList.add('on');}}if(bar)bar.style.width='3%';
     try{{
       let chosen=input&&input.files&&input.files[0];
       if(chosen){{const optimized=await optimizePhotoFile(chosen,status,text,bar);if(optimized!==chosen){{const dt=new DataTransfer();dt.items.add(optimized);input.files=dt.files;chosen=optimized;}}if(text)text.textContent='Fotoğraf yükleniyor…';}}
       else {{if(text)text.textContent='Değişiklikler kaydediliyor…';if(bar)bar.style.width='25%';}}
       const xhr=new XMLHttpRequest();xhr.open((form.method||'POST').toUpperCase(),form.action,true);
       xhr.upload.onprogress=function(e){{if(e.lengthComputable&&bar){{const pct=Math.max(25,Math.min(95,Math.round(e.loaded/e.total*100)));bar.style.width=pct+'%';if(text)text.textContent='Yükleniyor… %'+pct;}}}};
       xhr.onerror=function(){{fail('Sunucuya ulaşılamadı. Mobil internet / VPN bağlantısını kontrol edip tekrar deneyin.');}};
       xhr.ontimeout=function(){{fail('Yükleme zaman aşımına uğradı. Bağlantıyı kontrol edip tekrar deneyin.');}};xhr.timeout=120000;
       xhr.onload=function(){{if(xhr.status>=200&&xhr.status<400){{if(bar)bar.style.width='100%';if(text)text.textContent='Kaydedildi.';window.location.href=xhr.responseURL||form.action;}}else fail('Kaydetme başarısız oldu (HTTP '+xhr.status+'). Tekrar deneyin.');}};
       function fail(message){{form.dataset.submitting='0';if(btn){{btn.disabled=false;btn.textContent=btn.dataset.oldText||'Kaydet';}}if(status){{status.classList.add('error');status.classList.add('on');}}if(text)text.textContent=message;if(bar)bar.style.width='0';}}
       xhr.send(new FormData(form));
     }}catch(err){{form.dataset.submitting='0';if(btn){{btn.disabled=false;btn.textContent=btn.dataset.oldText||'Kaydet';}}if(status){{status.classList.add('error');status.classList.add('on');}}if(text)text.textContent='Fotoğraf hazırlanırken hata oluştu. Başka bir fotoğraf seçip tekrar deneyin.';if(bar)bar.style.width='0';console.error(err);}}
   }});
 }});
}}

function moneyRaw(v){{
  v=String(v||'').trim().replace(/[₺\\s]/g,'');
  if(!v)return '';
  const comma=v.indexOf(',');
  const whole=(comma>=0?v.slice(0,comma):v).replace(/\\D/g,'')||'0';
  const dec=comma>=0?v.slice(comma+1).replace(/\\D/g,'').slice(0,2):'';
  return whole+(comma>=0&&dec!==''?'.'+dec:'');
}}
function moneyGroupDigits(digits){{
  digits=String(digits||'').replace(/\\D/g,'').replace(/^0+(?=\\d)/,'');
  if(!digits)return '';
  return digits.replace(/\\B(?=(\\d{{3}})+(?!\\d))/g,'.');
}}
function bindSmartMoney(){{
  const names=new Set(['amount','cost','purchase_price','sale_price','sold_price','target_sale_price','daily_feed_cost','daily_care_cost']);
  document.querySelectorAll('input[name]').forEach(function(el){{
    if(!names.has(el.name)||el.dataset.moneyBound==='1')return;
    el.dataset.moneyBound='1';
    el.type='text';
    el.inputMode='numeric';
    el.autocomplete='off';
    el.classList.add('smart-money');

    // Existing database value -> Turkish display.
    if(el.value){{
      const initial=String(el.value).trim().replace(',', '.');
      const n=Number(initial);
      if(Number.isFinite(n)){{
        const fixed=Math.round(n*100)/100;
        const parts=String(fixed).split('.');
        const whole=moneyGroupDigits(parts[0]);
        const dec=(parts[1]||'').replace(/\\D/g,'').slice(0,2);
        el.value=whole+(dec?','+dec:'');
      }}
    }}

    // IMPORTANT: while typing, only digits are accepted as the lira amount.
    // The dots visible in the field are presentation only.
    el.addEventListener('input',function(){{
      const raw=el.value;
      const digits=raw.replace(/\\D/g,'');
      el.value=moneyGroupDigits(digits);
      try{{el.setSelectionRange(el.value.length,el.value.length);}}catch(e){{}}
    }});

    // Prevent punctuation typed by mobile/desktop keyboard from changing meaning.
    el.addEventListener('beforeinput',function(ev){{
      if(ev.data==='.'||ev.data===','){{
        ev.preventDefault();
      }}
    }});
  }});

  document.querySelectorAll('form').forEach(function(form){{
    if(form.dataset.moneySubmitBound==='1')return;
    form.dataset.moneySubmitBound='1';
    form.addEventListener('submit',function(){{
      form.querySelectorAll('input.smart-money').forEach(function(el){{
        el.value=String(el.value||'').replace(/\\D/g,'');
      }});
    }},true);
  }});
}}

function bindRationFloatingSummary(){{
  const panel=document.querySelector('.target-compare-sticky');
  if(!panel)return;
  const main=document.querySelector('.main');
  const holder=document.createElement('div');
  holder.className='target-compare-placeholder';
  panel.parentNode.insertBefore(holder,panel);
  const topbar=document.querySelector('.top');
  let anchorY=0, baseHeight=0, leftGap=0, rightGap=0;

  function captureBaseGeometry(){{
    if(panel.classList.contains('is-floating'))return;
    const r=panel.getBoundingClientRect();
    const mr=main?main.getBoundingClientRect():{{left:0,right:window.innerWidth}};
    leftGap=Math.max(0,Math.round(r.left-mr.left));
    rightGap=Math.max(0,Math.round(mr.right-r.right));
    anchorY=r.top+window.scrollY;
    baseHeight=Math.round(r.height);
  }}

  function applyHorizontalLock(){{
    const mr=main?main.getBoundingClientRect():{{left:0,right:window.innerWidth}};
    const left=Math.round(mr.left+leftGap);
    const right=Math.max(0,Math.round(window.innerWidth-mr.right+rightGap));
    panel.style.setProperty('left',left+'px','important');
    panel.style.setProperty('right',right+'px','important');
    panel.style.setProperty('width','auto','important');
    panel.style.setProperty('max-width','none','important');
  }}

  function clearFloating(){{
    panel.classList.remove('is-floating');
    ['left','right','top','width','height','max-width'].forEach(k=>panel.style.removeProperty(k));
    holder.classList.remove('active'); holder.style.height=''; holder.style.width='';
  }}

  function lockFloatingNow(){{
    if(!baseHeight)captureBaseGeometry();
    holder.style.height=baseHeight+'px'; holder.style.width='100%'; holder.classList.add('active');
    panel.classList.add('is-floating'); applyHorizontalLock();
    panel.style.setProperty('height',baseHeight+'px','important');
  }}

  function update(){{
    if(window.innerWidth<=1180){{clearFloating();captureBaseGeometry();return;}}
    const headerH=topbar?topbar.getBoundingClientRect().height:58;
    const trigger=window.scrollY+headerH+8>=anchorY;
    if(trigger){{
      if(!panel.classList.contains('is-floating'))lockFloatingNow(); else applyHorizontalLock();
      panel.style.setProperty('top',(headerH+6)+'px','important');
    }}else{{
      if(panel.classList.contains('is-floating'))clearFloating();
      captureBaseGeometry();
    }}
  }}

  captureBaseGeometry();update();
  window.addEventListener('scroll',update,{{passive:true}});
  window.addEventListener('resize',function(){{
    const wasFloating=panel.classList.contains('is-floating'); if(wasFloating)clearFloating();
    requestAnimationFrame(function(){{captureBaseGeometry();update();}});
  }});
}}
// DEV MOBILE UX: floating JS disabled; native sticky is used for ration summary.

document.addEventListener('DOMContentLoaded',bindRationFloatingSummary);

document.addEventListener('DOMContentLoaded',bindSmartMoney);

document.addEventListener('DOMContentLoaded',bindSmartPhotoForms);
</script>{DEV10_GLOBAL_FIX}</body></html>"""


PHOTO_MAX_SIDE=1024
PHOTO_WEBP_QUALITY=58
PHOTO_JPEG_QUALITY=56

def format_bytes(n):
    n=float(n or 0)
    for unit in ('B','KB','MB','GB'):
        if n<1024 or unit=='GB':
            return f'{n:.1f} {unit}' if unit!='B' else f'{int(n)} B'
        n/=1024

def get_setting(key,default=''):
    try:
        with db() as c:
            r=c.execute('select setting_value from settings where setting_key=?',(key,)).fetchone()
        return (r['setting_value'] if r else default) or default
    except Exception:
        return default

def set_setting_value(key,value):
    with db() as c:
        c.execute('insert or replace into settings(setting_key,setting_value) values(?,?)',(key,str(value or '')))

def configured_backup_dir():
    raw=(get_setting('backup_directory','') or '').strip()
    if not raw:return BACKUPS
    try:
        path=Path(os.path.expandvars(os.path.expanduser(raw)))
        path.mkdir(parents=True,exist_ok=True)
        probe=path/'.ciftlikpro_probe.tmp'
        probe.write_text('ok',encoding='utf-8')
        probe.unlink()
        return path
    except Exception:
        return BACKUPS

def uploads_storage_stats():
    total=count=0
    if UPLOADS.exists():
        for fp in UPLOADS.rglob('*'):
            if fp.is_file():
                try:
                    total+=fp.stat().st_size;count+=1
                except OSError:pass
    db_size=DB.stat().st_size if DB.exists() else 0
    return {'db_bytes':db_size,'upload_bytes':total,'upload_count':count,'total_bytes':db_size+total}

def optimized_webp_bytes(content):
    with Image.open(io.BytesIO(content)) as im:
        im=ImageOps.exif_transpose(im)
        if getattr(im,'is_animated',False):
            try:im.seek(0)
            except Exception:pass
        if im.mode not in ('RGB','RGBA'):im=im.convert('RGB')
        im.thumbnail((PHOTO_MAX_SIDE,PHOTO_MAX_SIDE),Image.Resampling.LANCZOS)
        out=io.BytesIO()
        im.save(out,'WEBP',quality=PHOTO_WEBP_QUALITY,method=6)
        return out.getvalue()

def save_optimized_upload(prefix,upload):
    content=upload.get('content') if isinstance(upload,dict) else None
    if not content:raise ValueError('Fotoğraf içeriği boş.')
    if len(content)>15*1024*1024:raise ValueError('Fotoğraf 15 MB sınırını aşıyor.')
    data=optimized_webp_bytes(content)
    name=f"{prefix}_{datetime.now().strftime('%Y%m%d_%H%M%S_%f')}.webp"
    UPLOADS.mkdir(parents=True,exist_ok=True)
    (UPLOADS/name).write_bytes(data)
    return name

def optimize_existing_uploads():
    if not UPLOADS.exists():return {'count':0,'saved':0,'before':0,'after':0}
    before=after=count=0
    for fp in list(UPLOADS.rglob('*')):
        if not fp.is_file() or fp.suffix.lower() not in ('.jpg','.jpeg','.png','.webp'):continue
        old_size=fp.stat().st_size;before+=old_size
        tmp=fp.with_name(fp.name+'.optimize-tmp')
        try:
            with Image.open(fp) as im:
                im=ImageOps.exif_transpose(im)
                im.thumbnail((PHOTO_MAX_SIDE,PHOTO_MAX_SIDE),Image.Resampling.LANCZOS)
                ext=fp.suffix.lower()
                if ext in ('.jpg','.jpeg'):
                    if im.mode!='RGB':im=im.convert('RGB')
                    im.save(tmp,'JPEG',quality=PHOTO_JPEG_QUALITY,optimize=True,progressive=True)
                elif ext=='.webp':
                    if im.mode not in ('RGB','RGBA'):im=im.convert('RGB')
                    im.save(tmp,'WEBP',quality=PHOTO_WEBP_QUALITY,method=6)
                else:
                    if im.mode not in ('RGB','RGBA'):im=im.convert('RGBA')
                    im.save(tmp,'PNG',optimize=True,compress_level=9)
            new_size=tmp.stat().st_size
            if new_size<old_size:
                os.replace(tmp,fp);after+=new_size;count+=1
            else:
                tmp.unlink(missing_ok=True);after+=old_size
        except Exception:
            tmp.unlink(missing_ok=True);after+=old_size
    return {'count':count,'saved':max(0,before-after),'before':before,'after':after}

def clean_text(v):
    if v is None:return ''
    t=str(v).strip()
    return '' if t.lower() in ('undefined','null','none') else t

def create_backup(label='manuel'):
    target_dir=configured_backup_dir()
    target_dir.mkdir(parents=True,exist_ok=True)
    ts=datetime.now().strftime('%Y%m%d_%H%M%S')
    name=f'CiftlikPro_Backup_{label}_{ts}.zip'
    dst=target_dir/name
    temp_db=DATA_ROOT/f'.snapshot_{ts}.db'
    with db() as src, sqlite3.connect(temp_db) as out:src.backup(out)
    manifest={'product':APP_NAME,'version':APP_VERSION,'created_at':datetime.now().isoformat(timespec='seconds'),'database':'ciftlik.db','includes_uploads':True,'label':label,'backup_directory':str(target_dir)}
    try:
        with zipfile.ZipFile(dst,'w',zipfile.ZIP_DEFLATED) as z:
            z.write(temp_db,'ciftlik.db')
            z.writestr('manifest.json',json.dumps(manifest,ensure_ascii=False,indent=2))
            if UPLOADS.exists():
                for fp in UPLOADS.rglob('*'):
                    if fp.is_file():z.write(fp,'uploads/'+str(fp.relative_to(UPLOADS)).replace('\\','/'))
    finally:
        gc.collect()
        for attempt in range(12):
            try:
                if temp_db.exists():temp_db.unlink()
                break
            except PermissionError:
                if attempt<11:time.sleep(.25)
            except FileNotFoundError:break
    with db() as c:c.execute('insert into backups(filename,created_at,size_bytes) values(?,?,?)',(name,datetime.now().strftime('%Y-%m-%d %H:%M:%S'),dst.stat().st_size))
    return name

def validate_backup_zip(zip_path):
    with zipfile.ZipFile(zip_path,'r') as z:
        names=set(z.namelist())
        if 'ciftlik.db' not in names:return False,'Yedekte ciftlik.db bulunamadı.',None
        for name in names:
            p=Path(name)
            if name.startswith('/') or '..' in p.parts:return False,'Güvensiz ZIP yolu tespit edildi.',None
        manifest={}
        if 'manifest.json' in names:
            try:manifest=json.loads(z.read('manifest.json').decode('utf-8'))
            except Exception:return False,'manifest.json okunamadı.',None
        with tempfile.TemporaryDirectory() as td:
            db_copy=Path(td)/'ciftlik.db';db_copy.write_bytes(z.read('ciftlik.db'))
            try:
                with sqlite3.connect(db_copy) as c:
                    if c.execute('pragma quick_check').fetchone()[0]!='ok':return False,'Yedek veritabanı bozuk.',None
            except Exception as exc:return False,'Yedek veritabanı açılamadı: '+str(exc),None
    return True,'Yedek geçerli.',manifest

def restore_backup_zip(zip_path):
    ok,message,manifest=validate_backup_zip(zip_path)
    if not ok:raise ValueError(message)
    emergency=create_backup('EmergencyBeforeRestore')
    with tempfile.TemporaryDirectory() as td:
        target=Path(td)
        with zipfile.ZipFile(zip_path,'r') as z:z.extractall(target)
        staged=DB.with_suffix('.restore.tmp');shutil.copy2(target/'ciftlik.db',staged);os.replace(staged,DB)
        restored=target/'uploads'
        if restored.exists():
            shutil.rmtree(UPLOADS,ignore_errors=True);shutil.copytree(restored,UPLOADS)
        else:UPLOADS.mkdir(exist_ok=True)
    init_db();return emergency,manifest

def daily_backup():
    BACKUPS.mkdir(exist_ok=True);today=date.today().strftime('%Y%m%d')
    if not any(BACKUPS.glob(f'CiftlikPro_Backup_otomatik_{today}_*.zip')):create_backup('otomatik')
    files=sorted(BACKUPS.glob('CiftlikPro_Backup_otomatik_*.zip'),key=lambda x:x.stat().st_mtime,reverse=True)
    for fp in files[30:]:
        try:fp.unlink()
        except:pass


def export_payload():
    with db() as c:
        return {'format':'ciftlik-suru-takip-v06','exportDate':datetime.now().isoformat(),'animals':[dict(r) for r in c.execute('select * from animals order by id')],'inseminations':[dict(r) for r in c.execute('select * from inseminations order by animal_id,attempt')],'estrus_records':[dict(r) for r in c.execute('select * from estrus_records order by animal_id,estrus_date')],'calves':[dict(r) for r in c.execute('select * from calves order by id')],'health':[dict(r) for r in c.execute('select * from health order by id')],'finance':[dict(r) for r in c.execute('select * from finance order by id')],'weights':[dict(r) for r in c.execute('select * from weights order by id')],'milk':[dict(r) for r in c.execute('select * from milk order by id')]}

def import_payload(payload,strategy='skip'):
    stats={'animals':0,'animals_updated':0,'inseminations':0,'calves':0,'finance':0,'health':0,'skipped':0,'errors':[]}
    animals=payload.get('herdData',payload.get('animals',[])) if isinstance(payload,dict) else []
    calf_data=payload.get('calfData',payload.get('calves',[])) if isinstance(payload,dict) else []
    finance_data=payload.get('financeData',payload.get('finance',[])) if isinstance(payload,dict) else []
    native_insems=payload.get('inseminations',[]) if isinstance(payload,dict) else []
    health_data=payload.get('health',[]) if isinstance(payload,dict) else []
    with db() as c:
        for a in animals:
            try:
                tag=clean_text(a.get('tagId',a.get('tag')))
                if not tag:stats['skipped']+=1;continue
                nickname=clean_text(a.get('description',a.get('nickname')))
                gender=clean_text(a.get('gender')) or 'Dişi'
                existing=c.execute('select id from animals where tag=?',(tag,)).fetchone()
                if existing:
                    aid=existing['id']
                    if strategy=='update':
                        c.execute('update animals set nickname=coalesce(nullif(?,''),nickname),gender=coalesce(nullif(?,''),gender),breed=coalesce(nullif(?,''),breed),birth_date=coalesce(nullif(?,''),birth_date),notes=coalesce(nullif(?,''),notes) where id=?',(nickname,gender,clean_text(a.get('breed')),clean_text(a.get('birth_date',a.get('birthDate'))),clean_text(a.get('notes')),aid));stats['animals_updated']+=1
                    else:stats['skipped']+=1
                else:
                    aid=c.execute('insert into animals(tag,nickname,gender,breed,birth_date,notes) values(?,?,?,?,?,?)',(tag,nickname,gender,clean_text(a.get('breed')),clean_text(a.get('birth_date',a.get('birthDate'))),clean_text(a.get('notes')))).lastrowid;stats['animals']+=1
                dates=[a.get('toh1'),a.get('toh2'),a.get('toh3')]
                latest=next((clean_text(x) for x in reversed(dates) if clean_text(x)),'')
                for i,d in enumerate(dates,1):
                    d=clean_text(d)
                    if not d:continue
                    result='Pozitif' if bool(a.get('isPregnant')) and d==latest else 'Belirsiz'
                    due=(date.fromisoformat(d)+timedelta(days=280)).isoformat() if result=='Pozitif' else ''
                    c.execute('insert or replace into inseminations(animal_id,attempt,insemination_date,pregnancy_result,due_date) values(?,?,?,?,?)',(aid,i,d,result,due));stats['inseminations']+=1
            except Exception as e:stats['errors'].append(f'Hayvan: {e}')
        for ins in native_insems:
            try:
                aid=ins.get('animal_id')
                if not aid and clean_text(ins.get('tag')):
                    r=c.execute('select id from animals where tag=?',(clean_text(ins.get('tag')),)).fetchone();aid=r['id'] if r else None
                if not aid:stats['skipped']+=1;continue
                c.execute('insert or replace into inseminations(animal_id,attempt,insemination_date,pregnancy_result,due_date) values(?,?,?,?,?)',(aid,ins.get('attempt',1),ins.get('insemination_date',''),ins.get('pregnancy_result','Belirsiz'),ins.get('due_date','')));stats['inseminations']+=1
            except Exception as e:stats['errors'].append(f'Tohumlama: {e}')
        for calf in calf_data:
            try:
                tag=clean_text(calf.get('tagId',calf.get('tag')));mtag=clean_text(calf.get('motherTagId',calf.get('mother_tag')))
                mother=c.execute("select id from animals where tag=? and gender='Dişi'",(mtag,)).fetchone()
                if not tag or not mother or c.execute('select 1 from calves where tag=?',(tag,)).fetchone():stats['skipped']+=1;continue
                c.execute('insert into calves(tag,mother_id,father_tag,birth_date,gender,notes) values(?,?,?,?,?,?)',(tag,mother['id'],clean_text(calf.get('fatherTagId',calf.get('father_tag'))),clean_text(calf.get('birthDate',calf.get('birth_date'))),clean_text(calf.get('gender')),clean_text(calf.get('notes'))));stats['calves']+=1
            except Exception as e:stats['errors'].append(f'Buzağı: {e}')
        for f in finance_data:
            try:
                typ='Gelir' if clean_text(f.get('type',f.get('tx_type'))).lower()=='gelir' else 'Gider';tag=clean_text(f.get('tagId',f.get('tag')));aid=None
                if tag and tag!='-':
                    r=c.execute('select id from animals where tag=?',(tag,)).fetchone();aid=r['id'] if r else None
                c.execute('insert into finance(tx_date,tx_type,category,amount,description,payment_method,animal_id,created_at) values(?,?,?,?,?,?,?,?)',(clean_text(f.get('date',f.get('tx_date'))),typ,clean_text(f.get('category')) or 'Diğer',float(f.get('amount') or 0),clean_text(f.get('description')),clean_text(f.get('payment_method')) or 'Belirtilmedi',aid,datetime.now().isoformat()));stats['finance']+=1
            except Exception as e:stats['errors'].append(f'Finans: {e}')
        for x in health_data:
            try:c.execute('insert into health(animal_id,kind,product,applied_date,next_date,cost,notes) values(?,?,?,?,?,?,?)',(x.get('animal_id'),x.get('kind'),x.get('product'),x.get('applied_date'),x.get('next_date'),float(x.get('cost') or 0),x.get('notes')));stats['health']+=1
            except Exception as e:stats['errors'].append(f'Sağlık: {e}')
    return stats

class App(BaseHTTPRequestHandler):
    def log_message(self,*a): pass
    def parse_cookie(self):
        c=cookies.SimpleCookie(self.headers.get('Cookie')); return c.get('sid').value if c.get('sid') else None
    def user(self): return SESSIONS.get(self.parse_cookie())
    def send_html(self,s,status=200,headers=None):
        b=s.encode('utf-8')
        try:
            self.send_response(status); self.send_header('Content-Type','text/html; charset=utf-8'); self.send_header('Content-Length',str(len(b)))
            for k,v in (headers or []):self.send_header(k,v)
            self.end_headers(); self.wfile.write(b)
        except (BrokenPipeError, ConnectionAbortedError, ConnectionResetError):
            # Tarayıcı yenileme/geri/sekme kapatma sırasında isteği iptal edebilir.
            # Bu durum uygulama hatası değildir; DEV konsoluna traceback basma.
            return
        except OSError as e:
            if getattr(e,'winerror',None) in (10053,10054): return
            raise
    def redirect(self,url,msg=''):
        # Mesaj query string'e, varsa #anchor'dan ÖNCE eklenmeli.
        # Aksi halde /rations?id=1%23... gibi bozuk URL oluşup sayfayı düşürebilir.
        base, sep, frag = url.partition('#')
        if msg:
            base += ('&' if '?' in base else '?')+'msg='+urllib.parse.quote(msg)
        target = base + (('#'+frag) if sep else '')
        self.send_response(303);self.send_header('Location',target);self.end_headers()
    def form(self):
        n=int(self.headers.get('Content-Length','0')); return {k:v[0] for k,v in urllib.parse.parse_qs(self.rfile.read(n).decode()).items()}
    def post_data(self):
        ctype=self.headers.get('Content-Type','')
        if ctype.startswith('multipart/form-data'):
            n=int(self.headers.get('Content-Length','0')); body=self.rfile.read(n)
            raw=(f'Content-Type: {ctype}\r\nMIME-Version: 1.0\r\n\r\n').encode()+body
            msg=BytesParser(policy=default).parsebytes(raw); data={}
            for part in msg.iter_parts():
                name=part.get_param('name',header='content-disposition'); filename=part.get_filename(); content=part.get_payload(decode=True) or b''
                if filename:data[name]={'filename':filename,'content':content}
                else:data[name]=content.decode(part.get_content_charset() or 'utf-8')
            return data
        return self.form()
    def require(self):
        if not self.user(): self.redirect('/login'); return False
        return True
    def is_admin(self):return bool(self.user() and self.user().get('role')=='admin')
    def require_admin(self):
        if not self.require():return False
        if not self.is_admin():self.redirect('/','Bu bölüm yalnızca yöneticilere açıktır.');return False
        return True
    def client_ip(self):return self.client_address[0] if self.client_address else ''
    def do_GET(self):
        ensure_archive_schema()
        p=urllib.parse.urlparse(self.path); path=p.path; q=urllib.parse.parse_qs(p.query); msg=q.get('msg',[''])[0]
        lic_ok,lic_payload,lic_msg=license_status()
        if path=='/license':
            payload=lic_payload or {}
            flash_html=('<div class="activation-alert">'+h(msg)+'</div>') if msg else ''
            status_class='license-ok' if lic_ok else 'license-bad'
            status_text='🟢 Lisans Aktif' if lic_ok else '🔴 Lisans Gerekli'
            active_details=''
            if lic_ok:
                exp=payload.get('expires_on') or 'Süresiz'
                if exp!='Süresiz': exp=fmt_date(exp)
                active_details='<div class="activation-details"><div><span>Lisans Sahibi</span><b>'+h(payload.get('licensee') or '-')+'</b></div><div><span>Lisans Türü</span><b>'+h(payload.get('license_type') or '-')+'</b></div><div><span>Geçerlilik</span><b>'+h(exp)+'</b></div></div><a class="activation-continue" href="/login">ÇiftlikPro\'ya Devam →</a>'
            upload_section='' if lic_ok else '''<form class="activation-form" method="post" action="/license-key-activate">
              <label class="device-label" for="license_key">Lisans Anahtarı</label>
              <textarea id="license_key" name="license_key" class="license-key-input" rows="4" placeholder="CFP-XXXXX-XXXXX-XXXXX-..." required></textarea>
              <button class="activation-submit" type="submit">🔓 ÇiftlikPro'yu Etkinleştir</button></form>
              <details class="file-fallback"><summary>Gelişmiş: .license dosyasıyla etkinleştir</summary>
              <form class="activation-form fallback-form" method="post" action="/license-activate" enctype="multipart/form-data">
              <label class="activation-upload" for="license_file"><span class="upload-icon">📄</span><span><b>Lisans dosyanızı seçin</b><small>Yedek aktivasyon yöntemi</small></span><span class="upload-button">Dosya Seç</span></label>
              <input id="license_file" type="file" name="license_file" accept=".license,application/json" required hidden><div id="selected-license" class="selected-license">Henüz dosya seçilmedi</div>
              <button class="activation-submit secondary-submit" type="submit">Dosyayla Etkinleştir</button></form></details>'''
            html='''<!doctype html><html lang="tr"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>ÇiftlikPro Aktivasyon</title>
<style>
*{box-sizing:border-box}body{margin:0;font-family:Inter,"Segoe UI",Arial,sans-serif;background:linear-gradient(145deg,#eef5f0 0%,#f8faf9 55%,#e7f1eb 100%);color:#173426;min-height:100vh}
.activation-page{min-height:100vh;display:flex;align-items:center;justify-content:center;padding:28px 16px}.activation-card{width:min(760px,100%);background:#fff;border:1px solid #dce9e0;border-radius:28px;box-shadow:0 24px 70px rgba(17,73,43,.14);overflow:hidden}
.activation-head{background:linear-gradient(135deg,#105b35,#18824a);padding:30px 34px;color:#fff}.activation-brand{display:flex;align-items:center;gap:11px;font-weight:800;font-size:19px}.activation-brand span:first-child{font-size:27px}
.activation-head h1{font-size:34px;line-height:1.1;margin:28px 0 8px}.activation-head p{margin:0;color:#dcefe4;font-size:15px}.activation-body{padding:32px 34px 34px}
.activation-status{display:inline-flex;padding:8px 13px;border-radius:999px;font-weight:800;font-size:14px;margin-bottom:22px}.license-ok{background:#e5f6eb;color:#14723d}.license-bad{background:#fff0ed;color:#ad3529}
.activation-alert{padding:13px 15px;border-radius:12px;background:#fff4dc;color:#7c5600;border:1px solid #f2d797;margin-bottom:18px}.device-label{font-size:13px;color:#698075;font-weight:700;margin-bottom:8px;text-transform:uppercase;letter-spacing:.5px}
.device-row{display:flex;gap:10px}.device-code{flex:1;background:#f0f7f2;border:1px solid #cfe2d5;border-radius:14px;padding:15px 17px;font:800 18px Consolas,monospace;letter-spacing:.7px;color:#145c36;overflow-wrap:anywhere}
.copy-btn{border:0;border-radius:14px;background:#e6f2e9;color:#176a3c;font-weight:800;padding:0 18px;cursor:pointer}.activation-help{color:#65786d;font-size:14px;line-height:1.55;margin:12px 0 24px}
.activation-upload{display:flex;align-items:center;gap:14px;border:2px dashed #b8d4c0;background:#f8fbf9;border-radius:18px;padding:18px;cursor:pointer}.upload-icon{font-size:28px}.activation-upload b,.activation-upload small{display:block}.activation-upload small{color:#77887f;margin-top:3px}.upload-button{margin-left:auto;background:#e4f1e8;color:#16693b;padding:9px 12px;border-radius:10px;font-weight:800}
.license-key-input{width:100%;resize:vertical;min-height:105px;border:1px solid #bfd5c6;border-radius:14px;background:#f8fbf9;padding:15px;font:700 14px Consolas,monospace;color:#174d31;outline:none;margin:0 0 14px}.selected-license{font-size:13px;color:#718177;margin:8px 2px 14px}.file-fallback{margin-top:18px;border-top:1px solid #e4ece7;padding-top:16px}.file-fallback summary{cursor:pointer;font-weight:700;color:#65786d}.fallback-form{margin-top:14px}.secondary-submit{background:#557267}.activation-submit,.activation-continue{width:100%;border:0;border-radius:14px;background:linear-gradient(135deg,#14763f,#199151);color:#fff;font-weight:800;font-size:16px;padding:15px 18px;cursor:pointer;text-decoration:none;display:block;text-align:center}
.activation-details{display:grid;grid-template-columns:repeat(3,1fr);gap:10px;margin:22px 0}.activation-details div{background:#f3f8f5;border-radius:13px;padding:13px}.activation-details span{display:block;color:#708278;font-size:12px}.activation-foot{text-align:center;color:#829087;font-size:12px;margin-top:22px}
@media(max-width:600px){.activation-page{padding:0}.activation-card{min-height:100vh;border-radius:0}.activation-head{padding:26px 22px}.activation-head h1{font-size:29px}.activation-body{padding:26px 22px}.device-row{display:block}.copy-btn{width:100%;padding:12px;margin-top:8px}.activation-details{grid-template-columns:1fr}}

/* HOTFIX4 UX: hedef kartları çalışma masasının üstünde, iki satır ve okunaklı sabit kokpit */
body:has(.workbench-shell) .workbench-shell{{display:block!important;padding:14px!important;overflow:visible!important}}
body:has(.workbench-shell) .workbench-shell>.target-workspace{{display:block!important}}
body:has(.workbench-shell) .target-controlbar{{display:grid!important;grid-template-columns:auto minmax(0,1fr)!important;gap:12px!important;margin:8px 0!important;padding:8px 10px!important}}
body:has(.workbench-shell) .target-compare-sticky{{position:sticky!important;top:62px!important;height:auto!important;max-height:none!important;overflow:visible!important;margin:8px 0 10px!important;padding:9px 10px!important;background:rgba(248,251,249,.985)!important;backdrop-filter:blur(8px)!important;z-index:40!important;box-shadow:0 8px 22px rgba(20,65,40,.14)!important}}
body:has(.workbench-shell) .target-compare-title{{height:auto!important;margin:0 2px 7px!important;font-size:13px!important}}
body:has(.workbench-shell) .target-compare-title>b{{font-size:14px!important}}
body:has(.workbench-shell) .target-compare-title span{{font-size:10px!important;max-width:none!important}}
body:has(.workbench-shell) .nutri-mini-grid{{position:static!important;display:grid!important;grid-template-columns:repeat(5,minmax(0,1fr))!important;grid-template-rows:none!important;grid-auto-rows:112px!important;height:auto!important;gap:7px!important;padding:0!important;margin:0!important;border:0!important;background:transparent!important;box-shadow:none!important;overflow:visible!important}}
body:has(.workbench-shell) .nutri-mini{{min-height:112px!important;height:112px!important;padding:7px 8px!important;display:grid!important;grid-template-rows:18px 1fr 34px!important;border-radius:10px!important;overflow:hidden!important}}
body:has(.workbench-shell) .nutri-card-title{{font-size:13px!important;line-height:18px!important;font-weight:900!important;margin:0!important}}
body:has(.workbench-shell) .nutri-compare-body{{display:grid!important;grid-template-columns:1fr 1fr!important;align-items:center!important;min-height:0!important}}
body:has(.workbench-shell) .nutri-side{{padding:2px 5px!important;min-width:0!important;text-align:center!important}}
body:has(.workbench-shell) .nutri-side span{{font-size:9px!important;line-height:1!important;font-weight:800!important}}
body:has(.workbench-shell) .nutri-side b{{font-size:18px!important;line-height:1.05!important;margin-top:4px!important;white-space:normal!important;overflow-wrap:anywhere!important}}
body:has(.workbench-shell) .nutri-card-footer{{min-height:34px!important;height:34px!important;padding:4px 3px 2px!important;display:grid!important;grid-template-columns:1fr!important;align-content:center!important;justify-items:center!important;gap:1px!important}}
body:has(.workbench-shell) .nutri-card-footer em{{font-size:12px!important;line-height:1!important;font-weight:900!important;white-space:nowrap!important}}
body:has(.workbench-shell) .nutri-card-footer .nutri-diff{{font-size:10px!important;line-height:1.05!important;margin:1px 0 0!important;white-space:nowrap!important;font-weight:800!important}}
body:has(.workbench-shell) #target-mini-cost{{grid-column:auto!important}}
body:has(.workbench-shell) #ration-workbench{{margin-top:0!important}}
@media(max-width:1350px){{body:has(.workbench-shell) .nutri-mini-grid{{grid-template-columns:repeat(4,minmax(0,1fr))!important}}}}
@media(max-width:1080px){{body:has(.workbench-shell) .nutri-mini-grid{{grid-template-columns:repeat(3,minmax(0,1fr))!important}}body:has(.workbench-shell) .target-compare-sticky{{position:static!important}}}}
@media(max-width:820px){{body:has(.workbench-shell) .nutri-mini-grid{{grid-template-columns:repeat(2,minmax(0,1fr))!important;grid-auto-rows:116px!important}}body:has(.workbench-shell) .nutri-mini{{height:116px!important;min-height:116px!important}}}}
</style></head><body><main class="activation-page"><section class="activation-card"><header class="activation-head"><div class="activation-brand"><span>🐄</span><span>ÇiftlikPro Enterprise</span></div><h1>🔐 Lisans Aktivasyonu</h1><p>Bu bilgisayarı güvenli bir ÇiftlikPro lisansıyla etkinleştirin.</p></header><div class="activation-body">'''
            html+=flash_html+'<div class="activation-status '+status_class+'">'+status_text+'</div><div class="device-label">Bu Bilgisayarın Cihaz Kimliği</div><div class="device-row"><div id="device-code" class="device-code">'+h(device_id())+'</div><button class="copy-btn" type="button" onclick="copyDevice()">Kopyala</button></div><p class="activation-help">Bu cihaz kodunu lisans yöneticisine iletin. Size verilen <b>CFP lisans anahtarını</b> aşağıdaki alana yapıştırın.</p>'+active_details+upload_section
            html+='''<div class="activation-foot">ÇiftlikPro Enterprise · Cihaza bağlı dijital lisans koruması</div></div></section></main>
<script>
function copyDevice(){var t=document.getElementById("device-code").innerText;navigator.clipboard.writeText(t).then(function(){var b=document.querySelector(".copy-btn");b.innerText="Kopyalandı ✓";setTimeout(function(){b.innerText="Kopyala"},1600)})}
var f=document.getElementById("license_file");if(f){f.addEventListener("change",function(){document.getElementById("selected-license").innerText=f.files.length?"Seçilen dosya: "+f.files[0].name:"Henüz dosya seçilmedi"})}
</script></body></html>'''
            return self.send_html(html)
        if not lic_ok and path not in ('/license','/license-activate','/license-key-activate'):
            return self.redirect('/license',lic_msg)
        if path=='/login':
            login_msg=('<div class="flash err">'+h(msg)+'</div>') if msg else ''
            return self.send_html(f'''<!doctype html><html lang="tr"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>ÇiftlikPro Enterprise · Giriş</title><style>{CSS}
            body{{min-height:100vh;margin:0;background:radial-gradient(circle at 15% 15%,#2f8b5b 0,#153d2a 38%,#0c261a 100%);display:grid;place-items:center;padding:24px;box-sizing:border-box}}
            .login-shell{{width:min(980px,100%);display:grid;grid-template-columns:1.05fr .95fr;background:#fff;border-radius:30px;overflow:hidden;box-shadow:0 30px 90px #06170f70}}
            .login-visual{{padding:54px;background:linear-gradient(145deg,#174a31,#23784b);color:#fff;display:flex;flex-direction:column;justify-content:space-between;min-height:510px}}
            .login-logo{{font-size:36px;font-weight:950;letter-spacing:-1px}}.login-visual h1{{font-size:42px;line-height:1.05;margin:24px 0 14px}}.login-visual p{{color:#dcece2;font-size:17px;line-height:1.6;max-width:430px}}.login-pills{{display:flex;flex-wrap:wrap;gap:8px}}.login-pills span{{padding:8px 11px;border-radius:999px;background:#ffffff17;border:1px solid #ffffff24;font-size:13px;font-weight:800}}
            .login-panel{{padding:54px 48px;display:flex;flex-direction:column;justify-content:center}}.login-panel h2{{font-size:30px;margin:0 0 8px;color:#173c29}}.login-sub{{color:#718276;margin-bottom:26px}}.login-version{{display:inline-flex;align-self:flex-start;background:#eef6f0;color:#557064;border-radius:999px;padding:7px 11px;font-weight:800;font-size:12px;margin-bottom:22px}}.login-form label{{display:block;font-size:13px;font-weight:850;margin:12px 0 6px;color:#263b30}}.login-form input{{width:100%;box-sizing:border-box;font-size:16px;padding:14px;border:1px solid #cfdcd2;border-radius:12px;background:#fbfdfb}}.password-wrap{{position:relative}}.password-wrap input{{padding-right:96px}}.password-toggle{{position:absolute;right:7px;top:50%;transform:translateY(-50%);border:0;background:#edf5ef;color:#176b3a;border-radius:9px;padding:8px 10px;font-weight:800;cursor:pointer}}.login-submit{{width:100%;margin-top:20px;min-height:50px;font-size:16px;border-radius:12px}}.forgot{{display:block;text-align:center;margin-top:16px;color:#176b3a;font-weight:800;text-decoration:none}}
            @media(max-width:760px){{body{{padding:8px;min-height:100dvh;align-items:start}}.login-shell{{grid-template-columns:1fr;border-radius:22px;margin:0 auto}}.login-visual{{min-height:auto;padding:20px 22px 18px}}.login-logo{{font-size:31px}}.login-visual h1{{font-size:28px;line-height:1.04;margin:12px 0 8px}}.login-visual p{{font-size:13.5px;line-height:1.45;margin:0 0 13px}}.login-pills{{gap:6px}}.login-pills span{{padding:6px 9px;font-size:11.5px}}.login-panel{{padding:22px 20px 24px}}.login-version{{margin-bottom:14px;padding:6px 9px;font-size:11px}}.login-panel h2{{font-size:27px}}.login-sub{{margin-bottom:17px;font-size:14px}}.login-form label{{margin:9px 0 5px}}.login-form input{{padding:12px;min-height:46px}}.password-wrap input{{padding-right:92px}}.password-toggle{{padding:7px 9px}}.login-submit{{margin-top:16px;min-height:48px}}.forgot{{margin-top:12px;padding-bottom:max(2px,env(safe-area-inset-bottom))}}}}
            
/* HOTFIX4 UX: hedef kartları çalışma masasının üstünde, iki satır ve okunaklı sabit kokpit */
body:has(.workbench-shell) .workbench-shell{{display:block!important;padding:14px!important;overflow:visible!important}}
body:has(.workbench-shell) .workbench-shell>.target-workspace{{display:block!important}}
body:has(.workbench-shell) .target-controlbar{{display:grid!important;grid-template-columns:auto minmax(0,1fr)!important;gap:12px!important;margin:8px 0!important;padding:8px 10px!important}}
body:has(.workbench-shell) .target-compare-sticky{{position:sticky!important;top:62px!important;height:auto!important;max-height:none!important;overflow:visible!important;margin:8px 0 10px!important;padding:9px 10px!important;background:rgba(248,251,249,.985)!important;backdrop-filter:blur(8px)!important;z-index:40!important;box-shadow:0 8px 22px rgba(20,65,40,.14)!important}}
body:has(.workbench-shell) .target-compare-title{{height:auto!important;margin:0 2px 7px!important;font-size:13px!important}}
body:has(.workbench-shell) .target-compare-title>b{{font-size:14px!important}}
body:has(.workbench-shell) .target-compare-title span{{font-size:10px!important;max-width:none!important}}
body:has(.workbench-shell) .nutri-mini-grid{{position:static!important;display:grid!important;grid-template-columns:repeat(5,minmax(0,1fr))!important;grid-template-rows:none!important;grid-auto-rows:112px!important;height:auto!important;gap:7px!important;padding:0!important;margin:0!important;border:0!important;background:transparent!important;box-shadow:none!important;overflow:visible!important}}
body:has(.workbench-shell) .nutri-mini{{min-height:112px!important;height:112px!important;padding:7px 8px!important;display:grid!important;grid-template-rows:18px 1fr 34px!important;border-radius:10px!important;overflow:hidden!important}}
body:has(.workbench-shell) .nutri-card-title{{font-size:13px!important;line-height:18px!important;font-weight:900!important;margin:0!important}}
body:has(.workbench-shell) .nutri-compare-body{{display:grid!important;grid-template-columns:1fr 1fr!important;align-items:center!important;min-height:0!important}}
body:has(.workbench-shell) .nutri-side{{padding:2px 5px!important;min-width:0!important;text-align:center!important}}
body:has(.workbench-shell) .nutri-side span{{font-size:9px!important;line-height:1!important;font-weight:800!important}}
body:has(.workbench-shell) .nutri-side b{{font-size:18px!important;line-height:1.05!important;margin-top:4px!important;white-space:normal!important;overflow-wrap:anywhere!important}}
body:has(.workbench-shell) .nutri-card-footer{{min-height:34px!important;height:34px!important;padding:4px 3px 2px!important;display:grid!important;grid-template-columns:1fr!important;align-content:center!important;justify-items:center!important;gap:1px!important}}
body:has(.workbench-shell) .nutri-card-footer em{{font-size:12px!important;line-height:1!important;font-weight:900!important;white-space:nowrap!important}}
body:has(.workbench-shell) .nutri-card-footer .nutri-diff{{font-size:10px!important;line-height:1.05!important;margin:1px 0 0!important;white-space:nowrap!important;font-weight:800!important}}
body:has(.workbench-shell) #target-mini-cost{{grid-column:auto!important}}
body:has(.workbench-shell) #ration-workbench{{margin-top:0!important}}
@media(max-width:1350px){{body:has(.workbench-shell) .nutri-mini-grid{{grid-template-columns:repeat(4,minmax(0,1fr))!important}}}}
@media(max-width:1080px){{body:has(.workbench-shell) .nutri-mini-grid{{grid-template-columns:repeat(3,minmax(0,1fr))!important}}body:has(.workbench-shell) .target-compare-sticky{{position:static!important}}}}
@media(max-width:820px){{body:has(.workbench-shell) .nutri-mini-grid{{grid-template-columns:repeat(2,minmax(0,1fr))!important;grid-auto-rows:116px!important}}body:has(.workbench-shell) .nutri-mini{{height:116px!important;min-height:116px!important}}}}
</style></head><body><div class="login-shell"><section class="login-visual"><div><div class="login-logo">🐄 ÇiftlikPro</div><h1>Çiftliğinizin kontrol merkezi.</h1><p>Sürü, üreme, sağlık, besi ve finans yönetimini tek yerde güvenle takip edin.</p></div><div class="login-pills"><span>🐄 Sürü Yönetimi</span><span>💉 Sağlık</span><span>🌸 Üreme</span><span>📈 Besi</span><span>₺ Finans</span></div></section><section class="login-panel"><div class="login-version">ÇiftlikPro Enterprise • V{APP_VERSION}</div><h2>Hoş geldiniz</h2><div class="login-sub">Devam etmek için hesabınızla giriş yapın.</div>{login_msg}<form method="post" class="login-form"><label>Kullanıcı adı</label><input name="username" autocomplete="username" required autofocus><label>Şifre</label><div class="password-wrap"><input id="loginPassword" type="password" name="password" autocomplete="current-password" required><button class="password-toggle" type="button" onclick="toggleLoginPassword(this)">👁 Göster</button></div><button class="btn login-submit">Giriş Yap →</button><a class="forgot" href="/forgot-password">🔑 Şifremi Unuttum</a></form></section></div><script>function toggleLoginPassword(btn){{var p=document.getElementById('loginPassword');var show=p.type==='password';p.type=show?'text':'password';btn.textContent=show?'🙈 Gizle':'👁 Göster';}}</script></body></html>''')
        if path=='/forgot-password':
            step=(q.get('step',['request'])[0] or 'request');rid=(q.get('id',[''])[0] or '');token=(q.get('token',[''])[0] or '')
            notice=f'<div class="flash">{h(msg)}</div>' if msg else ''
            base_style='''<style>
            body{background:linear-gradient(145deg,#eef5f0,#f9fbfa);min-height:100vh}.reset-shell{max-width:520px;margin:6vh auto;background:#fff;border:1px solid #dce8df;border-radius:24px;padding:28px;box-shadow:0 18px 48px rgba(22,72,45,.11)}
            .reset-brand{font-size:27px;font-weight:900;color:#183c2a;margin-bottom:5px}.reset-sub{color:#6c7f73;margin-bottom:22px}.reset-shell label{display:block;font-weight:800;font-size:13px;margin-top:12px}.reset-shell input{width:100%;padding:13px;border:1px solid #cbd9cf;border-radius:10px;margin-top:5px;font-size:16px}.reset-actions{display:flex;gap:8px;margin-top:18px;flex-wrap:wrap}.reset-actions .btn{flex:1}.code-box{text-align:center;letter-spacing:8px;font:900 24px Consolas,monospace}@media(max-width:600px){.reset-shell{margin:3vh 14px;padding:23px 20px}}</style>'''
            if step=='verify' and rid:
                content=f'''<div class="reset-brand">📧 Doğrulama Kodu</div><div class="reset-sub">E-postanıza gönderilen 6 haneli kodu girin. Kod 5 dakika geçerlidir.</div>{notice}<form method="post" action="/forgot-verify"><input type="hidden" name="reset_id" value="{h(rid)}"><label>6 Haneli Kod<input class="code-box" name="code" inputmode="numeric" pattern="[0-9]{{6}}" maxlength="6" autocomplete="one-time-code" required></label><div class="reset-actions"><button class="btn">Kodu Doğrula</button><a class="btn alt" href="/forgot-password">Başa Dön</a></div></form>'''
            elif step=='reset' and rid and token:
                content=f'''<div class="reset-brand">🔐 Yeni Şifre Belirle</div><div class="reset-sub">Doğrulama tamamlandı. Yeni şifrenizi oluşturun.</div>{notice}<form method="post" action="/forgot-reset"><input type="hidden" name="reset_id" value="{h(rid)}"><input type="hidden" name="token" value="{h(token)}"><label>Yeni Şifre<input type="password" name="password" minlength="8" required></label><label>Yeni Şifre Tekrar<input type="password" name="password_confirm" minlength="8" required></label><div class="reset-actions"><button class="btn">Şifreyi Değiştir</button></div></form>'''
            else:
                content=f'''<div class="reset-brand">🔑 Şifremi Unuttum</div><div class="reset-sub">Kullanıcı adınızı veya hesabınıza kayıtlı kurtarma e-postasını girin.</div>{notice}<form method="post" action="/forgot-password"><label>Kullanıcı Adı / Kurtarma E-postası<input name="identifier" autocomplete="username" required></label><div class="reset-actions"><button class="btn">📧 Kod Gönder</button><a class="btn alt" href="/login">Girişe Dön</a></div></form><p class="mut" style="margin-top:16px">Kod, kullanıcı hesabına tanımlı kurtarma e-postasına gönderilir.</p>'''
            return self.send_html(f'''<!doctype html><html lang="tr"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>ÇiftlikPro Şifre Kurtarma</title><style>{CSS}</style>{base_style}</head><body><div class="reset-shell">🐄 <b>ÇiftlikPro Enterprise</b><hr style="border:0;border-top:1px solid #e6eee8;margin:16px 0">{content}</div></body></html>''')
        if path.startswith('/uploads/'):
            name=os.path.basename(path.split('/uploads/',1)[1]); fp=UPLOADS/name
            if not fp.exists(): return self.send_html('Fotoğraf bulunamadı',404)
            ext=fp.suffix.lower(); ctype={'jpg':'image/jpeg','jpeg':'image/jpeg','png':'image/png','webp':'image/webp','gif':'image/gif'}.get(ext.lstrip('.'),'application/octet-stream')
            b=fp.read_bytes(); self.send_response(200); self.send_header('Content-Type',ctype); self.send_header('Content-Length',str(len(b))); self.end_headers(); self.wfile.write(b); return
        if path=='/logout':
            sid=self.parse_cookie(); SESSIONS.pop(sid,None); self.send_response(303);self.send_header('Set-Cookie','sid=; Max-Age=0; Path=/');self.send_header('Location','/login');self.end_headers();return
        if not self.require():return
        u=self.user()['username']
        if path=='/license-info':
            if not self.require_admin():return
            ok,payload,status=license_status();payload=payload or {}
            exp=payload.get('expires_on') or 'Süresiz'
            if exp!='Süresiz': exp=fmt_date(exp)
            badge_class='license-info-ok' if ok else 'license-info-bad'
            badge_text='🟢 Lisans Aktif' if ok else '🔴 Lisans Geçersiz'
            body=f'''<style>
.license-info-card{{background:#fff;border:1px solid #dfeae3;border-radius:22px;padding:26px;box-shadow:0 12px 35px rgba(20,80,48,.07)}}
.license-info-status{{display:inline-flex;padding:9px 14px;border-radius:999px;font-weight:800;margin-bottom:20px}}
.license-info-ok{{background:#e5f6eb;color:#14723d}}.license-info-bad{{background:#fff0ed;color:#ad3529}}
.license-info-grid{{display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:12px}}
.license-info-item{{background:#f5f9f6;border:1px solid #e2ece5;border-radius:14px;padding:15px}}
.license-info-item span{{display:block;color:#6d8075;font-size:12px;margin-bottom:5px;font-weight:700}}
.license-info-item b{{font-size:15px;color:#173b27;overflow-wrap:anywhere}}
.device-info{{grid-column:1/-1}}.device-row-info{{display:flex;align-items:center;gap:10px}}
.device-row-info b{{flex:1;font-family:Consolas,monospace}}
.copy-license-device{{border:0;border-radius:10px;background:#e2f1e7;color:#17683c;font-weight:800;padding:9px 13px;cursor:pointer}}
.license-note{{margin-top:18px;color:#6c7d73;line-height:1.55}}
@media(max-width:650px){{.license-info-grid{{grid-template-columns:1fr}}.device-info{{grid-column:auto}}.device-row-info{{align-items:stretch;flex-direction:column}}}}
</style>
<h1>🔐 Lisans Bilgileri</h1><div class="license-info-card">
<div class="license-info-status {badge_class}">{badge_text}</div>
<div class="license-info-grid">
<div class="license-info-item"><span>Lisans Sahibi</span><b>{h(payload.get('licensee') or '-')}</b></div>
<div class="license-info-item"><span>Ürün</span><b>{h(payload.get('product') or APP_NAME)}</b></div>
<div class="license-info-item"><span>Lisans Türü</span><b>{h(payload.get('license_type') or '-')}</b></div>
<div class="license-info-item"><span>Geçerlilik</span><b>{h(exp)}</b></div>
<div class="license-info-item device-info"><span>Cihaz Kimliği</span><div class="device-row-info"><b id="license-device-id">{h(device_id())}</b><button type="button" class="copy-license-device" onclick="copyLicenseDevice()">Kopyala</button></div></div>
<div class="license-info-item"><span>Durum</span><b>{h(status)}</b></div>
</div><div style="display:flex;gap:10px;flex-wrap:wrap;margin-top:18px"><a class="btn" href="/license-test">🧪 Lisans Aktivasyonunu Test Et</a><a class="btn secondary" href="/license-change">🔄 Lisansı Değiştir</a></div><p class="license-note">Lisans bu bilgisayarın cihaz kimliğine ve dijital imzaya bağlıdır. Lisans verisinde yapılan yetkisiz değişiklikler dijital imzayı geçersiz kılar.</p></div>
<script>function copyLicenseDevice(){{var e=document.getElementById("license-device-id");if(!e)return;navigator.clipboard.writeText(e.innerText).then(function(){{var b=document.querySelector(".copy-license-device");b.innerText="Kopyalandı ✓";setTimeout(function(){{b.innerText="Kopyala"}},1500)}})}}</script>'''
            return self.send_html(page('Lisans Bilgileri',body,'/license-info',u,msg))
        if path=='/license-test':
            if not self.require_admin():return
            body=f'''<h1>🧪 Lisans Aktivasyon Testi</h1><div class="card"><p class="mut">Bu ekran yeni CFP anahtarını doğrular; mevcut aktif lisansınızı değiştirmez.</p><div class="kv"><div><b>Cihaz Kimliği</b><span><code id="testDeviceId">{h(device_id())}</code> <button type="button" class="btn alt compact-btn" onclick="copyDeviceSimple('testDeviceId',this)">📋 Kopyala</button></span></div></div><form method="post" action="/license-test" class="form" style="margin-top:18px"><label class="full">Test Edilecek Lisans Anahtarı<textarea name="license_key" rows="5" placeholder="CFP-XXXXX-XXXXX-..." required></textarea></label><div class="full"><button class="btn">🧪 Anahtarı Doğrula</button> <a class="btn secondary" href="/license-info">Geri Dön</a></div></form></div>'''
            return self.send_html(page('Lisans Aktivasyon Testi',body,'/license-info',u,msg))
        if path=='/license-change':
            if not self.require_admin():return
            body=f'''<h1>🔄 Lisansı Değiştir</h1><div class="card"><div class="flash" style="background:#fff4dc;color:#7c5600">Bu işlem doğrulanan yeni lisansı mevcut aktif lisansın yerine kaydeder.</div><div class="kv"><div><b>Cihaz Kimliği</b><span><code id="changeDeviceId">{h(device_id())}</code> <button type="button" class="btn alt compact-btn" onclick="copyDeviceSimple('changeDeviceId',this)">📋 Kopyala</button></span></div></div><form method="post" action="/license-change" class="form" style="margin-top:18px"><label class="full">Yeni Lisans Anahtarı<textarea name="license_key" rows="5" placeholder="CFP-XXXXX-XXXXX-..." required></textarea></label><label class="full" style="display:flex;gap:8px;align-items:center"><input type="checkbox" name="confirm_change" value="yes" required style="width:auto"> Mevcut lisansın yeni lisansla değiştirileceğini onaylıyorum.</label><div class="full"><button class="btn">🔄 Yeni Lisansı Etkinleştir</button> <a class="btn secondary" href="/license-info">İptal</a></div></form></div>'''
            return self.send_html(page('Lisansı Değiştir',body,'/license-info',u,msg))
        if path=='/password-change':
            body='''<h1>Şifremi Değiştir</h1><div class="card"><form method="post" action="/password-change" class="form"><label>Mevcut Şifre<input type="password" name="current_password" required></label><label>Yeni Şifre<input type="password" name="new_password" minlength="8" required></label><label>Yeni Şifre Tekrar<input type="password" name="new_password_confirm" minlength="8" required></label><div class="full"><button class="btn">Şifreyi Değiştir</button></div></form></div>'''
            return self.send_html(page('Şifremi Değiştir',body,'/password-change',u,msg))
        if path=='/farm-profile':
            if not self.require_admin():return
            p=farm_profile()
            logo=p.get('farm_logo','')
            logo_html=(f'<img class="farm-logo-preview" src="{h(logo)}" alt="Çiftlik logosu">' if logo else '<div class="farm-logo-placeholder">🏡</div>')
            body=f'''<div class="settings-page-head"><h1>⚙️ ÇiftlikPro Ayarlar Merkezi</h1><p class="mut">Program, işletme, güvenlik, veri ve kullanıcı ayarlarını tek merkezden yönetin.</p></div>
            <div class="settings-groups">
              <section class="settings-group"><h3>⚙ Genel & İşletme</h3><a href="#farm-settings"><b>🏡 İşletme Bilgileri</b><span>Çiftlik adı, iletişim, adres ve logo</span></a><a href="/?edit=1"><b>▦ Dashboard Düzeni</b><span>Ana ekran kartlarını düzenle</span></a><a href="/performance-settings"><b>🐂 Besi Ayarları</b><span>Performans ve hedef profilleri</span></a></section>
              <section class="settings-group"><h3>🔐 Güvenlik & Kullanıcı</h3><a href="/users"><b>👥 Kullanıcı Yönetimi</b><span>Yetki, kullanıcı ve personel hesapları</span></a><a href="/password-change"><b>🔑 Şifre Değiştir</b><span>Aktif kullanıcı şifresi</span></a><a href="/audit-log"><b>📜 İşlem Günlüğü</b><span>Program içi işlem geçmişi</span></a></section>
              <section class="settings-group"><h3>🗄 Veri & İletişim</h3><a href="/backups"><b>💾 Yedekleme Merkezi</b><span>SQLite yedekleri ve veri güvenliği</span></a><a href="/data"><b>⇄ Veri Aktarımı</b><span>JSON içe/dışa aktarım</span></a><a href="/smtp-settings"><b>📧 E-posta / SMTP</b><span>Şifre kurtarma ve posta ayarları</span></a></section>
              <section class="settings-group"><h3>ℹ Sistem</h3><a href="/license-info"><b>🔐 Lisans Bilgileri</b><span>Lisans ve cihaz bilgileri</span></a><a href="/version-notes"><b>📝 Sürüm Notları</b><span>ÇiftlikPro değişiklik geçmişi</span></a><a href="/reports"><b>▥ Raporlar</b><span>Raporlama merkezine git</span></a></section>
            </div>
            <div class="card" id="farm-settings"><div class="farm-profile-head">{logo_html}<div><h2 style="margin:0 0 6px">İşletme Bilgileri</h2><p class="mut" style="margin:0">{h(farm_display_name(p))} · Dashboard ve raporlarda kullanılan temel bilgiler.</p></div></div></div>
            <div class="card" style="margin-top:14px"><form method="post" action="/farm-profile" enctype="multipart/form-data" class="form">
            <label>Çiftlik / İşletme Adı<input name="farm_name" value="{h(p.get('farm_name'))}" placeholder="Örn. Erdoğmuş Çiftliği"></label>
            <label>İşletme Sahibi<input name="owner_name" value="{h(p.get('owner_name'))}"></label>
            <label>Telefon<input name="phone" value="{h(p.get('phone'))}" inputmode="tel"></label>
            <label>E-posta<input type="email" name="email" value="{h(p.get('email'))}"></label>
            <label>İl<input name="province" value="{h(p.get('province'))}"></label>
            <label>İlçe<input name="district" value="{h(p.get('district'))}"></label>
            <label>İşletme Numarası<input name="business_no" value="{h(p.get('business_no'))}"></label>
            <label>Vergi / TC (isteğe bağlı)<input name="tax_or_tc" value="{h(p.get('tax_or_tc'))}"></label>
            <label class="full">Adres<textarea name="address" rows="3">{h(p.get('address'))}</textarea></label>
            <div class="full"><h2 style="margin:8px 0 0">Veteriner İletişim Bilgileri</h2></div>
            <label>Veteriner Adı<input name="vet_name" value="{h(p.get('vet_name'))}"></label>
            <label>Veteriner Telefonu<input name="vet_phone" value="{h(p.get('vet_phone'))}" inputmode="tel"></label>
            <label>Veteriner E-posta<input type="email" name="vet_email" value="{h(p.get('vet_email'))}"></label>
            <label>Çiftlik Logosu<input type="file" name="farm_logo_file" accept="image/jpeg,image/png,image/webp,.jpg,.jpeg,.png,.webp"><span class="camera-note">JPG, PNG veya WebP · En fazla 5 MB</span></label>
            <label class="full">Notlar<textarea name="notes" rows="4">{h(p.get('notes'))}</textarea></label>
            <label class="full" style="display:flex;flex-direction:row;align-items:center;gap:8px"><input type="checkbox" name="remove_logo" value="1" style="width:auto"> Mevcut logoyu kaldır</label>
            <div class="full"><button class="btn">💾 Çiftlik Profilini Kaydet</button></div>
            </form></div>'''
            return self.send_html(page('Ayarlar',body,'/farm-profile',u,msg))
        if path=='/smtp-settings':
            if not self.require_admin():return
            cfg=smtp_config();sec=cfg['security']
            body=f'''<h1>📧 E-posta / Şifre Kurtarma Ayarları</h1><div class="two"><div class="card"><h2>SMTP Ayarları</h2><form method="post" action="/smtp-settings" class="form"><label>SMTP Sunucusu<input name="smtp_host" value="{h(cfg['host'])}" required></label><label>Port<input type="number" name="smtp_port" value="{cfg['port']}" required></label><label>Güvenlik<select name="smtp_security"><option value="starttls" {'selected' if sec=='starttls' else ''}>STARTTLS</option><option value="ssl" {'selected' if sec=='ssl' else ''}>SSL/TLS</option><option value="none" {'selected' if sec=='none' else ''}>Yok</option></select></label><label>SMTP Kullanıcı<input name="smtp_username" value="{h(cfg['username'])}"></label><label>Gönderen E-posta<input type="email" name="smtp_sender" value="{h(cfg['sender'])}" required></label><label>SMTP / Uygulama Şifresi<input type="password" name="smtp_password" placeholder="Değiştirmeyecekseniz boş bırakın"></label><div class="full"><button class="btn">💾 Ayarları Kaydet</button></div></form></div><div class="card"><h2>Test E-postası</h2><p class="mut">Sağlayıcınız iki aşamalı doğrulama kullanıyorsa normal hesap şifresi yerine uygulama şifresi kullanın.</p><form method="post" action="/smtp-settings" class="form"><input type="hidden" name="action" value="test"><label class="full">Test Adresi<input type="email" name="test_email" required></label><div class="full"><button class="btn blue">📨 Test E-postası Gönder</button></div></form></div></div>'''
            return self.send_html(page('E-posta Ayarları',body,'/users',u,msg))
        if path=='/users':
            if not self.require_admin():return
            with db() as c:rows=c.execute('select id,username,full_name,recovery_email,role,active,last_login from users order by username').fetchall()
            trs=''.join(f'''<tr><td>{h(r["full_name"])}</td><td>{h(r["username"])}</td><td>{h(r["recovery_email"]) or '-'}</td><td>{'Yönetici' if r["role"]=='admin' else 'Personel'}</td><td>{'Aktif' if r["active"] else 'Pasif'}</td><td>{fmt_datetime(r["last_login"]) if r["last_login"] else '-'}</td><td><a class="btn alt" href="/users/edit?id={r["id"]}">Düzenle</a></td></tr>''' for r in rows)
            body=f'''<h1>Kullanıcı Yönetimi</h1><div class="two"><div class="card"><h2>Yeni Kullanıcı</h2><form method="post" action="/users/create" class="form"><label>Ad Soyad<input name="full_name" required></label><label>Kullanıcı Adı<input name="username" required></label><label>Kurtarma E-postası<input type="email" name="recovery_email" required></label><label>Şifre<input type="password" name="password" minlength="8" required></label><label>Rol<select name="role"><option value="personel">Personel</option><option value="admin">Yönetici</option></select></label><div class="full"><button class="btn">Kullanıcı Oluştur</button></div></form></div><div class="card"><h2>Güvenlik</h2><p class="mut">Şifreler PBKDF2-SHA256 ile saklanır. Şifre kurtarma kodları 5 dakika geçerli ve tek kullanımlıktır.</p><a class="btn alt" href="/smtp-settings">📧 E-posta Ayarları</a></div></div><div class="card" style="margin-top:14px"><table><tr><th>Ad Soyad</th><th>Kullanıcı</th><th>Kurtarma E-postası</th><th>Rol</th><th>Durum</th><th>Son Giriş</th><th>İşlem</th></tr>{trs}</table></div>'''
            return self.send_html(page('Kullanıcı Yönetimi',body,'/users',u,msg))
        if path=='/users/edit':
            if not self.require_admin():return
            uid=q.get('id',[''])[0]
            with db() as c:r=c.execute('select * from users where id=?',(uid,)).fetchone()
            if not r:return self.send_html('Kullanıcı bulunamadı',404)
            body=f'''<h1>Kullanıcı Düzenle</h1><div class="card"><form method="post" action="/users/update" class="form"><input type="hidden" name="id" value="{r["id"]}"><label>Ad Soyad<input name="full_name" value="{h(r["full_name"])}" required></label><label>Kurtarma E-postası<input type="email" name="recovery_email" value="{h(r["recovery_email"])}" required></label><label>Rol<select name="role"><option value="personel" {'selected' if r["role"]=='personel' else ''}>Personel</option><option value="admin" {'selected' if r["role"]=='admin' else ''}>Yönetici</option></select></label><label>Durum<select name="active"><option value="1" {'selected' if r["active"] else ''}>Aktif</option><option value="0" {'selected' if not r["active"] else ''}>Pasif</option></select></label><label>Yeni Şifre<input type="password" name="new_password" minlength="8"></label><div class="full"><button class="btn">Kaydet</button> <a class="btn alt" href="/users">İptal</a></div></form></div>'''
            return self.send_html(page('Kullanıcı Düzenle',body,'/users',u,msg))
        if path=='/audit-log':
            if not self.require_admin():return
            with db() as c:rows=c.execute('select * from audit_log order by id desc limit 300').fetchall()
            trs=''.join(f'<tr><td>{fmt_datetime(r["created_at"])}</td><td>{h(r["username"])}</td><td>{h(r["action"])}</td><td>{h(r["detail"])}</td><td>{h(r["ip_address"])}</td></tr>' for r in rows) or '<tr><td colspan=5>Kayıt yok.</td></tr>'
            body=f'''<h1>İşlem Günlüğü</h1><div class="card"><table><tr><th>Tarih</th><th>Kullanıcı</th><th>İşlem</th><th>Detay</th><th>IP</th></tr>{trs}</table></div>'''
            return self.send_html(page('İşlem Günlüğü',body,'/audit-log',u,msg))
        promote_mature_calves()
        if path=='/':
            _dash_t0=time.perf_counter()
            profile=farm_profile()
            farm_name=farm_display_name(profile)
            edit_dashboard=(q.get('edit',['0'])[0]=='1')
            dash_layout=dashboard_layout(u)
            with db() as c:
                animals=c.execute("select count(*) from animals where gender='Dişi' and status='Aktif'").fetchone()[0]
                males=c.execute("select count(*) from animals where gender='Erkek' and status='Aktif'").fetchone()[0]
                calves=c.execute('select count(*) from calves where promoted_animal_id is null').fetchone()[0]
                total_inc=c.execute("select coalesce(sum(amount),0) from finance where tx_type='Gelir'").fetchone()[0]
                total_exp=c.execute("select coalesce(sum(amount),0) from finance where tx_type='Gider'").fetchone()[0]
                pregnant=c.execute("select count(distinct animal_id) from inseminations where pregnancy_result='Pozitif'").fetchone()[0]
                active_total=animals+males+calves
                # HOTFIX 6.10: Dashboard ilk açılışında kullanılmayan erkek maliyet/performans
                # N+1 hesaplarını çalıştırma. Bu veriler Besi Performansı ekranında hesaplanır.
                due_rows=c.execute("select i.due_date,a.id,a.tag,a.nickname from inseminations i join animals a on a.id=i.animal_id where i.pregnancy_result='Pozitif' and i.due_date between ? and ? order by i.due_date limit 8",(date.today().isoformat(),(date.today()+timedelta(days=45)).isoformat())).fetchall()
                health_rows=c.execute("select h.id,h.next_date,h.kind,h.product,h.notes,h.animal_id,h.calf_id,a.id as adult_id,a.tag as animal_tag,ca.tag as calf_tag from health h left join animals a on a.id=h.animal_id left join calves ca on ca.id=h.calf_id where coalesce(h.next_date,'')<>'' and h.next_date<=? order by h.next_date limit 10",((date.today()+timedelta(days=30)).isoformat(),)).fetchall()
                pregnancy_vaccines=pregnancy_vaccine_tasks(c,horizon_days=7)
                estrus_dash_all=c.execute("select e.*,a.tag,a.nickname from estrus_records e join animals a on a.id=e.animal_id where a.gender='Dişi' and coalesce(a.status,'Aktif')='Aktif' order by e.estrus_date desc,e.id desc").fetchall()
                pregnant_ids={r[0] for r in c.execute("select distinct animal_id from inseminations where pregnancy_result='Pozitif' and animal_id is not null").fetchall()}
                estrus_dash_rows=[r for r in estrus_dash_all if r['animal_id'] not in pregnant_ids]
                month_defs=[]
                for n in range(5,-1,-1):
                    d=(date.today().replace(day=1)-timedelta(days=n*31)).replace(day=1); month_defs.append((d,d.strftime('%Y-%m')))
                month_keys=[x[1] for x in month_defs]
                finance_monthly={}
                if month_keys:
                    ph=','.join('?' for _ in month_keys)
                    for rr in c.execute(f"select substr(tx_date,1,7) ym,tx_type,coalesce(sum(amount),0) total from finance where substr(tx_date,1,7) in ({ph}) group by ym,tx_type",month_keys).fetchall():
                        finance_monthly[(rr['ym'],rr['tx_type'])]=rr['total']
                months=[(d.strftime('%m/%y'),finance_monthly.get((key,'Gelir'),0),finance_monthly.get((key,'Gider'),0)) for d,key in month_defs]
            estrus_latest={}
            for er in estrus_dash_rows:
                if er['animal_id'] not in estrus_latest: estrus_latest[er['animal_id']]=er
            estrus_upcoming=[]; today=date.today()
            for er in estrus_latest.values():
                cycle=next_estrus_cycle(c,er,today)
                if cycle and cycle['end']>=today and cycle['start']<=today+timedelta(days=30):
                    estrus_upcoming.append((cycle['start'],cycle['center'],cycle['end'],er,cycle['cycle_no']))
            estrus_upcoming.sort(key=lambda x:x[1])
            estrus_dashboard_cards=[]
            for es,ec,ee,er,cycle_no in estrus_upcoming[:8]:
                in_window=es<=today<=ee
                if in_window:
                    action=f'''<form method="post" action="/estrus-inseminate" onsubmit="return confirm('Bu hayvan bugün tohumlandı olarak Tohumlama kayıtlarına aktarılsın mı?')"><input type="hidden" name="estrus_id" value="{er['id']}"><button class="btn orange">🌱 Bugün Tohumlandı</button></form>'''
                else:
                    action=f'''<form method="post" action="/estrus-send" onsubmit="return confirm('Bu hayvan Tohumlama Takibi ekranına gönderilsin mi? Ana sayfadaki bu kızgınlık kartı kapanacaktır.')"><input type="hidden" name="estrus_id" value="{er['id']}"><input type="hidden" name="cycle_no" value="{cycle_no}"><button class="btn orange">🌱 Tohumlamaya Gönder</button></form>'''
                estrus_dashboard_cards.append(f'''<div class="alertitem {'estrus-window-now' if in_window else 'estrus-window-next'}"><b>🌸 <a class="taglink" href="/animal?id={er['animal_id']}">{h(er['tag'])} {h(er['nickname'])}</a></b><br><span class="mut">{fmt_date(es.isoformat())} – {fmt_date(ee.isoformat())} · En olası {fmt_date(ec.isoformat())}</span><div class="estrus-actions">{action}<form method="post" action="/estrus-skip" onsubmit="return confirm('Bu östrus dönemi atlandı olarak işaretlenecek. Emin misiniz?')"><input type="hidden" name="estrus_id" value="{er['id']}"><input type="hidden" name="cycle_no" value="{cycle_no}"><input type="hidden" name="return_to" value="/"><button class="btn alt">⏭️ Bu Östrusu Atla</button></form><a class="btn alt" href="/estrus">Kızgınlık Takibi</a></div></div>''')
            estrus_dashboard_html=''.join(estrus_dashboard_cards) or '<p class="mut">Önümüzdeki 30 gün için beklenen kızgınlık yok.</p>'
            net=total_inc-total_exp; maxv=max([max(x[1],x[2]) for x in months]+[1])
            bars=''.join(f'<div class="mini-col"><b title="Gelir {money(i)}" style="height:{max(2,int(i/maxv*100))}%"></b><i title="Gider {money(e)}" style="height:{max(2,int(e/maxv*100))}%"></i><span>{h(m)}</span></div>' for m,i,e in months)
            due_html=''.join(f'<div class="alertitem">🐄 <a class="taglink" href="/animal?id={r["id"]}">{h(r["tag"])} {h(r["nickname"])}</a><br><span class="mut">Tahmini doğum: {fmt_date(r["due_date"])}</span></div>' for r in due_rows) or '<p class="mut">45 gün içinde beklenen doğum yok.</p>'
            def health_task_html(r):
                tag=r["animal_tag"] or r["calf_tag"] or "Genel"
                try:days=(date.fromisoformat(r["next_date"])-date.today()).days
                except:days=9999
                if days<0:label=f'GECİKTİ · {abs(days)} gün';style='border-left-color:#c8392b;background:#fff1f0'
                elif days==0:label='BUGÜN';style='border-left-color:#e27b1f;background:#fff6e8'
                elif days<=3:label=f'{days} gün kaldı';style='border-left-color:#e27b1f;background:#fff6e8'
                else:label=f'{days} gün kaldı';style='border-left-color:#e2a21f;background:#fff9e8'
                action=''
                if str(r["kind"] or '')=='Aşı' and 'IKINCI_DOZ_PLAN' in str(r["notes"] or ''):
                    action=f'<form method="post" action="/health/second-dose-done" class="actions" style="margin-top:8px"><input type="hidden" name="source_id" value="{r["id"]}"><input type="hidden" name="return_to" value="/"><button class="btn">✅ 2. Doz Yapıldı</button><a class="btn alt" href="/health">Sağlığı Aç</a></form>'
                return f'<div class="alertitem" style="{style}"><b>💉 {h(tag)} · {h(r["kind"])}</b><br><span class="mut">{h(r["product"])} · {fmt_date(r["next_date"])} · {label}</span>{action}</div>'
            health_html=''.join(health_task_html(r) for r in health_rows) or '<p class="mut">30 gün içinde planlanan/geciken sağlık işlemi yok.</p>'
            def vaccine_task_html(t):
                if t['overdue']:
                    label=f"GECİKTİ · {abs(t['days_left'])} gün"; style='border-left-color:#c8392b;background:#fff1f0'
                elif t['today']:
                    label='BUGÜN YAPILMALI'; style='border-left-color:#e27b1f;background:#fff6e8'
                else:
                    label=f"{t['days_left']} gün kaldı"; style='border-left-color:#e2a21f;background:#fff9e8'
                return f'<div class="alertitem" style="{style}"><b>💉 {h(t["tag"])} · {t["month"]}. Ay Gebelik Aşısı</b><br><span class="mut">Planlanan: {fmt_date(t["task_date"])} · {label}</span><form method="post" action="/pregnancy-vaccine/done" class="actions" style="margin-top:8px"><input type="hidden" name="animal_id" value="{t["animal_id"]}"><input type="hidden" name="insemination_id" value="{t["insemination_id"]}"><input type="hidden" name="month" value="{t["month"]}"><input type="hidden" name="return_to" value="/"><button class="btn">✅ Aşı Yapıldı</button><a class="btn alt" href="/animal?id={t["animal_id"]}">Hayvanı Aç</a></form></div>'
            pregnancy_vaccine_html=''.join(vaccine_task_html(t) for t in pregnancy_vaccines) or '<p class="mut">7 gün içinde 7./8. ay gebelik aşısı görevi yok.</p>'
            dash_cards={
                'active_total':f'<a class="card stat metric green summary-link" href="/all-animals"><span class="metric-icon">🐄</span><span class="metric-title">Toplam Aktif Hayvan</span><b>{active_total}</b><small>Tüm hayvanları aç →</small></a>',
                'female':f'<a class="card stat metric green summary-link" href="/animals"><span class="metric-icon">🐮</span><span class="metric-title">Dişi Hayvan</span><b>{animals}</b><small>Listeyi aç →</small></a>',
                'male':f'<a class="card stat metric blue summary-link" href="/males"><span class="metric-icon">🐂</span><span class="metric-title">Erkek Hayvan</span><b>{males}</b><small>Listeyi aç →</small></a>',
                'pregnant':f'<a class="card stat metric orange summary-link" href="/inseminations"><span class="metric-icon">🤰</span><span class="metric-title">Gebe Hayvan</span><b>{pregnant}</b><small>Gebelikleri aç →</small></a>',
                'calves':f'<a class="card stat metric teal summary-link" href="/calves"><span class="metric-icon">🐮</span><span class="metric-title">Buzağı</span><b>{calves}</b><small>Listeyi aç →</small></a>',
                'due':f'<a class="card stat metric purple summary-link" href="#approaching-births"><span class="metric-icon">📅</span><span class="metric-title">Yaklaşan Doğum</span><b>{len(due_rows)}</b><small>Detaya git ↓</small></a>',
                'estrus':f'<a class="card stat metric green summary-link" href="#approaching-estrus"><span class="metric-icon">🌸</span><span class="metric-title">Yaklaşan Kızgınlık</span><b>{len(estrus_upcoming)}</b><small>Detaya git ↓</small></a>',
                'health_due':f'<a class="card stat metric teal summary-link" href="/health"><span class="metric-icon">💉</span><span class="metric-title">Yaklaşan Sağlık</span><b>{len(health_rows)}</b><small>Sağlığı aç →</small></a>',
                'income':f'<a class="card stat metric green summary-link" href="/finance?type=Gelir"><span class="metric-icon">📥</span><span class="metric-title">Toplam Gelir</span><b>{money(total_inc)}</b><small>Gelirleri aç →</small></a>',
                'expense':f'<a class="card stat metric red summary-link" href="/finance?type=Gider"><span class="metric-icon">📤</span><span class="metric-title">Toplam Gider</span><b>{money(total_exp)}</b><small>Giderleri aç →</small></a>',
                'net':f'<a class="card stat metric {"red" if net<0 else "green"} summary-link" href="/finance"><span class="metric-icon">⚖️</span><span class="metric-title">Net Durum</span><b>{money(net)}</b><small>Finansı aç →</small></a>',
            }
            card_meta={
                'active_total':('🐄','Toplam Aktif Hayvan','Sürüdeki toplam aktif kayıt'),
                'female':('🐮','Dişi Hayvan','Aktif dişi hayvan sayısı'),
                'male':('🐂','Erkek Hayvan','Aktif erkek hayvan sayısı'),
                'pregnant':('🤰','Gebe Hayvan','Pozitif gebelik kayıtları'),
                'calves':('🐮','Buzağı','Aktif buzağı kayıtları'),
                'due':('📅','Yaklaşan Doğum','Yaklaşan doğum sayısı'),
                'estrus':('🌸','Yaklaşan Kızgınlık','Takip penceresindeki hayvanlar'),
                'health_due':('💉','Yaklaşan Sağlık İşlemleri','30 gün içindeki planlı/gecikmiş sağlık işlemleri'),
                'income':('📥','Toplam Gelir','Tüm gelirlerin toplamı'),
                'expense':('📤','Toplam Gider','Tüm giderlerin toplamı'),
                'net':('⚖️','Net Durum','Gelir eksi gider'),
            }
            dash_slots=[]
            for slot,key in enumerate(dash_layout):
                card=dash_cards.get(key,'')
                if edit_dashboard:
                    if card:
                        card=card+f'<button type="button" class="dashboard-slot-plus" title="Bu kartı değiştir" onclick="openDashboardPicker({slot},\'{h(key)}\')">+</button>'
                    else:
                        card=f'''<div class="dashboard-empty-slot" onclick="openDashboardPicker({slot},'')"><div><div class="plus-icon">+</div><small>Bu yuvaya kart ekle</small></div></div>'''
                    cls='dashboard-slot editing'
                else:
                    if not card: continue
                    cls='dashboard-slot'
                dash_slots.append(f'<div class="{cls}">{card}</div>')
            dashboard_summary_html=''.join(dash_slots)
            gallery_choices=''.join(f'''<button type="button" class="dashboard-card-choice" data-key="{h(k)}" onclick="chooseDashboardCard('{h(k)}')"><span class="choice-icon">{card_meta[k][0]}</span><span><b>{h(card_meta[k][1])}</b><small>{h(card_meta[k][2])}</small></span></button>''' for k,_ in DASHBOARD_CARD_OPTIONS)
            dashboard_picker_html=f'''<div class="dashboard-picker-backdrop" id="dashboardPicker" onclick="if(event.target===this)closeDashboardPicker()"><div class="dashboard-picker"><div class="dashboard-picker-head"><div><h2>Dashboard Kartı Seç</h2><p>Seçtiğiniz kart bu yuvaya anında yerleşir.</p></div><button type="button" class="dashboard-picker-close" onclick="closeDashboardPicker()">×</button></div><div class="dashboard-card-gallery">{gallery_choices}</div><div class="dashboard-picker-footer"><button type="button" class="btn red" onclick="chooseDashboardCard('')">Yuvayı Boşalt</button><span class="mut">Daha sonra tekrar ekleyebilirsiniz.</span></div><form id="dashboardPickerForm" method="post" action="/dashboard-layout"><input type="hidden" name="slot" id="dashboardPickerSlot"><input type="hidden" name="card_key" id="dashboardPickerKey"></form></div></div><script>function openDashboardPicker(slot,current){{document.getElementById('dashboardPickerSlot').value=slot;document.getElementById('dashboardPicker').classList.add('open');document.querySelectorAll('.dashboard-card-choice').forEach(function(b){{b.classList.toggle('active',b.dataset.key===current);}});}}function closeDashboardPicker(){{document.getElementById('dashboardPicker').classList.remove('open');}}function chooseDashboardCard(key){{document.getElementById('dashboardPickerKey').value=key;document.getElementById('dashboardPickerForm').submit();}}document.addEventListener('keydown',function(e){{if(e.key==='Escape')closeDashboardPicker();}});</script>'''
            dashboard_logo=(f'<img class="farm-hero-logo" src="{h(profile.get("farm_logo"))}" alt="Çiftlik logosu">' if profile.get('farm_logo') else '')
            body=f'''<div class="hero"><a class="farm-hero home-hero-link" href="/" title="Ana Sayfa">{dashboard_logo}<div><h1>{h(farm_name)}</h1><div>Bugünün sürü, sağlık ve finans görünümü</div></div></a><div><a class="btn orange" href="/backup/create">💾 Hemen Yedek Al</a></div></div>
            <div class="dashboard-section-title"><h2>Dashboard Kartlarım</h2><span>{'Kartın üzerindeki + işaretine dokunarak değiştirebilirsiniz' if edit_dashboard else 'Size özel hızlı görünüm'}</span></div>
            <div class="grid summary-grid">{dashboard_summary_html}</div>{dashboard_picker_html if edit_dashboard else ''}
            <div class="dashboard-section-title today-title" id="approaching-estrus"><h2>📋 Bugünün İşleri</h2><span>Öncelikli üreme, sağlık ve finans takibi</span></div>
            <div class="today-work-grid">
              <section class="card today-work-card"><div class="today-work-head"><b>🌸 Yaklaşan Kızgınlık</b><span>{len(estrus_upcoming)}</span></div><div class="today-work-body"><div class="alertlist compact-alerts">{estrus_dashboard_html}</div></div><a class="today-work-action" href="/estrus">Kızgınlık Takibini Aç →</a></section>
              <section class="card today-work-card"><div class="today-work-head"><b>💉 Gebelik / Aşı Alarmı</b><span>{len(pregnancy_vaccines)}</span></div><div class="today-work-body"><div class="alertlist compact-alerts">{pregnancy_vaccine_html}</div></div><a class="today-work-action" href="/health">Sağlık Takibini Aç →</a></section>
              <section class="card today-work-card finance-today"><div class="today-work-head"><b>₺ Finans Özeti</b><span>{money(net)}</span></div><div class="today-finance-row"><span>Toplam Gelir</span><b>{money(total_inc)}</b></div><div class="today-finance-row"><span>Toplam Gider</span><b>{money(total_exp)}</b></div><div class="today-finance-row net"><span>Net Durum</span><b>{money(net)}</b></div><a class="today-work-action" href="/finance">Finans Detaylarını Aç →</a></section>
            </div>
            <div class="card" style="display:flex;align-items:center;justify-content:space-between;gap:18px;flex-wrap:wrap;margin-top:12px"><div><h2 style="margin:0 0 6px">🐂 Besi Performansı</h2><p class="mut" style="margin:0">Aktif ve kesilen erkekleri; alım tarihi, kesim tarihi, kilo performansı ve gerçekleşmiş maliyete göre inceleyin.</p></div><a class="btn blue" href="/performance">Besi Analizine Git →</a></div>
            <div class="two" style="margin-top:14px"><div class="card"><h2>Son 6 Ay Finans Eğilimi</h2><div class="mut">Yeşil: gelir · Kırmızı: gider</div><div class="mini-chart">{bars}</div></div><div class="card"><h2>Hızlı İşlemler</h2><p class="mut">Detaylı finans hareketleri Finans bölümünde tutulur.</p><div class="actions"><a class="btn blue" href="/finance">Finans Kaydı</a><a class="btn alt" href="/health">Sağlık Kaydı</a><a class="btn alt" href="/reports">Finans Raporları</a></div></div></div><div class="two" style="margin-top:14px"><div class="card" id="approaching-births"><h2>Yaklaşan Doğumlar</h2><div class="alertlist">{due_html}</div></div><div class="card"><h2>Yaklaşan Aşı / Sağlık</h2><div class="alertlist">{health_html}</div></div></div>'''
            with db() as c:
                month_key=date.today().strftime('%Y-%m')
                month_milk_income=c.execute("select coalesce(sum(amount),0) from finance where tx_type='Gelir' and category in ('Süt Satışı','Süt Geliri') and substr(tx_date,1,7)=?",(month_key,)).fetchone()[0]
                month_cut_income=c.execute("select coalesce(sum(amount),0) from finance where tx_type='Gelir' and category='Kesim Geliri' and substr(tx_date,1,7)=?",(month_key,)).fetchone()[0]
                recent_weights=c.execute("select count(*) from weights where measure_date>=?",((date.today()-timedelta(days=30)).isoformat(),)).fetchone()[0]
            body += f'''<div class="card" style="margin-top:14px"><h2>İşletme Özeti</h2><div class="grid business-summary-grid"><div class="card stat metric blue">Bu Ay Süt Geliri<b>{money(month_milk_income)}</b></div><div class="card stat metric green">Bu Ay Kesim Geliri<b>{money(month_cut_income)}</b></div><div class="card stat metric orange">30 Günlük Kilo Kaydı<b>{recent_weights}</b></div></div><div class="actions"><a class="btn" href="/animal-add">+ Hayvan Ekle</a><a class="btn alt" href="/reports">Raporları Aç</a></div></div>'''
            # V3.9.20: Dashboard uzunluğunu azaltan açılır/kapanır bölümler.
            fold_js='''<script>(function(){var titles=[].slice.call(document.querySelectorAll('.dashboard-section-title'));titles.forEach(function(t,i){if(i===0)return;var key='cp_dash_section_'+i;var saved=localStorage.getItem(key);var open=saved===null?(i<3):saved==='1';var nodes=[];for(var n=t.nextElementSibling;n&&!(n.classList&&n.classList.contains('dashboard-section-title'));n=n.nextElementSibling)nodes.push(n);t.style.cursor='pointer';var hint=t.querySelector('span');if(hint)hint.dataset.original=hint.textContent;function paint(){nodes.forEach(function(x){x.style.display=open?'':'none'});if(hint)hint.textContent=(open?'▲ ':'▼ ')+(hint.dataset.original||'');}t.addEventListener('click',function(){open=!open;localStorage.setItem(key,open?'1':'0');paint()});paint();});})();</script>'''
            # Büyük dashboard bloklarını başlıklarına göre istemci tarafında kompaktlaştır.
            body += fold_js
            print(f'[PERF] Dashboard hazır: {time.perf_counter()-_dash_t0:.3f} sn')
            return self.send_html(page('Profesyonel Dashboard',body,'/',u,msg))

        if path=='/cost-details':
            search=(q.get('search',[''])[0] or '').strip()
            status=(q.get('status',['all'])[0] or 'all').lower()
            if status not in ('all','active','cut'): status='all'
            try: per_page=int(q.get('per_page',['20'])[0])
            except: per_page=20
            if per_page not in (10,20,50): per_page=20
            try: page_no=max(1,int(q.get('page',['1'])[0]))
            except: page_no=1
            with db() as c:
                all_cost_rows=c.execute("select * from animals where gender='Erkek' and coalesce(status,'Aktif') in ('Aktif','Kesildi') order by tag").fetchall()
            filtered=[]
            needle=search.casefold()
            for r in all_cost_rows:
                rs=str(r['status'] or 'Aktif')
                if status=='active' and rs!='Aktif': continue
                if status=='cut' and rs!='Kesildi': continue
                hay=(str(r['tag'] or '')+' '+str(r['nickname'] or '')).casefold()
                if needle and needle not in hay: continue
                filtered.append(r)
            total_count=len(filtered)
            total_pages=max(1,(total_count+per_page-1)//per_page)
            if page_no>total_pages: page_no=total_pages
            offset=(page_no-1)*per_page
            shown=filtered[offset:offset+per_page]
            def cost_detail_tr(r):
                days,daily,operating,total=animal_cost_values(r)
                purchase=float(r['purchase_price'] or 0)
                start=r['purchase_date'] or r['birth_date'] or '-'
                is_cut=str(r['status'] or '')=='Kesildi'
                end=(r['exit_date'] or '-') if is_cut else 'Bugün'
                badge='<span class="perf-badge status-low">Kesildi</span>' if is_cut else '<span class="perf-badge status-good">Aktif</span>'
                note='<br><span class="mut">Kesim tarihinde donduruldu</span>' if is_cut else ''
                return f'<tr><td><a class="animal-tag-btn" title="Hayvan kartını aç" href="/animal?id={r["id"]}">{h(r["tag"])}</a><br><span class="mut">{h(r["nickname"])}</span></td><td>{badge}</td><td>{fmt_date(start)}</td><td>{fmt_date(end)}</td><td><b>{days} gün</b>{note}</td><td>{money(purchase)}</td><td>{money(operating)}</td><td><b>{money(total)}</b></td></tr>'
            detail_rows=''.join(cost_detail_tr(r) for r in shown) or '<tr><td colspan="8">Bu filtreye uygun hayvan bulunamadı.</td></tr>'
            active_rows=[r for r in filtered if str(r['status'] or '')=='Aktif']
            cut_rows=[r for r in filtered if str(r['status'] or '')=='Kesildi']
            purchase_sum=sum(float(r['purchase_price'] or 0) for r in filtered)
            operating_sum=sum(animal_cost_values(r)[2] for r in filtered)
            total_sum=sum(animal_cost_values(r)[3] for r in filtered)
            def page_url(n):
                return '/cost-details?'+urllib.parse.urlencode({'search':search,'status':status,'per_page':per_page,'page':n})
            pager=[]
            if page_no>1: pager.append(f'<a class="btn alt" href="{h(page_url(page_no-1))}">← Önceki</a>')
            pager.append(f'<span class="pill">Sayfa <b>{page_no}</b> / {total_pages}</span>')
            if page_no<total_pages: pager.append(f'<a class="btn alt" href="{h(page_url(page_no+1))}">Sonraki →</a>')
            status_options=''.join(f'<option value="{v}" {"selected" if status==v else ""}>{label}</option>' for v,label in [('all','Tümü'),('active','Sadece Aktif'),('cut','Sadece Kesilen')])
            per_options=''.join(f'<option value="{n}" {"selected" if per_page==n else ""}>{n} kayıt</option>' for n in (10,20,50))
            body=f'''<div class="hero"><div><h1>🐂 Erkek Hayvan Maliyet Detayı</h1><div>Dashboard kısa kalır; bütün maliyet kaynaklarını burada inceleyebilirsiniz.</div></div><div><a class="btn alt" href="/">← Dashboard</a></div></div>
            <div class="grid"><div class="card stat metric blue"><span class="metric-icon">🐂</span>Aktif Erkek<b>{len(active_rows)}</b></div><div class="card stat metric red"><span class="metric-icon">🥩</span>Kesilen Erkek<b>{len(cut_rows)}</b></div><div class="card stat metric orange"><span class="metric-icon">🌾</span>Yem + Bakım<b>{money(operating_sum)}</b></div><div class="card stat metric green"><span class="metric-icon">💰</span>Filtre Toplamı<b>{money(total_sum)}</b><small>Alış {money(purchase_sum)}</small></div></div>
            <div class="card" style="margin-top:14px"><form method="get" action="/cost-details" class="actions"><input name="search" value="{h(search)}" placeholder="Küpe veya takma ad ara" style="min-width:240px"><select name="status">{status_options}</select><select name="per_page">{per_options}</select><button class="btn">Filtrele</button><a class="btn alt" href="/cost-details">Temizle</a></form><p class="mut">{total_count} kayıt bulundu. Kesilen hayvanın maliyet günü kesim tarihinde durur; aktif hayvanın maliyeti bugüne kadar devam eder.</p></div>
            <div class="card" style="margin-top:14px"><div style="overflow-x:auto"><table><tr><th>Küpe</th><th>Durum</th><th>Başlangıç</th><th>Maliyet Bitişi</th><th>Maliyet Günü</th><th>Alış</th><th>Yem + Bakım</th><th>Toplam</th></tr>{detail_rows}</table></div><div class="actions" style="justify-content:center;margin-top:16px">{''.join(pager)}</div></div>'''
            return self.send_html(page('Maliyet Detayı',body,'',u,msg))

        if path=='/performance-settings':
            target=setting_float('male_min_daily_gain',1.0); ratio=setting_float('male_warning_ratio',0.90)
            body=f"""<h1>Besi Performans Ayarları</h1><div class="card setting-box"><form method="post" action="/performance-settings" class="form"><label>Minimum Günlük Canlı Ağırlık Artışı (kg/gün)<input type="number" min="0.01" step="0.01" name="male_min_daily_gain" value="{target:.2f}" required></label><label>Sarı Uyarı Başlangıcı (% hedef)<input type="number" min="1" max="100" step="1" name="warning_percent" value="{ratio*100:.0f}" required></label><div class="full"><p class="mut">Örnek: hedef 1,00 kg/gün ve sarı sınır %90 ise; 0,90-0,99 sarı, 0,90 altı kırmızı olur.</p><button class="btn">Ayarları Kaydet</button> <a class="btn alt" href="/performance">İptal</a></div></form></div>"""
            return self.send_html(page('Besi Performans Ayarları',body,path,u,msg))
        if path=='/paddocks':
            with db() as c:
                paddocks=c.execute("select * from paddocks where active=1 order by name").fetchall()
                rations=c.execute("select id,name from rations where active=1 order by name").fetchall()
                adults=c.execute("select id,tag,nickname,gender,paddock_id from animals where coalesce(status,'Aktif')='Aktif' order by tag").fetchall()
                calves=c.execute("select id,tag,nickname,gender,paddock_id from calves where promoted_animal_id is null order by tag").fetchall()
                rows=[]
                for pd in paddocks:
                    pop=paddock_population(pd['id'],c)
                    ar=c.execute("select pr.*,r.name ration_name from paddock_rations pr join rations r on r.id=pr.ration_id where pr.paddock_id=? and pr.active=1 and (pr.end_date is null or pr.end_date='') order by pr.id desc limit 1",(pd['id'],)).fetchone()
                    sm=ration_summary(ar['ration_id'],c) if ar else None
                    cap=int(pd['capacity'] or 0); doluluk=(pop/cap*100) if cap else 0
                    rows.append(f'''<tr><td><b>{h(pd['name'])}</b><div class="mut">{h(pd['code']) or '-'}</div></td><td>{h(pd['type']) or 'Genel'}</td><td>{pop}{('/'+str(cap)) if cap else ''}</td><td>{f'{doluluk:.0f}%' if cap else '-'}</td><td>{h(ar['ration_name']) if ar else '-'}</td><td>{(money(sm['cost'])+'/baş/gün · '+money(sm['cost']*pop)+'/padok/gün') if sm else '-'}</td><td>{h(pd['notes']) or '-'}</td></tr>''')
                pd_opts=''.join(f'<option value="{x["id"]}">{h(x["name"])}</option>' for x in paddocks)
                ration_opts=''.join(f'<option value="{x["id"]}">{h(x["name"])}</option>' for x in rations)
                animal_opts=''.join(f'<option value="animal:{x["id"]}">🐄 {h(x["tag"])} · {h(x["nickname"])} · {h(x["gender"])}</option>' for x in adults)+''.join(f'<option value="calf:{x["id"]}">🐮 {h(x["tag"])} · {h(x["nickname"])} · Buzağı</option>' for x in calves)
            body=f'''<h1>🏠 Padok Yönetimi</h1><p class="mut">Hayvanları padoklara yerleştirin; rasyonu padoka bağlayınca günlük yem ihtiyacı ve maliyet otomatik hesaplanır.</p>
            <div class="two"><div class="card"><h2>➕ Yeni Padok</h2><form method="post" action="/paddock/create" class="form"><label>Padok Adı<input name="name" required placeholder="Besi B-01"></label><label>Kod<input name="code" placeholder="B01"></label><label>Tür<select name="type"><option>Genel</option><option>Besi</option><option>Dişi</option><option>Buzağı</option><option>Doğum</option><option>Karantina</option></select></label><label>Kapasite<input type="number" min="0" name="capacity" value="0"></label><label class="full">Not<textarea name="notes"></textarea></label><div class="full"><button class="btn">Padoku Kaydet</button></div></form></div>
            <div class="card"><h2>🐄 Hayvanı Padoka Ata</h2><form method="post" action="/paddock/assign" class="form"><label class="full">Hayvan<select name="animal_ref" required><option value="">Seçin</option>{animal_opts}</select></label><label class="full">Padok<select name="paddock_id"><option value="">Padoksuz</option>{pd_opts}</select></label><label class="full">Taşıma Notu<input name="notes" placeholder="Grup değişimi"></label><div class="full"><button class="btn blue">Padoka Ata / Taşı</button></div></form></div></div>
            <div class="card" style="margin-top:14px"><h2>🥣 Padoka Rasyon Ata</h2><form method="post" action="/ration/assign" class="form"><label>Padok<select name="paddock_id" required><option value="">Seçin</option>{pd_opts}</select></label><label>Rasyon<select name="ration_id" required><option value="">Seçin</option>{ration_opts}</select></label><label>Başlangıç<input type="date" name="start_date" value="{date.today().isoformat()}" required></label><label>Not<input name="notes"></label><div class="full"><button class="btn orange">Rasyonu Padoka Ata</button> <a class="btn alt" href="/rations">Rasyon Yönetimi →</a></div></form></div>
            <div class="card" style="margin-top:14px;overflow:auto"><h2>Padoklar</h2><table><tr><th>Padok</th><th>Tür</th><th>Hayvan</th><th>Doluluk</th><th>Aktif Rasyon</th><th>Yem Maliyeti</th><th>Not</th></tr>{''.join(rows) if rows else '<tr><td colspan="7">Henüz padok tanımlanmadı.</td></tr>'}</table></div>'''
            return self.send_html(page('Padok Yönetimi',body,'/paddocks',u,msg))
        if path=='/feeds':
            search=(q.get('q',[''])[0] or '').strip()
            with db() as c:
                params=[]; where='where f.active=1'
                if search: where+=' and (f.name like ? or f.category like ?)';params=[f'%{search}%',f'%{search}%']
                feeds=c.execute(f'''select f.*,coalesce((select fp.price_per_kg from feed_prices fp where fp.feed_id=f.id and fp.effective_date<=? order by fp.effective_date desc,fp.id desc limit 1),0) price,
                    coalesce((select sum(case when st.tx_type in ('Giriş','Sayım +') then st.quantity_kg when st.tx_type in ('Çıkış','Tüketim','Sayım -') then -st.quantity_kg else 0 end) from feed_stock_transactions st where st.feed_id=f.id),0) stock
                    from feed_catalog f {where} order by f.category,f.name limit 250''',[date.today().isoformat()]+params).fetchall()
                allfeeds=c.execute("select id,name from feed_catalog where active=1 order by name").fetchall()
                daily_use={}
                active_pr=c.execute("select paddock_id,ration_id from paddock_rations where active=1 and (end_date is null or end_date='')").fetchall()
                for pr in active_pr:
                    pop=paddock_population(pr['paddock_id'],c)
                    for it in c.execute("select feed_id,kg_per_head_day from ration_items where ration_id=?",(pr['ration_id'],)).fetchall():
                        daily_use[it['feed_id']]=daily_use.get(it['feed_id'],0.0)+pop*float(it['kg_per_head_day'] or 0)
                opts=''.join(f'<option value="{x["id"]}">{h(x["name"])}</option>' for x in allfeeds)
                trs=''.join(f'''<tr class="feed-catalog-row" data-search="{h((str(r['name'])+' '+str(r['category'] or '')+' '+str(r['source'] or '')).casefold())}"><td><b>{h(r['name'])}</b><div class="mut">{h(r['category'])}</div><small class="mut">{h(r['source']) or '-'}</small></td><td>{float(r['dm_pct'] or 0):.1f}</td><td>{float(r['cp_pct'] or 0):.1f}</td><td>{float(r['ndf_pct'] or 0):.1f}</td><td>{float(r['me_mcal_kg'] or 0):.2f}</td><td>{float(r['ca_pct'] or 0):.2f}</td><td>{float(r['p_pct'] or 0):.2f}</td><td><b>{money(r['price'])}/kg</b></td><td>{float(r['stock'] or 0):,.1f} kg</td><td>{daily_use.get(r['id'],0):,.1f} kg</td><td>{(f"{float(r['stock'] or 0)/daily_use.get(r['id'],1):.0f} gün" if daily_use.get(r['id'],0)>0 else '-')}</td><td><div class="actions" style="flex-wrap:nowrap"><a class="btn alt compact-btn" href="/feed-edit?id={r['id']}">✏️ Düzenle</a><form method="post" action="/feed/delete" style="margin:0" onsubmit="return confirm('Bu yemi katalogdan kaldırmak istediğinize emin misiniz? Geçmiş rasyon ve fiyat kayıtları korunur.')"><input type="hidden" name="feed_id" value="{r['id']}"><button class="btn red compact-btn">🗑 Sil</button></form></div></td></tr>''' for r in feeds)
            body=f'''<h1>🌾 Yem Kataloğu & Stok</h1><p class="mut">Besin değerleri NASEM 2016 Beef + NASEM 2021 Dairy ile karşılaştırmalı güncelleniyor. Tam eşleşmeyen özel yemlerde mevcut referans korunur; kendi laboratuvar analizinizi Düzenle ile girebilirsiniz. Eski fiyatlar aktarılmadı.</p>
            <div class="grid"><div class="card stat metric"><span>Yem Kataloğu</span><b>{len(allfeeds)}</b></div><div class="card stat metric blue"><span>Gösterilen</span><b>{len(feeds)}</b></div><div class="card stat metric orange"><span>Fiyat Mantığı</span><b>Geçmişli</b><small>Her tarih kendi fiyatını korur</small></div></div>
            <div class="two" style="margin-top:14px"><div class="card"><h2>💰 Güncel Fiyat Gir</h2><form method="post" action="/feed/price" class="form"><label class="full">Yem<select name="feed_id" required><option value="">Seçin</option>{opts}</select></label><label>Tarih<input type="date" name="effective_date" value="{date.today().isoformat()}" required></label><label>₺ / kg<input type="number" step="0.0001" min="0" name="price_per_kg" required></label><label class="full">Not<input name="notes" placeholder="Tedarikçi / alım notu"></label><div class="full"><button class="btn">Fiyatı Kaydet</button></div></form></div>
            <div class="card"><h2>📦 Stok Hareketi</h2><form method="post" action="/feed/stock" class="form" id="feedStockForm"><label class="full">Yem<select name="feed_id" required><option value="">Seçin</option>{opts}</select></label><label>Tür<select name="tx_type" id="feedStockType"><option>Giriş</option><option>Çıkış</option><option>Tüketim</option><option>Sayım +</option><option>Sayım -</option></select></label><label>Miktar (kg)<input type="number" step="0.01" min="0.01" inputmode="decimal" name="quantity_kg" required></label><label>Tarih<input type="date" name="tx_date" value="{date.today().isoformat()}" required></label><label>Alış ₺/kg<input type="number" step="0.0001" min="0" name="unit_price" value="0"></label><label>Ödeme Yöntemi<select name="payment_method"><option>Nakit</option><option>Banka</option><option>Kredi Kartı</option><option>Vadeli</option></select></label><label class="full" id="feedFinanceAsk" style="padding:10px;background:#edf7f0;border:1px solid #cfe3d5;border-radius:10px"><input type="checkbox" name="post_to_finance" value="yes" style="width:auto;margin-right:8px"> Bu stok girişini <b>Finans → Gider / Yem</b> olarak da kaydet</label><label class="full">Not<input name="notes"></label><div class="full"><button class="btn blue">Stok Hareketini Kaydet</button></div></form></div></div>
            <div class="card" style="margin-top:14px"><details><summary><b>➕ Katalogda olmayan özel yem ekle</b></summary><form method="post" action="/feed/create" class="form" style="margin-top:14px"><label>Yem Adı<input name="name" required></label><label>Kategori<input name="category" value="Özel Yem"></label><label>KM %<input type="number" min="0" max="100" step="0.01" name="dm_pct"></label><label>HP % KM<input type="number" min="0" max="100" step="0.01" name="cp_pct"></label><label>NDF % KM<input type="number" min="0" max="100" step="0.01" name="ndf_pct"></label><label>eNDF etkinliği %<input type="number" min="0" max="100" step="0.01" name="effective_ndf_pct"></label><label>Nişasta % KM<input type="number" min="0" max="100" step="0.01" name="starch_pct"></label><label>ME Mcal/kg KM<input type="number" min="0" step="0.001" name="me_mcal_kg"></label><label>Ca % KM<input type="number" min="0" max="100" step="0.001" name="ca_pct"></label><label>P % KM<input type="number" min="0" max="100" step="0.001" name="p_pct"></label><details class="full"><summary><b>🏷 Üretici etiketi (ürün bazında)</b></summary><div class="form" style="margin-top:12px"><label>Etiket HP %<input type="number" min="0" max="100" step="0.01" name="label_cp_pct_as_fed"></label><label>Etiket ME kcal/kg<input type="number" min="0" step="1" name="label_me_kcal_kg_as_fed"></label><label>Etiket ham selüloz %<input type="number" min="0" max="100" step="0.01" name="label_crude_fiber_pct_as_fed"></label><label>Etiket ham yağ %<input type="number" min="0" max="100" step="0.01" name="label_fat_pct_as_fed"></label><label>Etiket ham kül %<input type="number" min="0" max="100" step="0.01" name="label_ash_pct_as_fed"></label><label>Etiket sodyum %<input type="number" min="0" max="100" step="0.01" name="label_sodium_pct_as_fed"></label></div><p class="mut">Bu alanlar çuval/üretici değerlerini aynen saklar. Solver alanları yukarıda KM bazında ayrıca girilir.</p></details><details class="full"><summary><b>🧪 İleri analiz ve etiket sınırları</b></summary><div class="form" style="margin-top:12px"><label>Nişasta rumen yıkılabilirliği %<input type="number" min="0" max="100" step="0.01" name="starch_degradability_pct"></label><label>NDF sindirilebilirliği %<input type="number" min="0" max="100" step="0.01" name="ndf_digestibility_pct"></label><label>İşleme biçimi<input name="processing_method" placeholder="Ezme / kırma / flake / öğütme"></label><label>Etiket alt doz kg/baş/gün<input type="number" min="0" step="0.01" name="solver_min_kg_day"></label><label>Etiket üst doz kg/baş/gün<input type="number" min="0" step="0.01" name="solver_max_kg_day"></label><label>Sınır kaynağı<input name="constraint_source" placeholder="Ürün etiketi / uzman / laboratuvar"></label></div><p class="mut">Bilinmeyen alanları 0 bırakın. Etiket üst dozu girilirse solver bu miktarı aşmaz.</p></details><div class="full"><button class="btn">Özel Yemi Ekle</button></div></form></details></div>
            <div class="card" style="margin-top:14px;overflow:auto"><form class="actions" id="feed-catalog-search-form"><input id="feed-catalog-search" name="q" value="{h(search)}" placeholder="Yem ara... yazdıkça filtrelenir" autocomplete="off"><button class="btn alt">🔎 Ara</button><a class="btn alt" href="/feeds">Temizle</a><span class="mut" id="feed-search-count"></span></form><table id="feed-catalog-table"><tr><th>Yem</th><th>KM%</th><th>HP%</th><th>NDF%</th><th>ME</th><th>Ca%</th><th>P%</th><th>Fiyat</th><th>Stok</th><th>Günlük Kullanım</th><th>Tahmini Yeterlilik</th><th>İşlem</th></tr>{trs or '<tr><td colspan="12">Kayıt bulunamadı.</td></tr>'}</table></div><script>(()=>{{const i=document.getElementById('feed-catalog-search'),rows=[...document.querySelectorAll('.feed-catalog-row')],count=document.getElementById('feed-search-count');if(!i)return;const norm=v=>(v||'').toLocaleLowerCase('tr-TR').normalize('NFD').replace(/[\u0300-\u036f]/g,'').replace(/ı/g,'i').replace(/ş/g,'s').replace(/ğ/g,'g').replace(/ü/g,'u').replace(/ö/g,'o').replace(/ç/g,'c');function run(){{const t=norm(i.value.trim());let n=0;rows.forEach(r=>{{const ok=!t||norm(r.dataset.search||r.textContent).includes(t);r.style.display=ok?'':'none';if(ok)n++;}});if(count)count.textContent=n+' yem gösteriliyor';}}i.addEventListener('input',run);run();}})();</script>'''
            return self.send_html(page('Yem Kataloğu',body,'/feeds',u,msg))
        if path=='/feed-edit':
            try: fid=int((q.get('id',['0'])[0] or 0))
            except: fid=0
            with db() as c: feed=c.execute('select * from feed_catalog where id=? and active=1',(fid,)).fetchone()
            if not feed:return self.redirect('/feeds','Yem bulunamadı.')
            def fv(k,dec=3):
                try:return f"{float(feed[k] or 0):.{dec}f}"
                except:return '0'
            body=f'''<div class="pro-form-head"><div><h1>✏️ Yemi Düzenle</h1><div class="mut">Sistem kataloğundaki veya sizin eklediğiniz yemin besin değerlerini güncelleyebilirsiniz.</div></div><a class="btn alt" href="/feeds">← Yem Kataloğu</a></div>
            <div class="card"><form method="post" action="/feed/edit" class="form">
            <input type="hidden" name="feed_id" value="{fid}"><label>Yem Adı<input name="name" value="{h(feed['name'])}" required></label><label>Kategori<input name="category" value="{h(feed['category'])}"></label>
            <label>KM %<input type="number" step="0.01" name="dm_pct" value="{fv('dm_pct',2)}"></label><label>HP % KM<input type="number" step="0.01" name="cp_pct" value="{fv('cp_pct',2)}"></label><label>NDF % KM<input type="number" step="0.01" name="ndf_pct" value="{fv('ndf_pct',2)}"></label><label>TDN % KM<input type="number" step="0.01" name="tdn_pct" value="{fv('tdn_pct',2)}"></label>
            <label>ME Mcal/kg KM<input type="number" step="0.001" name="me_mcal_kg" value="{fv('me_mcal_kg')}"></label><label>NEm Mcal/kg KM<input type="number" step="0.001" name="nem_mcal_kg" value="{fv('nem_mcal_kg')}"></label><label>NEg Mcal/kg KM<input type="number" step="0.001" name="neg_mcal_kg" value="{fv('neg_mcal_kg')}"></label>
            <label>Nişasta % KM<input type="number" step="0.01" name="starch_pct" value="{fv('starch_pct',2)}"></label><label>Yağ % KM<input type="number" step="0.01" name="fat_pct" value="{fv('fat_pct',2)}"></label><label>Kül % KM<input type="number" step="0.01" name="ash_pct" value="{fv('ash_pct',2)}"></label>
            <label>eNDF etkinliği %<input type="number" min="0" max="100" step="0.01" name="effective_ndf_pct" value="{fv('effective_ndf_pct',2)}"></label>
            <label>Ca % KM<input type="number" step="0.001" name="ca_pct" value="{fv('ca_pct')}"></label><label>P % KM<input type="number" step="0.001" name="p_pct" value="{fv('p_pct')}"></label><label>Mg % KM<input type="number" step="0.001" name="mg_pct" value="{fv('mg_pct')}"></label><label>K % KM<input type="number" step="0.001" name="k_pct" value="{fv('k_pct')}"></label><label>Na % KM<input type="number" step="0.001" name="na_pct" value="{fv('na_pct')}"></label><label>S % KM<input type="number" step="0.001" name="s_pct" value="{fv('s_pct')}"></label>
            <details class="full" open><summary><b>🏷 Üretici etiketi / ürün bazındaki değerler</b></summary><div class="form" style="margin-top:12px">
            <label>Etiket HP %<input type="number" min="0" max="100" step="0.01" name="label_cp_pct_as_fed" value="{fv('label_cp_pct_as_fed',2)}"></label><label>Etiket ME kcal/kg<input type="number" min="0" step="1" name="label_me_kcal_kg_as_fed" value="{fv('label_me_kcal_kg_as_fed',0)}"></label><label>Etiket ham selüloz %<input type="number" min="0" max="100" step="0.01" name="label_crude_fiber_pct_as_fed" value="{fv('label_crude_fiber_pct_as_fed',2)}"></label><label>Etiket ham yağ %<input type="number" min="0" max="100" step="0.01" name="label_fat_pct_as_fed" value="{fv('label_fat_pct_as_fed',2)}"></label><label>Etiket ham kül %<input type="number" min="0" max="100" step="0.01" name="label_ash_pct_as_fed" value="{fv('label_ash_pct_as_fed',2)}"></label><label>Etiket sodyum %<input type="number" min="0" max="100" step="0.01" name="label_sodium_pct_as_fed" value="{fv('label_sodium_pct_as_fed',2)}"></label>
            </div><p class="mut">Çuval veya üretici bilgisini aynen gösterir. Yukarıdaki HP/enerji/mineral alanları solverın kullandığı KM bazındaki değerlerdir.</p></details>
            <details class="full"><summary><b>🧪 İleri bilimsel analiz ve kesin ürün sınırları</b></summary><div class="form" style="margin-top:12px">
            <label>Nişasta rumen yıkılabilirliği %<input type="number" min="0" max="100" step="0.01" name="starch_degradability_pct" value="{fv('starch_degradability_pct',2)}"></label><label>NDF sindirilebilirliği %<input type="number" min="0" max="100" step="0.01" name="ndf_digestibility_pct" value="{fv('ndf_digestibility_pct',2)}"></label>
            <label>RDP % HP<input type="number" min="0" max="100" step="0.01" name="rdp_pct_cp" value="{fv('rdp_pct_cp',2)}"></label><label>RUP % HP<input type="number" min="0" max="100" step="0.01" name="rup_pct_cp" value="{fv('rup_pct_cp',2)}"></label>
            <label>INRA UFV<input type="number" min="0" step="0.001" name="inra_ufv" value="{fv('inra_ufv')}"></label><label>INRA PDI g/kg KM<input type="number" step="0.1" name="inra_pdi_g_kg_dm" value="{fv('inra_pdi_g_kg_dm',1)}"></label><label>INRA PDIA g/kg KM<input type="number" step="0.1" name="inra_pdia_g_kg_dm" value="{fv('inra_pdia_g_kg_dm',1)}"></label><label>INRA RPB g/kg KM<input type="number" step="0.1" name="inra_rpb_g_kg_dm" value="{fv('inra_rpb_g_kg_dm',1)}"></label><label>INRA doluluk birimi<input type="number" min="0" step="0.001" name="inra_fill_unit" value="{fv('inra_fill_unit')}"></label>
            <label>İşleme biçimi<input name="processing_method" value="{h(_rowval(feed,'processing_method',''))}" placeholder="Ezme / kırma / flake / öğütme"></label><label>Etiket alt doz kg/baş/gün<input type="number" min="0" step="0.01" name="solver_min_kg_day" value="{fv('solver_min_kg_day',2)}"></label><label>Etiket üst doz kg/baş/gün<input type="number" min="0" step="0.01" name="solver_max_kg_day" value="{fv('solver_max_kg_day',2)}"></label><label>Sınır kaynağı<input name="constraint_source" value="{h(_rowval(feed,'constraint_source',''))}" placeholder="Ürün etiketi / uzman / laboratuvar"></label>
            </div><p class="mut">0, “veri bilinmiyor” demektir. INRA alanları NASEM gereksinim hesabına karıştırılmaz; veri kapsamı ve yem doğrulaması için ayrı kullanılır. Etiket dozları kesin solver sınırıdır.</p></details>
            <label class="full">Kaynak / Referans<input name="source" value="{h(feed['source'])}" placeholder="NASEM 2016 / NASEM 2021 / Laboratuvar analizi / Kullanıcı girişi"></label>
            <div class="full"><button class="btn">Değişiklikleri Kaydet</button> <a class="btn alt" href="/feeds">İptal</a></div></form></div>'''
            return self.send_html(page('Yemi Düzenle',body,'/feeds',u,msg))
        if path=='/ration-prep-report':
            selected=int((q.get('id',['0'])[0] or 0))
            if not selected:
                return self.redirect('/rations','Rasyon seçilmedi.')
            with db() as c:
                rr=c.execute("select * from rations where id=? and active=1",(selected,)).fetchone()
                if not rr:
                    return self.redirect('/rations','Rasyon bulunamadı.')
                sm=ration_summary(selected,c)
                targets=ration_targets_for_record(rr)
            profile=farm_profile();logo=f'<img src="{h(profile.get("farm_logo"))}" alt="İşletme logosu">' if profile.get('farm_logo') else '<span class="prep-print-logo">🐄</span>'
            rows=''.join(f'''<tr class="prep-separate-row" data-kg="{float(x['kg_per_head_day'] or 0):.8f}" data-price="{float(x['price'] or 0):.8f}"><td><b>{h(x['name'])}</b></td><td>{float(x['kg_per_head_day'] or 0):.2f} kg</td><td class="prep-day">0,00 kg</td><td class="prep-period">0,00 kg</td><td class="prep-cost">₺0,00</td></tr>''' for x in sm['items']) or '<tr><td colspan="5">Rasyonda yem bulunmuyor.</td></tr>'
            body=f'''<div class="workbench-page-head"><div><a class="btn alt compact-btn" href="/ration-workbench?id={selected}">← Çalışma Masasına Dön</a><h1 style="margin:10px 0 4px">🧾 Rasyon Hazırlama / Toplam Yem Raporu</h1><p class="mut">Kaç hayvan ve kaç günlük yem hazırlayacağınızı girin; toplam ihtiyaç otomatik hesaplanır.</p></div></div>
            <div class="card ration-prep-report prep-separate-page">
              <div class="prep-print-header print-only"><div class="prep-print-brand">{logo}<div><h1>{h(farm_display_name(profile))}</h1><b>Rasyon Hazırlama / Toplam Yem Raporu</b></div></div><div><b>Rapor Tarihi</b><br>{date.today().strftime('%d/%m/%Y')}<br><b>İşletme No</b><br>{h(profile.get('business_no') or '-')}</div></div>
              <div class="prep-report-head"><div><h2 style="margin:0">🥣 {h(rr['name'])}</h2><span class="mut">Kaydedilmiş rasyon miktarları kullanılır.</span></div><div class="prep-report-controls no-print"><label>Hayvan Sayısı<input id="ration-animal-count" type="number" min="1" step="1" value="1" inputmode="numeric"></label><label>Dönem<select id="ration-period-days"><option value="1">1 Gün</option><option value="7">7 Gün</option><option value="30">30 Gün</option></select></label><button type="button" class="btn blue" id="ration-prep-print">🖨 Yazdır / PDF</button></div></div>
              <div class="prep-report-summary"><div><span>Hayvan Sayısı</span><b id="prep-headcount">1</b></div><div><span>Dönem</span><b id="prep-period-label">1 Gün</b></div><div><span>Toplam Yem</span><b id="prep-total-kg">{sm['as_fed_kg']:.2f} kg</b></div><div><span>Toplam Maliyet</span><b id="prep-total-cost">{money(sm['cost'])}</b></div></div>
              <div class="prep-report-summary prep-nutrition-summary"><div><span>Günlük KM</span><b>{sm['dm_kg']:.2f} kg/baş</b></div><div><span>Nişasta</span><b>%{sm['starch_pct_dm']:.1f} KM</b></div><div><span>Nişasta Miktarı</span><b>{sm['starch_kg']:.2f} kg/baş/gün</b></div><div><span>İdeal / Üst Sınır</span><b>%{targets['starch_min']:.0f}–{targets['starch_ideal_max']:.0f} / %{targets['starch_max']:.0f}</b></div></div>
              <div class="prep-report-table-wrap"><table class="prep-report-table"><thead><tr><th>Yem</th><th>kg/baş/gün</th><th>Toplam kg/gün</th><th>Dönem Toplamı</th><th>Dönem Maliyeti</th></tr></thead><tbody id="ration-prep-body">{rows}</tbody></table></div>
              <div class="mut prep-report-note">Rapor, kaydedilmiş rasyon reçetesine göre hazırlanır. Çalışma masasındaki kaydedilmemiş değişiklikleri raporlamadan önce kaydedin.</div>
              <div class="prep-print-footer print-only"><span>ÇiftlikPro Enterprise · {h(farm_display_name(profile))}</span><span>{date.today().strftime('%d/%m/%Y')}</span></div>
            </div>
            <style>.print-only{{display:none}}@page{{size:A4 portrait;margin:10mm}}@media print{{html,body{{height:auto!important;overflow:visible!important;background:#fff!important}}.top,.side,.erp-commandbar,.erp-tabs,.erp-statusbar,.no-print,.workbench-page-head{{display:none!important}}.layout{{display:block!important;min-height:0!important;padding:0!important}}.main{{margin:0!important;padding:0!important;min-height:0!important;background:#fff!important}}.card{{box-shadow:none!important}}.prep-separate-page{{border:0!important;padding:0!important;margin:0!important}}.print-only{{display:flex!important}}.prep-print-header{{align-items:center;justify-content:space-between;gap:20px;border-bottom:2px solid #176b3a;padding-bottom:9px;margin-bottom:12px;font-size:11px;text-align:right}}.prep-print-brand{{display:flex;align-items:center;gap:10px;text-align:left}}.prep-print-brand img,.prep-print-logo{{width:54px;height:54px;object-fit:contain;border-radius:7px}}.prep-print-logo{{display:grid!important;place-items:center;background:#edf6f0;font-size:28px}}.prep-print-brand h1{{font-size:19px;margin:0 0 3px}}.prep-report-summary{{grid-template-columns:repeat(4,1fr)!important;break-inside:avoid}}.prep-report-table-wrap{{overflow:visible!important}}.prep-report-table{{font-size:10px!important}}.prep-report-table thead{{display:table-header-group}}.prep-report-table tr{{break-inside:avoid;page-break-inside:avoid}}.prep-report-table th,.prep-report-table td{{padding:5px 6px!important}}.prep-print-footer{{justify-content:space-between;border-top:1px solid #d8e3db;margin-top:10px;padding-top:7px;font-size:9px;color:#68766d}}body{{print-color-adjust:exact;-webkit-print-color-adjust:exact}}}}</style>
            <script>(()=>{{const count=document.getElementById('ration-animal-count'),period=document.getElementById('ration-period-days'),rows=[...document.querySelectorAll('.prep-separate-row')];if(!count||!period)return;const nf=(n,d=2)=>Number(n||0).toLocaleString('tr-TR',{{minimumFractionDigits:d,maximumFractionDigits:d}}),money=n=>'₺'+nf(n,2);function update(){{const n=Math.max(1,parseInt(count.value||'1',10)||1),days=Math.max(1,parseInt(period.value||'1',10)||1);count.value=String(n);let totalKg=0,totalCost=0;rows.forEach(r=>{{const kg=parseFloat(r.dataset.kg||0)||0,price=parseFloat(r.dataset.price||0)||0,dayKg=kg*n,periodKg=dayKg*days,cost=periodKg*price;totalKg+=periodKg;totalCost+=cost;r.querySelector('.prep-day').textContent=nf(dayKg)+' kg';r.querySelector('.prep-period').textContent=nf(periodKg)+' kg';r.querySelector('.prep-cost').textContent=money(cost);}});document.getElementById('prep-headcount').textContent=String(n);document.getElementById('prep-period-label').textContent=days+' Gün';document.getElementById('prep-total-kg').textContent=nf(totalKg)+' kg';document.getElementById('prep-total-cost').textContent=money(totalCost);localStorage.setItem('cp-ration-headcount-{selected}',String(n));localStorage.setItem('cp-ration-period-{selected}',String(days));}}const savedN=parseInt(localStorage.getItem('cp-ration-headcount-{selected}')||'1',10),savedD=parseInt(localStorage.getItem('cp-ration-period-{selected}')||'1',10);if(savedN>0)count.value=String(savedN);if([1,7,30].includes(savedD))period.value=String(savedD);count.addEventListener('input',update);period.addEventListener('change',update);document.getElementById('ration-prep-print').addEventListener('click',()=>{{update();window.print();}});update();}})();</script>'''
            return self.send_html(page('Rasyon Hazırlama Raporu',body,'/rations',u,msg))
        if path in ('/rations','/ration-workbench'):
            is_workbench=(path=='/ration-workbench')
            selected=int((q.get('id',['0'])[0] or 0))
            if path=='/rations' and selected:
                keep={k:v[0] for k,v in q.items() if k!='msg' and v and v[0] not in ('',None)}
                keep['id']=str(selected)
                return self.redirect('/ration-workbench?'+urllib.parse.urlencode(keep),msg or '')
            with db() as c:
                rations=c.execute("select * from rations where active=1 order by name").fetchall()
                feeds=c.execute("""select f.id,f.name,f.category,f.dm_pct,f.cp_pct,f.ndf_pct,f.effective_ndf_pct,f.starch_pct,f.me_mcal_kg,f.ca_pct,f.p_pct, coalesce((select fp.price_per_kg from feed_prices fp where fp.feed_id=f.id and fp.effective_date<=? order by fp.effective_date desc,fp.id desc limit 1),0) price, coalesce((select sum(case when st.tx_type in ('Giriş','Sayım +') then st.quantity_kg when st.tx_type in ('Çıkış','Tüketim','Sayım -') then -st.quantity_kg else 0 end) from feed_stock_transactions st where st.feed_id=f.id),0) stock from feed_catalog f where f.active=1 order by f.name""",(date.today().isoformat(),)).fetchall()
                paddocks=c.execute("select id,name from paddocks where active=1 order by name").fetchall()
                cards=[]
                for r in rations:
                    sm=ration_summary(r['id'],c)
                    cards.append(f'''<a class="card ration-picker-card{' active' if selected==r['id'] else ''}" href="/ration-workbench?id={r['id']}"><div class="ration-picker-head"><h3>🥣 {h(r['name'])}</h3><span class="ration-picker-type">{('🥛 Süt' if ('ration_type' in r.keys() and (r['ration_type'] or '').startswith('Süt')) else '🥩 Besi')}</span></div><div class="ration-picker-main"><b>{sm['as_fed_kg']:.2f} kg</b><span>·</span><b>{money(sm['cost'])}/gün</b></div><div class="ration-picker-ratio">🌾 K/K <b>%{sm['roughage_pct_dm']:.0f} / %{sm['concentrate_pct_dm']:.0f}</b></div></a>''')
                detail=''
                if selected:
                    rr=c.execute("select * from rations where id=?",(selected,)).fetchone()
                    if rr:
                        sm=ration_summary(selected,c)
                        item_rows=''.join(f'''<tr class="ration-row" data-feed-name="{h(x['name'])}" data-dm="{float(x['dm_pct'] or 0):.8f}" data-cp="{float(x['cp_pct'] or 0):.8f}" data-ndf="{float(x['ndf_pct'] or 0):.8f}" data-endf="{float(x['effective_ndf_pct'] or 0):.8f}" data-starch="{_solver_starch_pct(x):.8f}" data-starch-deg="{float(_rowval(x,'starch_degradability_pct',0) or 0):.8f}" data-me="{float(x['me_mcal_kg'] or 0):.8f}" data-nem="{float(x['nem_mcal_kg'] or 0):.8f}" data-neg="{float(x['neg_mcal_kg'] or 0):.8f}" data-ca="{float(x['ca_pct'] or 0):.8f}" data-p="{float(x['p_pct'] or 0):.8f}" data-price="{float(x['price'] or 0):.8f}" data-group="{feed_group(x)}"><td><b>{h(x['name'])}</b></td><td><div class="ration-stepper"><button type="button" class="btn alt compact-btn qty-step" data-delta="-0.10">−</button><input class="ration-qty" type="number" min="0" step="0.01" name="item_{x['item_id']}" value="{float(x['kg_per_head_day']):.2f}" data-original="{float(x['kg_per_head_day']):.2f}"><button type="button" class="btn alt compact-btn qty-step" data-delta="0.10">+</button></div><small class="qty-delta mut"></small></td><td>{float(x['dm_pct'] or 0):.1f}%</td><td>{float(x['cp_pct'] or 0):.1f}%</td><td>{float(x['ndf_pct'] or 0):.1f}%</td><td>{money(x['price'])}/kg</td><td class="row-daily">{money(float(x['kg_per_head_day'])*float(x['price'] or 0))}</td><td><button type="button" class="btn red compact-btn qty-zero">Çıkar</button></td></tr>''' for x in sm['items']) or '<tr><td colspan="8">Henüz yem eklenmedi.</td></tr>'
                        feed_opts=''.join(f'<option value="{x["id"]}">{h(x["name"])}</option>' for x in feeds)
                        current_feed_qty={int(x['id']):float(x['kg_per_head_day'] or 0) for x in sm['items']}
                        quick_feed_rows=[]
                        for x in feeds:
                            fid=int(x['id']); price=float(x['price'] or 0); stock=float(x['stock'] or 0); cur=current_feed_qty.get(fid,0.0)
                            meta=f"{h(x['category'] or 'Yem')} · KM %{float(x['dm_pct'] or 0):.1f} · HP %{float(x['cp_pct'] or 0):.1f} · NDF %{float(x['ndf_pct'] or 0):.1f}"
                            price_txt=(money(price)+'/kg') if price>0 else 'Fiyat girilmemiş'
                            stock_txt=(f"{stock:,.0f} kg stok" if stock>0 else 'Stok bilinmiyor')
                            in_ration=(f" · Rasyonda {cur:.2f} kg" if cur>0 else '')
                            search_key=h((str(x['name'])+' '+str(x['category'] or '')).casefold())
                            quick_feed_rows.append(f"""<button type='button' class='quick-feed-result' data-feed-id='{fid}' data-feed-name='{h(x['name'])}' data-current='{cur:.2f}' data-search='{search_key}'><span><b>{h(x['name'])}</b><small>{meta}</small></span><span class='quick-feed-side'><b>{price_txt}</b><small>{stock_txt}{in_ration}</small></span></button>""")
                        quick_feed_html=''.join(quick_feed_rows)
                        pd_opts=''.join(f'<option value="{x["id"]}">{h(x["name"])}</option>' for x in paddocks)
                        add_recs=ration_addition_recommendations(rr,sm,c,30)
                        reduce_recs=ration_reduction_recommendations(rr,sm,c,8)
                        combo_recs=ration_combined_recommendations(rr,sm,reduce_recs,add_recs,5)
                        def price_label(fx):
                            return (money(fx['price'])+'/kg') if float(fx['price'] or 0)>0 else '<span class="mut">Fiyat girilmemiş</span>'
                        targets=ration_targets_for_record(rr)
                        add_top_html=''.join(f'''<tr><td><b>{h(fx['name'])}</b><div class="mut">{h(fx['category'])}</div></td><td><b>{h(ration_effect_text(targets,sm,ss,2))}</b><div class="mut">{h(reason)}</div>{('<div class="orange">⚠ '+h(warn)+'</div>') if warn else ''}</td><td>{price_label(fx)}</td><td>{(f"{float(fx['stock'] or 0):,.0f} kg" if float(fx['stock'] or 0)>0 else '<span class="mut">Stok yok/bilinmiyor</span>')}</td><td><a class="btn alt compact-btn" href="/ration-workbench?id={selected}&sim_feed={fx['id']}&sim_delta=0.50#smart-balance">+0,50 kg Simüle Et</a></td></tr>''' for score,fx,reason,warn,ss in add_recs[:8]) or '<tr><td colspan="5">Mevcut rasyonu belirgin biçimde iyileştiren ek yem adayı bulunmadı.</td></tr>'
                        add_all_html=''.join(f'''<tr><td><b>{h(fx['name'])}</b><div class="mut">{h(fx['category'])}</div></td><td>{h(ration_effect_text(targets,sm,ss,2))}</td><td>{price_label(fx)}</td><td>{(f"{float(fx['stock'] or 0):,.0f} kg" if float(fx['stock'] or 0)>0 else '-')}</td><td><a class="btn alt compact-btn" href="/ration-workbench?id={selected}&sim_feed={fx['id']}&sim_delta=0.50#smart-balance">Simüle Et</a></td></tr>''' for score,fx,reason,warn,ss in add_recs)
                        reduce_html=''.join(f'''<tr><td><b>{h(fx['name'])}</b></td><td><b>{delta:.2f} kg</b></td><td><b>{h(ration_effect_text(targets,sm,ss,2))}</b><div class="mut">{h(reason)}</div>{('<div class="orange">⚠ '+h(warn)+'</div>') if warn else ''}</td><td>{money(ss['cost'])}/baş/gün</td><td><a class="btn alt compact-btn" href="/ration-workbench?id={selected}&sim_feed={fx['id']}&sim_delta={delta:.2f}#smart-balance">Simüle Et</a></td></tr>''' for score,fx,delta,reason,warn,ss in reduce_recs) or '<tr><td colspan="5">Azaltılması genel dengeyi iyileştiren belirgin bir mevcut yem bulunmadı.</td></tr>'
                        combo_html=''.join(f'''<tr><td><b>{h(red[1]['name'])}</b> {red[2]:.2f} kg</td><td><b>{h(add[1]['name'])}</b> +0,50 kg</td><td>{h(effect)}</td><td><form method="post" action="/ration/apply-combo"><input type="hidden" name="ration_id" value="{selected}"><input type="hidden" name="red_feed_id" value="{red[1]['id']}"><input type="hidden" name="red_delta" value="{red[2]:.2f}"><input type="hidden" name="add_feed_id" value="{add[1]['id']}"><input type="hidden" name="add_delta" value="0.50"><button class="btn blue compact-btn">Birlikte Uygula</button></form></td></tr>''' for score,red,add,combo_ss,effect in combo_recs) or '<tr><td colspan="4">Şimdilik anlamlı bir azalt + ekle kombinasyonu bulunmadı.</td></tr>'
                        quick_solutions=[]
                        if reduce_recs:
                            score,fx,delta,reason,warn,ss=reduce_recs[0]
                            quick_solutions.append(f'''<div class="smart-solution"><div><b>✂️ Fazlalığı Azalt</b><small>{h(fx['name'])} {delta:.2f} kg</small></div><div class="effect">{h(ration_effect_text(targets,sm,ss,3))}</div><p>{h(reason)}</p><a class="btn alt compact-btn" href="/ration-workbench?id={selected}&sim_feed={fx['id']}&sim_delta={delta:.2f}#ration-workbench">Çalışma Masasında Dene</a></div>''')
                        if add_recs:
                            score,fx,reason,warn,ss=add_recs[0]
                            quick_solutions.append(f'''<div class="smart-solution"><div><b>➕ Eksik Tamamla</b><small>{h(fx['name'])} +0,50 kg</small></div><div class="effect">{h(ration_effect_text(targets,sm,ss,3))}</div><p>{h(reason)}</p><a class="btn alt compact-btn" href="/ration-workbench?id={selected}&sim_feed={fx['id']}&sim_delta=0.50#ration-workbench">Çalışma Masasında Dene</a></div>''')
                        if combo_recs:
                            score,red,add,combo_ss,effect=combo_recs[0]
                            quick_solutions.append(f'''<div class="smart-solution"><div><b>⚖️ Kombine Dengele</b><small>{h(red[1]['name'])} {red[2]:.2f} kg · {h(add[1]['name'])} +0,50 kg</small></div><div class="effect">{h(effect)}</div><p>Maliyet: {money(sm['cost'])} → <b>{money(combo_ss['cost'])}</b>/baş/gün</p><form class="combo-apply" method="post" action="/ration/apply-combo"><input type="hidden" name="ration_id" value="{selected}"><input type="hidden" name="red_feed_id" value="{red[1]['id']}"><input type="hidden" name="red_delta" value="{red[2]:.2f}"><input type="hidden" name="add_feed_id" value="{add[1]['id']}"><input type="hidden" name="add_delta" value="0.50"><button class="btn blue compact-btn">✅ İkisini Birlikte Uygula</button></form></div>''')
                        quick_solutions_html=''.join(quick_solutions) or '<div class="mut">Rasyon hedef aralıklara yakın; belirgin bir düzeltme önerisi oluşmadı.</div>'
                        sim_html=''
                        try: sim_feed=int((q.get('sim_feed',['0'])[0] or 0));sim_delta=float((q.get('sim_delta',['0.5'])[0] or .5))
                        except: sim_feed=0;sim_delta=.5
                        if sim_feed:
                            sf=c.execute('select name from feed_catalog where id=?',(sim_feed,)).fetchone();ss=ration_simulated_summary(selected,sim_feed,sim_delta,c)
                            if sf: sim_html=f'''<div class="card" style="margin-top:14px;border:2px solid #2f78d0"><h3>🧪 Kaydetmeden Simülasyon: +{sim_delta:.2f} kg {h(sf['name'])}</h3><div class="grid"><div class="card stat">KM<b>{sm['dm_kg']:.2f} → {ss['dm_kg']:.2f}</b></div><div class="card stat">HP<b>%{sm['cp_pct_dm']:.1f} → %{ss['cp_pct_dm']:.1f}</b></div><div class="card stat">ME<b>{sm['me_mcal']:.1f} → {ss['me_mcal']:.1f}</b></div><div class="card stat">NDF<b>%{sm['ndf_pct_dm']:.1f} → %{ss['ndf_pct_dm']:.1f}</b></div><div class="card stat">Kaba/Kesif<b>%{sm['roughage_pct_dm']:.0f}/%{sm['concentrate_pct_dm']:.0f} → %{ss['roughage_pct_dm']:.0f}/%{ss['concentrate_pct_dm']:.0f}</b></div><div class="card stat metric orange">Maliyet<b>{money(sm['cost'])} → {money(ss['cost'])}</b></div></div><div class="actions"><form method="post" action="/ration/apply-suggestion"><input type="hidden" name="ration_id" value="{selected}"><input type="hidden" name="feed_id" value="{sim_feed}"><input type="hidden" name="delta" value="{sim_delta}"><button class="btn blue">✅ Bu Değişikliği Uygula</button></form><a class="btn alt" href="/rations?id={selected}">↩ Vazgeç</a></div></div>'''
                        detail=f'''<div class="card workbench-shell" style="margin-top:14px"><div style="display:flex;justify-content:space-between;align-items:center;gap:12px;flex-wrap:wrap"><div><h2 style="margin:0">🥣 {h(rr['name'])}</h2><span class="mut">Hedef ↔ mevcut → çalışma masası → yem ekle → akıllı çözüm</span></div><details><summary><b>✏️ Rasyon Bilgileri</b></summary><form method="post" action="/ration/edit" class="form" style="margin-top:12px"><input type="hidden" name="ration_id" value="{selected}"><label>Rasyon Adı<input name="name" value="{h(rr['name'])}" required></label><label>Hedef Grup<input name="target_group" value="{h(rr['target_group'])}"></label><label class="full">Not<input name="notes" value="{h(rr['notes'])}"></label><div class="full"><button class="btn">Değişiklikleri Kaydet</button></div></form><form method="post" action="/ration/delete" style="margin-top:10px" onsubmit="return confirm('Bu rasyonu silmek istediğinize emin misiniz?');"><input type="hidden" name="ration_id" value="{selected}"><button class="btn red compact-btn">🗑 Rasyonu Sil</button></form></details></div>
                        {ration_requirement_panel(rr,sm)}
                        <div id="ration-workbench" class="card" style="margin-top:14px"><div class="workbench-head"><div><h3 style="margin:0">🌾 Rasyon Çalışma Masası</h3><span class="mut">Miktarı yazın veya −/+ kullanın. Hedef ↔ mevcut değerler üstte canlı güncellenir.</span></div><div class="workbench-actions"><button type="button" class="btn" onclick="document.getElementById('quick-feed-add').open=true;setTimeout(()=>document.getElementById('quick-feed-search')?.focus(),80)">➕ Yem Ekle</button><a class="btn alt" href="/ration-prep-report?id={selected}">🧾 Rasyonu Yazdır</a></div></div><form method="post" action="/ration/items-bulk" id="ration-bulk-form"><input type="hidden" name="ration_id" value="{selected}"><div class="ration-changebar compact-changebar"><span id="dirty-status" class="mut">Kaydedilmiş rasyon gösteriliyor.</span><button type="button" id="ration-reset" class="btn alt compact-btn" style="display:none">↩ Değişiklikleri Geri Al</button></div><div style="overflow:auto;margin-top:8px"><table class="ration-workbench-table"><thead><tr><th>Yem</th><th>Miktar kg/baş/gün</th><th>KM</th><th>HP</th><th>NDF</th><th>₺/kg</th><th>Günlük</th><th></th></tr></thead><tbody>{item_rows}</tbody></table></div><div class="ration-savebar"><button class="btn blue" id="ration-save" disabled>💾 Değişiklikleri Kaydet</button></div></form></div><script>(()=>{{const form=document.getElementById('ration-bulk-form');if(!form)return;const scrollKey='cp-ration-scroll';const rows=[...form.querySelectorAll('.ration-row')],save=document.getElementById('ration-save'),reset=document.getElementById('ration-reset'),status=document.getElementById('dirty-status');const base={{asfed:{sm['as_fed_kg']:.8f},dm:{sm['dm_kg']:.8f},cp:{sm['cp_pct_dm']:.8f},me:{sm['me_mcal']:.8f},ndf:{sm['ndf_pct_dm']:.8f},endf:{sm['endf_pct_dm']:.8f},ca:{sm['ca_g']:.8f},p:{sm['p_g']:.8f},cost:{sm['cost']:.8f},rough:{sm['roughage_pct_dm']:.8f},conc:{sm['concentrate_pct_dm']:.8f}}};const targetsLive={{dm:{targets['dmi_kg']:.8f},cp:{targets['cp_pct']:.8f},me:{targets['me_mcal_day']:.8f},ndfMin:{targets['ndf_min']:.8f},ndfMax:{targets['ndf_max']:.8f},endfMin:{targets.get('endf_min',0):.8f},starchMin:{targets.get('starch_min',0):.8f},starchIdeal:{targets.get('starch_ideal_max',100):.8f},starchMax:{targets.get('starch_max',100):.8f},ca:{targets['ca_g']:.8f},p:{targets['p_g']:.8f},roughMin:{targets['roughage_min']:.8f},roughMax:{targets['roughage_max']:.8f}}};const setText=(id,v)=>{{const e=document.getElementById(id);if(e)e.textContent=v;}};function targetStatus(actual,target,tol,upperTol){{upperTol=upperTol===undefined?tol:upperTol;const r=target?actual/target:1;if(r<1-tol)return '⚠️ Eksik %'+((1-r)*100).toFixed(0);if(r>1+upperTol)return '⬆️ Fazla %'+((r-1)*100).toFixed(0);return '✅ Uygun';}}function liveDiff(actual,target,unit,digits){{if(!target)return '—';const d=actual-target,p=d/target*100,sg=d>0?'+':'';return sg+d.toFixed(digits)+unit+' ('+sg+p.toFixed(0)+'%)';}}function updateTargetCard(key,current,status,diff){{setText('target-mini-'+key+'-current',current);setText('target-mini-'+key+'-status',status);setText('target-mini-'+key+'-diff',diff);const b=document.getElementById('target-mini-'+key);if(b){{b.classList.toggle('ok',status.indexOf('Uygun')>=0);b.classList.toggle('warn',status.indexOf('Uygun')<0);}}}}function updateTargetCards(dm,cp,me,ndf,endf,starch,starchKg,rapidStarch,coverage,ca,p,rough,conc,changed){{const ts=targetsLive;const dmS=targetStatus(dm,ts.dm,.10),cpS=targetStatus(cp,ts.cp,.05,.10),meS=targetStatus(me,ts.me,.08,.10),caS=targetStatus(ca,ts.ca,.10),pS=targetStatus(p,ts.p,.10),ndfS=(ndf>=ts.ndfMin&&ndf<=ts.ndfMax)?'✅ Uygun':(ndf<ts.ndfMin?'⚠️ Düşük':'⚠️ Yüksek'),rcS=(rough>=ts.roughMin&&rough<=ts.roughMax)?'✅ Uygun':(rough<ts.roughMin?'⚠️ Kaba yem düşük':'⚠️ Kaba yem yüksek'),starchS=(starch>=ts.starchMin&&starch<=ts.starchIdeal)?'✅ Uygun':(starch<ts.starchMin?'ℹ️ İdeal altı':(starch<=ts.starchIdeal+.5?'⚠️ Sınırda':(starch<=ts.starchMax?'⚠️ Sınıra yakın':'🔴 Yüksek')));updateTargetCard('dm',dm.toFixed(2)+' kg',dmS,liveDiff(dm,ts.dm,' kg',2));updateTargetCard('cp','%'+cp.toFixed(1),cpS,liveDiff(cp,ts.cp,' puan',1));updateTargetCard('me',me.toFixed(1),meS,liveDiff(me,ts.me,' Mcal',1));updateTargetCard('ndf','%'+ndf.toFixed(1),ndfS,ndfS.indexOf('Uygun')>=0?'Aralık içi':(ndf<ts.ndfMin?'Alt sınırın altında':'Üst sınırın üzerinde'));updateTargetCard('ca',ca.toFixed(0)+' g',caS,liveDiff(ca,ts.ca,' g',0));updateTargetCard('p',p.toFixed(0)+' g',pS,liveDiff(p,ts.p,' g',0));const mineral=document.getElementById('target-mini-mineral');if(mineral){{const mok=caS.indexOf('Uygun')>=0&&pS.indexOf('Uygun')>=0;mineral.classList.toggle('ok',mok);mineral.classList.toggle('warn',!mok);}}updateTargetCard('rc','%'+rough.toFixed(0)+' / %'+conc.toFixed(0),rcS,rcS.indexOf('Uygun')>=0?'Aralık içi':(rough<ts.roughMin?'Kaba yem düşük':'Kaba yem yüksek'));let riskScore=0;if(starch>ts.starchMax)riskScore+=2;else if(starch>ts.starchIdeal+.5)riskScore+=1;if(endf<ts.endfMin*.80)riskScore+=2;else if(endf<ts.endfMin)riskScore+=1;if(coverage>=.70&&rapidStarch>ts.starchIdeal*.75)riskScore+=1;const risk=riskScore>=3?'Yüksek':(riskScore>=1?'Orta':'Düşük'),confidence=coverage>=.85?'Yüksek':(coverage>=.50?'Orta':'Düşük'),riskS=risk==='Düşük'?'✅ Düşük':(risk==='Orta'?'⚠️ Orta':'🔴 Yüksek');setText('target-mini-starch-current','%'+starch.toFixed(1));setText('target-mini-starch-status',starchS);setText('target-mini-starch-diff',starchKg.toFixed(2)+' kg/baş');setText('target-mini-ph-current',risk);setText('target-mini-ph-status',riskS);setText('target-mini-ph-diff','Veri güveni '+confidence);const sr=document.getElementById('target-mini-starch-rumen');if(sr){{const ok=starchS.indexOf('Uygun')>=0&&risk==='Düşük';sr.classList.toggle('ok',ok);sr.classList.toggle('warn',!ok);}}const note=document.getElementById('target-live-note');if(note)note.textContent=changed?'🟡 Kaydedilmemiş taslak canlı analiz ediliyor. Hedef kartları çalışma masasıyla birlikte değişiyor.':'🌾 Çalışma masası ile canlı bağlı: miktar değiştikçe bu kartlar anında güncellenir.';}}const trMoney=n=>'₺'+n.toLocaleString('tr-TR',{{minimumFractionDigits:2,maximumFractionDigits:2}});const delta=(id,n,b,suffix='')=>{{const e=document.getElementById(id),d=n-b;e.textContent=Math.abs(d)<0.005?'':((d>0?'+':'')+d.toFixed(2)+suffix);e.style.color=d>0?'#17733d':d<0?'#b33a2b':''}};function calc(){{let asfed=0,dm=0,cpkg=0,ndfkg=0,endfkg=0,starchkg=0,rapidkg=0,knownstarch=0,me=0,ca=0,p=0,cost=0,roughdm=0,concdm=0,changed=0;rows.forEach(r=>{{const i=r.querySelector('.ration-qty'),kg=Math.max(0,parseFloat((i.value||'0').replace(',','.'))||0),orig=parseFloat(i.dataset.original||0),dmp=parseFloat(r.dataset.dm||0),dmkg=kg*dmp/100;asfed+=kg;dm+=dmkg;cpkg+=dmkg*parseFloat(r.dataset.cp||0)/100;ndfkg+=dmkg*parseFloat(r.dataset.ndf||0)/100;endfkg+=dmkg*parseFloat(r.dataset.ndf||0)/100*parseFloat(r.dataset.endf||0)/100;const skg=dmkg*parseFloat(r.dataset.starch||0)/100,sd=parseFloat(r.dataset.starchDeg||0);starchkg+=skg;if(sd>0){{knownstarch+=skg;rapidkg+=skg*sd/100;}}me+=dmkg*parseFloat(r.dataset.me||0);ca+=dmkg*parseFloat(r.dataset.ca||0)*10;p+=dmkg*parseFloat(r.dataset.p||0)*10;cost+=kg*parseFloat(r.dataset.price||0);if(r.dataset.group==='Kaba')roughdm+=dmkg;else if(r.dataset.group==='Kesif')concdm+=dmkg;const ch=Math.abs(kg-orig)>.0005;i.classList.toggle('ration-dirty',ch);r.querySelector('.qty-delta').textContent=ch?((kg-orig>0?'+':'')+(kg-orig).toFixed(2)+' kg'):'';if(ch)changed++;r.querySelector('.row-daily').textContent=trMoney(kg*parseFloat(r.dataset.price||0));}});const cp=dm?cpkg/dm*100:0,ndf=dm?ndfkg/dm*100:0,endf=dm?endfkg/dm*100:0,starch=dm?starchkg/dm*100:0,rapidStarch=dm?rapidkg/dm*100:0,coverage=starchkg?knownstarch/starchkg:1,rcdm=roughdm+concdm,rough=rcdm?roughdm/rcdm*100:0,conc=rcdm?concdm/rcdm*100:0;const dc=cost-base.cost;updateTargetCards(dm,cp,me,ndf,endf,starch,starchkg,rapidStarch,coverage,ca,p,rough,conc,changed);setText('target-mini-cost-current',trMoney(cost));setText('target-mini-cost-diff',Math.abs(cost-base.cost)<.005?'Değişiklik yok':((cost-base.cost>0?'+':'')+trMoney(cost-base.cost)));setText('target-mini-cost-status',cost<=base.cost?'💰 Maliyet düştü':'💰 Maliyet arttı');save.disabled=changed===0;reset.style.display=changed?'inline-flex':'none';status.className=changed?'ration-dirty-text':'mut';status.textContent=changed?('● '+changed+' yem kaleminde kaydedilmemiş değişiklik var'):'Kaydedilmiş rasyon gösteriliyor.';}}rows.forEach(r=>{{const i=r.querySelector('.ration-qty');i.addEventListener('input',calc);r.querySelectorAll('.qty-step').forEach(b=>b.onclick=(e)=>{{e.preventDefault();const y=window.scrollY;i.value=Math.max(0,(parseFloat(i.value)||0)+parseFloat(b.dataset.delta)).toFixed(2);calc();requestAnimationFrame(()=>window.scrollTo(0,y));}});r.querySelector('.qty-zero').onclick=(e)=>{{e.preventDefault();const y=window.scrollY;i.value='0.00';calc();requestAnimationFrame(()=>window.scrollTo(0,y));}};}});reset.onclick=()=>{{rows.forEach(r=>{{const i=r.querySelector('.ration-qty');i.value=parseFloat(i.dataset.original||0).toFixed(2);}});calc();}};form.addEventListener('submit',()=>sessionStorage.setItem(scrollKey,String(window.scrollY)));const saved=sessionStorage.getItem(scrollKey);if(saved!==null){{sessionStorage.removeItem(scrollKey);requestAnimationFrame(()=>window.scrollTo(0,parseFloat(saved)||0));}}calc();}})();</script>
                        <script id="dev413-science-live">(()=>{{const form=document.getElementById('ration-bulk-form');if(!form||!document.querySelector('.science-target-grid'))return;const rows=[...form.querySelectorAll('.ration-row')];const t={{weight:{targets['weight_kg']:.8f},sbw:{targets['sbw_kg']:.8f},ebw:{targets['ebw_kg']:.8f},age:{targets.get('age_months',0):.8f},adg:{targets['adg']:.8f},nemReq:{targets['nem_req_mcal']:.8f},negReq:{targets['neg_req_mcal']:.8f},gainCoefficient:{targets.get('gain_energy_coefficient',0.0635):.8f},cp:{targets['cp_pct']:.8f},ndfMin:{targets['ndf_min']:.8f},ndfMax:{targets['ndf_max']:.8f},endfMin:{targets.get('endf_min',0):.8f},starchMin:{targets.get('starch_min',0):.8f},starchIdeal:{targets.get('starch_ideal_max',100):.8f},starchMax:{targets.get('starch_max',100):.8f},ca:{targets['ca_g']:.8f},p:{targets['p_g']:.8f},roughMin:{targets['roughage_min']:.8f},roughMax:{targets['roughage_max']:.8f}}};const fmt=(n,d=1)=>Number(n||0).toLocaleString('tr-TR',{{minimumFractionDigits:d,maximumFractionDigits:d}});const set=(id,v)=>{{const e=document.getElementById(id);if(e)e.textContent=v;}};const level=s=>s.includes('🔴')?'bad':s.includes('⚠️')?'warn':s.includes('✅')?'ok':'info';function setRow(key,target,current,status,diff=''){{set('target-mini-'+key+'-target',target);set('target-mini-'+key+'-current',current);set('target-mini-'+key+'-status',status);set('target-mini-'+key+'-diff',diff);const el=document.getElementById('target-mini-'+key);if(el){{el.classList.remove('ok','warn','bad','info');el.classList.add(level(status));}}}}function symmetric(a,b){{const d=(a-b)/Math.max(b,.01);if(d<-.10)return '🔴 Eksik %'+fmt(Math.abs(d)*100,0);if(d>.10)return '⚠️ Yüksek %'+fmt(d*100,0);if(Math.abs(d)>.05)return '⚠️ Sınıra yakın';return '✅ Uygun';}}function minimum(a,b){{const d=(a-b)/Math.max(b,.01);if(d<-.10)return '🔴 Eksik %'+fmt(Math.abs(d)*100,0);if(d<-.005)return '⚠️ Hedef altı %'+fmt(Math.abs(d)*100,1);return '✅ Yeterli';}}function range(a,lo,hi){{return a<lo?'🔴 Düşük':(a>hi?'⚠️ Yüksek':'✅ Aralıkta');}}function calcScience(){{let dm=0,cpkg=0,ndfkg=0,endfkg=0,starchkg=0,rapidkg=0,knownstarch=0,me=0,nem=0,neg=0,ca=0,p=0,cost=0,roughdm=0,concdm=0,changed=0;rows.forEach(r=>{{const input=r.querySelector('.ration-qty'),kg=Math.max(0,parseFloat((input?.value||'0').replace(',','.'))||0),orig=parseFloat(input?.dataset.original||0),dmp=parseFloat(r.dataset.dm||0),d=kg*dmp/100;dm+=d;cpkg+=d*parseFloat(r.dataset.cp||0)/100;ndfkg+=d*parseFloat(r.dataset.ndf||0)/100;endfkg+=d*parseFloat(r.dataset.ndf||0)/100*parseFloat(r.dataset.endf||0)/100;const skg=d*parseFloat(r.dataset.starch||0)/100,sd=parseFloat(r.dataset.starchDeg||0);starchkg+=skg;if(sd>0){{knownstarch+=skg;rapidkg+=skg*sd/100;}}me+=d*parseFloat(r.dataset.me||0);nem+=d*parseFloat(r.dataset.nem||0);neg+=d*parseFloat(r.dataset.neg||0);ca+=d*parseFloat(r.dataset.ca||0)*10;p+=d*parseFloat(r.dataset.p||0)*10;cost+=kg*parseFloat(r.dataset.price||0);if(r.dataset.group==='Kaba')roughdm+=d;else if(r.dataset.group==='Kesif')concdm+=d;if(Math.abs(kg-orig)>.0005)changed++;}});const cp=dm?cpkg/dm*100:0,ndf=dm?ndfkg/dm*100:0,endf=dm?endfkg/dm*100:0,starch=dm?starchkg/dm*100:0,rapid=dm?rapidkg/dm*100:0,coverage=starchkg?knownstarch/starchkg:1,nemDensity=dm?nem/dm:0,negDensity=dm?neg/dm:0,rc=roughdm+concdm,rough=rc?roughdm/rc*100:0,conc=rc?concdm/rc*100:0;let ne=Math.max(.70,Math.min(nemDensity||1.60,2.50)),nema=Math.max(ne,.95),yearling=t.age>0?t.age>=12:t.weight>=300,pred=Math.pow(t.sbw,.75)*(.2435*nema-.0466*nema*nema-(yearling?.0869:.1128))/nema;pred=Math.max(t.weight*.018,Math.min(t.weight*.035,pred));let capacity=0;if(nemDensity>0&&negDensity>0){{const gain=Math.max(0,dm-t.nemReq/nemDensity)*negDensity;if(gain>0){{const ebg=Math.pow(gain/Math.max(t.gainCoefficient*Math.pow(t.ebw,.75),1e-9),1/1.097);capacity=Math.max(0,ebg/.956);if(cp<t.cp*.95)capacity*=Math.max(.70,cp/Math.max(t.cp*.95,.1));}}}}setRow('dm','≈ '+fmt(pred,2)+' kg',fmt(dm,2)+' kg',symmetric(dm,pred),(dm-pred>=0?'+':'')+fmt(dm-pred,2)+' kg');setRow('adg','≥ '+fmt(t.adg,2)+' kg',capacity?fmt(capacity,2)+' kg':'—',capacity?minimum(capacity,t.adg):'ℹ️ NEm/NEg verisi yok',(capacity-t.adg>=0?'+':'')+fmt(capacity-t.adg,2)+' kg');setRow('cp','≥ %'+fmt(t.cp,1),'%'+fmt(cp,1),minimum(cp,t.cp),'HP tabanı');setRow('ndf','%'+fmt(t.ndfMin,0)+'–'+fmt(t.ndfMax,0),'%'+fmt(ndf,1),range(ndf,t.ndfMin,t.ndfMax));setRow('endf','≥ %'+fmt(t.endfMin,1),'%'+fmt(endf,1),minimum(endf,t.endfMin));setRow('rc','Kaba %'+fmt(t.roughMin,0)+'–'+fmt(t.roughMax,0),'%'+fmt(rough,0)+' / %'+fmt(conc,0),range(rough,t.roughMin,t.roughMax));const starchS=starch>=t.starchMin&&starch<=t.starchIdeal?'✅ İdeal bant':(starch<t.starchMin?'ℹ️ İdeal altı':(starch<=t.starchIdeal+.5?'⚠️ Sınırda':(starch<=t.starchMax?'⚠️ Dikkat bandı':'🔴 Yüksek')));setRow('starch','İdeal %'+fmt(t.starchMin,0)+'–'+fmt(t.starchIdeal,0),'%'+fmt(starch,1),starchS,fmt(starchkg,2)+' kg/baş');let riskScore=0;if(starch>t.starchMax)riskScore+=2;else if(starch>t.starchIdeal+.5)riskScore++;if(endf<t.endfMin*.80)riskScore+=2;else if(endf<t.endfMin)riskScore++;if(coverage>=.70&&rapid>t.starchIdeal*.75)riskScore++;const risk=riskScore>=3?'Yüksek':(riskScore>=1?'Orta':'Düşük'),confidence=coverage>=.85?'Yüksek':(coverage>=.50?'Orta':'Düşük'),riskS=risk==='Düşük'?'✅ Düşük':(risk==='Orta'?'⚠️ Orta':'🔴 Yüksek');setRow('ph','Düşük',risk,riskS,'Veri güveni '+confidence);const caSoft=Math.max(t.ca*1.5,pred*10*.85),pSoft=Math.max(t.p*1.5,pred*10*.50),caHard=Math.max(t.ca*2.25,pred*10*1.20),pHard=Math.max(t.p*2.25,pred*10*.70);function mineral(a,min,soft,hard){{if(a<min*.90)return '🔴 Eksik %'+fmt((1-a/min)*100,0);if(a<min)return '⚠️ Minimuma yakın';if(a>hard)return '🔴 Güvenlik üstü';if(a>soft)return '⚠️ Yüksek';return '✅ Yeterli';}}setRow('ca','≥ '+fmt(t.ca,0)+' g',fmt(ca,0)+' g',mineral(ca,t.ca,caSoft,caHard));setRow('p','≥ '+fmt(t.p,0)+' g',fmt(p,0)+' g',mineral(p,t.p,pSoft,pHard));const ratio=p>0?ca/p:0,ratioS=ratio>=1.2&&ratio<=3?'✅ Dengeli':(ratio<1?'🔴 Düşük':'⚠️ Kontrol');setRow('cap','1,2–3,0',fmt(ratio,2),ratioS);setRow('cost','—','₺'+fmt(cost,2),changed?'💰 Taslak':'💰 Güncel',changed?changed+' kalem değişti':'Kaydedilmiş');set('target-mini-me-current',fmt(me,1));const note=document.getElementById('target-live-note');if(note)note.textContent=changed?'🟡 Kaydedilmemiş taslak; tüm göstergeler canlı yeniden hesaplandı.':'NASEM hayvan profili · Dinamik KM · NEm/NEg performansı';}}rows.forEach(r=>r.querySelector('.ration-qty')?.addEventListener('input',calcScience));form.addEventListener('click',e=>{{if(e.target.closest('.qty-step,.qty-zero,#ration-reset'))requestAnimationFrame(calcScience);}});calcScience();}})();</script>
                        <details id='quick-feed-add' class='card quick-feed-card' style='margin-top:14px;border:1px solid #cfe3d5' {('open' if q.get('feedadd',['0'])[0]=='1' else '')}><summary class='quick-feed-head' style='cursor:pointer'><div><h3 style='display:inline;margin:0'>➕ Rasyona Yem Ekle</h3><span class='mut' style='margin-left:8px'>Yem ekleyeceğiniz zaman açın</span></div><span class='pill'>{('🥛 Süt' if ('ration_type' in rr.keys() and (rr['ration_type'] or '').lower().startswith(('süt','sut'))) else '🥩 Besi')}</span></summary><div class='quick-feed-body'><div class='quick-feed-tools'><input id='quick-feed-search' type='search' placeholder='🔎 Yem ara: mısır, arpa, yonca, süt yemi...' autocomplete='off'><div class='quick-feed-shortcuts'><button type='button' class='btn alt compact-btn quick-filter' data-filter=''>Tümü</button><button type='button' class='btn alt compact-btn quick-filter' data-filter='mısır'>Mısır</button><button type='button' class='btn alt compact-btn quick-filter' data-filter='arpa'>Arpa</button><button type='button' class='btn alt compact-btn quick-filter' data-filter='yonca'>Yonca</button></div></div><div id='quick-feed-results' class='quick-feed-results'>{quick_feed_html}</div><form method='post' action='/ration/item' id='quick-feed-form' class='quick-feed-selected'><input type='hidden' name='ration_id' value='{selected}'><input type='hidden' name='keep_feed_add_open' value='1'><input type='hidden' name='feed_id' id='quick-feed-id' required><div><span class='mut'>Seçilen Yem</span><b id='quick-feed-name'>Önce yukarıdan yem seçin</b></div><div class='ration-stepper'><button type='button' class='btn alt compact-btn' id='quick-feed-minus'>−</button><input id='quick-feed-qty' type='number' step='0.01' min='0.01' name='kg_per_head_day' value='0.50' required><button type='button' class='btn alt compact-btn' id='quick-feed-plus'>+</button></div><button class='btn' id='quick-feed-submit' disabled>➕ Rasyona Ekle / Güncelle</button></form><div style='display:flex;justify-content:flex-end;margin-top:10px'><button type='button' class='btn alt compact-btn' id='quick-feed-close'>✕ Yem Ekleme Alanını Kapat</button></div></div></details><script>(()=>{{const box=document.getElementById('quick-feed-add');if(!box)return;const search=document.getElementById('quick-feed-search'),results=[...box.querySelectorAll('.quick-feed-result')],fid=document.getElementById('quick-feed-id'),fname=document.getElementById('quick-feed-name'),qty=document.getElementById('quick-feed-qty'),submit=document.getElementById('quick-feed-submit');const norm=v=>(v||'').toLocaleLowerCase('tr-TR').normalize('NFD').replace(/[\u0300-\u036f]/g,'').replace(/ı/g,'i').replace(/ş/g,'s').replace(/ğ/g,'g').replace(/ü/g,'u').replace(/ö/g,'o').replace(/ç/g,'c');function filter(){{const q=norm(search.value.trim());let shown=0;results.forEach(r=>{{const hay=norm((r.dataset.search||'')+' '+(r.textContent||''));const ok=!q||hay.includes(q);const vis=ok&&(q||shown<30);r.style.display=vis?'flex':'none';if(vis)shown++;}});}}search.addEventListener('input',filter);box.querySelectorAll('.quick-filter').forEach(b=>b.onclick=()=>{{search.value=b.dataset.filter||'';filter();search.focus();}});results.forEach(r=>r.onclick=()=>{{results.forEach(x=>x.classList.remove('selected'));r.classList.add('selected');fid.value=r.dataset.feedId;fname.textContent=r.dataset.feedName;const cur=parseFloat(r.dataset.current||0);qty.value=(cur>0?cur:0.50).toFixed(2);submit.disabled=false;}});document.getElementById('quick-feed-minus').onclick=()=>qty.value=Math.max(.01,(parseFloat(qty.value)||.5)-.10).toFixed(2);document.getElementById('quick-feed-plus').onclick=()=>qty.value=Math.max(.01,(parseFloat(qty.value)||.5)+.10).toFixed(2);document.getElementById('quick-feed-close').onclick=()=>{{box.open=false;history.replaceState(null,'',location.pathname+'?id={selected}#ration-workbench');}};box.addEventListener('toggle',()=>{{if(box.open)setTimeout(()=>search.focus(),50);}});filter();}})();</script>
                        
                        {sim_html}
                        <div id="smart-balance" class="card ration-section-collapse" style="margin-top:14px;border:1px solid #d8e4ff"><div style="display:flex;justify-content:space-between;align-items:center;gap:10px;flex-wrap:wrap"><div><h3 style="margin:0">🧠 Akıllı Dengeleme</h3><span class="mut">Hedef kartları çalışma masasıyla canlı bağlıdır. Akıllı çözüm önerileri kaydettiğiniz son rasyona göre yenilenir.</span></div></div><div class="smart-solution-grid">{quick_solutions_html}</div><details style="margin-top:12px"><summary><b>🔬 Tüm önerileri ve teknik analizi göster</b></summary><div style="margin-top:12px"><h3>✂️ Fazlalıkları Azaltmak İçin</h3><div style="overflow:auto"><table class="smart-tech-table"><tr><th>Mevcut Yem</th><th>Deneme</th><th>Beklenen Etki</th><th>Yeni Maliyet</th><th></th></tr>{reduce_html}</table></div><h3 style="margin-top:18px">➕ Eksikleri Tamamlamak İçin</h3><div style="overflow:auto"><table class="smart-tech-table"><tr><th>Yem</th><th>Beklenen Etki</th><th>Fiyat</th><th>Stok</th><th></th></tr>{add_top_html}</table></div><details style="margin-top:12px"><summary><b>📚 Tüm uygun yem adayları ({len(add_recs)})</b></summary><div style="overflow:auto;margin-top:8px"><table class="smart-tech-table"><tr><th>Yem</th><th>Beklenen Etki</th><th>Fiyat</th><th>Stok</th><th></th></tr>{add_all_html}</table></div></details><h3 style="margin-top:18px">⚖️ Kombine Dengeleme Fikirleri</h3><div style="overflow:auto"><table class="smart-tech-table"><tr><th>Azalt</th><th>Ekle</th><th>Beklenen Etki</th><th></th></tr>{combo_html}</table></div></div></details></div>
                        <div class="costbox"><b>Not:</b> ÇiftlikPro bu ekranda rasyonun besin içeriği ve maliyetini analiz eder. Nihai rasyon uygunluğu hayvanın canlı ağırlığı, yaş, sağlık ve hedef performansına göre veteriner/zooteknist tarafından değerlendirilmelidir.</div>
                        <details class="card" style="margin-top:14px"><summary><b>🏠 Padoka Ata</b></summary><form method="post" action="/ration/assign" class="actions" style="margin-top:12px"><input type="hidden" name="ration_id" value="{selected}"><select name="paddock_id" required><option value="">Padok seçin</option>{pd_opts}</select><input type="date" name="start_date" value="{date.today().isoformat()}" required><button class="btn orange">Padoka Ata</button></form></details></div>'''
            solve_feed_html=''.join(f'''<label class="solve-feed"><input type="checkbox" name="feed_{x["id"]}" value="1"><span><b>{h(x["name"])}</b><small>{h(x["category"] or "")} · KM %{float(x["dm_pct"] or 0):.0f} · HP %{float(x["cp_pct"] or 0):.1f} · NDF %{float(x["ndf_pct"] or 0):.1f}</small></span></label>''' for x in feeds)
            solve_assistant_html=''
            if (q.get('solve',[''])[0]=='1') and msg:
                solve_assistant_html=f'''<div class="solve-assistant-top" role="alert"><div class="solve-assistant-title"><span>🧠</span><div><b>Rasyon Asistanı</b><small>Çözüm için gereken düzeltme</small></div></div><div class="solve-assistant-text">{h(msg)}</div></div>'''
            solve_panel=f'''<div class="ration-drawer-backdrop" id="rationSolveBackdrop" data-close="rationSolveDrawer"></div><aside class="ration-drawer" id="rationSolveDrawer" aria-hidden="true"><div class="ration-drawer-head"><div><h2>🧠 Rasyon Çöz</h2><p>Besi veya süt hedefi + elinizdeki yemlerle akıllı çözüm</p></div><button type="button" class="ration-drawer-close" data-close="rationSolveDrawer">×</button></div><form method="post" action="/ration/solve" class="ration-drawer-form"><div class="ration-drawer-body">{solve_assistant_html}<div class="solve-type-switch"><label>Rasyon Türü<select name="ration_type" id="solve-ration-type"><option value="Besi">🥩 Besi Rasyonu</option><option value="Süt">🥛 Süt Rasyonu</option></select></label></div><div class="solve-target-grid"><label>Canlı Ağırlık (kg)<input type="number" name="target_weight_kg" min="150" max="900" step="1" value="450" required></label><label class="solve-beef-field">Hedef Günlük Artış (kg/gün)<input type="number" name="target_adg_kg" min="0.2" max="2.2" step="0.05" value="1.30"></label><label class="solve-beef-field">Yaş (ay, opsiyonel)<input type="number" name="target_age_months" min="0" max="36" step="1" placeholder="örn. 14"></label><label class="solve-beef-field">Besi Dönemi<select name="target_beef_phase"><option value="Otomatik">Otomatik (canlı ağırlığa göre)</option><option>Besi Başlangıç</option><option>Besi Geliştirme</option><option>Besi Bitirme</option></select></label><label class="solve-beef-field">Hayvan Tipi<select name="animal_type"><option>Besi Erkek</option><option>Düve</option><option>Genel Büyüyen Sığır</option></select></label><label class="solve-milk-field" style="display:none">Hedef Süt (kg/gün)<input type="number" name="target_milk_l" min="0" max="70" step="0.5" value="25"></label><div class="solve-milk-field mut" style="display:none">Aynı Yem Kataloğu kullanılır. Süt yağı/protein için saha varsayımları içeride uygulanır; temel kullanım yalnız canlı ağırlık + hedef süt miktarıdır.</div></div><div class="solve-section-title"><b>Elimdeki Yemler</b><span>Rasyon Ekle ile aynı ÇiftlikPro yem kataloğu kullanılır; min/max sınırlarını sistem otomatik belirler.</span></div><div class="solve-search-wrap"><input type="search" id="solve-feed-search" class="solve-search" placeholder="🔎 Yem ara... arpa, silaj, saman, küspe" autocomplete="off"></div><div class="solve-selected-box" id="solve-selected-box"><div class="solve-selected-head"><b>✅ Seçilen Yemler</b><span id="solve-selected-count-top">0 yem</span></div><div class="solve-selected-chips" id="solve-selected-chips"><span class="mut">Henüz yem seçilmedi.</span></div></div><div class="solve-feed-grid drawer-feed-grid" id="solve-feed-grid">{solve_feed_html}</div></div><div class="ration-drawer-foot"><span id="solve-selected-count">0 yem seçildi</span><span class="solve-progress" id="solve-progress"><i class="solve-spinner"></i><span>Rasyon çözülüyor...</span></span><button type="button" class="btn ghost" id="solve-clear">Temizle</button><button class="btn blue solve-submit" id="solve-submit">🧠 Rasyonu Çöz</button></div></form></aside>'''
            new_ration_panel='''<div class="ration-drawer-backdrop" id="rationAddBackdrop" data-close="rationAddDrawer"></div><aside class="ration-drawer" id="rationAddDrawer" aria-hidden="true"><div class="ration-drawer-head"><div><h2>➕ Rasyon Ekle</h2><p>Manuel reçete oluştur ve kaydet.</p></div><button type="button" class="ration-drawer-close" data-close="rationAddDrawer">×</button></div><div class="ration-drawer-body"><div class="drawer-section-card"><h3>🥩 Besi Rasyonu Oluştur</h3><form method="post" action="/ration/create" class="form"><input type="hidden" name="ration_type" value="Besi"><input type="hidden" name="target_group" value="Besi"><label>Rasyon Adı<input name="name" required placeholder="Besi 500 kg"></label><label>Ortalama Canlı Ağırlık (kg)<input type="number" min="150" max="900" step="1" name="target_weight_kg" value="450"></label><label>Hedef Günlük Artış (kg/gün)<input type="number" min="0.2" max="2.2" step="0.05" name="target_adg_kg" value="1.30"></label><label>Yaş (ay, opsiyonel)<input type="number" min="0" max="36" step="1" name="target_age_months"></label><label>Hayvan Tipi<select name="animal_type"><option>Besi Erkek</option><option>Düve</option><option>Genel Büyüyen Sığır</option></select></label><label class="full">Not<input name="notes"></label><div class="full"><button class="btn">Besi Rasyonunu Oluştur</button></div></form></div><div class="drawer-section-card"><h3>🥛 Süt Rasyonu Oluştur</h3><form method="post" action="/ration/create" class="form"><input type="hidden" name="ration_type" value="Süt"><input type="hidden" name="target_group" value="Sağmal"><input type="hidden" name="animal_type" value="Sağmal İnek"><label>Rasyon Adı<input name="name" required placeholder="Süt 25 L"></label><label>Ortalama Canlı Ağırlık (kg)<input type="number" min="350" max="900" step="1" name="target_weight_kg" value="650"></label><label>Hedef Süt (L/gün)<input type="number" min="0" max="70" step="0.5" name="target_milk_l" value="25"></label><label class="full">Not<input name="notes"></label><div class="full"><button class="btn blue">Süt Rasyonunu Oluştur</button></div></form></div></div></aside>'''
            action_cards='''<div class="ration-action-launchers"><button type="button" class="ration-launch-card add" data-open="rationAddDrawer"><span class="launch-icon">＋</span><span><b>Rasyon Ekle</b><small>Manuel reçete oluştur</small></span></button><button type="button" class="ration-launch-card solve" data-open="rationSolveDrawer"><span class="launch-icon">🧠</span><span><b>Rasyon Çöz</b><small>Besi: canlı ağırlık + artış · Süt: canlı ağırlık + süt</small></span></button></div>'''
            drawer_script='''<script>(()=>{const drawers={rationAddDrawer:'rationAddBackdrop',rationSolveDrawer:'rationSolveBackdrop'};function setDrawer(id,on){const d=document.getElementById(id),b=document.getElementById(drawers[id]);if(!d||!b)return;d.classList.toggle('open',on);b.classList.toggle('open',on);d.setAttribute('aria-hidden',on?'false':'true');document.body.classList.toggle('ration-drawer-open',on)}document.querySelectorAll('[data-open]').forEach(x=>x.addEventListener('click',()=>setDrawer(x.dataset.open,true)));document.querySelectorAll('[data-close]').forEach(x=>x.addEventListener('click',()=>setDrawer(x.dataset.close,false)));document.addEventListener('keydown',e=>{if(e.key==='Escape')Object.keys(drawers).forEach(id=>setDrawer(id,false))});const q=document.getElementById('solve-feed-search'),grid=document.getElementById('solve-feed-grid'),count=document.getElementById('solve-selected-count'),countTop=document.getElementById('solve-selected-count-top'),chips=document.getElementById('solve-selected-chips'),clear=document.getElementById('solve-clear'),form=document.querySelector('#rationSolveDrawer form');if(q&&grid){const KEY='cp-ration-solve-draft';const norm=v=>(v||'').toLocaleLowerCase('tr-TR').normalize('NFD').replace(/[\u0300-\u036f]/g,'').replace(/ı/g,'i').replace(/ş/g,'s').replace(/ğ/g,'g').replace(/ü/g,'u').replace(/ö/g,'o').replace(/ç/g,'c');const items=[...grid.querySelectorAll('.solve-feed')],checks=[...grid.querySelectorAll('input[type=checkbox]')];function saveDraft(){if(!form)return;const d={};form.querySelectorAll('input,select').forEach(el=>{if(!el.name)return;if(el.type==='checkbox')d[el.name]=el.checked;else d[el.name]=el.value});sessionStorage.setItem(KEY,JSON.stringify(d))}function restoreDraft(){try{const d=JSON.parse(sessionStorage.getItem(KEY)||'null');if(!d||!form)return;form.querySelectorAll('input,select').forEach(el=>{if(!el.name||!(el.name in d))return;if(el.type==='checkbox')el.checked=!!d[el.name];else el.value=d[el.name]})}catch(e){}}function updateCount(){const selected=checks.filter(c=>c.checked);const n=selected.length;if(count)count.textContent=n+' yem seçildi';if(countTop)countTop.textContent=n+' yem';if(chips){chips.innerHTML='';if(!n){const e=document.createElement('span');e.className='mut';e.textContent='Henüz yem seçilmedi.';chips.appendChild(e)}else selected.forEach(c=>{const label=c.closest('.solve-feed'),b=label?.querySelector('b'),small=label?.querySelector('small'),chip=document.createElement('button');chip.type='button';chip.className='solve-chip';chip.title='Seçimden çıkar';chip.innerHTML='<span>'+((b&&b.textContent)||'Yem')+'</span><small>'+((small&&small.textContent.split(' · ')[0])||'')+'</small><strong>×</strong>';chip.onclick=()=>{c.checked=false;updateCount();saveDraft()};chips.appendChild(chip)})}}q.addEventListener('input',()=>{const term=norm(q.value.trim());items.forEach(x=>x.classList.toggle('solve-filter-hidden',!!term&&!norm(x.textContent).includes(term)))});checks.forEach(c=>c.addEventListener('change',()=>{updateCount();saveDraft()}));if(form){form.querySelectorAll('input:not([type=checkbox]),select').forEach(el=>el.addEventListener('change',saveDraft));form.addEventListener('submit',()=>{saveDraft();const btn=document.getElementById('solve-submit'),prog=document.getElementById('solve-progress');if(btn){btn.disabled=true;btn.classList.add('solving');btn.textContent='⏳ Çözülüyor...'}if(prog)prog.classList.add('on')})}if(clear)clear.addEventListener('click',()=>{checks.forEach(c=>c.checked=false);q.value='';items.forEach(x=>x.classList.remove('solve-filter-hidden'));sessionStorage.removeItem(KEY);updateCount();q.focus()});restoreDraft();updateCount();const rt=document.getElementById('solve-ration-type');function syncSolveType(){const milk=rt&&rt.value==='Süt';document.querySelectorAll('.solve-beef-field').forEach(x=>x.style.display=milk?'none':'');document.querySelectorAll('.solve-milk-field').forEach(x=>x.style.display=milk?'':'none');const w=form&&form.querySelector('[name=target_weight_kg]');if(w){w.min=milk?'350':'150';if(milk&&Number(w.value)<350)w.value='650';}saveDraft()}if(rt){rt.addEventListener('change',syncSolveType);syncSolveType()}}if(new URLSearchParams(location.search).get('solve')==='1')setDrawer('rationSolveDrawer',true)})();</script>'''
            if is_workbench:
                if not selected or not detail:
                    return self.redirect('/rations','Rasyon bulunamadı veya seçilmedi.')
                body=f'''<div class="workbench-page-head"><div><a class="btn alt compact-btn" href="/rations">← Rasyonlara Dön</a><h1 class="ration-page-title" style="margin-top:10px">🧪 Rasyon Çalışma Masası</h1><p class="mut">Bu ekran yalnızca seçili rasyon üzerinde çalışmak içindir. Hedefler, yem miktarları, göreli asidoz riski ve maliyet birlikte değerlendirilir; klinik rumen pH tahmini yapılmaz.</p></div></div>{detail}{WORKBENCH_REFERENCE_UI_V3}'''
                return self.send_html(page('Rasyon Çalışma Masası',body,'/rations',u,msg))
            body=f'''<h1 class="ration-page-title">🥣 Rasyon Yönetimi</h1><p class="mut ration-page-subtitle">Rasyonlarınızı oluşturun, çözün ve yönetin. Bir rasyona tıklayınca ayrı Çalışma Masası açılır.</p><div class="ration-page-steps"><span>1️⃣ Rasyon oluştur / çöz</span><span>2️⃣ Rasyonu seç</span><span>3️⃣ Ayrı çalışma masasında düzenle</span></div>{action_cards}{new_ration_panel}{solve_panel}{drawer_script}<div class="ration-picker-grid">{''.join(cards) if cards else '<div class="card">Henüz rasyon oluşturulmadı.</div>'}</div>'''
            return self.send_html(page('Rasyon Yönetimi',body,'/rations',u,msg))
        if path=='/performance':
            status_filter=(q.get('status',[''])[0] or '').strip();scope=(q.get('scope',['all'])[0] or 'all').strip();search=(q.get('search',[''])[0] or '').strip()
            if scope not in ('all','active','completed'):scope='all'
            purchase_start=(q.get('purchase_start',[''])[0] or '').strip();purchase_end=(q.get('purchase_end',[''])[0] or '').strip();exit_start=(q.get('exit_start',[''])[0] or '').strip();exit_end=(q.get('exit_end',[''])[0] or '').strip();profit_filter=(q.get('profit',[''])[0] or '').strip()
            if profit_filter not in ('','profit','loss','pending'):profit_filter=''
            detail=[]
            with db() as c:
                rows=c.execute("select * from animals where gender='Erkek' and coalesce(status,'Aktif') in ('Aktif','Kesildi','Satıldı') order by tag").fetchall()
                for ar in rows:
                    st=str(ar['status'] or 'Aktif');completed=st in ('Kesildi','Satıldı');hay=(str(ar['tag'] or '')+' '+str(ar['nickname'] or '')+' '+str(ar['breed'] or '')).casefold()
                    if scope=='active' and st!='Aktif':continue
                    if scope=='completed' and not completed:continue
                    if search and search.casefold() not in hay:continue
                    pd=ar['purchase_date'] or '';ed=ar['exit_date'] or ''
                    if purchase_start and (not pd or pd<purchase_start):continue
                    if purchase_end and (not pd or pd>purchase_end):continue
                    if exit_start and (not completed or not ed or ed<exit_start):continue
                    if exit_end and (not completed or not ed or ed>exit_end):continue
                    perf=male_weight_performance(ar['id'],c)
                    if status_filter and perf['status']!=status_filter:continue
                    days,daily,operating,total_cost=animal_cost_values(ar)
                    revenue=float(c.execute("select coalesce(sum(amount),0) from finance where animal_id=? and tx_type='Gelir' and category in ('Kesim Geliri','Hayvan Satışı')",(ar['id'],)).fetchone()[0] or 0)
                    if completed and revenue<=0 and float(ar['sold_price'] or 0)>0:revenue=float(ar['sold_price'] or 0)
                    net_profit=(revenue-total_cost) if completed and revenue>0 else None;profit_pct=((net_profit/total_cost)*100) if net_profit is not None and total_cost>0 else None
                    if profit_filter=='profit' and not(net_profit is not None and net_profit>0):continue
                    if profit_filter=='loss' and not(net_profit is not None and net_profit<0):continue
                    if profit_filter=='pending' and net_profit is not None:continue
                    wr=c.execute('select measure_date,weight from weights where animal_id=? order by measure_date,id',(ar['id'],)).fetchall();first=wr[0] if wr else None;last=wr[-1] if wr else None
                    gain=(float(last['weight'])-float(first['weight'])) if first and last and len(wr)>=2 else None
                    detail.append((ar,perf,days,operating,total_cost,gain,revenue,net_profit,profit_pct))
            active_count=sum(1 for x in detail if str(x[0]['status'] or '')=='Aktif');completed_count=len(detail)-active_count
            purchase_total=sum(float(x[0]['purchase_price'] or 0) for x in detail);operating_total=sum(x[3] for x in detail);total_cost=sum(x[4] for x in detail)
            realized_revenue=sum(x[6] for x in detail if x[7] is not None);realized_cost=sum(x[4] for x in detail if x[7] is not None);realized_profit=sum(x[7] for x in detail if x[7] is not None);margin=((realized_profit/realized_cost)*100) if realized_cost>0 else None
            daily_vals=[x[1]['daily'] for x in detail if x[1].get('daily') is not None];avg_daily=(sum(daily_vals)/len(daily_vals)) if daily_vals else None
            labels={'good':('Hedefte','status-good'),'watch':('Takip','status-watch'),'low':('Düşük','status-low'),'none':('Veri Yetersiz','status-none')};trs=''
            for ar,perf,days,operating,cost,gain,revenue,profit,pct in detail:
                st=str(ar['status'] or 'Aktif');completed=st in ('Kesildi','Satıldı');label,cls=labels[perf['status']];daily_text=f"{perf['daily']:.3f} kg/gün" if perf.get('daily') is not None else '-';gain_text=f'{gain:+.1f} kg' if gain is not None else '-'
                if profit is None:profit_html='<span class="profit-pill wait">Henüz gerçekleşmedi</span>';pct_text='—'
                else:profit_html=f'<span class="profit-pill {"good" if profit>=0 else "bad"}">{"+" if profit>0 else ""}{money(profit)}</span>';pct_text=f'{pct:+.1f}%' if pct is not None else '—'
                rev=money(revenue) if completed and revenue>0 else ('Gelir bekleniyor' if completed else '—');ed=fmt_date(ar['exit_date'] or '-') if completed else 'Devam ediyor'
                trs+=f'<tr><td><a class="taglink" href="/animal?id={ar["id"]}">{h(ar["tag"])}</a><br><span class="mut">{h(ar["nickname"])}</span></td><td>{h(st)}</td><td>{fmt_date(ar["purchase_date"] or "-")}</td><td>{ed}</td><td>{days} gün</td><td>{gain_text}<br><span class="mut">{daily_text}</span></td><td>{money(float(ar["purchase_price"] or 0))}</td><td>{money(operating)}</td><td><b>{money(cost)}</b></td><td><b>{rev}</b></td><td>{profit_html}</td><td><b>{pct_text}</b></td><td><span class="perf-badge {cls}">{label}</span></td></tr>'
            trs=trs or '<tr><td colspan="13">Seçilen filtrelerde hayvan bulunamadı.</td></tr>';margin_text=f'{margin:+.1f}%' if margin is not None else '—';avg_text=f'{avg_daily:.3f} kg/gün' if avg_daily is not None else '—';net_cls='green' if realized_profit>=0 else 'red'
            body=f'''<div class="perf-hero"><div><h1>🐂 Besi Performansı & Kârlılık Merkezi</h1><p>Canlı ağırlık performansı, gerçek maliyet ve satış/kesim gelirini tek ekranda karşılaştırın.</p></div><a class="btn" href="/performance-settings">⚙️ Performans Eşiği</a></div>
            <div class="perf-tabs"><a class="perf-tab {'active' if scope=='all' else ''}" href="/performance?scope=all">📊 Tümü</a><a class="perf-tab {'active' if scope=='active' else ''}" href="/performance?scope=active">🐂 Aktif Besi</a><a class="perf-tab {'active' if scope=='completed' else ''}" href="/performance?scope=completed">💰 Tamamlanan Besiler</a></div>
            <div class="card perf-filter-card"><div class="perf-filter-head"><div><h2>🔎 Akıllı Filtreler</h2><span class="mut">Küpe, tarih, performans ve kârlılık durumuna göre daraltın.</span></div><a class="btn alt" href="/performance">Tümünü Temizle</a></div><form method="get" class="perf-filter-grid"><input type="hidden" name="scope" value="{h(scope)}"><label>Küpe / Hayvan Ara<input name="search" value="{h(search)}" placeholder="Küpe, takma ad veya ırk"></label><label>Alım Başlangıç<input type="date" name="purchase_start" value="{h(purchase_start)}"></label><label>Alım Bitiş<input type="date" name="purchase_end" value="{h(purchase_end)}"></label><label>Çıkış Başlangıç<input type="date" name="exit_start" value="{h(exit_start)}"></label><label>Çıkış Bitiş<input type="date" name="exit_end" value="{h(exit_end)}"></label><label>Performans<select name="status"><option value="">Tümü</option><option value="good" {'selected' if status_filter=='good' else ''}>Hedefte</option><option value="watch" {'selected' if status_filter=='watch' else ''}>Takip</option><option value="low" {'selected' if status_filter=='low' else ''}>Düşük</option><option value="none" {'selected' if status_filter=='none' else ''}>Veri Yetersiz</option></select></label><label>Kârlılık<select name="profit"><option value="">Tümü</option><option value="profit" {'selected' if profit_filter=='profit' else ''}>Kârlı</option><option value="loss" {'selected' if profit_filter=='loss' else ''}>Zararda</option><option value="pending" {'selected' if profit_filter=='pending' else ''}>Henüz Gerçekleşmedi</option></select></label><div class="perf-filter-actions"><button class="btn blue">Filtrele</button><a class="btn alt" href="/performance?scope={h(scope)}">Sıfırla</a></div></form></div>
            <div class="perf-summary"><div class="card stat metric blue">Seçilen Hayvan<b>{len(detail)}</b><small>{active_count} aktif · {completed_count} tamamlanan</small></div><div class="card stat metric orange">Toplam Maliyet<b>{money(total_cost)}</b><small>Alış {money(purchase_total)} · Yem/Bakım {money(operating_total)}</small></div><div class="card stat metric green">Gerçekleşen Gelir<b>{money(realized_revenue)}</b><small>Kesim + hayvan satış gelirleri</small></div><div class="card stat metric {net_cls}">Gerçekleşen Net Kâr<b>{money(realized_profit)}</b><small>Geliri kayıtlı tamamlanan hayvanlar</small></div><div class="card stat metric purple">Gerçekleşen Kâr Oranı<b>{margin_text}</b><small>Maliyet üzerinden · Ort. artış {avg_text}</small></div></div>
            <div class="card" style="margin-top:14px"><h2>Hayvan Bazında Kârlılık Analizi</h2><p class="mut">Kesilen/satılan hayvanın maliyeti çıkış tarihinde donar. Finansta hayvana dağıtılan Kesim Geliri / Hayvan Satışı otomatik eşleşir.</p><div class="perf-table-wrap"><table class="performance-table"><thead><tr><th>Hayvan</th><th>Durum</th><th>Alım</th><th>Çıkış</th><th>Besi Süresi</th><th>Kilo Performansı</th><th>Alış Maliyeti</th><th>Yem + Bakım</th><th>Toplam Maliyet</th><th>Gerçekleşen Gelir</th><th>Net Kâr/Zarar</th><th>Kâr %</th><th>Besi Performansı</th></tr></thead><tbody>{trs}</tbody></table></div></div>'''
            return self.send_html(page('Besi Performansı & Kârlılık',body,path,u,msg))
        if path=='/animal-edit':
            aid=q.get('id',[''])[0]
            with db() as c:
                rec=c.execute('select * from animals where id=?',(aid,)).fetchone()
            if not rec:return self.send_html('Hayvan bulunamadı',404)
            cancel='/animals' if rec['gender']=='Dişi' else '/males'
            body=f'''<h1>Hayvan Düzenle</h1><div class="card"><form method="post" action="/animal-edit" enctype="multipart/form-data" class="form" data-smart-photo-form="1"><input type="hidden" name="id" value="{rec["id"]}"><label>Küpe No<input name="tag" required value="{h(rec["tag"])}"></label><label>Takma Ad<input name="nickname" value="{h(rec["nickname"])}"></label><label>Cinsiyet<select name="gender"><option value="Dişi" {'selected' if rec["gender"]=='Dişi' else ''}>Dişi</option><option value="Erkek" {'selected' if rec["gender"]=='Erkek' else ''}>Erkek</option></select></label><label>Irk<input name="breed" value="{h(rec["breed"])}"></label><label>Doğum Tarihi<input type="date" name="birth_date" value="{h(rec["birth_date"])}"></label><label>Padok / Ahır<input name="paddock" value="{h(rec["paddock"])}"></label><label>Fotoğrafı Değiştir<input type="file" name="photo_file" accept="image/*"><span class="camera-note">Telefonda kamera veya galeriden seçim yapabilirsiniz. Büyük fotoğraflar otomatik küçültülür.</span><div class="photo-upload-status" data-upload-status><span data-upload-text>Fotoğraf hazırlanıyor…</span><div class="upload-progress"><div class="upload-progress-bar" data-upload-bar></div></div></div></label><input type="hidden" name="photo_url" value="{h(rec["photo_url"])}"><label>Durum<select name="status"><option value="Aktif" {'selected' if rec["status"]=='Aktif' else ''}>Aktif</option><option value="Satıldı" {'selected' if rec["status"]=='Satıldı' else ''}>Satıldı</option><option value="Kesildi" {'selected' if rec["status"]=='Kesildi' else ''}>Kesildi</option></select></label><label>Satış Fiyatı<input type="number" step="0.01" name="sold_price" value="{h(rec["sold_price"])}"></label><label>Alış Tarihi<input type="date" name="purchase_date" value="{h(rec["purchase_date"])}"></label><label>Alış Fiyatı (TL)<input type="number" min="0" step="0.01" name="purchase_price" value="{h(rec["purchase_price"])}"></label><label>Alış Kilosu (kg)<input type="number" min="0" step="0.1" name="purchase_weight" value="{h(rec["purchase_weight"])}"></label><label>Günlük Yem/Rasyon (TL)<input type="number" min="0" step="0.01" name="daily_feed_cost" value="{h(rec["daily_feed_cost"])}"></label><label>Günlük Bakım (TL)<input type="number" min="0" step="0.01" name="daily_care_cost" value="{h(rec["daily_care_cost"])}"></label><label>Hedef Satış Fiyatı (TL)<input type="number" min="0" step="0.01" name="target_sale_price" value="{h(rec["target_sale_price"])}"></label><label class="full">Not<textarea name="notes">{h(rec["notes"])}</textarea></label><div class="full"><button class="btn">Değişiklikleri Kaydet</button> <a class="btn alt" href="{cancel}">İptal</a></div></form></div>'''
            return self.send_html(page('Hayvan Düzenle',body,cancel,u,msg))
        if path=='/calf-edit':
            cid=q.get('id',[''])[0]
            with db() as c:
                rec=c.execute('select * from calves where id=?',(cid,)).fetchone()
                mothers=c.execute("select id,tag,nickname from animals where gender='Dişi' and coalesce(status,'Aktif')='Aktif' order by tag").fetchall()
            if not rec:return self.send_html('Buzağı bulunamadı',404)
            opts=''.join(f'<option value="{m["id"]}" {"selected" if rec["mother_id"]==m["id"] else ""}>{h(m["tag"])} - {h(m["nickname"])}</option>' for m in mothers)
            body=f'''<h1>Buzağı Düzenle</h1><div class="card"><form method="post" action="/calf-edit" enctype="multipart/form-data" class="form" data-smart-photo-form="1">
            <input type="hidden" name="id" value="{rec["id"]}"><input type="hidden" name="photo_url" value="{h(rec["photo_url"])}">
            <label>Buzağı Küpesi<input name="tag" required value="{h(rec["tag"])}"></label><label>Takma Ad<input name="nickname" value="{h(rec["nickname"])}"></label>
            <label>Anne<select name="mother_id" required>{opts}</select></label><label>Baba Küpesi<input name="father_tag" value="{h(rec["father_tag"])}"></label>
            <label>Doğum Tarihi<input type="date" name="birth_date" required value="{h(rec["birth_date"])}"></label>
            <label>Cinsiyet<select name="gender"><option value="Dişi" {'selected' if rec["gender"]=='Dişi' else ''}>Dişi</option><option value="Erkek" {'selected' if rec["gender"]=='Erkek' else ''}>Erkek</option></select></label>
            <label>Irk<input name="breed" value="{h(rec["breed"])}"></label><label>Padok / Ahır<input name="paddock" value="{h(rec["paddock"])}"></label>
            <label>Alış Tarihi<input type="date" name="purchase_date" value="{h(rec["purchase_date"])}"></label><label>Alış Fiyatı (TL)<input type="number" step="0.01" min="0" name="purchase_price" value="{float(rec["purchase_price"] or 0)}"></label>
            <label>Alış Ödeme Yöntemi<select name="purchase_payment_method"><option {"selected" if rec["purchase_payment_method"]=="Nakit" else ""}>Nakit</option><option {"selected" if rec["purchase_payment_method"]=="Banka" else ""}>Banka</option><option {"selected" if rec["purchase_payment_method"]=="Kredi Kartı" else ""}>Kredi Kartı</option><option {"selected" if rec["purchase_payment_method"]=="Vadeli" else ""}>Vadeli</option></select></label>
            <label>Fotoğraf Yükle<input type="file" name="photo_file" accept="image/*"></label>
            <label class="full">Not<textarea name="notes">{h(rec["notes"])}</textarea></label>
            <div class="full"><button class="btn">Buzağıyı Güncelle</button> <a class="btn alt" href="/calf?id={cid}">İptal</a></div></form></div>'''
            return self.send_html(page('Buzağı Düzenle',body,'/calves',u,msg))
        if path=='/animal-add':
            with db() as c:
                mothers=c.execute("select tag,nickname from animals where gender='Dişi' and coalesce(status,'Aktif')='Aktif' order by tag").fetchall()
                breeds=[r[0] for r in c.execute("select distinct breed from animals where trim(coalesce(breed,''))<>'' order by breed").fetchall()]
                paddocks=[r[0] for r in c.execute("select distinct paddock from animals where trim(coalesce(paddock,''))<>'' order by paddock").fetchall()]
            mother_options=''.join(f'<option value="{h(r["tag"])}">{h(r["nickname"])}</option>' for r in mothers)
            breed_options=''.join(f'<option value="{h(x)}">' for x in breeds)
            paddock_options=''.join(f'<option value="{h(x)}">' for x in paddocks)
            body=f'''<div class="pro-form-head"><div><h1>Hayvan Ekle</h1><div class="mut">Tek formdan dişi, erkek veya buzağı kaydı oluşturun.</div></div><span id="recordTypeBadge" class="type-chip">Dişi Hayvan</span></div><div class="card"><form method="post" action="/animal-add" enctype="multipart/form-data" class="form" data-smart-photo-form="1"><label>Kayıt Türü<select id="recordType" name="record_type" required onchange="toggleAnimalFields()"><option value="Dişi">Dişi Hayvan</option><option value="Erkek">Erkek Hayvan</option><option value="Buzağı">Buzağı</option></select></label><label>Küpe No<input name="tag" required autocomplete="off"></label><label>Takma Ad<input name="nickname"></label><label class="adult-only">Irk<input name="breed" list="breedOptions"><datalist id="breedOptions">{breed_options}</datalist></label><label>Doğum Tarihi<input type="date" name="birth_date"></label><label class="adult-only">Padok / Ahır<input name="paddock" list="paddockOptions"><datalist id="paddockOptions">{paddock_options}</datalist></label><label class="adult-only">Fotoğraf Yükle / Kameradan veya Galeriden Seç<input type="file" name="photo_file" accept="image/*"><span class="camera-note">Telefonda kamera veya galeriden seçim yapabilirsiniz. Büyük fotoğraflar otomatik küçültülür.</span><div class="photo-upload-status" data-upload-status><span data-upload-text>Fotoğraf hazırlanıyor…</span><div class="upload-progress"><div class="upload-progress-bar" data-upload-bar></div></div></div></label><label class="adult-only">Alış Tarihi<input type="date" name="purchase_date"></label><label class="adult-only">Alış Fiyatı (TL)<input type="number" min="0" step="0.01" name="purchase_price"></label><label class="adult-only">Alış Ödeme Yöntemi<select name="purchase_payment_method"><option>Nakit</option><option>Banka</option><option>Kredi Kartı</option><option>Vadeli</option></select></label><label class="male-only" style="display:none">Alış Kilosu (kg)<input type="number" min="0" step="0.1" name="purchase_weight"></label><label class="male-only" style="display:none">Günlük Yem/Rasyon (TL)<input type="number" min="0" step="0.01" name="daily_feed_cost"></label><label class="male-only" style="display:none">Günlük Bakım (TL)<input type="number" min="0" step="0.01" name="daily_care_cost"></label><label class="male-only" style="display:none">Hedef Satış Fiyatı (TL)<input type="number" min="0" step="0.01" name="target_sale_price"></label><div class="female-pregnancy full" id="femalePregnancyBox"><div class="card" style="background:#f7fbf8;border:1px solid #d7eadc"><h3 style="margin-top:0">🤰 Üreme Durumu</h3><div class="form"><label>Hayvanın Durumu<select id="entryPregnancyStatus" name="entry_pregnancy_status" onchange="toggleEntryPregnancy()"><option value="Bos">Boş / Gebe Değil</option><option value="Gebe">Gebe</option><option value="Bilinmiyor">Bilinmiyor</option></select></label><label id="pregnancyInfoModeLabel" style="display:none">Gebelik Bilgisi<select id="pregnancyInfoMode" name="pregnancy_info_mode" onchange="toggleEntryPregnancy()"><option value="date">Son Tohumlama Tarihi Biliniyor</option><option value="age">Sadece Gebelik Yaşı Biliniyor</option></select></label><label id="knownInseminationLabel" style="display:none">Son Tohumlama Tarihi<input type="date" name="known_insemination_date"></label><label id="pregnancyAgeLabel" style="display:none">Gebelik Yaşı (Ay)<input type="number" name="pregnancy_age_months" min="1" max="9" step="0.5" placeholder="Örn. 6"></label><label id="pregnancyEntryDateLabel" style="display:none">Gebelik Bilgisinin Tarihi<input type="date" name="pregnancy_entry_date" value="{date.today().isoformat()}"></label><div class="full mut" id="pregnancyEntryHint" style="display:none">Satın alınırken gebe olduğu ayrıca kaydedilir. Yalnız gebelik ayı biliniyorsa tahmini doğum yaklaşık hesaplanır.</div></div></div></div><label class="calf-only" style="display:none">Buzağı Cinsiyeti<select name="calf_gender"><option>Dişi</option><option>Erkek</option></select></label><label class="calf-only" style="display:none">Anne Küpesi<input name="mother_tag" list="motherTagOptions"><datalist id="motherTagOptions">{mother_options}</datalist></label><label class="calf-only" style="display:none">Baba Küpesi<input name="father_tag"></label><label class="full">Not<textarea name="notes"></textarea></label><div class="full"><button class="btn">Kaydı Oluştur</button> <a class="btn alt" href="/">İptal</a></div></form></div><script>document.addEventListener('DOMContentLoaded',function(){{toggleAnimalFields();}});</script>'''
            return self.send_html(page('Hayvan Ekle',body,'/animal-add',u,msg))

        if path=='/all-animals':
            with db() as c:
                adults=c.execute("select id,tag,nickname,gender,breed,paddock,birth_date from animals where coalesce(status,'Aktif')='Aktif' order by tag").fetchall()
                calves_all=c.execute("select id,tag,nickname,gender,breed,paddock,birth_date from calves where promoted_animal_id is null order by tag").fetchall()
            paddocks=sorted({str(r['paddock'] or '').strip() for r in list(adults)+list(calves_all) if str(r['paddock'] or '').strip()},key=lambda x:x.casefold())
            combined=[]
            for r in adults:
                kind='Dişi' if r['gender']=='Dişi' else 'Erkek'
                row=f'<tr class="data-row" data-kind="{h(kind)}" data-paddock="{h(r["paddock"])}"><td><a class="animal-tag-btn" title="Hayvan kartını aç" href="/animal?id={r["id"]}">{h(r["tag"])}</a></td><td>{h(r["nickname"]) or "-"}</td><td>{h(kind)}</td><td>{h(r["breed"]) or "-"}</td><td>{h(r["paddock"]) or "-"}</td><td>{age_text(r["birth_date"])}</td></tr>'
                combined.append((str(r['tag'] or '').casefold(),row))
            for r in calves_all:
                row=f'<tr class="data-row" data-kind="Buzağı" data-paddock="{h(r["paddock"])}"><td><a class="animal-tag-btn" title="Buzağı kartını aç" href="/calf?id={r["id"]}">{h(r["tag"])}</a></td><td>{h(r["nickname"]) or "-"}</td><td>Buzağı - {h(r["gender"])}</td><td>{h(r["breed"]) or "-"}</td><td>{h(r["paddock"]) or "-"}</td><td>{age_text(r["birth_date"])}</td></tr>'
                combined.append((str(r['tag'] or '').casefold(),row))
            combined.sort(key=lambda x:x[0]); trs=''.join(x[1] for x in combined)
            paddock_options=''.join(f'<option value="{h(x)}">{h(x)}</option>' for x in paddocks)
            body=f'''<h1>🐄 Tüm Aktif Hayvanlar</h1><p class="mut">Aktif dişi, erkek ve buzağılar tek listede. Satılan ve kesilen hayvanlar gösterilmez.</p><div class="card"><div class="form" style="grid-template-columns:2fr 1fr 1fr"><label>Arama<input id="allAnimalSearch" type="search" placeholder="Küpe, takma ad, ırk veya padok yazın..." autocomplete="off"></label><label>Padok / Ahır<select id="allAnimalPaddock"><option value="">Tüm Padoklar</option>{paddock_options}</select></label><label>Tür<select id="allAnimalKind"><option value="">Tümü</option><option>Dişi</option><option>Erkek</option><option>Buzağı</option></select></label></div></div><div id="allAnimalEmpty" class="empty-state">Filtrelere uyan aktif hayvan bulunamadı.</div><div class="card"><table id="allAnimalTable" class="mobile-animal-table all-animal-table"><thead><tr><th>Küpe</th><th>Takma Ad</th><th>Tür</th><th>Irk</th><th>Padok</th><th>Yaş</th></tr></thead><tbody>{trs}</tbody></table></div><script>document.addEventListener('DOMContentLoaded',function(){{var search=document.getElementById('allAnimalSearch'),paddock=document.getElementById('allAnimalPaddock'),kind=document.getElementById('allAnimalKind'),rows=[...document.querySelectorAll('#allAnimalTable tr.data-row')],empty=document.getElementById('allAnimalEmpty');function norm(v){{return (v||'').toLocaleLowerCase('tr-TR').trim()}}function apply(){{var q=norm(search.value),p=norm(paddock.value),k=norm(kind.value),shown=0;rows.forEach(function(r){{var ok=(!q||norm(r.innerText).includes(q))&&(!p||norm(r.dataset.paddock)===p)&&(!k||norm(r.dataset.kind)===k);r.style.display=ok?'':'none';if(ok)shown++;}});empty.style.display=shown?'none':'block';}}search.addEventListener('input',apply);paddock.addEventListener('change',apply);kind.addEventListener('change',apply);apply();}});</script>'''
            return self.send_html(page('Tüm Aktif Hayvanlar',body,'/all-animals',u,msg))

        if path=='/animals':
            edit=q.get('edit',[''])[0]
            term=q.get('q',[''])[0].strip()
            with db() as c:
                if term:
                    like=f"%{term}%"
                    rows=c.execute(
                        "select * from animals where gender='Dişi' and coalesce(status,'Aktif')='Aktif' and (tag like ? or nickname like ? or breed like ? or paddock like ?) order by tag",
                        (like,like,like,like)
                    ).fetchall()
                else:
                    rows=c.execute("select * from animals where gender='Dişi' and coalesce(status,'Aktif')='Aktif' order by tag").fetchall()
                rec=c.execute('select * from animals where id=?',(edit,)).fetchone() if edit else None
            trs=''.join('<tr><td><a class="animal-tag-btn" title="Hayvan kartını aç" href="/animal?id={0}">{1}</a></td><td>{2}</td><td>{3}</td><td>{4}</td><td>{5}</td><td>{6}</td><td><a class="btn alt" href="/animal-edit?id={0}">Düzenle</a>{7}</td></tr>'.format(r['id'],h(r['tag']),h(r['nickname']),h(r['gender']),h(r['breed']),h(r['paddock']),age_text(r['birth_date']),(' <a class="btn" href="/inseminations?animal='+str(r['id'])+'">Tohumlama</a>' if r['gender']=='Dişi' else '')+' <form class="inline-form" method="post" action="/animal-delete" onsubmit="return confirm(\'Bu hayvan ve bağlı kayıtları kalıcı olarak silmek istediğinize emin misiniz?\')"><input type="hidden" name="id" value="'+str(r['id'])+'"><button class="btn red">Sil</button></form>') for r in rows)
            search_options=''.join(f'<option value="{h(r["tag"])}">{h(r["nickname"])}</option>' for r in rows)
            table_rows=trs.replace('<tr>','<tr class="data-row">')
            body=f'''<h1>Dişi Hayvanlar</h1><div class="livebox"><input id="femaleLiveSearch" type="search" placeholder="Küpe, takma ad, ırk veya padok yazın..." autocomplete="off"><button type="button" class="btn alt live-clear" onclick="document.getElementById('femaleLiveSearch').value='';document.getElementById('femaleLiveSearch').dispatchEvent(new Event('input'))">Temizle</button></div><div id="femaleEmpty" class="empty-state">Eşleşen dişi hayvan bulunamadı.</div><div class="card"><table id="femaleLiveTable" class="mobile-animal-table female-table"><thead><tr><th>Küpe</th><th>Takma Ad</th><th>Cinsiyet</th><th>Irk</th><th>Padok</th><th>Yaş</th><th>İşlem</th></tr></thead><tbody>{table_rows}</tbody></table></div><script>document.addEventListener('DOMContentLoaded',function(){{liveTableFilter('femaleLiveSearch','femaleLiveTable','femaleEmpty');}});</script>'''
            return self.send_html(page('Hayvanlar',body,'/animals',u,msg))
        if path=='/males':
            edit=q.get('edit',[''])[0]
            term=q.get('q',[''])[0].strip()
            with db() as c:
                if term:
                    like=f"%{term}%"
                    rows=c.execute(
                        "select * from animals where gender='Erkek' and coalesce(status,'Aktif')='Aktif' and (tag like ? or nickname like ? or breed like ? or paddock like ?) order by tag",
                        (like,like,like,like)
                    ).fetchall()
                else:
                    rows=c.execute("select * from animals where gender='Erkek' and coalesce(status,'Aktif')='Aktif' order by tag").fetchall()
                rec=c.execute("select * from animals where id=? and gender='Erkek'",(edit,)).fetchone() if edit else None
            male_rows=[]
            for r in rows:
                days,daily,accumulated,current=animal_cost_values(r)
                male_rows.append(f'<tr><td><a class="animal-tag-btn" title="Hayvan kartını aç" href="/animal?id={r["id"]}">{h(r["tag"])}</a></td><td>{h(r["nickname"])}</td><td>{h(r["breed"])}</td><td>{h(r["paddock"])}</td><td>{days} gün</td><td>{money(r["purchase_price"])}</td><td><b>{money(current)}</b></td><td>{money(float(r['target_sale_price'] or 0)-current) if float(r['target_sale_price'] or 0)>0 else '-'}</td><td><a class="btn alt" href="/animal-edit?id={r["id"]}">Düzenle</a> <form class="inline-form" method="post" action="/animal-delete" onsubmit="return confirm(\'Bu hayvan ve bağlı kayıtları kalıcı olarak silmek istediğinize emin misiniz?\')"><input type="hidden" name="id" value="{r["id"]}"><button class="btn red">Sil</button></form></td></tr>')
            trs=''.join(male_rows) or '<tr><td colspan=8>Erkek hayvan kaydı yok</td></tr>'
            search_options=''.join(f'<option value="{h(r["tag"])}">{h(r["nickname"])}</option>' for r in rows)
            table_rows=trs.replace('<tr>','<tr class="data-row">')
            body=f'''<h1>Erkek Hayvanlar</h1><div class="livebox"><input id="maleLiveSearch" type="search" placeholder="Küpe, takma ad, ırk veya padok yazın..." autocomplete="off"><button type="button" class="btn alt live-clear" onclick="document.getElementById('maleLiveSearch').value='';document.getElementById('maleLiveSearch').dispatchEvent(new Event('input'))">Temizle</button></div><div id="maleEmpty" class="empty-state">Eşleşen erkek hayvan bulunamadı.</div><div class="card"><p class="mut">10 ayını dolduran erkek buzağılar otomatik olarak bu listeye geçer.</p><table id="maleLiveTable" class="mobile-animal-table male-table"><thead><tr><th>Küpe</th><th>Takma Ad</th><th>Irk</th><th>Padok</th><th>Bizde Kalma</th><th>Alış</th><th>Anlık Maliyet</th><th>Hedef Kâr</th><th>İşlem</th></tr></thead><tbody>{table_rows}</tbody></table></div><script>document.addEventListener('DOMContentLoaded',function(){{liveTableFilter('maleLiveSearch','maleLiveTable','maleEmpty');}});</script>'''
            return self.send_html(page('Erkek Hayvanlar',body,'/males',u,msg))
        if path in ('/archive/sold','/archive/slaughtered'):
            status='Satıldı' if path=='/archive/sold' else 'Kesildi'
            title='Satılan Hayvanlar' if status=='Satıldı' else 'Kesilen Hayvanlar'
            with db() as c:
                rows=c.execute("select * from animals where status=? order by exit_date desc,tag",(status,)).fetchall()
            trs=''.join(
                f'<tr><td><a class="animal-tag-btn" title="Hayvan kartını aç" href="/animal?id={r["id"]}">{h(r["tag"])}</a></td>'
                f'<td>{h(r["nickname"])}</td><td>{h(r["gender"])}</td><td>{h(r["breed"])}</td>'
                f'<td>{fmt_date(r["exit_date"])}</td><td>{h(r["exit_reason"])}</td><td>{money(r["sold_price"])}</td></tr>'
                for r in rows
            ) or '<tr><td colspan=7>Kayıt yok.</td></tr>'
            body=f'<h1>{title}</h1><div class="card"><p class="mut">Bu hayvanların geçmiş kayıtları silinmez; yalnızca aktif sürü listesinden çıkarılır.</p><table class="mobile-animal-table archive-animal-table"><tr><th>Küpe</th><th>Takma Ad</th><th>Cinsiyet</th><th>Irk</th><th>Çıkış Tarihi</th><th>Neden</th><th>Satış/Kesim Tutarı</th></tr>{trs}</table></div>'
            return self.send_html(page(title,body,path,u,msg))
        if path=='/pregnancy-edit':
            aid=q.get('animal_id',[''])[0]
            with db() as c:
                a=c.execute("select * from animals where id=? and gender='Dişi'",(aid,)).fetchone()
                rec=current_pregnancy_record(c,aid) if a else None
            if not a:return self.redirect('/animals','Dişi hayvan bulunamadı.')
            if not rec:return self.redirect('/animal?id='+str(aid),'Düzenlenecek aktif gebelik kaydı bulunamadı.')
            if not str(a['pregnancy_source'] or '').startswith('Satın Alındığında Gebe'):
                return self.redirect('/animal?id='+str(aid),'Bu düzenleme ekranı dışarıdan gebe alınan hayvanlar içindir.')
            source=str(a['pregnancy_source'] or '')
            mode='date' if 'Tohumlama Tarihi' in source else 'age'
            ref=(a['pregnancy_entry_date'] or a['purchase_date'] or date.today().isoformat())
            age=float(a['pregnancy_age_months_at_entry'] or 0)
            body=f'''<div class="actions"><a class="btn alt" href="/animal?id={aid}">← Hayvan Kartına Dön</a></div><h1>🤰 Gebelik Bilgisini Düzenle</h1>
            <div class="card"><p class="mut">Dışarıdan gebe alınan hayvanın tahmini gebelik bilgisini düzeltin. Kaydedildiğinde hayvan kartı ve Tohumlama kaydı birlikte güncellenir.</p>
            <form method="post" action="/pregnancy-edit" class="form"><input type="hidden" name="animal_id" value="{aid}">
            <label>Hayvan<input value="{h(a['tag'])} · {h(a['nickname'])}" disabled></label>
            <label>Gebelik Bilgisi<select name="pregnancy_info_mode" id="pregEditMode" onchange="syncPregEditMode()"><option value="age" {'selected' if mode=='age' else ''}>Gebelik Yaşı Biliniyor</option><option value="date" {'selected' if mode=='date' else ''}>Son Tohumlama Tarihi Biliniyor</option></select></label>
            <label id="pregEditAge">Yaklaşık Gebelik Yaşı (Ay)<input type="number" min="0.5" max="9" step="0.1" name="pregnancy_age_months" value="{age if age>0 else ''}"></label>
            <label id="pregEditDate">Son Tohumlama Tarihi<input type="date" name="known_insemination_date" value="{h(rec['insemination_date'])}"></label>
            <label>Bilginin Geçerli Olduğu Tarih<input type="date" name="pregnancy_entry_date" value="{h(ref)}" required></label>
            <div class="full"><button class="btn">💾 Gebelik Bilgisini Güncelle</button></div></form></div>
            <script>function syncPregEditMode(){{const m=document.getElementById('pregEditMode').value;document.getElementById('pregEditAge').style.display=m==='age'?'block':'none';document.getElementById('pregEditDate').style.display=m==='date'?'block':'none';}}syncPregEditMode();</script>'''
            return self.send_html(page('Gebelik Bilgisini Düzenle',body,'/animals',u,msg))
        if path=='/animal':
            aid=q.get('id',[''])[0]
            with db() as c:
                a=c.execute('select * from animals where id=?',(aid,)).fetchone()
                if not a:return self.send_html('Hayvan bulunamadı',404)
                ins=c.execute('select * from inseminations where animal_id=? order by attempt',(aid,)).fetchall()
                health=c.execute('select * from health where animal_id=? order by applied_date desc',(aid,)).fetchall()
                fin=c.execute('select * from finance where animal_id=? order by tx_date desc',(aid,)).fetchall()
                weights=c.execute('select * from weights where animal_id=? order by measure_date desc',(aid,)).fetchall()
                milk=c.execute('select * from milk where animal_id=? order by measure_date desc',(aid,)).fetchall()
                calves=c.execute('select * from calves where mother_id=? order by birth_date desc',(aid,)).fetchall()
                photos=c.execute('select * from animal_photos where animal_id=? order by created_at desc',(aid,)).fetchall()
            latest=ins[-1] if ins else None
            with db() as pc:
                active_preg=current_pregnancy_record(pc,aid) if a['gender']=='Dişi' else None
            preg=(active_preg['pregnancy_result'] if active_preg else (latest['pregnancy_result'] if latest else 'Kayıt yok'))
            due=(active_preg['due_date'] if active_preg else '')
            cls='pos' if active_preg else 'neg' if str(preg).strip().lower()=='negatif' else ''
            pregnancy_panel=''
            pregnancy_line=f'<p class="preg {cls}">Gebelik: {h(preg)}{(" · Tahmini doğum "+fmt_date(due)) if due else ""}</p>'
            if active_preg and due:
                try:
                    due_day=date.fromisoformat(due);raw_days=(due_day-date.today()).days
                    start_day=date.fromisoformat(active_preg['insemination_date']) if active_preg['insemination_date'] else date.today()-timedelta(days=max(0,280-max(0,raw_days)))
                    preg_days=max(0,min(280,(date.today()-start_day).days));preg_months=min(9,preg_days/(280/9));month_text=f'{preg_months:.1f}'.replace('.',',')
                    remain_text=f'{raw_days} gün' if raw_days>=0 else f'{abs(raw_days)} gün geçti';progress=max(0,min(100,round((preg_days/280)*100)))
                    pregnancy_panel=(f'<div style="margin:18px 0;padding:16px;border:1px solid #cfe5d6;border-radius:18px;background:linear-gradient(180deg,#f7fcf8,#eef8f1)">'
                        f'<div style="display:flex;justify-content:space-between;gap:10px;align-items:center;flex-wrap:wrap;margin-bottom:12px"><b style="font-size:18px;color:#176b3a">🤰 Gebelik Durumu</b><span class="status-badge status-preg">Pozitif</span></div>'
                        f'<div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(140px,1fr));gap:10px"><div class="pill" style="display:block;text-align:center;padding:12px"><span class="mut">Gebelik Süresi</span><br><b style="font-size:18px">≈ {month_text} ay</b></div><div class="pill" style="display:block;text-align:center;padding:12px"><span class="mut">Doğuma Kalan</span><br><b style="font-size:18px">{remain_text}</b></div><div class="pill" style="display:block;text-align:center;padding:12px"><span class="mut">Tahmini Doğum</span><br><b style="font-size:18px">{fmt_date(due)}</b></div></div>'
                        f'<div style="margin-top:14px"><div style="display:flex;justify-content:space-between;font-size:13px;color:#667a6d"><span>Gebelik ilerlemesi</span><b>%{progress}</b></div><div style="height:10px;background:#dcebe1;border-radius:999px;overflow:hidden;margin-top:6px"><div style="height:100%;width:{progress}%;background:#238a50;border-radius:999px"></div></div></div>'
                        + (f'<div style="margin-top:14px"><a class="btn alt" href="/pregnancy-edit?animal_id={aid}">✏️ Gebelik Bilgisini Düzenle</a></div>' if str(a["pregnancy_source"] or "").startswith("Satın Alındığında Gebe") else '')
                        + '</div>')
                    pregnancy_line=''
                except Exception:pass
            total_cost=sum(r['amount'] for r in fin if r['tx_type']=='Gider')+sum(r['cost'] or 0 for r in health)
            latest_weight=weights[0]['weight'] if weights else None
            first_weight=weights[-1]['weight'] if weights else (a['purchase_weight'] or None)
            try:
                first_date=date.fromisoformat(weights[-1]['measure_date']) if weights else date.fromisoformat(a['purchase_date'])
                perf_days=max(1,(date.fromisoformat(weights[0]['measure_date'])-first_date).days) if latest_weight is not None and first_weight is not None else 0
            except Exception: perf_days=0
            weight_gain=(float(latest_weight)-float(first_weight)) if latest_weight is not None and first_weight is not None else None
            daily_gain=(weight_gain/perf_days) if weight_gain is not None and perf_days>0 else None
            latest_milk=milk[0]['liters'] if milk else None
            total_income=sum(r['amount'] for r in fin if r['tx_type']=='Gelir')
            net_value=total_income-total_cost
            stay_days,daily_cost,accumulated_cost,current_cost=animal_cost_values(a)
            feed_ctx=animal_current_feed_context(a)
            target_profit=float(a['target_sale_price'] or 0)-current_cost if float(a['target_sale_price'] or 0)>0 else None
            period_perf=male_weight_performance(aid) if a['gender']=='Erkek' else None
            perf_labels={'good':('Hedefte / Üstünde','status-good'),'watch':('Takip Edilmeli','status-watch'),'low':('Düşük Artış','status-low'),'none':('Veri Yetersiz','status-none')}
            perf_label,perf_class=perf_labels[period_perf['status']] if period_perf else ('','')
            chart_html=weight_chart_svg(list(reversed(weights))) if a['gender']=='Erkek' else ''
            purchase_summary=(f'<div class="costbox"><h3>Canlı Anlık Maliyet ve Performans</h3><div class="quick-metrics"><span class="pill">Alış Fiyatı<br><b>{money(a["purchase_price"])}</b></span><span class="pill">Bizde Kaldığı Süre<br><b>{stay_days} gün</b></span><span class="pill">Birikmiş Yem + Bakım<br><b>{money(accumulated_cost)}</b></span><span class="pill">Anlık Toplam Maliyet<br><b>{money(current_cost)}</b></span><span class="pill">Toplam Kilo Artışı<br><b>{(str(round(weight_gain,1))+" kg") if weight_gain is not None else "-"}</b></span><span class="pill">Günlük Kilo Artışı<br><b>{(str(round(daily_gain,3))+" kg/gün") if daily_gain is not None else "-"}</b></span><span class="pill">Hedef Satış<br><b>{money(a["target_sale_price"]) if float(a["target_sale_price"] or 0)>0 else "-"}</b></span><span class="pill">Hedef Kâr<br><b>{money(target_profit) if target_profit is not None else "-"}</b></span></div><p class="mut">Günlük yem/rasyon: {money(feed_ctx["feed_cost"])} · Günlük bakım: {money(a["daily_care_cost"])} · Günlük toplam: {money(daily_cost)}</p><p class="mut">{("🌾 Padok rasyonu: "+h(feed_ctx["ration_name"])+" · "+fmt_date(feed_ctx["start_date"])+" tarihinden itibaren") if feed_ctx["source"]=="ration" else "Manuel sabit yem maliyeti kullanılıyor (padokta aktif rasyon yok)."}</p></div>') if a['gender']=='Erkek' else ''
            sale_box=(f'<div class="card" style="margin-top:14px"><h2>Erkek Hayvan Satışı</h2><p class="mut">Satış kaydı oluşturulduğunda hayvan Satılan Hayvanlar arşivine alınır ve net kâr otomatik hesaplanır.</p><form method="post" action="/animal/sale" class="form" onsubmit="return confirm(\'Bu hayvanı satıldı olarak işaretlemek istediğinize emin misiniz?\')"><input type="hidden" name="animal_id" value="{aid}"><label>Satış Tarihi<input type="date" name="sale_date" required value="{date.today().isoformat()}"></label><label>Satış Fiyatı (TL)<input type="number" name="sale_price" min="0" step="0.01" required value="{h(a["target_sale_price"])}"></label><label>Satış Kilosu (kg)<input type="number" name="sale_weight" min="0" step="0.1" value="{h(latest_weight)}"></label><label>Alıcı / Açıklama<input name="description"></label><div class="full"><button class="btn orange">Satışı Tamamla</button></div></form></div>') if a['gender']=='Erkek' and a['status']=='Aktif' else ''
            photo=f'<img class="photo" src="{h(a["photo_url"])}">' if a['photo_url'] else '<div class="photo">🐄</div>'
            gallery=''.join(f'<figure><img src="/uploads/{h(r["filename"])}"><figcaption>{h(r["caption"])}<br>{fmt_datetime(r["created_at"])}</figcaption></figure>' for r in photos) or '<p class="mut">Henüz fotoğraf yüklenmedi.</p>'
            itr=''.join(f'<tr><td>{r["attempt"]}</td><td>{fmt_date(r["insemination_date"])}</td><td>{h(r["pregnancy_result"])}</td><td>{fmt_date(r["due_date"])}</td></tr>' for r in ins) or '<tr><td colspan=4>Kayıt yok</td></tr>'
            htr=''.join(f'<tr><td>{fmt_date(r["applied_date"])}</td><td>{h(r["kind"])}</td><td>{h(r["product"])}</td><td>{money(r["cost"])}</td></tr>' for r in health) or '<tr><td colspan=4>Kayıt yok</td></tr>'
            weight_chron=list(reversed(weights)); wrows=[]
            for i,r in enumerate(weight_chron):
                gain_txt=daily_txt=monthly_txt='-'
                if i>0:
                    prev=weight_chron[i-1]
                    try:
                        wd=(date.fromisoformat(r['measure_date'])-date.fromisoformat(prev['measure_date'])).days
                        wg=float(r['weight'])-float(prev['weight'])
                        gain_txt=f'{wg:+.1f} kg'
                        if wd>0:
                            dd=wg/wd;daily_txt=f'{dd:.3f} kg/gün';monthly_txt=f'{dd*30:.1f} kg'
                    except Exception:pass
                wrows.append(f'<tr><td>{fmt_date(r["measure_date"])}</td><td>{r["weight"]} kg</td><td>{gain_txt}</td><td>{daily_txt}</td><td>{monthly_txt}</td><td>{h(r["notes"])}</td></tr>')
            wtr=''.join(reversed(wrows)) or '<tr><td colspan=6>Kayıt yok</td></tr>'
            mtr=''.join(f'<tr><td>{fmt_date(r["measure_date"])}</td><td>{r["liters"]} L</td><td>{h(r["notes"])}</td></tr>' for r in milk) or '<tr><td colspan=3>Kayıt yok</td></tr>'
            ctr=''.join(f'<tr><td>{h(r["tag"])}</td><td>{fmt_date(r["birth_date"])}</td><td>{h(r["gender"])}</td></tr>' for r in calves) or '<tr><td colspan=3>Kayıt yok</td></tr>'
            back='/males' if a['gender']=='Erkek' else '/animals'; edit_url='/animal-edit?id='+str(aid)
            body=f'''<div class="actions"><a class="btn alt" href="{back}">← Hayvanlara Dön</a><a class="btn" href="{edit_url}">Bilgileri Düzenle</a><a class="btn blue" href="/animal/print?id={aid}">Kimlik Kartını Yazdır</a></div><div class="card profile">{photo}<div><h1>{h(a['tag'])}</h1><h2>{h(a['nickname'])}</h2><span class="pill">{h(a['gender'])}</span><span class="pill">{h(a['breed'])}</span><span class="pill">Padok: {h(a['paddock']) or '-'}</span><span class="pill">Durum: {h(a['status'])}</span>{pregnancy_panel}{pregnancy_line}<div class="quick-metrics"><span class="pill">Yaş<br><b>{age_text(a['birth_date'])}</b></span><span class="pill">Son Kilo<br><b>{(str(latest_weight)+' kg') if latest_weight is not None else '-'}</b></span><span class="pill">Son Süt<br><b>{(str(latest_milk)+' L') if latest_milk is not None else '-'}</b></span><span class="pill">Net Değer<br><b>{money(net_value)}</b></span></div><p>Toplam masraf: <b>{money(total_cost)}</b> · Buzağı: <b>{len(calves)}</b></p>{purchase_summary}<p>{h(a['notes'])}</p></div></div><div class="two" style="margin-top:14px"><div class="card"><h2>Tohumlama ve Gebelik</h2><table><tr><th>Deneme</th><th>Tarih</th><th>Sonuç</th><th>Tahmini Doğum</th></tr>{itr}</table></div><div class="card"><h2>Buzağıları</h2><table><tr><th>Küpe</th><th>Doğum</th><th>Cinsiyet</th></tr>{ctr}</table></div></div><div class="two" style="margin-top:14px"><div class="card"><h2>{'Aylık Tartım ve Besi Performansı' if a['gender']=='Erkek' else 'Kilo Geçmişi'}</h2>{(f'<div class="costbox"><span class="perf-badge {perf_class}">{perf_label}</span><div class="quick-metrics"><span class="pill">Son Dönem Artışı<br><b>{period_perf["gain"]:+.1f} kg</b></span><span class="pill">Tartım Aralığı<br><b>{period_perf["days"]} gün</b></span><span class="pill">Günlük Artış<br><b>{period_perf["daily"]:.3f} kg/gün</b></span><span class="pill">30 Günlük Tahmin<br><b>{period_perf["monthly"]:.1f} kg</b></span></div></div>' if period_perf and period_perf['daily'] is not None else '<p class="mut">Performans hesabı için en az iki tartım girin.</p>') if a['gender']=='Erkek' else ''}<form method="post" action="/animal/weight" class="actions"><input type="hidden" name="animal_id" value="{aid}"><input type="date" name="measure_date" required value="{date.today().isoformat()}"><input type="number" step="0.1" name="weight" placeholder="kg" required><input name="notes" placeholder="Not"><button class="btn">Tartım Ekle</button></form>{chart_html}<table style="margin-top:12px"><tr><th>Tarih</th><th>Kilo</th><th>Fark</th><th>Günlük Artış</th><th>30 Günlük</th><th>Not</th></tr>{wtr}</table></div><div class="card"><h2>Süt Verimi</h2><form method="post" action="/animal/milk" class="actions"><input type="hidden" name="animal_id" value="{aid}"><input type="date" name="measure_date" required value="{date.today().isoformat()}"><input type="number" step="0.1" name="liters" placeholder="Litre" required><input name="notes" placeholder="Not"><button class="btn">Ekle</button></form><table><tr><th>Tarih</th><th>Litre</th><th>Not</th></tr>{mtr}</table></div></div>{sale_box}<div class="card" style="margin-top:14px"><h2>Fotoğraf Galerisi</h2><form method="post" action="/animal/photo" enctype="multipart/form-data" class="uploadbox"><input type="hidden" name="animal_id" value="{aid}"><label>Fotoğraf seç veya telefondan çek<input type="file" name="photo_file" accept="image/*" required></label><input name="caption" placeholder="Açıklama (isteğe bağlı)"><button class="btn">Fotoğrafı Yükle</button><div class="camera-note">Mobil tarayıcıda arka kamera açılır. Fotoğraflar uygulama klasöründeki uploads dizininde saklanır; bu klasörü de düzenli kopyalayın.</div></form><div class="gallery" style="margin-top:14px">{gallery}</div></div><div class="card" style="margin-top:14px"><h2>Sağlık Geçmişi</h2><table><tr><th>Tarih</th><th>Tür</th><th>İşlem</th><th>Maliyet</th></tr>{htr}</table></div>'''
            return self.send_html(page('Hayvan Kartı',body,'/animals',u,msg))
        if path=='/animal/print':
            aid=q.get('id',[''])[0]
            with db() as c:a=c.execute('select * from animals where id=?',(aid,)).fetchone(); ins=c.execute('select * from inseminations where animal_id=? order by attempt',(aid,)).fetchall()
            if not a:return self.send_html('Hayvan bulunamadı',404)
            latest=ins[-1] if ins else None
            return self.send_html(f'''<!doctype html><html lang="tr"><head><meta charset="utf-8"><title>ÇiftlikPro Hayvan Kartı</title><style>body{{font-family:Arial;padding:30px}}.box{{border:2px solid #176b3a;border-radius:16px;padding:24px;max-width:700px}}h1{{color:#176b3a}}table{{width:100%;border-collapse:collapse}}td{{padding:8px;border-bottom:1px solid #ddd}}@media print{{button{{display:none}}}}
/* HOTFIX4 UX: hedef kartları çalışma masasının üstünde, iki satır ve okunaklı sabit kokpit */
body:has(.workbench-shell) .workbench-shell{{display:block!important;padding:14px!important;overflow:visible!important}}
body:has(.workbench-shell) .workbench-shell>.target-workspace{{display:block!important}}
body:has(.workbench-shell) .target-controlbar{{display:grid!important;grid-template-columns:auto minmax(0,1fr)!important;gap:12px!important;margin:8px 0!important;padding:8px 10px!important}}
body:has(.workbench-shell) .target-compare-sticky{{position:sticky!important;top:62px!important;height:auto!important;max-height:none!important;overflow:visible!important;margin:8px 0 10px!important;padding:9px 10px!important;background:rgba(248,251,249,.985)!important;backdrop-filter:blur(8px)!important;z-index:40!important;box-shadow:0 8px 22px rgba(20,65,40,.14)!important}}
body:has(.workbench-shell) .target-compare-title{{height:auto!important;margin:0 2px 7px!important;font-size:13px!important}}
body:has(.workbench-shell) .target-compare-title>b{{font-size:14px!important}}
body:has(.workbench-shell) .target-compare-title span{{font-size:10px!important;max-width:none!important}}
body:has(.workbench-shell) .nutri-mini-grid{{position:static!important;display:grid!important;grid-template-columns:repeat(5,minmax(0,1fr))!important;grid-template-rows:none!important;grid-auto-rows:112px!important;height:auto!important;gap:7px!important;padding:0!important;margin:0!important;border:0!important;background:transparent!important;box-shadow:none!important;overflow:visible!important}}
body:has(.workbench-shell) .nutri-mini{{min-height:112px!important;height:112px!important;padding:7px 8px!important;display:grid!important;grid-template-rows:18px 1fr 34px!important;border-radius:10px!important;overflow:hidden!important}}
body:has(.workbench-shell) .nutri-card-title{{font-size:13px!important;line-height:18px!important;font-weight:900!important;margin:0!important}}
body:has(.workbench-shell) .nutri-compare-body{{display:grid!important;grid-template-columns:1fr 1fr!important;align-items:center!important;min-height:0!important}}
body:has(.workbench-shell) .nutri-side{{padding:2px 5px!important;min-width:0!important;text-align:center!important}}
body:has(.workbench-shell) .nutri-side span{{font-size:9px!important;line-height:1!important;font-weight:800!important}}
body:has(.workbench-shell) .nutri-side b{{font-size:18px!important;line-height:1.05!important;margin-top:4px!important;white-space:normal!important;overflow-wrap:anywhere!important}}
body:has(.workbench-shell) .nutri-card-footer{{min-height:34px!important;height:34px!important;padding:4px 3px 2px!important;display:grid!important;grid-template-columns:1fr!important;align-content:center!important;justify-items:center!important;gap:1px!important}}
body:has(.workbench-shell) .nutri-card-footer em{{font-size:12px!important;line-height:1!important;font-weight:900!important;white-space:nowrap!important}}
body:has(.workbench-shell) .nutri-card-footer .nutri-diff{{font-size:10px!important;line-height:1.05!important;margin:1px 0 0!important;white-space:nowrap!important;font-weight:800!important}}
body:has(.workbench-shell) #target-mini-cost{{grid-column:auto!important}}
body:has(.workbench-shell) #ration-workbench{{margin-top:0!important}}
@media(max-width:1350px){{body:has(.workbench-shell) .nutri-mini-grid{{grid-template-columns:repeat(4,minmax(0,1fr))!important}}}}
@media(max-width:1080px){{body:has(.workbench-shell) .nutri-mini-grid{{grid-template-columns:repeat(3,minmax(0,1fr))!important}}body:has(.workbench-shell) .target-compare-sticky{{position:static!important}}}}
@media(max-width:820px){{body:has(.workbench-shell) .nutri-mini-grid{{grid-template-columns:repeat(2,minmax(0,1fr))!important;grid-auto-rows:116px!important}}body:has(.workbench-shell) .nutri-mini{{height:116px!important;min-height:116px!important}}}}
</style></head><body><button onclick="print()">Yazdır / PDF Kaydet</button><div class="box"><h1>🐄 ÇiftlikPro Hayvan Kimlik Kartı</h1><table><tr><td>Küpe</td><td><b>{h(a['tag'])}</b></td></tr><tr><td>Takma Ad</td><td>{h(a['nickname'])}</td></tr><tr><td>Cinsiyet / Irk</td><td>{h(a['gender'])} / {h(a['breed'])}</td></tr><tr><td>Doğum / Yaş</td><td>{fmt_date(a['birth_date'])} / {age_text(a['birth_date'])}</td></tr><tr><td>Padok</td><td>{h(a['paddock'])}</td></tr><tr><td>Gebelik</td><td>{h(latest['pregnancy_result'] if latest else 'Kayıt yok')}</td></tr><tr><td>Tahmini Doğum</td><td>{fmt_date(latest['due_date'] if latest else '')}</td></tr><tr><td>Not</td><td>{h(a['notes'])}</td></tr></table></div></body></html>''')
        if path=='/calves':
            edit=q.get('edit',[''])[0]
            term=q.get('q',[''])[0].strip()
            with db() as c:
                mothers=c.execute("select id,tag,nickname from animals where gender='Dişi' order by tag").fetchall()
                if term:
                    like=f"%{term}%"
                    rows=c.execute(
                        '''select calves.*,animals.tag mother_tag,animals.nickname mother_name
                           from calves join animals on animals.id=calves.mother_id
                           where calves.promoted_animal_id is null and
                           (calves.tag like ? or animals.tag like ? or animals.nickname like ?)
                           order by calves.birth_date desc''',
                        (like,like,like)
                    ).fetchall()
                else:
                    rows=c.execute(
                        'select calves.*,animals.tag mother_tag,animals.nickname mother_name from calves join animals on animals.id=calves.mother_id where calves.promoted_animal_id is null order by calves.birth_date desc'
                    ).fetchall()
                rec=c.execute('select * from calves where id=?',(edit,)).fetchone() if edit else None
            opts=''.join(f'<option value="{m["id"]}" {"selected" if rec and rec["mother_id"]==m["id"] else ""}>{h(m["tag"])} - {h(m["nickname"])}</option>' for m in mothers)
            trs=''.join(f'<tr><td><a class="animal-tag-btn" title="Buzağı kartını aç" href="/calf?id={r["id"]}">{h(r["tag"])}</a></td><td>{h(r["mother_tag"])} {h(r["mother_name"])}</td><td>{h(r["father_tag"])}</td><td>{h(fmt_date(r["birth_date"]))}</td><td>{age_text(r["birth_date"])}</td><td>{h(r["gender"])}</td><td><a class="btn alt" href="/calf-edit?id={r["id"]}">Düzenle</a> <form class="inline-form" method="post" action="/calf-delete" onsubmit="return confirm(\'Bu buzağı kaydını kalıcı olarak silmek istediğinize emin misiniz?\')"><input type="hidden" name="id" value="{r["id"]}"><button class="btn red">Sil</button></form></td></tr>' for r in rows)
            search_options=''.join(f'<option value="{h(r["tag"])}">{h(r["mother_tag"])}</option>' for r in rows)
            table_rows=trs.replace('<tr>','<tr class="data-row">')
            body=f'''<h1>Buzağılar</h1><div class="livebox"><input id="calfLiveSearch" type="search" placeholder="Buzağı küpesi, anne küpesi veya anne adı yazın..." autocomplete="off"><button type="button" class="btn alt live-clear" onclick="document.getElementById('calfLiveSearch').value='';document.getElementById('calfLiveSearch').dispatchEvent(new Event('input'))">Temizle</button></div><div id="calfEmpty" class="empty-state">Eşleşen buzağı bulunamadı.</div><div class="card"><table id="calfLiveTable" class="mobile-animal-table calf-table"><thead><tr><th>Küpe</th><th>Anne</th><th>Baba</th><th>Doğum</th><th>Yaş</th><th>Cinsiyet</th><th>İşlem</th></tr></thead><tbody>{table_rows}</tbody></table></div><script>document.addEventListener('DOMContentLoaded',function(){{liveTableFilter('calfLiveSearch','calfLiveTable','calfEmpty');}});</script>'''
            return self.send_html(page('Buzağılar',body,'/calves',u,msg))
        if path=='/calf':
            cid=q.get('id',[''])[0]
            with db() as c:
                calf=c.execute('select calves.*,animals.tag mother_tag,animals.nickname mother_name from calves join animals on animals.id=calves.mother_id where calves.id=?',(cid,)).fetchone()
                if not calf:return self.send_html('Buzağı bulunamadı',404)
                health_rows=c.execute('select * from health where calf_id=? order by applied_date desc,id desc',(cid,)).fetchall()
                weight_rows=c.execute('select * from calf_weights where calf_id=? order by measure_date desc,id desc',(cid,)).fetchall()
            promoted=''
            if calf['promoted_animal_id']:promoted=f'<p class="flash">Bu kayıt 10 ayını doldurduğu için yetişkin karta aktarıldı. <a class="taglink" href="/animal?id={calf["promoted_animal_id"]}">Yeni hayvan kartını aç</a></p>'
            icon='🐮' if calf['gender']=='Dişi' else '🐂'
            photo=f'<img src="{h(calf["photo_url"])}" alt="Buzağı">' if calf['photo_url'] else icon
            last_weight=float(weight_rows[0]['weight']) if weight_rows else None
            health_html=''.join(f'<tr><td>{fmt_date(r["applied_date"])}</td><td>{h(r["kind"])}</td><td>{h(r["product"])}</td><td>{fmt_date(r["next_date"])}</td><td>{h(r["notes"])}</td></tr>' for r in health_rows) or '<tr><td colspan="5">Henüz sağlık/tedavi kaydı yok.</td></tr>'
            weight_html=''.join(f'<tr><td>{fmt_date(r["measure_date"])}</td><td><b>{float(r["weight"]):.1f} kg</b></td><td>{h(r["notes"])}</td></tr>' for r in weight_rows) or '<tr><td colspan="3">Henüz tartım kaydı yok.</td></tr>'
            body=f'''<div class="actions"><a class="btn alt" href="/calves">← Buzağılara Dön</a><a class="btn" href="/calf-edit?id={cid}">Düzenle</a></div>{promoted}
            <div class="card profile"><div class="photo">{photo}</div><div><h1>{h(calf["tag"])}</h1><span class="pill">{h(calf["gender"])}</span><span class="pill">Yaş: {age_text(calf["birth_date"])}</span>
            <p>Takma ad: <b>{h(calf["nickname"]) or "-"}</b></p><p>Irk: <b>{h(calf["breed"]) or "-"}</b> · Padok: <b>{h(calf["paddock"]) or "-"}</b></p>
            <p>Doğum tarihi: <b>{fmt_date(calf["birth_date"])}</b></p><p>Anne: <a class="taglink" href="/animal?id={calf["mother_id"]}">{h(calf["mother_tag"])} {h(calf["mother_name"])}</a></p><p>Baba: <b>{h(calf["father_tag"]) or "-"}</b></p>
            <p>Son kilo: <b>{f"{last_weight:.1f} kg" if last_weight is not None else "-"}</b></p><p>Alış: <b>{fmt_date(calf["purchase_date"]) or "-"}</b> · <b>{money(calf["purchase_price"])}</b></p><p>{h(calf["notes"])}</p></div></div>
            <div class="grid" style="margin-top:14px"><div class="card"><h2>📷 Fotoğraf</h2><form method="post" action="/calf/photo" enctype="multipart/form-data" class="form" data-smart-photo-form="1"><input type="hidden" name="calf_id" value="{cid}"><label class="full">Kamera / Galeri<input type="file" name="photo_file" accept="image/*" required></label><div class="full"><button class="btn">Fotoğrafı Yükle</button></div></form></div>
            <div class="card"><h2>⚖️ Kilo / Gelişim</h2><form method="post" action="/calf/weight" class="form"><input type="hidden" name="calf_id" value="{cid}"><label>Tarih<input type="date" name="measure_date" value="{date.today().isoformat()}" required></label><label>Kilo (kg)<input type="number" step="0.1" min="0.1" name="weight" required></label><label class="full">Not<input name="notes"></label><div class="full"><button class="btn">Tartımı Kaydet</button></div></form></div></div>
            <div class="card" style="margin-top:14px"><h2>💉 Sağlık / Tedavi Geçmişi</h2><p><a class="btn" href="/health">Sağlık Kaydı Ekle</a></p><div class="tablewrap"><table><tr><th>Tarih</th><th>Tür</th><th>Ürün/İşlem</th><th>Sonraki</th><th>Not</th></tr>{health_html}</table></div></div>
            <div class="card" style="margin-top:14px"><h2>⚖️ Kilo Geçmişi</h2><div class="tablewrap"><table><tr><th>Tarih</th><th>Kilo</th><th>Not</th></tr>{weight_html}</table></div></div>'''
            return self.send_html(page('Buzağı Kartı',body,'/calves',u,msg))
        if path=='/estrus-edit':
            eid=q.get('id',[''])[0]
            with db() as c:
                rec=c.execute("select e.*,a.tag,a.nickname from estrus_records e join animals a on a.id=e.animal_id where e.id=?",(eid,)).fetchone()
            if not rec:return self.redirect('/estrus','Kızgınlık kaydı bulunamadı.')
            body=f'''<h1>✏️ Kızgınlık Kaydını Düzenle</h1><div class="card"><form method="post" action="/estrus-edit" class="form"><input type="hidden" name="id" value="{rec['id']}"><label>Hayvan<input value="{h(rec['tag'])} · {h(rec['nickname'])}" disabled></label><label>Kızgınlık Tarihi<input type="date" name="estrus_date" max="{date.today().isoformat()}" value="{h(rec['estrus_date'])}" required></label><label class="full">Gözlenen Belirtiler<input name="signs" value="{h(rec['signs'])}" placeholder="Gözlenen belirtiler"></label><label class="full">Not<textarea name="notes" rows="3">{h(rec['notes'])}</textarea></label><div class="full actions"><button class="btn">💾 Değişiklikleri Kaydet</button><a class="btn alt" href="/estrus">İptal</a></div></form></div>'''
            return self.send_html(page('Kızgınlık Kaydı Düzenle',body,'/estrus',u,msg))
        if path=='/estrus':
            with db() as c:
                female_rows=c.execute("select id,tag,nickname from animals where gender='Dişi' and coalesce(status,'Aktif')='Aktif' order by tag").fetchall()
                females=[a for a in female_rows if not is_currently_pregnant(c,a['id'])]
                rows=c.execute("select e.*,a.tag,a.nickname from estrus_records e join animals a on a.id=e.animal_id order by e.estrus_date desc,e.id desc").fetchall()
                latest={}
                for r in rows:
                    if r['animal_id'] not in latest: latest[r['animal_id']]=r
            today=date.today(); upcoming=[]
            with db() as c:
                for aid,r in latest.items():
                    if is_currently_pregnant(c,aid):continue
                    cycle=next_estrus_cycle(c,r,today)
                    if cycle and cycle['end']>=today and cycle['start']<=today+timedelta(days=30):
                        upcoming.append((cycle['start'],cycle['center'],cycle['end'],r,cycle['cycle_no']))
            upcoming.sort(key=lambda x:x[1])
            # V3.9.17: Kızgınlık seçicisinde takma adı boş hayvanlarda görünen etiket ile
            # gizli animal_id eşleşmesi aynı kaynaktan üretilir. Böylece "KÜPE · " / "KÜPE"
            # farkı nedeniyle geçerli dişinin reddedilmesi engellenir.
            estrus_picker_rows=[{'id':r['id'],'tag':str(r['tag'] or ''),'nickname':str(r['nickname'] or '')} for r in females]
            for x in estrus_picker_rows:
                x['label']=x['tag'] + ((' · '+x['nickname']) if x['nickname'] else '')
            opts=''.join(f'<option value="{h(x["label"])}"></option>' for x in estrus_picker_rows)
            estrus_picker_data=json.dumps(estrus_picker_rows,ensure_ascii=False)
            cards=[]
            for a,center,e,r,cycle_no in upcoming:
                color='#e27b1f' if a<=today<=e else '#238a50'
                if a<=today<=e:
                    action=f'''<form method="post" action="/estrus-inseminate" onsubmit="return confirm('Bu hayvan bugün tohumlandı olarak Tohumlama kayıtlarına aktarılsın mı?')"><input type="hidden" name="estrus_id" value="{r['id']}"><button class="btn orange">🌱 Bugün Tohumlandı</button></form>'''
                else:
                    action=f'''<form method="post" action="/estrus-send" onsubmit="return confirm('Bu hayvan Tohumlama Takibi ekranına gönderilsin mi? Bu kızgınlık kartı kapanacaktır.')"><input type="hidden" name="estrus_id" value="{r['id']}"><input type="hidden" name="cycle_no" value="{cycle_no}"><button class="btn orange">🌱 Tohumlamaya Gönder</button></form><span class="mut" style="align-self:center">Pencere {fmt_date(a.isoformat())} tarihinde başlıyor</span>'''
                cards.append(f'''<div class="alertitem" style="border-left-color:{color}"><b>🐄 <a class="taglink" href="/animal?id={r['animal_id']}">{h(r['tag'])} {h(r['nickname'])}</a></b><br><span class="mut">Beklenen pencere: {fmt_date(a.isoformat())} – {fmt_date(e.isoformat())} · En olası: {fmt_date(center.isoformat())}</span><div class="estrus-actions">{action}<form method="post" action="/estrus-skip" onsubmit="return confirm('Bu östrus dönemi atlandı olarak işaretlenecek. Emin misiniz?')"><input type="hidden" name="estrus_id" value="{r['id']}"><input type="hidden" name="cycle_no" value="{cycle_no}"><input type="hidden" name="return_to" value="/estrus"><button class="btn alt">⏭️ Bu Östrusu Atla</button></form><a class="btn alt" href="/estrus-edit?id={r['id']}">✏️ Düzenle</a></div></div>''')
            cards_html=''.join(cards) or '<p class="mut">Önümüzdeki 14 gün için beklenen kızgınlık kaydı yok.</p>'
            history=[]
            for r in rows:
                try:
                    d=date.fromisoformat(r['estrus_date']); center=d+timedelta(days=21); window=f'{fmt_date((d+timedelta(days=18)).isoformat())} – {fmt_date((d+timedelta(days=24)).isoformat())}'
                except Exception: center=None; window='-'
                inseminate_action=''
                decision_badge=''
                try:
                    decisions=c.execute("select cycle_no,decision,decision_date from estrus_decisions where estrus_id=? order by cycle_no desc",(r['id'],)).fetchall()
                    skipped=[x for x in decisions if str(x['decision'])=='Atlandı']
                    inseminated=[x for x in decisions if str(x['decision'])=='Tohumlandı']
                    if inseminated:
                        x=inseminated[0]
                        decision_badge=f'''<div class="estrus-decision-badge estrus-done">🌱 Tohumlandı<br><small>{fmt_date(x['decision_date'])}</small></div>'''
                    elif skipped:
                        x=skipped[0]
                        next_center=d+timedelta(days=21*(int(x['cycle_no'])+1))
                        decision_badge=f'''<div class="estrus-decision-badge estrus-skipped">⏭️ Sonraki Döngüye Ertelendi<br><small>Tahmini: {fmt_date(next_center.isoformat())}</small></div>'''
                    elif 0 <= (today-d).days <= 1:
                        inseminate_action=f'''<form method="post" action="/estrus-inseminate" class="inline-form" onsubmit="return confirm('Bu hayvan bugün tohumlandı olarak Tohumlama kayıtlarına aktarılsın mı?')"><input type="hidden" name="estrus_id" value="{r['id']}"><button class="btn orange">🌱 Bugün Tohumlandı</button></form>'''
                except Exception: pass
                history.append(f'''<tr class="data-row"><td><a class="taglink" href="/animal?id={r['animal_id']}">{h(r['tag'])} {h(r['nickname'])}</a></td><td>{fmt_date(r['estrus_date'])}</td><td>{h(r['signs']) or '-'}</td><td>{window}</td><td>{fmt_date(center.isoformat()) if center else '-'}</td><td>{h(r['notes']) or '-'}</td><td>{decision_badge}{inseminate_action}<a class="btn alt" href="/estrus-edit?id={r['id']}">✏️ Düzenle</a><form method="post" action="/estrus-delete" class="inline-form" onsubmit="return confirm('Bu kızgınlık kaydı silinsin mi?')"><input type="hidden" name="id" value="{r['id']}"><button class="btn red">Sil</button></form></td></tr>''')
            history_html=''.join(history) or '<tr><td colspan="7">Henüz kayıt yok.</td></tr>'
            body=f'''<div class="estrus-erp-shell"><h1>🌸 Kızgınlık & Tohumlama Takibi</h1><p class="mut">Dişi hayvanların gözlenen kızgınlıklarını kaydedin. Sistem 18–24 günlük takip penceresini ve 21. günü merkez tahmin olarak gösterir. Tahminler gözlem planlaması içindir.</p></div>
            <div class="two"><div class="card"><h2>Yeni Kızgınlık Kaydı</h2><form method="post" action="/estrus" class="form" onsubmit="var b=this.querySelector('button[type=submit]');if(b.disabled)return false;b.disabled=true;b.textContent='Kaydediliyor…';return true;"><label>Dişi Hayvan<input id="estrusAnimalSearch" type="search" list="estrusAnimalOptions" placeholder="Küpe veya takma ad yazın..." autocomplete="off" required><input type="hidden" name="animal_id" id="estrusAnimalId"><datalist id="estrusAnimalOptions">{opts}</datalist><span class="mut">Gebe hayvanlar doğum ve buzağı kaydı tamamlanana kadar listelenmez.</span></label><label>Kızgınlık Tarihi<input type="date" name="estrus_date" max="{today.isoformat()}" value="{today.isoformat()}" required></label><label class="full">Gözlenen Belirtiler<input name="signs" placeholder="Örn. üzerine atlamaya izin verme, huzursuzluk, şeffaf akıntı"></label><label class="full">Not<textarea name="notes" rows="3" placeholder="Ek gözlemler..."></textarea></label><div class="full"><button class="btn">💾 Kızgınlığı Kaydet</button></div></form></div><div class="card"><h2>📅 Yaklaşan Kızgınlıklar</h2>{cards_html}</div></div>
            <div class="card" style="margin-top:14px"><h2>Kızgınlık Geçmişi</h2><div class="livebox"><input id="estrusLiveSearch" type="search" placeholder="Küpe, takma ad, belirti veya not ara..." autocomplete="off"><button type="button" class="btn alt live-clear" onclick="document.getElementById('estrusLiveSearch').value='';document.getElementById('estrusLiveSearch').dispatchEvent(new Event('input'))">Temizle</button></div><div id="estrusEmpty" class="empty-state">Eşleşen kızgınlık kaydı bulunamadı.</div><div style="overflow:auto"><table id="estrusLiveTable" class="estrus-table"><tr><th>Hayvan</th><th>Gözlem Tarihi</th><th>Belirtiler</th><th>18–24 Gün Penceresi</th><th>21. Gün</th><th>Not</th><th>İşlem</th></tr>{history_html}</table></div></div>
            <script>const estrusPicker={estrus_picker_data};const estrusSearch=document.getElementById('estrusAnimalSearch'),estrusId=document.getElementById('estrusAnimalId');function normEstrus(v){{return (v||'').trim().toLocaleLowerCase('tr-TR');}}function syncEstrusAnimal(){{const v=normEstrus(estrusSearch.value);const m=estrusPicker.find(x=>normEstrus(x.label)===v||normEstrus(x.tag)===v);estrusId.value=m?m.id:'';}}estrusSearch.addEventListener('input',syncEstrusAnimal);estrusSearch.addEventListener('change',syncEstrusAnimal);document.addEventListener('DOMContentLoaded',function(){{syncEstrusAnimal();liveTableFilter('estrusLiveSearch','estrusLiveTable','estrusEmpty');}});</script>'''
            return self.send_html(page('Kızgınlık Takibi',body,'/estrus',u,msg))
        if path=='/inseminations':
            aid=q.get('animal',[''])[0]
            estrus_id=q.get('estrus',[''])[0]
            with db() as c:
                female_rows=c.execute("select id,tag,nickname from animals where gender='Dişi' and coalesce(status,'Aktif')='Aktif' order by tag").fetchall()
                females=[a for a in female_rows if not is_currently_pregnant(c,a['id'])]
                all_rows=c.execute('''select i.*,a.tag,a.nickname,a.pregnancy_source from inseminations i join animals a on a.id=i.animal_id order by a.tag,i.attempt,i.insemination_date''').fetchall()
                estrus_context=c.execute('select * from estrus_records where id=? and animal_id=?',(estrus_id,aid)).fetchone() if estrus_id and aid else None
            grouped={}
            for r in all_rows:grouped.setdefault(r['animal_id'],[]).append(r)
            waiting=sum(1 for records in grouped.values() if str(records[-1]['pregnancy_result'] or '').strip().lower() in ('bekleniyor',''))
            pregnant=sum(1 for records in grouped.values() if is_pregnant_value(records[-1]['pregnancy_result']))
            third_attempt=sum(1 for records in grouped.values() if int(records[-1]['attempt'] or 0)>=3)
            month_prefix=date.today().strftime('%Y-%m')
            this_month=sum(1 for r in all_rows if str(r['insemination_date'] or '').startswith(month_prefix))
            next_attempts={a['id']:(max([int(x['attempt']) for x in grouped.get(a['id'],[])],default=0)+1) for a in females}
            selected=next((a for a in females if str(a['id'])==aid),None)
            selected_label=(f"{selected['tag']} · {selected['nickname']}" if selected else '')
            picker_options=''.join(f'<option value="{h(a["tag"])} · {h(a["nickname"])}" data-id="{a["id"]}" data-attempt="{next_attempts[a["id"]]}"></option>' for a in females)
            picker_data=json.dumps([{'id':a['id'],'tag':a['tag'] or '', 'nickname':a['nickname'] or '', 'attempt':next_attempts[a['id']]} for a in females],ensure_ascii=False)
            latest_rows=[]
            for animal_id,records in grouped.items():
                latest=records[-1]
                result=str(latest['pregnancy_result'] or 'Belirsiz')
                if is_pregnant_value(result):
                    external=str(latest['pregnancy_source'] or '').startswith('Satın Alındığında Gebe')
                    badge='<span class="status-badge status-preg">'+('Gebe · Dışarıdan Gebe' if external else 'Gebe')+'</span>'
                elif result.strip().lower()=='negatif':badge='<span class="status-badge status-neg">Gebe Değil</span>'
                elif result.strip().lower()=='bekleniyor':badge='<span class="status-badge status-wait">Kontrol Bekliyor</span>'
                else:badge='<span class="status-badge status-unknown">Belirsiz</span>'
                hist=[]
                for rec in reversed(records):
                    rr=str(rec['pregnancy_result'] or 'Belirsiz')
                    if is_pregnant_value(rr):
                        external=str(rec['pregnancy_source'] or '').startswith('Satın Alındığında Gebe')
                        rb='<span class="status-badge status-preg">'+('Gebe · Dışarıdan' if external else 'Gebe')+'</span>'
                    elif rr.strip().lower()=='negatif':rb='<span class="status-badge status-neg">Gebe Değil</span>'
                    elif rr.strip().lower()=='bekleniyor':rb='<span class="status-badge status-wait">Kontrol Bekliyor</span>'
                    else:rb='<span class="status-badge status-unknown">Belirsiz</span>'
                    hist.append(f'''<tr><td>{rec["attempt"]}. deneme</td><td>{h(fmt_date(rec["insemination_date"]))}<div class="mut" style="margin-top:4px">🐂 {h(rec['bull_tag']) or 'Küpe yok'}{(' · '+h(rec['bull_name'])) if rec['bull_name'] else ''}<br>👤 {h(rec['inseminator']) or 'Tohumlayan belirtilmedi'}</div></td><td>{rb}</td><td>{h(fmt_date(rec["due_date"])) or '—'}</td><td><div class="row-actions"><a class="btn alt compact-btn" href="/insemination-edit?id={rec['id']}">✏️ Düzenle</a><form method="post" action="/insemination-delete" class="inline-form" onsubmit="return confirm('Bu tohumlama kaydı silinsin mi?')"><input type="hidden" name="id" value="{rec['id']}"><button class="btn red compact-btn">Sil</button></form></div></td></tr>''')
                history=''.join(hist)
                status_sort=('1' if is_pregnant_value(result) else '2' if result.strip().lower()=='bekleniyor' else '3' if result.strip().lower()=='negatif' else '4')
                latest_rows.append(f'''<tr class="data-row" data-animal="{h((str(latest['tag'] or '')+' '+str(latest['nickname'] or '')).lower())}" data-attempt="{int(latest['attempt'] or 0)}" data-insem="{h(latest['insemination_date'] or '')}" data-status="{status_sort}" data-due="{h(latest['due_date'] or '')}" data-history="{len(records)}"><td><a class="taglink" href="/animal?id={animal_id}">{h(latest['tag'])}</a><div class="mut">{h(latest['nickname'])}</div></td><td>{latest['attempt']}. Deneme</td><td>{h(fmt_date(latest['insemination_date']))}<div class="mut" style="margin-top:4px">🐂 {h(latest['bull_tag']) or 'Küpe yok'}{(' · '+h(latest['bull_name'])) if latest['bull_name'] else ''}<br>👤 {h(latest['inseminator']) or 'Tohumlayan belirtilmedi'}</div></td><td>{badge}</td><td>{h(fmt_date(latest['due_date'])) or '—'}</td><td><details><summary>Geçmiş ({len(records)})</summary><div style="overflow:auto"><table class="insem-history"><tr><th>Deneme</th><th>Tarih</th><th>Sonuç</th><th>Tahmini Doğum</th><th>İşlem</th></tr>{history}</table></div></details></td></tr>''')
            table_rows=''.join(latest_rows)
            estrus_info=''
            if estrus_context:
                try:
                    ed=date.fromisoformat(estrus_context['estrus_date']); es=ed+timedelta(days=18); ee=ed+timedelta(days=24)
                    estrus_info=f'<div class="flash">🌸 Bu hayvan Kızgınlık Takibi ekranından gönderildi. Beklenen pencere: <b>{fmt_date(es.isoformat())} – {fmt_date(ee.isoformat())}</b>. Tohumlama gerçekleştiğinde tarihi seçip kaydedin.</div>'
                except Exception:
                    estrus_info='<div class="flash">🌸 Bu hayvan Kızgınlık Takibi ekranından gönderildi.</div>'
            body=f'''<div class="insem-head"><div><h1>🐄 Üreme Takip Merkezi</h1><div class="mut">Tohumlama kayıtlarını hayvan bazında yönetin, gebelik sonucunu güncelleyin.</div></div></div>
            {estrus_info}
            <div class="grid insem-stats"><div class="card stat metric blue">Kontrol Bekleyen<b>{waiting}</b><small>Son kaydı sonuç bekleyen</small></div><div class="card stat metric green">Gebe<b>{pregnant}</b><small>Son sonucu pozitif olan</small></div><div class="card stat metric orange">3. Denemede<b>{third_attempt}</b><small>Yakın takip gereken</small></div><div class="card stat metric purple">Bu Ay Tohumlanan<b>{this_month}</b><small>{date.today().strftime('%m/%Y')}</small></div></div>
            <div class="card"><h2>Yeni Tohumlama</h2><form id="inseminationForm" method="post" action="/inseminations" class="form"><input type="hidden" name="estrus_id" value="{h(estrus_id if estrus_context else '')}"><label>Dişi Hayvan<div class="animal-picker"><input id="inseminationAnimalSearch" value="{h(selected_label)}" placeholder="Küpe veya takma ad yazın..." autocomplete="off" inputmode="search" required><div id="inseminationAnimalSuggestions" class="animal-suggestions" role="listbox" aria-label="Eşleşen dişi hayvanlar"></div></div><datalist id="inseminationAnimalOptions">{picker_options}</datalist><input type="hidden" id="inseminationAnimalId" name="animal_id" value="{h(aid if selected else '')}"><div class="animal-picker-note">Küpe veya takma ad yazın; eşleşen hayvanlar anında aşağıda görünür.</div></label><label>Deneme<div id="attemptPreview" class="attempt-preview">{(str(next_attempts[selected['id']])+'. Deneme') if selected else 'Hayvan seçildiğinde otomatik belirlenecek'}</div></label><label>Tohumlama Tarihi<input id="inseminationDate" type="date" name="insemination_date" required max="{date.today().isoformat()}"><div id="futureWarning" class="future-warning">Gelecek tarihli tohumlama kaydı girilemez.</div></label><label>Baba Küpe / Boğa No<input name="bull_tag" maxlength="80" placeholder="Örn. TR-BOGA-123"></label><label>Boğa Adı<input name="bull_name" maxlength="120" placeholder="Kullanılan boğanın adı"></label><label>Tohumlayan<input name="inseminator" maxlength="120" placeholder="Veteriner / teknisyen / kişi"></label><label>İlk Durum<div class="attempt-preview">Kontrol Bekliyor</div><input type="hidden" name="pregnancy_result" value="Bekleniyor"></label><div class="full"><button class="btn">💾 Tohumlamayı Kaydet</button></div></form></div>
            <div class="card" style="margin-top:14px"><h2>Hayvan Bazında Tohumlama Geçmişi</h2><p class="mut">Her hayvan tek satırda gösterilir. “Geçmiş” bağlantısından tüm denemeleri açabilir ve kayıtları düzenleyebilirsiniz.</p><div class="insem-search" style="margin:12px 0 14px"><input id="inseminationLiveSearch" type="search" placeholder="Kayıtlarda küpe veya takma ad ara..." autocomplete="off"><button type="button" class="btn alt live-clear" onclick="document.getElementById('inseminationLiveSearch').value='';document.getElementById('inseminationLiveSearch').dispatchEvent(new Event('input'))">Temizle</button></div><div id="insemEmpty" class="insem-empty">Eşleşen kayıt bulunamadı.</div><div style="overflow:auto"><table id="inseminationLiveTable" class="insem-table sortable-insem"><thead><tr><th><button type="button" class="sort-head" data-sort="animal">Hayvan <span>↕</span></button></th><th><button type="button" class="sort-head" data-sort="attempt">Son Deneme <span>↕</span></button></th><th><button type="button" class="sort-head active" data-sort="insem" data-dir="desc">Son Tohumlama <span>↓</span></button></th><th><button type="button" class="sort-head" data-sort="status">Durum <span>↕</span></button></th><th><button type="button" class="sort-head" data-sort="due">Tahmini Doğum <span>↕</span></button></th><th><button type="button" class="sort-head" data-sort="history">Geçmiş / İşlem <span>↕</span></button></th></tr></thead><tbody>{table_rows}</tbody></table></div></div>
            <script>
            document.addEventListener('DOMContentLoaded',function(){{
              liveTableFilter('inseminationLiveSearch','inseminationLiveTable','insemEmpty');
              var sortTable=document.getElementById('inseminationLiveTable'),sortBody=sortTable.querySelector('tbody');
              function sortInsem(key,dir){{
                var rows=Array.from(sortBody.querySelectorAll('tr.data-row'));
                rows.sort(function(a,b){{
                  var av=a.dataset[key]||'',bv=b.dataset[key]||'';
                  if(key==='attempt'||key==='status'||key==='history'){{av=parseInt(av||'0');bv=parseInt(bv||'0');return dir==='asc'?av-bv:bv-av;}}
                  var cmp=av.localeCompare(bv,'tr',{{numeric:true,sensitivity:'base'}});
                  return dir==='asc'?cmp:-cmp;
                }});
                rows.forEach(function(r){{sortBody.appendChild(r);}});
              }}
              document.querySelectorAll('.sort-head').forEach(function(btn){{
                btn.addEventListener('click',function(){{
                  var key=btn.dataset.sort,dir=(btn.dataset.dir==='asc'?'desc':'asc');
                  if(!btn.dataset.dir && (key==='insem'||key==='due'||key==='attempt'||key==='history'))dir='desc';
                  document.querySelectorAll('.sort-head').forEach(function(x){{x.classList.remove('active');x.removeAttribute('data-dir');x.querySelector('span').textContent='↕';}});
                  btn.classList.add('active');btn.dataset.dir=dir;btn.querySelector('span').textContent=dir==='asc'?'↑':'↓';
                  sortInsem(key,dir);
                }});
              }});
              sortInsem('insem','desc');
              var input=document.getElementById('inseminationAnimalSearch'),hidden=document.getElementById('inseminationAnimalId'),preview=document.getElementById('attemptPreview'),form=document.getElementById('inseminationForm'),dt=document.getElementById('inseminationDate'),warn=document.getElementById('futureWarning'),suggestions=document.getElementById('inseminationAnimalSuggestions');
              var animals={picker_data};
              function norm(v){{return (v||'').toLocaleLowerCase('tr-TR').trim();}}
              function chooseAnimal(a){{input.value=a.tag+(a.nickname?' · '+a.nickname:'');hidden.value=String(a.id);preview.textContent=(parseInt(a.attempt)>3)?'3 deneme sınırına ulaşıldı':(a.attempt+'. Deneme');suggestions.classList.remove('open');suggestions.innerHTML='';}}
              function exactAnimal(){{var v=norm(input.value),found=null;animals.forEach(function(a){{var label=norm(a.tag+(a.nickname?' · '+a.nickname:''));if(v===label||v===norm(a.tag))found=a;}});return found;}}
              function renderSuggestions(){{var q=norm(input.value);hidden.value='';preview.textContent='Hayvan seçildiğinde otomatik belirlenecek';if(!q){{suggestions.classList.remove('open');suggestions.innerHTML='';return;}}var hits=animals.filter(function(a){{return norm(a.tag).indexOf(q)!==-1||norm(a.nickname).indexOf(q)!==-1||norm(a.tag+' '+a.nickname).indexOf(q)!==-1;}}).slice(0,8);if(!hits.length){{suggestions.innerHTML='<div class="animal-suggestion"><span>Eşleşen aktif dişi hayvan bulunamadı.</span></div>';suggestions.classList.add('open');return;}}suggestions.innerHTML='';hits.forEach(function(a){{var b=document.createElement('button');b.type='button';b.className='animal-suggestion';b.innerHTML='<b>'+a.tag+'</b><span>'+(a.nickname||'Takma ad yok')+' · '+a.attempt+'. deneme</span>';b.addEventListener('click',function(){{chooseAnimal(a);}});suggestions.appendChild(b);}});suggestions.classList.add('open');var exact=exactAnimal();if(exact){{hidden.value=String(exact.id);preview.textContent=(parseInt(exact.attempt)>3)?'3 deneme sınırına ulaşıldı':(exact.attempt+'. Deneme');}}}}
              function syncAnimal(){{var exact=exactAnimal();if(exact){{hidden.value=String(exact.id);preview.textContent=(parseInt(exact.attempt)>3)?'3 deneme sınırına ulaşıldı':(exact.attempt+'. Deneme');}}}}
              input.addEventListener('input',renderSuggestions);input.addEventListener('focus',renderSuggestions);input.addEventListener('change',syncAnimal);document.addEventListener('click',function(e){{if(!e.target.closest('.animal-picker'))suggestions.classList.remove('open');}});syncAnimal();
              dt.addEventListener('change',function(){{var bad=dt.value&&dt.value>'{date.today().isoformat()}';warn.style.display=bad?'block':'none';}});
              form.addEventListener('submit',function(e){{syncAnimal();if(!hidden.value){{e.preventDefault();alert('Lütfen listeden geçerli bir dişi hayvan seçin.');return;}}if(dt.value>'{date.today().isoformat()}'){{e.preventDefault();warn.style.display='block';alert('Gelecek tarihli tohumlama kaydı girilemez.');}}}});
            }});
            </script>'''
            return self.send_html(page('Üreme Takip Merkezi',body,'/inseminations',u,msg))
        if path=='/insemination-edit':
            iid=q.get('id',[''])[0]
            with db() as c:rec=c.execute('''select i.*,a.tag,a.nickname from inseminations i join animals a on a.id=i.animal_id where i.id=?''',(iid,)).fetchone()
            if not rec:return self.redirect('/inseminations','Tohumlama kaydı bulunamadı.')
            result=str(rec['pregnancy_result'] or 'Bekleniyor')
            body=f'''<div class="actions"><a class="btn alt" href="/inseminations">← Tohumlamalara Dön</a></div><h1>Tohumlama Kaydını Düzenle</h1><div class="card"><form method="post" action="/insemination-edit" class="form"><input type="hidden" name="id" value="{rec['id']}"><label>Hayvan<div class="attempt-preview">{h(rec['tag'])} · {h(rec['nickname'])}</div></label><label>Deneme<div class="attempt-preview">{rec['attempt']}. Deneme</div></label><label>Tohumlama Tarihi<input type="date" name="insemination_date" required max="{date.today().isoformat()}" value="{h(rec['insemination_date'])}"></label><label>Baba Küpe / Boğa No<input name="bull_tag" maxlength="80" value="{h(rec['bull_tag'])}" placeholder="Örn. TR-BOGA-123"></label><label>Boğa Adı<input name="bull_name" maxlength="120" value="{h(rec['bull_name'])}" placeholder="Kullanılan boğanın adı"></label><label>Tohumlayan<input name="inseminator" maxlength="120" value="{h(rec['inseminator'])}" placeholder="Veteriner / teknisyen / kişi"></label><label>Gebelik Sonucu<select name="pregnancy_result"><option value="Bekleniyor" {'selected' if result=='Bekleniyor' else ''}>Kontrol Bekliyor</option><option value="Pozitif" {'selected' if is_pregnant_value(result) else ''}>Gebe</option><option value="Negatif" {'selected' if result=='Negatif' else ''}>Gebe Değil</option><option value="Belirsiz" {'selected' if result=='Belirsiz' else ''}>Belirsiz</option></select></label><div class="full"><button class="btn">Değişiklikleri Kaydet</button> <a class="btn alt" href="/inseminations">İptal</a></div></form></div>'''
            return self.send_html(page('Tohumlama Düzenle',body,'/inseminations',u,msg))
        if path=='/health-edit':
            hid=q.get('id',[''])[0]
            with db() as c:
                r=c.execute("select h.*,a.tag as animal_tag,ca.tag as calf_tag from health h left join animals a on a.id=h.animal_id left join calves ca on ca.id=h.calf_id where h.id=?",(hid,)).fetchone()
            if not r:return self.redirect('/health','Sağlık kaydı bulunamadı.')
            subject=r['animal_tag'] or r['calf_tag'] or '-'
            body=f'''<h1>Sağlık Kaydını Düzenle</h1><div class="card"><form method="post" action="/health-edit" class="form">
            <input type="hidden" name="id" value="{r["id"]}">
            <label>Hayvan / Buzağı<input value="{h(subject)}" disabled></label>
            <label>Tür<select name="kind"><option {"selected" if r["kind"]=="Aşı" else ""}>Aşı</option><option {"selected" if r["kind"]=="İlaç" else ""}>İlaç</option><option {"selected" if r["kind"]=="Muayene" else ""}>Muayene</option></select></label>
            <label>Ürün/İşlem<input name="product" value="{h(r["product"])}" required></label>
            <label>Uygulama Tarihi<input type="date" name="applied_date" value="{h(r["applied_date"])}" required></label>
            <label>Sonraki Tarih<input type="date" name="next_date" value="{h(r["next_date"])}"></label>
            <label>Maliyet<input type="number" step="0.01" min="0" name="cost" value="{float(r["cost"] or 0)}"></label>
            <label class="full">Not<textarea name="notes">{h(r["notes"])}</textarea></label>
            <div class="full actions"><button class="btn">Değişiklikleri Kaydet</button><a class="btn alt" href="/health">İptal</a></div>
            </form></div>'''
            return self.send_html(page('Sağlık Kaydı Düzenle',body,'/health',u,msg))
        if path=='/health':
            with db() as c:
                active_animals=c.execute("select id,tag,nickname,gender,paddock_id from animals where coalesce(status,'Aktif')='Aktif' order by tag").fetchall()
                active_calves=c.execute("select id,tag,gender,paddock_id from calves where promoted_animal_id is null order by tag").fetchall()
                paddocks=c.execute("select id,name,code from paddocks where active=1 order by name").fetchall()
                rows=c.execute("select h.*,a.tag as animal_tag,a.nickname,c.tag as calf_tag from health h left join animals a on a.id=h.animal_id left join calves c on c.id=h.calf_id order by h.applied_date desc,h.id desc").fetchall()
                legacy_plans=c.execute("select h.*,a.tag as animal_tag,a.nickname,c.tag as calf_tag from health h left join animals a on a.id=h.animal_id left join calves c on c.id=h.calf_id where coalesce(h.next_date,'')<>'' order by h.next_date,h.id").fetchall()
                task_rows=c.execute("""select t.*,hc.kind,hc.product,hc.scope_type,hc.paddock_id,p.name as paddock_name,a.tag as animal_tag,a.nickname,c.tag as calf_tag
                    from health_tasks t join health_courses hc on hc.id=t.course_id left join paddocks p on p.id=hc.paddock_id
                    left join animals a on a.id=t.animal_id left join calves c on c.id=t.calf_id
                    where t.status='Bekliyor' order by t.planned_date,t.course_id,t.dose_no,t.day_no,t.application_no,t.id""").fetchall()
            subject_items=[]
            for a in active_animals:
                label=str(a['tag'])+' · '+(str(a['nickname'] or '').strip() or str(a['gender'] or 'Hayvan'))
                subject_items.append(('A:'+str(a['id']),label))
            for c in active_calves: subject_items.append(('C:'+str(c['id']),str(c['tag'])+' · Buzağı · '+str(c['gender'] or '')))
            subject_items.sort(key=lambda x:x[1].casefold())
            subject_json=json.dumps([{'key':k,'label':v} for k,v in subject_items],ensure_ascii=False)
            paddock_options=''.join(f'<option value="{r["id"]}">{h(r["name"])}{(" · "+h(r["code"])) if r["code"] else ""}</option>' for r in paddocks)
            health_groups={}
            for r in rows:
                key=('A',r['animal_id']) if r['animal_id'] else ('C',r['calf_id']); tag=r['animal_tag'] or r['calf_tag'] or '-'; nickname=(r['nickname'] or '').strip() if r['animal_id'] else 'Buzağı'
                health_groups.setdefault(key,{'tag':tag,'nickname':nickname,'rows':[]})['rows'].append(r)
            group_cards=[]
            for g in health_groups.values():
                inner=''.join(f'<tr><td>{fmt_date(r["applied_date"])}</td><td>{h(r["kind"])}</td><td>{h(r["product"])}</td><td>{fmt_date(r["next_date"])}</td><td>{money(r["cost"])}</td><td><a class="btn alt" href="/health-edit?id={r["id"]}">Düzenle</a> <form method="post" action="/health-delete" class="inline-form" onsubmit="return confirm(\'Bu sağlık kaydı silinsin mi?\')"><input type="hidden" name="id" value="{r["id"]}"><button class="btn red">Sil</button></form></td></tr>' for r in g['rows'])
                latest=g['rows'][0]['applied_date'] if g['rows'] else ''
                group_cards.append(f'''<details class="health-group"><summary><span><b>🐄 {h(g['tag'])}</b> · {h(g['nickname'])}</span><span class="mut">{len(g['rows'])} kayıt · Son işlem {fmt_date(latest)} ▾</span></summary><div class="tablewrap"><table><tr><th>Tarih</th><th>Tür</th><th>Ürün/İşlem</th><th>Sonraki</th><th>Maliyet</th><th>İşlem</th></tr>{inner}</table></div></details>''')
            grouped_health_html=''.join(group_cards) or '<p class="mut">Henüz sağlık kaydı yok.</p>'
            def due_badge(due):
                try:days=(date.fromisoformat(due)-date.today()).days
                except:days=9999
                if days<0:return '🔴 '+str(abs(days))+' gün gecikti'
                if days==0:return '🟠 Bugün'
                if days<=3:return '🟠 '+str(days)+' gün kaldı'
                return '🟡 '+str(days)+' gün kaldı'
            task_groups={}
            for r in task_rows:
                key=('P',r['course_id'],r['planned_date'],r['dose_no'],r['day_no'],r['application_no']) if r['scope_type']=='paddock' else ('S',r['id'])
                task_groups.setdefault(key,[]).append(r)
            task_cards=[]
            for key,items in task_groups.items():
                r=items[0];badge=due_badge(r['planned_date'])
                detail=(f'{int(r["dose_no"] or 1)} / {int(r["dose_total"] or 1)}. doz' if r['kind']=='Aşı' else f'{int(r["day_no"] or 1)} / {int(r["day_total"] or 1)}. gün · {int(r["application_no"] or 1)} / {int(r["applications_per_day"] or 1)} uygulama')
                if r['scope_type']=='paddock':
                    title='🏠 '+str(r['paddock_name'] or 'Padok');meta=f'{len(items)} hayvan · {detail}'
                    action=f'''<form method="post" action="/health/task-batch-done"><input type="hidden" name="course_id" value="{r['course_id']}"><input type="hidden" name="planned_date" value="{h(r['planned_date'])}"><input type="hidden" name="dose_no" value="{int(r['dose_no'] or 1)}"><input type="hidden" name="day_no" value="{int(r['day_no'] or 1)}"><input type="hidden" name="application_no" value="{int(r['application_no'] or 1)}"><button class="btn">✅ {len(items)} Hayvan Yapıldı</button></form>'''
                else:
                    title='🐄 '+str(r['animal_tag'] or r['calf_tag'] or '-');meta=detail
                    action=f'''<form method="post" action="/health/task-done"><input type="hidden" name="task_id" value="{r['id']}"><button class="btn">✅ Yapıldı</button></form>'''
                task_cards.append(f'''<div class="health-plan-card"><div class="health-plan-main"><div><b>{h(title)}</b><div class="health-plan-product">{h(r['product'])}</div><div class="mut">{h(meta)} · {fmt_date(r['planned_date'])}</div></div><span class="health-due">{badge}</span></div><div class="health-plan-action">{action}</div></div>''')
            for r in legacy_plans:
                tag=r['animal_tag'] or r['calf_tag'] or '-';badge=due_badge(r['next_date'])
                action=(f'''<form method="post" action="/health/second-dose-done"><input type="hidden" name="source_id" value="{r['id']}"><input type="hidden" name="return_to" value="/health"><button class="btn">✅ 2. Doz Yapıldı</button></form>''' if str(r['kind'] or '')=='Aşı' and 'IKINCI_DOZ_PLAN' in str(r['notes'] or '') else f'''<form method="post" action="/health/plan-done"><input type="hidden" name="source_id" value="{r['id']}"><input type="hidden" name="return_to" value="/health"><button class="btn">✅ Yapıldı</button></form>''')
                task_cards.append(f'''<div class="health-plan-card"><div class="health-plan-main"><div><b>🐄 {h(tag)}</b><div class="health-plan-product">{h(r['product'])}</div><div class="mut">Eski plan · {fmt_date(r['next_date'])}</div></div><span class="health-due">{badge}</span></div><div class="health-plan-action">{action}</div></div>''')
            planned_html=''.join(task_cards) or '<p class="mut">Planlanmış sağlık işlemi yok.</p>'
            body=f'''<style>.health-mode-box{{background:#f7fbf8;border:1px solid #d7eadc;border-radius:16px;padding:14px 16px}}.health-plan-list{{display:grid;gap:10px}}.health-plan-card{{border:1px solid #dfe9e2;border-radius:15px;padding:14px;background:#fff}}.health-plan-main{{display:flex;justify-content:space-between;gap:12px;align-items:flex-start}}.health-plan-product{{font-weight:800;margin:5px 0}}.health-due{{white-space:nowrap;font-weight:800}}.health-plan-action{{margin-top:12px}}.health-plan-action form{{margin:0}}@media(max-width:700px){{.health-plan-main{{display:block}}.health-due{{display:inline-block;margin-top:8px}}.health-plan-action .btn{{width:100%}}.health-group summary{{align-items:flex-start;flex-direction:column}}}}</style><h1>Sağlık</h1>
            <div class="card"><form method="post" class="form" id="healthForm"><label>Uygulama Kapsamı<select name="scope_type" id="healthScope"><option value="single">Tek Hayvan / Buzağı</option><option value="paddock">Padok Bazında Aşılama</option></select></label><label>Tür<select name="kind" id="healthKind"><option>Aşı</option><option>İlaç</option><option>Muayene</option></select></label>
            <div class="full" id="singleSubjectBox"><label>Hayvan / Buzağı</label><div style="position:relative"><input type="search" id="healthSubjectSearch" placeholder="Küpe veya takma ad yazın…" autocomplete="off"><input type="hidden" name="subject_key" id="healthSubjectKey"><div id="healthSubjectResults" style="display:none;position:absolute;left:0;right:0;top:100%;z-index:30;background:#fff;border:1px solid #d7e4da;border-radius:12px;max-height:280px;overflow:auto;box-shadow:0 12px 28px #173b2822"></div></div><div class="mut">Aktif hayvanlar ve buzağılar listelenir.</div></div>
            <div class="full" id="paddockSubjectBox" style="display:none"><label>Padok<select name="paddock_id" id="healthPaddock"><option value="">Padok seçin…</option>{paddock_options}</select></label><div class="mut">Plan oluşturulduğu anda padoktaki aktif hayvan listesi sabitlenir. Sonradan padoka giren hayvanlar bu plana eklenmez.</div></div>
            <label>Ürün/İşlem<input name="product" required placeholder="Örn. Şap aşısı / antibiyotik"></label><label>Başlangıç / Uygulama Tarihi<input type="date" name="applied_date" id="healthAppliedDate" required value="{date.today().isoformat()}"></label>
            <div class="full health-mode-box" id="vaccinePlanBox"><div class="form"><label>Kaç Doz?<input type="number" name="dose_count" id="doseCount" min="1" max="10" value="1"></label><label>Dozlar Arası Gün<input type="number" name="dose_interval_days" id="doseInterval" min="1" max="365" value="15"></label></div><div class="mut">Örn. 3 doz / 15 gün arayla. Tüm doz tarihleri otomatik planlanır.</div></div>
            <div class="full health-mode-box" id="medicinePlanBox" style="display:none"><div class="form"><label>Tedavi Kaç Gün?<input type="number" name="treatment_days" min="1" max="60" value="5"></label><label>Günde Kaç Uygulama?<input type="number" name="times_per_day" min="1" max="6" value="1"></label></div><div class="mut">Her uygulama ayrı ayrı “Yapıldı” işaretlenir; geciken uygulamalar otomatik görünür.</div></div>
            <label id="nextDateLabel" style="display:none">Sonraki Tarih<input type="date" name="next_date" id="healthNextDate"></label><label>Uygulama Başı / Hayvan Başı Maliyet<input type="number" step="0.01" min="0" name="cost" value="0"></label><label class="full">Not<textarea name="notes"></textarea></label><div class="full"><button class="btn" id="healthSubmit">💾 Planı Oluştur</button></div></form></div>
            <div class="card" style="margin-top:14px"><h2>💉 Planlanan Aşı / Tedavi İşlemleri</h2><div class="health-plan-list">{planned_html}</div></div><div class="card" style="margin-top:14px"><h2>🩺 Hayvan Bazlı Sağlık Dosyaları</h2><p class="mut">Tamamlanan tedavi, aşı ve muayeneler hayvanın gerçek sağlık geçmişine işlenir.</p>{grouped_health_html}</div>
            <script>const healthSubjects={subject_json};const hs=document.getElementById('healthSubjectSearch'),hk=document.getElementById('healthSubjectKey'),hr=document.getElementById('healthSubjectResults');function renderHealthSubjects(){{const q=(hs.value||'').toLocaleLowerCase('tr-TR').trim();const found=healthSubjects.filter(x=>!q||x.label.toLocaleLowerCase('tr-TR').includes(q)).slice(0,40);hr.innerHTML=found.map((x,i)=>'<button type="button" data-index="'+i+'" style="display:block;width:100%;text-align:left;border:0;border-bottom:1px solid #eef3ef;background:#fff;padding:12px 14px;font:inherit;cursor:pointer">'+x.label+'</button>').join('');hr.style.display=found.length?'block':'none';hr.querySelectorAll('button').forEach(b=>b.onclick=function(){{const x=found[parseInt(this.dataset.index)];hk.value=x.key;hs.value=x.label;hr.style.display='none';}});}}hs.addEventListener('input',function(){{hk.value='';renderHealthSubjects();}});hs.addEventListener('focus',renderHealthSubjects);document.addEventListener('click',e=>{{if(!hr.contains(e.target)&&e.target!==hs)hr.style.display='none';}});const scope=document.getElementById('healthScope'),kind=document.getElementById('healthKind'),single=document.getElementById('singleSubjectBox'),paddock=document.getElementById('paddockSubjectBox'),vbox=document.getElementById('vaccinePlanBox'),mbox=document.getElementById('medicinePlanBox'),nlabel=document.getElementById('nextDateLabel'),submit=document.getElementById('healthSubmit'),doseCount=document.getElementById('doseCount'),doseInterval=document.getElementById('doseInterval');function syncHealthForm(){{if(scope.value==='paddock'&&kind.value!=='Aşı')kind.value='Aşı';single.style.display=scope.value==='single'?'block':'none';paddock.style.display=scope.value==='paddock'?'block':'none';vbox.style.display=kind.value==='Aşı'?'block':'none';mbox.style.display=kind.value==='İlaç'?'block':'none';nlabel.style.display=kind.value==='Muayene'?'block':'none';doseInterval.closest('label').style.display=(parseInt(doseCount.value||1)>1)?'block':'none';submit.textContent=(kind.value==='Muayene')?'💾 Sağlık Kaydını Kaydet':'💾 Planı Oluştur';}}scope.addEventListener('change',syncHealthForm);kind.addEventListener('change',syncHealthForm);doseCount.addEventListener('input',syncHealthForm);syncHealthForm();document.getElementById('healthForm').addEventListener('submit',function(e){{if(scope.value==='single'&&!hk.value){{e.preventDefault();alert('Lütfen listeden bir hayvan veya buzağı seçin.');hs.focus();return;}}if(scope.value==='paddock'&&!document.getElementById('healthPaddock').value){{e.preventDefault();alert('Lütfen bir padok seçin.');}}}});</script>'''
            return self.send_html(page('Sağlık',body,'/health',u,msg))
        if path=='/finance/edit':
            record_id=int(q.get('id',['0'])[0])
            with db() as c:
                r=c.execute('select f.*,a.tag,a.nickname from finance f left join animals a on a.id=f.animal_id where f.id=?',(record_id,)).fetchone()
                animals=c.execute('select id,tag,nickname,status from animals order by tag').fetchall()
                linked=c.execute('''select l.*,fc.name feed_name,st.tx_date stock_date,st.notes stock_notes from feed_finance_links l join feed_catalog fc on fc.id=l.feed_id left join feed_stock_transactions st on st.id=l.stock_tx_id where l.finance_id=?''',(record_id,)).fetchone()
            if not r:return self.redirect('/finance','Finans kaydı bulunamadı.')
            animal_options='<option value="">Hayvan seçmeden kaydet</option>'+''.join(
                '<option value="{0}" {1}>{2} · {3} · {4}</option>'.format(
                    a["id"],'selected' if r["animal_id"]==a["id"] else '',h(a["tag"]),h(a["nickname"]),h(a["status"])
                ) for a in animals
            )
            categories=['Süt Satışı','Hayvan Satışı','Kesim Geliri','Buzağı Satışı','Destekleme','Yem','Veteriner','İlaç','Aşı','Saman','Elektrik','Yakıt','İşçilik','Hayvan Alımı','Diğer']
            category_options=''.join('<option {0}>{1}</option>'.format('selected' if r["category"]==x else '',h(x)) for x in categories)
            linked_html=''
            if linked:
                linked_html=f'''<div class="linked-feed-box"><h3>🔗 Bağlı Yem Stok Hareketi</h3><div class="mut">Bu finans kaydı <b>{h(linked['feed_name'])}</b> stok girişiyle bağlıdır. Miktar veya birim fiyatı burada değiştirirseniz stok kaydı da birlikte güncellenir.</div><div class="linked-feed-grid"><label>Yem<input value="{h(linked['feed_name'])}" disabled></label><label>Miktar (kg)<input type="number" step="0.01" min="0.01" name="linked_feed_qty" id="linkedFeedQty" value="{linked['quantity_kg']}" required></label><label>Birim Fiyat (₺/kg)<input type="number" step="0.0001" min="0.0001" name="linked_feed_unit" id="linkedFeedUnit" value="{linked['unit_price']}" required></label></div><div class="linked-total"><span>Finansa kaydedilecek yeni toplam</span><b id="linkedFeedTotal">{money(float(linked['quantity_kg'])*float(linked['unit_price']))}</b></div><input type="hidden" name="linked_feed" value="yes"></div>'''
            body=f'''<h1>Finans Kaydını Düzenle</h1><div class="card"><form method="post" action="/finance/edit" class="form">
            <input type="hidden" name="id" value="{r["id"]}">
            <label>Tarih<input type="date" name="tx_date" value="{h(r["tx_date"])}" required></label>
            <label>Tür<select name="tx_type"><option {"selected" if r["tx_type"]=="Gelir" else ""}>Gelir</option><option {"selected" if r["tx_type"]=="Gider" else ""}>Gider</option></select></label>
            <label>Kategori<select name="category" id="financeCategory">{category_options}</select></label>
            <label>Tutar<input type="number" step="0.01" min="0" name="amount" value="{r["amount"]}" required></label>
            <label>Ödeme<select name="payment_method"><option {"selected" if r["payment_method"]=="Nakit" else ""}>Nakit</option><option {"selected" if r["payment_method"]=="Banka" else ""}>Banka</option><option {"selected" if r["payment_method"]=="Kredi Kartı" else ""}>Kredi Kartı</option><option {"selected" if r["payment_method"]=="Vadeli" else ""}>Vadeli</option></select></label>
            <label>İlgili Hayvan<select name="animal_id" id="financeAnimal">{animal_options}</select></label>
            <label class="full">Açıklama<input name="description" value="{h(r["description"])}"></label>
            {linked_html}
            <div class="full" id="statusWarning" style="display:none;padding:12px;border-radius:10px;background:#fff3cd;color:#664d03"><b>Uyarı:</b> Satış veya kesim seçilirse hayvan aktif sürüden çıkarılır. Kategori değiştirilirse durum yeniden hesaplanır.</div>
            <div class="full"><button class="btn">Değişiklikleri Kaydet</button> <a class="btn alt" href="/finance">İptal</a></div>
            </form></div>'''
            return self.send_html(page('Finans Düzenle',body,path,u,msg))
        if path=='/finance':
            start=q.get('start',[date.today().replace(day=1).isoformat()])[0]; end=q.get('end',[date.today().isoformat()])[0]; typ=q.get('type',[''])[0]; category=q.get('category',[''])[0]
            sql='''select f.*,a.tag,
            case when f.category='Süt Satışı'
                 then coalesce((select group_concat(a2.tag, ', ') from finance_animals fa join animals a2 on a2.id=fa.animal_id where fa.finance_id=f.id and fa.relation_type='Süt Satışı'),a.tag)
                 else a.tag end as related_tags
            from finance f left join animals a on a.id=f.animal_id where tx_date between ? and ?'''; args=[start,end]
            if typ: sql+=' and tx_type=?'; args.append(typ)
            if category: sql+=' and category=?'; args.append(category)
            sql+=' order by tx_date desc,id desc'
            with db() as c:
                animals=c.execute("select id,tag,nickname,gender from animals where coalesce(status,'Aktif')='Aktif' order by tag").fetchall()
                purchase_linked_ids={int(r['animal_id']) for r in c.execute("select distinct animal_id from finance where category='Hayvan Alımı' and animal_id is not null").fetchall()}
                milk_females=[a for a in animals if str(a['gender'] or '')=='Dişi']
                categories=c.execute("select distinct category from finance where coalesce(category,'')<>'' order by category").fetchall()
                finance_feeds=c.execute("select id,name from feed_catalog where active=1 order by name").fetchall()
                rows=c.execute(sql,args).fetchall()
                inc=sum(float(r['amount'] or 0) for r in rows if r['tx_type']=='Gelir'); exp=sum(float(r['amount'] or 0) for r in rows if r['tx_type']=='Gider')
            opts=''.join(f'<option value="{a["id"]}">{h(a["tag"])} - {h(a["nickname"])}</option>' for a in animals)
            bulk_cards=''.join(f'''<label class="bulk-row" data-search="{h((str(a["tag"])+" "+str(a["nickname"] or "")).lower())}"><input type="checkbox" class="bulk-check" value="{a["id"]}" onchange="syncBulkSelection()"><span class="tag">🐄 {h(a["tag"])}</span><span class="nick">{h(a["nickname"]) or "Takma ad yok"}</span></label>''' for a in animals)
            milk_cards=''.join(f'''<label class="bulk-row milk-row" data-search="{h((str(a["tag"])+" "+str(a["nickname"] or "")).lower())}"><input type="checkbox" class="milk-check" value="{a["id"]}" onchange="syncMilkSelection()"><span class="tag">🥛 {h(a["tag"])}</span><span class="nick">{h(a["nickname"]) or "Takma ad yok"}</span></label>''' for a in milk_females)
            category_opts=''.join(f'<option value="{h(r["category"])}" {"selected" if category==r["category"] else ""}>{h(r["category"])}</option>' for r in categories)
            finance_feed_opts=''.join(f'<option value="{r["id"]}">{h(r["name"])}</option>' for r in finance_feeds)
            trs=''.join(
                '<tr><td>{0}</td><td>{1}</td><td>{2}</td><td>{3}</td><td>{4}</td><td>{5}</td><td>{6}</td><td><b>{7}</b></td><td><div class="finance-actions"><a class="btn alt" href="/finance/edit?id={8}">Düzenle</a><form method="post" action="/finance/delete" onsubmit="return confirm(\'Bu finans kaydı silinsin mi?\')"><input type="hidden" name="id" value="{8}"><button class="btn danger">Sil</button></form></div></td></tr>'.format(
                    fmt_date(r["tx_date"]),h(r["tx_type"]),h(r["category"]),h(r["description"]),h(r["related_tags"]),h(r["animal_status_action"]) or "-",h(r["payment_method"]),money(r["amount"]),r["id"]
                ) for r in rows
            )
            body=f'''<h1>Finans</h1><div class="grid"><div class="card stat">Gelir<b>{money(inc)}</b></div><div class="card stat">Gider<b>{money(exp)}</b></div><div class="card stat">Net<b>{money(inc-exp)}</b></div></div><div class="finance-primary-actions"><button type="button" class="btn finance-new-btn" onclick="openFinanceDrawer()">➕ Yeni Finans Kaydı</button><span class="mut">Kayıtlar ve filtreler öncelikli görünür.</span></div><div id="financeDrawerBackdrop" class="finance-drawer-backdrop" onclick="closeFinanceDrawer(event)"></div><aside id="financeDrawer" class="finance-drawer" aria-hidden="true"><div class="finance-drawer-head"><div><span class="mut">FİNANS</span><h2 style="margin:3px 0">➕ Yeni Finans Kaydı</h2><span class="mut">Kaydı oluşturun; bitince listenize dönün.</span></div><button type="button" class="finance-drawer-close" onclick="closeFinanceDrawer()">×</button></div><div class="finance-drawer-body"><div class="card finance-entry-card"><form method="post" class="form" id="financeCreateForm">
<label>Tarih<input type="date" name="tx_date" required value="{date.today().isoformat()}"></label>
<label>Tür<select name="tx_type" id="tx"><option>Gelir</option><option>Gider</option></select></label>
<label>Kategori<select name="category" id="financeCategory"><option>Süt Satışı</option><option>Hayvan Satışı</option><option>Kesim Geliri</option><option>Buzağı Satışı</option><option>Destekleme</option><option>Yem</option><option>Veteriner</option><option>İlaç</option><option>Aşı</option><option>Saman</option><option>Elektrik</option><option>Yakıt</option><option>İşçilik</option><option>Hayvan Alımı</option><option>Diğer</option></select></label>
<label>Toplam Tutar<input type="text" inputmode="decimal" name="amount" id="financeAmount" placeholder="Örn. 200.000 veya 200000" required></label>
<label>Ödeme Yöntemi<select name="payment_method"><option>Nakit</option><option>Banka</option><option>Kredi Kartı</option><option>Vadeli</option></select></label>
<label id="singleAnimalLabel">İlgili Hayvan<select name="animal_id" id="financeAnimal"><option value="">Yok</option>{opts}</select></label><div class="full" id="financeFeedBox" style="display:none;padding:12px;background:#f0f7f2;border:1px solid #d5e7da;border-radius:11px"><div class="form"><label>Yem Kataloğu<select name="feed_id" id="financeFeed"><option value="">Yem seçin…</option>{finance_feed_opts}</select></label><label>Miktar (kg)<input type="number" step="0.01" min="0.01" name="feed_quantity_kg" id="financeFeedQty"></label><label>Alış ₺/kg<input type="number" step="0.0001" min="0" name="feed_unit_price" id="financeFeedUnit"></label><div class="full mut">Gider / Yem kaydedildiğinde seçilen yem stoğuna otomatik giriş yapılır.</div></div></div>
<input type="hidden" name="animal_ids" id="bulkAnimalIds" value="">
<div class="full bulk-animal-box" id="bulkAnimalBox"><div class="bulk-picker"><div class="bulk-picker-head"><div><h3 style="margin:0">🐄 İlgili Hayvanlar</h3><div class="mut">İlgili hayvanları seçin.</div></div><input class="bulk-search" id="bulkSearch" placeholder="Küpe veya takma ad ara…" oninput="filterBulkAnimals()"></div><div class="bulk-list" id="bulkList">{bulk_cards}</div><div class="bulk-summary"><span class="pill">Seçilen <b id="bulkCount">0</b> hayvan</span><span class="pill"><span id="bulkShareLabel">Hayvan Başı Gelir</span> <b id="bulkShare">₺0,00</b></span><button type="button" class="btn alt" onclick="clearBulkAnimals()">Seçimi Temizle</button></div><div class="bulk-selected-preview" id="bulkSelectedPreview">Henüz hayvan seçilmedi.</div></div></div><input type="hidden" name="milk_animal_ids" id="milkAnimalIds" value=""><div class="full bulk-animal-box" id="milkAnimalBox" style="display:none"><div class="bulk-picker"><div class="bulk-picker-head"><div><h3 style="margin:0">🥛 Süt Gelirine Dahil Dişi Hayvanlar</h3><div class="mut">Yalnızca aktif dişi hayvanlar gösterilir. Toplam süt geliri bölünmez; seçilen hayvanlar kayda ilişkilendirilir.</div></div><input class="bulk-search" id="milkSearch" placeholder="Dişi küpe veya takma ad ara…" oninput="filterMilkAnimals()"></div><div class="bulk-list" id="milkList">{milk_cards}</div><div class="bulk-summary"><span class="pill">Seçilen <b id="milkCount">0</b> dişi</span><span class="pill">Toplam gelir <b id="milkTotal">₺0,00</b></span><button type="button" class="btn alt" onclick="clearMilkAnimals()">Seçimi Temizle</button></div><div class="bulk-selected-preview" id="milkSelectedPreview">Henüz dişi hayvan seçilmedi.</div></div></div>
<label class="full">Açıklama<input name="description"></label>
<div class="full" id="statusWarning" style="display:none;padding:12px;border-radius:10px;background:#fff3cd;color:#664d03"><b>Uyarı:</b> Seçilen hayvanlar işlem türüne göre aktif sürüden çıkarılır; geçmiş bilgileri silinmez. Toplam tutar seçilen hayvan sayısına göre otomatik dağıtılır.</div>
<div class="full finance-savebar"><button type="submit" class="btn" id="financeSubmitBtn">💾 Finans Kaydını Kaydet</button><span class="mut" id="financeSaveHint"></span></div></form></div></div></aside><div class="card finance-filter-card" id="financeRecords" style="margin-top:14px"><div class="finance-filter-title"><div><h2>🔎 Finans Filtreleri</h2><span class="mut">Tarih, işlem türü ve kategoriye göre kayıtları daraltın.</span></div><span class="filter-count-pill">{len(rows)} kayıt</span></div><form method="get" class="finance-toolbar finance-toolbar-modern"><label><span>📅 Başlangıç</span><input type="date" name="start" value="{h(start)}"></label><label><span>📅 Bitiş</span><input type="date" name="end" value="{h(end)}"></label><label><span>↕️ Tür</span><select name="type"><option value="">Gelir + Gider</option><option {'selected' if typ=='Gelir' else ''}>Gelir</option><option {'selected' if typ=='Gider' else ''}>Gider</option></select></label><label><span>🏷️ Kategori</span><select name="category"><option value="">Tüm Kategoriler</option>{category_opts}</select></label><div class="finance-filter-actions"><button class="btn blue">🔎 Filtrele</button><a class="btn alt" href="/finance">↺ Temizle</a><a class="btn export-btn" href="/finance/export?start={urllib.parse.quote(start)}&end={urllib.parse.quote(end)}&type={urllib.parse.quote(typ)}&category={urllib.parse.quote(category)}">⬇ CSV</a></div></form><div class="finance-table-wrap"><table class="finance-table"><tr><th>Tarih</th><th>Tür</th><th>Kategori</th><th>Açıklama</th><th>Hayvan</th><th>Durum</th><th>Ödeme</th><th>Tutar</th><th>İşlem</th></tr>{trs}</table></div></div>'''
            body += f'''<script>
            function isBulkFinance(){{
              const t=document.getElementById('tx').value;
              const c=document.getElementById('financeCategory').value;
              return t==='Gelir' && (c==='Hayvan Satışı' || c==='Kesim Geliri');
            }}
            function isMilkFinance(){{return document.getElementById('tx').value==='Gelir' && document.getElementById('financeCategory').value==='Süt Satışı';}}
            function refreshFinanceFeed(){{const on=document.getElementById('tx').value==='Gider'&&document.getElementById('financeCategory').value==='Yem';const b=document.getElementById('financeFeedBox'),ff=document.getElementById('financeFeed'),fq=document.getElementById('financeFeedQty');if(b)b.style.display=on?'block':'none';if(ff)ff.required=on;if(fq)fq.required=on;}}
            function parseMoneyInput(v){{v=String(v||'').trim().replace(/\\s/g,'').replace(/₺/g,'');if(!v)return 0;if(v.includes(',')){{v=v.replace(/\\./g,'').replace(',','.');}}else{{const parts=v.split('.');if(parts.length>1&&parts.slice(1).every(x=>x.length===3))v=parts.join('');}}const n=Number(v);return Number.isFinite(n)?n:0;}}
            function formatTRY(v){{return new Intl.NumberFormat('tr-TR',{{style:'currency',currency:'TRY'}}).format(v||0);}}
            function selectedChecks(){{return Array.from(document.querySelectorAll('.bulk-check:checked'));}}
            function syncBulkSelection(){{
              const checks=selectedChecks();
              checks.forEach(x=>x.closest('.bulk-row').classList.add('selected'));
              document.querySelectorAll('.bulk-check:not(:checked)').forEach(x=>x.closest('.bulk-row').classList.remove('selected'));
              document.getElementById('bulkAnimalIds').value=checks.map(x=>x.value).join(',');
              document.getElementById('bulkCount').textContent=checks.length;
              const total=parseMoneyInput(document.getElementById('financeAmount').value);
              document.getElementById('bulkShare').textContent=formatTRY(checks.length ? total/checks.length : 0);
              const shareLabel=document.getElementById('bulkShareLabel');
              if(shareLabel) shareLabel.textContent=checks.length===1?'Hayvan Geliri':'Hayvan Başı Gelir';
              const labels=checks.slice(0,5).map(x=>x.closest('.bulk-row').querySelector('.tag').textContent.trim());
              document.getElementById('bulkSelectedPreview').textContent=checks.length ? labels.join(' · ')+(checks.length>5?' · +'+(checks.length-5)+' diğer':'') : 'Henüz hayvan seçilmedi.';
              document.getElementById('financeSaveHint').textContent=(isBulkFinance()&&checks.length)
                ? (checks.length===1 ? 'Seçilen hayvana '+formatTRY(total)+' yazılacak' : checks.length+' hayvana otomatik dağıtılacak')
                : '';
            }}
            function syncMilkSelection(){{
              const checks=Array.from(document.querySelectorAll('.milk-check:checked'));
              document.querySelectorAll('.milk-row').forEach(x=>x.classList.toggle('selected',x.querySelector('.milk-check').checked));
              document.getElementById('milkAnimalIds').value=checks.map(x=>x.value).join(',');document.getElementById('milkCount').textContent=checks.length;
              document.getElementById('milkTotal').textContent=formatTRY(parseMoneyInput(document.getElementById('financeAmount').value));
              const labels=checks.slice(0,8).map(x=>x.closest('.milk-row').querySelector('.tag').textContent.trim());
              document.getElementById('milkSelectedPreview').textContent=checks.length?labels.join(' · ')+(checks.length>8?' · +'+(checks.length-8)+' diğer':''):'Henüz dişi hayvan seçilmedi.';
              if(isMilkFinance())document.getElementById('financeSaveHint').textContent=checks.length+' dişi hayvan süt gelirine ilişkilendirilecek';
            }}
            function clearMilkAnimals(){{document.querySelectorAll('.milk-check').forEach(x=>x.checked=false);syncMilkSelection();}}
            function filterMilkAnimals(){{const q=(document.getElementById('milkSearch').value||'').toLocaleLowerCase('tr-TR').trim();document.querySelectorAll('.milk-row').forEach(row=>{{row.style.display=!q||row.dataset.search.toLocaleLowerCase('tr-TR').includes(q)?'grid':'none';}});}}
            function refreshPurchaseAnimalOptions(){{
              const category=document.getElementById('financeCategory').value;
              const select=document.getElementById('financeAnimal');
              const purchaseMode=category==='Hayvan Alımı';
              Array.from(select.options).forEach(function(opt){{
                if(!opt.value)return;
                const used=opt.dataset.purchaseLinked==='1';
                opt.hidden=purchaseMode&&used;
                opt.disabled=purchaseMode&&used;
              }});
              if(select.selectedOptions.length&&select.selectedOptions[0].disabled)select.value='';
            }}
            function refreshBulkFinance(){{
              const on=isBulkFinance(),milk=isMilkFinance();
              document.getElementById('bulkAnimalBox').style.display=on?'block':'none';document.getElementById('milkAnimalBox').style.display=milk?'block':'none';
              document.getElementById('singleAnimalLabel').style.display=(on||milk)?'none':'block';document.getElementById('financeAnimal').required=false;
              document.getElementById('statusWarning').style.display=on?'block':'none';refreshPurchaseAnimalOptions();syncBulkSelection();syncMilkSelection();
            }}
            function clearBulkAnimals(){{
              document.querySelectorAll('.bulk-check').forEach(x=>x.checked=false);
              syncBulkSelection();
            }}
            function filterBulkAnimals(){{
              const q=(document.getElementById('bulkSearch').value||'').toLocaleLowerCase('tr-TR').trim();
              document.querySelectorAll('.bulk-row').forEach(row=>{{row.style.display=!q||row.dataset.search.toLocaleLowerCase('tr-TR').includes(q)?'grid':'none';}});
            }}
            document.getElementById('tx').addEventListener('change',function(){{refreshBulkFinance();refreshFinanceFeed();}});
            document.getElementById('financeCategory').addEventListener('change',function(){{refreshBulkFinance();refreshFinanceFeed();}});
            document.getElementById('financeAmount').addEventListener('input',function(){{syncBulkSelection();syncMilkSelection();}});
            document.getElementById('financeCreateForm').addEventListener('submit',function(e){{
              syncBulkSelection();
              if(isBulkFinance()) document.getElementById('financeAnimal').required=false;
              if(isBulkFinance() && selectedChecks().length===0){{
                e.preventDefault();
                alert('Hayvan satışı veya kesim geliri için en az bir hayvan seçmelisiniz.');
                return false;
              }}
              if(isMilkFinance() && document.querySelectorAll('.milk-check:checked').length===0){{
                e.preventDefault();
                alert('Süt satışı için en az bir aktif dişi hayvan seçmelisiniz.');
                return false;
              }}
              if(this.dataset.submitting==='1'){{
                e.preventDefault();
                return false;
              }}
              this.dataset.submitting='1';
              const submitBtn=this.querySelector('button[type=submit]');
              if(submitBtn){{
                submitBtn.disabled=true;
                submitBtn.textContent='⏳ Kaydediliyor…';
              }}
            }});
            refreshBulkFinance();refreshFinanceFeed();
            function setFinanceDrawer(open){{const d=document.getElementById('financeDrawer'),b=document.getElementById('financeDrawerBackdrop');if(!d||!b)return;d.classList.toggle('open',open);b.classList.toggle('open',open);d.setAttribute('aria-hidden',open?'false':'true');document.body.style.overflow=open?'hidden':'';}}
function openFinanceDrawer(){{setFinanceDrawer(true);}}
function closeFinanceDrawer(ev){{if(ev&&ev.target!==document.getElementById('financeDrawerBackdrop'))return;setFinanceDrawer(false);}}
document.addEventListener('keydown',e=>{{if(e.key==='Escape')setFinanceDrawer(false);}});
setTimeout(()=>setFinanceDrawer(false),0);
</script>'''
            return self.send_html(page('Finans',body,'/finance',u,msg))
        if path=='/reports':
            profile=farm_profile(); farm_name=farm_display_name(profile)
            start=q.get('start',[(date.today()-timedelta(days=365)).isoformat()])[0]; end=q.get('end',[date.today().isoformat()])[0]
            animal_group=q.get('animal_group',['all'])[0];animal_status=q.get('animal_status',['Aktif'])[0]
            animal_search=q.get('animal_search',[''])[0].strip();animal_paddock=q.get('animal_paddock',[''])[0].strip()
            report_rows=animal_report_rows(animal_group,animal_status,animal_search,animal_paddock)
            report_columns=animal_report_selected_columns(q);report_column_keys=[x[0] for x in report_columns]
            report_query=urllib.parse.urlencode({'animal_group':animal_group,'animal_status':animal_status,'animal_search':animal_search,'animal_paddock':animal_paddock,'columns_mode':'custom','columns':report_column_keys},doseq=True)
            with db() as c:
                sums=c.execute('select tx_type,category,sum(amount) total,count(*) cnt from finance where tx_date between ? and ? group by tx_type,category order by tx_type, total desc',(start,end)).fetchall(); monthly=c.execute("select substr(tx_date,1,7) m, sum(case when tx_type='Gelir' then amount else 0 end) inc, sum(case when tx_type='Gider' then amount else 0 end) exp from finance where tx_date between ? and ? group by m order by m",(start,end)).fetchall();paddocks=[r['name'] for r in c.execute("select name from paddocks where active=1 order by name").fetchall()]
            inc=sum(r['total'] for r in sums if r['tx_type']=='Gelir');exp=sum(r['total'] for r in sums if r['tx_type']=='Gider'); maxv=max([max(r['inc'],r['exp']) for r in monthly] or [1])
            bars=''.join(f'<div style="flex:1;display:flex;align-items:end;gap:2px;height:170px"><div class="bar" style="height:{max(2,r["inc"]/maxv*150)}px"><i>{int(r["inc"])}</i></div><div class="bar" style="height:{max(2,r["exp"]/maxv*150)}px;background:linear-gradient(#e76d5b,#b9382b)"><i>{int(r["exp"])}</i></div><span style="position:absolute"></span><small style="position:absolute;margin-top:175px">{h(r["m"])}</small></div>' for r in monthly)
            trs=''.join(f'<tr><td>{h(r["tx_type"])}</td><td>{h(r["category"])}</td><td>{r["cnt"]}</td><td>{money(r["total"])}</td></tr>' for r in sums)
            animal_ths=''.join(f'<th>{h(label)}</th>' for _,label,_,_ in report_columns)
            animal_trs=''.join('<tr>'+''.join(f'<td>{"<b>" if key=="tag" else ""}{h(animal_report_display_value(r,key,i))}{"</b>" if key=="tag" else ""}</td>' for key,_,_,_ in report_columns)+'</tr>' for i,r in enumerate(report_rows,1)) or f'<tr><td colspan="{len(report_columns)}">Seçilen filtrelerde hayvan bulunamadı.</td></tr>'
            animal_mobile_cards=''.join(f'''<article class="animal-mobile-card"><div class="animal-mobile-head">{'<span class="animal-mobile-index">#'+str(i)+'</span>' if 'row_no' in report_column_keys else ''}<strong>{h(r['tag'])}</strong>{'<span class="animal-mobile-group">'+h(r['group'])+'</span>' if 'group' in report_column_keys else ''}</div><div class="animal-mobile-fields">{''.join(f'<div><small>{h(label)}</small><span>{h(animal_report_display_value(r,key,i))}</span></div>' for key,label,_,_ in report_columns if key not in ('row_no','tag','group'))}</div></article>''' for i,r in enumerate(report_rows,1)) or '<div class="animal-mobile-empty">Seçilen filtrelerde hayvan bulunamadı.</div>'
            column_checks=''.join((f'<label class="report-column-check locked"><input type="hidden" name="columns" value="tag"><input type="checkbox" checked disabled><span>{h(label)}</span></label>' if key=='tag' else f'<label class="report-column-check"><input type="checkbox" name="columns" value="{h(key)}" {"checked" if key in report_column_keys else ""}><span>{h(label)}</span></label>') for key,label,_,_ in ANIMAL_REPORT_COLUMNS)
            paddock_options='<option value="">Tüm Padoklar</option>'+''.join(f'<option value="{h(x)}" {"selected" if x==animal_paddock else ""}>{h(x)}</option>' for x in paddocks)
            female_count=sum(1 for r in report_rows if r['group']=='Dişi');male_count=sum(1 for r in report_rows if r['group']=='Erkek');calf_count=sum(1 for r in report_rows if r['group']=='Buzağı')
            body=f'''<style>
            .report-quick-actions{{display:flex;flex-wrap:wrap;gap:8px;margin:0 0 12px}}.animal-import-callout{{border-left:5px solid #176b3a!important;background:linear-gradient(135deg,#f4fbf6,#fff)!important}}.animal-import-callout h2{{margin:0 0 5px}}.report-column-picker{{margin-top:12px;border:1px solid #cfe0d5;border-radius:7px;background:#f8fbf9}}.report-column-picker>summary{{cursor:pointer;padding:11px 12px;font-weight:800;color:#173b28;display:flex;justify-content:space-between;gap:8px}}.report-column-content{{padding:0 12px 12px}}.report-column-grid{{display:grid;grid-template-columns:repeat(6,minmax(120px,1fr));gap:7px}}.report-column-check{{display:flex;align-items:center;gap:7px;border:1px solid #dce6df;background:#fff;padding:8px;border-radius:6px;cursor:pointer}}.report-column-check input{{width:17px;height:17px;margin:0}}.report-column-check.locked{{background:#edf4ef;color:#516359;cursor:default}}.animal-mobile-list{{display:none}}.animal-mobile-card{{border:1px solid #d6e2da;border-radius:9px;background:#fff;padding:11px;box-shadow:0 1px 3px #1232}}.animal-mobile-head{{display:flex;align-items:center;gap:8px;padding-bottom:8px;border-bottom:1px solid #e3ebe5}}.animal-mobile-head strong{{font-size:15px;color:#143d28;overflow-wrap:anywhere}}.animal-mobile-index{{font-size:11px;color:#64746a}}.animal-mobile-group{{margin-left:auto;background:#e8f4ec;color:#176b3a;border-radius:99px;padding:4px 8px;font-size:11px;font-weight:800}}.animal-mobile-fields{{display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:9px 12px;padding-top:9px}}.animal-mobile-fields div{{min-width:0}}.animal-mobile-fields small{{display:block;color:#6c7b72;font-size:10px;margin-bottom:2px}}.animal-mobile-fields span{{display:block;font-size:12px;font-weight:700;color:#263c30;overflow-wrap:anywhere}}.animal-mobile-empty{{padding:20px;text-align:center;color:#68786f;background:#fff;border:1px dashed #ccd9d0;border-radius:8px}}
            @media(max-width:700px){{.main>h1:first-of-type{{font-size:20px!important}}.report-quick-actions{{display:grid;grid-template-columns:1fr 1fr}}.report-quick-actions .btn{{text-align:center;padding:10px 7px!important}}.animal-import-callout{{padding:13px!important}}.animal-import-callout .form{{display:block!important}}.animal-import-callout label,.animal-import-callout button{{width:100%}}#animal-report .finance-toolbar-modern{{display:grid!important;grid-template-columns:1fr!important;gap:9px!important}}#animal-report .finance-filter-actions{{display:grid!important;grid-template-columns:1fr 1fr!important}}.report-column-grid{{grid-template-columns:repeat(2,minmax(0,1fr))}}.report-column-check{{padding:9px 7px;font-size:12px}}#animal-report>.grid{{grid-template-columns:repeat(2,minmax(0,1fr))!important}}#animal-report>.grid .stat{{min-height:74px!important;padding:9px!important}}.report-actions{{display:grid!important;grid-template-columns:1fr 1fr!important;gap:7px!important}}.report-actions .btn{{text-align:center;white-space:normal;padding:10px 6px!important}}.report-screen-table{{display:none!important}}.animal-mobile-list{{display:grid;gap:9px;margin-top:10px}}}}
            </style><h1>{h(farm_name)} · Raporlar</h1>
            <div class="report-quick-actions"><a class="btn blue" href="#animal-import">📥 Excel/PDF'den Hayvan Aktar</a><a class="btn" href="#animal-report">🐄 Tüm Hayvanlar Raporu</a><a class="btn alt" target="_blank" href="/reports/animals.pdf?{report_query}">📄 Hızlı PDF</a></div>
            <div class="card animal-import-callout" id="animal-import"><h2>📥 Excel / PDF'den Hayvan İçe Aktar</h2><p class="mut">Bakanlık “İşletmede Bulunan Sığır ve Manda Türü Hayvan Raporu” PDF'leri ile XLSX/CSV tabloları desteklenir. Önce güvenli önizleme açılır; mükerrer küpeler aktarılmaz.</p><form method="post" action="/animals/import-preview" enctype="multipart/form-data" class="form"><label class="full">Hayvan listesi seçin<input type="file" name="animal_file" accept=".xlsx,.csv,.pdf,application/pdf,application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,text/csv" required></label><div class="full"><button class="btn blue">📂 Dosyayı Oku ve Önizle</button></div></form><p class="mut" style="font-size:12px">Taranmış fotoğraf PDF yerine mümkünse sistemden alınmış orijinal PDF veya Excel kullanın.</p></div>
            <div class="card" id="animal-report" style="margin-top:14px"><div class="filter-title"><div><h2 style="margin:0">🐄 Tüm Hayvanlar Raporu</h2><p class="mut" style="margin:4px 0 0">Filtreleri ve görünecek sütunları seçin; aynı görünümü PDF veya Excel'e aktarın.</p></div><span class="pill">{len(report_rows)} kayıt</span></div>
              <form class="finance-toolbar-modern" style="margin-top:14px">
                <input type="hidden" name="start" value="{h(start)}"><input type="hidden" name="end" value="{h(end)}"><input type="hidden" name="columns_mode" value="custom">
                <label><span>Hayvan Grubu</span><select name="animal_group"><option value="all" {"selected" if animal_group=='all' else ""}>Tüm Hayvanlar</option><option value="female" {"selected" if animal_group=='female' else ""}>Dişi Hayvanlar</option><option value="male" {"selected" if animal_group=='male' else ""}>Erkek Hayvanlar</option><option value="calves" {"selected" if animal_group=='calves' else ""}>Buzağılar</option></select></label>
                <label><span>Durum</span><select name="animal_status"><option {"selected" if animal_status=='Aktif' else ""}>Aktif</option><option {"selected" if animal_status=='Satıldı' else ""}>Satıldı</option><option {"selected" if animal_status=='Kesildi' else ""}>Kesildi</option><option {"selected" if animal_status=='Tümü' else ""}>Tümü</option></select></label>
                <label><span>Padok</span><select name="animal_paddock">{paddock_options}</select></label>
                <label><span>Arama</span><input name="animal_search" value="{h(animal_search)}" placeholder="Küpe, ırk, anne, padok..."></label>
                <details class="report-column-picker full" open><summary><span>☑ Raporda Gösterilecek Sütunlar</span><span>{len(report_columns)} seçili</span></summary><div class="report-column-content"><div class="report-column-grid">{column_checks}</div><p class="mut" style="margin:8px 0 0">Küpe No zorunludur. Seçiminiz ekran, mobil kartlar, PDF, web önizleme ve Excel'e birlikte uygulanır.</p></div></details>
                <div class="finance-filter-actions"><button class="btn">Seçimleri Uygula</button><a class="btn alt" href="/reports#animal-report">Varsayılana Dön</a></div>
              </form>
              <div class="grid" style="margin-top:12px"><div class="card stat metric green">Listelenen Toplam<b>{len(report_rows)}</b></div><div class="card stat metric blue">Dişi<b>{female_count}</b></div><div class="card stat metric orange">Erkek<b>{male_count}</b></div><div class="card stat metric teal">Buzağı<b>{calf_count}</b></div></div>
              <div class="actions report-actions"><a class="btn blue" target="_blank" href="/reports/animals.pdf?{report_query}">📄 Temiz PDF</a><a class="btn alt web-preview-btn" target="_blank" href="/reports/animals/print?{report_query}">👁 Web Önizleme</a><a class="btn" href="/reports/animals.xlsx?{report_query}">📊 Excel'e Aktar</a><a class="btn orange" href="#animal-import">📥 Hayvan Aktar</a></div>
              <div class="tablewrap report-screen-table" style="max-height:520px;overflow:auto"><table><thead><tr>{animal_ths}</tr></thead><tbody>{animal_trs}</tbody></table></div><div class="animal-mobile-list">{animal_mobile_cards}</div>
            </div>
            <h2 style="margin-top:22px">💰 Finans Raporları</h2><div class="card"><form class="actions"><label>Başlangıç <input type="date" name="start" value="{start}"></label><label>Bitiş <input type="date" name="end" value="{end}"></label><button class="btn">Raporla</button><a class="btn blue" href="/reports/export?start={start}&end={end}">Rapor CSV</a></form></div><div class="grid" style="margin-top:14px"><div class="card stat">Toplam Gelir<b>{money(inc)}</b></div><div class="card stat">Toplam Gider<b>{money(exp)}</b></div><div class="card stat">Net Sonuç<b>{money(inc-exp)}</b></div><div class="card stat">Gider/Gelir Oranı<b>{(exp/inc*100 if inc else 0):.1f}%</b></div></div><div class="two" style="margin-top:14px"><div class="card"><h2>Aylık Gelir / Gider</h2><p class="mut">Yeşil: gelir · Kırmızı: gider</p><div class="chart">{bars or '<p>Kayıt yok</p>'}</div></div><div class="card"><h2>Kategori Özeti</h2><table><tr><th>Tür</th><th>Kategori</th><th>Adet</th><th>Toplam</th></tr>{trs}</table></div></div>'''
            return self.send_html(page('Raporlar',body,'/reports',u,msg))
        if path=='/reports/animals.xlsx':
            profile=farm_profile();group=q.get('animal_group',['all'])[0];status=q.get('animal_status',['Aktif'])[0];search=q.get('animal_search',[''])[0];paddock=q.get('animal_paddock',[''])[0]
            columns=animal_report_selected_columns(q);rows=animal_report_rows(group,status,search,paddock);b=animal_report_xlsx(rows,profile,columns);name=f'hayvan_raporu_{date.today().strftime("%Y%m%d")}.xlsx'
            self.send_response(200);self.send_header('Content-Type','application/vnd.openxmlformats-officedocument.spreadsheetml.sheet');self.send_header('Content-Disposition',f'attachment; filename="{name}"');self.send_header('Content-Length',str(len(b)));self.end_headers();self.wfile.write(b);return
        if path=='/reports/animals.pdf':
            profile=farm_profile();group=q.get('animal_group',['all'])[0];status=q.get('animal_status',['Aktif'])[0];search=q.get('animal_search',[''])[0];paddock=q.get('animal_paddock',[''])[0]
            columns=animal_report_selected_columns(q);rows=animal_report_rows(group,status,search,paddock);subtitle=' · '.join(x for x in [status+' Kayıtlar',{'female':'Dişi Hayvanlar','male':'Erkek Hayvanlar','calves':'Buzağılar'}.get(group,'Tüm Hayvanlar'),('Padok: '+paddock if paddock else '')] if x);b=animal_report_pdf(rows,profile,subtitle,columns);name=f'hayvan_raporu_{date.today().strftime("%Y%m%d")}.pdf'
            self.send_response(200);self.send_header('Content-Type','application/pdf');self.send_header('Content-Disposition',f'inline; filename="{name}"');self.send_header('Content-Length',str(len(b)));self.end_headers();self.wfile.write(b);return
        if path=='/reports/animals/print':
            profile=farm_profile();group=q.get('animal_group',['all'])[0];status=q.get('animal_status',['Aktif'])[0];search=q.get('animal_search',[''])[0];paddock=q.get('animal_paddock',[''])[0];columns=animal_report_selected_columns(q);rows=animal_report_rows(group,status,search,paddock)
            logo=f'<img src="{h(profile.get("farm_logo"))}" alt="İşletme logosu">' if profile.get('farm_logo') else '<div class="logo-mark">🐄</div>'
            column_keys=[x[0] for x in columns];ths=''.join(f'<th>{h(label)}</th>' for _,label,_,_ in columns)
            trs=''.join('<tr>'+''.join(f'<td>{"<b>" if key=="tag" else ""}{h(animal_report_display_value(r,key,i))}{"</b>" if key=="tag" else ""}</td>' for key,_,_,_ in columns)+'</tr>' for i,r in enumerate(rows,1)) or f'<tr><td colspan="{len(columns)}">Kayıt bulunamadı.</td></tr>'
            mobile_cards=''.join(f'''<article class="mobile-report-card"><div class="mobile-report-head">{'<span>#'+str(i)+'</span>' if 'row_no' in column_keys else ''}<strong>{h(r['tag'])}</strong>{'<em>'+h(r['group'])+'</em>' if 'group' in column_keys else ''}</div><div class="mobile-report-fields">{''.join(f'<div><small>{h(label)}</small><b>{h(animal_report_display_value(r,key,i))}</b></div>' for key,label,_,_ in columns if key not in ('row_no','tag','group'))}</div></article>''' for i,r in enumerate(rows,1)) or '<p class="mobile-empty">Kayıt bulunamadı.</p>'
            clean_query=urllib.parse.urlencode({'animal_group':group,'animal_status':status,'animal_search':search,'animal_paddock':paddock,'columns_mode':'custom','columns':column_keys},doseq=True)
            subtitle=' · '.join(x for x in [status+' Kayıtlar',{'female':'Dişi','male':'Erkek','calves':'Buzağı'}.get(group,'Tüm Hayvanlar'),('Padok: '+paddock if paddock else '')] if x)
            html=f'''<!doctype html><html lang="tr"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>Tüm Hayvanlar Raporu</title><style>
            *{{box-sizing:border-box}}body{{font-family:Arial,sans-serif;color:#183025;margin:0;background:#eef3ef}}.toolbar{{max-width:1400px;margin:12px auto;display:flex;gap:8px}}button,.toolbar a{{border:0;border-radius:8px;padding:10px 14px;background:#176b3a;color:white;text-decoration:none;font-weight:700;cursor:pointer;text-align:center}}.toolbar a:first-child{{background:#1468a0}}.sheet{{max-width:1400px;margin:0 auto 20px;background:#fff;padding:22px;box-shadow:0 8px 25px #0002}}.head{{display:flex;justify-content:space-between;gap:20px;border-bottom:2px solid #176b3a;padding-bottom:12px;margin-bottom:12px}}.brand{{display:flex;gap:14px;align-items:center}}.brand img,.logo-mark{{width:70px;height:70px;object-fit:contain;border-radius:8px}}.logo-mark{{display:grid;place-items:center;background:#edf6f0;font-size:36px}}h1{{font-size:22px;margin:0 0 5px}}.mut{{color:#617168;font-size:12px}}.meta{{text-align:right;font-size:12px;line-height:1.6}}.summary{{display:flex;gap:18px;margin:8px 0 12px;font-size:12px}}table{{width:100%;border-collapse:collapse;font-size:10px}}th,td{{border:1px solid #d8e3db;padding:5px 6px;text-align:left}}th{{background:#e9f3ec}}tbody tr:nth-child(even){{background:#f8faf8}}.foot{{margin-top:10px;font-size:10px;color:#68766d;display:flex;justify-content:space-between}}.mobile-report-list{{display:none}}@page{{size:A4 landscape;margin:9mm}}
            @media screen and (max-width:700px){{body{{background:#f2f5f3}}.toolbar{{position:sticky;top:0;z-index:10;margin:0;padding:8px;background:#fff;box-shadow:0 2px 10px #0002;display:grid;grid-template-columns:1fr 1fr}}.toolbar a,.toolbar button{{padding:11px 6px;font-size:12px}}.toolbar a:last-child{{grid-column:1/-1}}.sheet{{margin:0;padding:14px 12px 22px;box-shadow:none}}.head{{display:grid;grid-template-columns:1fr;gap:10px}}.brand{{align-items:flex-start}}.brand img,.logo-mark{{width:54px;height:54px;font-size:28px;flex:0 0 54px}}h1{{font-size:19px}}.brand b{{font-size:13px}}.meta{{text-align:left;background:#f3f7f4;border-radius:7px;padding:9px;display:grid;grid-template-columns:1fr 1fr;gap:3px 8px}}.summary{{display:grid;grid-template-columns:1fr 1fr;gap:6px;margin:10px 0}}.summary>*{{background:#eef5f0;border-radius:6px;padding:8px}}table{{display:none}}.mobile-report-list{{display:grid;gap:9px}}.mobile-report-card{{background:#fff;border:1px solid #d4e1d8;border-radius:9px;padding:11px;box-shadow:0 1px 3px #00000012}}.mobile-report-head{{display:flex;align-items:center;gap:7px;padding-bottom:8px;border-bottom:1px solid #e3eae5}}.mobile-report-head span{{color:#6a786f;font-size:11px}}.mobile-report-head strong{{font-size:15px;overflow-wrap:anywhere}}.mobile-report-head em{{margin-left:auto;background:#e7f4eb;color:#176b3a;border-radius:99px;padding:4px 8px;font-style:normal;font-size:10px;font-weight:700}}.mobile-report-fields{{display:grid;grid-template-columns:1fr 1fr;gap:9px 12px;padding-top:9px}}.mobile-report-fields div{{min-width:0}}.mobile-report-fields small{{display:block;color:#6a786f;font-size:10px;margin-bottom:2px}}.mobile-report-fields b{{display:block;font-size:12px;overflow-wrap:anywhere}}.mobile-empty{{padding:20px;text-align:center;background:#f7f9f7;border-radius:8px}}.foot{{gap:10px;flex-wrap:wrap}}}}
            @media print{{body{{background:#fff}}.toolbar,.mobile-report-list{{display:none!important}}.sheet{{max-width:none;margin:0;padding:0;box-shadow:none}}table{{display:table!important}}thead{{display:table-header-group}}tr{{break-inside:avoid;page-break-inside:avoid}}.head{{break-after:avoid}}}}
            </style></head><body><div class="toolbar"><a href="/reports/animals.pdf?{clean_query}">📄 Temiz PDF'yi Aç</a><button onclick="window.print()">🖨 Yazdır</button><a href="/reports#animal-report">← Raporlara Dön</a></div><main class="sheet"><header class="head"><div class="brand">{logo}<div><h1>{h(farm_display_name(profile))}</h1><b>İşletmede Bulunan Hayvanlar Raporu</b><div class="mut">{h(subtitle)}</div></div></div><div class="meta"><span><b>Rapor Tarihi:</b> {date.today().strftime('%d/%m/%Y')}</span><span><b>İşletme No:</b> {h(profile.get('business_no') or '-')}</span><span><b>İşletme Sahibi:</b> {h(profile.get('owner_name') or '-')}</span><span>{h(profile.get('province') or '')}{' / '+h(profile.get('district')) if profile.get('district') else ''}</span></div></header><div class="summary"><b>Toplam Hayvan: {len(rows)}</b><span>Dişi: {sum(1 for r in rows if r['group']=='Dişi')}</span><span>Erkek: {sum(1 for r in rows if r['group']=='Erkek')}</span><span>Buzağı: {sum(1 for r in rows if r['group']=='Buzağı')}</span></div><div class="mobile-report-list">{mobile_cards}</div><table><thead><tr>{ths}</tr></thead><tbody>{trs}</tbody></table><footer class="foot"><span>ÇiftlikPro Enterprise · {h(farm_display_name(profile))}</span><span>{len(rows)} kayıt</span></footer></main></body></html>'''
            return self.send_html(html)
        if path=='/animals/import-preview':
            token=q.get('token',[''])[0]
            with ANIMAL_IMPORT_LOCK:preview=ANIMAL_IMPORT_PREVIEWS.get(token)
            if not preview or preview.get('username')!=u:return self.redirect('/reports#animal-import','İçe aktarma önizlemesi bulunamadı veya süresi doldu.')
            rows=preview['rows'];ready=sum(1 for r in rows if r['valid']);errors=len(rows)-ready;warnings=sum(1 for r in rows if r['state']=='warning')
            trs=''.join(f'''<tr class="import-{r['state']}"><td>{r['row_no']}</td><td><b>{h(r['tag']) or '-'}</b></td><td>{h(r['record_type'])}</td><td>{h(r['breed']) or '-'}</td><td>{h(r['gender']) or '-'}</td><td>{fmt_date(r['birth_date']) or '-'}</td><td>{h(r['mother_tag']) or '-'}</td><td>{' · '.join(h(x) for x in r['issues']) or 'Hazır'}</td></tr>''' for r in rows[:500])
            more=f'<p class="mut">İlk 500 satır gösteriliyor; toplam {len(rows)} satır doğrulandı.</p>' if len(rows)>500 else ''
            body=f'''<h1>📥 Hayvan İçe Aktarma Önizlemesi</h1><div class="grid"><div class="card stat metric blue">Dosya Satırı<b>{len(rows)}</b></div><div class="card stat metric green">Aktarılabilir<b>{ready}</b></div><div class="card stat metric orange">Uyarılı<b>{warnings}</b></div><div class="card stat metric red">Atlanacak<b>{errors}</b></div></div><div class="card" style="margin-top:14px"><h2>{h(preview['filename'])}</h2><p class="mut">Yeşil satırlar hazır, sarı satırlar uyarılı fakat aktarılabilir, kırmızı satırlar aktarılmaz. Mevcut kayıtların üzerine yazılmaz.</p><style>.import-ready{{background:#eef9f1!important}}.import-warning{{background:#fff7df!important}}.import-error{{background:#fdebea!important}}.import-preview-table td{{font-size:12px}}</style><div class="tablewrap" style="max-height:58vh;overflow:auto"><table class="import-preview-table"><thead><tr><th>Satır</th><th>Küpe</th><th>Kayıt Grubu</th><th>Irk</th><th>Cinsiyet</th><th>Doğum</th><th>Ana No</th><th>Kontrol</th></tr></thead><tbody>{trs}</tbody></table></div>{more}<div class="actions"><form method="post" action="/animals/import-confirm" onsubmit="return confirm('{ready} geçerli hayvan kaydı içe aktarılsın mı?')"><input type="hidden" name="token" value="{h(token)}"><button class="btn blue" {"" if ready else "disabled"}>✅ {ready} Geçerli Kaydı İçe Aktar</button></form><a class="btn alt" href="/reports#animal-import">İptal</a></div></div>'''
            return self.send_html(page('Hayvan İçe Aktarma',body,'/reports',u,msg))
        if path=='/data':
            body="""<h1>Veri Aktarımı</h1><div class='card' style='margin-bottom:14px;border-left:5px solid #176b3a;background:linear-gradient(135deg,#f2faf4,#fff)'><h2 style='margin-top:0'>📥 Excel / PDF'den Hayvan Aktar</h2><p class='mut'>Bakanlık hayvan raporu PDF'sini, Excel (.xlsx) veya CSV listesini okuyup güvenli önizleme ile toplu hayvan kaydı oluşturur.</p><div class='actions'><a class='btn blue' href='/reports#animal-import'>Dosya Seçme Ekranını Aç</a></div></div><div class='two'><div class='card'><h2>JSON'dan İçe Aktar</h2><p class='mut'>Eski sistem yedeklerini ve V0.6 dışa aktarımlarını destekler. İçe aktarmadan önce otomatik veritabanı yedeği alınır.</p><form method='post' action='/data/import' enctype='multipart/form-data' class='form'><label class='full'>JSON dosyası<input type='file' name='json_file' accept='.json,application/json' required></label><label>Çakışan küpeler<select name='strategy'><option value='skip'>Atla (önerilen)</option><option value='update'>Mevcut kaydı güncelle</option></select></label><div class='full'><button class='btn'>İçe Aktar</button></div></form></div><div class='card'><h2>Dışa Aktar</h2><p>Tüm hayvan, tohumlama, buzağı, sağlık ve finans kayıtlarını tek JSON dosyasına aktarır.</p><div class='actions'><a class='btn blue' href='/data/export'>JSON Yedeğini İndir</a><a class='btn alt' href='/backups'>SQLite Yedekleri</a></div><hr><p class='mut'>JSON taşınabilir veri yedeğidir. SQLite yedeği uygulamanın birebir veritabanı kopyasıdır.</p></div></div>"""
            return self.send_html(page('Veri Aktarımı',body,'/data',u,msg))
        if path=='/data/export':
            b=json.dumps(export_payload(),ensure_ascii=False,indent=2).encode('utf-8');name=f'ciftlik_json_yedek_{datetime.now().strftime("%Y%m%d_%H%M%S")}.json'
            self.send_response(200);self.send_header('Content-Type','application/json; charset=utf-8');self.send_header('Content-Disposition',f'attachment; filename="{name}"');self.send_header('Content-Length',str(len(b)));self.end_headers();self.wfile.write(b);return
        if path=='/version-notes':
            body='''<h1>📝 Sürüm Notları</h1><div class="card"><h2>ÇiftlikPro Enterprise · v3.9.20</h2><p class="mut">Rasyon motoru Besi_V5.02.xlsm içindeki Solver yaklaşımı incelenerek kısıt-öncelikli yapıya taşındı.</p><ul><li><b>Besi Rasyonu:</b> Seçilen yemlerin miktarları aynı anda optimize edilir; KM, HP, ME ve kaba/kesif hedefleri yaklaşık ±%3,5 saha toleransıyla birlikte değerlendirilir.</li><li><b>Bilimsel katman:</b> NASEM gereksinim çekirdeği korunur; INRA yem değerleri yalnız veri varsa ayrı doğrulama katmanında kullanılır.</li><li><b>Rumen güvenliği:</b> NDF/eNDF, toplam ve rumende yıkılabilir nişasta birlikte yorumlanır. Gösterge göreli risktir; klinik pH tahmini değildir.</li><li><b>Ürün sınırları:</b> Evrensel tahıl/fabrika yemi yüzdesi uygulanmaz; girilmiş etiket veya uzman doz sınırı kesin kısıttır.</li><li><b>Maliyet:</b> Besleme, güvenlik ve fizibilite hedeflerinden sonra son seçim kriteridir.</li><li><b>Arayüz:</b> Kullanıcıya HOTFIX / DEV / PORT gibi teknik etiketler gösterilmez.</li></ul><p class="mut">Kaynak geçmişi CHANGELOG.md içinde tutulur.</p></div>'''
            return self.send_html(page('Sürüm Notları',body,'/version-notes',u,msg))
        if path=='/backups':
            if not self.require_admin():return
            target_dir=configured_backup_dir()
            raw_dir=(get_setting('backup_directory','') or '').strip()
            disk_files=[fp for fp in target_dir.glob('CiftlikPro_Backup_*.zip') if fp.is_file()]
            disk_files.sort(key=lambda x:x.stat().st_mtime,reverse=True)
            trs=''.join(
                f'<tr><td>{fmt_datetime(datetime.fromtimestamp(fp.stat().st_mtime).strftime("%Y-%m-%d %H:%M:%S"))}</td><td>{h(fp.name)}</td><td>{format_bytes(fp.stat().st_size)}</td><td><a class="btn blue" href="/backup/download?file={urllib.parse.quote(fp.name)}">İndir</a> <form method="post" action="/backup/restore-existing" class="inline-form" onsubmit="return confirm(\'Bu yedek geri yüklensin mi?\')"><input type="hidden" name="file" value="{h(fp.name)}"><button class="btn orange">Geri Yükle</button></form> <a class="btn red" href="/backup/delete?file={urllib.parse.quote(fp.name)}" onclick="return confirm(\'Bu yedek silinsin mi?\')">Sil</a></td></tr>'
                for fp in disk_files[:100]
            ) or '<tr><td colspan=4>Seçili klasörde yedek bulunamadı.</td></tr>'
            stats=uploads_storage_stats()
            body=f'''<h1>Yedekleme Merkezi</h1>
            <div class="grid" style="margin-bottom:14px">
              <div class="card stat"><span class="mut">Veritabanı</span><b>{format_bytes(stats["db_bytes"])}</b></div>
              <div class="card stat"><span class="mut">Fotoğraflar</span><b>{format_bytes(stats["upload_bytes"])}</b><small>{stats["upload_count"]} dosya</small></div>
              <div class="card stat"><span class="mut">Toplam Veri</span><b>{format_bytes(stats["total_bytes"])}</b></div>
              <div class="card stat"><span class="mut">Yedek Klasörü</span><b style="font-size:14px;word-break:break-all">{h(str(target_dir))}</b></div>
            </div>
            <div class="two">
              <div class="card"><h2>Yedekleme Ayarları</h2>
                <form method="post" action="/backup/settings" class="form">
                  <label class="full">Yedek Klasörü<input name="backup_directory" value="{h(raw_dir or str(BACKUPS))}" placeholder="D:\\CiftlikPro_Yedekler"></label>
                  <div class="full mut">D:, harici disk veya erişilebilir ağ klasörü kullanabilirsiniz. Program yolu hatırlar.</div>
                  <div class="full"><button class="btn">Klasörü Kaydet ve Test Et</button></div>
                </form>
                <hr style="border:0;border-top:1px solid #e2ece5;margin:18px 0">
                <h3>Fotoğraf Optimizasyonu</h3>
                <p class="mut">Yeni fotoğraflar otomatik olarak 1024 px WebP biçiminde saklanır.</p>
                <form method="post" action="/photos/optimize-existing" onsubmit="return confirm('Mevcut fotoğraflar küçültülsün mü? Önce güvenlik yedeği alınacaktır.')"><button class="btn orange">Mevcut Fotoğrafları Optimize Et</button></form>
              </div>
              <div class="card"><h2>Tam Yedek Al</h2><p>Veritabanı ve fotoğraflar seçili klasörde tek ZIP dosyasında saklanır.</p><a class="btn orange" href="/backup/create">Şimdi Yedek Al</a>
                <hr style="border:0;border-top:1px solid #e2ece5;margin:18px 0">
                <h2>Dosyadan Geri Yükle</h2><form method="post" action="/backup/restore" enctype="multipart/form-data"><input type="file" name="backup_file" accept=".zip" required><label style="display:block;margin:12px 0"><input type="checkbox" name="confirm_restore" value="yes" required> Mevcut verilerin değiştirileceğini kabul ediyorum.</label><button class="btn red">Yedeği Geri Yükle</button></form>
              </div>
            </div>
            <div class="card" style="margin-top:14px"><h2>Seçili Klasördeki Yedekler</h2><div class="tablewrap"><table><tr><th>Tarih</th><th>Dosya</th><th>Boyut</th><th>İşlem</th></tr>{trs}</table></div></div>'''
            return self.send_html(page('Yedekleme Merkezi',body,'/backups',u,msg))
        if path=='/backup/create':
            if not self.require_admin():return
            name=create_backup('manuel');audit(u,'Yedek oluşturdu',name,self.client_ip());return self.redirect('/backups','Tam yedek oluşturuldu.')
        if path=='/backup/download':
            if not self.require_admin():return
            name=os.path.basename(q.get('file',[''])[0]);fp=configured_backup_dir()/name
            if not fp.exists():return self.send_html('Dosya bulunamadı',404)
            b=fp.read_bytes();self.send_response(200);self.send_header('Content-Type','application/zip');self.send_header('Content-Disposition',f'attachment; filename="{name}"');self.send_header('Content-Length',str(len(b)));self.end_headers();self.wfile.write(b);return
        if path=='/backup/delete':
            if not self.require_admin():return
            name=os.path.basename(q.get('file',[''])[0]);fp=configured_backup_dir()/name
            if fp.exists():fp.unlink()
            with db() as c:c.execute('delete from backups where filename=?',(name,))
            audit(u,'Yedek sildi',name,self.client_ip());return self.redirect('/backups','Yedek silindi.')
        if path in ('/finance/export','/reports/export'):
            start=q.get('start',['0000-01-01'])[0];end=q.get('end',['9999-12-31'])[0];typ=q.get('type',[''])[0];category=q.get('category',[''])[0]
            sql='select tx_date,tx_type,category,amount,description,payment_method from finance where tx_date between ? and ?';args=[start,end]
            if typ: sql+=' and tx_type=?'; args.append(typ)
            if category: sql+=' and category=?'; args.append(category)
            with db() as c: rows=c.execute(sql,args).fetchall()
            out=io.StringIO();w=csv.writer(out,delimiter=';');w.writerow(['Tarih','Tür','Kategori','Tutar','Açıklama','Ödeme Yöntemi']);w.writerows(rows);b=('\ufeff'+out.getvalue()).encode('utf-8')
            self.send_response(200);self.send_header('Content-Type','text/csv; charset=utf-8');self.send_header('Content-Disposition',f'attachment; filename="finans_{start}_{end}.csv"');self.send_header('Content-Length',str(len(b)));self.end_headers();self.wfile.write(b);return
        self.send_html('Sayfa bulunamadı',404)
    def do_POST(self):
        ensure_archive_schema()
        path=urllib.parse.urlparse(self.path).path
        f={}
        try:
            f=self.post_data()
        except Exception as exc:
            return self.redirect('/','Form verisi okunamadı: '+str(exc))
        if path=='/license-test':
            if not self.require_admin():return
            key=(f.get('license_key') or '').strip()
            try: raw=license_key_to_bytes(key)
            except Exception:return self.redirect('/license-test','❌ Lisans anahtarı eksik, bozuk veya okunamadı. Anahtarı yeniden kopyalayıp yapıştırın.')
            ok,payload,status=validate_license_bytes(raw);payload=payload or {}
            if not ok:return self.redirect('/license-test','❌ '+str(status))
            exp=payload.get('expires_on') or 'Süresiz'
            if exp!='Süresiz':exp=fmt_date(exp)
            body=f'''<h1>🧪 Lisans Test Sonucu</h1><div class="card"><p><span class="license-ok">✅ Anahtar Geçerli</span></p><div class="kv"><div><b>Dijital İmza</b><span>✅ Doğru</span></div><div><b>Cihaz Kimliği</b><span>✅ Eşleşiyor · {h(device_id())}</span></div><div><b>Lisans Sahibi</b><span>{h(payload.get('licensee') or '-')}</span></div><div><b>Lisans Türü</b><span>{h(payload.get('license_type') or '-')}</span></div><div><b>Geçerlilik</b><span>{h(exp)}</span></div></div><p class="mut">Test başarılı. Mevcut aktif lisans değiştirilmedi.</p><a class="btn" href="/license-test">Başka Anahtar Test Et</a> <a class="btn secondary" href="/license-info">Lisans Bilgilerine Dön</a></div>'''
            return self.send_html(page('Lisans Test Sonucu',body,'/license-info',u,''))
        if path=='/license-change':
            if not self.require_admin():return
            if (f.get('confirm_change') or '')!='yes':return self.redirect('/license-change','Değişiklik için onay kutusunu işaretleyin.')
            key=(f.get('license_key') or '').strip()
            try:raw=license_key_to_bytes(key)
            except Exception:return self.redirect('/license-change','❌ Lisans anahtarı eksik, bozuk veya okunamadı. Anahtarı yeniden kopyalayıp yapıştırın.')
            ok,payload,status=validate_license_bytes(raw)
            if not ok:return self.redirect('/license-change','❌ '+str(status))
            tmp=LICENSE_FILE.with_suffix('.license.tmp');tmp.write_bytes(raw);os.replace(tmp,LICENSE_FILE)
            return self.redirect('/license-info','✅ Yeni lisans başarıyla etkinleştirildi.')
        if path=='/license-key-activate':
            key=(f.get('license_key') or '').strip()
            try: raw=license_key_to_bytes(key)
            except Exception: return self.redirect('/license','Lisans anahtarı okunamadı veya geçersiz.')
            ok,payload,status=validate_license_bytes(raw)
            if not ok:return self.redirect('/license',status)
            tmp=LICENSE_FILE.with_suffix('.license.tmp');tmp.write_bytes(raw);os.replace(tmp,LICENSE_FILE)
            return self.redirect('/login','Lisans başarıyla etkinleştirildi.')
        if path=='/license-activate':
            item=f.get('license_file')
            if not isinstance(item,dict) or not item.get('content'):
                return self.redirect('/license','Lütfen geçerli bir lisans dosyası seçin.')
            raw=item['content']
            ok,payload,status=validate_license_bytes(raw)
            if not ok:return self.redirect('/license',status)
            tmp=LICENSE_FILE.with_suffix('.license.tmp');tmp.write_bytes(raw);os.replace(tmp,LICENSE_FILE)
            return self.redirect('/login','Lisans başarıyla etkinleştirildi.')
        if path=='/forgot-password':
            identifier=(f.get('identifier') or '').strip()
            if not identifier:return self.redirect('/forgot-password','Kullanıcı adı veya e-posta girin.')
            with db() as c:user=c.execute("select * from users where active=1 and (lower(username)=lower(?) or lower(coalesce(recovery_email,''))=lower(?)) limit 1",(identifier,identifier)).fetchone()
            if not user or not (user['recovery_email'] or '').strip():return self.redirect('/forgot-password','Bu hesap için kurtarma e-postası tanımlı değil veya hesap bulunamadı.')
            with db() as c:recent=c.execute('select created_at from password_reset_codes where user_id=? order by id desc limit 1',(user['id'],)).fetchone()
            if recent:
                try:
                    if (datetime.now()-datetime.fromisoformat(recent['created_at'])).total_seconds()<60:return self.redirect('/forgot-password','Yeni kod istemeden önce 60 saniye bekleyin.')
                except Exception:pass
            code=f'{secrets.randbelow(1000000):06d}';salt=secrets.token_hex(16);expires=(datetime.now()+timedelta(minutes=5)).isoformat(timespec='seconds');created=datetime.now().isoformat(timespec='seconds')
            with db() as c:
                c.execute('insert into password_reset_codes(user_id,code_hash,salt,expires_at,attempts,used,created_at,ip_address) values(?,?,?,?,0,0,?,?)',(user['id'],reset_code_hash(salt,code),salt,expires,created,self.client_ip()));rid=c.execute('select last_insert_rowid()').fetchone()[0]
            try:send_reset_email(user['recovery_email'],user['full_name'],code)
            except Exception as exc:
                with db() as c:c.execute('update password_reset_codes set used=1 where id=?',(rid,))
                return self.redirect('/forgot-password','E-posta gönderilemedi. Yönetici SMTP ayarlarını kontrol etsin: '+str(exc))
            audit(user['username'],'Şifre sıfırlama kodu istedi','',self.client_ip());return self.redirect(f'/forgot-password?step=verify&id={rid}','Doğrulama kodu kurtarma e-postasına gönderildi.')
        if path=='/forgot-verify':
            rid=int(f.get('reset_id') or 0);code=(f.get('code') or '').strip()
            with db() as c:r=c.execute('select pr.*,u.username from password_reset_codes pr join users u on u.id=pr.user_id where pr.id=?',(rid,)).fetchone()
            if not r:return self.redirect('/forgot-password','Şifre sıfırlama isteği bulunamadı.')
            if r['used'] or int(r['attempts'] or 0)>=3:return self.redirect('/forgot-password','Bu doğrulama isteği artık kullanılamaz.')
            try:
                if datetime.now()>datetime.fromisoformat(r['expires_at']):
                    with db() as c:c.execute('update password_reset_codes set used=1 where id=?',(rid,))
                    return self.redirect('/forgot-password','Doğrulama kodunun süresi doldu. Yeni kod isteyin.')
            except Exception:return self.redirect('/forgot-password','Doğrulama isteği geçersiz.')
            if not hmac.compare_digest(reset_code_hash(r['salt'],code),r['code_hash']):
                attempts=int(r['attempts'] or 0)+1
                with db() as c:c.execute('update password_reset_codes set attempts=?,used=? where id=?',(attempts,1 if attempts>=3 else 0,rid))
                return self.redirect(f'/forgot-password?step=verify&id={rid}',f'Kod hatalı. Kalan deneme: {max(0,3-attempts)}')
            token=secrets.token_urlsafe(32);token_exp=(datetime.now()+timedelta(minutes=10)).isoformat(timespec='seconds')
            with db() as c:c.execute('update password_reset_codes set reset_token_hash=?,reset_token_expires=? where id=?',(reset_token_hash(token),token_exp,rid))
            return self.redirect(f'/forgot-password?step=reset&id={rid}&token={urllib.parse.quote(token)}','Kod doğrulandı.')
        if path=='/forgot-reset':
            rid=int(f.get('reset_id') or 0);token=f.get('token') or '';pw=f.get('password') or '';pw2=f.get('password_confirm') or ''
            if len(pw)<8:return self.redirect('/forgot-password','Yeni şifre en az 8 karakter olmalıdır.')
            if pw!=pw2:return self.redirect('/forgot-password','Yeni şifreler eşleşmiyor.')
            with db() as c:r=c.execute('select pr.*,u.username from password_reset_codes pr join users u on u.id=pr.user_id where pr.id=?',(rid,)).fetchone()
            if not r or r['used'] or not r['reset_token_hash']:return self.redirect('/forgot-password','Şifre sıfırlama bağlantısı geçersiz.')
            try:
                if datetime.now()>datetime.fromisoformat(r['reset_token_expires']):return self.redirect('/forgot-password','Şifre sıfırlama bağlantısının süresi doldu.')
            except Exception:return self.redirect('/forgot-password','Şifre sıfırlama bağlantısı geçersiz.')
            if not hmac.compare_digest(reset_token_hash(token),r['reset_token_hash']):return self.redirect('/forgot-password','Şifre sıfırlama bağlantısı geçersiz.')
            with db() as c:c.execute('update users set password=?,password_changed_at=? where id=?',(password_hash(pw),datetime.now().strftime('%Y-%m-%d %H:%M:%S'),r['user_id']));c.execute('update password_reset_codes set used=1 where id=?',(rid,))
            audit(r['username'],'Şifresini e-posta doğrulamasıyla sıfırladı','',self.client_ip());return self.redirect('/login','Şifreniz yenilendi. Yeni şifrenizle giriş yapabilirsiniz.')
        if path=='/login':
            ok,_,status=license_status()
            if not ok:return self.redirect('/license',status)
            username=(f.get('username') or '').strip()
            with db() as c:r=c.execute('select * from users where username=?',(username,)).fetchone()
            if not r or not r['active'] or not password_verify(f.get('password',''),r['password']):audit(username or 'bilinmeyen','Başarısız giriş','',self.client_ip());return self.redirect('/login','Kullanıcı adı veya şifre hatalı ya da hesap pasif.')
            if not str(r['password']).startswith('pbkdf2_sha256$'):
                with db() as c:c.execute('update users set password=?,password_changed_at=? where id=?',(password_hash(f.get('password','')),datetime.now().strftime('%Y-%m-%d %H:%M:%S'),r['id']))
            now=datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            with db() as c:c.execute('update users set last_login=? where id=?',(now,r['id']))
            sid=secrets.token_urlsafe(24);SESSIONS[sid]={'id':r['id'],'username':r['username'],'role':r['role'],'full_name':r['full_name']};audit(r['username'],'Oturum açtı','',self.client_ip())
            self.send_response(303);self.send_header('Set-Cookie',f'sid={sid}; HttpOnly; SameSite=Lax; Path=/');self.send_header('Location','/');self.end_headers();return
        if not self.require():return
        current=self.user();username=current['username']
        if path=='/animals/import-preview':
            upload=f.get('animal_file')
            if not isinstance(upload,dict) or not upload.get('filename') or not upload.get('content'):
                return self.redirect('/reports#animal-import','Lütfen XLSX, CSV veya PDF hayvan listesi seçin.')
            try:
                raw_rows=parse_animal_import_file(upload['filename'],upload['content'])
                rows=prepare_animal_import(upload['filename'],raw_rows)
            except Exception as exc:
                return self.redirect('/reports#animal-import','Dosya okunamadı: '+str(exc))
            token=secrets.token_urlsafe(24);now=time.time()
            with ANIMAL_IMPORT_LOCK:
                for old_token,item in list(ANIMAL_IMPORT_PREVIEWS.items()):
                    if now-float(item.get('created_at') or 0)>7200:ANIMAL_IMPORT_PREVIEWS.pop(old_token,None)
                ANIMAL_IMPORT_PREVIEWS[token]={'username':username,'filename':os.path.basename(upload['filename']),'created_at':now,'rows':rows}
            audit(username,'Hayvan içe aktarma önizlemesi',f'{os.path.basename(upload["filename"])} · {len(rows)} satır',self.client_ip())
            return self.redirect('/animals/import-preview?token='+urllib.parse.quote(token))
        if path=='/animals/import-confirm':
            token=(f.get('token') or '').strip()
            with ANIMAL_IMPORT_LOCK:preview=ANIMAL_IMPORT_PREVIEWS.get(token)
            if not preview or preview.get('username')!=username:return self.redirect('/reports#animal-import','İçe aktarma önizlemesi bulunamadı veya süresi doldu.')
            valid=[r for r in preview['rows'] if r.get('valid')]
            if not valid:return self.redirect('/reports#animal-import','Aktarılabilir hayvan kaydı bulunamadı.')
            try:create_backup('hayvan_aktarim_oncesi')
            except Exception as exc:return self.redirect('/animals/import-preview?token='+urllib.parse.quote(token),'Güvenlik yedeği alınamadığı için aktarım başlatılmadı: '+str(exc))
            imported=0;skipped=0;calf_rows=[]
            try:
                with db() as c:
                    for row in valid:
                        if row['record_type']=='Buzağı':calf_rows.append(row);continue
                        if c.execute('select 1 from animals where tag=?',(row['tag'],)).fetchone() or c.execute('select 1 from calves where tag=?',(row['tag'],)).fetchone():skipped+=1;continue
                        paddock_id=None
                        if row['paddock']:
                            pd=c.execute('select id from paddocks where name=? and active=1',(row['paddock'],)).fetchone();paddock_id=pd['id'] if pd else None
                        c.execute('''insert into animals(tag,nickname,gender,breed,birth_date,notes,paddock,paddock_id,photo_url,sold_price,status,purchase_date,purchase_price)
                                     values(?,?,?,?,?,?,?,?,?,0,'Aktif',?,0)''',(row['tag'],row['nickname'],row['gender'],row['breed'],row['birth_date'],row['notes'],row['paddock'],paddock_id,'',row['arrival_date']))
                        imported+=1
                    for row in calf_rows:
                        if c.execute('select 1 from animals where tag=?',(row['tag'],)).fetchone() or c.execute('select 1 from calves where tag=?',(row['tag'],)).fetchone():skipped+=1;continue
                        mother=c.execute("select id from animals where tag=? and gender='Dişi' and coalesce(status,'Aktif')='Aktif'",(row['mother_tag'],)).fetchone()
                        if not mother:skipped+=1;continue
                        paddock_id=None
                        if row['paddock']:
                            pd=c.execute('select id from paddocks where name=? and active=1',(row['paddock'],)).fetchone();paddock_id=pd['id'] if pd else None
                        c.execute('''insert into calves(tag,nickname,mother_id,father_tag,birth_date,gender,breed,paddock,paddock_id,photo_url,purchase_date,purchase_price,purchase_payment_method,notes)
                                     values(?,?,?,?,?,?,?,?,?,?,?,0,'Nakit',?)''',(row['tag'],row['nickname'],mother['id'],'',row['birth_date'],row['gender'],row['breed'],row['paddock'],paddock_id,'',row['arrival_date'],row['notes']))
                        imported+=1
            except Exception as exc:
                return self.redirect('/animals/import-preview?token='+urllib.parse.quote(token),'Aktarım tamamlanamadı; veritabanı değişiklikleri geri alındı: '+str(exc))
            with ANIMAL_IMPORT_LOCK:ANIMAL_IMPORT_PREVIEWS.pop(token,None)
            audit(username,'Dosyadan hayvan aktardı',f'{preview["filename"]} · {imported} aktarıldı · {skipped} atlandı',self.client_ip())
            return self.redirect('/reports#animal-report',f'✅ {imported} hayvan başarıyla aktarıldı. {skipped} mükerrer veya ilişkisiz kayıt atlandı.')
        if path=='/paddock/create':
            name=(f.get('name') or '').strip()
            if not name:return self.redirect('/paddocks','Padok adı zorunludur.')
            try:
                cap=max(0,int(float(f.get('capacity') or 0)))
            except Exception:cap=0
            try:
                with db() as c:c.execute('insert into paddocks(name,code,type,capacity,notes,active,created_at) values(?,?,?,?,?,1,?)',(name,(f.get('code') or '').strip(),(f.get('type') or 'Genel').strip(),cap,(f.get('notes') or '').strip(),datetime.now().isoformat(timespec='seconds')))
            except sqlite3.IntegrityError:return self.redirect('/paddocks','Bu padok adı zaten kayıtlı.')
            audit(username,'Padok oluşturdu',name,self.client_ip());return self.redirect('/paddocks','Padok oluşturuldu.')
        if path=='/paddock/assign':
            ref=(f.get('animal_ref') or '').strip();pid=int(f.get('paddock_id') or 0) or None
            if ':' not in ref:return self.redirect('/paddocks','Hayvan seçin.')
            source,raw_id=ref.split(':',1)
            if source not in ('animal','calf'):return self.redirect('/paddocks','Geçersiz hayvan türü.')
            aid=int(raw_id or 0);table='animals' if source=='animal' else 'calves'
            with db() as c:
                rec=c.execute(f'select id,tag,paddock_id from {table} where id=?',(aid,)).fetchone()
                if not rec:return self.redirect('/paddocks','Hayvan bulunamadı.')
                if pid and not c.execute('select id from paddocks where id=? and active=1',(pid,)).fetchone():return self.redirect('/paddocks','Padok bulunamadı.')
                old=rec['paddock_id'];sync_paddock_text(c,source,aid,pid)
                if old!=pid:c.execute('insert into paddock_history(animal_source,animal_id,from_paddock_id,to_paddock_id,moved_at,notes) values(?,?,?,?,?,?)',(source,aid,old,pid,datetime.now().isoformat(timespec='seconds'),(f.get('notes') or '').strip()))
            audit(username,'Hayvan padok taşıma',f'{rec["tag"]} -> {pid or "Padoksuz"}',self.client_ip());return self.redirect('/paddocks','Hayvanın padoku güncellendi.')
        if path=='/feed/create':
            name=(f.get('name') or '').strip()
            if not name:return self.redirect('/feeds','Yem adı zorunludur.')
            def num(k):
                try:return float(f.get(k) or 0)
                except:return 0.0
            def pct(k):return max(0.0,min(100.0,num(k)))
            label_min=max(0.0,num('solver_min_kg_day'));label_max=max(0.0,num('solver_max_kg_day'))
            if label_min>0 and label_max>0 and label_min>label_max:
                return self.redirect('/feeds','Etiket alt dozu üst dozdan büyük olamaz.')
            try:
                with db() as c:c.execute('''insert into feed_catalog(
                    name,category,dm_pct,ndf_pct,effective_ndf_pct,cp_pct,starch_pct,me_mcal_kg,ca_pct,p_pct,
                    label_cp_pct_as_fed,label_me_kcal_kg_as_fed,label_crude_fiber_pct_as_fed,
                    label_fat_pct_as_fed,label_ash_pct_as_fed,label_sodium_pct_as_fed,
                    starch_degradability_pct,ndf_digestibility_pct,processing_method,
                    solver_min_kg_day,solver_max_kg_day,constraint_source,source,active
                ) values(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,1)''',(
                    name,(f.get('category') or 'Özel Yem').strip(),pct('dm_pct'),pct('ndf_pct'),
                    pct('effective_ndf_pct'),pct('cp_pct'),pct('starch_pct'),max(0.0,num('me_mcal_kg')),
                    pct('ca_pct'),pct('p_pct'),pct('label_cp_pct_as_fed'),max(0.0,num('label_me_kcal_kg_as_fed')),
                    pct('label_crude_fiber_pct_as_fed'),pct('label_fat_pct_as_fed'),
                    pct('label_ash_pct_as_fed'),pct('label_sodium_pct_as_fed'),
                    pct('starch_degradability_pct'),pct('ndf_digestibility_pct'),
                    (f.get('processing_method') or '').strip(),label_min,label_max,
                    (f.get('constraint_source') or '').strip(),'Kullanıcı girişi'))
            except sqlite3.IntegrityError:return self.redirect('/feeds','Bu yem zaten katalogda var.')
            audit(username,'Yem kataloğuna ekledi',name,self.client_ip());return self.redirect('/feeds','Özel yem kataloğa eklendi.')
        if path=='/feed/edit':
            try: fid=int(f.get('feed_id') or 0)
            except:return self.redirect('/feeds','Geçersiz yem kaydı.')
            name=(f.get('name') or '').strip()
            if fid<=0 or not name:return self.redirect('/feeds','Yem adı zorunludur.')
            def num(k):
                try:return float(f.get(k) or 0)
                except:return 0.0
            def pct(k):return max(0.0,min(100.0,num(k)))
            label_min=max(0.0,num('solver_min_kg_day'));label_max=max(0.0,num('solver_max_kg_day'))
            if label_min>0 and label_max>0 and label_min>label_max:
                return self.redirect('/feed-edit?id='+str(fid),'Etiket alt dozu üst dozdan büyük olamaz.')
            try:
                with db() as c:
                    old=c.execute('select name from feed_catalog where id=? and active=1',(fid,)).fetchone()
                    if not old:return self.redirect('/feeds','Yem bulunamadı.')
                    c.execute('''update feed_catalog set name=?,category=?,dm_pct=?,ndf_pct=?,effective_ndf_pct=?,cp_pct=?,tdn_pct=?,me_mcal_kg=?,nem_mcal_kg=?,neg_mcal_kg=?,starch_pct=?,fat_pct=?,ash_pct=?,ca_pct=?,p_pct=?,mg_pct=?,k_pct=?,na_pct=?,s_pct=?,label_cp_pct_as_fed=?,label_me_kcal_kg_as_fed=?,label_crude_fiber_pct_as_fed=?,label_fat_pct_as_fed=?,label_ash_pct_as_fed=?,label_sodium_pct_as_fed=?,starch_degradability_pct=?,ndf_digestibility_pct=?,rdp_pct_cp=?,rup_pct_cp=?,inra_ufv=?,inra_pdi_g_kg_dm=?,inra_pdia_g_kg_dm=?,inra_rpb_g_kg_dm=?,inra_fill_unit=?,processing_method=?,solver_min_kg_day=?,solver_max_kg_day=?,constraint_source=?,source=? where id=?''',(
                        name,(f.get('category') or 'Özel Yem').strip(),
                        pct('dm_pct'),pct('ndf_pct'),pct('effective_ndf_pct'),pct('cp_pct'),pct('tdn_pct'),
                        max(0.0,num('me_mcal_kg')),max(0.0,num('nem_mcal_kg')),max(0.0,num('neg_mcal_kg')),
                        pct('starch_pct'),pct('fat_pct'),pct('ash_pct'),pct('ca_pct'),pct('p_pct'),pct('mg_pct'),pct('k_pct'),pct('na_pct'),pct('s_pct'),
                        pct('label_cp_pct_as_fed'),max(0.0,num('label_me_kcal_kg_as_fed')),
                        pct('label_crude_fiber_pct_as_fed'),pct('label_fat_pct_as_fed'),
                        pct('label_ash_pct_as_fed'),pct('label_sodium_pct_as_fed'),
                        pct('starch_degradability_pct'),pct('ndf_digestibility_pct'),pct('rdp_pct_cp'),pct('rup_pct_cp'),
                        max(0.0,num('inra_ufv')),max(0.0,num('inra_pdi_g_kg_dm')),max(0.0,num('inra_pdia_g_kg_dm')),num('inra_rpb_g_kg_dm'),max(0.0,num('inra_fill_unit')),
                        (f.get('processing_method') or '').strip(),label_min,label_max,(f.get('constraint_source') or '').strip(),
                        (f.get('source') or 'Kullanıcı tarafından güncellendi').strip(),fid))
            except sqlite3.IntegrityError:return self.redirect('/feed-edit?id='+str(fid),'Bu yem adı başka bir kayıtta kullanılıyor.')
            audit(username,'Yem kataloğunu düzenledi',f'{old["name"]} -> {name}',self.client_ip());return self.redirect('/feeds','Yem bilgileri güncellendi.')
        if path=='/feed/delete':
            try: fid=int(f.get('feed_id') or 0)
            except:return self.redirect('/feeds','Geçersiz yem kaydı.')
            with db() as c:
                feed=c.execute('select name from feed_catalog where id=? and active=1',(fid,)).fetchone()
                if not feed:return self.redirect('/feeds','Yem bulunamadı.')
                c.execute('update feed_catalog set active=0 where id=?',(fid,))
            audit(username,'Yemi katalogdan kaldırdı',feed['name'],self.client_ip());return self.redirect('/feeds','Yem katalogdan kaldırıldı. Geçmiş rasyon/fiyat/stok kayıtları korundu.')
        if path=='/feed/price':
            try:fid=int(f.get('feed_id') or 0);price=float(f.get('price_per_kg') or 0)
            except:return self.redirect('/feeds','Yem ve fiyat bilgisi geçersiz.')
            if fid<=0 or price<0:return self.redirect('/feeds','Yem ve fiyat bilgisi geçersiz.')
            d=(f.get('effective_date') or date.today().isoformat()).strip()
            with db() as c:
                feed=c.execute('select name from feed_catalog where id=?',(fid,)).fetchone()
                if not feed:return self.redirect('/feeds','Yem bulunamadı.')
                c.execute('insert into feed_prices(feed_id,effective_date,price_per_kg,notes) values(?,?,?,?)',(fid,d,price,(f.get('notes') or '').strip()))
            audit(username,'Yem fiyatı girdi',f'{feed["name"]}: {price}',self.client_ip());return self.redirect('/feeds','Yem fiyatı kaydedildi.')
        if path=='/feed/stock':
            try:fid=int(f.get('feed_id') or 0);qty=float(f.get('quantity_kg') or 0);unit=float(f.get('unit_price') or 0)
            except:return self.redirect('/feeds','Stok bilgisi geçersiz.')
            typ=(f.get('tx_type') or 'Giriş').strip()
            if fid<=0 or qty<=0 or typ not in ('Giriş','Çıkış','Tüketim','Sayım +','Sayım -'):return self.redirect('/feeds','Stok bilgisi geçersiz.')
            d=(f.get('tx_date') or date.today().isoformat()).strip()
            with db() as c:
                feed=c.execute('select name from feed_catalog where id=?',(fid,)).fetchone()
                if not feed:return self.redirect('/feeds','Yem bulunamadı.')
                c.execute('insert into feed_stock_transactions(feed_id,tx_date,tx_type,quantity_kg,unit_price,notes) values(?,?,?,?,?,?)',(fid,d,typ,qty,unit,(f.get('notes') or '').strip()))
                stock_id=c.execute('select last_insert_rowid()').fetchone()[0]
                if typ=='Giriş' and unit>0:c.execute('insert into feed_prices(feed_id,effective_date,price_per_kg,notes) values(?,?,?,?)',(fid,d,unit,'Stok girişinden otomatik fiyat'))
                finance_created=False
                if typ=='Giriş' and (f.get('post_to_finance') or '')=='yes':
                    if unit<=0:return self.redirect('/feeds','Finansa aktarım için alış ₺/kg 0’dan büyük olmalıdır.')
                    amount=round(qty*unit,2);desc=(f.get('notes') or '').strip();desc=(desc+' · ' if desc else '')+f'Yem stok alımı: {feed["name"]} · {qty:g} kg × {unit:g} TL'
                    c.execute('insert into finance(tx_date,tx_type,category,amount,description,payment_method,animal_id,created_at,animal_status_action) values(?,?,?,?,?,?,?,?,?)',(d,'Gider','Yem',amount,desc,f.get('payment_method') or 'Nakit',None,datetime.now().isoformat(),''))
                    finance_id=c.execute('select last_insert_rowid()').fetchone()[0]
                    c.execute('insert into feed_finance_links(feed_id,stock_tx_id,finance_id,quantity_kg,unit_price,created_at) values(?,?,?,?,?,?)',(fid,stock_id,finance_id,qty,unit,datetime.now().isoformat()))
                    finance_created=True
            audit(username,'Yem stok hareketi',f'{feed["name"]} {typ} {qty} kg',self.client_ip());return self.redirect('/feeds','Stok hareketi kaydedildi.'+(' Finans gideri de otomatik oluşturuldu.' if finance_created else ''))
        if path=='/ration/solve':
            rtype=(f.get('ration_type') or 'Besi').strip()
            try:
                w=float(f.get('target_weight_kg') or (650 if rtype=='Süt' else 450));adg=float(f.get('target_adg_kg') or 1.3);age=float(f.get('target_age_months') or 0);animal_type=(f.get('animal_type') or 'Besi Erkek').strip();phase_override=(f.get('target_beef_phase') or 'Otomatik').strip();milk=float(f.get('target_milk_l') or 25)
            except:return self.redirect('/rations?solve=1','Rasyon hedef bilgileri geçersiz.')
            if rtype=='Süt':
                if not (350<=w<=900 and 0<=milk<=70):return self.redirect('/rations?solve=1','Süt rasyonu hedefleri izin verilen aralıkta değil.')
            elif not (150<=w<=900 and .2<=adg<=2.2):return self.redirect('/rations?solve=1','Besi hedef değerleri izin verilen aralıkta değil.')
            ids=[]
            for k,v in f.items():
                if k.startswith('feed_') and str(v)=='1':
                    try:ids.append(int(k[5:]))
                    except:pass
            if len(ids)<2:return self.redirect('/rations?solve=1','Rasyon çözmek için en az 2 yem seçin.')
            with db() as c:
                marks=','.join('?'*len(ids)); rows=c.execute(f"select f.*,coalesce((select fp.price_per_kg from feed_prices fp where fp.feed_id=f.id and fp.effective_date<=? order by fp.effective_date desc,fp.id desc limit 1),0) price from feed_catalog f where f.active=1 and f.id in ({marks})",(date.today().isoformat(),*ids)).fetchall()
                if rtype=='Süt':
                    solved,t,warn=solve_smart_dairy_ration(rows,w,milk)
                else:
                    solved,t,warn=solve_smart_ration(rows,w,adg,animal_type,age,phase_override)
                if not solved:return self.redirect('/rations?solve=1',warn or 'Seçilen yemlerle çözüm üretilemedi.')
                qty,m,bounds=solved; stamp=datetime.now().strftime('%d%m-%H%M%S')
                if rtype=='Süt':
                    risk=m.get('rumen_risk',{}).get('level','Belirsiz')
                    name=f'Akıllı Süt {w:.0f}kg {milk:.1f}kg {stamp}'; note=f'Süt_V5.01 / INRA-NASEM referanslı süt hedef motoru · KM {t["dmi_kg"]:.2f} kg · HP %{t["cp_pct"]:.1f} · ME {t["me_mcal_day"]:.1f} Mcal · Göreli asidoz riski {risk} (klinik pH tahmini değildir)' + ((' · '+warn) if warn else '')
                    cur=c.execute('insert into rations(name,target_group,notes,active,created_at,target_weight_kg,target_adg_kg,animal_type,ration_type,target_milk_l,milk_fat_pct,milk_protein_pct,target_age_months) values(?,?,?,1,?,?,?,?,?,?,?,?,?)',(name,'Sağmal',note,datetime.now().isoformat(timespec='seconds'),w,0.0,'Sağmal İnek','Süt',milk,t['milk_fat_pct'],t['milk_protein_pct'],0));rid=cur.lastrowid
                else:
                    risk=m.get('rumen_risk',{}).get('level','Belirsiz')
                    name=f'Akıllı Çözüm {w:.0f}kg {adg:.2f} GCAA {stamp}'; note=f'Besi Hayvanı İhtiyaç Motoru V2 · NASEM 2016 dinamik DMI + net enerji çekirdeği · NEm {t["nem_req_mcal"]:.1f} · NEg {t["neg_req_mcal"]:.1f} Mcal · Göreli rumen asidoz riski: {risk} (klinik pH tahmini değildir)' + ((' · '+warn) if warn else '')
                    cur=c.execute('insert into rations(name,target_group,notes,active,created_at,target_weight_kg,target_adg_kg,animal_type,ration_type,target_milk_l,milk_fat_pct,milk_protein_pct,target_age_months,target_beef_phase) values(?,?,?,1,?,?,?,?,?,?,?,?,?,?)',(name,'Besi',note,datetime.now().isoformat(timespec='seconds'),w,adg,animal_type,'Besi',25,3.8,3.2,age,phase_override));rid=cur.lastrowid
                for feed,kg in zip(rows,qty):
                    if kg>=.01:c.execute('insert into ration_items(ration_id,feed_id,kg_per_head_day) values(?,?,?)',(rid,int(feed['id']),kg));record_ration_item_history(c,rid,int(feed['id']),kg,notes='Hotfix6_16 Akıllı Rasyon')
            result_head=(f'🧠 Rasyon çözüldü · Göreli rumen riski {m.get("rumen_risk",{}).get("level","Belirsiz")}.' if rtype!='Süt' else f'🧠 Süt rasyonu çözüldü · Göreli asidoz riski {m.get("rumen_risk",{}).get("level","Belirsiz")} (klinik pH tahmini değildir).')
            audit(username,'Akıllı rasyon çözdü',name,self.client_ip());return self.redirect('/rations?id='+str(rid),result_head+' '+(warn or 'Otomatik sınırlar içinde çözüm üretildi.'))
        if path=='/ration/create':
            name=(f.get('name') or '').strip()
            if not name:return self.redirect('/rations','Rasyon adı zorunludur.')
            try:
                with db() as c:
                    rtype=(f.get('ration_type') or 'Besi').strip(); cur=c.execute('insert into rations(name,target_group,notes,active,created_at,target_weight_kg,target_adg_kg,animal_type,ration_type,target_milk_l,milk_fat_pct,milk_protein_pct,target_age_months) values(?,?,?,1,?,?,?,?,?,?,?,?,?)',(name,(f.get('target_group') or ('Sağmal' if rtype=='Süt' else 'Besi')).strip(),(f.get('notes') or '').strip(),datetime.now().isoformat(timespec='seconds'),float(f.get('target_weight_kg') or (650 if rtype=='Süt' else 450)),float(f.get('target_adg_kg') or 1.3),(f.get('animal_type') or ('Sağmal İnek' if rtype=='Süt' else 'Besi Erkek')).strip(),rtype,float(f.get('target_milk_l') or 25),3.8,3.2,float(f.get('target_age_months') or 0)));rid=cur.lastrowid
            except sqlite3.IntegrityError:return self.redirect('/rations','Bu rasyon adı zaten kayıtlı.')
            audit(username,'Rasyon oluşturdu',name,self.client_ip());return self.redirect('/rations?id='+str(rid),'Rasyon oluşturuldu. Şimdi yemleri ekleyin.')
        if path=='/ration/target':
            try:
                rid=int(f.get('ration_id') or 0);rtype=(f.get('ration_type') or 'Besi').strip();w=float(f.get('target_weight_kg') or (650 if rtype=='Süt' else 450));adg=float(f.get('target_adg_kg') or 1.3);milk=float(f.get('target_milk_l') or 25);fat=float(f.get('milk_fat_pct') or 3.8);mprot=float(f.get('milk_protein_pct') or 3.2);age=float(f.get('target_age_months') or 0);phase_override=(f.get('target_beef_phase') or 'Otomatik').strip()
            except:return self.redirect('/rations','Hedef bilgileri geçersiz.')
            if rid<=0 or not (150<=w<=900):return self.redirect('/rations?id='+str(rid),'Canlı ağırlık aralık dışında.')
            if rtype=='Süt':
                if not (0<=milk<=70):return self.redirect('/rations?id='+str(rid),'Hedef süt miktarı aralık dışında.')
                with db() as c:c.execute('update rations set ration_type=?,target_weight_kg=?,target_milk_l=?,milk_fat_pct=?,milk_protein_pct=?,animal_type=? where id=?',('Süt',w,milk,fat,mprot,'Sağmal İnek',rid))
                return self.redirect('/rations?id='+str(rid),'Süt rasyonu hedefi güncellendi.')
            if not (0.2<=adg<=2.2):return self.redirect('/rations?id='+str(rid),'Hedef artış aralık dışında.')
            with db() as c:c.execute('update rations set ration_type=?,target_weight_kg=?,target_adg_kg=?,animal_type=?,target_age_months=?,target_beef_phase=? where id=?',('Besi',w,adg,(f.get('animal_type') or 'Besi Erkek').strip(),age,phase_override,rid))
            return self.redirect('/rations?id='+str(rid),'Besi rasyonu hedefi güncellendi.')
        if path=='/ration/delete':
            try:rid=int(f.get('ration_id') or 0)
            except:return self.redirect('/rations','Geçersiz rasyon.')
            if rid<=0:return self.redirect('/rations','Geçersiz rasyon.')
            with db() as c:
                rr=c.execute('select name from rations where id=? and active=1',(rid,)).fetchone()
                if not rr:return self.redirect('/rations','Rasyon bulunamadı.')
                active_pd=c.execute('''select p.name from paddock_rations pr join paddocks p on p.id=pr.paddock_id where pr.ration_id=? and pr.active=1 and (pr.end_date is null or pr.end_date='' or pr.end_date>=?) order by p.name''',(rid,date.today().isoformat())).fetchall()
                if active_pd:return self.redirect('/rations','Bu rasyon aktif olarak '+', '.join(x['name'] for x in active_pd)+' padokunda kullanılıyor. Önce padok atamasını kaldırın.')
                hist=c.execute('select 1 from paddock_rations where ration_id=? limit 1',(rid,)).fetchone()
                if hist:
                    c.execute('update rations set active=0 where id=?',(rid,))
                    audit(username,'Rasyonu arşivledi',rr['name'],self.client_ip())
                    return self.redirect('/rations','Rasyon geçmiş padok kayıtları korunarak arşivlendi.')
                c.execute('delete from ration_item_history where ration_id=?',(rid,))
                c.execute('delete from ration_items where ration_id=?',(rid,))
                c.execute('delete from rations where id=?',(rid,))
            audit(username,'Rasyonu sildi',rr['name'],self.client_ip());return self.redirect('/rations','Rasyon silindi.')
        if path=='/ration/edit':
            try:rid=int(f.get('ration_id') or 0)
            except:return self.redirect('/rations','Geçersiz rasyon.')
            name=(f.get('name') or '').strip()
            if not name:return self.redirect('/rations?id='+str(rid),'Rasyon adı zorunludur.')
            with db() as c:c.execute('update rations set name=?,target_group=?,notes=? where id=?',(name,(f.get('target_group') or 'Besi').strip(),(f.get('notes') or '').strip(),rid))
            return self.redirect('/rations?id='+str(rid),'Rasyon bilgileri güncellendi.')
        if path=='/ration/items-bulk':
            try:rid=int(f.get('ration_id') or 0)
            except:return self.redirect('/rations','Geçersiz rasyon.')
            if rid<=0:return self.redirect('/rations','Geçersiz rasyon.')
            with db() as c:
                valid={int(r['id']) for r in c.execute('select id from ration_items where ration_id=?',(rid,)).fetchall()}
                for key,val in f.items():
                    if not key.startswith('item_'):continue
                    try:iid=int(key[5:]);kg=max(0.0,float(str(val).replace(',','.')))
                    except:continue
                    if iid not in valid:continue
                    item=c.execute('select feed_id from ration_items where id=? and ration_id=?',(iid,rid)).fetchone()
                    if not item:continue
                    if kg<0.001:c.execute('delete from ration_items where id=? and ration_id=?',(iid,rid))
                    else:c.execute('update ration_items set kg_per_head_day=? where id=? and ration_id=?',(round(kg,3),iid,rid))
                    record_ration_item_history(c,rid,item['feed_id'],0 if kg<0.001 else round(kg,3),notes='Çalışma masası')
            return self.redirect('/rations?id='+str(rid),'Rasyon miktarları kaydedildi.')
        if path=='/ration/item-adjust':
            try:rid=int(f.get('ration_id') or 0);iid=int(f.get('item_id') or 0);delta=float(f.get('delta') or 0)
            except:return self.redirect('/rations','Geçersiz rasyon kalemi.')
            with db() as c:
                row=c.execute('select kg_per_head_day from ration_items where id=? and ration_id=?',(iid,rid)).fetchone()
                if not row:return self.redirect('/rations?id='+str(rid),'Yem kalemi bulunamadı.')
                new=max(0,float(row['kg_per_head_day'] or 0)+delta)
                item=c.execute('select feed_id from ration_items where id=?',(iid,)).fetchone()
                if new<0.001:c.execute('delete from ration_items where id=?',(iid,))
                else:c.execute('update ration_items set kg_per_head_day=? where id=?',(round(new,3),iid))
                if item:record_ration_item_history(c,rid,item['feed_id'],0 if new<0.001 else round(new,3),notes='Hızlı miktar ayarı')
            return self.redirect('/rations?id='+str(rid),'Rasyon miktarı güncellendi.')
        if path=='/ration/apply-combo':
            try:
                rid=int(f.get('ration_id') or 0);rfid=int(f.get('red_feed_id') or 0);rd=float(f.get('red_delta') or 0);afid=int(f.get('add_feed_id') or 0);ad=float(f.get('add_delta') or 0)
            except:return self.redirect('/rations','Kombine öneri uygulanamadı.')
            if rid<=0 or rfid<=0 or afid<=0 or rd>=0 or ad<=0:return self.redirect('/rations?id='+str(rid),'Kombine öneri geçersiz.')
            with db() as c:
                row=c.execute('select id,kg_per_head_day from ration_items where ration_id=? and feed_id=?',(rid,rfid)).fetchone()
                if not row:return self.redirect('/rations?id='+str(rid),'Azaltılacak yem rasyonda bulunamadı.')
                newkg=max(0.0,float(row['kg_per_head_day'] or 0)+rd)
                if newkg<0.001:c.execute('delete from ration_items where id=?',(row['id'],))
                else:c.execute('update ration_items set kg_per_head_day=? where id=?',(round(newkg,3),row['id']))
                record_ration_item_history(c,rid,rfid,0 if newkg<0.001 else round(newkg,3),notes='Kombine akıllı dengeleme - azalt')
                addrow=c.execute('select id,kg_per_head_day from ration_items where ration_id=? and feed_id=?',(rid,afid)).fetchone()
                if addrow:
                    addkg=max(0.0,float(addrow['kg_per_head_day'] or 0)+ad);c.execute('update ration_items set kg_per_head_day=? where id=?',(round(addkg,3),addrow['id']))
                else:
                    addkg=ad;c.execute('insert into ration_items(ration_id,feed_id,kg_per_head_day) values(?,?,?)',(rid,afid,round(addkg,3)))
                record_ration_item_history(c,rid,afid,round(addkg,3),notes='Kombine akıllı dengeleme - ekle')
            return self.redirect('/rations?id='+str(rid)+'#smart-balance','Kombine dengelemedeki iki değişiklik birlikte uygulandı.')
        if path=='/ration/apply-suggestion':
            try:rid=int(f.get('ration_id') or 0);fid=int(f.get('feed_id') or 0);delta=float(f.get('delta') or 0)
            except:return self.redirect('/rations','Öneri uygulanamadı.')
            if rid<=0 or fid<=0 or abs(delta)<0.001:return self.redirect('/rations?id='+str(rid),'Öneri miktarı geçersiz.')
            with db() as c:
                row=c.execute('select id,kg_per_head_day from ration_items where ration_id=? and feed_id=?',(rid,fid)).fetchone()
                if row:
                    newkg=max(0.0,float(row['kg_per_head_day'] or 0)+delta)
                    if newkg<0.001:c.execute('delete from ration_items where id=?',(row['id'],))
                    else:c.execute('update ration_items set kg_per_head_day=? where id=?',(round(newkg,3),row['id']))
                    record_ration_item_history(c,rid,fid,0 if newkg<0.001 else round(newkg,3),notes='Akıllı dengeleme önerisi')
                elif delta>0:
                    c.execute('insert into ration_items(ration_id,feed_id,kg_per_head_day) values(?,?,?)',(rid,fid,round(delta,3)))
                    record_ration_item_history(c,rid,fid,round(delta,3),notes='Akıllı dengeleme önerisi')
                else:return self.redirect('/rations?id='+str(rid),'Azaltılacak yem rasyonda bulunamadı.')
            return self.redirect('/rations?id='+str(rid)+'#smart-balance','Simülasyondaki değişiklik rasyona uygulandı.')
        if path=='/ration/item':
            try:rid=int(f.get('ration_id') or 0);fid=int(f.get('feed_id') or 0);kg=float(f.get('kg_per_head_day') or 0)
            except:return self.redirect('/rations','Rasyon kalemi geçersiz.')
            if rid<=0 or fid<=0 or kg<=0:return self.redirect('/rations?id='+str(rid),'Rasyon kalemi geçersiz.')
            with db() as c:
                c.execute('''insert into ration_items(ration_id,feed_id,kg_per_head_day) values(?,?,?) on conflict(ration_id,feed_id) do update set kg_per_head_day=excluded.kg_per_head_day''',(rid,fid,kg))
                record_ration_item_history(c,rid,fid,kg,notes='Rasyona yem ekle/güncelle')
            keep=(f.get('keep_feed_add_open') or '').strip()=='1'
            return self.redirect('/rations?id='+str(rid)+('&feedadd=1#ration-workbench' if keep else '#ration-workbench'),'Rasyon kalemi güncellendi.')
        if path=='/ration/item-delete':
            try:iid=int(f.get('id') or 0);rid=int(f.get('ration_id') or 0)
            except:return self.redirect('/rations','Geçersiz kayıt.')
            with db() as c:
                item=c.execute('select feed_id from ration_items where id=? and ration_id=?',(iid,rid)).fetchone()
                c.execute('delete from ration_items where id=?',(iid,))
                if item:record_ration_item_history(c,rid,item['feed_id'],0,notes='Yem rasyondan çıkarıldı')
            return self.redirect('/rations?id='+str(rid),'Yem rasyondan çıkarıldı.')
        if path=='/ration/assign':
            try:pid=int(f.get('paddock_id') or 0);rid=int(f.get('ration_id') or 0)
            except:return self.redirect('/paddocks','Padok ve rasyon seçin.')
            if pid<=0 or rid<=0:return self.redirect('/paddocks','Padok ve rasyon seçin.')
            start=(f.get('start_date') or date.today().isoformat()).strip()
            with db() as c:
                if not c.execute('select id from paddocks where id=? and active=1',(pid,)).fetchone():return self.redirect('/paddocks','Padok bulunamadı.')
                if not c.execute('select id from rations where id=? and active=1',(rid,)).fetchone():return self.redirect('/paddocks','Rasyon bulunamadı.')
                c.execute("update paddock_rations set active=0,end_date=? where paddock_id=? and active=1 and (end_date is null or end_date='')",((date.fromisoformat(start)-timedelta(days=1)).isoformat(),pid))
                c.execute('insert into paddock_rations(paddock_id,ration_id,start_date,end_date,active,notes) values(?,?,?,NULL,1,?)',(pid,rid,start,(f.get('notes') or '').strip()))
            audit(username,'Padoka rasyon atadı',f'Padok {pid} / Rasyon {rid}',self.client_ip());return self.redirect('/paddocks','Rasyon padoka atandı.')
        if path=='/smtp-settings':
            if not self.require_admin():return
            action=(f.get('action') or 'save').strip()
            if action=='test':
                target=(f.get('test_email') or '').strip()
                if not target:return self.redirect('/smtp-settings','Test e-posta adresi girin.')
                try:send_reset_email(target,'SMTP Test','123456');return self.redirect('/smtp-settings','✅ Test e-postası gönderildi.')
                except Exception as exc:return self.redirect('/smtp-settings','Test e-postası gönderilemedi: '+str(exc))
            setting_set('smtp_host',(f.get('smtp_host') or '').strip());setting_set('smtp_port',(f.get('smtp_port') or '587').strip());setting_set('smtp_security',(f.get('smtp_security') or 'starttls').strip());setting_set('smtp_username',(f.get('smtp_username') or '').strip());setting_set('smtp_sender',(f.get('smtp_sender') or '').strip())
            if (f.get('smtp_password') or '').strip():setting_set('smtp_password',f.get('smtp_password').strip())
            audit(username,'SMTP ayarlarını güncelledi','',self.client_ip());return self.redirect('/smtp-settings','E-posta ayarları kaydedildi.')

        if path=='/dashboard-layout':
            try: slot=int(f.get('slot','-1'))
            except Exception: slot=-1
            valid={x[0] for x in DASHBOARD_CARD_OPTIONS}
            card_key=(f.get('card_key') or '').strip()
            if slot<0 or slot>=8:return self.redirect('/?edit=1','Geçersiz Dashboard yuvası.')
            if card_key and card_key not in valid:return self.redirect('/?edit=1','Geçersiz Dashboard kartı.')
            layout=dashboard_layout(username)
            layout[slot]=card_key
            with db() as c:
                c.execute("insert into settings(setting_key,setting_value) values(?,?) on conflict(setting_key) do update set setting_value=excluded.setting_value",('dashboard_layout_'+username,','.join(layout)))
            return self.redirect('/?edit=1','Dashboard kartı güncellendi.')
        if path=='/farm-profile':
            if not self.require_admin():return
            text_keys=('farm_name','owner_name','phone','email','province','district','address','business_no','tax_or_tc','vet_name','vet_phone','vet_email','notes')
            with db() as c:
                for key in text_keys:
                    val=(f.get(key,'') or '').strip()
                    c.execute("insert into settings(setting_key,setting_value) values(?,?) on conflict(setting_key) do update set setting_value=excluded.setting_value",(key,val))
                current=c.execute("select setting_value from settings where setting_key='farm_logo'").fetchone()
                current_logo=current['setting_value'] if current else ''
                if f.get('remove_logo')=='1':
                    if current_logo.startswith('/uploads/'):
                        try:(UPLOADS/os.path.basename(current_logo)).unlink(missing_ok=True)
                        except Exception:pass
                    c.execute("insert into settings(setting_key,setting_value) values('farm_logo','') on conflict(setting_key) do update set setting_value=''")
                    current_logo=''
                upload=f.get('farm_logo_file')
                if isinstance(upload,dict) and upload.get('filename') and upload.get('content'):
                    content=upload['content']
                    if len(content)>5*1024*1024:return self.redirect('/farm-profile','Logo dosyası 5 MB sınırını aşıyor.')
                    ext=Path(upload['filename']).suffix.lower()
                    if ext not in ('.jpg','.jpeg','.png','.webp'):return self.redirect('/farm-profile','Logo yalnızca JPG, PNG veya WebP olabilir.')
                    UPLOADS.mkdir(parents=True,exist_ok=True)
                    if current_logo.startswith('/uploads/'):
                        try:(UPLOADS/os.path.basename(current_logo)).unlink(missing_ok=True)
                        except Exception:pass
                    filename='farm_logo'+('.jpg' if ext=='.jpeg' else ext)
                    (UPLOADS/filename).write_bytes(content)
                    logo_url='/uploads/'+filename
                    c.execute("insert into settings(setting_key,setting_value) values('farm_logo',?) on conflict(setting_key) do update set setting_value=excluded.setting_value",(logo_url,))
            audit(username,'Çiftlik profilini güncelledi',(f.get('farm_name') or 'ÇiftlikPro').strip(),self.client_ip())
            return self.redirect('/farm-profile','Çiftlik profili başarıyla kaydedildi.')
        if path=='/password-change':
            np=f.get('new_password','')
            if len(np)<8:return self.redirect('/password-change','Yeni şifre en az 8 karakter olmalıdır.')
            if np!=f.get('new_password_confirm',''):return self.redirect('/password-change','Yeni şifreler eşleşmiyor.')
            with db() as c:r=c.execute('select * from users where username=?',(username,)).fetchone()
            if not r or not password_verify(f.get('current_password',''),r['password']):return self.redirect('/password-change','Mevcut şifre hatalı.')
            with db() as c:c.execute('update users set password=?,password_changed_at=? where id=?',(password_hash(np),datetime.now().strftime('%Y-%m-%d %H:%M:%S'),r['id']))
            audit(username,'Şifresini değiştirdi','',self.client_ip());return self.redirect('/password-change','Şifreniz başarıyla değiştirildi.')
        if path=='/users/create':
            if not self.require_admin():return
            uname=(f.get('username') or '').strip();password=f.get('password','');role=f.get('role','personel')
            if len(password)<8:return self.redirect('/users','Şifre en az 8 karakter olmalıdır.')
            try:
                with db() as c:c.execute('insert into users(username,password,role,full_name,recovery_email,active,password_changed_at) values(?,?,?,?,?,1,?)',(uname,password_hash(password),role,f.get('full_name',''),(f.get('recovery_email') or '').strip(),datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
            except sqlite3.IntegrityError:return self.redirect('/users','Bu kullanıcı adı zaten kullanılıyor.')
            audit(username,'Kullanıcı oluşturdu',uname+' · '+role,self.client_ip());return self.redirect('/users','Kullanıcı oluşturuldu.')
        if path=='/users/update':
            if not self.require_admin():return
            uid=int(f.get('id') or 0);role=f.get('role','personel');active=1 if f.get('active')=='1' else 0
            with db() as c:r=c.execute('select * from users where id=?',(uid,)).fetchone()
            if not r:return self.redirect('/users','Kullanıcı bulunamadı.')
            if r['role']=='admin' and r['active'] and (role!='admin' or not active) and active_admin_count()<=1:return self.redirect('/users/edit?id='+str(uid),'Son aktif yönetici değiştirilemez.')
            with db() as c:
                c.execute('update users set full_name=?,recovery_email=?,role=?,active=? where id=?',(f.get('full_name',''),(f.get('recovery_email') or '').strip(),role,active,uid))
                if f.get('new_password'):
                    if len(f['new_password'])<8:return self.redirect('/users/edit?id='+str(uid),'Yeni şifre en az 8 karakter olmalıdır.')
                    c.execute('update users set password=?,password_changed_at=? where id=?',(password_hash(f['new_password']),datetime.now().strftime('%Y-%m-%d %H:%M:%S'),uid))
            audit(username,'Kullanıcı güncelledi',r['username'],self.client_ip());return self.redirect('/users','Kullanıcı güncellendi.')
        if path=='/backup/restore':
            if not self.require_admin():return
            upload=f.get('backup_file')
            if f.get('confirm_restore')!='yes':return self.redirect('/backups','Geri yükleme onayı verilmedi.')
            if not upload or not isinstance(upload,dict):return self.redirect('/backups','ZIP dosyası seçilmedi.')
            temp_path=BACKUPS/('.incoming_'+secrets.token_hex(8)+'.zip');temp_path.write_bytes(upload.get('content',b''))
            try:
                emergency,manifest=restore_backup_zip(temp_path);audit(username,'Yedek geri yükledi',upload.get('filename','')+' · '+emergency,self.client_ip());SESSIONS.clear();return self.redirect('/login','Yedek geri yüklendi. Yeniden giriş yapın.')
            except Exception as exc:return self.redirect('/backups','Geri yükleme başarısız: '+str(exc))
            finally:
                if temp_path.exists():temp_path.unlink()

        if path=='/animal-delete':
            if not self.require_admin():return
            aid=(f.get('id') or '').strip()
            try:
                with db() as c:
                    rec=c.execute('select tag,gender from animals where id=?',(aid,)).fetchone()
                    if not rec:return self.redirect('/','Hayvan kaydı bulunamadı.')
                    photo_rows=c.execute('select filename from animal_photos where animal_id=?',(aid,)).fetchall()
                    for table in ('inseminations','health','finance','weights','milk','animal_photos'):
                        c.execute(f'delete from {table} where animal_id=?',(aid,))
                    c.execute('delete from calves where mother_id=?',(aid,))
                    c.execute('delete from animals where id=?',(aid,))
                for pr in photo_rows:
                    fp=UPLOADS/pr['filename']
                    if fp.exists():
                        try:fp.unlink()
                        except Exception:pass
                audit(username,'Hayvan sildi',rec['tag'],self.client_ip())
                return self.redirect('/males' if rec['gender']=='Erkek' else '/animals','Hayvan ve bağlı kayıtları silindi.')
            except Exception as exc:return self.redirect('/','Silme hatası: '+str(exc))
        if path=='/calf-delete':
            if not self.require_admin():return
            cid=(f.get('id') or '').strip()
            try:
                with db() as c:
                    rec=c.execute('select tag from calves where id=?',(cid,)).fetchone()
                    if not rec:return self.redirect('/calves','Buzağı kaydı bulunamadı.')
                    c.execute('delete from health where calf_id=?',(cid,))
                    c.execute('delete from calf_weights where calf_id=?',(cid,))
                    c.execute('delete from calf_photos where calf_id=?',(cid,))
                    c.execute('delete from calves where id=?',(cid,))
                audit(username,'Buzağı sildi',rec['tag'],self.client_ip())
                return self.redirect('/calves','Buzağı kaydı silindi.')
            except Exception as exc:return self.redirect('/calves','Silme hatası: '+str(exc))

        if path=='/animal-edit':
            aid=(f.get('id') or '').strip()
            try:
                with db() as c:
                    rec=c.execute('select * from animals where id=?',(aid,)).fetchone()
                    if not rec:return self.redirect('/','Hayvan kaydı bulunamadı.')
                    tag=(f.get('tag') or '').strip()
                    duplicate=c.execute('select id from animals where tag=? and id<>?',(tag,aid)).fetchone()
                    calf_duplicate=c.execute('select id from calves where tag=?',(tag,)).fetchone()
                    if duplicate or calf_duplicate:return self.redirect('/animal-edit?id='+aid,'Bu küpe numarası başka bir kayıtta kullanılıyor.')
                    photo_url=f.get('photo_url','')
                    upload=f.get('photo_file')
                    if upload and isinstance(upload,dict) and upload.get('content'):
                        ext=Path(upload['filename']).suffix.lower()
                        if ext not in ('.jpg','.jpeg','.png','.webp','.gif'):return self.redirect('/animal-edit?id='+aid,'Desteklenmeyen fotoğraf biçimi.')
                        if len(upload['content'])>10*1024*1024:return self.redirect('/animal-edit?id='+aid,'Fotoğraf 10 MB sınırını aşıyor.')
                        name=save_optimized_upload('animal_edit',upload)
                        photo_url='/uploads/'+name
                    gender=f.get('gender') if f.get('gender') in ('Dişi','Erkek') else rec['gender']
                    c.execute('update animals set tag=?,nickname=?,gender=?,breed=?,birth_date=?,notes=?,paddock=?,photo_url=?,sold_price=?,status=?,purchase_date=?,purchase_price=?,purchase_weight=?,daily_feed_cost=?,daily_care_cost=?,target_sale_price=? where id=?',
                              (tag,f.get('nickname',''),gender,f.get('breed',''),f.get('birth_date',''),f.get('notes',''),f.get('paddock',''),photo_url,float(f.get('sold_price') or 0),f.get('status') or 'Aktif',f.get('purchase_date',''),float(f.get('purchase_price') or 0),float(f.get('purchase_weight') or 0),float(f.get('daily_feed_cost') or 0),float(f.get('daily_care_cost') or 0),float(f.get('target_sale_price') or 0),aid))
                audit(username,'Hayvan düzenledi',tag,self.client_ip())
                return self.redirect('/animals' if gender=='Dişi' else '/males','Hayvan başarıyla güncellendi.')
            except sqlite3.IntegrityError:
                return self.redirect('/animal-edit?id='+aid,'Bu küpe numarası zaten kullanılıyor.')
            except Exception as exc:
                return self.redirect('/animal-edit?id='+aid,'Güncelleme hatası: '+str(exc))
        if path=='/calf-edit':
            cid=(f.get('id') or '').strip()
            try:
                with db() as c:
                    rec=c.execute('select * from calves where id=?',(cid,)).fetchone()
                    if not rec:return self.redirect('/calves','Buzağı kaydı bulunamadı.')
                    mother=c.execute("select id from animals where id=? and gender='Dişi' and coalesce(status,'Aktif')='Aktif'",(f.get('mother_id'),)).fetchone()
                    if not mother:return self.redirect('/calf-edit?id='+cid,'Anne olarak aktif bir dişi hayvan seçilmelidir.')
                    tag=(f.get('tag') or '').strip()
                    if c.execute('select id from calves where tag=? and id<>?',(tag,cid)).fetchone() or c.execute('select id from animals where tag=?',(tag,)).fetchone():return self.redirect('/calf-edit?id='+cid,'Bu küpe numarası başka bir kayıtta kullanılıyor.')
                    photo_url=f.get('photo_url') or rec['photo_url'] or '';upload=f.get('photo_file')
                    if upload and isinstance(upload,dict) and upload.get('content'):
                        ext=Path(upload['filename']).suffix.lower()
                        if ext not in ('.jpg','.jpeg','.png','.webp','.gif'):return self.redirect('/calf-edit?id='+cid,'Desteklenmeyen fotoğraf biçimi.')
                        if len(upload['content'])>10*1024*1024:return self.redirect('/calf-edit?id='+cid,'Fotoğraf 10 MB sınırını aşıyor.')
                        name=save_optimized_upload(f'calf_{cid}',upload);photo_url='/uploads/'+name
                        c.execute('insert into calf_photos(calf_id,filename,created_at,caption) values(?,?,?,?)',(cid,name,datetime.now().strftime('%Y-%m-%d %H:%M:%S'),'Profil fotoğrafı'))
                    c.execute('update calves set tag=?,nickname=?,mother_id=?,father_tag=?,birth_date=?,gender=?,breed=?,paddock=?,photo_url=?,purchase_date=?,purchase_price=?,purchase_payment_method=?,notes=? where id=?',
                        (tag,f.get('nickname',''),f.get('mother_id'),f.get('father_tag',''),f.get('birth_date',''),f.get('gender','Dişi'),f.get('breed',''),f.get('paddock',''),photo_url,f.get('purchase_date',''),float(f.get('purchase_price') or 0),f.get('purchase_payment_method') or 'Nakit',f.get('notes',''),cid))
                audit(username,'Buzağı düzenledi',tag,self.client_ip());return self.redirect('/calf?id='+cid,'Buzağı başarıyla güncellendi.')
            except sqlite3.IntegrityError:return self.redirect('/calf-edit?id='+cid,'Bu küpe numarası zaten kullanılıyor.')
            except Exception as exc:return self.redirect('/calf-edit?id='+cid,'Güncelleme hatası: '+str(exc))
        if path=='/animal-add':
            kind=(f.get('record_type') or '').strip();tag=(f.get('tag') or '').strip()
            if not tag:return self.redirect('/animal-add','Küpe numarası zorunludur.')
            try:
                with db() as c:
                    if c.execute('select 1 from animals where tag=?',(tag,)).fetchone() or c.execute('select 1 from calves where tag=?',(tag,)).fetchone():return self.redirect('/animal-add','Bu küpe numarası zaten kayıtlı.')
                    if kind in ('Dişi','Erkek'):
                        photo_url='';upload=f.get('photo_file')
                        if upload and isinstance(upload,dict) and upload.get('content'):
                            ext=Path(upload['filename']).suffix.lower()
                            if ext not in ('.jpg','.jpeg','.png','.webp','.gif'):return self.redirect('/animal-add','Desteklenmeyen fotoğraf biçimi.')
                            if len(upload['content'])>10*1024*1024:return self.redirect('/animal-add','Fotoğraf 10 MB sınırını aşıyor.')
                            name=save_optimized_upload('animal_new',upload);photo_url='/uploads/'+name
                        cur=c.execute('insert into animals(tag,nickname,gender,breed,birth_date,notes,paddock,photo_url,sold_price,status,purchase_date,purchase_price,purchase_weight,daily_feed_cost,daily_care_cost,target_sale_price) values(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)',(tag,f.get('nickname',''),kind,f.get('breed',''),f.get('birth_date',''),f.get('notes',''),f.get('paddock',''),photo_url,0,'Aktif',f.get('purchase_date',''),float(f.get('purchase_price') or 0),float(f.get('purchase_weight') or 0) if kind=='Erkek' else 0,float(f.get('daily_feed_cost') or 0) if kind=='Erkek' else 0,float(f.get('daily_care_cost') or 0) if kind=='Erkek' else 0,float(f.get('target_sale_price') or 0) if kind=='Erkek' else 0))
                        aid=cur.lastrowid
                        if photo_url:c.execute('insert into animal_photos(animal_id,filename,created_at,caption) values(?,?,?,?)',(aid,photo_url.split('/uploads/',1)[1],datetime.now().strftime('%Y-%m-%d %H:%M:%S'),'Profil fotoğrafı'))
                        if kind=='Dişi' and (f.get('entry_pregnancy_status') or '')=='Gebe':
                            mode=(f.get('pregnancy_info_mode') or 'date').strip()
                            entry_date=(f.get('pregnancy_entry_date') or date.today().isoformat()).strip()
                            try:ref_day=date.fromisoformat(entry_date)
                            except:ref_day=date.today();entry_date=ref_day.isoformat()
                            age_months=0.0
                            if mode=='date':
                                ins_date=(f.get('known_insemination_date') or '').strip()
                                if not ins_date:return self.redirect('/animal-add','Gebe hayvan için son tohumlama tarihini girin.')
                                try:ins_day=date.fromisoformat(ins_date)
                                except:return self.redirect('/animal-add','Tohumlama tarihi geçersiz.')
                                due=(ins_day+timedelta(days=280)).isoformat()
                                result='Gebe (Satın Alındığında · Tohumlama Tarihi Biliniyor)'
                                source='Satın Alındığında Gebe · Tohumlama Tarihi'
                            else:
                                try:age_months=float(f.get('pregnancy_age_months') or 0)
                                except:age_months=0
                                if age_months<=0 or age_months>9:return self.redirect('/animal-add','Gebelik yaşını 1-9 ay arasında girin.')
                                elapsed=round(age_months*(280/9))
                                ins_day=ref_day-timedelta(days=elapsed);ins_date=ins_day.isoformat()
                                due=(ins_day+timedelta(days=280)).isoformat()
                                result='Gebe (Satın Alındığında · Gebelik Yaşından Tahmini)'
                                source='Satın Alındığında Gebe · Yaklaşık'
                            c.execute('insert into inseminations(animal_id,attempt,insemination_date,pregnancy_result,due_date) values(?,?,?,?,?)',(aid,1,ins_date,result,due))
                            c.execute('update animals set pregnancy_source=?,pregnancy_age_months_at_entry=?,pregnancy_entry_date=? where id=?',(source,age_months,entry_date,aid))

                        purchase_price=float(f.get('purchase_price') or 0)
                        if purchase_price>0:
                            purchase_date=(f.get('purchase_date') or date.today().isoformat()).strip()
                            payment=(f.get('purchase_payment_method') or 'Nakit').strip()
                            desc=f'Otomatik hayvan alımı · {tag}'
                            c.execute('insert into finance(tx_date,tx_type,category,amount,description,payment_method,animal_id,created_at,animal_status_action) values(?,?,?,?,?,?,?,?,?)',
                                      (purchase_date,'Gider','Hayvan Alımı',purchase_price,desc,payment,aid,datetime.now().isoformat(),''))
                            audit(username,'Hayvan alımı finansa otomatik işlendi',f'{tag} · {money(purchase_price)} · {payment}',self.client_ip())

                        return self.redirect('/animals' if kind=='Dişi' else '/males',kind+' hayvan başarıyla kaydedildi.'+(' Alış bedeli Finans > Hayvan Alımı giderine otomatik işlendi.' if purchase_price>0 else ''))
                    if kind=='Buzağı':
                        mt=(f.get('mother_tag') or '').strip();bd=(f.get('birth_date') or '').strip()
                        if not mt:return self.redirect('/animal-add','Buzağı kaydı için anne küpesi zorunludur.')
                        if not bd:return self.redirect('/animal-add','Buzağı kaydı için doğum tarihi zorunludur.')
                        mother=c.execute("select id from animals where tag=? and gender='Dişi' and coalesce(status,'Aktif')='Aktif'",(mt,)).fetchone()
                        if not mother:return self.redirect('/animal-add','Anne küpesi aktif dişi hayvanlarda bulunamadı.')
                        c.execute('insert into calves(tag,mother_id,father_tag,birth_date,gender,notes) values(?,?,?,?,?,?)',(tag,mother['id'],f.get('father_tag',''),bd,f.get('calf_gender','Dişi'),f.get('notes','')))
                        return self.redirect('/calves','Buzağı başarıyla kaydedildi.')
                return self.redirect('/animal-add','Geçersiz kayıt türü.')
            except sqlite3.IntegrityError:return self.redirect('/animal-add','Bu küpe numarası zaten kayıtlı.')
            except Exception as exc:return self.redirect('/animal-add','Kayıt hatası: '+str(exc))

        if path=='/backup/settings':
            if not self.require_admin():return
            raw=(f.get('backup_directory') or '').strip()
            if not raw:
                set_setting_value('backup_directory','')
                return self.redirect('/backups','Yedek klasörü varsayılan konuma döndürüldü.')
            try:
                target=Path(os.path.expandvars(os.path.expanduser(raw)))
                target.mkdir(parents=True,exist_ok=True)
                probe=target/'.ciftlikpro_write_test.tmp'
                probe.write_text('ok',encoding='utf-8');probe.unlink()
            except Exception as exc:
                return self.redirect('/backups','Yedek klasörü kullanılamıyor: '+str(exc))
            set_setting_value('backup_directory',str(target))
            audit(username,'Yedek klasörünü değiştirdi',str(target),self.client_ip())
            return self.redirect('/backups','Yedek klasörü kaydedildi ve yazma testi başarılı.')
        if path=='/backup/restore-existing':
            if not self.require_admin():return
            name=os.path.basename((f.get('file') or '').strip())
            fp=configured_backup_dir()/name
            if not fp.exists():return self.redirect('/backups','Yedek dosyası bulunamadı.')
            try:
                restore_backup_zip(fp)
                audit(username,'Klasörden yedek geri yükledi',name,self.client_ip())
                return self.redirect('/backups','Yedek başarıyla geri yüklendi.')
            except Exception as exc:
                return self.redirect('/backups','Geri yükleme hatası: '+str(exc))
        if path=='/photos/optimize-existing':
            if not self.require_admin():return
            try:
                safety=create_backup('FotoOptimizasyonOncesi')
                result=optimize_existing_uploads()
                audit(username,'Fotoğrafları optimize etti',f"{result['count']} dosya · {format_bytes(result['saved'])} kazanım",self.client_ip())
                return self.redirect('/backups',f"Optimizasyon tamamlandı: {result['count']} fotoğraf küçültüldü, {format_bytes(result['saved'])} alan kazanıldı. Güvenlik yedeği: {safety}")
            except Exception as exc:
                return self.redirect('/backups','Fotoğraf optimizasyonu başarısız: '+str(exc))
        if path=='/data/import':
            try:
                upload=f.get('json_file')
                if not upload or not isinstance(upload,dict):return self.redirect('/data','JSON dosyası seçilmedi.')
                payload=json.loads(upload['content'].decode('utf-8-sig'))
                create_backup('import_oncesi')
                stats=import_payload(payload,f.get('strategy','skip'))
                summary=f"Aktarım tamamlandı: {stats['animals']} yeni hayvan, {stats['animals_updated']} güncellenen, {stats['inseminations']} tohumlama, {stats['calves']} buzağı, {stats['finance']} finans, {stats['skipped']} atlanan"
                if stats['errors']:summary+=f", {len(stats['errors'])} hata"
                return self.redirect('/data',summary)
            except Exception as e:return self.redirect('/data','İçe aktarma hatası: '+str(e))
        if path=='/calf/photo':
            cid=(f.get('calf_id') or '').strip()
            try:
                upload=f.get('photo_file')
                if not upload or not isinstance(upload,dict):return self.redirect('/calf?id='+cid,'Fotoğraf seçilmedi.')
                ext=Path(upload['filename']).suffix.lower()
                if ext not in ('.jpg','.jpeg','.png','.webp','.gif'):return self.redirect('/calf?id='+cid,'Desteklenmeyen fotoğraf biçimi.')
                if len(upload['content'])>10*1024*1024:return self.redirect('/calf?id='+cid,'Fotoğraf 10 MB sınırını aşıyor.')
                name=save_optimized_upload(f'calf_{cid}',upload)
                with db() as c:
                    if not c.execute('select 1 from calves where id=?',(cid,)).fetchone():return self.redirect('/calves','Buzağı bulunamadı.')
                    c.execute('insert into calf_photos(calf_id,filename,created_at,caption) values(?,?,?,?)',(cid,name,datetime.now().strftime('%Y-%m-%d %H:%M:%S'),f.get('caption','')))
                    c.execute('update calves set photo_url=? where id=?',('/uploads/'+name,cid))
                return self.redirect('/calf?id='+cid,'Fotoğraf başarıyla yüklendi.')
            except Exception as e:return self.redirect('/calf?id='+cid,'Fotoğraf yükleme hatası: '+str(e))
        if path=='/calf/weight':
            cid=(f.get('calf_id') or '').strip()
            try:
                weight=float(f.get('weight') or 0)
                if weight<=0:return self.redirect('/calf?id='+cid,'Kilo sıfırdan büyük olmalıdır.')
                with db() as c:
                    if not c.execute('select 1 from calves where id=?',(cid,)).fetchone():return self.redirect('/calves','Buzağı bulunamadı.')
                    existing=c.execute('select id from calf_weights where calf_id=? and measure_date=?',(cid,f.get('measure_date'))).fetchone()
                    if existing:c.execute('update calf_weights set weight=?,notes=? where id=?',(weight,f.get('notes',''),existing['id']))
                    else:c.execute('insert into calf_weights(calf_id,measure_date,weight,notes) values(?,?,?,?)',(cid,f.get('measure_date'),weight,f.get('notes','')))
                return self.redirect('/calf?id='+cid,'Buzağı tartımı kaydedildi.')
            except Exception as e:return self.redirect('/calf?id='+cid,'Tartım kayıt hatası: '+str(e))
        if path=='/animal/photo':
            try:
                upload=f.get('photo_file'); aid=f.get('animal_id','')
                if not upload or not isinstance(upload,dict): return self.redirect('/animal?id='+aid,'Fotoğraf seçilmedi.')
                ext=Path(upload['filename']).suffix.lower()
                if ext not in ('.jpg','.jpeg','.png','.webp','.gif'): return self.redirect('/animal?id='+aid,'Desteklenmeyen fotoğraf biçimi.')
                if len(upload['content'])>10*1024*1024: return self.redirect('/animal?id='+aid,'Fotoğraf 10 MB sınırını aşıyor.')
                name=save_optimized_upload(f'animal_{aid}',upload)
                with db() as c:
                    c.execute('insert into animal_photos(animal_id,filename,created_at,caption) values(?,?,?,?)',(aid,name,datetime.now().strftime('%Y-%m-%d %H:%M:%S'),f.get('caption','')))
                    c.execute('update animals set photo_url=? where id=?',('/uploads/'+name,aid))
                return self.redirect('/animal?id='+aid,'Fotoğraf başarıyla yüklendi.')
            except Exception as e: return self.redirect('/animal?id='+f.get('animal_id',''),'Fotoğraf yükleme hatası: '+str(e))
        try:
            with db() as c:
                if path=='/animal/sale':
                    aid=f['animal_id']; sale_price=float(f['sale_price']); sale_date=f['sale_date']; sale_weight=float(f.get('sale_weight') or 0)
                    a=c.execute("select * from animals where id=? and gender='Erkek'",(aid,)).fetchone()
                    if not a:return self.redirect('/males','Erkek hayvan bulunamadı.')
                    if a['status']!='Aktif':return self.redirect('/animal?id='+aid,'Bu hayvan daha önce aktif sürüden çıkarılmış.')
                    days,daily,accumulated,current=animal_cost_values(a); profit=sale_price-current
                    desc=(f.get('description') or '').strip()
                    detail=f"{desc} | Satış kilosu: {sale_weight:.1f} kg | Anlık maliyet: {current:.2f} TL | Net kâr/zarar: {profit:.2f} TL"
                    c.execute('insert into finance(tx_date,tx_type,category,amount,description,payment_method,animal_id,created_at,animal_status_action) values(?,?,?,?,?,?,?,?,?)',(sale_date,'Gelir','Hayvan Satışı',sale_price,detail,'Nakit',aid,datetime.now().isoformat(),'Satıldı'))
                    c.execute('update animals set status=?,exit_date=?,exit_reason=?,sold_price=? where id=?',('Satıldı',sale_date,'Hayvan Satışı',sale_price,aid))
                    if sale_weight>0:c.execute('insert into weights(animal_id,measure_date,weight,notes) values(?,?,?,?)',(aid,sale_date,sale_weight,'Satış kilosu'))
                    return self.redirect('/archive/sold','Satış tamamlandı. Net kâr/zarar: '+money(profit))
                if path=='/performance-settings':
                    target=max(0.01,float(f['male_min_daily_gain'])); ratio=max(0.01,min(1.0,float(f['warning_percent'])/100.0))
                    c.execute("insert or replace into settings(setting_key,setting_value) values('male_min_daily_gain',?)",(str(target),))
                    c.execute("insert or replace into settings(setting_key,setting_value) values('male_warning_ratio',?)",(str(ratio),))
                    return self.redirect('/performance','Besi performans eşikleri güncellendi.')
                if path=='/animal/weight':
                    aid=f['animal_id']; measure=f['measure_date']; weight=float(f['weight'])
                    if weight<=0:return self.redirect('/animal?id='+aid,'Kilo sıfırdan büyük olmalıdır.')
                    existing=c.execute('select id from weights where animal_id=? and measure_date=?',(aid,measure)).fetchone()
                    if existing:c.execute('update weights set weight=?,notes=? where id=?',(weight,f.get('notes'),existing['id']));msg='Aynı tarihteki tartım güncellendi.'
                    else:c.execute('insert into weights(animal_id,measure_date,weight,notes) values(?,?,?,?)',(aid,measure,weight,f.get('notes')));msg='Tartım kaydı eklendi.'
                    return self.redirect('/animal?id='+aid,msg)
                if path=='/animal/milk':
                    c.execute('insert into milk(animal_id,measure_date,liters,notes) values(?,?,?,?)',(f['animal_id'],f['measure_date'],float(f['liters']),f.get('notes')));return self.redirect('/animal?id='+f['animal_id'],'Süt kaydı eklendi.')
                if path in ('/animals','/males'):
                    upload=f.get('photo_file'); photo_url=f.get('photo_url','')
                    if upload and isinstance(upload,dict) and upload.get('content'):
                        ext=Path(upload['filename']).suffix.lower()
                        if ext not in ('.jpg','.jpeg','.png','.webp','.gif'): return self.redirect(path,'Desteklenmeyen fotoğraf biçimi.')
                        if len(upload['content'])>10*1024*1024: return self.redirect(path,'Fotoğraf 10 MB sınırını aşıyor.')
                        name=save_optimized_upload('animal_new',upload); photo_url='/uploads/'+name
                    vals=(f['tag'],f.get('nickname'),f['gender'],f.get('breed'),f.get('birth_date'),f.get('notes'),f.get('paddock'),photo_url,float(f.get('sold_price') or 0),f.get('status') or 'Aktif',f.get('purchase_date',''),float(f.get('purchase_price') or 0),float(f.get('purchase_weight') or 0),float(f.get('daily_feed_cost') or 0),float(f.get('daily_care_cost') or 0),float(f.get('target_sale_price') or 0))
                    if f.get('id'):
                        c.execute('update animals set tag=?,nickname=?,gender=?,breed=?,birth_date=?,notes=?,paddock=?,photo_url=?,sold_price=?,status=?,purchase_date=?,purchase_price=?,purchase_weight=?,daily_feed_cost=?,daily_care_cost=?,target_sale_price=? where id=?',vals+(f['id'],)); aid=f['id']
                    else:
                        cur=c.execute('insert into animals(tag,nickname,gender,breed,birth_date,notes,paddock,photo_url,sold_price,status,purchase_date,purchase_price,purchase_weight,daily_feed_cost,daily_care_cost,target_sale_price) values(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)',vals); aid=cur.lastrowid
                    if photo_url.startswith('/uploads/'):
                        fname=photo_url.split('/uploads/',1)[1]
                        exists=c.execute('select 1 from animal_photos where animal_id=? and filename=?',(aid,fname)).fetchone()
                        if not exists:c.execute('insert into animal_photos(animal_id,filename,created_at,caption) values(?,?,?,?)',(aid,fname,datetime.now().strftime('%Y-%m-%d %H:%M:%S'),'Profil fotoğrafı'))
                    return self.redirect('/males' if path=='/males' else '/animals','Hayvan kaydedildi.')
                if path=='/calves':
                    m=c.execute("select id from animals where id=? and gender='Dişi'",(f['mother_id'],)).fetchone()
                    if not m:return self.redirect('/calves','Anne olarak yalnızca kayıtlı dişi hayvan seçilebilir.')
                    vals=(f['tag'],f['mother_id'],f.get('father_tag'),f['birth_date'],f.get('gender'),f.get('notes'))
                    if f.get('id'):c.execute('update calves set tag=?,mother_id=?,father_tag=?,birth_date=?,gender=?,notes=? where id=?',vals+(f['id'],));msg='Buzağı güncellendi.'
                    else:c.execute('insert into calves(tag,mother_id,father_tag,birth_date,gender,notes) values(?,?,?,?,?,?)',vals);msg='Buzağı kaydedildi.'
                    promote_mature_calves(); return self.redirect('/calves',msg)
                if path=='/estrus':
                    aid=f.get('animal_id',''); a=c.execute("select id,tag from animals where id=? and gender='Dişi' and coalesce(status,'Aktif')='Aktif'",(aid,)).fetchone()
                    if not a:return self.redirect('/estrus','Lütfen geçerli ve aktif bir dişi hayvan seçin.')
                    if is_currently_pregnant(c,aid):return self.redirect('/estrus',f'{a["tag"]} şu anda gebe görünüyor. Doğum ve buzağı kaydı tamamlanmadan yeni kızgınlık kaydı açılamaz.')
                    estrus_date=f.get('estrus_date','')
                    try: estrus_day=date.fromisoformat(estrus_date)
                    except Exception:return self.redirect('/estrus','Geçerli bir kızgınlık tarihi girin.')
                    if estrus_day>date.today():return self.redirect('/estrus','Gelecek tarihli kızgınlık kaydı girilemez.')
                    duplicate=c.execute('select id from estrus_records where animal_id=? and estrus_date=?',(aid,estrus_date)).fetchone()
                    if duplicate:return self.redirect('/estrus',f'{a["tag"]} için {fmt_date(estrus_date)} tarihinde zaten kızgınlık kaydı var. Yeni kayıt açılmadı; mevcut kaydı Düzenle ile değiştirebilirsiniz.')
                    c.execute('insert into estrus_records(animal_id,estrus_date,signs,notes,created_at) values(?,?,?,?,?)',(aid,estrus_date,f.get('signs',''),f.get('notes',''),datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
                    audit(username,'Kızgınlık kaydı eklendi',f'{a["tag"]} · {estrus_date}',self.client_ip())
                    return self.redirect('/estrus',f'{a["tag"]} için kızgınlık kaydı eklendi. Beklenen yeni pencere {fmt_date((estrus_day+timedelta(days=18)).isoformat())} – {fmt_date((estrus_day+timedelta(days=24)).isoformat())}.')
                if path=='/estrus-edit':
                    eid=f.get('id','')
                    rec=c.execute('select e.*,a.tag from estrus_records e join animals a on a.id=e.animal_id where e.id=?',(eid,)).fetchone()
                    if not rec:return self.redirect('/estrus','Kızgınlık kaydı bulunamadı.')
                    estrus_date=f.get('estrus_date','')
                    try: estrus_day=date.fromisoformat(estrus_date)
                    except Exception:return self.redirect('/estrus-edit?id='+str(eid),'Geçerli bir kızgınlık tarihi girin.')
                    if estrus_day>date.today():return self.redirect('/estrus-edit?id='+str(eid),'Gelecek tarihli kızgınlık kaydı girilemez.')
                    duplicate=c.execute('select id from estrus_records where animal_id=? and estrus_date=? and id<>?',(rec['animal_id'],estrus_date,eid)).fetchone()
                    if duplicate:return self.redirect('/estrus-edit?id='+str(eid),f'{rec["tag"]} için {fmt_date(estrus_date)} tarihinde başka bir kızgınlık kaydı zaten var.')
                    c.execute('update estrus_records set estrus_date=?,signs=?,notes=? where id=?',(estrus_date,f.get('signs',''),f.get('notes',''),eid))
                    audit(username,'Kızgınlık kaydı güncellendi',f'{rec["tag"]} · {estrus_date}',self.client_ip())
                    return self.redirect('/estrus','Kızgınlık kaydı güncellendi.')
                if path=='/estrus-send':
                    eid=f.get('estrus_id','')
                    try: cycle_no=int(f.get('cycle_no','1'))
                    except Exception: cycle_no=1
                    rec=c.execute('select e.*,a.tag from estrus_records e join animals a on a.id=e.animal_id where e.id=?',(eid,)).fetchone()
                    if not rec:return self.redirect('/estrus','Kızgınlık kaydı bulunamadı.')
                    if is_currently_pregnant(c,rec['animal_id']):return self.redirect('/estrus',f'{rec["tag"]} şu anda gebe görünüyor. Tohumlama takibine gönderilemedi.')
                    current_cycle=next_estrus_cycle(c,rec,date.today())
                    if not current_cycle or current_cycle['cycle_no']!=cycle_no:
                        return self.redirect('/estrus','Bu östrus dönemi daha önce sonuçlandırılmış veya artık aktif değil.')
                    c.execute("insert or replace into estrus_decisions(estrus_id,cycle_no,decision,decision_date,notes) values(?,?,?,?,?)",(eid,cycle_no,'Tohumlamaya Gönderildi',date.today().isoformat(),'Hayvan Tohumlama Takibi ekranına gönderildi; gerçek tohumlama kaydı bekleniyor.'))
                    audit(username,'Kızgınlık tohumlama takibine gönderildi',f'{rec["tag"]} · {cycle_no}. tahmini döngü',self.client_ip())
                    return self.redirect(f'/inseminations?animal={rec["animal_id"]}&estrus={rec["id"]}',f'{rec["tag"]} Tohumlama Takibi ekranına gönderildi. Kızgınlık kartı kapatıldı; tohumlama gerçekleştiğinde tarihi kaydedin.')
                if path=='/estrus-skip':
                    eid=f.get('estrus_id','')
                    try: cycle_no=int(f.get('cycle_no','1'))
                    except Exception: cycle_no=1
                    rec=c.execute('select e.*,a.tag from estrus_records e join animals a on a.id=e.animal_id where e.id=?',(eid,)).fetchone()
                    if not rec:return self.redirect('/estrus','Kızgınlık kaydı bulunamadı.')
                    current=next_estrus_cycle(c,rec,date.today())
                    if not current or current['cycle_no']!=cycle_no:return self.redirect(f.get('return_to') or '/estrus','Bu östrus dönemi daha önce sonuçlandırılmış veya artık aktif değil.')
                    c.execute("insert or replace into estrus_decisions(estrus_id,cycle_no,decision,decision_date,notes) values(?,?,?,?,?)",(eid,cycle_no,'Atlandı',date.today().isoformat(),'Kullanıcı bu östrus dönemini atladı.'))
                    audit(username,'Östrus atlandı',f'{rec["tag"]} · {cycle_no}. tahmini döngü',self.client_ip())
                    return self.redirect(f.get('return_to') or '/estrus',f'{rec["tag"]} için bu östrus dönemi atlandı. Sonraki döngü otomatik takip edilecek.')
                if path=='/estrus-inseminate':
                    eid=f.get('estrus_id','')
                    rec=c.execute('select e.*,a.tag from estrus_records e join animals a on a.id=e.animal_id where e.id=?',(eid,)).fetchone()
                    if not rec:return self.redirect('/estrus','Kızgınlık kaydı bulunamadı.')
                    if is_currently_pregnant(c,rec['animal_id']):return self.redirect('/estrus',f'{rec["tag"]} şu anda gebe görünüyor. Yeni tohumlama işlemi engellendi.')
                    try:
                        observed=date.fromisoformat(rec['estrus_date']); start=observed+timedelta(days=18); end=observed+timedelta(days=24)
                    except Exception:return self.redirect('/estrus','Kızgınlık tarihi geçersiz.')
                    today=date.today()
                    actual_window=observed<=today<=observed+timedelta(days=1)
                    predicted_window=start<=today<=end
                    if not (actual_window or predicted_window):return self.redirect('/estrus',f'Tohumlama aktarımı yalnızca kaydedilen kızgınlık günü/ertesi gün veya beklenen kızgınlık penceresinde ({fmt_date(start.isoformat())} – {fmt_date(end.isoformat())}) yapılabilir.')
                    same_day=c.execute('select id,attempt from inseminations where animal_id=? and insemination_date=?',(rec['animal_id'],today.isoformat())).fetchone()
                    if same_day:return self.redirect('/inseminations?animal='+str(rec['animal_id']),f'{rec["tag"]} için bugün zaten tohumlama kaydı mevcut. Mükerrer kayıt oluşturulmadı.')
                    latest=c.execute('select max(attempt) from inseminations where animal_id=?',(rec['animal_id'],)).fetchone()[0] or 0
                    attempt=int(latest)+1
                    if attempt>3:return self.redirect('/inseminations?animal='+str(rec['animal_id']),'Bu hayvan için 3 tohumlama denemesi zaten kayıtlı. Önce mevcut kayıtları düzenleyin.')
                    cur=c.execute('insert into inseminations(animal_id,attempt,insemination_date,pregnancy_result,due_date) values(?,?,?,?,?)',(rec['animal_id'],attempt,today.isoformat(),'Bekleniyor',''))
                    cycle_no=0 if actual_window else max(1,round((today-observed).days/21))
                    c.execute("insert or replace into estrus_decisions(estrus_id,cycle_no,decision,decision_date,insemination_id,notes) values(?,?,?,?,?,?)",(rec['id'],cycle_no,'Tohumlamaya Gönderildi',today.isoformat(),cur.lastrowid,'Kızgınlık Takibi üzerinden doğrudan aktarıldı.'))
                    audit(username,'Kızgınlıktan tohumlamaya aktarıldı',f'{rec["tag"]} · {attempt}. deneme · {today.isoformat()}',self.client_ip())
                    return self.redirect('/inseminations?animal='+str(rec['animal_id']),f'{rec["tag"]} kızgınlık takibinden {attempt}. tohumlama olarak aktarıldı. Gebelik sonucu Bekleniyor durumunda.')
                if path=='/estrus-delete':
                    rec=c.execute('select e.*,a.tag from estrus_records e join animals a on a.id=e.animal_id where e.id=?',(f.get('id',''),)).fetchone()
                    if not rec:return self.redirect('/estrus','Kızgınlık kaydı bulunamadı.')
                    c.execute('delete from estrus_records where id=?',(rec['id'],)); audit(username,'Kızgınlık kaydı silindi',f'{rec["tag"]} · {rec["estrus_date"]}',self.client_ip())
                    return self.redirect('/estrus','Kızgınlık kaydı silindi.')
                if path=='/inseminations':
                    aid=f.get('animal_id','')
                    a=c.execute("select id,tag from animals where id=? and gender='Dişi' and coalesce(status,'Aktif')='Aktif'",(aid,)).fetchone()
                    if not a:return self.redirect('/inseminations','Lütfen geçerli ve aktif bir dişi hayvan seçin.')
                    if is_currently_pregnant(c,aid):return self.redirect('/inseminations',f'{a["tag"]} şu anda gebe görünüyor. Doğum ve buzağı kaydı tamamlanmadan yeni tohumlama kaydı açılamaz.')
                    ins_date=f.get('insemination_date','')
                    try:ins_day=date.fromisoformat(ins_date)
                    except Exception:return self.redirect('/inseminations','Geçerli bir tohumlama tarihi girin.')
                    if ins_day>date.today():return self.redirect('/inseminations','Gelecek tarihli tohumlama kaydı girilemez.')
                    latest=c.execute('select max(attempt) from inseminations where animal_id=?',(aid,)).fetchone()[0] or 0
                    attempt=int(latest)+1
                    if attempt>3:return self.redirect('/inseminations?animal='+str(aid),'Bu hayvan için 3 tohumlama denemesi zaten kayıtlı. Önce mevcut kayıtları düzenleyin.')
                    bull_tag=(f.get('bull_tag') or '').strip();bull_name=(f.get('bull_name') or '').strip();inseminator=(f.get('inseminator') or '').strip()
                    cur=c.execute('insert into inseminations(animal_id,attempt,insemination_date,pregnancy_result,due_date,bull_tag,bull_name,inseminator) values(?,?,?,?,?,?,?,?)',(aid,attempt,ins_date,'Bekleniyor','',bull_tag,bull_name,inseminator))
                    linked_estrus=f.get('estrus_id','')
                    if linked_estrus:
                        er=c.execute('select * from estrus_records where id=? and animal_id=?',(linked_estrus,aid)).fetchone()
                        if er:
                            cyc=next_estrus_cycle(c,er,date.fromisoformat(ins_date))
                            if cyc:c.execute("insert or replace into estrus_decisions(estrus_id,cycle_no,decision,decision_date,insemination_id,notes) values(?,?,?,?,?,?)",(er['id'],cyc['cycle_no'],'Tohumlamaya Gönderildi',ins_date,cur.lastrowid,'Tohumlama ekranından kaydedildi.'))
                    audit(username,'Tohumlama kaydı eklendi',f'{a["tag"]} · {attempt}. deneme · {ins_date}',self.client_ip())
                    return self.redirect('/inseminations?animal='+str(aid),f'{a["tag"]} için {attempt}. tohumlama kaydedildi.')
                if path=='/pregnancy-edit':
                    aid=int(f.get('animal_id') or 0)
                    a=c.execute("select * from animals where id=? and gender='Dişi'",(aid,)).fetchone()
                    if not a:return self.redirect('/animals','Dişi hayvan bulunamadı.')
                    if not str(a['pregnancy_source'] or '').startswith('Satın Alındığında Gebe'):
                        return self.redirect('/animal?id='+str(aid),'Bu işlem yalnız dışarıdan gebe alınan hayvanlar için kullanılabilir.')
                    rec=current_pregnancy_record(c,aid)
                    if not rec:return self.redirect('/animal?id='+str(aid),'Aktif gebelik kaydı bulunamadı.')
                    mode=(f.get('pregnancy_info_mode') or 'age').strip()
                    ref=(f.get('pregnancy_entry_date') or date.today().isoformat()).strip()
                    try:ref_day=date.fromisoformat(ref)
                    except:return self.redirect('/pregnancy-edit?animal_id='+str(aid),'Bilgi tarihi geçersiz.')
                    age_months=0.0
                    if mode=='date':
                        ins_date=(f.get('known_insemination_date') or '').strip()
                        try:ins_day=date.fromisoformat(ins_date)
                        except:return self.redirect('/pregnancy-edit?animal_id='+str(aid),'Son tohumlama tarihi geçersiz.')
                        if ins_day>ref_day:return self.redirect('/pregnancy-edit?animal_id='+str(aid),'Tohumlama tarihi bilgi tarihinden sonra olamaz.')
                        due=(ins_day+timedelta(days=280)).isoformat()
                        result='Gebe (Satın Alındığında · Tohumlama Tarihi Biliniyor)'
                        source='Satın Alındığında Gebe · Tohumlama Tarihi'
                    else:
                        try:age_months=float(f.get('pregnancy_age_months') or 0)
                        except:age_months=0
                        if age_months<=0 or age_months>9:return self.redirect('/pregnancy-edit?animal_id='+str(aid),'Gebelik yaşını 0,5-9 ay arasında girin.')
                        elapsed=round(age_months*(280/9));ins_day=ref_day-timedelta(days=elapsed);ins_date=ins_day.isoformat()
                        due=(ins_day+timedelta(days=280)).isoformat()
                        result='Gebe (Satın Alındığında · Gebelik Yaşından Tahmini)'
                        source='Satın Alındığında Gebe · Yaklaşık'
                    c.execute('update inseminations set insemination_date=?,pregnancy_result=?,due_date=? where id=?',(ins_date,result,due,rec['id']))
                    c.execute('update animals set pregnancy_source=?,pregnancy_age_months_at_entry=?,pregnancy_entry_date=? where id=?',(source,age_months,ref,aid))
                    audit(username,'Dışarıdan gebe hayvan bilgisini güncelledi',f'{a["tag"]} · Tahmini doğum {due}',self.client_ip())
                    return self.redirect('/animal?id='+str(aid),'Gebelik bilgisi güncellendi; hayvan kartı ve Tohumlama kaydı senkronlandı.')
                if path=='/insemination-edit':
                    iid=f.get('id','')
                    rec=c.execute('select i.*,a.tag from inseminations i join animals a on a.id=i.animal_id where i.id=?',(iid,)).fetchone()
                    if not rec:return self.redirect('/inseminations','Tohumlama kaydı bulunamadı.')
                    ins_date=f.get('insemination_date','')
                    try:ins_day=date.fromisoformat(ins_date)
                    except Exception:return self.redirect('/insemination-edit?id='+str(iid),'Geçerli bir tarih girin.')
                    if ins_day>date.today():return self.redirect('/insemination-edit?id='+str(iid),'Gelecek tarihli tohumlama kaydı girilemez.')
                    result=f.get('pregnancy_result','Bekleniyor')
                    if result not in ('Bekleniyor','Pozitif','Negatif','Belirsiz'):result='Belirsiz'
                    animal=c.execute('select pregnancy_source from animals where id=?',(rec['animal_id'],)).fetchone()
                    external=animal and str(animal['pregnancy_source'] or '').startswith('Satın Alındığında Gebe')
                    if external and result=='Pozitif':
                        result='Gebe (Satın Alındığında · Tohumlama Tarihi Biliniyor)'
                    due=(ins_day+timedelta(days=280)).isoformat() if is_pregnant_value(result) else ''
                    bull_tag=(f.get('bull_tag') or '').strip();bull_name=(f.get('bull_name') or '').strip();inseminator=(f.get('inseminator') or '').strip()
                    c.execute('update inseminations set insemination_date=?,pregnancy_result=?,due_date=?,bull_tag=?,bull_name=?,inseminator=? where id=?',(ins_date,result,due,bull_tag,bull_name,inseminator,iid))
                    audit(username,'Tohumlama kaydı güncellendi',f'{rec["tag"]} · {rec["attempt"]}. deneme',self.client_ip())
                    return self.redirect('/inseminations?animal='+str(rec['animal_id']),'Tohumlama kaydı güncellendi.')
                if path=='/insemination-delete':
                    iid=f.get('id','')
                    rec=c.execute('select i.*,a.tag from inseminations i join animals a on a.id=i.animal_id where i.id=?',(iid,)).fetchone()
                    if not rec:return self.redirect('/inseminations','Tohumlama kaydı bulunamadı.')
                    c.execute('delete from inseminations where id=?',(iid,))
                    remaining=c.execute('select id,attempt from inseminations where animal_id=? order by attempt,insemination_date,id',(rec['animal_id'],)).fetchall()
                    for idx,row in enumerate(remaining,1):
                        if row['attempt']!=idx:c.execute('update inseminations set attempt=? where id=?',(idx,row['id']))
                    audit(username,'Tohumlama kaydı silindi',f'{rec["tag"]} · {rec["attempt"]}. deneme',self.client_ip())
                    return self.redirect('/inseminations?animal='+str(rec['animal_id']),'Tohumlama kaydı silindi ve deneme sırası yeniden düzenlendi.')
                if path=='/pregnancy-vaccine/done':
                    aid=int(f['animal_id']); ins_id=int(f['insemination_id']); month=int(f['month'])
                    if month not in (7,8):return self.redirect('/','Geçersiz gebelik aşı görevi.')
                    ins=c.execute("select i.*,a.tag from inseminations i join animals a on a.id=i.animal_id where i.id=? and i.animal_id=?",(ins_id,aid)).fetchone()
                    if not ins or not is_pregnant_value(ins['pregnancy_result']):return self.redirect('/','Gebelik kaydı bulunamadı veya aktif değil.')
                    token=f'GEBELIK_ASI|{ins_id}|{month}'
                    existing=c.execute("select id from health where animal_id=? and notes like ? limit 1",(aid,token+'%')).fetchone()
                    if not existing:
                        product=f'{month}. Ay Gebelik Aşısı'
                        notes=token+' | Dashboard gebelik aşı alarmından tamamlandı.'
                        c.execute('insert into health(animal_id,kind,product,applied_date,next_date,cost,notes) values(?,?,?,?,?,?,?)',(aid,'Aşı',product,date.today().isoformat(),'',0,notes))
                        audit(username,'Gebelik aşısı yapıldı',f'{ins["tag"]} · {month}. ay',self.client_ip())
                    target=f.get('return_to') or ('/animal?id='+str(aid))
                    return self.redirect(target,f'{ins["tag"]} · {month}. ay gebelik aşısı sağlık geçmişine kaydedildi.')
                if path=='/health-edit':
                    hid=int(f.get('id') or 0)
                    old=c.execute('select * from health where id=?',(hid,)).fetchone()
                    if not old:return self.redirect('/health','Sağlık kaydı bulunamadı.')
                    c.execute('update health set kind=?,product=?,applied_date=?,next_date=?,cost=?,notes=? where id=?',
                              (f.get('kind'),f.get('product'),f.get('applied_date'),f.get('next_date') or '',float(f.get('cost') or 0),f.get('notes') or '',hid))
                    audit(username,'Sağlık kaydı düzenledi',f'#{hid} · {f.get("product")}',self.client_ip())
                    return self.redirect('/health','Sağlık kaydı güncellendi.')
                if path=='/health-delete':
                    hid=int(f.get('id') or 0)
                    old=c.execute('select * from health where id=?',(hid,)).fetchone()
                    if not old:return self.redirect('/health','Sağlık kaydı bulunamadı.')
                    c.execute('delete from health where id=?',(hid,))
                    audit(username,'Sağlık kaydı sildi',f'#{hid} · {old["product"]}',self.client_ip())
                    return self.redirect('/health','Sağlık kaydı silindi.')
                if path=='/health/plan-done':
                    source_id=int(f.get('source_id') or 0)
                    src=c.execute("select h.*,a.tag as animal_tag,ca.tag as calf_tag from health h left join animals a on a.id=h.animal_id left join calves ca on ca.id=h.calf_id where h.id=?",(source_id,)).fetchone()
                    if not src:return self.redirect('/health','Planlanan sağlık kaydı bulunamadı.')
                    if not src['next_date']:return self.redirect(f.get('return_to') or '/health','Bu plan daha önce tamamlanmış.')
                    tag=src['animal_tag'] or src['calf_tag'] or '-'; actual=date.today().isoformat(); planned=src['next_date']
                    product=str(src['product'] or '')
                    if src['kind']=='Aşı' and 'IKINCI_DOZ_PLAN' in str(src['notes'] or ''): product += ' · 2. Doz'
                    notes=f'Planlanan sağlık işlemi tamamlandı | Kaynak sağlık #{source_id} | Planlanan {planned}'
                    c.execute('insert into health(animal_id,calf_id,kind,product,applied_date,next_date,cost,notes) values(?,?,?,?,?,?,?,?)',(src['animal_id'],src['calf_id'],src['kind'],product,actual,'',0,notes))
                    c.execute("update health set next_date='',notes=coalesce(notes,'')||? where id=?",(' | Tamamlandı: '+actual,source_id))
                    audit(username,'Planlanan sağlık işlemini tamamladı',f'{tag} · {product}',self.client_ip())
                    return self.redirect(f.get('return_to') or '/health',f'{tag} · işlem {fmt_date(actual)} tarihinde yapıldı olarak kaydedildi.')
                if path=='/health/second-dose-done':
                    source_id=int(f.get('source_id') or 0)
                    src=c.execute("select h.*,a.tag as animal_tag,ca.tag as calf_tag from health h left join animals a on a.id=h.animal_id left join calves ca on ca.id=h.calf_id where h.id=?",(source_id,)).fetchone()
                    if not src:return self.redirect('/health','Planlanan aşı kaydı bulunamadı.')
                    if not src['next_date']:return self.redirect('/health','Bu ikinci doz daha önce tamamlanmış veya plan kapatılmış.')
                    if src['kind']!='Aşı' or 'IKINCI_DOZ_PLAN' not in str(src['notes'] or ''):return self.redirect('/health','Bu kayıt ikinci doz planı değil.')
                    tag=src['animal_tag'] or src['calf_tag'] or '-'
                    actual=date.today().isoformat()
                    product=str(src['product'] or '')+' · 2. Doz'
                    notes=f'2. doz tamamlandı | Kaynak sağlık #{source_id} | Planlanan {src["next_date"]}'
                    c.execute('insert into health(animal_id,calf_id,kind,product,applied_date,next_date,cost,notes) values(?,?,?,?,?,?,?,?)',(src['animal_id'],src['calf_id'],'Aşı',product,actual,'',0,notes))
                    c.execute("update health set next_date='',notes=coalesce(notes,'')||? where id=?",(' | 2. doz tamamlandı: '+actual,source_id))
                    audit(username,'Aşı ikinci dozu tamamladı',f'{tag} · {product}',self.client_ip())
                    return self.redirect(f.get('return_to') or '/health',f'{tag} · 2. doz sağlık geçmişine kaydedildi.')
                if path=='/health/task-done':
                    tid=int(f.get('task_id') or 0)
                    t=c.execute("""select t.*,hc.kind,hc.product from health_tasks t join health_courses hc on hc.id=t.course_id where t.id=?""",(tid,)).fetchone()
                    if not t:return self.redirect('/health','Planlanan işlem bulunamadı.')
                    if t['status']!='Bekliyor':return self.redirect('/health','Bu uygulama daha önce tamamlanmış.')
                    actual=date.today().isoformat();product=str(t['product'] or '')
                    if t['kind']=='Aşı':product+=f' · {int(t["dose_no"] or 1)}. Doz'
                    elif t['kind']=='İlaç':product+=f' · Gün {int(t["day_no"] or 1)}/{int(t["day_total"] or 1)} · Uygulama {int(t["application_no"] or 1)}/{int(t["applications_per_day"] or 1)}'
                    notes=f'Planlı uygulama tamamlandı | Plan #{t["course_id"]} | Planlanan {t["planned_date"]}'
                    c.execute('insert into health(animal_id,calf_id,kind,product,applied_date,next_date,cost,notes) values(?,?,?,?,?,?,?,?)',(t['animal_id'],t['calf_id'],t['kind'],product,actual,'',float(t['cost'] or 0),notes))
                    c.execute("update health_tasks set status='Tamamlandı',completed_date=? where id=?",(actual,tid))
                    if float(t['cost'] or 0)>0:c.execute('insert into finance(tx_date,tx_type,category,amount,description,payment_method,animal_id,created_at) values(?,?,?,?,?,?,?,?)',(actual,'Gider',t['kind'],float(t['cost'] or 0),product,'Nakit',t['animal_id'],datetime.now().isoformat()))
                    audit(username,'Planlı sağlık uygulaması tamamlandı',f'Görev #{tid} · {product}',self.client_ip())
                    return self.redirect('/health','Sağlık uygulaması Yapıldı olarak kaydedildi.')
                if path=='/health/task-batch-done':
                    course_id=int(f.get('course_id') or 0);planned=(f.get('planned_date') or '').strip();dose_no=int(f.get('dose_no') or 1);day_no=int(f.get('day_no') or 1);app_no=int(f.get('application_no') or 1)
                    tasks=c.execute("""select t.*,hc.kind,hc.product from health_tasks t join health_courses hc on hc.id=t.course_id where t.course_id=? and t.planned_date=? and t.dose_no=? and t.day_no=? and t.application_no=? and t.status='Bekliyor'""",(course_id,planned,dose_no,day_no,app_no)).fetchall()
                    if not tasks:return self.redirect('/health','Bu toplu uygulama daha önce tamamlanmış veya bulunamadı.')
                    actual=date.today().isoformat();count=0
                    for t in tasks:
                        product=str(t['product'] or '')
                        if t['kind']=='Aşı':product+=f' · {int(t["dose_no"] or 1)}. Doz'
                        else:product+=f' · Gün {int(t["day_no"] or 1)}/{int(t["day_total"] or 1)} · Uygulama {int(t["application_no"] or 1)}/{int(t["applications_per_day"] or 1)}'
                        notes=f'Padok bazlı plan tamamlandı | Plan #{course_id} | Planlanan {planned}'
                        c.execute('insert into health(animal_id,calf_id,kind,product,applied_date,next_date,cost,notes) values(?,?,?,?,?,?,?,?)',(t['animal_id'],t['calf_id'],t['kind'],product,actual,'',float(t['cost'] or 0),notes))
                        c.execute("update health_tasks set status='Tamamlandı',completed_date=? where id=?",(actual,t['id']))
                        if float(t['cost'] or 0)>0:c.execute('insert into finance(tx_date,tx_type,category,amount,description,payment_method,animal_id,created_at) values(?,?,?,?,?,?,?,?)',(actual,'Gider',t['kind'],float(t['cost'] or 0),product,'Nakit',t['animal_id'],datetime.now().isoformat()))
                        count+=1
                    audit(username,'Padok sağlık uygulaması tamamlandı',f'Plan #{course_id} · {count} hayvan',self.client_ip())
                    return self.redirect('/health',f'{count} hayvan için uygulama Yapıldı olarak sağlık geçmişine işlendi.')
                if path=='/health':
                    scope=(f.get('scope_type') or 'single').strip();kind=(f.get('kind') or '').strip();product=(f.get('product') or '').strip();applied=(f.get('applied_date') or '').strip();notes=(f.get('notes') or '').strip()
                    if not product:return self.redirect('/health','Ürün / işlem adı zorunludur.')
                    try:date.fromisoformat(applied)
                    except:return self.redirect('/health','Uygulama tarihi geçersiz.')
                    try:cost=max(0,float(f.get('cost') or 0))
                    except:cost=0
                    targets=[];paddock_id=None
                    if scope=='paddock':
                        if kind!='Aşı':return self.redirect('/health','Padok bazlı plan şu anda yalnız aşı için kullanılabilir.')
                        try:paddock_id=int(f.get('paddock_id') or 0)
                        except:paddock_id=0
                        pd=c.execute('select id,name from paddocks where id=? and active=1',(paddock_id,)).fetchone()
                        if not pd:return self.redirect('/health','Geçerli bir padok seçin.')
                        for r in c.execute("select id,tag from animals where paddock_id=? and coalesce(status,'Aktif')='Aktif'",(paddock_id,)).fetchall():targets.append((r['id'],None,r['tag']))
                        for r in c.execute("select id,tag from calves where paddock_id=? and promoted_animal_id is null",(paddock_id,)).fetchall():targets.append((None,r['id'],r['tag']))
                        if not targets:return self.redirect('/health','Seçilen padokta aktif hayvan bulunamadı.')
                    else:
                        subject=(f.get('subject_key') or '').strip()
                        if ':' not in subject:return self.redirect('/health','Lütfen geçerli bir hayvan veya buzağı seçin.')
                        stype,sid=subject.split(':',1)
                        try:sid=int(sid)
                        except:return self.redirect('/health','Hayvan seçimi geçersiz.')
                        if stype=='A':
                            rec=c.execute("select id,tag from animals where id=? and coalesce(status,'Aktif')='Aktif'",(sid,)).fetchone()
                            if not rec:return self.redirect('/health','Seçilen hayvan aktif sürüde değil.')
                            targets=[(sid,None,rec['tag'])]
                        elif stype=='C':
                            rec=c.execute("select id,tag from calves where id=? and promoted_animal_id is null",(sid,)).fetchone()
                            if not rec:return self.redirect('/health','Seçilen buzağı aktif kayıtlarda değil.')
                            targets=[(None,sid,rec['tag'])]
                        else:return self.redirect('/health','Hayvan seçimi geçersiz.')
                    if kind in ('Aşı','İlaç'):
                        if kind=='Aşı':
                            try:dose_count=max(1,min(10,int(f.get('dose_count') or 1)))
                            except:dose_count=1
                            try:interval=max(1,min(365,int(f.get('dose_interval_days') or 15)))
                            except:interval=15
                            treatment_days=1;times_per_day=1
                        else:
                            try:treatment_days=max(1,min(60,int(f.get('treatment_days') or 1)))
                            except:treatment_days=1
                            try:times_per_day=max(1,min(6,int(f.get('times_per_day') or 1)))
                            except:times_per_day=1
                            dose_count=1;interval=0
                        cur=c.execute('''insert into health_courses(kind,product,scope_type,paddock_id,start_date,treatment_days,times_per_day,dose_count,interval_days,cost_per_application,notes,created_at) values(?,?,?,?,?,?,?,?,?,?,?,?)''',(kind,product,scope,paddock_id,applied,treatment_days,times_per_day,dose_count,interval,cost,notes,datetime.now().isoformat(timespec='seconds')))
                        course_id=cur.lastrowid;start_day=date.fromisoformat(applied);task_count=0
                        for animal_id,calf_id,tag in targets:
                            if kind=='Aşı':
                                for dn in range(1,dose_count+1):
                                    planned=(start_day+timedelta(days=(dn-1)*interval)).isoformat()
                                    c.execute('''insert into health_tasks(course_id,animal_id,calf_id,planned_date,dose_no,dose_total,day_no,day_total,application_no,applications_per_day,status,cost,notes) values(?,?,?,?,?,?,?,?,?,?,?,?,?)''',(course_id,animal_id,calf_id,planned,dn,dose_count,1,1,1,1,'Bekliyor',cost,notes));task_count+=1
                            else:
                                for day_no in range(1,treatment_days+1):
                                    planned=(start_day+timedelta(days=day_no-1)).isoformat()
                                    for app_no in range(1,times_per_day+1):
                                        c.execute('''insert into health_tasks(course_id,animal_id,calf_id,planned_date,dose_no,dose_total,day_no,day_total,application_no,applications_per_day,status,cost,notes) values(?,?,?,?,?,?,?,?,?,?,?,?,?)''',(course_id,animal_id,calf_id,planned,1,1,day_no,treatment_days,app_no,times_per_day,'Bekliyor',cost,notes));task_count+=1
                        audit(username,'Sağlık planı oluşturdu',f'{kind} · {product} · {len(targets)} hayvan · {task_count} uygulama',self.client_ip())
                        if scope=='paddock':return self.redirect('/health',f'{len(targets)} hayvan için {dose_count} dozluk padok aşı planı oluşturuldu.')
                        if kind=='Aşı':return self.redirect('/health',f'{dose_count} dozluk aşı planı oluşturuldu. Her doz Yapıldı olarak işaretlenebilir.')
                        return self.redirect('/health',f'{treatment_days} gün × günde {times_per_day} uygulamalık tedavi planı oluşturuldu.')
                    if scope!='single':return self.redirect('/health','Muayene kaydı tek hayvan üzerinden oluşturulmalıdır.')
                    animal_id,calf_id,subject_tag=targets[0];next_date=(f.get('next_date') or '').strip()
                    c.execute('insert into health(animal_id,calf_id,kind,product,applied_date,next_date,cost,notes) values(?,?,?,?,?,?,?,?)',(animal_id,calf_id,kind,product,applied,next_date,cost,notes))
                    if cost>0:c.execute('insert into finance(tx_date,tx_type,category,amount,description,payment_method,animal_id,created_at) values(?,?,?,?,?,?,?,?)',(applied,'Gider',kind,cost,product,'Nakit',animal_id,datetime.now().isoformat()))
                    audit(username,'Sağlık kaydı oluşturdu',f'{subject_tag} · {kind} · {product}',self.client_ip())
                    return self.redirect('/health','Sağlık kaydı oluşturuldu.')
                if path=='/finance/edit':
                    record_id=int(f['id'])
                    old=c.execute('select * from finance where id=?',(record_id,)).fetchone()
                    if not old:return self.redirect('/finance','Finans kaydı bulunamadı.')
                    category=f['category']; animal_id=f.get('animal_id') or None
                    action='Satıldı' if category=='Hayvan Satışı' else 'Kesildi' if category=='Kesim Geliri' else ''
                    if action and not animal_id:return self.redirect(f'/finance/edit?id={record_id}','Satış veya kesim için ilgili hayvan seçilmelidir.')
                    old_animal_id=old['animal_id']
                    link=c.execute('select * from feed_finance_links where finance_id=?',(record_id,)).fetchone()
                    new_amount=float(f['amount'])
                    if link and (f.get('linked_feed') or '')=='yes':
                        try:
                            linked_qty=float(f.get('linked_feed_qty') or 0);linked_unit=float(f.get('linked_feed_unit') or 0)
                        except:return self.redirect(f'/finance/edit?id={record_id}','Bağlı yem miktarı veya birim fiyatı geçersiz.')
                        if linked_qty<=0 or linked_unit<=0:return self.redirect(f'/finance/edit?id={record_id}','Bağlı yem miktarı ve birim fiyatı 0’dan büyük olmalıdır.')
                        new_amount=round(linked_qty*linked_unit,2)
                        c.execute('update feed_stock_transactions set tx_date=?,quantity_kg=?,unit_price=? where id=?',(f['tx_date'],linked_qty,linked_unit,link['stock_tx_id']))
                        c.execute('update feed_finance_links set quantity_kg=?,unit_price=? where finance_id=?',(linked_qty,linked_unit,record_id))
                        c.execute('insert into feed_prices(feed_id,effective_date,price_per_kg,notes) values(?,?,?,?)',(link['feed_id'],f['tx_date'],linked_unit,'Bağlı Finans/Stok düzenlemesinden güncellendi'))
                    c.execute(
                        'update finance set tx_date=?,tx_type=?,category=?,amount=?,description=?,payment_method=?,animal_id=?,animal_status_action=? where id=?',
                        (f['tx_date'],f['tx_type'],category,new_amount,f.get('description'),f.get('payment_method'),animal_id,action,record_id)
                    )
                    if category!='Süt Satışı':
                        c.execute('delete from finance_animals where finance_id=?',(record_id,))
                    recalculate_animal_exit_status(c,old_animal_id)
                    if animal_id!=old_animal_id:recalculate_animal_exit_status(c,animal_id)
                    return self.redirect('/finance','Finans kaydı güncellendi.' + (' Bağlı yem stok hareketi de birlikte güncellendi.' if link else ''))
                if path in ('/finance/delete','/finance-delete'):
                    record_id=int(f.get('id') or 0)
                    old=c.execute('select * from finance where id=?',(record_id,)).fetchone()
                    if not old:return self.redirect('/finance','Finans kaydı bulunamadı.')
                    animal_id=old['animal_id']
                    link=c.execute('select stock_tx_id from feed_finance_links where finance_id=?',(record_id,)).fetchone()
                    if link and link['stock_tx_id']:c.execute('delete from feed_stock_transactions where id=?',(link['stock_tx_id'],))
                    c.execute('delete from feed_finance_links where finance_id=?',(record_id,))
                    c.execute('delete from finance_animals where finance_id=?',(record_id,))
                    c.execute('delete from finance where id=?',(record_id,))
                    if animal_id:recalculate_animal_exit_status(c,animal_id)
                    return self.redirect('/finance','Finans kaydı silindi.')
                if path=='/finance':
                    category=f['category']; tx_type=f.get('tx_type','Gelir')
                    action='Satıldı' if category=='Hayvan Satışı' else 'Kesildi' if category=='Kesim Geliri' else ''
                    amount=round(float(f['amount']),2)
                    if amount<=0:return self.redirect('/finance','Tutar 0’dan büyük olmalıdır.')
                    milk_mode=tx_type=='Gelir' and category=='Süt Satışı'
                    if milk_mode:
                        raw_ids=[x.strip() for x in (f.get('milk_animal_ids') or '').split(',') if x.strip()];animal_ids=[]
                        for x in raw_ids:
                            try:aid=int(x)
                            except:continue
                            if aid not in animal_ids:animal_ids.append(aid)
                        if not animal_ids:return self.redirect('/finance','Süt satışı için en az bir aktif dişi hayvan seçilmelidir.')
                        placeholders=','.join('?' for _ in animal_ids);selected=c.execute(f"select id,tag,gender,status from animals where id in ({placeholders})",animal_ids).fetchall()
                        if len(selected)!=len(animal_ids):return self.redirect('/finance','Seçilen dişi hayvanlardan biri bulunamadı.')
                        if any(str(r['gender'] or '')!='Dişi' or str(r['status'] or 'Aktif')!='Aktif' for r in selected):return self.redirect('/finance','Süt satışına yalnız aktif dişi hayvanlar bağlanabilir. Sayfayı yenileyin.')
                        fingerprint=finance_request_fingerprint(username,f)
                        if not claim_request_once(c,fingerprint,15):
                            return self.redirect('/finance','⚠️ Aynı finans kaydı ikinci kez gönderildi; mükerrer kayıt engellendi.')
                        created=datetime.now().isoformat();tags=', '.join(str(r['tag']) for r in selected);desc=(f.get('description') or '').strip();relation_note=f'Süt satışı · {len(animal_ids)} dişi: {tags}'
                        c.execute('insert into finance(tx_date,tx_type,category,amount,description,payment_method,animal_id,created_at,animal_status_action) values(?,?,?,?,?,?,?,?,?)',(f['tx_date'],tx_type,category,amount,(desc+' · ' if desc else '')+relation_note,f.get('payment_method'),None,created,''));finance_id=c.execute('select last_insert_rowid()').fetchone()[0]
                        for aid in animal_ids:c.execute('insert or ignore into finance_animals(finance_id,animal_id,relation_type) values(?,?,?)',(finance_id,aid,'Süt Satışı'))
                        audit(username,'Süt satışını dişi hayvanlara ilişkilendirdi',f'{len(animal_ids)} dişi · {money(amount)}',self.client_ip());return self.redirect('/finance',f'Süt satışı kaydedildi ve {len(animal_ids)} aktif dişi hayvana ilişkilendirildi.')
                    bulk_mode=tx_type=='Gelir' and category in ('Hayvan Satışı','Kesim Geliri')
                    if bulk_mode:
                        raw_ids=[x.strip() for x in (f.get('animal_ids') or '').split(',') if x.strip()]
                        animal_ids=[]
                        for x in raw_ids:
                            try: aid=int(x)
                            except: continue
                            if aid not in animal_ids: animal_ids.append(aid)
                        if not animal_ids:return self.redirect('/finance','Hayvan satışı veya kesim geliri için en az bir hayvan seçilmelidir.')
                        placeholders=','.join('?' for _ in animal_ids)
                        selected=c.execute(f"select id,tag,nickname,status from animals where id in ({placeholders})",animal_ids).fetchall()
                        if len(selected)!=len(animal_ids):return self.redirect('/finance','Seçilen hayvanlardan biri bulunamadı.')
                        inactive=[r for r in selected if str(r['status'] or 'Aktif')!='Aktif']
                        if inactive:return self.redirect('/finance','Seçilen hayvanlardan biri artık aktif sürüde değil. Sayfayı yenileyip tekrar deneyin.')
                        fingerprint=finance_request_fingerprint(username,f)
                        if not claim_request_once(c,fingerprint,15):
                            return self.redirect('/finance','⚠️ Aynı finans kaydı ikinci kez gönderildi; mükerrer kayıt engellendi.')
                        total_cents=int(round(amount*100)); n=len(animal_ids)
                        base=total_cents//n; remainder=total_cents-(base*n)
                        description=(f.get('description') or '').strip()
                        batch_note=f'{category}: {n} hayvan · toplam {money(amount)}'
                        created=datetime.now().isoformat()
                        for idx,aid in enumerate(animal_ids):
                            cents=base+(1 if idx<remainder else 0)
                            share=cents/100.0
                            desc=(description+' · ' if description else '')+batch_note
                            c.execute('insert into finance(tx_date,tx_type,category,amount,description,payment_method,animal_id,created_at,animal_status_action) values(?,?,?,?,?,?,?,?,?)',(f['tx_date'],tx_type,category,share,desc,f.get('payment_method'),aid,created,action))
                            c.execute('update animals set status=?,exit_date=?,exit_reason=?,sold_price=? where id=?',(action,f['tx_date'],category,share,aid))
                        audit(username,'Finans geliri hayvanlara dağıtıldı',f'{category} · {n} hayvan · {money(amount)}',self.client_ip())
                        if n==1:
                            return self.redirect('/finance',f'{money(amount)} gelir seçilen hayvana kaydedildi.')
                        return self.redirect('/finance',f'{n} hayvana toplam {money(amount)} gelir otomatik dağıtıldı. Hayvan başı {money(amount/n)}.')
                    animal_id=f.get('animal_id') or None
                    if action and not animal_id:return self.redirect('/finance','Hayvan satışı veya kesim geliri için ilgili hayvan seçilmelidir.')
                    if category=='Hayvan Alımı' and animal_id:
                        existing_purchase=c.execute("select id from finance where category='Hayvan Alımı' and animal_id=? limit 1",(animal_id,)).fetchone()
                        if existing_purchase:return self.redirect('/finance','Bu hayvan daha önce Hayvan Alımı gideriyle ilişkilendirilmiş. İkinci alış kaydı engellendi.')
                    fingerprint=finance_request_fingerprint(username,f)
                    if not claim_request_once(c,fingerprint,15):
                        return self.redirect('/finance','⚠️ Aynı finans kaydı ikinci kez gönderildi; mükerrer kayıt engellendi.')
                    c.execute('insert into finance(tx_date,tx_type,category,amount,description,payment_method,animal_id,created_at,animal_status_action) values(?,?,?,?,?,?,?,?,?)',(f['tx_date'],tx_type,category,amount,f.get('description'),f.get('payment_method'),animal_id,datetime.now().isoformat(),action))
                    finance_id=c.execute('select last_insert_rowid()').fetchone()[0]
                    stock_created=False
                    if tx_type=='Gider' and category=='Yem' and (f.get('feed_id') or '').strip():
                        try:feed_id=int(f.get('feed_id'));feed_qty=float(f.get('feed_quantity_kg') or 0);feed_unit=float(f.get('feed_unit_price') or 0)
                        except:return self.redirect('/finance','Yem stok bilgisi geçersiz.')
                        feedrow=c.execute('select name from feed_catalog where id=? and active=1',(feed_id,)).fetchone()
                        if not feedrow or feed_qty<=0:return self.redirect('/finance','Yem ve miktar bilgisi zorunludur.')
                        if feed_unit<=0:feed_unit=round(amount/feed_qty,4)
                        c.execute('insert into feed_stock_transactions(feed_id,tx_date,tx_type,quantity_kg,unit_price,notes) values(?,?,?,?,?,?)',(feed_id,f['tx_date'],'Giriş',feed_qty,feed_unit,'Finans giderinden otomatik stok girişi'))
                        stock_id=c.execute('select last_insert_rowid()').fetchone()[0]
                        c.execute('insert into feed_prices(feed_id,effective_date,price_per_kg,notes) values(?,?,?,?)',(feed_id,f['tx_date'],feed_unit,'Finans giderinden otomatik fiyat'))
                        c.execute('insert into feed_finance_links(feed_id,stock_tx_id,finance_id,quantity_kg,unit_price,created_at) values(?,?,?,?,?,?)',(feed_id,stock_id,finance_id,feed_qty,feed_unit,datetime.now().isoformat()))
                        stock_created=True
                    if action:c.execute('update animals set status=?,exit_date=?,exit_reason=?,sold_price=? where id=?',(action,f['tx_date'],category,amount,animal_id))
                    return self.redirect('/finance','Finans kaydı eklendi.' + (' Yem stoğu da otomatik artırıldı.' if stock_created else '') + (' Hayvan aktif sürüden çıkarıldı.' if action else ''))
        except sqlite3.IntegrityError as e:return self.redirect(path,'Aynı küpe numarası daha önce kaydedilmiş olabilir.')
        except Exception as e:
            if path in ('/finance/delete','/finance-delete'):
                return self.redirect('/finance','Finans kaydı silinemedi: '+str(e))
            return self.redirect(path,'Hata: '+str(e))

def local_ip():
    try:
        s=socket.socket(socket.AF_INET,socket.SOCK_DGRAM);s.connect(('8.8.8.8',80));ip=s.getsockname()[0];s.close();return ip
    except:return '127.0.0.1'

if __name__=='__main__':
    init_db(); ensure_archive_schema(); promote_mature_calves(); daily_backup(); print(f'Yerel: http://127.0.0.1:{PORT}/login');print(f'Ağ: http://{local_ip()}:{PORT}/login');ThreadingHTTPServer(('0.0.0.0',PORT),App).serve_forever()


# DEV4.3 — desktop-only ration workbench refinement; solver logic untouched
DEV43_DESKTOP_RATION_CSS = r'''
<style id="dev43-desktop-ration-polish">
@media(min-width:901px){
  /* FINAL/LOCKED brand placement: old cow wordmark, just above Dashboard, left-biased */
  .side{padding-top:0!important}
  .erp-side-brand{height:58px!important;min-height:58px!important;margin:0 0 4px!important;padding:0 12px 0 18px!important;display:flex!important;align-items:center!important;justify-content:flex-start!important;text-align:left!important;font-size:18px!important;line-height:1!important;box-sizing:border-box!important;background:transparent!important;border-bottom:1px solid rgba(255,255,255,.14)!important}

  body.erp-ration-reference .main{background:#f3f6f4!important;padding:10px 12px 34px!important}
  body.erp-ration-reference .erp-ration-layout{grid-template-columns:226px minmax(0,1fr)!important;gap:10px!important;align-items:start!important}
  body.erp-ration-reference .erp-ration-left{position:sticky!important;top:116px!important;max-height:calc(100vh - 146px)!important}
  body.erp-ration-reference .erp-panel,body.erp-ration-reference #ration-workbench,body.erp-ration-reference .target-compare-sticky{border:1px solid #d4e0d7!important;box-shadow:0 2px 8px rgba(21,61,40,.045)!important;border-radius:8px!important}
  body.erp-ration-reference .erp-panel-head{min-height:32px!important;padding:6px 9px!important;background:#f7faf8!important;border-bottom:1px solid #dde7e0!important}
  body.erp-ration-reference .erp-panel-body{padding:8px!important}
  body.erp-ration-reference #quick-feed-add .quick-feed-results{max-height:500px!important}
  body.erp-ration-reference #quick-feed-add .quick-feed-result{padding:7px 8px!important}
  body.erp-ration-reference #quick-feed-add .quick-feed-result:hover{background:#edf7f0!important}

  body.erp-ration-reference .target-compare-sticky{margin-top:8px!important;overflow:hidden!important}
  body.erp-ration-reference .target-mini-grid{gap:5px!important}
  body.erp-ration-reference .target-mini-card{min-height:88px!important;border-radius:7px!important}

  body.erp-ration-reference #ration-workbench{margin-top:9px!important;padding:0!important;overflow:hidden!important;background:#fff!important}
  body.erp-ration-reference #ration-workbench .workbench-head{min-height:42px!important;padding:8px 10px!important;margin:0!important;background:#f8fbf9!important;border-bottom:1px solid #dbe6de!important}
  body.erp-ration-reference #ration-workbench .workbench-head h3{font-size:14px!important}
  body.erp-ration-reference #ration-workbench .workbench-head .mut{font-size:10px!important}
  body.erp-ration-reference #ration-workbench .workbench-actions{gap:6px!important}
  body.erp-ration-reference #ration-workbench .workbench-actions .btn{padding:7px 10px!important;font-size:10px!important;border-radius:6px!important}
  body.erp-ration-reference #ration-bulk-form{padding:0 9px 9px!important}
  body.erp-ration-reference .ration-changebar{margin:0 -9px 0!important;padding:7px 9px!important;border-bottom:1px solid #edf1ee!important;background:#fff!important}

  /* Executive desktop summary: visible before table, not buried at page bottom */
  body.erp-ration-reference .desktop-ration-summary{display:grid!important;grid-template-columns:repeat(6,minmax(0,1fr));gap:0;margin:0 -9px 7px!important;border-bottom:1px solid #d6e2d9!important;background:#edf4ef!important}
  body.erp-ration-reference .desktop-ration-summary>div{background:#f8fbf9;padding:7px 9px;min-width:0;border-right:1px solid #dbe5de!important}
  body.erp-ration-reference .desktop-ration-summary>div:last-child{border-right:0!important}
  body.erp-ration-reference .desktop-ration-summary span{display:block;font-size:8px;color:#6a7a70;text-transform:uppercase;letter-spacing:.35px;font-weight:900}
  body.erp-ration-reference .desktop-ration-summary b{display:block;margin-top:2px;font-size:12px;color:#173d28;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
  body.erp-ration-reference .desktop-ration-summary .rumen-ok b{color:#16864a}
  body.erp-ration-reference .desktop-ration-summary .rumen-warn b{color:#b76b00}

  body.erp-ration-reference .ration-workbench-table{font-size:10.5px!important;border-collapse:separate!important;border-spacing:0!important;width:100%!important}
  body.erp-ration-reference .ration-workbench-table thead th{position:sticky;top:0;z-index:4;background:#eaf2ec!important;color:#294a37!important;font-weight:900!important;border-top:1px solid #d4dfd7!important;border-bottom:1px solid #cbd8cf!important;padding:6px 5px!important;white-space:nowrap!important}
  body.erp-ration-reference .ration-workbench-table thead th:first-child{left:0;z-index:6!important}
  body.erp-ration-reference .ration-workbench-table tbody td{padding:5px 5px!important;border-bottom:1px solid #e4ebe6!important;vertical-align:middle!important}
  body.erp-ration-reference .ration-workbench-table tbody tr:nth-child(even){background:#f7faf8!important}
  body.erp-ration-reference .ration-workbench-table tbody tr:hover{background:#e9f4ed!important}
  body.erp-ration-reference .ration-workbench-table tbody td:first-child{position:sticky;left:0;z-index:2;background:inherit!important;font-weight:800!important;min-width:210px!important;max-width:280px!important}
  body.erp-ration-reference .ration-workbench-table .ration-qty{width:54px!important;height:28px!important;padding:2px 4px!important;text-align:center!important;font-weight:900!important}
  body.erp-ration-reference .ration-workbench-table .qty-step{width:25px!important;height:28px!important;padding:0!important}
  body.erp-ration-reference .ration-workbench-table .qty-zero{padding:5px 8px!important}
  body.erp-ration-reference .ration-savebar{position:static!important;bottom:auto!important;z-index:8!important;display:flex!important;justify-content:flex-end!important;background:#fff!important;border-top:1px solid #dce6df!important;margin:6px -9px -9px!important;padding:7px 9px!important;backdrop-filter:none!important}
  body.erp-ration-reference .ration-savebar .btn{padding:8px 12px!important;border-radius:6px!important}

  body.erp-ration-reference #smart-balance{margin-top:10px!important;padding:12px!important;border-color:#d4e0d7!important;background:#fff!important}
  body.erp-ration-reference #smart-balance h3{font-size:14px!important}
  body.erp-ration-reference #smart-balance .mut{font-size:10px!important}
  body.erp-ration-reference .smart-solution-grid{grid-template-columns:repeat(3,minmax(0,1fr))!important;gap:8px!important;margin-top:8px!important}
  body.erp-ration-reference .smart-solution{border-left:3px solid #218a50!important;border-radius:7px!important;padding:9px!important}
  body.erp-ration-reference .smart-solution .effect{background:#f0f7f2!important;border:1px solid #dce9df!important}
}
@media(max-width:1200px) and (min-width:901px){
 body.erp-ration-reference .desktop-ration-summary{grid-template-columns:repeat(3,minmax(0,1fr))!important}
 body.erp-ration-reference .smart-solution-grid{grid-template-columns:1fr!important}
}
</style>
<script id="dev43-desktop-ration-polish-script">/* DEV4.5: redundant desktop ration summary removed; Hedef ↔ Mevcut cards are authoritative. */</script>
'''

_old_page_dev43 = page

def page(title, body, path='/', user='admin', flash=''):
    html = _old_page_dev43(title, body, path, user, flash)
    # Apply desktop ration polish only on the ration module; brand placement is global desktop.
    if path == '/rations':
        return html.replace('</body>', DEV43_DESKTOP_RATION_CSS + '</body>')
    brand_only = DEV43_DESKTOP_RATION_CSS.split('<script id="dev43-desktop-ration-polish-script">')[0]
    return html.replace('</body>', brand_only + '</body>')



# DEV4.9 — mobile toggle removed + desktop logo lowered; solver logic untouched
DEV47_UI_FIX = r"""
<style id="dev47-mobile-ration-authoritative">
/* Desktop logo: Dashboard üstündeki ayrılmış yeşil alanda görünür ve sabit. */
@media(min-width:901px){
  /* Brandı üst komut çubuğunun altına indir: Dashboard satırının hemen üstünde tam görünür. */
  #sideMenu.side{padding-top:0!important;overflow:visible!important}
  #sideMenu.side > .erp-side-brand{position:relative!important;top:auto!important;left:auto!important;right:auto!important;z-index:2!important;height:38px!important;min-height:38px!important;margin:14px 0 2px!important;padding:0 10px 0 18px!important;display:flex!important;align-items:center!important;justify-content:flex-start!important;gap:6px!important;text-align:left!important;font-size:14px!important;line-height:1!important;box-sizing:border-box!important;background:transparent!important;border:0!important;border-bottom:1px solid rgba(255,255,255,.14)!important;color:#fff!important;visibility:visible!important;opacity:1!important;overflow:visible!important;transform:none!important}
  #sideMenu.side > .erp-side-brand b{display:inline!important;color:#fff!important;font-size:14px!important;line-height:1!important;white-space:nowrap!important}
  #sideMenu.side > .erp-side-brand + a{margin-top:0!important}
}

@media(max-width:900px){
  /* Masaüstü kokpit kurallarını mobilde kesin olarak sıfırla. */
  body:has(.workbench-shell) .workbench-shell{display:block!important;grid-template-columns:none!important;grid-template-rows:none!important;padding:8px!important;gap:0!important;overflow:visible!important;position:relative!important}
  body:has(.workbench-shell) .workbench-shell>.target-workspace{display:block!important}
  body:has(.workbench-shell) .workbench-shell>#ration-workbench,
  body:has(.workbench-shell) .workbench-shell>#ration-workbench~*:not(script){display:block!important;grid-column:auto!important;grid-row:auto!important;min-width:0!important;width:100%!important}

  /* Mobilde hedef ayar formu yer kaplamaz; düzenleme Rasyon Bilgileri alanından yapılır. */
  .mobile-target-toggle{display:none!important}
  body:has(.workbench-shell) .target-controlbar{display:none!important;position:static!important;width:100%!important;margin:0!important;padding:0!important;grid-column:auto!important;grid-row:auto!important}
  body:has(.workbench-shell) .target-controlbar .target-head{display:block!important;margin:0 0 9px!important}
  body:has(.workbench-shell) .target-controlbar .target-head h3{font-size:16px!important;white-space:normal!important;margin:0 0 5px!important}
  body:has(.workbench-shell) .target-controlbar .target-context{display:block!important;white-space:normal!important;overflow-wrap:anywhere!important;margin:0!important;padding:7px 8px!important;font-size:12px!important;line-height:1.35!important}
  body:has(.workbench-shell) .target-controlbar .target-form{display:grid!important;grid-template-columns:1fr 1fr!important;gap:8px!important;width:100%!important;margin-top:9px!important}
  body:has(.workbench-shell) .target-controlbar .target-form label{min-width:0!important;width:100%!important;font-size:11px!important}
  body:has(.workbench-shell) .target-controlbar .target-form input,
  body:has(.workbench-shell) .target-controlbar .target-form select{width:100%!important;min-width:0!important;box-sizing:border-box!important;min-height:42px!important;height:auto!important;font-size:14px!important;padding:8px!important}
  body:has(.workbench-shell) .target-controlbar .compact-target-btn{grid-column:1/-1!important;width:100%!important;min-height:44px!important;height:auto!important}

  /* Hedef ↔ Mevcut: telefonda yatay kart şeridi, metinler sıkışmaz. */
  body:has(.workbench-shell) .target-compare-sticky,
  body:has(.workbench-shell) .target-compare-sticky.is-floating{position:static!important;left:auto!important;top:auto!important;width:100%!important;height:auto!important;max-height:none!important;overflow:hidden!important;margin:8px 0 12px!important;padding:9px!important;background:#f8fbf9!important;border:1px solid #cfe3d5!important;border-radius:14px!important;box-shadow:none!important;backdrop-filter:none!important;z-index:auto!important;box-sizing:border-box!important}
  body:has(.workbench-shell) .target-compare-title{height:auto!important;min-height:24px!important;margin:0 2px 7px!important;padding:0!important;display:flex!important;align-items:center!important;font-size:14px!important}
  body:has(.workbench-shell) .target-compare-title>b{font-size:15px!important}
  body:has(.workbench-shell) .target-compare-title span{display:none!important}
  body:has(.workbench-shell) .nutri-mini-grid{display:flex!important;grid-template-columns:none!important;grid-template-rows:none!important;grid-auto-rows:auto!important;gap:7px!important;width:100%!important;height:auto!important;padding:0 0 4px!important;margin:0!important;overflow-x:auto!important;overflow-y:hidden!important;scroll-snap-type:x proximity;scrollbar-width:thin!important}
  body:has(.workbench-shell) .nutri-mini.nutri-compare-card,
  body:has(.workbench-shell) .nutri-mini{flex:0 0 150px!important;width:150px!important;height:116px!important;min-height:116px!important;min-width:150px!important;display:grid!important;grid-template-rows:26px 1fr 36px!important;padding:0!important;border-radius:12px!important;overflow:hidden!important;scroll-snap-align:start;background:#fff!important}
  body:has(.workbench-shell) .nutri-card-title{height:26px!important;min-height:26px!important;padding:6px 8px!important;font-size:13px!important;line-height:1!important}
  body:has(.workbench-shell) .nutri-compare-body{display:grid!important;grid-template-columns:1fr 1fr!important;min-height:0!important}
  body:has(.workbench-shell) .nutri-side{padding:4px 3px!important;min-width:0!important;display:flex!important;flex-direction:column!important;justify-content:center!important;align-items:center!important;text-align:center!important}
  body:has(.workbench-shell) .nutri-side span{font-size:9px!important;white-space:nowrap!important}
  body:has(.workbench-shell) .nutri-side b{font-size:18px!important;line-height:1.05!important;white-space:nowrap!important;overflow:hidden!important;text-overflow:ellipsis!important;max-width:100%!important}
  body:has(.workbench-shell) .nutri-card-footer{height:36px!important;min-height:36px!important;padding:4px!important;display:grid!important;grid-template-rows:auto auto!important;align-content:center!important;justify-items:center!important}
  body:has(.workbench-shell) .nutri-card-footer em{font-size:11px!important;white-space:nowrap!important;overflow:hidden!important;text-overflow:ellipsis!important;max-width:100%!important}
  body:has(.workbench-shell) .nutri-card-footer .nutri-diff{font-size:9px!important;white-space:nowrap!important;overflow:hidden!important;text-overflow:ellipsis!important;max-width:100%!important}
  body:has(.workbench-shell) #target-mini-cost{grid-column:auto!important}

  /* Çalışma masası mobil kart listesi. */
  body:has(.workbench-shell) #ration-workbench{display:block!important;width:100%!important;margin:0!important;padding:12px!important;border-radius:14px!important;overflow:visible!important;box-sizing:border-box!important}
  body:has(.workbench-shell) #ration-workbench .workbench-head{display:block!important;padding:0 0 10px!important;background:transparent!important;border:0!important}
  body:has(.workbench-shell) #ration-workbench .workbench-head h3{font-size:22px!important;line-height:1.1!important;margin:0!important}
  body:has(.workbench-shell) #ration-workbench .workbench-head .mut{display:block!important;font-size:13px!important;line-height:1.35!important;margin-top:4px!important}
  body:has(.workbench-shell) #ration-workbench .workbench-actions{display:grid!important;grid-template-columns:1fr 1fr!important;gap:8px!important;width:100%!important;margin-top:10px!important}
  body:has(.workbench-shell) #ration-workbench .workbench-actions .btn{width:100%!important;min-height:46px!important;font-size:14px!important;padding:9px!important}
  body:has(.workbench-shell) #ration-workbench form{padding:0!important}
  body:has(.workbench-shell) #ration-workbench form>div[style*="overflow:auto"]{overflow:visible!important;margin-top:7px!important}
  body:has(.workbench-shell) .ration-workbench-table,
  body:has(.workbench-shell) .ration-workbench-table tbody{display:block!important;width:100%!important;min-width:0!important}
  body:has(.workbench-shell) .ration-workbench-table thead{display:none!important}
  body:has(.workbench-shell) .ration-workbench-table tr.ration-row{display:grid!important;grid-template-columns:1fr 1fr!important;grid-template-areas:'name name' 'qty qty' 'price daily' 'remove remove'!important;gap:8px 10px!important;width:100%!important;min-width:0!important;margin:0 0 10px!important;padding:12px!important;border:1px solid #dce8df!important;border-radius:14px!important;background:#fff!important;box-shadow:none!important;box-sizing:border-box!important}
  body:has(.workbench-shell) .ration-workbench-table tr.ration-row td{display:none!important;position:static!important;left:auto!important;border:0!important;padding:0!important;width:auto!important;min-width:0!important;max-width:none!important;background:transparent!important;font-size:12px!important}
  body:has(.workbench-shell) .ration-workbench-table tr.ration-row td:nth-child(1){display:block!important;grid-area:name!important;font-size:15px!important;line-height:1.25!important;padding-bottom:7px!important;border-bottom:1px solid #e6eee8!important;overflow-wrap:anywhere!important}
  body:has(.workbench-shell) .ration-workbench-table tr.ration-row td:nth-child(2){display:block!important;grid-area:qty!important}
  body:has(.workbench-shell) .ration-workbench-table tr.ration-row td:nth-child(6){display:flex!important;grid-area:price!important;align-items:center!important;padding:8px!important;background:#f1f6f2!important;border-radius:9px!important}
  body:has(.workbench-shell) .ration-workbench-table tr.ration-row td:nth-child(6)::before{content:'Fiyat';display:block;margin-right:auto;color:#718077;font-size:11px;font-weight:800}
  body:has(.workbench-shell) .ration-workbench-table tr.ration-row td:nth-child(7){display:flex!important;grid-area:daily!important;align-items:center!important;justify-content:flex-end!important;padding:8px!important;background:#f1f6f2!important;border-radius:9px!important}
  body:has(.workbench-shell) .ration-workbench-table tr.ration-row td:nth-child(7)::before{content:'Günlük';display:block;margin-right:auto;color:#718077;font-size:11px;font-weight:800}
  body:has(.workbench-shell) .ration-workbench-table tr.ration-row td:nth-child(8){display:block!important;grid-area:remove!important;text-align:right!important}
  body:has(.workbench-shell) .ration-workbench-table .ration-stepper{display:grid!important;grid-template-columns:52px minmax(100px,1fr) 52px!important;gap:8px!important;width:100%!important;align-items:center!important}
  body:has(.workbench-shell) .ration-workbench-table .ration-stepper .btn{width:52px!important;height:48px!important;min-height:48px!important;padding:0!important;font-size:22px!important}
  body:has(.workbench-shell) .ration-workbench-table .ration-qty{width:100%!important;height:48px!important;min-width:0!important;font-size:20px!important;padding:8px!important;text-align:center!important;box-sizing:border-box!important}
  body:has(.workbench-shell) .ration-workbench-table .qty-zero{width:100%!important;min-height:44px!important;font-size:14px!important}
  body:has(.workbench-shell) .ration-savebar{position:sticky!important;bottom:34px!important;z-index:45!important;width:100%!important;margin:8px 0 0!important;padding:8px!important;background:rgba(255,255,255,.96)!important;border-radius:12px!important;box-shadow:0 -5px 18px rgba(18,60,37,.09)!important;box-sizing:border-box!important}
  body:has(.workbench-shell) .ration-savebar .btn{width:100%!important;min-height:50px!important;font-size:16px!important}

  /* Mobilde yem havuzu çalışma alanını işgal etmez; Yem Ekle düğmesi mevcut drawer/details alanını açar. */
  body:has(.workbench-shell) .erp-ration-layout{display:block!important;grid-template-columns:none!important}
  body:has(.workbench-shell) .erp-ration-left{display:none!important}
}
@media(max-width:520px){
  body:has(.workbench-shell) .target-controlbar .target-form{grid-template-columns:1fr!important}
  body:has(.workbench-shell) .nutri-mini.nutri-compare-card,
  body:has(.workbench-shell) .nutri-mini{flex-basis:142px!important;width:142px!important;min-width:142px!important}
}
</style>
<script id="dev49-mobile-toggle-cleanup">
(function(){
  function clean(){
    document.querySelectorAll('.mobile-target-toggle').forEach(function(el){ el.remove(); });
  }
  if(document.readyState==='loading'){document.addEventListener('DOMContentLoaded',clean);}else{clean();}
  setTimeout(clean,250);
})();
</script>

"""

_old_page_dev47 = page

def page(title, body, path='/', user='admin', flash=''):
    html = _old_page_dev47(title, body, path, user, flash)
    return html.replace('</body>', DEV47_UI_FIX + '</body>')


# DEV4.10 — desktop sidebar/header alignment only; solver/mobile ration untouched
DEV410_SIDEBAR_BRAND_FIX = r"""
<style id="dev410-sidebar-brand-final">
@media(min-width:901px){
  /* Keep the 34px utility header above everything, then start the green sidebar below it. */
  #sideMenu.side{
    top:var(--erp-top)!important;
    bottom:var(--erp-status)!important;
    padding-top:0!important;
    overflow-y:auto!important;
    overflow-x:hidden!important;
  }
  /* Brand occupies the same vertical band as the desktop command bar, so it can never sit under the white header. */
  #sideMenu.side > .erp-side-brand{
    position:relative!important;
    inset:auto!important;
    height:var(--erp-cmd)!important;
    min-height:var(--erp-cmd)!important;
    margin:0 0 4px!important;
    padding:0 18px!important;
    display:flex!important;
    align-items:center!important;
    justify-content:flex-start!important;
    gap:7px!important;
    box-sizing:border-box!important;
    background:transparent!important;
    border:0!important;
    border-bottom:1px solid rgba(255,255,255,.14)!important;
    color:#fff!important;
    font-size:17px!important;
    line-height:1!important;
    text-align:left!important;
    visibility:visible!important;
    opacity:1!important;
    transform:none!important;
    overflow:visible!important;
    white-space:nowrap!important;
  }
  #sideMenu.side > .erp-side-brand b{
    display:inline!important;
    color:#fff!important;
    font-size:17px!important;
    line-height:1!important;
    white-space:nowrap!important;
  }
  #sideMenu.side > .erp-side-brand + a{margin-top:0!important}
}
</style>
"""

_old_page_dev410 = page

def page(title, body, path='/', user='admin', flash=''):
    html = _old_page_dev410(title, body, path, user, flash)
    return html.replace('</body>', DEV410_SIDEBAR_BRAND_FIX + '</body>')
