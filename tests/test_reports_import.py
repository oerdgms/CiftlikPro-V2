import io
import os
import sys
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook
from pypdf import PdfReader


TEST_DATA = tempfile.TemporaryDirectory(prefix="ciftlikpro-report-test-")
os.environ["CIFTLIKPRO_DATA_DIR"] = TEST_DATA.name
sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "app"))
import server  # noqa: E402


class ReportImportTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        server.init_db()

    def test_official_pdf_row_parser(self):
        text = "TR380001234567 1 SIĞIR Simental DİŞİ 20/05/2021 TR380009876543 26/10/2021"
        rows = server._official_pdf_rows(text)
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["tag"], "TR380001234567")
        self.assertEqual(rows[0]["gender"], "DİŞİ")

    def test_xlsx_parse_prepare_and_export(self):
        wb = Workbook(); ws = wb.active
        ws.append(["KÜPE NO", "SÜRÜ NO", "TÜR", "IRK", "CİNSİYET", "DOĞUM TARİHİ", "ANA NO", "GELİŞ TARİHİ"])
        ws.append(["TR580000000101", 1, "SIĞIR", "Simental", "DİŞİ", "01/01/2020", "", "01/02/2020"])
        ws.append(["TR580000000102", 1, "SIĞIR", "Simental", "ERKEK", "01/08/2026", "TR580000000101", "02/08/2026"])
        data = io.BytesIO(); wb.save(data); wb.close()

        raw = server.parse_animal_import_file("hayvanlar.xlsx", data.getvalue())
        prepared = server.prepare_animal_import("hayvanlar.xlsx", raw)
        self.assertEqual(len(prepared), 2)
        self.assertTrue(all(row["valid"] for row in prepared))
        self.assertEqual(prepared[1]["record_type"], "Buzağı")

        export = server.animal_report_xlsx([], {"farm_name": "Test Çiftliği", "business_no": "TR1"})
        check = load_workbook(io.BytesIO(export), read_only=True)
        self.assertEqual(check.active["A1"].value, "Test Çiftliği · Tüm Hayvanlar Raporu")
        check.close()

    def test_report_combines_active_adults_and_calves(self):
        with server.db() as con:
            mother_id = con.execute(
                "insert into animals(tag,nickname,gender,breed,birth_date,notes,status) values(?,?,?,?,?,?,'Aktif')",
                ("TR580000000201", "Anne", "Dişi", "Simental", "2020-01-01", ""),
            ).lastrowid
            con.execute(
                "insert into calves(tag,mother_id,father_tag,birth_date,gender,notes,breed) values(?,?,?,?,?,?,?)",
                ("TR580000000202", mother_id, "", "2026-08-01", "Erkek", "", "Simental"),
            )
        rows = server.animal_report_rows()
        tags = {row["tag"] for row in rows}
        self.assertIn("TR580000000201", tags)
        self.assertIn("TR580000000202", tags)

    def test_direct_pdf_is_landscape_and_repeats_table_header(self):
        sample=[]
        for index in range(75):
            sample.append({'tag':f'TR5800{index:08d}','nickname':'Örnek Hayvan','group':'Dişi','species':'Sığır','breed':'Simental','gender':'Dişi','birth_date':'2024-01-01','mother_tag':'TR580000000001','arrival_date':'2024-02-01','paddock':'SA1','status':'Aktif'})
        content=server.animal_report_pdf(sample,{'farm_name':'Test Çiftliği','business_no':'TR1','owner_name':'Osman','province':'Sivas','district':'Şarkışla'},'Aktif Kayıtlar · Tüm Hayvanlar')
        reader=PdfReader(io.BytesIO(content))
        self.assertGreater(len(reader.pages),1)
        for page in reader.pages:
            self.assertGreater(float(page.mediabox.width),float(page.mediabox.height))
            self.assertIn('Küpe No',page.extract_text())

    def test_selected_columns_apply_to_xlsx_and_pdf(self):
        sample=[{'tag':'TR580000000301','nickname':'Kara','group':'Dişi','species':'Sığır','breed':'Simental','gender':'Dişi','birth_date':'2024-01-01','mother_tag':'','arrival_date':'2024-02-01','paddock':'SA1','status':'Aktif'}]
        query={'columns_mode':['custom'],'columns':['tag','nickname','paddock']}
        columns=server.animal_report_selected_columns(query)
        self.assertEqual([item[0] for item in columns],['tag','nickname','paddock'])

        content=server.animal_report_xlsx(sample,{'farm_name':'Test Çiftliği','business_no':'TR1'},columns)
        book=load_workbook(io.BytesIO(content),read_only=True);sheet=book.active
        self.assertEqual([sheet.cell(4,index).value for index in range(1,4)],['Küpe No','Takma Ad','Padok'])
        self.assertIsNone(sheet.cell(4,4).value);book.close()

        pdf=server.animal_report_pdf(sample,{'farm_name':'Test Çiftliği','business_no':'TR1'},'Aktif Kayıtlar',columns)
        text='\n'.join(page.extract_text() or '' for page in PdfReader(io.BytesIO(pdf)).pages)
        self.assertIn('Küpe No',text);self.assertIn('Takma Ad',text);self.assertIn('Padok',text)
        self.assertNotIn('Doğum Tarihi',text)

    def test_beef_starch_targets_follow_finishing_phase(self):
        expected={
            'Besi Başlangıç':(20.0,24.0,28.0),
            'Besi Geliştirme':(23.0,27.0,30.0),
            'Besi Bitirme':(25.0,29.0,31.0),
        }
        for phase,values in expected.items():
            target=server.beef_starch_targets(phase)
            self.assertEqual((target['starch_min'],target['starch_ideal_max'],target['starch_max']),values)

    def test_manual_phase_override_controls_all_solver_limits(self):
        limits=server.beef_phase_limits(250,1.4,6.5,0,'Besi Bitirme')
        self.assertEqual(limits['phase'],'Besi Bitirme')
        self.assertEqual((limits['starch_min'],limits['starch_ideal_max'],limits['starch_max']),(25.0,29.0,31.0))
        self.assertEqual((limits['roughage_min'],limits['roughage_max']),(30.0,40.0))

    def test_ration_target_panel_contains_combined_starch_rumen_card(self):
        rr={'id':1,'ration_type':'Besi','target_weight_kg':500,'target_adg_kg':1.5,'animal_type':'Besi Erkek','target_age_months':0,'target_beef_phase':'Besi Bitirme'}
        sm={'dm_kg':12.0,'cp_pct_dm':13.0,'me_mcal':30.0,'ndf_pct_dm':29.0,'endf_pct_dm':15.0,'ca_g':60.0,'p_g':35.0,'roughage_pct_dm':35.0,'concentrate_pct_dm':65.0,'cost':100.0,'starch_pct_dm':27.0,'starch_kg':3.24,'rapid_starch_pct_dm':20.0,'starch_degradability_coverage':1.0}
        html=server.ration_requirement_panel(rr,sm)
        self.assertIn('Rumen Dengesi',html)
        self.assertIn('Göreli asidoz riski',html)
        self.assertIn('klinik pH değildir',html)
        self.assertIn('target-mini-starch-current',html)
        self.assertIn('%25–29',html)

    def test_intact_male_uses_nasem_growing_bull_profile(self):
        # 520.833 kg canlı ağırlık yaklaşık 500 kg SBW'dir. Değerler NASEM
        # Chapter 20 Errata Table 20-2'nin 500 kg / 1.2 kg ADG sütunlarıdır.
        target=server.ration_requirement_targets(500/0.96,1.2,'Besi Erkek',0,'Otomatik')
        self.assertEqual(target['animal_profile'],'bull')
        self.assertAlmostEqual(target['nem_req_mcal'],9.4,places=1)
        self.assertAlmostEqual(target['neg_req_mcal'],4.5,places=1)
        self.assertAlmostEqual(target['mp_maint_g'],402,places=0)
        self.assertAlmostEqual(target['mp_gain_g'],358,places=0)
        self.assertAlmostEqual(target['ca_g'],42.4,places=1)
        self.assertAlmostEqual(target['p_g'],22.7,places=1)

    def test_500_kg_130_bull_checkpoint_matches_card_requirements(self):
        target=server.ration_requirement_targets(500,1.30,'Besi Erkek',12,'Otomatik')
        self.assertEqual(target['animal_profile'],'bull')
        self.assertAlmostEqual(target['nem_req_mcal'],9.076,places=2)
        self.assertAlmostEqual(target['neg_req_mcal'],4.775,places=2)
        self.assertAlmostEqual(target['mp_req_g'],774.6,places=1)
        self.assertAlmostEqual(target['cp_pct'],9.96,places=1)
        self.assertAlmostEqual(target['ca_g'],44.485,places=1)
        self.assertAlmostEqual(target['p_g'],23.305,places=1)

    def test_heifer_and_castrated_male_use_table_20_1_profile(self):
        target=server.ration_requirement_targets(500/0.96,1.2,'Düve',0,'Otomatik')
        self.assertEqual(target['animal_profile'],'steer_heifer')
        self.assertAlmostEqual(target['nem_req_mcal'],8.1,places=1)
        self.assertAlmostEqual(target['neg_req_mcal'],6.5,places=1)
        self.assertAlmostEqual(target['mp_gain_g'],267,places=0)
        castrated=server.ration_requirement_targets(500/0.96,1.2,'Kastre Erkek',0,'Otomatik')
        self.assertEqual(castrated['animal_profile'],'steer_heifer')

    def test_scientific_target_panel_uses_minimums_and_adg_capacity(self):
        rr={'id':1,'ration_type':'Besi','target_weight_kg':500,'target_adg_kg':1.3,'animal_type':'Besi Erkek','target_age_months':0,'target_beef_phase':'Otomatik'}
        sm={'dm_kg':11.4,'cp_pct_dm':12.2,'me_mcal':29.0,'nem_mcal':19.2,'neg_mcal':12.0,
            'nem_density':1.68,'neg_density':1.05,'ndf_pct_dm':30.0,'endf_pct_dm':15.0,
            'ca_g':56.0,'p_g':36.0,'roughage_pct_dm':35.0,'concentrate_pct_dm':65.0,
            'cost':100.0,'starch_pct_dm':28.0,'starch_kg':3.19,'rapid_starch_pct_dm':18.0,
            'starch_degradability_coverage':1.0}
        html=server.ration_requirement_panel(rr,sm)
        self.assertIn('Tosun / Boğa',html)
        self.assertIn('GCAA kapasitesi',html)
        self.assertIn('HP bir tabandır',html)
        self.assertIn('Minimum gereksinim',html)
        self.assertNotIn('ME (referans)',html)

    def test_413_small_adg_deficit_is_not_rendered_green(self):
        rr={'id':1,'ration_type':'Besi','target_weight_kg':280,'target_adg_kg':1.30,
            'animal_type':'Besi Erkek','target_age_months':12,'target_beef_phase':'Besi Başlangıç'}
        sm={'dm_kg':6.77,'cp_pct_dm':16.1,'me_mcal':20.0,'nem_mcal':11.0,'neg_mcal':6.0,
            'nem_density':1.62,'neg_density':.89,'ndf_pct_dm':33.1,'endf_pct_dm':22.6,
            'ca_g':78.0,'p_g':30.0,'roughage_pct_dm':40.0,'concentrate_pct_dm':60.0,
            'cost':104.08,'starch_pct_dm':24.2,'starch_kg':1.64,'rapid_starch_pct_dm':2.0,
            'starch_degradability_coverage':.07}
        with patch.object(server,'_achievable_adg',return_value=1.24):
            html=server.ration_requirement_panel(rr,sm)
        self.assertIn('⚠️ Hedef altı %4.6',html)
        self.assertNotIn("id='target-mini-adg-status'>✅ Yeterli",html)

    def test_413_starch_rounding_buffer_does_not_raise_rumen_risk(self):
        metrics={'starch_pct_dm':24.2,'endf_pct_dm':22.6,'rapid_starch_pct_dm':2.0,
                 'starch_degradability_coverage':.07}
        limits={'starch_ideal_max':24.0,'starch_max':28.0,'endf_min':12.0}
        risk=server._rumen_risk_assessment(metrics,limits)
        self.assertEqual(risk['level'],'Düşük')
        self.assertEqual(risk['confidence'],'Düşük')

    def test_413_assistant_distinguishes_roughage_share_from_effective_fiber(self):
        targets=server.ration_requirement_targets(280,1.30,'Besi Erkek',12,'Besi Başlangıç')
        limits=server.beef_phase_limits(280,1.30,6.54,12,'Besi Başlangıç')
        metrics={'predicted_dmi_kg':6.54,'achievable_adg_kg':1.24,'cp_pct_dm':16.1,
                 'endf_pct_dm':22.6,'roughage_pct_dm':40.0,'starch_pct_dm':24.2,
                 'ca_g':78.0,'p_g':30.0}
        feeds=[{'name':'SIĞIR SÜT YEMİ','category':'Ticari Karma Yem','solver_max_kg_day':0}]
        advice=server._ration_assistant(feeds,metrics,targets,limits)
        self.assertIn('Kaba yem KM oranı faz koridorunun altında',advice)
        self.assertNotIn('Etkili lif düşük',advice)
        self.assertIn('enerji açığı yönünde artırılmalı',advice)
        self.assertNotIn('etiket üst dozu girilmediğinden',advice)

    def test_413_long_ration_result_is_collapsed(self):
        message='🧠 Rasyon çözüldü. '+('Ayrıntılı bilimsel kontrol mesajı. '*20)
        html=server.page('Rasyon Çalışma Masası','<p>İçerik</p>','/rations','admin',message)
        self.assertIn('class="flash ration-result-flash"',html)
        self.assertIn('Ayrıntılar',html)
        self.assertNotIn('class="flash ration-result-flash" open',html)

    def test_cp_surplus_is_not_a_feasibility_blocker(self):
        targets=server.ration_requirement_targets(500,1.3,'Besi Erkek',0,'Otomatik')
        metrics={'dm_kg':11.6,'predicted_dmi_kg':11.6,'cp_pct_dm':12.2,'me_mcal':28.0,
                 'achievable_adg_kg':1.32,'ndf_pct_dm':30.0,'endf_pct_dm':13.0,
                 'starch_pct_dm':28.0,'rapid_starch_pct_dm':18.0,'starch_degradability_coverage':1.0,
                 'roughage_pct_dm':35.0,'ca_g':56.0,'p_g':36.0,'nem_density':1.7,'neg_density':1.1}
        limits=server.beef_phase_limits(500,1.3,11.6,0,'Otomatik')
        report=server._solver_feasibility_report(metrics,targets,limits)
        self.assertFalse(any(item.startswith('HP') for item in report['blockers']))

    def test_balance_score_treats_reasonable_cp_and_mineral_supply_as_minimums(self):
        targets=server.ration_requirement_targets(500,1.3,'Besi Erkek',12,'Otomatik')
        base={'dm_kg':11.4,'cp_pct_dm':targets['cp_pct'],'me_mcal':25.0,'nem_density':1.68,'neg_density':1.05,
              'ndf_pct_dm':30.0,'roughage_pct_dm':35.0,'starch_pct_dm':28.0,
              'ca_g':targets['ca_g'],'p_g':targets['p_g']}
        supplied={**base,'cp_pct_dm':12.2,'ca_g':56.0,'p_g':30.0}
        self.assertAlmostEqual(server.ration_balance_error(targets,base),server.ration_balance_error(targets,supplied),places=6)

    def test_feasibility_uses_diet_specific_predicted_dmi(self):
        targets=server.ration_requirement_targets(500,1.3,'Besi Erkek',0,'Otomatik')
        metrics={'dm_kg':11.6,'predicted_dmi_kg':11.6,'cp_pct_dm':10.5,'me_mcal':24.0,
                 'achievable_adg_kg':1.3,'ndf_pct_dm':30.0,'endf_pct_dm':13.0,
                 'starch_pct_dm':28.0,'rapid_starch_pct_dm':18.0,'starch_degradability_coverage':1.0,
                 'roughage_pct_dm':35.0,'ca_g':50.0,'p_g':30.0,'nem_density':1.68,'neg_density':1.05}
        limits=server.beef_phase_limits(500,1.3,11.6,0,'Otomatik')
        report=server._solver_feasibility_report(metrics,targets,limits)
        self.assertFalse(any(item.startswith('KM') for item in report['blockers']))

    def test_commercial_feed_has_no_invented_percent_cap_but_obeys_label(self):
        feed={'name':'SIĞIR SÜT YEMİ','category':'Ticari Karma Yem','dm_pct':88.35,'ndf_pct':22.04,
              'effective_ndf_pct':27.4,'cp_pct':18.78,'starch_pct':79.92,'me_mcal_kg':2.76}
        self.assertEqual(server._solver_feed_role(feed),'commercial')
        self.assertEqual(server._solver_starch_pct(feed),32.0)
        _,upper=server.smart_feed_bounds(feed,300,7.93,1.0,0,'Besi Geliştirme')
        self.assertGreater(upper*.8835/7.93,.35)

        labeled={**feed,'solver_min_kg_day':0.75,'solver_max_kg_day':2.25}
        lower,upper=server.smart_feed_bounds(labeled,300,7.93,1.0,0,'Besi Geliştirme')
        self.assertEqual((lower,upper),(0.75,2.25))

    def test_417_starch_hard_limit_rejects_ration_even_when_effective_fiber_is_enough(self):
        targets={'dmi_kg':8.0,'cp_pct':12.0,'me_mcal_day':20.0,'ndf_min':25.0,'ca_g':35.0,'p_g':20.0}
        limits={'starch_ideal_max':27.0,'starch_max':30.0,'endf_min':11.5,'roughage_min':37.0,'roughage_max':43.0}
        metrics={'dm_kg':8.0,'cp_pct_dm':12.0,'me_mcal':20.0,'ndf_pct_dm':30.0,'endf_pct_dm':13.0,
                 'starch_pct_dm':32.0,'rapid_starch_pct_dm':0.0,'starch_degradability_coverage':0.0,
                 'roughage_pct_dm':40.0,'ca_g':35.0,'p_g':20.0}
        report=server._solver_feasibility_report(metrics,targets,limits)
        self.assertEqual(report['status'],'unsafe')
        self.assertTrue(any('faz üst sınırı' in item for item in report['unsafe']))
        self.assertEqual(report['rumen_risk']['level'],'Orta')

    def test_high_starch_and_low_effective_fiber_are_rejected_together(self):
        targets={'dmi_kg':8.0,'cp_pct':12.0,'me_mcal_day':20.0,'ndf_min':25.0,'ca_g':35.0,'p_g':20.0}
        limits={'starch_ideal_max':27.0,'starch_max':30.0,'endf_min':11.5,'roughage_min':37.0,'roughage_max':43.0}
        metrics={'dm_kg':8.0,'cp_pct_dm':12.0,'me_mcal':20.0,'ndf_pct_dm':28.0,'endf_pct_dm':7.0,
                 'starch_pct_dm':33.0,'rapid_starch_pct_dm':27.0,'starch_degradability_coverage':1.0,
                 'roughage_pct_dm':40.0,'ca_g':35.0,'p_g':20.0}
        report=server._solver_feasibility_report(metrics,targets,limits)
        self.assertEqual(report['status'],'unsafe')
        self.assertFalse(report['rumen_risk']['clinical_ph_prediction'])

    def test_solver_does_not_hide_growth_target_by_overfeeding_dry_matter(self):
        targets={'dmi_kg':8.0,'cp_pct':12.0,'me_mcal_day':20.0,'adg':1.3,'ndf_min':25.0,'ca_g':35.0,'p_g':20.0}
        limits={'starch_ideal_max':27.0,'starch_max':30.0,'endf_min':11.5,'roughage_min':37.0,'roughage_max':43.0}
        metrics={'dm_kg':9.2,'cp_pct_dm':12.0,'me_mcal':20.0,'achievable_adg_kg':1.28,
                 'ndf_pct_dm':30.0,'endf_pct_dm':13.0,'starch_pct_dm':27.0,'rapid_starch_pct_dm':20.0,
                 'starch_degradability_coverage':1.0,'roughage_pct_dm':40.0,'ca_g':35.0,'p_g':20.0}
        report=server._solver_feasibility_report(metrics,targets,limits)
        self.assertEqual(report['status'],'infeasible')
        self.assertTrue(any(item.startswith('KM ') for item in report['blockers']))

    def test_inra_reference_fields_are_separate_and_loaded_for_exact_feed(self):
        with server.db() as con:
            row=con.execute("select starch_degradability_pct,inra_ufv,inra_pdi_g_kg_dm,processing_method from feed_catalog where name='ARPA EZMESİ'").fetchone()
        self.assertIsNotNone(row)
        self.assertEqual(row['starch_degradability_pct'],86.0)
        self.assertEqual(row['inra_ufv'],1.11)
        self.assertEqual(row['inra_pdi_g_kg_dm'],87.0)
        self.assertEqual(row['processing_method'],'Ezme')

    def test_grain_mix_share_is_calculated_on_dry_matter(self):
        feeds=[
            {'name':'ARPA EZMESİ','category':'Kesif Yemler','dm_pct':90.0,'ndf_pct':18.0,'cp_pct':12.0,'starch_pct':57.0},
            {'name':'BUĞDAY EZMESİ','category':'Kesif Yemler','dm_pct':80.0,'ndf_pct':12.0,'cp_pct':13.0,'starch_pct':63.0},
        ]
        total,wheat,barley=server._grain_mix_dm(feeds,[2.0,1.0])
        self.assertAlmostEqual(total,2.6)
        self.assertAlmostEqual(wheat/total,0.8/2.6)
        self.assertAlmostEqual(barley/total,1.8/2.6)

    def test_hp_and_me_accept_reasonable_surplus_but_not_deficit(self):
        self.assertIn('Uygun',server.nutrient_status(109,100,.05,.10)[0])
        self.assertIn('Fazla',server.nutrient_status(111,100,.05,.10)[0])
        self.assertIn('Eksik',server.nutrient_status(94,100,.05,.10)[0])

    def test_414_wheat_is_capped_at_thirty_percent_of_ration_dm(self):
        wheat={'name':'BUĞDAY EZMESİ','category':'Kesif Yemler','dm_pct':85.7,
               'ndf_pct':12.5,'cp_pct':13.5,'starch_pct':68.0}
        _,upper=server.smart_feed_bounds(wheat,500,10.25,1.3,12,'Besi Bitirme')
        self.assertLessEqual(upper*.857,10.25*.30+1e-8)

    def test_414_wheat_dominance_has_quality_penalty(self):
        barley={'name':'ARPA EZMESİ','category':'Kesif Yemler','dm_pct':88.7,
                'ndf_pct':18.6,'cp_pct':11.8,'starch_pct':58.0,'me_mcal_kg':3.0}
        wheat={'name':'BUĞDAY EZMESİ','category':'Kesif Yemler','dm_pct':85.7,
               'ndf_pct':12.5,'cp_pct':13.5,'starch_pct':68.0,'me_mcal_kg':3.1}
        dominated=server._feed_quality_penalty([barley,wheat],[.5,1.5],10.0)
        balanced=server._feed_quality_penalty([barley,wheat],[1.0,1.0],10.0)
        self.assertGreater(dominated,balanced)

    def test_415_wheat_over_forty_percent_of_grain_is_unsafe(self):
        targets={'dmi_kg':10.0,'cp_pct':12.0,'me_mcal_day':25.0,'adg':1.3,
                 'ndf_min':25.0,'ca_g':40.0,'p_g':22.0}
        limits={'starch_ideal_max':27.0,'starch_max':30.0,'endf_min':11.5,
                'roughage_min':37.0,'roughage_max':43.0}
        metrics={'dm_kg':10.0,'predicted_dmi_kg':10.0,'cp_pct_dm':12.0,'me_mcal':25.0,
                 'achievable_adg_kg':1.3,'ndf_pct_dm':30.0,'endf_pct_dm':14.0,
                 'starch_pct_dm':27.0,'rapid_starch_pct_dm':18.0,'starch_degradability_coverage':1.0,
                 'roughage_pct_dm':40.0,'ca_g':40.0,'p_g':22.0,'wheat_grain_pct_dm':40.1}
        report=server._solver_feasibility_report(metrics,targets,limits)
        self.assertEqual(report['status'],'unsafe')
        self.assertTrue(any('buğday tahıl KM payı' in x for x in report['unsafe']))

    def test_416_beef_profile_penalizes_dairy_feed_but_not_beef_feed(self):
        feeds=[
            {'name':'SUNAR KARDELEN 19.27 SÜT YEMİ','category':'Ticari Karma Yem','dm_pct':88.35},
            {'name':'SUNAR 15.26 GELİŞTİRME BESİ YEMİ','category':'Ticari Karma Yem','dm_pct':88.35},
        ]
        self.assertGreater(server._commercial_profile_penalty(feeds,[1.0,1.0],'Besi Erkek',8.0),0)
        self.assertEqual(server._commercial_profile_penalty(feeds,[0.0,2.0],'Besi Erkek',8.0),0)
        self.assertGreater(server._commercial_profile_penalty(feeds,[0.0,2.0],'Süt İneği',18.0),0)

    def test_416_commercial_feed_kind_is_not_guessed_for_neutral_mix(self):
        neutral={'name':'İŞLETME ÖZEL KARMA YEMİ','category':'Ticari Karma Yem','dm_pct':88.0}
        self.assertEqual(server._commercial_feed_kind(neutral),'neutral')

    def test_417_beef_feed_selected_closes_dairy_feed_bound(self):
        feeds=[
            {'name':'SUNAR KARDELEN 19.27 SÜT YEMİ','category':'Ticari Karma Yem'},
            {'name':'SUNAR 15.26 GELİŞTİRME BESİ YEMİ','category':'Ticari Karma Yem'},
            {'name':'ARPA EZMESİ','category':'Kesif Yemler'},
        ]
        adjusted=server._apply_commercial_profile_bounds(feeds,[(0,5),(0,5),(0,5)],'Besi Erkek')
        self.assertEqual(adjusted,[(0.0,0.0),(0,5),(0,5)])

    def test_417_three_percent_growth_deficit_is_infeasible(self):
        targets={'dmi_kg':8.0,'cp_pct':12.0,'me_mcal_day':20.0,'adg':1.40,'ndf_min':25.0,'ca_g':35.0,'p_g':20.0}
        limits={'starch_ideal_max':27.0,'starch_max':30.0,'endf_min':11.5,'roughage_min':37.0,'roughage_max':43.0}
        metrics={'dm_kg':8.0,'cp_pct_dm':12.0,'me_mcal':20.0,'achievable_adg_kg':1.36,
                 'ndf_pct_dm':30.0,'endf_pct_dm':13.0,'starch_pct_dm':27.0,'rapid_starch_pct_dm':18.0,
                 'starch_degradability_coverage':1.0,'roughage_pct_dm':40.0,'ca_g':35.0,'p_g':20.0}
        report=server._solver_feasibility_report(metrics,targets,limits)
        self.assertEqual(report['status'],'infeasible')
        self.assertTrue(any(item.startswith('GCAA kapasitesi') for item in report['blockers']))

    def test_414_commercial_reference_profiles_are_complete_and_normalized(self):
        names=('BUZAĞI BAŞLANGIÇ YEMİ','SUNAR BUZAĞI BÜYÜTME ÖZEL DÖNEM YEMİ',
               'SUNAR KARDELEN 19.27 SÜT YEMİ',
               'SIĞIR BESİ YEMİ,13,2700','SIĞIR BESİ YEMİ,14,2800',
               'SUNAR 15.26 GELİŞTİRME BESİ YEMİ','SIĞIR BESİ YEMİ,14,2600')
        with server.db() as con:
            rows=con.execute('select * from feed_catalog where name in ('+','.join('?' for _ in names)+')',names).fetchall()
        self.assertEqual(len(rows),len(names))
        for row in rows:
            self.assertEqual(row['category'],'Ticari Karma Yem')
            self.assertGreater(row['dm_pct'],0);self.assertGreater(row['cp_pct'],0)
            self.assertGreater(row['ndf_pct'],0);self.assertGreater(row['me_mcal_kg'],0)
            self.assertGreater(row['nem_mcal_kg'],0);self.assertGreater(row['neg_mcal_kg'],0)
            self.assertGreater(row['ca_pct'],0);self.assertGreater(row['p_pct'],0)
            self.assertLessEqual(row['starch_pct'],35.0)
            self.assertTrue(row['constraint_source'])

    def test_418_sunar_beef_real_label_profile_and_dose(self):
        with server.db() as con:
            row=con.execute("select * from feed_catalog where name='SUNAR 15.26 GELİŞTİRME BESİ YEMİ'").fetchone()
        self.assertIsNotNone(row)
        self.assertAlmostEqual(row['cp_pct'],16.978,places=3)
        self.assertAlmostEqual(row['me_mcal_kg'],2.943,places=3)
        self.assertAlmostEqual(row['nem_mcal_kg'],1.984,places=3)
        self.assertAlmostEqual(row['neg_mcal_kg'],1.333,places=3)
        self.assertAlmostEqual(row['fat_pct'],3.396,places=3)
        self.assertAlmostEqual(row['ash_pct'],8.749,places=3)
        self.assertAlmostEqual(row['na_pct'],.306,places=3)
        self.assertEqual(row['label_cp_pct_as_fed'],15.0)
        self.assertEqual(row['label_me_kcal_kg_as_fed'],2600.0)
        self.assertEqual(row['label_crude_fiber_pct_as_fed'],9.27)
        self.assertEqual(row['label_fat_pct_as_fed'],3.0)
        self.assertEqual(row['label_ash_pct_as_fed'],7.73)
        self.assertEqual(row['label_sodium_pct_as_fed'],.27)
        self.assertEqual(row['solver_max_kg_day'],10.0)
        self.assertIn('20/08/2026',row['constraint_source'])
        self.assertIn('referans tahminidir',row['source'])

    def test_418_sunar_kardelen_real_label_profile_and_dose(self):
        with server.db() as con:
            row=con.execute("select * from feed_catalog where name='SUNAR KARDELEN 19.27 SÜT YEMİ'").fetchone()
        self.assertIsNotNone(row)
        self.assertAlmostEqual(row['cp_pct'],21.505,places=3)
        self.assertAlmostEqual(row['me_mcal_kg'],3.056,places=3)
        self.assertAlmostEqual(row['nem_mcal_kg'],2.078,places=3)
        self.assertAlmostEqual(row['neg_mcal_kg'],1.413,places=3)
        self.assertAlmostEqual(row['fat_pct'],3.962,places=3)
        self.assertAlmostEqual(row['ash_pct'],7.799,places=3)
        self.assertAlmostEqual(row['na_pct'],.374,places=3)
        self.assertEqual(row['label_cp_pct_as_fed'],19.0)
        self.assertEqual(row['label_me_kcal_kg_as_fed'],2700.0)
        self.assertEqual(row['label_crude_fiber_pct_as_fed'],9.07)
        self.assertEqual(row['label_fat_pct_as_fed'],3.50)
        self.assertEqual(row['label_ash_pct_as_fed'],6.89)
        self.assertEqual(row['label_sodium_pct_as_fed'],.33)
        self.assertEqual((row['solver_min_kg_day'],row['solver_max_kg_day']),(6.0,12.0))
        self.assertEqual(server.dairy_feed_bounds(row,650,20.0),(6.0,12.0))
        self.assertIn('19/08/2026',row['source'])
        self.assertIn('referans tahminidir',row['source'])

    def test_415_sunar_calf_grower_is_named_and_classified_without_fake_label_claim(self):
        with server.db() as con:
            row=con.execute("select * from feed_catalog where name='SUNAR BUZAĞI BÜYÜTME ÖZEL DÖNEM YEMİ'").fetchone()
        self.assertIsNotNone(row)
        self.assertTrue(server._is_commercial_compound_feed(row))
        self.assertIn('60-120 gün',row['source'])
        self.assertIn('referans tahminidir',row['source'])
        self.assertIn('sayısal etiket',row['constraint_source'])

    def test_418_sunar_name_migration_preserves_feed_identity_and_history_link(self):
        with server.db() as con:
            row=con.execute("select id from feed_catalog where name='SUNAR KARDELEN 19.27 SÜT YEMİ'").fetchone()
            feed_id=row['id']
            beef=con.execute("select id from feed_catalog where name='SUNAR 15.26 GELİŞTİRME BESİ YEMİ'").fetchone()
            beef_id=beef['id']
            ration_id=con.execute("insert into rations(name,target_group,notes,active,created_at) values(?,?,?,?,?)",
                                  ('Sunar geçiş testi','Test','',1,'2026-08-29T00:00:00')).lastrowid
            con.execute('insert into ration_items(ration_id,feed_id,kg_per_head_day) values(?,?,?)',(ration_id,feed_id,1.0))
            stock_id=con.execute("insert into feed_stock_transactions(feed_id,tx_date,tx_type,quantity_kg,unit_price,notes) values(?,?,?,?,?,?)",
                                 (beef_id,'2026-08-31','Giriş',50,1,'Sunar geçiş testi')).lastrowid
            con.execute("update feed_catalog set name='SUNAR KARDELEN SÜT YEMİ,19,2700',source='Sunar resmi ürün adı ve 2020 katalog test profili' where id=?",(feed_id,))
            con.execute("update feed_catalog set name='ÇUKOYEM GELİŞTİRME BESİ YEMİ,15,2650',source='Sunar/Çukoyem ürün etiketi test profili' where id=?",(beef_id,))
        server.init_db()
        with server.db() as con:
            migrated=con.execute('select * from feed_catalog where id=?',(feed_id,)).fetchone()
            migrated_beef=con.execute('select * from feed_catalog where id=?',(beef_id,)).fetchone()
            link=con.execute('select feed_id from ration_items where ration_id=?',(ration_id,)).fetchone()
            stock_link=con.execute('select feed_id from feed_stock_transactions where id=?',(stock_id,)).fetchone()
            con.execute('delete from feed_stock_transactions where id=?',(stock_id,))
            con.execute('delete from ration_item_history where ration_id=?',(ration_id,))
            con.execute('delete from ration_items where ration_id=?',(ration_id,))
            con.execute('delete from rations where id=?',(ration_id,))
        self.assertEqual(migrated['name'],'SUNAR KARDELEN 19.27 SÜT YEMİ')
        self.assertEqual(link['feed_id'],feed_id)
        self.assertEqual(migrated['label_cp_pct_as_fed'],19.0)
        self.assertAlmostEqual(migrated['cp_pct'],21.505,places=3)
        self.assertEqual((migrated['solver_min_kg_day'],migrated['solver_max_kg_day']),(6.0,12.0))
        self.assertEqual(migrated_beef['name'],'SUNAR 15.26 GELİŞTİRME BESİ YEMİ')
        self.assertEqual(stock_link['feed_id'],beef_id)
        self.assertEqual(migrated_beef['label_crude_fiber_pct_as_fed'],9.27)
        self.assertAlmostEqual(migrated_beef['me_mcal_kg'],2.943,places=3)

    def test_414_high_cp_and_energy_capacity_are_explained_as_surplus(self):
        rr={'id':1,'ration_type':'Besi','target_weight_kg':500,'target_adg_kg':1.30,
            'animal_type':'Besi Erkek','target_age_months':12,'target_beef_phase':'Besi Bitirme'}
        sm={'dm_kg':10.25,'cp_pct_dm':13.6,'me_mcal':30.0,'nem_mcal':20.0,'neg_mcal':12.0,
            'nem_density':1.9,'neg_density':1.1,'ndf_pct_dm':30.7,'endf_pct_dm':14.7,
            'ca_g':84.0,'p_g':42.0,'roughage_pct_dm':40.0,'concentrate_pct_dm':60.0,
            'cost':104.09,'starch_pct_dm':27.0,'starch_kg':2.77,'rapid_starch_pct_dm':18.0,
            'starch_degradability_coverage':1.0}
        with patch.object(server,'_achievable_adg',return_value=1.41):
            html=server.ration_requirement_panel(rr,sm)
        self.assertIn('Enerjiye göre GCAA kapasitesi',html)
        self.assertIn('Gerçekleşen GCAA tahmini değildir',html)
        self.assertIn('⚠️ Yüksek %',html)
        self.assertIn('Kapasite +%8',html)

    def test_414_desktop_save_bar_does_not_cover_last_feed_row(self):
        css=server.DEV43_DESKTOP_RATION_CSS
        self.assertIn('.ration-savebar{position:static!important;bottom:auto!important',css)

    def test_419_limited_solution_is_only_a_small_wheat_share_deviation(self):
        targets={'dmi_kg':10.0,'cp_pct':12.0,'me_mcal_day':25.0,'adg':1.3,
                 'ndf_min':25.0,'ca_g':40.0,'p_g':22.0}
        limits={'starch_ideal_max':27.0,'starch_max':30.0,'endf_min':11.5,
                'roughage_min':37.0,'roughage_max':43.0,'grain_max':30.0}
        base={'dm_kg':10.0,'predicted_dmi_kg':10.0,'cp_pct_dm':12.0,'me_mcal':25.0,
              'achievable_adg_kg':1.3,'ndf_pct_dm':30.0,'endf_pct_dm':14.0,
              'starch_pct_dm':27.0,'rapid_starch_pct_dm':18.0,'starch_degradability_coverage':1.0,
              'roughage_pct_dm':40.0,'grain_pct_dm':20.0,'ca_g':40.0,'p_g':22.0}
        small=dict(base,wheat_grain_pct_dm=31.0)
        large=dict(base,wheat_grain_pct_dm=32.1)
        self.assertEqual(server._solver_feasibility_report(small,targets,limits)['status'],'limited')
        self.assertEqual(server._solver_feasibility_report(large,targets,limits)['status'],'infeasible')

    def test_419_total_grain_phase_cap_is_a_hard_save_gate(self):
        targets={'dmi_kg':10.0,'cp_pct':12.0,'me_mcal_day':25.0,'adg':1.3,
                 'ndf_min':25.0,'ca_g':40.0,'p_g':22.0}
        limits={'starch_ideal_max':27.0,'starch_max':30.0,'endf_min':11.5,
                'roughage_min':37.0,'roughage_max':43.0,'grain_max':30.0}
        metrics={'dm_kg':10.0,'predicted_dmi_kg':10.0,'cp_pct_dm':12.0,'me_mcal':25.0,
                 'achievable_adg_kg':1.3,'ndf_pct_dm':30.0,'endf_pct_dm':14.0,
                 'starch_pct_dm':27.0,'rapid_starch_pct_dm':18.0,'starch_degradability_coverage':1.0,
                 'roughage_pct_dm':40.0,'grain_pct_dm':30.1,'wheat_grain_pct_dm':20.0,
                 'ca_g':40.0,'p_g':22.0}
        report=server._solver_feasibility_report(metrics,targets,limits)
        self.assertEqual(report['status'],'unsafe')
        self.assertTrue(any('toplam tahıl' in item for item in report['unsafe']))

    def test_419_sunar_label_limits_survive_solver_bounds(self):
        with server.db() as con:
            beef=con.execute("select * from feed_catalog where name='SUNAR 15.26 GELİŞTİRME BESİ YEMİ'").fetchone()
            dairy=con.execute("select * from feed_catalog where name='SUNAR KARDELEN 19.27 SÜT YEMİ'").fetchone()
        self.assertLessEqual(server.smart_feed_bounds(beef,500,11.0,1.4,12,'Besi Bitirme')[1],10.0)
        self.assertEqual(server.dairy_feed_bounds(dairy,650,20.0),(6.0,12.0))

    def test_419_field_scenarios_250_350_500_beef_and_25l_dairy(self):
        beef_names=['ARPA SAMANI','ARPA EZMESİ','BUĞDAY EZMESİ',
                    "YONCA KURU OTU, KM'de %17-19 HP, %40-44 NDF",
                    'MISIR SİLAJI, %29-33 KM','SUNAR 15.26 GELİŞTİRME BESİ YEMİ']
        dairy_names=["YONCA KURU OTU, KM'de %17-19 HP, %40-44 NDF",
                     'MISIR SİLAJI, %29-33 KM','ARPA EZMESİ','SUNAR KARDELEN 19.27 SÜT YEMİ']
        with server.db() as con:
            beef=[con.execute('select * from feed_catalog where name=?',(name,)).fetchone() for name in beef_names]
            dairy=[con.execute('select * from feed_catalog where name=?',(name,)).fetchone() for name in dairy_names]
        for weight in (250,350,500):
            with self.subTest(weight=weight):
                solved,_,message=server.solve_smart_ration(beef,weight,1.4,'Besi Erkek',12,'Otomatik')
                self.assertIsNotNone(solved,message)
                quantities,metrics,_=solved
                self.assertTrue(all(value>0 for value in quantities))
                self.assertLessEqual(quantities[-1],10.0)
                self.assertLessEqual(metrics['wheat_grain_pct_dm'],32.0)
                self.assertIn(metrics['feasibility']['status'],('feasible','limited'))
        solved,_,message=server.solve_smart_dairy_ration(dairy,650,25)
        self.assertIsNotNone(solved,message)
        quantities,_,_=solved
        self.assertGreaterEqual(quantities[-1],6.0)
        self.assertLessEqual(quantities[-1],12.0)

    def test_4191_field_260kg_eight_feed_set_solves_inside_roughage_corridor(self):
        names=['ARPA EZMESİ','ARPA SAMANI','BUĞDAY KEPEĞİ',
               'SOYA KÜSPESİ, SOLVENT, %44','SUNAR 15.26 GELİŞTİRME BESİ YEMİ',
               'SUNAR BUZAĞI BÜYÜTME ÖZEL DÖNEM YEMİ',
               "YONCA KURU OTU, KM'de %17-19 HP, %40-44 NDF"]
        with server.db() as con:
            base=[dict(con.execute('select * from feed_catalog where name=?',(name,)).fetchone()) for name in names]
        cob=dict(base[1])
        # Gerçek saha yedeğindeki kullanıcı girişi: ME mevcut, NEm/NEg/TDN/eNDF
        # boş. Solver bunları sıfır enerji saymamalı; güvenli çalışma değerini türetmeli.
        cob.update({'name':'MISIR KOÇANI SİLAJI','category':'Özel Yem','dm_pct':35.0,
                    'ndf_pct':40.0,'effective_ndf_pct':0.0,'tdn_pct':0.0,
                    'me_mcal_kg':2.65,'nem_mcal_kg':0.0,'neg_mcal_kg':0.0,
                    'cp_pct':9.0,'starch_pct':0.0,'solver_min_kg_day':0.0,
                    'solver_max_kg_day':0.0})
        feeds=base[:3]+[cob]+base[3:]
        solved,targets,message=server.solve_smart_ration(feeds,260,1.4,'Besi Erkek',10,'Otomatik')
        self.assertIsNotNone(solved,message)
        quantities,metrics,_=solved
        self.assertTrue(all(value>0 for value in quantities))
        self.assertLessEqual(quantities[3],5.0)
        self.assertGreaterEqual(metrics['roughage_pct_dm'],targets['roughage_min'])
        self.assertLessEqual(metrics['roughage_pct_dm'],targets['roughage_max'])
        self.assertIn(metrics['feasibility']['status'],('feasible','limited'))

    def test_4192_missing_net_energy_is_derived_from_entered_me(self):
        feed={'name':'MISIR KOÇANI SİLAJI','category':'Özel Yem','dm_pct':35.0,
              'ndf_pct':40.0,'effective_ndf_pct':0.0,'tdn_pct':0.0,
              'me_mcal_kg':2.65,'nem_mcal_kg':0.0,'neg_mcal_kg':0.0}
        self.assertGreater(server._solver_nutrient(feed,'nem_mcal_kg'),1.7)
        self.assertGreater(server._solver_nutrient(feed,'neg_mcal_kg'),1.1)
        self.assertGreater(server._solver_nutrient(feed,'tdn_pct'),70.0)
        self.assertEqual(server._solver_nutrient(feed,'effective_ndf_pct'),70.0)

    def test_4193_live_ration_rows_use_same_normalized_energy_as_solver(self):
        source=Path(server.__file__).read_text(encoding='utf-8')
        self.assertIn('data-nem="{_solver_nutrient(x,\'nem_mcal_kg\'):.8f}"',source)
        self.assertIn('data-neg="{_solver_nutrient(x,\'neg_mcal_kg\'):.8f}"',source)
        self.assertNotIn('data-nem="{float(x[\'nem_mcal_kg\'] or 0):.8f}"',source)


if __name__ == "__main__":
    unittest.main()
